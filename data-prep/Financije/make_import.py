# -*- coding: utf-8 -*-
import sys, re, os
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT  = os.path.join(SCRIPT_DIR, "Financije 2026_3.xlsx")
OUTPUT = os.path.join(SCRIPT_DIR, "Financije_ZaSasu_import.xlsx")
AREA   = "Financije_1"

# ── Fills / fonts ─────────────────────────────────────
PINK_FILL   = PatternFill("solid", fgColor="FFE6F0")
BLUE_FILL   = PatternFill("solid", fgColor="E6F2FF")
ORANGE_FILL = PatternFill("solid", fgColor="FFC000")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
LEG_FILL    = PatternFill("solid", fgColor="7030A0")
SEP_FILL    = PatternFill("solid", fgColor="FFD0E0")
WHITE_FONT  = Font(color="FFFFFF", bold=True)
TITLE_FONT  = Font(bold=True, size=12)
BOLD_FONT   = Font(bold=True)
THIN        = Side(style='thin')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
L_ALIGN     = Alignment(horizontal='left',   vertical='center')
C_ALIGN     = Alignment(horizontal='center', vertical='center')
R_ALIGN     = Alignment(horizontal='right',  vertical='center')

def col_letter(n):
    s = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

# ── Attr column definitions (Area > CategoryPath > AttrName order) ─────────
# (category_path_no_area, attr_name, data_type)
ATTR_COLS = [
    ("Rashodi",                                                      "Iznos",          "number"),
    ("Rashodi",                                                      "Račun",          "suggest"),
    ("Rashodi",                                                      "Valuta",         "suggest"),
    ("Rashodi > Darovi i pokloni",                                   "Prigoda",        "text"),
    ("Rashodi > Darovi i pokloni",                                   "Za koga",        "text"),
    ("Rashodi > Digitalne pretplate",                                "Naziv",          "suggest"),
    ("Rashodi > Hrana i namirnice",                                  "Dućan",          "suggest"),
    ("Rashodi > Odgoda plaćanja / Rate",                             "Naziv",          "text"),
    ("Rashodi > Odgoda plaćanja / Rate",                             "Rata",           "text"),
    ("Rashodi > Ostali rashodi",                                     "Opis",           "text"),
    ("Rashodi > Osobna kupovina > Odjeća i obuća",                   "Dućan",          "text"),
    ("Rashodi > Osobna kupovina > Ostala kupovina",                  "Opis",           "text"),
    ("Rashodi > Prijevoz i auto > Gorivo",                           "Stanica",        "text"),
    ("Rashodi > Prijevoz i auto > Gorivo",                           "Vrsta",          "suggest"),
    ("Rashodi > Prijevoz i auto > Ostali troškovi auta",             "Vrsta",          "suggest"),
    ("Rashodi > Restoran i kava",                                    "Naziv",          "text"),
    ("Rashodi > Stanovanje > Komunalije",                            "Objekt",         "suggest"),
    ("Rashodi > Stanovanje > Komunalije",                            "Vrsta",          "suggest"),
    ("Rashodi > Stanovanje > Pričuva i Holding",                     "Objekt",         "suggest"),
    ("Rashodi > Stanovanje > Pričuva i Holding",                     "Vrsta",          "suggest"),
    ("Rashodi > Telekomunikacije",                                   "Operater",       "suggest"),
    ("Rashodi > Zdravlje i higijena > Dopunsko i životno osiguranje","Korisnik",       "text"),
    ("Rashodi > Zdravlje i higijena > Dopunsko i životno osiguranje","Osiguravatelj",  "suggest"),
    ("Rashodi > Zdravlje i higijena > Liječnici i ljekarna",         "Vrsta",          "suggest"),
    ("Rashodi > Zdravlje i higijena > Mirovinski fond",              "Fond",           "suggest"),
    ("Rashodi > Zdravlje i higijena > Mirovinski fond",              "Osoba",          "suggest"),
    ("Transferi",                                                    "Iznos",          "number"),
    ("Transferi",                                                    "Izvor",          "suggest"),
    ("Transferi",                                                    "Napomena",       "text"),
    ("Transferi",                                                    "Odredište",      "suggest"),
    ("Transferi",                                                    "Valuta",         "suggest"),
]

FIXED_COUNT    = 8   # A-H
ATTR_COL_START = FIXED_COUNT + 1  # 9 → col I

# Build lookup: (cat_path, attr_name) → col index (1-based)
ATTR_KEY_TO_COL = {
    (cat, attr): ATTR_COL_START + i
    for i, (cat, attr, _) in enumerate(ATTR_COLS)
}

# ── Hierarchy: which paths are ancestors of which ──────────────────────────
def ancestors(path):
    """Return all ancestor paths for a category path (including itself)."""
    parts = [p.strip() for p in path.split('>')]
    result = []
    for i in range(1, len(parts) + 1):
        result.append(' > '.join(parts[:i]))
    return set(result)

def is_relevant(attr_cat_path, event_cat_path):
    """True if attr_cat_path is an ancestor-or-self of event_cat_path."""
    return attr_cat_path in ancestors(event_cat_path)

# ── Mapping ────────────────────────────────────────────────────────────────
NACIN_TO_RACUN = {
    'Master': 'Mastercard (Koka)',
    'Visa':   'Visa (Saša)',
    'Zaba':   'Kokin tekući (Zaba)',
    'RF':     'Raiffeisen (Saša)',
}

def categorize(sto, nacin, iznos):
    """Returns (category_path, attrs_dict, comment).
    attrs_dict keys: (cat_path, attr_name).
    """
    sto   = str(sto).strip()   if sto   else ''
    nacin = str(nacin).strip() if nacin else ''
    racun = NACIN_TO_RACUN.get(nacin, nacin)

    # TRANSFERI — card payments & cash
    card_dest = {'Visa': 'Visa (Saša)', 'Master': 'Mastercard (Koka)',
                 'Mastercard': 'Mastercard (Koka)'}
    if sto in card_dest:
        return ('Transferi', {
            ('Transferi','Iznos'): iznos, ('Transferi','Valuta'): 'EUR',
            ('Transferi','Izvor'): racun, ('Transferi','Odredište'): card_dest[sto],
        }, sto)
    if sto == 'Koka':
        return ('Transferi', {
            ('Transferi','Iznos'): iznos, ('Transferi','Valuta'): 'EUR',
            ('Transferi','Izvor'): racun, ('Transferi','Odredište'): 'Kokin tekući (Zaba)',
            ('Transferi','Napomena'): 'Saša → Koka',
        }, sto)
    if sto in ('Gotovina', 'Cash', 'Sašin novčanik'):
        return ('Transferi', {
            ('Transferi','Iznos'): iznos, ('Transferi','Valuta'): 'EUR',
            ('Transferi','Izvor'): racun, ('Transferi','Odredište'): 'Gotovina',
            ('Transferi','Napomena'): sto,
        }, sto)

    # RASHODI base (Iznos, Račun, Valuta on L1)
    base = {
        ('Rashodi','Iznos'): iznos,
        ('Rashodi','Račun'): racun,
        ('Rashodi','Valuta'): 'EUR',
    }

    # HRANA
    hrana = {'Konzum':'Konzum','Konzum dostava':'Konzum dostava','Spar':'Spar',
             'Studenac':'Studenac','Lidl':'Lidl','Mlinar':'Mlinar','Voćarna':'Voćarna',
             'Bofrost':'Bofrost','Igomat':'Igomat','Pekara':'Pekara','Kraš':'Kraš',
             'Nespresso':'Nespresso','Kruh':None,'Hrana':None}
    if sto in hrana:
        p = 'Rashodi > Hrana i namirnice'
        a = {**base}
        if hrana[sto]: a[(p,'Dućan')] = hrana[sto]
        return (p, a, sto)

    # DIGITALNE PRETPLATE
    dig = {'Claude':'Claude','Youtube':'YouTube','HBO':'HBO','Disney':'Disney+',
           'Spotify':'Spotify','Apple':'Apple','Apple Cloud':'Apple','iCloude':'Apple',
           'Cloud':'Apple','Sky':'SkyShowtime','SkyShowtime':'SkyShowtime',
           'Skyshow':'SkyShowtime','Google':'Google','Audible SS':'Audible',
           'Audible DPS':'Audible','Audible':'Audible','SS audible':'Audible',
           'Jutarnji list':'Jutarnji list','Netdomena Igor':'Netdomena',
           'Amazon Prime':'PrimeVideo','PrimeVideo':'PrimeVideo'}
    if sto in dig:
        p = 'Rashodi > Digitalne pretplate'
        return (p, {**base,(p,'Naziv'):dig[sto]}, sto)

    # GORIVO
    if sto in ('Ina','Petrol','Gorivo','Gorivo dizel'):
        p = 'Rashodi > Prijevoz i auto > Gorivo'
        a = {**base,(p,'Vrsta'):'Dizel' if 'dizel' in sto.lower() else 'Benzin'}
        if sto in ('Ina','Petrol'): a[(p,'Stanica')] = sto
        return (p, a, sto)

    # PRIJEVOZ ostalo
    prij = {'Parking':'Parking','Taxi':'Taxi','Bolt':'Bolt',
            'Mjesečni parking':'Parking','Lacetti parking':'Parking',
            'Carglass':'Servis','Tehnički C5':'Servis','Šatrak':'Servis'}
    if sto in prij:
        p = 'Rashodi > Prijevoz i auto > Ostali troškovi auta'
        return (p, {**base,(p,'Vrsta'):prij[sto]}, sto)

    # KOMUNALIJE
    komu = {'Bulatova HEP':('HEP','Bulatova'),'Bulatova plin':('Plin','Bulatova'),
            'Plin razlika':('Plin','Bulatova')}
    if sto in komu:
        p = 'Rashodi > Stanovanje > Komunalije'
        v,o = komu[sto]
        return (p, {**base,(p,'Vrsta'):v,(p,'Objekt'):o}, sto)

    # PRIČUVA I HOLDING
    pric = {'Saša Holding':('Holding','Saša'),'Nataša Holding':('Holding','Nataša')}
    if sto in pric:
        p = 'Rashodi > Stanovanje > Pričuva i Holding'
        v,o = pric[sto]
        return (p, {**base,(p,'Vrsta'):v,(p,'Objekt'):o}, sto)

    # TELEKOMUNIKACIJE
    if sto in ('T-com','T-mobile'):
        p = 'Rashodi > Telekomunikacije'
        return (p, {**base,(p,'Operater'):sto}, sto)

    # DOPUNSKO / ŽIVOTNO osiguranje
    osig = {'Passsport':'Passport','PassSport':'Passport','Ljekarma':'Passport',
            'Generali police':'Generali'}
    if sto in osig:
        p = 'Rashodi > Zdravlje i higijena > Dopunsko i životno osiguranje'
        return (p, {**base,(p,'Osiguravatelj'):osig[sto]}, sto)

    # LIJEČNICI I LJEKARNA
    lijec = {'Biberon':'Ljekarna','Ljekarna':'Ljekarna','HLK':'HLK',
             'D-vitamin':'Ljekarna','Yasenka':'Ljekarna','Lijekovi za mamu':'Ljekarna'}
    if sto in lijec:
        p = 'Rashodi > Zdravlje i higijena > Liječnici i ljekarna'
        return (p, {**base,(p,'Vrsta'):lijec[sto]}, sto)

    # MIROVINSKI FOND (PP ...)
    if sto.startswith('PP '):
        p = 'Rashodi > Zdravlje i higijena > Mirovinski fond'
        a = {**base,(p,'Fond'):'Dopunsko'}
        if 'SS' in sto or 'Saša' in sto: a[(p,'Osoba')] = 'Saša'
        elif 'DPS' in sto or 'Koka' in sto: a[(p,'Osoba')] = 'Koka'
        return (p, a, sto)

    # RATE (pattern "Naziv X/Y") — before restoran so "Allianz X/10" lands here first
    rate_m = re.match(r'^(.+?)\s+(\d+/\d+)$', sto)
    if rate_m:
        naziv = rate_m.group(1).strip()
        rata  = rate_m.group(2)
        if naziv == 'Allianz':
            p = 'Rashodi > Zdravlje i higijena > Dopunsko i životno osiguranje'
            return (p, {**base,(p,'Osiguravatelj'):'Allianz'}, sto)
        p = 'Rashodi > Odgoda plaćanja / Rate'
        return (p, {**base,(p,'Naziv'):naziv,(p,'Rata'):rata}, sto)

    # RESTORAN I KAVA
    rest = {'Pizzeria':'Pizzeria','Afrodita':'Afrodita','Dubravica':'Dubravica',
            'Vidikovac':'Vidikovac','Fisherija':'Fisherija',
            'Restoran Time':'Restoran Time','Veronika':'Veronika',
            'Maslina':'Maslina','Picek':'Picek','Nautic pizza':'Nautic pizza',
            'Mullef':'Mullef','Chipoteka':'Chipoteka','Kava':None,'Batak':None}
    if sto in rest:
        p = 'Rashodi > Restoran i kava'
        a = {**base}
        if rest[sto]: a[(p,'Naziv')] = rest[sto]
        return (p, a, sto)

    # OSOBNA — Odjeća
    if sto in ('Galeb gaće','Decathlon'):
        p = 'Rashodi > Osobna kupovina > Odjeća i obuća'
        return (p, {**base,(p,'Dućan'):sto.split()[0]}, sto)

    # OSOBNA — Ostala kupovina
    ostala = {'Temu','DM','Kreatin','Gitara','Miš za komp','Myprotein',
              'Čaše','Ikea','Video game museum','Korica','Nordletics',
              'Purex','Cinestar','Body shop','GLS','Paket','Sljeme','Tisak'}
    if sto in ostala:
        p = 'Rashodi > Osobna kupovina > Ostala kupovina'
        return (p, {**base,(p,'Opis'):sto}, sto)

    # DAROVI
    if sto == 'Igor':
        p = 'Rashodi > Darovi i pokloni'
        return (p, {**base,(p,'Za koga'):'Igor'}, sto)

    # CATCH-ALL
    p = 'Rashodi > Ostali rashodi'
    return (p, {**base,(p,'Opis'):sto}, sto)


# ── Read source ─────────────────────────────────────────────────────────────
src_wb = openpyxl.load_workbook(INPUT, data_only=True)
src_ws = src_wb.worksheets[2]   # 'Za Sašu'

rows = []
for r in range(2, src_ws.max_row + 1):
    datum  = src_ws.cell(r, 1).value
    nacin  = src_ws.cell(r, 2).value
    sto    = src_ws.cell(r, 3).value
    iznos  = src_ws.cell(r, 4).value
    if not datum or not sto: continue
    rows.append((datum, nacin, sto, iznos))

print(f"Source rows: {len(rows)}")

# ── Build output ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Events'

row = 1

# ── LEGEND title ────────────────────────────────────────────────────────────
c = ws.cell(row, 1, 'ATTRIBUTE LEGEND:')
c.font = TITLE_FONT
ws.cell(row, 3).value = 'see Structure sheet for more details'
ws.cell(row, 3).font  = Font(italic=True, color='666666')
row += 1

# ── LEGEND header ───────────────────────────────────────────────────────────
for ci, h in enumerate(['Col','Area','Category_Path','Attribute','Type','Unit'], 1):
    c = ws.cell(row, ci, h)
    c.fill = LEG_FILL; c.font = WHITE_FONT; c.border = BORDER; c.alignment = C_ALIGN
row += 1

# ── LEGEND rows ─────────────────────────────────────────────────────────────
legend_start = row
for i, (cat_path, attr_name, dtype) in enumerate(ATTR_COLS):
    letter = col_letter(ATTR_COL_START + i)
    short  = cat_path.split(' > ')[-1]
    # Legend col C = category path WITHOUT area name (matches full_path in categoriesDict)
    data   = [letter, AREA, cat_path, attr_name, dtype, 'EUR' if (attr_name=='Iznos' and dtype=='number') else '']
    is_sep = (i == 0) or (cat_path != ATTR_COLS[i-1][0])
    for ci, v in enumerate(data, 1):
        c = ws.cell(row, ci, v if v else None)
        c.fill   = SEP_FILL if is_sep else PINK_FILL
        c.font   = BOLD_FONT if is_sep else Font()
        c.border = BORDER
        c.alignment = L_ALIGN
    row += 1
legend_end = row - 1

# Collapse non-separator legend rows
for r in range(legend_start, legend_end + 1):
    # Determine if this is a separator row
    i = r - legend_start
    cat_path = ATTR_COLS[i][0]
    is_sep = (i == 0) or (cat_path != ATTR_COLS[i-1][0])
    if not is_sep:
        ws_row = ws.row_dimensions[r]
        ws_row.outlineLevel = 1
        ws_row.hidden = True

row += 1  # blank separator

# ── EVENT DATA title ─────────────────────────────────────────────────────────
event_title_row = row
ws.cell(row, 1, 'EVENT DATA:').font = TITLE_FONT
ws.cell(row, 3, 'Summ (if relevant) ->').alignment = Alignment(horizontal='right')
row += 1

# ── EVENT DATA header ────────────────────────────────────────────────────────
event_header_row = row
fixed_hdrs = ['event_id','Area','Category_Path','event_date','session_start','created_at','User','leaf comment']
attr_hdrs  = [f"{attr} ({cat.split(' > ')[-1]})" for cat,attr,_ in ATTR_COLS]
for ci, h in enumerate(fixed_hdrs + attr_hdrs, 1):
    c = ws.cell(row, ci, h)
    c.fill = HEADER_FILL; c.font = WHITE_FONT; c.border = BORDER; c.alignment = C_ALIGN
row += 1

event_data_start = row

# ── DATA ROWS ────────────────────────────────────────────────────────────────
unmapped = {}
for datum, nacin, sto, iznos in rows:
    cat_path, attrs, comment = categorize(sto, nacin, iznos)

    # Normalise date
    if isinstance(datum, datetime):
        ev_date = datetime(datum.year, datum.month, datum.day)
    else:
        ev_date = None  # skip bad dates

    # Fixed cols A-H
    fixed_vals = [
        None,           # A event_id (empty = new)
        AREA,           # B Area
        cat_path,       # C Category_Path WITHOUT area name (matches full_path in categoriesDict)
        ev_date,        # D event_date
        '09:00',        # E session_start
        '09:00:01',     # F created_at
        '',             # G User
        comment,        # H comment (original 'Što')
    ]
    for ci, v in enumerate(fixed_vals, 1):
        c = ws.cell(row, ci, v)
        c.border = BORDER; c.alignment = L_ALIGN
        if ci <= 3:
            c.fill = PINK_FILL
        elif ci == 4:
            c.fill = BLUE_FILL
            c.number_format = 'YYYY-MM-DD'
        elif ci == 7:
            c.fill = PINK_FILL
        else:
            c.fill = BLUE_FILL

    # Attr cols I+
    for i, (ac_path, attr_name, dtype) in enumerate(ATTR_COLS):
        col_num = ATTR_COL_START + i
        key     = (ac_path, attr_name)
        val     = attrs.get(key)
        relevant = is_relevant(ac_path, cat_path)
        c = ws.cell(row, col_num, val)
        c.border    = BORDER
        c.fill      = BLUE_FILL if relevant else ORANGE_FILL
        c.alignment = R_ALIGN if dtype == 'number' else L_ALIGN
        if dtype == 'number' and val is not None:
            c.number_format = '0.##'

    # Track unmapped (Ostali rashodi catch-all)
    if cat_path == 'Rashodi > Ostali rashodi':
        unmapped[sto] = unmapped.get(sto, 0) + 1

    ws.row_dimensions[row].height = 18
    row += 1

event_data_end = row - 1

# SUBTOTAL for Iznos cols
for i, (ac_path, attr_name, dtype) in enumerate(ATTR_COLS):
    if dtype != 'number': continue
    col_num = ATTR_COL_START + i
    letter  = col_letter(col_num)
    c = ws.cell(event_title_row, col_num)
    c.value     = f'=SUBTOTAL(9,{letter}{event_data_start}:{letter}{event_data_end})'
    c.alignment = R_ALIGN

# Autofilter
ws.auto_filter.ref = f"A{event_header_row}:{col_letter(ATTR_COL_START+len(ATTR_COLS)-1)}{event_data_end}"

# Freeze panes
ws.freeze_panes = ws.cell(event_data_start, ATTR_COL_START)

# Column widths
widths = {'A':10,'B':14,'C':42,'D':12,'E':9,'F':10,'G':22,'H':28}
for col_letter_k, w in widths.items():
    ws.column_dimensions[col_letter_k].width = w
for i in range(len(ATTR_COLS)):
    ws.column_dimensions[col_letter(ATTR_COL_START+i)].width = 14

# Group col G (User)
ws.column_dimensions['G'].outlineLevel = 1

# ── STRUCTURE SHEET ──────────────────────────────────────────────────────────
# Required so that ExcelImportModal 'confirm-structure' flow triggers.
# importStructureExcel detects header by 'type' + 'categorypath' columns.
# CategoryPath in Structure sheet = WITH area name (unlike Events col C).

ss = wb.create_sheet('Structure')

STR_HDRS = ['Type','CategoryPath','Sort','AttrName','Slug','AttrType',
            'IsRequired','Val.Type','Default','ValMax','Unit','TextOptions',
            'DependsOn','WhenValue','Description']

# Suggest options for reuse
RACUN_OPTS = 'Kokin tekući (Zaba)|Mastercard (Koka)|Sašin tekući (PBZ)|Visa (Saša)|Raiffeisen (Saša)|Gotovina|KEKS (Saša)|Ašo (Saša)'
RACUN_ODRED = 'Kokin tekući (Zaba)|Mastercard (Koka)|Sašin tekući (PBZ)|Visa (Saša)|Raiffeisen (Saša)|Gotovina'

# (type, fullPath, attrName, slug, attrType, valType, default, unit, textOptions)
STRUCT = [
    # Area
    ('Area',      f'{AREA}',                                                                     '','','','','','',''),
    # Prihodi L1 + attrs
    ('Category',  f'{AREA} > Prihodi',                                                           '','','','','','',''),
    ('Attribute', f'{AREA} > Prihodi',  'Iznos',  'iznos',  'number',  'none', '',    'EUR', ''),
    ('Attribute', f'{AREA} > Prihodi',  'Račun',  'racun',  'suggest', 'suggest','','',   RACUN_OPTS),
    ('Attribute', f'{AREA} > Prihodi',  'Valuta', 'valuta', 'suggest', 'suggest','EUR','', 'EUR|HRK|USD'),
    ('Category',  f'{AREA} > Prihodi > Mirovina',                                                '','','','','','',''),
    ('Attribute', f'{AREA} > Prihodi > Mirovina',              'Stup',  'stup',  'suggest','suggest','','','Redovna|I stup|II stup|III stup'),
    ('Category',  f'{AREA} > Prihodi > Plaća i dodaci',                                         '','','','','','',''),
    ('Attribute', f'{AREA} > Prihodi > Plaća i dodaci',        'Vrsta', 'vrsta', 'suggest','suggest','','','Plaća|Prijevoz|Prehrana|Regres|Božićnica|Bonus|Naknada'),
    ('Category',  f'{AREA} > Prihodi > Najam — Anja',                                           '','','','','','',''),
    ('Attribute', f'{AREA} > Prihodi > Najam — Anja',          'Rata',  'rata',  'text',   'none','','',''),
    ('Category',  f'{AREA} > Prihodi > Honorar i freelance',                                     '','','','','','',''),
    ('Attribute', f'{AREA} > Prihodi > Honorar i freelance',   'Izvor', 'izvor', 'text',   'none','','',''),
    ('Category',  f'{AREA} > Prihodi > Povrat troškova',                                         '','','','','','',''),
    ('Attribute', f'{AREA} > Prihodi > Povrat troškova',       'Osoba', 'osoba', 'suggest','suggest','','','Zoran|Anja|Ostalo'),
    ('Attribute', f'{AREA} > Prihodi > Povrat troškova',       'Opis',  'opis',  'text',   'none','','',''),
    ('Category',  f'{AREA} > Prihodi > Ostali prihodi',                                          '','','','','','',''),
    ('Attribute', f'{AREA} > Prihodi > Ostali prihodi',        'Opis',  'opis',  'text',   'none','','',''),
    # Rashodi L1 + attrs
    ('Category',  f'{AREA} > Rashodi',                                                            '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi',  'Iznos',  'iznos',  'number',  'none', '',    'EUR', ''),
    ('Attribute', f'{AREA} > Rashodi',  'Račun',  'racun',  'suggest', 'suggest','','',   RACUN_OPTS),
    ('Attribute', f'{AREA} > Rashodi',  'Valuta', 'valuta', 'suggest', 'suggest','EUR','', 'EUR|HRK|USD'),
    # Stanovanje
    ('Category',  f'{AREA} > Rashodi > Stanovanje',                                              '','','','','','',''),
    ('Category',  f'{AREA} > Rashodi > Stanovanje > Komunalije',                                 '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Stanovanje > Komunalije',         'Objekt','objekt','suggest','suggest','','','Bulatova|Medulićeva|Kućište|Nena|Mama'),
    ('Attribute', f'{AREA} > Rashodi > Stanovanje > Komunalije',         'Vrsta', 'vrsta', 'suggest','suggest','','','HEP|Plin|Voda|Grijanje|Struja|Electrocoin'),
    ('Category',  f'{AREA} > Rashodi > Stanovanje > Pričuva i Holding',                         '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Stanovanje > Pričuva i Holding',  'Objekt','objekt','suggest','suggest','','','Nataša|Saša|Nena|Mama|Medulićeva|Kućište'),
    ('Attribute', f'{AREA} > Rashodi > Stanovanje > Pričuva i Holding',  'Vrsta', 'vrsta', 'suggest','suggest','','','Pričuva|Holding'),
    ('Category',  f'{AREA} > Rashodi > Stanovanje > Osiguranje nekretnine',                     '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Stanovanje > Osiguranje nekretnine','Osiguravatelj','osiguravatelj','text','none','','',''),
    ('Attribute', f'{AREA} > Rashodi > Stanovanje > Osiguranje nekretnine','Objekt',       'objekt',       'text','none','','',''),
    # Telekomunikacije
    ('Category',  f'{AREA} > Rashodi > Telekomunikacije',                                        '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Telekomunikacije',   'Operater','operater','suggest','suggest','','','T-com|T-mobile|A1'),
    # Prijevoz
    ('Category',  f'{AREA} > Rashodi > Prijevoz i auto',                                         '','','','','','',''),
    ('Category',  f'{AREA} > Rashodi > Prijevoz i auto > Gorivo',                                '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Prijevoz i auto > Gorivo',        'Stanica','stanica','text',   'none','','',''),
    ('Attribute', f'{AREA} > Rashodi > Prijevoz i auto > Gorivo',        'Vrsta',  'vrsta',  'suggest','suggest','','','Dizel|Benzin'),
    ('Category',  f'{AREA} > Rashodi > Prijevoz i auto > Ostali troškovi auta',                  '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Prijevoz i auto > Ostali troškovi auta','Vrsta','vrsta','suggest','suggest','','','Parking|HAK|Osiguranje|Servis|Registracija|Taxi|Bolt|Kazna'),
    # Zdravlje
    ('Category',  f'{AREA} > Rashodi > Zdravlje i higijena',                                     '','','','','','',''),
    ('Category',  f'{AREA} > Rashodi > Zdravlje i higijena > Dopunsko i životno osiguranje',     '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Zdravlje i higijena > Dopunsko i životno osiguranje','Korisnik',     'korisnik',     'text',   'none','','',''),
    ('Attribute', f'{AREA} > Rashodi > Zdravlje i higijena > Dopunsko i životno osiguranje','Osiguravatelj','osiguravatelj','suggest','suggest','','','Allianz|Triglav|Generali|Wustenrot|Jadransko|Passport'),
    ('Category',  f'{AREA} > Rashodi > Zdravlje i higijena > Liječnici i ljekarna',              '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Zdravlje i higijena > Liječnici i ljekarna','Vrsta','vrsta','suggest','suggest','','','HLK|Liječnik|Ljekarna|Laboratorij|Optika|Passport|Stomatolog'),
    ('Category',  f'{AREA} > Rashodi > Zdravlje i higijena > Mirovinski fond',                   '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Zdravlje i higijena > Mirovinski fond','Fond', 'fond', 'suggest','suggest','','','I stup|II stup|III stup|Dopunsko'),
    ('Attribute', f'{AREA} > Rashodi > Zdravlje i higijena > Mirovinski fond','Osoba','osoba','suggest','suggest','','','Koka|Saša'),
    # Porezi
    ('Category',  f'{AREA} > Rashodi > Porezi i javne obveze',                                   '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Porezi i javne obveze','Vrsta','vrsta','suggest','suggest','','','Porez i prirez|APN|Porez na dohodak|Porez na nekretninu|Komunalni doprinos'),
    # Hrana
    ('Category',  f'{AREA} > Rashodi > Hrana i namirnice',                                       '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Hrana i namirnice','Dućan','ducan','suggest','suggest','','','Konzum|Spar|Studenac|Lidl|Plodine|Mlinar|Voćarna|Dostava|Temu|Konzum dostava|Igomat|Bofrost|Pekara|Kraš|Nespresso'),
    # Digitalne pretplate
    ('Category',  f'{AREA} > Rashodi > Digitalne pretplate',                                     '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Digitalne pretplate','Naziv','naziv','suggest','suggest','','','Claude|YouTube|HBO|Disney+|Spotify|Audible|Apple|Netflix|PrimeVideo|SkyShowtime|Google|Jutarnji list|Netdomena'),
    # Rate
    ('Category',  f'{AREA} > Rashodi > Odgoda plaćanja / Rate',                                  '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Odgoda plaćanja / Rate','Naziv','naziv','text','none','','',''),
    ('Attribute', f'{AREA} > Rashodi > Odgoda plaćanja / Rate','Rata', 'rata', 'text','none','','',''),
    # Osobna kupovina
    ('Category',  f'{AREA} > Rashodi > Osobna kupovina',                                          '','','','','','',''),
    ('Category',  f'{AREA} > Rashodi > Osobna kupovina > Odjeća i obuća',                         '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Osobna kupovina > Odjeća i obuća','Dućan','ducan','text','none','','',''),
    ('Category',  f'{AREA} > Rashodi > Osobna kupovina > Ostala kupovina',                        '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Osobna kupovina > Ostala kupovina','Opis','opis','text','none','','',''),
    # Restoran
    ('Category',  f'{AREA} > Rashodi > Restoran i kava',                                          '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Restoran i kava','Naziv','naziv','text','none','','',''),
    # Darovi
    ('Category',  f'{AREA} > Rashodi > Darovi i pokloni',                                         '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Darovi i pokloni','Prigoda', 'prigoda', 'text','none','','',''),
    ('Attribute', f'{AREA} > Rashodi > Darovi i pokloni','Za koga', 'za-koga', 'text','none','','',''),
    # Ostali rashodi
    ('Category',  f'{AREA} > Rashodi > Ostali rashodi',                                           '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Ostali rashodi','Opis','opis','text','none','','',''),
    # Transferi L1 + attrs
    ('Category',  f'{AREA} > Transferi',                                                           '','','','','','',''),
    ('Attribute', f'{AREA} > Transferi','Iznos',     'iznos',     'number', 'none',   '', 'EUR', ''),
    ('Attribute', f'{AREA} > Transferi','Izvor',     'izvor',     'suggest','suggest','','',     RACUN_ODRED),
    ('Attribute', f'{AREA} > Transferi','Napomena',  'napomena',  'text',   'none',   '','',     ''),
    ('Attribute', f'{AREA} > Transferi','Odredište', 'odrediste', 'suggest','suggest','','',     RACUN_ODRED),
    ('Attribute', f'{AREA} > Transferi','Valuta',    'valuta',    'suggest','suggest','EUR','',  'EUR|HRK|USD'),
]

# Write header
for ci, h in enumerate(STR_HDRS, 1):
    c = ss.cell(1, ci, h)
    c.fill = HEADER_FILL; c.font = WHITE_FONT; c.border = BORDER; c.alignment = C_ALIGN

# Write data rows
for ri, (typ, path, aname, slug, atype, valtype, defval, unit, textopts) in enumerate(STRUCT, 2):
    row_data = [typ, path, '', aname, slug, atype, '', valtype, defval, '', unit, textopts, '', '', '']
    for ci, v in enumerate(row_data, 1):
        c = ss.cell(ri, ci, v if v else None)
        c.border = BORDER; c.alignment = L_ALIGN

ss.column_dimensions['A'].width = 12
ss.column_dimensions['B'].width = 62
ss.column_dimensions['D'].width = 18
ss.column_dimensions['E'].width = 16
ss.column_dimensions['F'].width = 10
ss.column_dimensions['H'].width = 10
ss.column_dimensions['L'].width = 55

wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")
print(f"Events rows: {event_data_end - event_data_start + 1}")
print(f"Attr columns: {len(ATTR_COLS)} (I–{col_letter(ATTR_COL_START+len(ATTR_COLS)-1)})")
print(f"Structure rows: {len(STRUCT)} ({sum(1 for r in STRUCT if r[0]=='Area')} areas, "
      f"{sum(1 for r in STRUCT if r[0]=='Category')} categories, "
      f"{sum(1 for r in STRUCT if r[0]=='Attribute')} attributes)")
print(f"\nCatch-all (Ostali rashodi) — provjeri mapiranje:")
for k, v in sorted(unmapped.items(), key=lambda x: -x[1]):
    print(f"  {v:3}x  {k!r}")
