"""
garmin_activities_to_xlsx.py
============================
Generates a roundtrip-compatible Excel workbook (Events + Structure sheets)
from Garmin Connect summarized activities export JSON.

Output area name: Fitness_Garmin (configurable via --area)
Structure: read from Sport_structure.xlsx (area "Fitness" renamed to --area)
Events:    mapped from Garmin summarizedActivities JSON files

Usage:
  python garmin_activities_to_xlsx.py
      [--garmin-dir  "path/to/DataFromGarmin"]
      [--structure   "path/to/Sport_structure.xlsx"]
      [--out         "Fitness_Garmin_import.xlsx"]
      [--area        "Fitness_Garmin"]
      [--from-date   "2015-01-01"]
      [--to-date     "2025-12-31"]
      [--email       "user@example.com"]
      [--geocode]              Enable reverse geocoding via Nominatim (slow first run)
      [--geocode-cache PATH]   JSON cache file path (default: Tools/geocode_cache.json)

Unit assumptions (Garmin summarizedActivities JSON):
  duration       ms  -> minutes  (/ 60000)
  distance       cm  -> km       (/ 100000)
  elevationGain  cm  -> m        (/ 100)
  avgSpeed       m/s
  startTimeLocal epoch ms (local tz, no conversion needed)

Pace format: "MM:SS" text (e.g. "6:06") — standard running notation.
  NOTE: Sport_structure.xlsx defines pace as 'number'. Before importing,
  change pace row data_type from 'number' to 'text' in the Structure sheet
  (or the script will auto-patch the copied Structure sheet).
"""

import json, os, sys, argparse, time, urllib.request
from pathlib import Path
from datetime import datetime, timezone
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────
# Default paths
# ─────────────────────────────────────────────────────────────────
SCRIPT_DIR  = Path(__file__).parent
REPO_ROOT   = SCRIPT_DIR.parent.parent
DATA_PREP   = REPO_ROOT / "Claude-temp_R/Data_preparation"

# GARMIN_DATA_DIR env var overrides default (set if moved outside repo)
# e.g.  set GARMIN_DATA_DIR=C:\0_Sasa\GarminData
_env_garmin = os.environ.get("GARMIN_DATA_DIR")
DEFAULT_GARMIN    = Path(_env_garmin) if _env_garmin else DATA_PREP / "DataFromGarmin"
DEFAULT_STRUCTURE = DATA_PREP / "Sport_structure.xlsx"
DEFAULT_OUT       = DATA_PREP / "Fitness_Garmin_import.xlsx"

# ─────────────────────────────────────────────────────────────────
# Styles (match app export palette)
# ─────────────────────────────────────────────────────────────────
PINK_FILL   = PatternFill("solid", fgColor="FFE6F0")
BLUE_FILL   = PatternFill("solid", fgColor="E6F2FF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
LEGEND_FILL = PatternFill("solid", fgColor="7030A0")
SEP_FILL    = PatternFill("solid", fgColor="FFD0E0")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT  = Font(bold=True, size=12)

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────────────────────────
# Column definitions
# ─────────────────────────────────────────────────────────────────

# Fixed cols A-H (must match FIXED_COLUMNS in excelExport.ts)
FIXED_COLS = ["event_id", "Area", "Category_Path", "event_date",
              "session_start", "created_at", "User", "comment"]

# Attribute definitions — (slug, name, data_type, unit, description, category_path)
# category_path = WITHOUT area name
ATTR_DEFS = [
    # ── Activity (parent, shared by all leaf types) ──────────────────────
    ("wormup-notes",       "Wormup_notes",         "text",   "",       "Warmup, stretching notes",              "Activity"),
    ("duration",           "duration",              "number", "min",    "Duration in minutes",                   "Activity"),
    ("hr-avg",             "hr_avg",                "number", "bpm",    "Average heart rate",                    "Activity"),
    ("hr-max",             "hr_max",                "number", "bpm",    "Max heart rate",                        "Activity"),
    ("kcal",               "kcal",                  "number", "kcal",   "Total calories burned",                 "Activity"),
    ("aerobic-effect",     "aerobic_effect",        "number", "",       "Aerobic Training Effect (0–5, Garmin)", "Activity"),
    ("anaerobic-effect",   "anaerobic_effect",      "number", "",       "Anaerobic Training Effect (0–5)",       "Activity"),
    ("training-load",      "training_load",         "number", "",       "Training Load (Garmin)",                "Activity"),
    ("intensity",          "intensity",             "text",   "",       "Subjective intensity: recovery/light/moderate/hard", "Activity"),
    ("mood",               "mood",                  "text",   "",       "Mood: :-(  :-/  :-)  :-D",              "Activity"),
    ("location",           "location",              "text",   "",       "Location: Zagreb, Maksimir / Sljeme / gym name...", "Activity"),
    # ── Outdoor leaf ─────────────────────────────────────────────────────
    ("outdoor-type",       "Outdoor_type",          "text",   "",       "Activity type: Hiking/Cycling/Run",     "Activity > Outdoor"),
    ("trening-type",       "Trening_type",          "text",   "",       "Training type: easy/long/interval/tempo/fartlek/race", "Activity > Outdoor"),
    ("distance",           "distance",              "number", "km",     "Distance in km",                        "Activity > Outdoor"),
    ("pace",               "pace",                  "text",   "min/km", "Average pace MM:SS format (e.g. 6:06)", "Activity > Outdoor"),
    ("terrain",            "terrain",               "text",   "",       "Terrain: ravno/brda/mix",               "Activity > Outdoor"),
    ("total-ascent",       "total_ascent",          "number", "m",      "Total elevation gain in metres",        "Activity > Outdoor"),
    # ── Gym > Cardio leaf ────────────────────────────────────────────────
    ("cardio-type",        "Cardio_type",           "text",   "",       "Z2/tempo/interval/wormup",              "Activity > Gym > Cardio"),
    ("equipment",          "equipment",             "text",   "",       "Equipment: orb/erg/gir/istezanje",      "Activity > Gym > Cardio"),
    ("intervals-description", "intervals_description","text", "",       "Interval description e.g. 8×30sec",    "Activity > Gym > Cardio"),
    # ── Gym > Strength leaf ──────────────────────────────────────────────
    ("strength-type",      "Strength_type",         "text",   "",       "Upp/Low/Core/wormup/test",              "Activity > Gym > Strength"),
    ("exercise-name",      "exercise_name",         "text",   "",       "Exercise name",                         "Activity > Gym > Strength"),
    ("sets-reps",          "sets_reps",             "text",   "",       'Sets and reps e.g. "3×10"',             "Activity > Gym > Strength"),
    ("weight-info",        "weight_info",           "text",   "",       "Weight information",                    "Activity > Gym > Strength"),
]

# slug → column index (I = 9, J = 10, …)
ATTR_COL_INDEX = {slug: 9 + i for i, (slug, *_) in enumerate(ATTR_DEFS)}
ATTR_COL_COUNT = len(ATTR_DEFS)

# ─────────────────────────────────────────────────────────────────
# Activity type mapping
# ─────────────────────────────────────────────────────────────────
OUTDOOR_TYPES = {"running", "trail_running", "cycling", "walking"}
STRENGTH_TYPES = {"strength_training"}
SKIP_TYPES = {"stop_watch", "sailing_v2", "boating_v2"}

OUTDOOR_TYPE_MAP = {
    "running":       "Run",
    "trail_running": "Run",
    "cycling":       "Cycling",
    "walking":       "Hiking",
}
# Location strings Garmin uses for Zagreb area but aren't real neighbourhoods
# Maps wrong name → normalised replacement
LOCATION_NORMALISE = {
    "Gornji Čehi":  "Zagreb",
    "Donji Čehi":   "Zagreb",
    "Čehi":         "Zagreb",
    "Stenjevec":    "Zagreb",
    "Podsused":     "Zagreb",
    "Vrapče":       "Zagreb",
    "Sesvete":      "Zagreb",
    "Susedgrad":    "Zagreb",
}

def normalise_location(text: str, geocoded: str = "") -> str:
    """Replace wrong location labels. If geocoded is provided, use it instead."""
    if geocoded:
        for wrong in LOCATION_NORMALISE:
            if wrong in text:
                return text.replace(wrong, geocoded)
        return text
    for wrong, right in LOCATION_NORMALISE.items():
        text = text.replace(wrong, right)
    return text


EQUIPMENT_MAP = {
    "elliptical":        "orb",
    "indoor_rowing":     "erg",
    "treadmill_running": None,
    "other":             None,
}

# ─────────────────────────────────────────────────────────────────
# Load Garmin JSON files
# ─────────────────────────────────────────────────────────────────
def load_garmin_activities(garmin_dir: Path) -> list:
    fitness_dir = garmin_dir / "DI_CONNECT/DI-Connect-Fitness"
    if not fitness_dir.exists():
        print(f"ERROR: fitness dir not found: {fitness_dir}")
        sys.exit(1)

    all_acts = []
    for path in sorted(fitness_dir.glob("*_summarizedActivities.json")):
        for enc in ("utf-8", "utf-8-sig", "latin-1"):
            try:
                with open(path, encoding=enc) as f:
                    data = json.load(f)
                # Each file is a list with one dict containing 'summarizedActivitiesExport'
                if isinstance(data, list):
                    for item in data:
                        if isinstance(item, dict) and "summarizedActivitiesExport" in item:
                            all_acts.extend(item["summarizedActivitiesExport"])
                        elif isinstance(item, dict):
                            all_acts.append(item)
                elif isinstance(data, dict):
                    for v in data.values():
                        if isinstance(v, list):
                            all_acts.extend(v)
                break
            except (UnicodeDecodeError, json.JSONDecodeError):
                continue

    print(f"Loaded {len(all_acts)} Garmin activities from {fitness_dir}")
    return all_acts


# ─────────────────────────────────────────────────────────────────
# Pace: seconds/km → "MM:SS" text
# ─────────────────────────────────────────────────────────────────
def pace_to_mmss(duration_min: float, dist_km: float) -> str | None:
    if not duration_min or not dist_km or dist_km <= 0:
        return None
    total_sec = (duration_min * 60.0) / dist_km
    # Sanity check: ignore implausible paces (<1 min/km or >30 min/km)
    if total_sec < 60 or total_sec > 1800:
        return None
    mm = int(total_sec // 60)
    ss = int(total_sec % 60)
    return f"{mm:02d}:{ss:02d}"


# ─────────────────────────────────────────────────────────────────
# Reverse geocoding via Nominatim (OpenStreetMap) — optional
# ─────────────────────────────────────────────────────────────────
_geo_cache: dict = {}
_geo_cache_path: Path | None = None

def load_geo_cache(cache_path: Path) -> None:
    global _geo_cache, _geo_cache_path
    _geo_cache_path = cache_path
    if cache_path.exists():
        with open(cache_path, encoding="utf-8") as f:
            _geo_cache = json.load(f)
        print(f"Geocode cache loaded: {len(_geo_cache)} entries from {cache_path}")

def save_geo_cache() -> None:
    if _geo_cache_path:
        with open(_geo_cache_path, "w", encoding="utf-8") as f:
            json.dump(_geo_cache, f, ensure_ascii=False, indent=2)

def _nominatim_lookup(lat: float, lon: float) -> str:
    url = (
        f"https://nominatim.openstreetmap.org/reverse"
        f"?lat={lat}&lon={lon}&format=json&zoom=18&addressdetails=1"
    )
    req = urllib.request.Request(url, headers={"User-Agent": "events-tracker-garmin/1.0"})
    with urllib.request.urlopen(req, timeout=8) as resp:
        data = json.loads(resp.read())
    addr = data.get("address", {})

    # city_district = "Zagreb", city = "Grad Zagreb" (too broad) — prefer city_district
    city = (addr.get("city_district")
            or addr.get("town")
            or addr.get("village")
            or "")
    # Nearby villages that are really Zagreb → normalise
    if city.lower() in ("gornji čehi", "donji čehi", "stenjevec", "čehi", "podsused"):
        city = "Zagreb"
    # Fallback: if city_district missing, use city and strip "Grad " prefix
    if not city:
        city = addr.get("city", "")
        if city.lower().startswith("grad "):
            city = city[5:]   # "Grad Zagreb" → "Zagreb"

    # suburb = "Gradska četvrt Maksimir" → strip prefix → "Maksimir"
    suburb = (addr.get("suburb") or addr.get("quarter") or addr.get("neighbourhood") or "")
    for prefix in ("Gradska četvrt ", "Gradska "):
        if suburb.startswith(prefix):
            suburb = suburb[len(prefix):]
            break

    if city and suburb:
        return f"{city}, {suburb}"
    return city or data.get("display_name", "")[:50]

def reverse_geocode(lat: float | None, lon: float | None) -> str:
    """Return location string, using cache. Respects Nominatim 1 req/sec limit."""
    if not lat or not lon:
        return ""
    # Cache key: rounded to 3 decimal places (~111m grid)
    key = f"{lat:.3f},{lon:.3f}"
    if key in _geo_cache:
        return _geo_cache[key]
    try:
        result = _nominatim_lookup(lat, lon)
        _geo_cache[key] = result
        time.sleep(1.1)   # Nominatim rate limit
    except Exception as e:
        print(f"  Geocode error ({key}): {e}", flush=True)
        result = ""
        _geo_cache[key] = result  # cache failure too, avoid retry
    return result


# ─────────────────────────────────────────────────────────────────
# Map one Garmin activity → import row dict
# ─────────────────────────────────────────────────────────────────
def map_activity(act: dict, area_name: str, user_email: str, geocode: bool = False) -> dict | None:
    act_type = act.get("activityType", "other")

    if act_type in SKIP_TYPES:
        return None

    # ── Date/time from startTimeLocal (epoch ms, local tz) ──
    ts_local = act.get("startTimeLocal") or act.get("startTimeGmt") or act.get("beginTimestamp", 0)
    try:
        ts_sec = float(ts_local) / 1000.0
        dt = datetime.utcfromtimestamp(ts_sec)
    except (TypeError, ValueError, OSError):
        return None

    date_str    = dt.strftime("%Y-%m-%d")
    time_hhmm   = dt.strftime("%H:%M")
    time_hhmmss = dt.strftime("%H:%M:%S")

    # ── Duration: ms → minutes ──
    duration_ms  = act.get("duration") or 0
    duration_min = round(float(duration_ms) / 60000.0, 1) if duration_ms else None

    # ── HR ──
    hr_avg = act.get("avgHr")
    hr_max = act.get("maxHr")
    hr_avg = round(hr_avg) if hr_avg else None
    hr_max = round(hr_max) if hr_max else None

    # ── Calories ──
    kcal = act.get("calories")
    kcal = round(kcal) if kcal else None

    # ── Training effects ──
    aerobic   = act.get("aerobicTrainingEffect")
    anaerobic = act.get("anaerobicTrainingEffect")
    t_load    = act.get("activityTrainingLoad")
    aerobic   = round(aerobic, 1)   if aerobic   else None
    anaerobic = round(anaerobic, 1) if anaerobic else None
    t_load    = round(t_load, 1)    if t_load    else None

    # ── Garmin activity name — normalise wrong location labels ──
    garmin_name = normalise_location(act.get("name", "") or "")

    # ── Category + leaf-specific attrs ──
    if act_type in OUTDOOR_TYPES:
        cat_path = "Activity > Outdoor"

        dist_cm  = float(act.get("distance") or 0)
        dist_km  = round(dist_cm / 100000.0, 2) if dist_cm > 5000 else None  # skip if <50m (GPS noise)
        elev_cm  = float(act.get("elevationGain") or 0)
        ascent_m = round(elev_cm / 100.0) if elev_cm > 0 else None
        pace     = pace_to_mmss(duration_min, dist_km) if dist_km else None

        # Location: GPS reverse geocode if enabled, else use Garmin locationName
        lat = act.get("startLatitude")
        lon = act.get("startLongitude")
        if geocode:
            location = reverse_geocode(lat, lon) or ""
            # Also normalise the activity name using the geocoded location
            garmin_name = normalise_location(garmin_name, location)
        else:
            location = act.get("locationName") or ""
            location  = normalise_location(location)

        leaf_vals = {
            "outdoor-type":  OUTDOOR_TYPE_MAP.get(act_type, "Run"),
            "trening-type":  None,
            "distance":      dist_km,
            "pace":          pace,
            "terrain":       None,
            "total-ascent":  ascent_m,
        }

    elif act_type in STRENGTH_TYPES:
        cat_path  = "Activity > Gym > Strength"
        location  = normalise_location(act.get("locationName") or "")
        leaf_vals = {
            "strength-type": None,
            "exercise-name": None,
            "sets-reps":     None,
            "weight-info":   None,
        }

    else:  # elliptical, indoor_rowing, treadmill, other
        cat_path = "Activity > Gym > Cardio"
        location  = normalise_location(act.get("locationName") or "")
        leaf_vals = {
            "cardio-type":           None,
            "equipment":             EQUIPMENT_MAP.get(act_type),
            "intervals-description": None,
        }

    location_val = location.strip() if location else None

    parent_vals = {
        "wormup-notes":    None,
        "duration":        duration_min,
        "hr-avg":          hr_avg,
        "hr-max":          hr_max,
        "kcal":            kcal,
        "aerobic-effect":  aerobic,
        "anaerobic-effect":anaerobic,
        "training-load":   t_load,
        "intensity":       None,
        "mood":            None,
        "location":        location_val,
    }

    # Comment: Garmin name + location (both for now, easy to review in Excel)
    if location_val and location_val not in garmin_name:
        comment = f"{garmin_name} [{location_val}]" if garmin_name else location_val
    else:
        comment = garmin_name

    return {
        "area":       area_name,
        "cat_path":   cat_path,
        "date":       date_str,
        "time":       time_hhmm,
        "time_sec":   time_hhmmss,
        "user_email": user_email,
        "comment":    comment,
        "attrs":      {**parent_vals, **leaf_vals},
    }


# ─────────────────────────────────────────────────────────────────
# Write ATTRIBUTE LEGEND section
# ─────────────────────────────────────────────────────────────────
def write_legend(ws, area_name: str, start_row: int) -> int:
    """Write ATTRIBUTE LEGEND. Returns next row after legend."""
    row = start_row

    # Title
    ws.cell(row, 1, "ATTRIBUTE LEGEND:").font = TITLE_FONT
    ws.cell(row, 3, "see Structure sheet for more details").font = Font(italic=True, color="666666")
    row += 1

    # Header
    legend_headers = ["Col", "Area", "Category_Path", "Attribute", "Type", "Unit", "Description"]
    for ci, h in enumerate(legend_headers, 1):
        c = ws.cell(row, ci, h)
        c.fill, c.font, c.border = LEGEND_FILL, HEADER_FONT, THIN
        c.alignment = Alignment(horizontal="center")
    row += 1

    # One row per attr def
    for i, (slug, name, data_type, unit, desc, cat_path) in enumerate(ATTR_DEFS):
        col_letter = get_column_letter(9 + i)   # I, J, K, …
        row_data   = [col_letter, area_name, cat_path, name, data_type, unit, desc]
        is_first_of_cat = (i == 0) or (ATTR_DEFS[i-1][5] != cat_path)

        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row, ci, val or None)
            c.fill   = SEP_FILL if is_first_of_cat else PINK_FILL
            c.border = THIN
            if is_first_of_cat:
                c.font = Font(bold=True)
        row += 1

    return row   # points to blank row after legend


# ─────────────────────────────────────────────────────────────────
# Write EVENT DATA section
# ─────────────────────────────────────────────────────────────────
def write_event_data(ws, rows: list, start_row: int) -> None:
    """Write EVENT DATA title + header + data rows."""
    row = start_row

    # Summary placeholder rows (Max / Min / Sum) — importer skips them
    for label in ["Max (if relevant) ->", "Min (if relevant) ->", "Summ (if relevant) ->"]:
        c = ws.cell(row, 8, label)   # col H
        c.font = Font(italic=True, color="666666")
        c.alignment = Alignment(horizontal="right")
        row += 1

    # EVENT DATA title
    ws.cell(row, 1, "EVENT DATA:").font = TITLE_FONT
    row += 1

    # Header row
    all_headers = FIXED_COLS + [name for (_, name, *_) in ATTR_DEFS]
    for ci, h in enumerate(all_headers, 1):
        c = ws.cell(row, ci, h)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, THIN
        c.alignment = Alignment(horizontal="center")
    row += 1

    # Data rows
    for ev in rows:
        # Fixed cols
        ws.cell(row, 1, "")                    # event_id — empty = new event
        ws.cell(row, 2, ev["area"])
        ws.cell(row, 3, ev["cat_path"])
        ws.cell(row, 4, ev["date"])
        ws.cell(row, 5, ev["time"])
        ws.cell(row, 6, ev["time_sec"])
        ws.cell(row, 7, ev["user_email"])
        ws.cell(row, 8, ev["comment"])

        # Attr cols (I onwards)
        for i, (slug, *_) in enumerate(ATTR_DEFS):
            val = ev["attrs"].get(slug)
            if val is not None:
                ws.cell(row, 9 + i, val)

        # Light blue tint on all data cells
        for ci in range(1, 9 + ATTR_COL_COUNT + 1):
            ws.cell(row, ci).fill = BLUE_FILL

        row += 1


# ─────────────────────────────────────────────────────────────────
# Copy Structure sheet from Sport_structure.xlsx
# ─────────────────────────────────────────────────────────────────
def copy_structure_sheet(wb_out, structure_path: Path, area_name: str) -> None:
    if not structure_path.exists():
        print(f"WARNING: Structure file not found: {structure_path} — skipping Structure sheet")
        return

    wb_src = openpyxl.load_workbook(structure_path, data_only=False)
    if "Structure" not in wb_src.sheetnames:
        print("WARNING: No 'Structure' sheet found in source file — skipping")
        return

    ws_src = wb_src["Structure"]
    ws_dst = wb_out.create_sheet("Structure")

    ORIGINAL_AREA = "Fitness"   # area name in the source file to replace

    for row in ws_src.iter_rows():
        for cell in row:
            dst = ws_dst.cell(row=cell.row, column=cell.column)

            val = cell.value
            # Replace area name in Category_Path column (col E = 5) and Area col (col C = 3)
            if isinstance(val, str) and ORIGINAL_AREA in val:
                val = val.replace(ORIGINAL_AREA, area_name)

            # Keep formulas as strings (data_only=False), but replace IFERROR area refs
            if isinstance(val, str) and val.startswith("="):
                val = val.replace(f'"{ORIGINAL_AREA}"', f'"{area_name}"')
                # Column E formulas reference area name via string matching on path
                # The IFERROR LEFT/FIND formulas don't need changes — they derive from col E

            dst.value = val

            # Copy styles (shallow)
            if cell.has_style:
                from copy import copy as _copy
                dst.font      = _copy(cell.font)
                dst.fill      = _copy(cell.fill)
                dst.border    = _copy(cell.border)
                dst.alignment = _copy(cell.alignment)

    # Copy column widths
    for col_letter, cd in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_letter].width = cd.width

    # Auto-patch 1: pace attr_type number → text (we output "MM:SS" strings)
    for row in ws_dst.iter_rows():
        vals = [str(c.value or "") for c in row]
        if len(vals) >= 9 and vals[6] == "pace" and vals[8] == "number":
            row[8].value = "text"

    # Auto-patch 2: append 'location' attribute row to Fitness > Activity
    # Structure sheet columns (1-indexed):
    # A=Type  B=IsLeaf  C=Area(formula)  D=SharedWith  E=CategoryPath
    # F=Sort  G=AttrName  H=Slug  I=AttrType  J=IsRequired
    # K=Val.Type  L=Default  M=Val.Max  N=Unit  O=TextOptions/Val.Min
    # P=DependsOn  Q=WhenValue  R=Description
    suggest_opts = (
        "Zagreb|Zagreb, Maksimir|Zagreb, Jarun|Zagreb, Sava nasip|"
        "Zagreb, Dotrsčina|Zagreb, Bundek|Zagreb, Šalata|"
        "Sljeme|Varaždin|Gorski kotar"
    )
    cat_path_with_area = f"{area_name} > Activity"
    loc_row = [
        "Attribute", "",
        f'=IFERROR(LEFT(E{ws_dst.max_row+1},FIND(" > ",E{ws_dst.max_row+1})-1),E{ws_dst.max_row+1})',
        "",
        cat_path_with_area,
        11,                 # sort_order (after mood=10)
        "location",         # AttrName
        "location",         # Slug
        "text",             # AttrType
        "FALSE",            # IsRequired
        "suggest",          # Val.Type
        "",                 # Default
        "",                 # Val.Max
        "",                 # Unit
        suggest_opts,       # TextOptions (pipe-separated)
        "",                 # DependsOn
        "",                 # WhenValue
        "Location of activity (Zagreb, Maksimir / Sljeme / gym name...)",
    ]
    ws_dst.append(loc_row)

    print(f"Structure sheet: {ws_src.max_row} copied rows + location attr appended")


# ─────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--garmin-dir",  default=str(DEFAULT_GARMIN))
    parser.add_argument("--structure",   default=str(DEFAULT_STRUCTURE))
    parser.add_argument("--out",         default=str(DEFAULT_OUT))
    parser.add_argument("--area",        default="Fitness_Garmin")
    parser.add_argument("--from-date",   default="2015-01-01")
    parser.add_argument("--to-date",     default="2099-12-31")
    parser.add_argument("--email",         default="")
    parser.add_argument("--geocode",       action="store_true",
                        help="Enable reverse geocoding via Nominatim (slow first run, cached)")
    parser.add_argument("--geocode-cache", default="",
                        help="Path to geocode cache JSON (default: Tools/geocode_cache.json)")
    args = parser.parse_args()

    garmin_dir     = Path(args.garmin_dir)
    structure_path = Path(args.structure)
    out_path       = Path(args.out)
    area_name      = args.area
    from_date      = args.from_date
    to_date        = args.to_date
    user_email     = args.email
    do_geocode     = args.geocode
    cache_path     = Path(args.geocode_cache) if args.geocode_cache else SCRIPT_DIR / "geocode_cache.json"

    if do_geocode:
        load_geo_cache(cache_path)
        print("Geocoding enabled — new coords: ~1 req/sec (progress shown)")

    # Load and map activities
    raw_acts = load_garmin_activities(garmin_dir)
    mapped   = []
    skipped  = 0
    geo_hits = 0
    geo_miss = 0

    for i, act in enumerate(raw_acts):
        date_ts = act.get("startTimeLocal") or act.get("beginTimestamp", 0)
        try:
            dt = datetime.utcfromtimestamp(float(date_ts) / 1000)
            date_str = dt.strftime("%Y-%m-%d")
        except Exception:
            skipped += 1
            continue

        if date_str < from_date or date_str > to_date:
            continue

        row = map_activity(act, area_name, user_email, geocode=do_geocode)
        if row:
            mapped.append(row)
            # Progress for geocoding
            if do_geocode and act.get("activityType") in OUTDOOR_TYPES:
                key = f"{act.get('startLatitude', 0):.3f},{act.get('startLongitude', 0):.3f}"
                if key in _geo_cache and _geo_cache[key]:
                    geo_hits += 1
                else:
                    geo_miss += 1
                if (geo_hits + geo_miss) % 50 == 0:
                    print(f"  Geocoded {geo_hits+geo_miss} outdoor activities "
                          f"(cache hits: {geo_hits}, new: {geo_miss})", flush=True)
        else:
            skipped += 1

    if do_geocode:
        save_geo_cache()
        print(f"Geocode cache saved: {len(_geo_cache)} entries -> {cache_path}")

    # Sort by date + time
    mapped.sort(key=lambda r: (r["date"], r["time"]))

    # Stats
    cat_counts: dict[str, int] = defaultdict(int)
    for r in mapped:
        cat_counts[r["cat_path"]] += 1

    print(f"\nMapped {len(mapped)} activities (skipped {skipped})")
    for cat, cnt in sorted(cat_counts.items()):
        print(f"  {cat}: {cnt}")

    # ── Build workbook ──
    wb = openpyxl.Workbook()
    ws_events = wb.active
    ws_events.title = "Events"

    # Legend
    next_row = write_legend(ws_events, area_name, start_row=1)
    next_row += 1   # blank separator row

    # Event data
    write_event_data(ws_events, mapped, start_row=next_row)

    # Column widths (approximate)
    col_widths = {1: 10, 2: 14, 3: 28, 4: 12, 5: 10, 6: 10, 7: 26, 8: 30}
    for ci, w in col_widths.items():
        ws_events.column_dimensions[get_column_letter(ci)].width = w
    for i in range(ATTR_COL_COUNT):
        ws_events.column_dimensions[get_column_letter(9 + i)].width = 14

    # Freeze panes at H (after fixed cols)
    ws_events.freeze_panes = "I1"

    # Structure sheet
    copy_structure_sheet(wb, structure_path, area_name)

    # HelpEvents stub (so importer doesn't complain)
    wb.create_sheet("HelpEvents")

    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    print(f"Events sheet: {len(mapped)} rows")
    print(f"Open in Excel, verify a few rows, then import via app.")
    print()
    print("NOTES:")
    print("  - session_start = Garmin startTimeLocal (local time, no timezone conversion)")
    print("  - distance:       Garmin cm / 100000 -> km  (centimeters, NOT meters)")
    print("  - elevationGain:  Garmin cm / 100 -> m        (verify a known trail)")
    print("  - intensity/mood/trening-type/terrain: empty - fill manually after import")
    print("  - cardio-type (Z2/interval/etc.): empty - fill manually after import")
    print("  - Gym > Strength events: all leaf attrs empty - fill manually")


if __name__ == "__main__":
    main()
