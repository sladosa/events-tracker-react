"""
make_health_structure.py
Generates Health_structure_import.xlsx — Structure Import file for
the Events Tracker "Health" area.

Run: python make_health_structure.py
Output: Health_structure_import.xlsx  (same folder)

Structure:
  Area:    Health
  L1:      Medical
  Leaf:    Lab Results   (numeric attrs + Lab/Status suggest)
  Leaf:    Medical Visit (Doktor/Vrsta suggest + Iznos/Napomena)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Column definitions (must match structureImport.ts header lookup) ──────────
HEADERS = [
    "Type",              # A
    "IsLeaf",            # B
    "Area",              # C
    "SharedWith",        # D
    "CategoryPath",      # E  ← key: "Health > Medical > Lab Results"
    "Sort",              # F
    "AttrName",          # G
    "Slug",              # H  (leave empty → auto-generated)
    "AttrType",          # I  number | text | boolean | datetime
    "IsRequired",        # J  TRUE | FALSE
    "Val.Type",          # K  suggest | none
    "Default",           # L
    "Val.Max (no)",      # M
    "Unit",              # N
    "TextOptions/Val.Min", # O  pipe-separated options for suggest
    "DependsOn",         # P
    "WhenValue",         # Q
    "Description",       # R
]

# Column indices (1-based) — matches HEADERS list above
COL = {h: i+1 for i, h in enumerate(HEADERS)}

LAB_PATH   = "Health > Medical > Lab Results"
VISIT_PATH = "Health > Medical > Medical Visit"

# ── Data rows ──────────────────────────────────────────────────────────────────
# Each dict: keys are column header names; missing keys = empty cell
ROWS = [
    # ── Area ──────────────────────────────────────────────────────────────────
    {"Type": "Area",     "CategoryPath": "Health"},

    # ── L1 category ───────────────────────────────────────────────────────────
    {"Type": "Category", "CategoryPath": "Health > Medical", "Sort": 1},

    # ── Lab Results leaf ──────────────────────────────────────────────────────
    {"Type": "Category", "CategoryPath": LAB_PATH, "Sort": 1},

    # Numeric attributes — Lab Results
    {"Type":"Attribute","CategoryPath":LAB_PATH,"Sort":10,
     "AttrName":"Zeljezo","AttrType":"number","Unit":"µmol/L",
     "Description":"Ref: 9–30 µmol/L (Iron / Fe)"},

    {"Type":"Attribute","CategoryPath":LAB_PATH,"Sort":20,
     "AttrName":"Eritrociti","AttrType":"number","Unit":"10¹²/L",
     "Description":"Ref: 4.34–5.72 × 10¹²/L (RBC)"},

    {"Type":"Attribute","CategoryPath":LAB_PATH,"Sort":30,
     "AttrName":"Hemoglobin","AttrType":"number","Unit":"g/L",
     "Description":"Ref: 138–175 g/L"},

    {"Type":"Attribute","CategoryPath":LAB_PATH,"Sort":40,
     "AttrName":"Feritin","AttrType":"number","Unit":"µg/L",
     "Description":"Ref: 20–400 µg/L (Ferritin)"},

    {"Type":"Attribute","CategoryPath":LAB_PATH,"Sort":50,
     "AttrName":"Kreatinin","AttrType":"number","Unit":"µmol/L",
     "Description":"Ref: 64–104 µmol/L"},

    {"Type":"Attribute","CategoryPath":LAB_PATH,"Sort":60,
     "AttrName":"Kolesterol","AttrType":"number","Unit":"mmol/L",
     "Description":"Ref: <5.2 mmol/L (Total cholesterol)"},

    {"Type":"Attribute","CategoryPath":LAB_PATH,"Sort":70,
     "AttrName":"Kolesterol-LDL","AttrType":"number","Unit":"mmol/L",
     "Description":"Ref: <3.0 mmol/L (LDL cholesterol)"},

    {"Type":"Attribute","CategoryPath":LAB_PATH,"Sort":80,
     "AttrName":"PSA","AttrType":"number","Unit":"µg/L",
     "Description":"Ref: 0–4 µg/L (Prostate-Specific Antigen)"},

    {"Type":"Attribute","CategoryPath":LAB_PATH,"Sort":90,
     "AttrName":"F/T ratio","AttrType":"number","Unit":"-",
     "Description":"Free PSA / Total PSA ratio"},

    # Suggest attributes — Lab Results
    {"Type":"Attribute","CategoryPath":LAB_PATH,"Sort":100,
     "AttrName":"Lab","AttrType":"text","Val.Type":"suggest",
     "TextOptions/Val.Min":"Synevo|Bates|Bolnica Sestre|Ostalo",
     "Description":"Laboratory where the test was done"},

    {"Type":"Attribute","CategoryPath":LAB_PATH,"Sort":110,
     "AttrName":"Status","AttrType":"text","Val.Type":"suggest",
     "TextOptions/Val.Min":"Normalno|Visoko|Nisko|Mješovito",
     "Description":"Overall assessment of results"},

    # ── Medical Visit leaf ────────────────────────────────────────────────────
    {"Type": "Category", "CategoryPath": VISIT_PATH, "Sort": 2},

    {"Type":"Attribute","CategoryPath":VISIT_PATH,"Sort":10,
     "AttrName":"Doktor","AttrType":"text","Val.Type":"suggest",
     "TextOptions/Val.Min":"Bates|Filipovic|Galovic|Mihaljevic|Runjaninovic|Ostalo",
     "Description":"Doctor or specialist seen"},

    {"Type":"Attribute","CategoryPath":VISIT_PATH,"Sort":20,
     "AttrName":"Vrsta","AttrType":"text","Val.Type":"suggest",
     "TextOptions/Val.Min":"Kontrola|UZV|RTG|EKG|MR|Operacija|Stomatolog|Ostalo",
     "Description":"Type of medical visit or procedure"},

    {"Type":"Attribute","CategoryPath":VISIT_PATH,"Sort":30,
     "AttrName":"Iznos","AttrType":"number","Unit":"EUR",
     "Description":"Cost of the visit or procedure"},

    {"Type":"Attribute","CategoryPath":VISIT_PATH,"Sort":40,
     "AttrName":"Napomena","AttrType":"text",
     "Description":"Notes: diagnosis, therapy, doctor recommendations"},
]

# ── Build workbook ─────────────────────────────────────────────────────────────

def make_fill(hex_rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_rgb)

FILL_HEADER = make_fill("4472C4")
FILL_AREA   = make_fill("D9E1F2")
FILL_CAT    = make_fill("EBF0FB")
FILL_ATTR   = make_fill("FFFFFF")

def write_structure_excel(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Structure"   # ← importer looks for sheet named "Structure"

    # ── Rows 1-6: info / legend (importer skips until header found) ────────────
    ws["A1"] = "Health Area — Structure Import"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Generated by make_health_structure.py"
    ws["A3"] = "Import via Structure tab → Import button"
    ws["A4"] = "Non-destructive: only adds new structure, never deletes."
    ws["A5"] = ""
    ws["A6"] = "INFO"

    # ── Row 7: header ──────────────────────────────────────────────────────────
    HEADER_ROW = 7
    for col_name, col_idx in COL.items():
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")

    # ── Row 8+: data ───────────────────────────────────────────────────────────
    for row_i, row_data in enumerate(ROWS):
        excel_row = HEADER_ROW + 1 + row_i
        row_type  = row_data.get("Type", "")

        fill = FILL_ATTR
        if row_type == "Area":     fill = FILL_AREA
        elif row_type == "Category": fill = FILL_CAT

        bold_row = row_type in ("Area", "Category")

        for col_name, col_idx in COL.items():
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill = fill
            if bold_row and col_name in ("Type", "CategoryPath"):
                cell.font = Font(bold=True)

    # ── Column widths ──────────────────────────────────────────────────────────
    widths = {
        "Type": 10, "IsLeaf": 7, "Area": 7, "SharedWith": 10,
        "CategoryPath": 42, "Sort": 6, "AttrName": 20, "Slug": 20,
        "AttrType": 10, "IsRequired": 10, "Val.Type": 10,
        "Default": 9, "Val.Max (no)": 9, "Unit": 10,
        "TextOptions/Val.Min": 45,
        "DependsOn": 15, "WhenValue": 12, "Description": 50,
    }
    for col_name, width in widths.items():
        ws.column_dimensions[get_column_letter(COL[col_name])].width = width

    # Freeze pane at G8 (match structure export convention)
    ws.freeze_panes = "G8"

    wb.save(path)
    print(f"Saved: {path}")
    print(f"  Rows: {len(ROWS)} data rows + 1 header")
    print(f"  Sheet: 'Structure'")
    print()
    print("Structure preview:")
    for r in ROWS:
        t = r.get("Type","")
        p = r.get("CategoryPath","")
        a = r.get("AttrName","")
        indent = "  " if t=="Category" else ("    " if t=="Attribute" else "")
        label = f"{p}" if t in ("Area","Category") else f"  [{a}]  ({r.get('AttrType','')} {r.get('Unit','')})"
        print(f"  {indent}{t:10s}  {label}")


if __name__ == "__main__":
    import os
    out = os.path.join(os.path.dirname(__file__), "Health_structure_import.xlsx")
    write_structure_excel(out)
