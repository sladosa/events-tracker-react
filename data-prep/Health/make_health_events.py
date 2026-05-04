# -*- coding: utf-8 -*-
"""
make_health_events.py
Read Bloodwork.xlsx and generate Health_events_import.xlsx
for import via Events Tracker → Activities → Import Excel.

Source:  Health/Bloodwork.xlsx  (sheet "Krv", rows 7+)
Output:  Health/Health_events_import.xlsx

Logic:
  - tip == "zdravstveni" only; skip all other tips
  - Row has any numeric value (Zeljezo..F/T) → Lab Results event
    comment = out-of-range summary (e.g. "Kolesterol H · Feritin L") + Pregled if any
  - Row has no numerics but has Pregled text → Medical Visit event
    Napomena = Pregled
  - Row has no numerics and no Pregled → skip

Run:  Health\run.bat make_health_events.py
"""

import sys
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT  = os.path.join(SCRIPT_DIR, "Bloodwork.xlsx")
OUTPUT = os.path.join(SCRIPT_DIR, "Health_events_import.xlsx")
AREA   = "Health"

# ── Styles ────────────────────────────────────────────────────────────────────
THIN        = Side(style='thin')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
L_ALIGN     = Alignment(horizontal='left',   vertical='center')
C_ALIGN     = Alignment(horizontal='center', vertical='center')
R_ALIGN     = Alignment(horizontal='right',  vertical='center')
BOLD_FONT   = Font(bold=True)
TITLE_FONT  = Font(bold=True, size=12)
WHITE_FONT  = Font(color="FFFFFF", bold=True)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")   # blue — fixed cols header
LEG_FILL    = PatternFill("solid", fgColor="7030A0")   # purple — legend header
LAB_FILL    = PatternFill("solid", fgColor="E2EFDA")   # green — Lab Results attr cols
VISIT_FILL  = PatternFill("solid", fgColor="DDEBF7")   # blue  — Medical Visit attr cols
OWN_FILL    = PatternFill("solid", fgColor="FFF2CC")   # amber — attrs NOT relevant for this row
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
ROW_FILL    = PatternFill("solid", fgColor="FAFAFA")

def make_fill(hex_rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_rgb)

def col_letter(n: int) -> str:
    s = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

# ── Attribute column definitions ──────────────────────────────────────────────
# Order: Lab Results numeric → Lab Results suggest → Medical Visit suggest → number → text
# (category_path without area, attr_name, data_type, unit)
ATTR_COLS = [
    ("Medical > Lab Results",  "Zeljezo",        "number",  "µmol/L"),
    ("Medical > Lab Results",  "Eritrociti",     "number",  "10¹²/L"),
    ("Medical > Lab Results",  "Hemoglobin",     "number",  "g/L"),
    ("Medical > Lab Results",  "Feritin",        "number",  "µg/L"),
    ("Medical > Lab Results",  "Kreatinin",      "number",  "µmol/L"),
    ("Medical > Lab Results",  "Kolesterol",     "number",  "mmol/L"),
    ("Medical > Lab Results",  "Kolesterol-LDL", "number",  "mmol/L"),
    ("Medical > Lab Results",  "PSA",            "number",  "µg/L"),
    ("Medical > Lab Results",  "F/T ratio",      "number",  "-"),
    ("Medical > Lab Results",  "Lab",            "suggest", ""),
    ("Medical > Medical Visit","Doktor",          "suggest", ""),
    ("Medical > Medical Visit","Vrsta",           "suggest", ""),
    ("Medical > Medical Visit","Iznos",           "number",  "EUR"),
    ("Medical > Medical Visit","Napomena",        "text",    ""),
]

FIXED_COUNT    = 8             # A-H
ATTR_COL_START = FIXED_COUNT + 1  # 9 → col I

# Lookup: attr_name for each attr column index
ATTR_COL_IDX = {attr_name: ATTR_COL_START + i for i, (_, attr_name, _, _) in enumerate(ATTR_COLS)}

LAB_PATH   = "Medical > Lab Results"
VISIT_PATH = "Medical > Medical Visit"

# ── Bloodwork column positions (1-based) ──────────────────────────────────────
COL_DATUM    = 2   # B
COL_TIP      = 3   # C
COL_PREGLED  = 4   # D
COL_ZELJEZO  = 5   # E
COL_ERITROC  = 6   # F
COL_HEMOGLOB = 7   # G
COL_FERITIN  = 8   # H
COL_KREATIN  = 9   # I
COL_KOLEST   = 10  # J  (Kolesterol-ukupni → "Kolesterol")
COL_KOLEST_L = 11  # K  (Kolesterol-LDL)
COL_PSA      = 12  # L
COL_FT       = 13  # M  (F/T → "F/T ratio")

# Mapping: (Bloodwork col, app attr name)
LAB_MAPPING = [
    (COL_ZELJEZO,  "Zeljezo"),
    (COL_ERITROC,  "Eritrociti"),
    (COL_HEMOGLOB, "Hemoglobin"),
    (COL_FERITIN,  "Feritin"),
    (COL_KREATIN,  "Kreatinin"),
    (COL_KOLEST,   "Kolesterol"),
    (COL_KOLEST_L, "Kolesterol-LDL"),
    (COL_PSA,      "PSA"),
    (COL_FT,       "F/T ratio"),
]

# ── Reference ranges (min, max) — None = no bound on that side ───────────────
# F/T ratio excluded: threshold depends on PSA level (complex, skip auto-flag)
REF_RANGES = {
    "Zeljezo":        (9.0,   30.0),
    "Eritrociti":     (4.34,  5.72),
    "Hemoglobin":     (138.0, 175.0),
    "Feritin":        (20.0,  400.0),
    "Kreatinin":      (64.0,  104.0),
    "Kolesterol":     (None,  5.2),
    "Kolesterol-LDL": (None,  3.0),
    "PSA":            (None,  4.0),
}

def range_flags(nums: dict) -> str:
    """Return 'Kolesterol H · Feritin L' for out-of-range values. Empty if all OK."""
    flags = []
    for attr_name, value in nums.items():
        ref = REF_RANGES.get(attr_name)
        if ref is None:
            continue
        lo, hi = ref
        if hi is not None and value > hi:
            flags.append(f"{attr_name} H")
        elif lo is not None and value < lo:
            flags.append(f"{attr_name} L")
    return " · ".join(flags)

# ── Read source ───────────────────────────────────────────────────────────────
print(f"Reading: {INPUT}")
src_wb = openpyxl.load_workbook(INPUT, data_only=True)
src_ws = src_wb["Krv"]

class EventRow:
    def __init__(self, datum, cat_path, comment, attrs):
        self.datum    = datum     # datetime
        self.cat_path = cat_path  # "Medical > Lab Results" or "Medical > Medical Visit"
        self.comment  = comment   # leaf comment (Pregled for Lab; empty for Visit)
        self.attrs    = attrs     # dict: attr_name → value

rows = []
skipped_tip = 0
skipped_nodate = 0
skipped_empty = 0

for r in range(7, src_ws.max_row + 1):
    tip    = src_ws.cell(r, COL_TIP).value
    if tip != 'zdravstveni':
        skipped_tip += 1
        continue

    datum   = src_ws.cell(r, COL_DATUM).value
    pregled = src_ws.cell(r, COL_PREGLED).value
    pregled = str(pregled).strip() if pregled else ''

    if datum is None:
        skipped_nodate += 1
        print(f"  WARN row {r}: no date, skipping — {pregled[:60]!r}")
        continue

    # Normalise datum to date only
    try:
        if isinstance(datum, datetime):
            date_val = datum.date()
        elif isinstance(datum, str):
            # Handle DD.MM.YYYY format
            parts = datum.strip().split('.')
            if len(parts) == 3:
                from datetime import date as _date
                date_val = _date(int(parts[2]), int(parts[1]), int(parts[0]))
            else:
                raise ValueError(f"Unknown date format: {datum!r}")
        else:
            date_val = datum  # assume date object
    except (ValueError, TypeError) as e:
        skipped_nodate += 1
        print(f"  WARN row {r}: invalid date {datum!r}, skipping — {e}")
        continue

    # Check for numeric values E-M
    nums = {}
    for bw_col, attr_name in LAB_MAPPING:
        v = src_ws.cell(r, bw_col).value
        if isinstance(v, (int, float)):
            nums[attr_name] = v

    if nums:
        # Lab Results event — comment: out-of-range flags + Pregled if any
        flags = range_flags(nums)
        parts = [p for p in (flags, pregled) if p]
        rows.append(EventRow(
            datum    = date_val,
            cat_path = LAB_PATH,
            comment  = " | ".join(parts),
            attrs    = nums,
        ))
    elif pregled:
        # Medical Visit event — Napomena = Pregled text
        rows.append(EventRow(
            datum    = date_val,
            cat_path = VISIT_PATH,
            comment  = '',
            attrs    = {"Napomena": pregled},
        ))
    else:
        skipped_empty += 1

print(f"  zdravstveni rows: {len(rows)} importable "
      f"({sum(1 for r in rows if r.cat_path == LAB_PATH)} Lab Results, "
      f"{sum(1 for r in rows if r.cat_path == VISIT_PATH)} Medical Visit)")
print(f"  skipped: {skipped_tip} non-zdravstveni, {skipped_nodate} no-date, {skipped_empty} empty")

# ── Build output workbook ─────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Events'

row = 1

# ── ATTRIBUTE LEGEND title ────────────────────────────────────────────────────
c = ws.cell(row, 1, 'ATTRIBUTE LEGEND:')
c.font = TITLE_FONT
ws.cell(row, 3).value = 'see Structure sheet for more details'
ws.cell(row, 3).font  = Font(italic=True, color='666666')
row += 1

# ── ATTRIBUTE LEGEND header ───────────────────────────────────────────────────
for ci, h in enumerate(['Col', 'Area', 'Category_Path', 'Attribute', 'Type', 'Unit'], 1):
    c = ws.cell(row, ci, h)
    c.fill = LEG_FILL; c.font = WHITE_FONT; c.border = BORDER; c.alignment = C_ALIGN
row += 1

# ── ATTRIBUTE LEGEND rows ─────────────────────────────────────────────────────
for i, (cat_path, attr_name, dtype, unit) in enumerate(ATTR_COLS):
    letter = col_letter(ATTR_COL_START + i)
    is_visit = cat_path == VISIT_PATH
    fill = VISIT_FILL if is_visit else LAB_FILL
    data = [letter, AREA, cat_path, attr_name, dtype, unit]
    for ci, v in enumerate(data, 1):
        c = ws.cell(row, ci, v if v else None)
        c.fill = fill; c.border = BORDER; c.alignment = L_ALIGN
    row += 1

row += 1  # blank separator before EVENT DATA

# ── EVENT DATA title ──────────────────────────────────────────────────────────
event_title_row = row
ws.cell(row, 1, 'EVENT DATA:').font = TITLE_FONT
row += 1

# ── EVENT DATA header ─────────────────────────────────────────────────────────
event_header_row = row
fixed_hdrs = ['event_id', 'Area', 'Category_Path', 'event_date', 'session_start',
              'created_at', 'User', 'leaf comment']
attr_hdrs  = [f"{attr} ({cat.split(' > ')[-1]})" for cat, attr, _, _ in ATTR_COLS]
for ci, h in enumerate(fixed_hdrs + attr_hdrs, 1):
    c = ws.cell(row, ci, h)
    c.fill = HEADER_FILL; c.font = WHITE_FONT; c.border = BORDER; c.alignment = C_ALIGN
row += 1

event_data_start = row

# ── DATA ROWS ─────────────────────────────────────────────────────────────────
for ev in rows:
    date_str = ev.datum.isoformat()  # "YYYY-MM-DD"

    # Fixed cols A-H
    fixed_vals = [
        None,           # A event_id (empty = new)
        AREA,           # B Area
        ev.cat_path,    # C Category_Path (NO area name)
        date_str,       # D event_date
        '08:00',        # E session_start
        '',             # F created_at
        '',             # G User (own event)
        ev.comment,     # H leaf comment
    ]
    for ci, v in enumerate(fixed_vals, 1):
        c = ws.cell(row, ci, v if v else None)
        c.border = BORDER; c.alignment = L_ALIGN

    # Attr cols I+
    for i, (cat_path, attr_name, dtype, _) in enumerate(ATTR_COLS):
        col_num = ATTR_COL_START + i
        val     = ev.attrs.get(attr_name)
        is_relevant = (cat_path == ev.cat_path)

        c = ws.cell(row, col_num, val)
        c.border    = BORDER
        c.fill      = (VISIT_FILL if cat_path == VISIT_PATH else LAB_FILL) if is_relevant else OWN_FILL
        c.alignment = R_ALIGN if (dtype == 'number' and val is not None) else L_ALIGN
        if dtype == 'number' and val is not None:
            c.number_format = '0.##'

    row += 1

event_data_end = row - 1

# ── Autofilter + freeze ───────────────────────────────────────────────────────
last_col = col_letter(ATTR_COL_START + len(ATTR_COLS) - 1)
ws.auto_filter.ref = f"A{event_header_row}:{last_col}{event_data_end}"
ws.freeze_panes    = ws.cell(event_data_start, ATTR_COL_START)

# ── Column widths ─────────────────────────────────────────────────────────────
widths = {'A':10, 'B':8, 'C':32, 'D':12, 'E':9, 'F':10, 'G':14, 'H':50}
for ltr, w in widths.items():
    ws.column_dimensions[ltr].width = w
for i in range(len(ATTR_COLS)):
    ws.column_dimensions[col_letter(ATTR_COL_START + i)].width = 14

ws.row_dimensions[event_header_row].height = 30

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
lab_count   = sum(1 for ev in rows if ev.cat_path == LAB_PATH)
visit_count = sum(1 for ev in rows if ev.cat_path == VISIT_PATH)
print(f"\nSaved: {OUTPUT}")
print(f"  Total events:  {len(rows)}")
print(f"  Lab Results:   {lab_count}")
print(f"  Medical Visit: {visit_count}")
print(f"  Attr columns:  {len(ATTR_COLS)} (I–{last_col})")
