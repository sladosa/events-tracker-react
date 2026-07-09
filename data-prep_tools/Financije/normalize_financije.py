# -*- coding: utf-8 -*-
"""
normalize_financije.py  (S107, 2026-07-09)
==========================================
Korak 1+2 pipeline-a iz FINANCIJE_MIGRACIJA.md (§8):
  - čita 'Financije 2026.xlsx' (koka EU, sasa EU, Za Sašu)
  - normalizira OBA računa u jedinstveni itemizirani model (§3)
  - primjenjuje odluke D1/D1a/D2/D4/D8/D9
  - klasificira Tip/Podtip (rules-first, §6) s pouzdanošću
  - generira REVIEW Excel s dependent dropdownima (INDIRECT, bez VBA)

NE piše ništa u bazu. Output je za ručni pregled (Saša/Koka) — tek odobreni
review ide u generiranje app-import datoteke (korak 4).

Model po redu (buduci event, leaf 'Transakcija'):
  Racun | event_date | Datum naplate | Smjer | Izvor | Uplata | Isplata |
  Stanje | Napomena | Tip | Podtip | Rate? | Broj rata | Status

Izvori po računu (D4):
  - Koka:  'koka EU' 1:1 (Kokin tekući → Izvor=Racun; Mastercard → Izvor=Mastercard,
           event_date = kol G datum kupovine, Datum naplate = kol C).
           'Za Sašu' Master/Zaba redovi se NE uvoze — služe samo kao izvor labela
           (match po datum+iznos → 'Što' puni Napomenu praznih redova).
  - Saša:  'sasa EU' 1:1 (lump 'Visa' → Tip=Transfer drži Stanje; ostalo Izvor=Racun;
           'Za Sašu' RF match → labela).
           'Za Sašu' Visa redovi = itemizirane kartične stavke (Izvor=Visa, uvoze se).
           'Za Sašu' Gotovina → Izvor=Cash, Tip=Razno (D2).

Pokretanje:
  PYTHONUTF8=1 C:/0_Sasa/events-tracker/venv/Scripts/python.exe normalize_financije.py
"""

import hashlib
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding='utf-8')

# ── Paths / config ─────────────────────────────────────────────────────────────
DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")
INPUT    = DATA_DIR / "Financije 2026.xlsx"
OUTPUT   = DATA_DIR / f"Financije_review_{datetime.now():%Y%m%d_%H%M}.xlsx"

RACUN_KOKA = 'Kokin tekući ZABA'
RACUN_SASA = 'Sašin tekući RF'

TODAY    = date.today()
MIN_DATE = date(2023, 1, 1)
MAX_DATE = date(2027, 12, 31)

# ── Taksonomija (§5 + D2) ──────────────────────────────────────────────────────
TAXONOMY: dict[str, list[str]] = {
    'Razno':          ['Odjeća/obuća', 'Pokloni', 'Kave/jelo vani', 'Temu', 'Taksi', 'Kino/Kazalište/Muzeji'],
    'Povrat':         ['Anja'],
    'auto C5':        ['gorivo', 'registracija', 'parking', 'popravci'],
    'Mirovina':       ['Saša', 'Koka'],
    'Ostavine':       ['Advokati'],
    'Zdravlje':       ['Medical', 'Lječnička komora', 'PP', 'PassSport', 'Sportski rekviziti'],
    'Putovanja':      ['Karte', 'Smještaj', 'Restoran'],
    'Informatika':    ['T-mobile', 'T-com', 'HP', 'Saša projekti', 'Disney', 'Sky', 'Prime', 'HBOmax',
                       'Youtube', 'AudibleKoka', 'AudibleSasa', 'Cloud backup', 'Microsoft'],
    'Domaćinstvo':    ['Struja', 'Voda', 'Holding (smeće)', 'Plin', 'Bankovni troškovi',
                       'Popravci i održavanje', 'Investicije', 'Povrat Nataša', 'Povrat Zoran'],
    'auto Lacetti':   ['gorivo', 'registracija', 'parking', 'popravci'],
    'Transfer':       [],
    'Ostali prihodi': [],
    'Namirnice':      ['Hrana i ostalo'],   # D2 (novi Tip)
    'N/A':            [],
}
TIP_LIST = list(TAXONOMY.keys())

# ── Klasifikacijska pravila (§6) ───────────────────────────────────────────────
# (regex na normaliziranu labelu, Tip, Podtip, pouzdanost, napomena-za-alternativu)
# Redoslijed = prioritet (prvi match pobjeđuje). Labela = Opis/Što bez "X/N" sufiksa.
VISOKA, SREDNJA, NISKA = 'VISOKA', 'SREDNJA', 'NISKA'

RULES: list[tuple[str, str, str, str, str]] = [
    # Namirnice (D2)
    (r'\b(konzum|spar|interspar|studenac|biberon|pekara|kruh|plodine|lidl|kaufland|tommy|billa|dubravica|dm)\b',
        'Namirnice', 'Hrana i ostalo', VISOKA, ''),
    # Mirovine (podtip po računu se dopuni poslije)
    (r'mirovinski doprinos', 'Mirovina', '', SREDNJA, 'doprinos — provjeri'),
    (r'\bmirovina\b',        'Mirovina', '', VISOKA, ''),
    # Informatika
    (r't-com',            'Informatika', 'T-com', VISOKA, ''),
    (r't-mobile',         'Informatika', 'T-mobile', VISOKA, ''),
    (r'\bhp\b',           'Informatika', 'HP', VISOKA, ''),
    (r'disney',           'Informatika', 'Disney', VISOKA, ''),
    (r'\bsky\b',          'Informatika', 'Sky', VISOKA, ''),
    (r'prime|amazon',     'Informatika', 'Prime', SREDNJA, 'Amazon možda roba, ne Prime'),
    (r'hbo',              'Informatika', 'HBOmax', VISOKA, ''),
    (r'youtube',          'Informatika', 'Youtube', VISOKA, ''),
    (r'audible\s*(ss|sasa|saša)', 'Informatika', 'AudibleSasa', VISOKA, ''),
    (r'audible\s*koka',   'Informatika', 'AudibleKoka', VISOKA, ''),
    (r'audible',          'Informatika', 'AudibleSasa', NISKA, 'AudibleKoka?'),
    (r'claude|anthropic|openai|cursor', 'Informatika', 'Saša projekti', VISOKA, ''),
    (r'microsoft',        'Informatika', 'Microsoft', VISOKA, ''),
    (r'icloud|backblaze|dropbox|google (one|storage)', 'Informatika', 'Cloud backup', SREDNJA, ''),
    (r'spotify|netflix|apple|steam', 'Informatika', '', NISKA, 'podtip ne postoji u taksonomiji'),
    # Domaćinstvo
    (r'hep|struja',       'Domaćinstvo', 'Struja', VISOKA, ''),
    (r'\bplin\b',         'Domaćinstvo', 'Plin', VISOKA, ''),
    (r'\bvoda\b|vodoopskrba', 'Domaćinstvo', 'Voda', VISOKA, ''),
    (r'nataša holding|natasa holding', 'Domaćinstvo', 'Povrat Nataša', VISOKA, ''),
    (r'zoran',            'Domaćinstvo', 'Povrat Zoran', VISOKA, ''),
    (r'holding',          'Domaćinstvo', 'Holding (smeće)', VISOKA, ''),
    (r'e-zaba|naknada',   'Domaćinstvo', 'Bankovni troškovi', VISOKA, ''),
    (r'pričuva|pricuva',  'Domaćinstvo', '', SREDNJA, 'pričuva — podtip?'),
    (r'popravak|majstor|vodoinstalater', 'Domaćinstvo', 'Popravci i održavanje', SREDNJA, ''),
    (r'ikea|emmezeta|lesnina', 'Domaćinstvo', 'Investicije', SREDNJA, 'namještaj — Investicije?'),
    # Auto (D8: default C5)
    (r'gorivo|\bina\b|benzin|petrol|tifon|crodux', 'auto C5', 'gorivo', SREDNJA, 'D8 default — Lacetti?'),
    (r'parking|parkiranje|park&|garaža', 'auto C5', 'parking', SREDNJA, 'D8 default — Lacetti?'),
    (r'registracija|tehnički pregled', 'auto C5', 'registracija', SREDNJA, 'D8 default — Lacetti?'),
    (r'carglass|vulkanizer|autoservis', 'auto C5', 'popravci', SREDNJA, 'D8 default — Lacetti?'),
    (r'\bhak\b',          'auto C5', '', NISKA, 'HAK — koji auto/što?'),
    # Razno
    (r'temu',             'Razno', 'Temu', VISOKA, ''),
    (r'taxi|taksi|bolt|uber', 'Razno', 'Taksi', VISOKA, ''),
    (r'kava|kave|caffe|restoran|pizzeria|mcdonald|kfc|dostava', 'Razno', 'Kave/jelo vani', SREDNJA, ''),
    (r'kino|kazalište|kazaliste|muzej', 'Razno', 'Kino/Kazalište/Muzeji', VISOKA, ''),
    (r'pokloni?|dar\b',   'Razno', 'Pokloni', SREDNJA, ''),
    (r'odjeća|odjeca|obuća|obuca|h&m|zara|c&a|deichmann', 'Razno', 'Odjeća/obuća', SREDNJA, ''),
    # Zdravlje
    (r'ljekarn|lijek|pharmalog|apoteka|dopunsko|ordinacija|doktor|poliklinika|bolnica|dental|zubar',
        'Zdravlje', 'Medical', VISOKA, ''),
    (r'passsport|passport', 'Zdravlje', 'PassSport', VISOKA, ''),
    (r'\bpp\b|posmrtna',  'Zdravlje', 'PP', VISOKA, ''),
    (r'lječnička komora|ljecnicka', 'Zdravlje', 'Lječnička komora', VISOKA, ''),
    (r'multisport|kreatin|proteini|decathlon|sportski', 'Zdravlje', 'Sportski rekviziti', SREDNJA, ''),
    # Putovanja
    (r'airbnb|booking|hotel|hostel', 'Putovanja', 'Smještaj', VISOKA, ''),
    (r'ryanair|croatia airlines|easyjet|wizz|aviokart|avionske|flixbus|vlak karta',
        'Putovanja', 'Karte', VISOKA, ''),
    (r'amsterdam|putovanje', 'Putovanja', '', SREDNJA, 'trip trošak — podtip?'),
    # Povrat / Ostavine
    (r'\banja\b',         'Povrat', 'Anja', VISOKA, ''),
    (r'advokat|odvjetni', 'Ostavine', 'Advokati', VISOKA, ''),
    # Transferi / financijski promet (nisu trošak — D3)
    (r'^\s*(pbz\s+)?visa\b|^\s*master(card)?\b', 'Transfer', '', VISOKA, 'lump/kartica skidanje'),
    (r'\bkeks\b|revolut|electrocoin|dionice|aircash', 'Transfer', '', SREDNJA, 'prijenos/investicija'),
    (r'gotovina|cash|bankomat|atm', 'Transfer', '', VISOKA, 'podizanje gotovine (D2)'),
    (r'\bašo\b|\baso\b',  'Transfer', '', SREDNJA, 'transfer osobi?'),
    (r'saša uplata|koka uplata|prijenos|transfer', 'Transfer', '', SREDNJA, ''),
    # Osiguranja / porezi — nema Tipa u taksonomiji → review
    (r'životno|zivotno|allianz|croatia osiguranje|generali|police', 'N/A', '', NISKA, 'osiguranje — nema Tipa u taksonomiji'),
    (r'porez|prirez|\bapn\b', 'N/A', '', NISKA, 'porez — nema Tipa u taksonomiji'),
]

# Prihodi (Smjer=Uplata) koji nisu mirovina/povrat
INCOME_RULES: list[tuple[str, str, str, str, str]] = [
    (r'plaća|placa|regres|božićnica|bozicnica|kamate|dividenda', 'Ostali prihodi', '', VISOKA, ''),
]

RE_RATE = re.compile(r'(.*?)\s*(\d{1,3})\s*/\s*(\d{1,3})\s*$')


def norm_label(s: str | None) -> str:
    return (str(s).strip() if s not in (None, '') else '')


def parse_rate(label: str) -> tuple[str, int | None, int | None]:
    """'Konzum 2/6' → ('Konzum', 2, 6); bez ratea → (label, None, None)."""
    m = RE_RATE.match(label)
    if not m:
        return label, None, None
    base = m.group(1).strip() or label
    return base, int(m.group(2)), int(m.group(3))


def classify(label: str, smjer: str, racun: str) -> tuple[str, str, str, str]:
    """→ (tip, podtip, pouzdanost, alternativa)"""
    if not label:
        return 'N/A', '', 'NEMA', 'nema opisa — ručno ili ostaje N/A'
    low = label.lower()

    if smjer == 'Uplata':
        for rx, tip, pod, conf, alt in INCOME_RULES:
            if re.search(rx, low):
                return tip, pod, conf, alt

    for rx, tip, pod, conf, alt in RULES:
        if re.search(rx, low):
            if tip == 'Mirovina' and not pod:
                pod = 'Saša' if racun == RACUN_SASA else 'Koka'
            return tip, pod, conf, alt

    if smjer == 'Uplata':
        return 'Ostali prihodi', '', 'NISKA', 'uplata bez pravila — provjeri'
    return 'N/A', '', 'NEMA', 'nepoznata labela — ručno'


def source_key(racun: str, d: date, seq: int, iznos, opis: str) -> str:
    raw = f'{racun}|{d.isoformat()}|{seq}|{iznos}|{opis}'
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()[:12]


def to_date(val) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if val in (None, ''):
        return None
    m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d+)\.?$', str(val).strip())
    if m:
        dd, mm, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yy < 100:
            yy += 2000
        try:
            return date(yy, mm, dd)
        except ValueError:
            return None
    return None


def num(val) -> float | None:
    if val in (None, ''):
        return None
    try:
        f = round(float(val), 2)
        return f if f != 0 else None
    except (TypeError, ValueError):
        return None


# ── Normalizirani red ──────────────────────────────────────────────────────────

class Row(dict):
    """Jedan normalizirani event-kandidat."""


def make_row(*, racun, event_date, datum_naplate, smjer, izvor, uplata, isplata,
             stanje, napomena, status, src, problem='', label_src='') -> Row:
    base, rata_x, rata_n = parse_rate(napomena)
    tip, podtip, conf, alt = classify(base, smjer, racun)
    if rata_n and alt == '':
        alt = f'rata {rata_x}/{rata_n}'
    return Row(
        racun=racun, event_date=event_date, datum_naplate=datum_naplate,
        smjer=smjer, izvor=izvor, uplata=uplata, isplata=isplata, stanje=stanje,
        napomena=napomena, tip=tip, podtip=podtip, conf=conf, alt=alt,
        rate=bool(rata_n), broj_rata=rata_n, status=status,
        src=src, problem=problem, label_src=label_src,
    )


# ── Učitavanje i normalizacija ─────────────────────────────────────────────────

def main() -> None:
    print(f'Čitam {INPUT} ...')
    wb = openpyxl.load_workbook(INPUT, data_only=True)
    problems: list[list] = []   # [sheet, row, opis problema, akcija]

    # ---- Za Sašu (labele + Sašine Visa/Cash stavke) ----
    zs = wb['Za Sašu']
    zs_rows: list[dict] = []
    for r in range(2, zs.max_row + 1):
        d     = to_date(zs.cell(r, 1).value)
        nacin = norm_label(zs.cell(r, 2).value)
        sto   = norm_label(zs.cell(r, 3).value)
        eu    = num(zs.cell(r, 4).value)
        if d is None and not nacin and not sto and eu is None:
            continue
        if nacin.lower() == 'visa':
            nacin = 'Visa'          # tipfeler 'vISA'
        if d is None:
            problems.append(['Za Sašu', r, f'neparsiv datum {zs.cell(r,1).value!r}', 'red PRESKOČEN'])
            continue
        zs_rows.append({'row': r, 'datum': d, 'nacin': nacin, 'sto': sto, 'eu': eu, 'used': False})

    # label lookup: (nacin, datum, iznos) → queue Za Sašu redova
    zs_lookup: dict[tuple, list[dict]] = defaultdict(list)
    for z in zs_rows:
        if z['eu'] is not None:
            zs_lookup[(z['nacin'], z['datum'], round(z['eu'], 2))].append(z)

    def take_label(nacin: str, d: date | None, iznos) -> tuple[str, str]:
        """Nađi (i potroši) Za Sašu red istog načina+iznosa, datum ±2 dana → (labela, info)."""
        if d is None or iznos is None:
            return '', ''
        for delta in (0, 1, -1, 2, -2):
            q = zs_lookup.get((nacin, d + timedelta(days=delta), round(iznos, 2)), [])
            for z in q:
                if not z['used'] and z['sto']:
                    z['used'] = True
                    return z['sto'], f'Za Sašu:{z["row"]}'
        return '', ''

    rows: list[Row] = []

    # ---- koka EU ----
    ws = wb['koka EU']
    seq_per_day: Counter = Counter()
    last_valid = MIN_DATE
    for r in range(2, ws.max_row + 1):
        tip_rac = norm_label(ws.cell(r, 1).value)
        opis    = norm_label(ws.cell(r, 2).value)
        d       = to_date(ws.cell(r, 3).value)
        uplata  = num(ws.cell(r, 4).value)
        isplata = num(ws.cell(r, 5).value)
        stanje  = ws.cell(r, 6).value
        kupnja  = to_date(ws.cell(r, 7).value)

        if not tip_rac and not opis and d is None and uplata is None and isplata is None:
            continue

        problem = ''
        if not tip_rac:
            problem = 'prazan Tip računa — pretpostavljen Kokin tekući'
            problems.append(['koka EU', r, 'prazan Tip računa', 'pretpostavljen Kokin tekući'])
            tip_rac = 'Kokin tekući'

        if d is None or not (MIN_DATE <= d <= MAX_DATE):
            problems.append(['koka EU', r, f'datum {ws.cell(r,3).value!r} izvan raspona/neparsiv', f'fallback {last_valid}'])
            problem = (problem + '; ' if problem else '') + f'datum → fallback {last_valid}'
            d = last_valid
        else:
            last_valid = d

        smjer = 'Uplata' if (uplata and not isplata) else 'Isplata' if isplata else 'PROVJERI'
        if smjer == 'PROVJERI':
            problems.append(['koka EU', r, 'ni Uplata ni Isplata (ili oboje)', 'Smjer=PROVJERI'])

        if tip_rac == 'Mastercard':
            if kupnja is None or not (MIN_DATE - timedelta(days=90) <= kupnja <= MAX_DATE):
                if kupnja is not None:
                    problems.append(['koka EU', r, f'datum kupovine (G) {kupnja} sumnjiv', 'event_date = datum naplate'])
                kupnja = None
            event_date    = kupnja or d          # D1: kupovina; fallback naplata
            datum_naplate = d
            izvor         = 'Mastercard'
            label, lsrc = (opis, '') if opis else take_label('Master', kupnja or d, isplata or uplata)
        else:
            event_date    = d
            datum_naplate = None
            izvor         = 'Racun'
            label, lsrc = (opis, '') if opis else take_label('Zaba', d, isplata or uplata)

        seq_key = (RACUN_KOKA, event_date)
        seq_per_day[seq_key] += 1
        status = 'Planiran' if event_date > TODAY else 'Izvrsen'

        row = make_row(
            racun=RACUN_KOKA, event_date=event_date, datum_naplate=datum_naplate,
            smjer=smjer, izvor=izvor, uplata=uplata, isplata=isplata,
            stanje=num(stanje), napomena=label, status=status,
            src=f'koka EU:{r}', problem=problem, label_src=lsrc,
        )
        row['skey'] = source_key(RACUN_KOKA, event_date, seq_per_day[seq_key], isplata or uplata or 0, label)
        rows.append(row)

    # ---- sasa EU ----
    ws = wb['sasa EU']
    last_valid = MIN_DATE
    for r in range(2, ws.max_row + 1):
        opis    = norm_label(ws.cell(r, 2).value)
        d       = to_date(ws.cell(r, 3).value)
        uplata  = num(ws.cell(r, 4).value)
        isplata = num(ws.cell(r, 5).value)
        stanje  = ws.cell(r, 6).value

        if not opis and d is None and uplata is None and isplata is None:
            continue

        problem = ''
        if d is None or not (MIN_DATE <= d <= MAX_DATE):
            problems.append(['sasa EU', r, f'datum {ws.cell(r,3).value!r} izvan raspona/neparsiv', f'fallback {last_valid}'])
            problem = f'datum → fallback {last_valid}'
            d = last_valid
        else:
            last_valid = d

        smjer = 'Uplata' if (uplata and not isplata) else 'Isplata' if isplata else 'PROVJERI'
        if smjer == 'PROVJERI':
            problems.append(['sasa EU', r, 'ni Uplata ni Isplata (ili oboje)', 'Smjer=PROVJERI'])

        label, lsrc = (opis, '') if opis else take_label('RF', d, isplata or uplata)

        seq_key = (RACUN_SASA, d)
        seq_per_day[seq_key] += 1
        status = 'Planiran' if d > TODAY else 'Izvrsen'

        row = make_row(
            racun=RACUN_SASA, event_date=d, datum_naplate=None,
            smjer=smjer, izvor='Racun', uplata=uplata, isplata=isplata,
            stanje=num(stanje), napomena=label, status=status,
            src=f'sasa EU:{r}', problem=problem, label_src=lsrc,
        )
        row['skey'] = source_key(RACUN_SASA, d, seq_per_day[seq_key], isplata or uplata or 0, label)
        rows.append(row)

    # ---- Za Sašu → Sašine itemizirane stavke (Visa + Gotovina) ----
    # 'Visa'/'vISA' = kartične stavke (uvoze se); red čiji je 'Što' sam lump
    # ('Visa'/'Master') se preskače — lump već postoji u sasa EU.
    for z in zs_rows:
        if z['nacin'] not in ('Visa', 'Gotovina'):
            continue                      # Master/Zaba/RF = samo izvor labela (D4)
        if z['sto'].lower() in ('visa', 'master', 'mastercard'):
            continue                      # lump duplikat
        d = z['datum']
        status = 'Planiran' if d > TODAY else 'Izvrsen'
        izvor  = 'Cash' if z['nacin'] == 'Gotovina' else 'Visa'

        seq_key = (RACUN_SASA, d)
        seq_per_day[seq_key] += 1

        row = make_row(
            racun=RACUN_SASA, event_date=d, datum_naplate=None,
            smjer='Isplata', izvor=izvor, uplata=None, isplata=z['eu'],
            stanje=None, napomena=z['sto'], status=status,
            src=f'Za Sašu:{z["row"]}', problem='', label_src='',
        )
        if izvor == 'Cash':               # D2: cash plaćanja → Razno
            row['tip'], row['podtip'] = 'Razno', ''
            row['conf'], row['alt'] = 'SREDNJA', 'cash plaćanje (D2 → Razno)'
        row['skey'] = source_key(RACUN_SASA, d, seq_per_day[seq_key], z['eu'] or 0, z['sto'])
        rows.append(row)
        z['used'] = True

    # ---- Nepodudareni Za Sašu Master/Zaba/RF redovi → Problemi ----
    for z in zs_rows:
        if z['used'] or z['nacin'] not in ('Master', 'Zaba', 'RF'):
            continue
        if z['sto'].lower() in ('visa', 'master', 'mastercard'):
            continue
        problems.append(['Za Sašu', z['row'],
                         f"{z['nacin']} '{z['sto']}' {z['eu']}€ {z['datum']} bez para u izvodu",
                         'labela neiskorištena (možda budući/rate red)'])

    rows.sort(key=lambda x: (x['racun'], x['event_date'], x['src']))
    write_review(rows, problems)


# ── Review Excel ───────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill('solid', fgColor='4472C4')
ORIG_FILL  = PatternFill('solid', fgColor='F2F2F2')
EDIT_FILL  = PatternFill('solid', fgColor='FFF2CC')
WHITE_BOLD = Font(color='FFFFFF', bold=True)
THIN       = Side(style='thin')
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RED_FILL   = PatternFill('solid', fgColor='FFC7CE')
YEL_FILL   = PatternFill('solid', fgColor='FFEB9C')

COLS = [
    ('Racun', 18), ('event_date', 11), ('Datum naplate', 12), ('Smjer', 9),
    ('Izvor', 11), ('Uplata', 9), ('Isplata', 9), ('Stanje', 10),
    ('Napomena', 26), ('Tip', 14), ('Podtip', 18), ('Pouzdanost', 11),
    ('Alternativa / nap.', 26), ('Rate?', 6), ('Broj rata', 8), ('Status', 9),
    ('Izvor reda', 12), ('Labela iz', 11), ('Problem', 22), ('source_key', 14),
]
COL_TIP, COL_PODTIP, COL_CONF = 10, 11, 12   # 1-based indeksi


def sanitize_name(s: str) -> str:
    s = s.replace('ć', 'c').replace('č', 'c').replace('š', 's').replace('ž', 'z').replace('đ', 'd')
    s = s.replace('Ć', 'C').replace('Č', 'C').replace('Š', 'S').replace('Ž', 'Z').replace('Đ', 'D')
    return re.sub(r'[^A-Za-z0-9_]', '_', s)


def write_review(rows: list[Row], problems: list[list]) -> None:
    out = openpyxl.Workbook()

    # ---- Liste (hidden) — named ranges za dependent dropdown ----
    lst = out.active
    lst.title = 'Liste'
    lst.cell(1, 1, 'Tip')
    for i, tip in enumerate(TIP_LIST, 2):
        lst.cell(i, 1, tip)
    out.defined_names.add(_dn(out, 'TipList', f'Liste!$A$2:$A${1 + len(TIP_LIST)}'))

    col = 2
    for tip, podtips in TAXONOMY.items():
        values = podtips if podtips else ['—']
        lst.cell(1, col, tip)
        for i, p in enumerate(values, 2):
            lst.cell(i, col, p)
        ltr = get_column_letter(col)
        out.defined_names.add(_dn(out, f'Tip_{sanitize_name(tip)}', f'Liste!${ltr}$2:${ltr}${1 + len(values)}'))
        col += 1
    lst.sheet_state = 'veryHidden'

    # ---- Review ----
    ws = out.create_sheet('Review')
    for c, (h, w) in enumerate(COLS, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
        ws.column_dimensions[get_column_letter(c)].width = w

    for r_i, row in enumerate(rows, 2):
        vals = [
            row['racun'], row['event_date'], row['datum_naplate'], row['smjer'],
            row['izvor'], row['uplata'], row['isplata'], row['stanje'],
            row['napomena'], row['tip'], row['podtip'], row['conf'],
            row['alt'], 'DA' if row['rate'] else '', row['broj_rata'], row['status'],
            row['src'], row['label_src'], row['problem'], row['skey'],
        ]
        for c_i, v in enumerate(vals, 1):
            cell = ws.cell(r_i, c_i, v)
            cell.border = BORDER
            if c_i in (COL_TIP, COL_PODTIP):
                cell.fill = EDIT_FILL
            elif c_i >= 17:
                cell.fill = ORIG_FILL
            if isinstance(v, date):
                cell.number_format = 'YYYY-MM-DD'
            if c_i in (6, 7, 8):
                cell.number_format = '#,##0.00'

    last = len(rows) + 1
    tip_col_ltr    = get_column_letter(COL_TIP)
    podtip_col_ltr = get_column_letter(COL_PODTIP)

    # Dropdown Tip (named range)
    dv_tip = DataValidation(type='list', formula1='=TipList', allowBlank=True, showErrorMessage=False)
    ws.add_data_validation(dv_tip)
    dv_tip.add(f'{tip_col_ltr}2:{tip_col_ltr}{last}')

    # Dependent dropdown Podtip — INDIRECT("Tip_"& sanitized(Tip)).
    # VAŽNO: Excel DV formula limit je 255 znakova → SUBSTITUTE lanac pokriva SAMO
    # znakove koji se stvarno pojavljuju u imenima Tipova: space, '/', 'ć'.
    # (Ako se ikad doda Tip s drugim specijalnim znakom, proširi i sanitize_name i ovo.)
    sub = f'{tip_col_ltr}2'
    for a, b in [(' ', '_'), ('/', '_'), ('ć', 'c')]:
        sub = f'SUBSTITUTE({sub},"{a}","{b}")'
    dv_pod = DataValidation(type='list', formula1=f'INDIRECT("Tip_"&{sub})', allowBlank=True, showErrorMessage=False)
    ws.add_data_validation(dv_pod)
    dv_pod.add(f'{podtip_col_ltr}2:{podtip_col_ltr}{last}')

    # CF: Podtip ne pripada Tipu → crveno; prazan Tip/N-A + NEMA pouzdanost → žuto
    ws.conditional_formatting.add(
        f'{podtip_col_ltr}2:{podtip_col_ltr}{last}',
        FormulaRule(
            formula=[f'AND({podtip_col_ltr}2<>"",{podtip_col_ltr}2<>"—",ISERROR(MATCH({podtip_col_ltr}2,INDIRECT("Tip_"&{sub}),0)))'],
            fill=RED_FILL,
        ),
    )
    ws.conditional_formatting.add(
        f'{tip_col_ltr}2:{tip_col_ltr}{last}',
        FormulaRule(formula=[f'OR({tip_col_ltr}2="",{tip_col_ltr}2="N/A")'], fill=YEL_FILL),
    )

    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}{last}'
    ws.freeze_panes = 'A2'

    # ---- Problemi ----
    pws = out.create_sheet('Problemi')
    for c, h in enumerate(['Sheet', 'Red', 'Problem', 'Akcija'], 1):
        cell = pws.cell(1, c, h)
        cell.fill, cell.font = HDR_FILL, WHITE_BOLD
    for r_i, p in enumerate(problems, 2):
        for c_i, v in enumerate(p, 1):
            pws.cell(r_i, c_i, v)
    for c, w in enumerate([10, 6, 60, 40], 1):
        pws.column_dimensions[get_column_letter(c)].width = w

    # ---- Statistika ----
    sws = out.create_sheet('Statistika')
    tip_cnt  = Counter((r['tip'] or '(prazno)') for r in rows)
    conf_cnt = Counter(r['conf'] for r in rows)
    racun_cnt = Counter(r['racun'] for r in rows)
    sws.cell(1, 1, 'REVIEW STATISTIKA').font = Font(bold=True, size=12)
    r_i = 3
    sws.cell(r_i, 1, 'Po računu:').font = Font(bold=True); r_i += 1
    for k, v in racun_cnt.most_common():
        sws.cell(r_i, 1, k); sws.cell(r_i, 2, v); r_i += 1
    r_i += 1
    sws.cell(r_i, 1, 'Po Tipu:').font = Font(bold=True); r_i += 1
    for k, v in tip_cnt.most_common():
        sws.cell(r_i, 1, k); sws.cell(r_i, 2, v); r_i += 1
    r_i += 1
    sws.cell(r_i, 1, 'Po pouzdanosti:').font = Font(bold=True); r_i += 1
    for k, v in conf_cnt.most_common():
        sws.cell(r_i, 1, k); sws.cell(r_i, 2, v); r_i += 1
    sws.column_dimensions['A'].width = 28

    # ---- Pomoć ----
    hws = out.create_sheet('Pomoć')
    help_lines = [
        'FINANCIJE REVIEW — upute (S107)',
        '',
        'Ovaj file je REVIEW korak migracije — NIŠTA još nije u bazi.',
        'Pregledaj/ispravi Tip i Podtip (žute kolone), ostalo NE diraj.',
        '',
        '• Tip: dropdown (14 opcija, uklj. novi "Namirnice" — D2).',
        '• Podtip: dropdown se prilagođava odabranom Tipu (bez makroa).',
        '  Ako promijeniš Tip, stari Podtip se NE briše sam — oboji se CRVENO',
        '  ako ne pripada novom Tipu. "—" znači: Tip nema podtipove.',
        '• Pouzdanost: VISOKA/SREDNJA/NISKA/NEMA — filtriraj NISKA+NEMA za ručni rad.',
        '  ŽUTO na Tip = prazan ili N/A → treba odluka.',
        '• Rate "X/N" su parsirane u Rate?/Broj rata; svaka rata = svoj red.',
        '• auto C5 vs Lacetti: sve gorivo/parking je default C5 (D8) — ispravi Lacetti ručno.',
        '• Izvor=Racun redovi drže Stanje; Tip=Transfer se isključuje iz analize potrošnje (D3).',
        '• Kolone "Izvor reda"/"Labela iz": odakle red dolazi (sheet:red u Kokinom Excelu).',
        '• Sheet "Problemi": datumi izvan raspona, prazni smjerovi, nepodudarene labele.',
        '',
        'VAŽNO — velika rupa u podacima: ~82% Mastercard redova (2023–2025-06) nema',
        'nikakav opis pa su Tip=N/A, Pouzdanost=NEMA. Odluka: ostaviti N/A (iznosi i',
        'dalje ulaze u saldo) ili ručno klasificirati odabrane velike iznose.',
    ]
    for i, line in enumerate(help_lines, 1):
        hws.cell(i, 1, line)
        if i == 1:
            hws.cell(i, 1).font = Font(bold=True, size=13)
    hws.column_dimensions['A'].width = 90

    out.save(OUTPUT)

    # ---- Konzolni sažetak ----
    print(f'\n✔ Review Excel: {OUTPUT}')
    print(f'  Redova ukupno: {len(rows)}')
    for k, v in Counter(r["racun"] for r in rows).items():
        print(f'    {k}: {v}')
    print('  Pouzdanost:', dict(Counter(r['conf'] for r in rows)))
    print('  Top Tip:', Counter(r['tip'] for r in rows).most_common(8))
    print(f'  Problema: {len(problems)} (sheet Problemi)')


def _dn(wb, name: str, ref: str):
    from openpyxl.workbook.defined_name import DefinedName
    return DefinedName(name, attr_text=ref)


if __name__ == '__main__':
    main()
