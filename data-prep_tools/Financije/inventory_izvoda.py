# -*- coding: utf-8 -*-
"""
inventory_izvoda.py  (S107d, 2026-07-13)
========================================
Sredi kaos u `data-prep_data/Financije/izvodi/` (generička download imena,
duplikati, pomiješani formati) i izvuci SVE transakcije u jedan Excel:

  1. Skenira izvodi/ rekurzivno (uklj. dodatno_stiglo/), md5 dedup po SADRŽAJU
     → duplikati se sele u izvodi/duplikati/ (ništa se ne briše!)
  2. Klasificira PDF po SADRŽAJU (ne po imenu):
       ZABA     — izvadak tekućeg računa ("Jedinstveni izvadak građana")
       MC       — ZABA Mastercard izvod kartice ("Obavijest o učinjenim troškovima")
       PBZVISA  — PBZ Card Visa Gold mjesečni račun (SPECIFIKACIJA PROMETA)
       (bez tekst-sloja / nepoznato → ostaje na mjestu, samo u Manifest)
  3. Parsira transakcije (parseri iz enrich_from_izvoda.py), period = mjesec
     s najviše transakcija → preimenuje u PREFIX_YYYY-MM.pdf i premjesti u
     izvodi/Analizirani_izvodi/
  4. Piše izvodi/Izvodi_transakcije.xlsx:
       'Transakcije' — sve transakcije (enrich_from_izvoda.py čita OVO umjesto PDF-ova)
       'Manifest'    — svaki file: novo ime, original, tip, period, #tx, status
     ('Nematchano' sheet dopiše enrich_from_izvoda.py nakon matcha na Review)

Idempotentno — slobodno pokretati više puta; već analizirani fajlovi se samo
re-parsiraju u svježi Excel. Pokretanje:
  Financije\\run.bat inventory_izvoda.py          → sve
  Financije\\run.bat inventory_izvoda.py --dry    → samo report, bez premještanja/pisanja
"""

import hashlib
import re
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
import pdfplumber

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, str(Path(__file__).parent))

from enrich_from_izvoda import (BORDER, HDR_FILL, IZVODI_DIR, SOURCE_TYPES,
                                TX_XLSX, WHITE_BOLD)

ANALIZIRANI = IZVODI_DIR / 'Analizirani_izvodi'
DUPLIKATI   = IZVODI_DIR / 'duplikati'

RE_GOOD_NAME = re.compile(r'^(ZABA|MC|PBZVISA|RF)_\d{4}-\d{2}[a-z]?\.pdf$')


def classify(path: Path) -> str:
    """Tip izvoda iz SADRŽAJA 1. stranice (imena fajlova su generička/nepouzdana).
    Bez tekst-sloja → OCR vrha stranice (RF izvodi su vektorske krivulje)."""
    try:
        with pdfplumber.open(path) as pdf:
            txt = pdf.pages[0].extract_text() or ''
    except Exception as e:
        print(f'  ✗ {path.name}: PDF se ne da otvoriti ({e})')
        return 'ERROR'
    if len(txt.strip()) < 40:
        try:
            from rf_ocr import ocr_page1_head
            head = ocr_page1_head(path)
        except Exception as e:
            print(f'  ✗ {path.name}: OCR klasifikacija pala ({e})')
            return 'NOTEXT'
        if 'RAIFFEISEN' in head and 'IZVADAK' in head:
            return 'RF'
        return 'NOTEXT'
    if 'REFERENCIJA DATUM OPIS TRANSAKCIJE' in txt:
        return 'MC'
    if 'PBZ Card' in txt:
        return 'PBZVISA'
    if 'IZVADAK' in txt and 'Zagreba' in txt:
        return 'ZABA'
    return 'UNKNOWN'


def period_from_dates(dates: list, tip: str) -> str:
    """Period za ime fajla. RF: mjesec PRVE transakcije (RBA izvadak ide od
    sredine do sredine mjeseca → mode bi dva uzastopna izvatka znao staviti u
    isti mjesec). Ostali: mjesec s najviše transakcija (robusno na PBZ RATA
    retke sa starim datumima kupovine)."""
    if tip == 'RF':
        return f'{min(dates):%Y-%m}'
    months = Counter(f'{d:%Y-%m}' for d in dates)
    return months.most_common(1)[0][0]


def coverage_period(txs: list[dict], tip: str) -> str:
    return period_from_dates([t['date'] for t in txs], tip)


def month_gaps(periods: list[str]) -> list[str]:
    """Mjeseci koji fale između min i max perioda (za report pokrivenosti)."""
    if len(periods) < 2:
        return []
    ys, ms = int(periods[0][:4]), int(periods[0][5:7])
    ye, me = int(periods[-1][:4]), int(periods[-1][5:7])
    have = set(periods)
    gaps, y, m = [], ys, ms
    while (y, m) < (ye, me):
        m += 1
        if m > 12:
            y, m = y + 1, 1
        p = f'{y:04d}-{m:02d}'
        if p not in have:
            gaps.append(p)
    return gaps


def unique_target(tip: str, period: str, md5: str, taken: dict[Path, str],
                  primary: Path) -> Path | None:
    """PREFIX_YYYY-MM.pdf; kolizija s DRUGIM sadržajem → sufiks b, c...
    Vraća None ako isti sadržaj već postoji na targetu (source = duplikat)."""
    for suffix in [''] + [chr(c) for c in range(ord('b'), ord('z') + 1)]:
        target = ANALIZIRANI / f'{tip}_{period}{suffix}.pdf'
        if target == primary:          # file je već na svom mjestu
            return target
        existing = taken.get(target)
        if existing is None and not target.exists():
            return target
        existing_md5 = existing or hashlib.md5(target.read_bytes()).hexdigest()
        if existing_md5 == md5:
            return None
    raise RuntimeError(f'Previše kolizija za {tip}_{period}')


def load_cache() -> tuple[dict, dict]:
    """Prethodni Izvodi_transakcije.xlsx kao keš: md5[:8] → manifest info +
    transakcije po fajlu. Skupi parseri (RF = OCR, ~25 s/str.) se tako
    plaćaju samo jednom po fajlu; rerun je jeftin."""
    if not TX_XLSX.exists():
        return {}, {}
    wb = openpyxl.load_workbook(TX_XLSX, read_only=True)
    tx_by_file: dict[str, list[tuple]] = defaultdict(list)
    if 'Transakcije' in wb.sheetnames:
        for r in wb['Transakcije'].iter_rows(min_row=2, values_only=True):
            d = r[0].date() if hasattr(r[0], 'date') else r[0]
            tx_by_file[str(r[8])].append((d,) + tuple(r[1:10]))
    man_by_md5: dict[str, dict] = {}
    if 'Manifest' in wb.sheetnames:
        for r in wb['Manifest'].iter_rows(min_row=2, values_only=True):
            if r[4] and r[4] > 0:   # samo fajlovi s parsiranim transakcijama
                man_by_md5[str(r[5])] = {'file': str(r[0]), 'tip': str(r[2]),
                                         'period': str(r[3]), 'ntx': int(r[4])}
    wb.close()
    return man_by_md5, tx_by_file


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    # --reparse a,b,c → fajlovi čije keširano ime sadrži neki od substringova
    # se parsiraju ISPOČETKA (npr. nakon poboljšanja OCR-a): --reparse RF_2026-06
    reparse: list[str] = []
    if '--reparse' in args:
        idx = args.index('--reparse')
        if idx + 1 >= len(args):
            sys.exit('✗ --reparse traži argument (substring imena, zarezom odvojeni)')
        reparse = [s.strip().lower() for s in args[idx + 1].split(',') if s.strip()]
    print(f'Izvodi: {IZVODI_DIR}{"  [DRY RUN]" if dry else ""}'
          f'{"  [REPARSE: " + ",".join(reparse) + "]" if reparse else ""}\n')
    cache_man, cache_tx = load_cache()

    # 1. Skupi PDF-ove (bez duplikati/) + md5 dedup po sadržaju
    pdfs = [p for p in IZVODI_DIR.rglob('*.pdf') if DUPLIKATI not in p.parents]
    if not pdfs:
        sys.exit(f'✗ Nema PDF-ova u {IZVODI_DIR}')
    by_hash: dict[str, list[Path]] = defaultdict(list)
    for p in pdfs:
        by_hash[hashlib.md5(p.read_bytes()).hexdigest()].append(p)

    def primary_key(p: Path) -> tuple:
        loc = 0 if p.parent == ANALIZIRANI else (1 if p.parent == IZVODI_DIR else 2)
        return (loc, 0 if RE_GOOD_NAME.match(p.name) else 1, p.name)

    manifest: list[dict] = []          # {file, original, lokacija, tip, period, ntx, md5, status}
    all_rows: list[tuple] = []         # retci za Transakcije sheet
    duplicates: list[tuple[Path, str]] = []
    taken: dict[Path, str] = {}        # planirani targeti (za dry i kolizije u istom runu)
    content_sigs: dict[tuple, str] = {}  # potpis transakcija → file (isti izvod, drugi bajtovi)

    for md5, group in sorted(by_hash.items(), key=lambda kv: str(kv[1][0])):
        group.sort(key=primary_key)
        primary, dups = group[0], group[1:]
        duplicates.extend((d, md5) for d in dups)
        rel = str(primary.relative_to(IZVODI_DIR))

        cached = cache_man.get(md5[:8])
        if cached and any(s in cached['file'].lower() for s in reparse):
            cached = None            # forsiraj svježi parse
        if cached and cached['file'] in cache_tx:
            rows = cache_tx[cached['file']]
            tip = cached['tip']
            sig = (tip, tuple(sorted((str(r[0]), r[3], r[2], r[1]) for r in rows)))
            if sig in content_sigs:
                duplicates.append((primary, md5))
                print(f'  ⚠ {rel}: iste transakcije kao {content_sigs[sig]} '
                      f'(drugi bajtovi, isti izvod) → duplikat')
                continue
            content_sigs[sig] = cached['file']
            # period/ime se RE-računa iz keširanih datuma — promjena pravila
            # imenovanja se tako primijeni bez ponovnog parsiranja (OCR!)
            period = period_from_dates([r[0] for r in rows], tip)
            target = unique_target(tip, period, md5, taken, primary)
            if target is None:
                target = primary
            status = 'ok (keš)'
            if primary != target:
                if not dry:
                    ANALIZIRANI.mkdir(exist_ok=True)
                    shutil.move(str(primary), str(target))
                status = (f'preimenovan ← {cached["file"]} (keš)'
                          if primary.parent == ANALIZIRANI else f'premješten ← {rel} (keš)')
            taken[target] = md5
            if target.name != cached['file']:
                rows = [r[:8] + (target.name, str(r[9]).replace(cached['file'], target.name))
                        for r in rows]
            all_rows.extend(rows)
            manifest.append(dict(file=target.name, original=rel, tip=tip,
                                 period=period, ntx=cached['ntx'],
                                 md5=md5[:8], status=status))
            continue

        tip = classify(primary)
        if tip in ('NOTEXT', 'UNKNOWN', 'ERROR'):
            status = {'NOTEXT': 'bez tekst-sloja — treba OCR ili CSV export',
                      'UNKNOWN': 'nepoznat tip — pogledati ručno',
                      'ERROR': 'PDF se ne da otvoriti'}[tip]
            manifest.append(dict(file=rel, original=rel, tip=tip, period='',
                                 ntx=0, md5=md5[:8], status=status))
            continue

        parser, racun, izvor = SOURCE_TYPES[tip]
        try:
            txs = parser(primary)
        except Exception as e:
            manifest.append(dict(file=rel, original=rel, tip=tip, period='',
                                 ntx=0, md5=md5[:8], status=f'greška pri parsiranju: {e}'))
            continue
        if not txs:
            manifest.append(dict(file=rel, original=rel, tip=tip, period='',
                                 ntx=0, md5=md5[:8], status='0 transakcija — ostaje na mjestu'))
            continue

        sig = (tip, tuple(sorted((str(t['date']), t['smjer'], t['iznos'], t['opis'])
                                 for t in txs)))
        if sig in content_sigs:
            duplicates.append((primary, md5))
            print(f'  ⚠ {rel}: iste transakcije kao {content_sigs[sig]} '
                  f'(drugi bajtovi, isti izvod) → duplikat')
            continue

        period = coverage_period(txs, tip)
        target = unique_target(tip, period, md5, taken, primary)
        if target is None:            # identičan sadržaj već analiziran pod tim imenom
            duplicates.append((primary, md5))
            continue
        taken[target] = md5
        content_sigs[sig] = target.name

        status = 'ok'
        if primary != target:
            if not dry:
                ANALIZIRANI.mkdir(exist_ok=True)
                shutil.move(str(primary), str(target))
            status = f'premješten ← {rel}'
        for t in txs:                 # src referencira FINALNO ime
            t['src'] = t['src'].replace(primary.name, target.name, 1)
            all_rows.append((t['date'], t['opis'], t['iznos'], t['smjer'],
                             t.get('kartica', ''), racun, izvor, tip,
                             target.name, t['src']))
        manifest.append(dict(file=target.name, original=rel, tip=tip, period=period,
                             ntx=len(txs), md5=md5[:8], status=status))

    # 2. Duplikati → izvodi/duplikati/
    for dup, md5 in duplicates:
        rel = str(dup.relative_to(IZVODI_DIR))
        target = DUPLIKATI / dup.name
        n = 1
        while target.exists():
            n += 1
            target = DUPLIKATI / f'{dup.stem} ({n}){dup.suffix}'
        if not dry:
            DUPLIKATI.mkdir(exist_ok=True)
            shutil.move(str(dup), str(target))
        manifest.append(dict(file=f'duplikati/{target.name}', original=rel, tip='DUP',
                             period='', ntx=0, md5=md5[:8],
                             status='duplikat po sadržaju — premješten'))

    # 3. Izvodi_transakcije.xlsx
    all_rows.sort(key=lambda r: (r[7], r[0]))
    if not dry:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Transakcije'
        headers = ('Datum', 'Opis', 'Iznos', 'Smjer', 'Kartica', 'Racun',
                   'Izvor', 'Tip', 'File', 'Src')
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
        for w, col in zip((12, 62, 11, 9, 26, 18, 12, 9, 20, 24), 'ABCDEFGHIJ'):
            ws.column_dimensions[col].width = w
        for r, row in enumerate(all_rows, 2):
            for c, v in enumerate(row, 1):
                ws.cell(r, c, v)
            ws.cell(r, 1).number_format = 'DD.MM.YYYY'
            ws.cell(r, 3).number_format = '#,##0.00'
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:J{max(2, len(all_rows) + 1)}'

        wm = wb.create_sheet('Manifest')
        mh = ('File', 'Original', 'Tip', 'Period', 'Transakcija', 'MD5', 'Status')
        for c, h in enumerate(mh, 1):
            cell = wm.cell(1, c, h)
            cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
        for w, col in zip((24, 46, 9, 10, 12, 11, 44), 'ABCDEFG'):
            wm.column_dimensions[col].width = w
        manifest.sort(key=lambda m: (m['tip'], m['period'], m['file']))
        for r, m in enumerate(manifest, 2):
            for c, k in enumerate(('file', 'original', 'tip', 'period', 'ntx', 'md5', 'status'), 1):
                wm.cell(r, c, m[k])
        wm.freeze_panes = 'A2'
        wm.auto_filter.ref = f'A1:G{max(2, len(manifest) + 1)}'
        try:
            wb.save(TX_XLSX)
        except PermissionError:
            sys.exit(f'✗ Ne mogu snimiti {TX_XLSX.name} — zatvori ga u Excelu i ponovi.')

    # 4. Report: pokrivenost po tipu + rupe
    print(f'Fajlova: {len(pdfs)}  |  unikatnih: {len(by_hash)}  |  duplikata: {len(duplicates)}')
    per_tip: dict[str, list[str]] = defaultdict(list)
    for m in manifest:
        if m['period']:
            per_tip[m['tip']].append(m['period'])
    for tip in sorted(per_tip):
        periods = sorted(per_tip[tip])
        ntx = sum(m['ntx'] for m in manifest if m['tip'] == tip)
        gaps = month_gaps(periods)
        print(f'\n{tip}: {len(periods)} izvoda, {ntx} transakcija, '
              f'{periods[0]} → {periods[-1]}')
        if gaps:
            print(f'  ⚠ RUPE u pokrivenosti: {", ".join(gaps)}')
    skipped = [m for m in manifest if m['tip'] in ('NOTEXT', 'UNKNOWN', 'ERROR')]
    if skipped:
        print(f'\nPreskočeno ({len(skipped)}):')
        for m in skipped:
            print(f'  {m["original"]}: {m["status"]}')
    if not dry:
        print(f'\n✔ {TX_XLSX.name}: {len(all_rows)} transakcija, {len(manifest)} manifest redova')
        print('  Sljedeći korak: run.bat enrich_from_izvoda.py --dry')


if __name__ == '__main__':
    main()
