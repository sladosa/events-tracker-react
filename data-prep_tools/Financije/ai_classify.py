"""
ai_classify.py — AI klasifikacija Tip/Podtip za Financije Review.

Dva moda:
  --eval   NASLIJEPO na već klasificiranim redcima → mjeri točnost. NE PIŠE NIŠTA.
  --run    klasificira N/A retke i piše Tip_AI/Podtip_AI/Pouzdanost_AI/AI run.

Zašto eval daje DVA broja (v. NEXT_SESSION_PROMPT O9a):
  Od ~2580 klasificiranih redaka s tekstom, ~1414 je klasificirao `apply_rules.py`
  (keyword pravilo) — model ih pogađa trivijalno jer JE keyword u opisu. Pošten broj
  je slaganje na ručno labeliranim redcima (Kokine/Sašine labele).

Model se bira tako da EVAL VRTI ISTI MODEL I ISTI PROMPT koji ide u pravi run —
inače je izmjereni broj lažan.

MODEL NIKAD NE PIŠE U `Tip`/`Podtip`. Izmjereno je 81,5 % točnih parova na ručnim
labelama (`visoka` pouzdanost: 95 % na 57 % redaka) — to je razina "model predlaže,
čovjek potvrđuje". Prijenos AI stupaca u prave je zaseban, svjestan korak.

Pokretanje:
  # eval (Review NE mora biti zatvoren — samo čita)
  python ai_classify.py --eval --sample 600 --effort high

  # run: plan bez ijednog API poziva
  python ai_classify.py --run --dry
  # run: prava predikcija na 30 redaka, ali BEZ pisanja u Review (proba prompta)
  python ai_classify.py --run --dry --limit 30 --effort high
  # run: pravi upis (Review mora biti ZATVOREN; radi backup .pre-aiclass-*)
  python ai_classify.py --run --effort high [--resume] [--only-text]

Predikcije se uvijek spremaju u ai_predictions.jsonl, pa `--resume` poslije --dry
runa ne plaća isti redak dvaput.
"""
from __future__ import annotations

import argparse
import collections
import json
import os
import random
import re
import shutil
import sys
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
DATA = HERE.parent.parent / 'data-prep_data' / 'Financije'
ENV_FILE = HERE.parent.parent / '.env.local'

MODEL = 'claude-sonnet-5'
# 25, ne 40: potpunost odgovora pada s effortom (pri `--effort high` vratilo se 550
# od 600 poslanih redaka, uz uredan stop_reason). Manji batch = manje dozivanja.
BATCH = 25
UNKNOWN = 'NEPOZNATO'

# Bumpaj kad se promijeni prompt — predikcije iz starije verzije se NE recikliraju.
PROMPT_VER = 'v3-tvrda-pravila'
CONTEXT_FILE = 'AI_KONTEKST_pitanja.txt'      # Sašini odgovori, idu doslovno u prompt
PRED_FILE = 'ai_predictions.jsonl'            # append-only store, ključ = source_key
EVAL_SEED = 20260726                          # zamrznut uzorak → runovi su usporedivi

# Cijena po milijun tokena (Sonnet 5 uvodna cijena do 2026-08-31).
PRICE_IN, PRICE_OUT = 2.00, 10.00


# ── Pomoćno ──────────────────────────────────────────────────────────────────

def load_api_key() -> str:
    key = os.environ.get('ANTHROPIC_API_KEY')
    if key:
        return key
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding='utf-8').splitlines():
            m = re.match(r'\s*ANTHROPIC_API_KEY\s*=\s*(.+)\s*$', line)
            if m:
                return m.group(1).strip().strip('"').strip("'")
    sys.exit('✗ ANTHROPIC_API_KEY nije nađen (ni u okolini ni u .env.local).')


def pick_review() -> Path:
    cands = [p for p in DATA.glob('Financije_review_*.xlsx')
             if '.pre-' not in p.name and not p.name.startswith('~$')]
    if not cands:
        sys.exit(f'✗ Nema Review filea u {DATA}')
    return max(cands, key=lambda p: p.stat().st_mtime)


def header_map(ws) -> dict[str, int]:
    return {str(c.value).strip(): i
            for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
            if c.value}


def clean(v) -> str:
    return str(v).strip() if v is not None else ''


# ── Učitavanje ───────────────────────────────────────────────────────────────

def load_taxonomy(wb) -> list[str]:
    """Vraća listu valjanih 'Tip | Podtip' parova (bez N/A bucketa)."""
    ws = wb['Taksonomija']
    pairs, seen = [], set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        tip, pod = clean(r[0]), clean(r[1]) if len(r) > 1 else ''
        if not tip or tip == 'N/A':
            continue
        p = f'{tip} | {pod}'
        if p not in seen:
            seen.add(p)
            pairs.append(p)
    return pairs


def load_context() -> str:
    """Sašini odgovori o kraticama i granicama — idu doslovno u system prompt.

    Namjerno se NE parsira: Saša je odgovarao inline uz stavke, ne iza "ODGOVOR:",
    i svaki parser bi nešto pojeo. Cijeli tekst je ~15 KB i kešira se.
    """
    p = DATA / CONTEXT_FILE
    if not p.exists():
        print(f'⚠ Nema {CONTEXT_FILE} — vrtim BEZ konteksta (slabiji rezultat).')
        return ''
    return p.read_text(encoding='utf-8')


def load_rows(wb, want: str = 'classified') -> list[dict]:
    """Retci Reviewa u obliku koji ide modelu.

    want='classified' → već klasificirani retci S TEKSTOM (eval set, nosi 'istina').
    want='na'         → N/A retci (Tip prazan ili 'N/A') = kandidati za --run.
                        Retci bez teksta se NE izbacuju ovdje (v. --only-text) — model
                        za njih vraća NEPOZNATO, a poziv je svjesna odluka, ne slučaj.
    """
    ws = wb['Review']
    H = header_map(ws)
    need = ['Tip', 'Podtip', 'Napomena', 'Izvod opis', 'Izvor', 'Racun',
            'Uplata', 'Isplata', 'event_date', 'Pravilo run', 'Pouzdanost', 'source_key',
            'Labela iz']
    missing = [c for c in need if c not in H]
    if missing:
        sys.exit(f'✗ Review sheetu nedostaju kolone: {missing}')

    out = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tip = clean(r[H['Tip']])
        is_na = (not tip) or tip == 'N/A'
        if is_na != (want == 'na'):
            continue
        napomena, izvod = clean(r[H['Napomena']]), clean(r[H['Izvod opis']])
        labela_iz = clean(r[H['Labela iz']])
        if want == 'classified':
            if not (napomena or izvod):
                continue
            # Labelu je prenio apply_ai.py iz Tip_AI → eval bi mjerio model protiv
            # njegovog vlastitog izlaza. Takvi retci NIKAD ne ulaze u eval set.
            if labela_iz.startswith('AI:'):
                continue
        iznos = r[H['Isplata']] or r[H['Uplata']] or 0
        d = r[H['event_date']]
        out.append({
            'row': i,
            # Ključ za store je source_key, NE broj retka — retci se pomiču pri re-sortu.
            'key': clean(r[H['source_key']]) or f'row:{i}',
            'datum': d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d),
            'racun': clean(r[H['Racun']]),
            'izvor': clean(r[H['Izvor']]),
            'smjer': 'isplata' if r[H['Isplata']] else 'uplata',
            'iznos': round(float(iznos), 2) if iznos else 0.0,
            'napomena': napomena,
            'izvod': izvod,
            'ima_tekst': bool(napomena or izvod),
            'istina': f'{tip} | {clean(r[H["Podtip"]])}' if not is_na else '',
            'izvor_labele': ('ai' if labela_iz.startswith('AI:') else
                             'pravilo' if clean(r[H['Pravilo run']]) else 'rucno'),
        })
    return out


# ── Poziv modela ─────────────────────────────────────────────────────────────

SYSTEM = """Klasificiraš bankovne transakcije jednog hrvatskog kućanstva (Saša i Koka) \
u zadanu taksonomiju Tip/Podtip.

Pravila:
- Odgovori TOČNO jednim parom iz zadane liste, doslovno prepisanim.
- Ako opis stvarno ne dopušta zaključak, vrati "{unknown}". Bolje to nego pogoditi nasumično.
- Podtipovi sa sufiksom _Koka / _Sasa označavaju osobu. Kokin račun je "Kokin tekući ZABA", \
Sašin je "Sašin tekući RF" — ali to je samo naznaka, ne pravilo: oboje su kupovali objema karticama.
- "Namirnice | Hrana i ostalo" pokriva i drogerije (DM, Müller) — ustaljena konvencija ovog kućanstva.
- Prijenosi između vlastitih računa i podizanje gotovine idu u "Transfer".
- conf: "visoka" kad je merchant jasno prepoznatljiv, "srednja" kad je vjerojatno ali ne sigurno, \
"niska" kad pogađaš.
- Ako je zapis čista interna bilješka (osobno ime, kratica, ime + broj bez konteksta) i ništa \
u kontekstu ispod ga ne objašnjava — vrati "{unknown}" s conf "niska". NEMOJ pogađati \
samouvjereno; na prvom evalu je model za "Igor kune" i "Zoran povrat" tvrdio "Domaćinstvo | Struja" \
s visokom pouzdanošću, što je gore od priznanja da ne zna.

TVRDA PRAVILA — imaju prednost pred svime, uključujući kontekst na dnu.
Izvedena su iz Sašinih odgovora i iz stvarnih grešaka na prethodnom evalu.

 1. OSOBA KOD PRIHODA: kod uplata (mirovina, plaća, povrat) osobu određuje RAČUN —
    "Kokin tekući ZABA" → Koka, "Sašin tekući RF" → Saša. Kod KUPOVINA račun NE
    određuje osobu (oboje su kupovali objema karticama).
 2. AUDIBLE: ako u opisu piše AUDIBLE, Podtip je uvijek Audible_* i NIKAD Kindle.
    Iznos < 10 € → Audible_Koka; > 10 € (tipično ~16 €) → Audible_Sasa.
    Kindle_Koka samo ako u opisu doslovno piše KINDLE.
 3. RADNIČKA 49 = Sašino radno mjesto:
    · BIBERON (bilo gdje) → uvijek "Projekti | Sasa_Informatika".
    · KONZUM + RADNIČKA + isplata < 30 € i BEZ "RATA" → "Projekti | Sasa_Informatika"
      (ručak uz posao). Svaki drugi Konzum → "Namirnice | Hrana i ostalo".
 4. RATE: "RATA nn/mm" ili "ime N/M" znači N-tu ratu od ukupno M. Kategoriju određuje
    MERCHANT, ne činjenica da je rata. Konzum na rate ostaje Namirnice (velika kupovina
    razbijena na rate NIJE ručak, pa ne ide u Projekti).
 5. TRANSFER vs OSTALI PRIHODI: prijenos vlastitog novca ("Saša uplata", "PBZ Visa",
    naplata kartice, KEKS između njih dvoje) → "Transfer | izmedju racuna".
    "Ostali prihodi" je samo za STVARNI vanjski prihod koji nije mirovina ni povrat.
 6. AUTO kad se ne zna koji: default je "auto C5" (Koka ga vozi). "auto Lacetti" samo
    ako opis to izričito kaže (npr. oznaka SS = Saša Sladoljev).
 7. Ako Podtip nosi sufiks _Koka/_Sasa, a u Napomeni stoji ime ("Saša multisport"),
    ime određuje sufiks. To vrijedi SAMO za Podtipove sa sufiksom — "Saša Holding"
    je i dalje "Domaćinstvo | Holding (smeće)", jer taj Podtip nema osobu.

Dostupni parovi (Tip | Podtip):
{pairs}
{context}"""

CONTEXT_HDR = """

================================================================================
KONTEKST OD VLASNIKA PODATAKA (Saša) — kako ovo kućanstvo stvarno bilježi troškove.
Ovo je autoritativno: gdje se kosi s tvojom intuicijom, vrijedi ovo.
Tekst je izvorni odgovor na pitanja, uključujući primjere redaka.
================================================================================
"""


def render(rows: list[dict]) -> str:
    lines = []
    for r in rows:
        txt = ' / '.join(x for x in (r['napomena'], r['izvod']) if x)
        lines.append(f"[{r['row']}] {r['datum']} · {r['racun']} · {r['izvor']} · "
                     f"{r['smjer']} {r['iznos']:.2f} € · {txt}")
    return '\n'.join(lines)


def schema(pairs: list[str]) -> dict:
    return {
        'type': 'object',
        'properties': {
            'results': {
                'type': 'array',
                'items': {
                    'type': 'object',
                    'properties': {
                        'id': {'type': 'integer'},
                        'par': {'type': 'string', 'enum': pairs + [UNKNOWN]},
                        'conf': {'type': 'string', 'enum': ['visoka', 'srednja', 'niska']},
                    },
                    'required': ['id', 'par', 'conf'],
                    'additionalProperties': False,
                },
            }
        },
        'required': ['results'],
        'additionalProperties': False,
    }


def is_fatal(e: Exception) -> bool:
    """Greške koje ponavljanje ne rješava: prazan kredit, loš ključ, neispravan zahtjev.

    Naučeno na pravom runu: kad je kredit pao usred posla, svaki od preostalih batcheva
    bi inače odradio 4 pokušaja s backoffom prije nego što odustane.
    """
    if getattr(e, 'status_code', None) in (400, 401, 403):
        return True
    return 'credit balance' in str(e).lower()


def _call(client, model: str, pairs: list[str], rows: list[dict],
          effort: str, context: str = '') -> tuple[dict, dict]:
    """Jedan poziv. Vraća ({row: (par, conf)}, usage)."""
    sys_prompt = SYSTEM.format(unknown=UNKNOWN, pairs='\n'.join(pairs),
                               context=(CONTEXT_HDR + context) if context else '')
    ids = [r['row'] for r in rows]
    # Opseg se navodi EKSPLICITNO: Sonnet 5 čita upute doslovno i ne generalizira
    # instrukciju s prve stavke na ostale (v. migration guide). Bez ovoga je pri
    # effort=low vraćao 1 rezultat na 40 redaka, uz uredan stop_reason=end_turn.
    user = (f'Ispod je {len(rows)} transakcija.\n'
            f'Vrati TOČNO {len(rows)} rezultata — po jedan za SVAKI navedeni ID, '
            f'istim redoslijedom. Nemoj preskočiti nijedan i nemoj stati na prvom.\n'
            f'ID-evi: {", ".join(str(i) for i in ids)}\n\n' + render(rows))

    last = None
    for attempt in range(4):
        try:
            resp = client.messages.create(
                model=model,
                max_tokens=16000,
                system=[{'type': 'text', 'text': sys_prompt,
                         'cache_control': {'type': 'ephemeral'}}],
                thinking={'type': 'disabled'},
                output_config={'effort': effort,
                               'format': {'type': 'json_schema', 'schema': schema(pairs)}},
                messages=[{'role': 'user', 'content': user}],
            )
            break
        except Exception as e:                                   # noqa: BLE001
            if is_fatal(e):        # prazan kredit / loš zahtjev — retry je čisti gubitak
                raise
            last = e
            time.sleep(2 ** attempt + random.random())
    else:
        raise RuntimeError(f'4 pokušaja neuspješna: {last}')

    if resp.stop_reason == 'refusal':
        raise RuntimeError(f'refusal: {getattr(resp, "stop_details", None)}')

    text = ''.join(b.text for b in resp.content if b.type == 'text')
    data = json.loads(text)
    wanted = set(ids)
    canon = {p.lower(): p for p in pairs}          # structured-output enum NIJE obvezujuć:
    got = {}                                       # vraćalo je 'Hrana I ostalo' (veliko I)
    for x in data['results']:
        if int(x['id']) not in wanted:
            continue
        par = x['par']
        if par != UNKNOWN:
            par = canon.get(par.lower(), par)
        got[int(x['id'])] = (par, x['conf'])
    u = resp.usage
    return got, {
        'in': u.input_tokens,
        'out': u.output_tokens,
        'cache_w': getattr(u, 'cache_creation_input_tokens', 0) or 0,
        'cache_r': getattr(u, 'cache_read_input_tokens', 0) or 0,
    }


def classify(client, model: str, pairs: list[str], rows: list[dict],
             effort: str = 'medium', context: str = '') -> tuple[dict, dict]:
    """Poziv + dopuna: nepotpun odgovor se NE prihvaća tiho, nego se doziva.

    Tihi manjak je najgori mogući ishod — izgleda kao uspješan run s manjim N.
    """
    got, usage = _call(client, model, pairs, rows, effort, context)
    remaining = [r for r in rows if r['row'] not in got]
    for _ in range(3):
        if not remaining:
            break
        more, u = _call(client, model, pairs, remaining, effort, context)
        got.update(more)
        for k, v in u.items():
            usage[k] += v
        nxt = [r for r in remaining if r['row'] not in got]
        if len(nxt) == len(remaining):          # nema napretka → ne petljaj
            break
        remaining = nxt
    if remaining:
        print(f'\n  ⚠ {len(remaining)} redaka bez odgovora nakon dopuna '
              f'(redovi: {[r["row"] for r in remaining][:8]}…)')
    return got, usage


# ── Izvještaj ────────────────────────────────────────────────────────────────

def report(rows: list[dict], pred: dict, usage: dict, model: str) -> None:
    buckets = collections.defaultdict(lambda: {'n': 0, 'hit': 0, 'tip_hit': 0, 'unk': 0})
    conf_stats = collections.defaultdict(lambda: {'n': 0, 'hit': 0})
    confusion = collections.Counter()
    misses = []

    for r in rows:
        p = pred.get(r['row'])
        if not p:
            continue
        par, conf = p
        b = buckets[r['izvor_labele']]
        a = buckets['SVE']
        for x in (b, a):
            x['n'] += 1
        if par == UNKNOWN:
            for x in (b, a):
                x['unk'] += 1
            continue
        ok = par == r['istina']
        tip_ok = par.split(' | ')[0] == r['istina'].split(' | ')[0]
        for x in (b, a):
            x['hit'] += ok
            x['tip_hit'] += tip_ok
        conf_stats[conf]['n'] += 1
        conf_stats[conf]['hit'] += ok
        if not ok:
            confusion[(r['istina'], par)] += 1
            misses.append((r['row'], r['izvor_labele'], conf, r['istina'], par,
                           (r['napomena'] or r['izvod'])[:58]))

    def pct(a, b):
        return f'{a / b * 100:5.1f}%' if b else '    —'

    print('\n' + '=' * 78)
    print(f'  EVAL — {model}   (naslijepo, ništa nije pisano u Review)')
    print('=' * 78)
    print(f'\n{"skup":<22}{"N":>6}{"točan par":>12}{"točan Tip":>12}{"NEPOZNATO":>12}')
    print('-' * 78)
    for k, lbl in (('rucno', 'RUČNE labele ⭐'), ('pravilo', 'pravilo (napuhano)'), ('SVE', 'ukupno')):
        b = buckets.get(k)
        if not b or not b['n']:
            continue
        star = '  ← odluka na ovome' if k == 'rucno' else ''
        print(f'{lbl:<22}{b["n"]:>6}{pct(b["hit"], b["n"]):>12}'
              f'{pct(b["tip_hit"], b["n"]):>12}{pct(b["unk"], b["n"]):>12}{star}')

    print(f'\n{"pouzdanost modela":<22}{"N":>6}{"točan par":>12}')
    print('-' * 78)
    for c in ('visoka', 'srednja', 'niska'):
        s = conf_stats.get(c)
        if s and s['n']:
            print(f'{c:<22}{s["n"]:>6}{pct(s["hit"], s["n"]):>12}')

    if confusion:
        print(f'\nNajčešća neslaganja (istina → model), top 15:')
        print('-' * 78)
        for (t, p), n in confusion.most_common(15):
            print(f'{n:>4}×  {t[:33]:<34} → {p[:33]}')

    tot_in = usage['in'] + usage['cache_w'] + usage['cache_r']
    cost = ((usage['in'] + usage['cache_w'] * 1.25 + usage['cache_r'] * 0.1) / 1e6 * PRICE_IN
            + usage['out'] / 1e6 * PRICE_OUT)
    print(f'\nTokeni: {tot_in:,} ulaz ({usage["cache_r"]:,} iz keša) · {usage["out"]:,} izlaz')
    print(f'STVARNI TROŠAK: ${cost:.2f}')

    out = DATA / 'ai_eval_neslaganja.tsv'
    with out.open('w', encoding='utf-8') as f:
        f.write('red\tizvor_labele\tconf\tistina\tmodel\ttekst\n')
        for m in sorted(misses, key=lambda x: (x[1], x[0])):
            f.write('\t'.join(str(x) for x in m) + '\n')
    print(f'Neslaganja ({len(misses)}) → {out.name}   ← ovo je i detektor postojećih grešaka')


# ── Upis u Review (--run) ────────────────────────────────────────────────────

AI_COLS = ['Tip_AI', 'Podtip_AI', 'Pouzdanost_AI', 'AI run']
AI_HIDDEN = {'Pouzdanost_AI', 'AI run'}          # collapsed grupa (odluka S107m)
AI_WIDTH = {'Tip_AI': 14, 'Podtip_AI': 17, 'Pouzdanost_AI': 13, 'AI run': 17}

HDR_FILL = PatternFill('solid', fgColor='7030A0')   # ljubičasto = AI, da se ne miješa
WHITE_BOLD = Font(color='FFFFFF', bold=True)        # s plavim ljudskim kolonama
_THIN = Side(style='thin')
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def ensure_ai_columns(ws) -> dict[str, int]:
    """AI kolone odmah desno od `Podtip`, 1-based indeksi. Idempotentno.

    Umeće se DESNO od J/K (Tip/Podtip) pa data validation i conditional formatting,
    koji žive na J2:J* / K2:K*, ostaju netaknuti — openpyxl ih ionako ne bi pomaknuo.
    Zato je smjer umetanja bitan i provjerava se na kraju.
    """
    H = header_map(ws)
    have = [c for c in AI_COLS if c in H]
    if len(have) == len(AI_COLS):
        return {c: H[c] + 1 for c in AI_COLS}          # header_map je 0-based
    if have:
        sys.exit(f'✗ Postoji samo dio AI kolona: {have}. Očekujem sve četiri ili nijednu — '
                 f'obriši te kolone pa ponovi.')
    if 'Podtip' not in H:
        sys.exit('✗ Review nema kolonu "Podtip" — ne znam gdje umetnuti AI kolone.')

    at = H['Podtip'] + 2                                # 1-based, odmah iza Podtipa
    n = len(AI_COLS)

    # insert_cols pomiče ćelije, ali NE i column_dimensions (širine/outline ostale bi
    # na starim slovima i raspale bi se po smislu). Zato ih prenosimo ručno.
    snap = {}
    for i in range(1, ws.max_column + 1):
        L = get_column_letter(i)
        if L in ws.column_dimensions:
            d = ws.column_dimensions[L]
            snap[i] = (d.width, d.outlineLevel, d.hidden)   # customWidth je izveden, read-only

    ws.insert_cols(at, n)

    for L in list(ws.column_dimensions):
        del ws.column_dimensions[L]
    for i, (w, lvl, hid) in snap.items():
        d = ws.column_dimensions[get_column_letter(i if i < at else i + n)]
        d.width, d.outlineLevel, d.hidden = w, lvl, hid

    for k, name in enumerate(AI_COLS):
        c = at + k
        cell = ws.cell(1, c, name)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
        ws.column_dimensions[get_column_letter(c)].width = AI_WIDTH[name]
        d = ws.column_dimensions[get_column_letter(c)]
        if name in AI_HIDDEN:
            d.outlineLevel, d.hidden = 1, True

    if ws.auto_filter.ref:
        first = ws.auto_filter.ref.split(':')[0]
        ws.auto_filter.ref = f'{first}:{get_column_letter(ws.max_column)}{ws.max_row}'

    _check_dv_alignment(ws)
    print(f'✔ Kreirane kolone {", ".join(AI_COLS)} '
          f'({get_column_letter(at)}–{get_column_letter(at + n - 1)}); '
          f'{", ".join(sorted(AI_HIDDEN))} u collapsed grupi')
    return {name: at + k for k, name in enumerate(AI_COLS)}


def _check_dv_alignment(ws) -> None:
    """Dropdowni Tip/Podtip moraju i dalje sjediti na kolonama Tip/Podtip."""
    H = header_map(ws)
    want = {get_column_letter(H['Tip'] + 1): 'Tip', get_column_letter(H['Podtip'] + 1): 'Podtip'}
    for dv in ws.data_validations.dataValidation:
        cols = {re.match(r'([A-Z]+)', str(rng)).group(1) for rng in dv.sqref.ranges}
        for col in cols:
            if col not in want:
                print(f'⚠ Data validation ({dv.formula1}) sjedi na koloni {col}, '
                      f'a to više nije Tip/Podtip — provjeri raspored kolona!')


def write_predictions(ws, col: dict[str, int], rows: list[dict],
                      pred: dict, stamp: str, effort: str) -> int:
    """Piše SAMO u AI kolone. `Tip`/`Podtip` se ne diraju — nikad, ni pri visokoj conf."""
    mark = f'{stamp} · {PROMPT_VER} · {effort}'
    written = 0
    for r in rows:
        p = pred.get(r['row'])
        if not p:
            continue
        par, conf = p
        tip, pod = (('', '') if par == UNKNOWN else
                    (par.split(' | ')[0], par.split(' | ')[1] if ' | ' in par else ''))
        ws.cell(r['row'], col['Tip_AI'], tip or UNKNOWN)
        ws.cell(r['row'], col['Podtip_AI']).value = pod or None
        ws.cell(r['row'], col['Pouzdanost_AI'], conf)
        ws.cell(r['row'], col['AI run'], mark)
        written += 1
    return written


def run_report(rows: list[dict], pred: dict, usage: dict, model: str, effort: str) -> None:
    conf_n = collections.Counter()
    pairs_n = collections.Counter()
    unk = 0
    for r in rows:
        p = pred.get(r['row'])
        if not p:
            continue
        par, conf = p
        conf_n[conf] += 1
        if par == UNKNOWN:
            unk += 1
        else:
            pairs_n[par] += 1

    got = sum(conf_n.values())
    print('\n' + '=' * 78)
    print(f'  RUN — {model} · effort {effort}')
    print('=' * 78)
    print(f'\nPoslano {len(rows)} · vraćeno {got}'
          + (f'  ⚠ BEZ ODGOVORA: {len(rows) - got}' if got < len(rows) else '  (potpuno)'))
    print(f'NEPOZNATO (model priznaje da ne zna): {unk}')
    print(f'\n{"pouzdanost":<14}{"N":>7}{"udio":>9}')
    print('-' * 32)
    for c in ('visoka', 'srednja', 'niska'):
        if conf_n[c]:
            print(f'{c:<14}{conf_n[c]:>7}{conf_n[c] / got * 100:>8.1f}%')
    print(f'\nNajčešći predloženi parovi, top 15:')
    print('-' * 60)
    for par, n in pairs_n.most_common(15):
        print(f'{n:>5}×  {par}')

    if usage:
        cost = ((usage['in'] + usage['cache_w'] * 1.25 + usage['cache_r'] * 0.1) / 1e6 * PRICE_IN
                + usage['out'] / 1e6 * PRICE_OUT)
        print(f'\nTokeni: {usage["in"] + usage["cache_w"] + usage["cache_r"]:,} ulaz '
              f'({usage["cache_r"]:,} iz keša) · {usage["out"]:,} izlaz')
        print(f'STVARNI TROŠAK: ${cost:.2f}')


# ── Main ─────────────────────────────────────────────────────────────────────

def store_load(model: str, effort: str) -> dict[str, tuple[str, str]]:
    """Predikcije iz ranijih runova s ISTIM (prompt_ver, model, effort). Ključ = source_key."""
    p = DATA / PRED_FILE
    if not p.exists():
        return {}
    out: dict[str, tuple[str, str]] = {}
    for line in p.read_text(encoding='utf-8').splitlines():
        if not line.strip():
            continue
        try:
            d = json.loads(line)
        except json.JSONDecodeError:
            continue
        if (d.get('prompt_ver'), d.get('model'), d.get('effort')) == (PROMPT_VER, model, effort):
            out[d['key']] = (d['par'], d['conf'])            # kasniji zapis pobjeđuje
    return out


def store_append(recs: list[dict]) -> None:
    with (DATA / PRED_FILE).open('a', encoding='utf-8') as f:
        for r in recs:
            f.write(json.dumps(r, ensure_ascii=False) + '\n')


def stratified(rows: list[dict], n: int) -> list[dict]:
    """Zamrznut uzorak, pola ručnih / pola pravilo — da su runovi međusobno usporedivi."""
    rnd = random.Random(EVAL_SEED)
    out: list[dict] = []
    for src in ('rucno', 'pravilo'):
        pool = sorted((r for r in rows if r['izvor_labele'] == src), key=lambda r: r['key'])
        out += rnd.sample(pool, min(n // 2, len(pool)))
    return out


# Iz izmjerenih runova (600 redaka · effort high · $0,73). Samo za procjenu u --dry.
EST_PER_ROW = 0.0013


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument('--eval', action='store_true',
                    help='naslijepo na klasificiranim redcima; NE PIŠE u Review')
    ap.add_argument('--run', action='store_true',
                    help='klasificiraj N/A retke i upiši Tip_AI/Podtip_AI/Pouzdanost_AI/AI run')
    ap.add_argument('--dry', action='store_true',
                    help='--run bez pisanja u Review. Bez --limit ne zove ni model (samo plan); '
                         's --limit N odradi pravu predikciju na N redaka i pokaže uzorak.')
    ap.add_argument('--only-text', action='store_true',
                    help='--run samo na retke koji imaju Napomenu ili Izvod opis')
    ap.add_argument('--sample', type=int, default=0,
                    help='zamrznut stratificiran uzorak od N (samo --eval)')
    ap.add_argument('--limit', type=int, default=0, help='N nasumičnih redaka (smoke test)')
    ap.add_argument('--resume', action='store_true',
                    help='preskoči retke koje store već ima za isti prompt_ver+model+effort')
    ap.add_argument('--only-conf', default='',
                    help='ponovi SAMO retke te pouzdanosti iz storea, npr. niska,srednja')
    ap.add_argument('--workers', type=int, default=5)
    ap.add_argument('--model', default=MODEL)
    ap.add_argument('--effort', default='medium', choices=['low', 'medium', 'high', 'xhigh'])
    args = ap.parse_args()

    if args.eval == args.run:
        ap.error('odaberi točno jedan mod: --eval ILI --run')
    if args.eval and (args.dry or args.only_text):
        ap.error('--dry i --only-text vrijede samo uz --run (eval ionako ništa ne piše)')
    if args.run and args.sample:
        ap.error('--sample je stratificiran po izvoru labele — vrijedi samo uz --eval')

    review = pick_review()
    # data_only=True bi pri spremanju formule pretvorio u vrijednosti → samo za eval.
    wb = openpyxl.load_workbook(review, data_only=args.eval)
    pairs = load_taxonomy(wb)
    rows = load_rows(wb, 'na' if args.run else 'classified')
    if args.eval:
        wb.close()
    context = load_context()

    n_all = len(rows)
    n_text = sum(1 for r in rows if r['ima_tekst'])
    if args.run and args.only_text:
        rows = [r for r in rows if r['ima_tekst']]
    if args.sample:
        rows = stratified(rows, args.sample)
    elif args.limit:
        rnd = random.Random(EVAL_SEED)
        rows = list(rows)
        rnd.shuffle(rows)
        rows = rows[:args.limit]

    cached = store_load(args.model, args.effort)
    todo = rows
    if args.only_conf:
        want = {c.strip() for c in args.only_conf.split(',')}
        todo = [r for r in rows if cached.get(r['key'], ('', ''))[1] in want]
    elif args.resume:
        todo = [r for r in rows if r['key'] not in cached]

    print(f'Review: {review.name}')
    print(f'Taksonomija: {len(pairs)} parova · '
          + (f'kontekst {len(context):,} znakova' if context else 'kontekst: NEMA'))
    if args.eval:
        src = collections.Counter(r['izvor_labele'] for r in rows)
        print(f'Eval set: {len(rows)} redaka  (ručno {src["rucno"]} · pravilo {src["pravilo"]})')
    else:
        print(f'N/A retci: {n_all} ukupno  (s tekstom {n_text} · bez teksta {n_all - n_text})')
        print(f'Za klasifikaciju: {len(rows)} redaka'
              + ('  [--only-text]' if args.only_text else '')
              + (f'  [--limit {args.limit}]' if args.limit else ''))
    print(f'Store: {len(cached)} predikcija za {PROMPT_VER}/{args.model}/{args.effort}'
          f'  →  zovem model za {len(todo)}')
    print(f'Model: {args.model} · effort {args.effort} · batch {BATCH} · '
          f'{args.workers} paralelno')

    if args.run and args.dry and not args.limit:
        H = header_map(wb['Review'])
        stanje = ('postoje' if all(c in H for c in AI_COLS)
                  else f'kreiraju se desno od "Podtip" ({", ".join(AI_COLS)})')
        print(f'\n[DRY — PLAN, bez ijednog API poziva]')
        print(f'  AI kolone: {stanje}; '
              f'{", ".join(sorted(AI_HIDDEN))} idu u collapsed grupu')
        print(f'  Tip/Podtip se NE diraju ni u jednom slučaju.')
        print(f'  Procjena troška za {len(todo)} redaka: ~${len(todo) * EST_PER_ROW:.2f}')
        print(f'\n  Proba prompta bez pisanja:  --run --dry --limit 30 --effort {args.effort}')
        print(f'  Pravi upis:                 --run --effort {args.effort} --resume')
        return
    print()

    pred = {r['row']: cached[r['key']] for r in rows if r['key'] in cached}
    usage: collections.Counter = collections.Counter()

    if todo:
        import anthropic

        batches = [todo[i:i + BATCH] for i in range(0, len(todo), BATCH)]
        client = anthropic.Anthropic(api_key=load_api_key())
        by_row = {r['row']: r for r in todo}
        stamp = datetime.now().isoformat(timespec='seconds')
        run_id = f'{stamp}-{PROMPT_VER}'
        done = 0

        def work(b):
            return classify(client, args.model, pairs, b, args.effort, context)

        # Pao batch NE ruši cijeli run: ono što je plaćeno i dobiveno se zadrži i upiše,
        # a ostatak se dovrši s --resume. (Naučeno: kredit je pao na 19/64 batcheva i
        # cijeli je posao propao pri izlasku, iako je 491 predikcija bila u storeu.)
        failed_rows, fatal = 0, None
        with ThreadPoolExecutor(max_workers=args.workers) as ex:
            futs = {ex.submit(work, b): b for b in batches}
            for fut in as_completed(futs):
                try:
                    got, u = fut.result()
                except Exception as e:                           # noqa: BLE001
                    failed_rows += len(futs[fut])
                    if is_fatal(e):
                        fatal = e
                        for f in futs:
                            f.cancel()
                        break
                    print(f'\n  ⚠ batch pao ({e}) — nastavljam s ostalima')
                    continue
                pred.update(got)
                usage.update(u)
                store_append([
                    {'key': by_row[row]['key'], 'row': row, 'run_id': run_id,
                     'prompt_ver': PROMPT_VER, 'model': args.model, 'effort': args.effort,
                     'mode': 'run' if args.run else 'eval',
                     'par': par, 'conf': conf, 'ts': stamp}
                    for row, (par, conf) in got.items() if row in by_row
                ])
                done += 1
                print(f'\r  batch {done}/{len(batches)}', end='', flush=True)
        print()
        if fatal:
            print(f'\n✗ PREKID: {fatal}')
            print(f'  Zadržavam {len(pred)} predikcija; ostatak dovrši s --resume '
                  f'(store ih ne plaća ponovo).')
        elif failed_rows:
            print(f'\n⚠ {failed_rows} redaka je ostalo bez odgovora zbog palih batcheva.')

    if args.eval:
        report(rows, pred, dict(usage), args.model)
        print(f'Predikcije → {PRED_FILE}   ({len(pred)} u ovom izvještaju)')
        return

    run_report(rows, pred, dict(usage), args.model, args.effort)
    print(f'Predikcije → {PRED_FILE}   ({len(pred)} u ovom izvještaju)')

    if args.dry:
        print('\nUzorak (redak · pouzdanost · prijedlog · tekst) — prvih 25:')
        print('-' * 78)
        for r in sorted(rows, key=lambda r: r['row']):
            p = pred.get(r['row'])
            if not p:
                continue
            txt = (r['napomena'] or r['izvod'])[:40]
            print(f'{r["row"]:>6}  {p[1]:<8} {p[0][:36]:<38} {txt}')
        print(f'\n✔ [DRY] Ništa nije pisano u Review. Predikcije su u storeu — '
              f'pravi run s --resume ih ne plaća ponovo.')
        return

    if not pred:
        print('\n✔ Nema nijedne predikcije — ništa za snimiti, Review netaknut.')
        return

    ws = wb['Review']
    col = ensure_ai_columns(ws)
    written = write_predictions(ws, col, rows, pred,
                                datetime.now().strftime('%Y-%m-%d %H:%M'), args.effort)
    backup = review.with_name(f'{review.stem}.pre-aiclass-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        sys.exit(f'✗ Zatvori Review u Excelu i ponovi. (Backup: {backup.name})')
    print(f'\n✔ Snimljeno: {written} redaka u Tip_AI/Podtip_AI/Pouzdanost_AI/AI run. '
          f'Tip/Podtip netaknuti.')
    print(f'  Backup: {backup.name}')


if __name__ == '__main__':
    main()
