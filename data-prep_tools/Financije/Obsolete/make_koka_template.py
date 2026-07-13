# -*- coding: utf-8 -*-
"""
Generira Koka_Struktura_Template.xlsx — template koji Koka popunjava/editira.
Format je čitljiv i parseable: kad Koka vrati editiran file, isti kod
može čitati njenu verziju i generirati make_import.py strukturu.

Format sheeta "Struktura":
  A: Kategorija  — leaf naziv (npr. "Svakodnevni")
  B: Atribut     — naziv atributa
  C: Opcije      — comma-separated dropdown opcije; "(tekst)" za slobodan unos; "(broj EUR)" za number
  D: Ovisi o     — slug roditeljskog atributa (za DependsOn); prazno = nema ovisnosti
  E: Za vrijednost — specifična vrijednost roditelja (npr. "Hrana"); prazno = bez ovisnosti
  F: Napomena    — za Koku

Posebni retci (col A):
  "=== RASHODI ===" i sl.  → header sekcija, zanemariti pri parsiranju
  "(L1) Rashodi"            → L1 atributi koji se pojavljuju na SVIM rashodima
"""
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
sys.stdout.reconfigure(encoding='utf-8')

import os
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(SCRIPT_DIR, "Koka_Struktura_Template.xlsx")

# ── Stilovi ─────────────────────────────────────────────────────────────────
THIN  = Side(style='thin')
BRD   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
L     = Alignment(horizontal='left', vertical='center', wrap_text=True)
C     = Alignment(horizontal='center', vertical='center')

def fill(hex_):  return PatternFill("solid", fgColor=hex_)
def font(hex_, bold=False, sz=11): return Font(color=hex_, bold=bold, size=sz)

HDR_FILL = fill("4472C4"); HDR_FONT = font("FFFFFF", bold=True)
L1_FILL  = fill("E2EFDA"); L1_FONT  = font("375623", bold=True)  # svjetlozelena
SEC_FILL = fill("BDD7EE"); SEC_FONT = font("1F4E79", bold=True)  # plava header
DEP_FILL = fill("FFF2CC")  # žuta za DependsOn retke
REG_FILL = fill("FFFFFF")
NOTE_FILL= fill("F2F2F2")

HEADERS = ["Kategorija", "Atribut", "Opcije / Vrijednosti", "Ovisi o atributu", "Za vrijednost", "Napomena"]

# ── Sadržaj ─────────────────────────────────────────────────────────────────
# (kategorija, atribut, opcije, ovisi_o, za_vrijednost, napomena)
# Kategorija koja počinje s "==" je section header (preskočiti pri parsiranju)
# Kategorija koja počinje s "(L1)" je L1 atribut (zajednički za sve u grani)

ROWS = [
    # ── L1 RASHODI ─────────────────────────────────────────────────────────
    ("== RASHODI ==", "", "", "", "", "Svi troškovi — ove 3 vrijednosti se pojavljuju na svakom rashodu"),
    ("(L1) Rashodi", "Iznos",  "(broj EUR)",  "", "", "Iznos troška u EUR"),
    ("(L1) Rashodi", "Račun",  "Mastercard (Koka), Visa (Saša), Kokin tekući (Zaba), Raiffeisen (Saša), Gotovina", "", "", "S kojeg računa ili kartice"),
    ("(L1) Rashodi", "Valuta", "EUR, HRK, USD", "", "", "Obično EUR — možeš ostaviti prazno ako uvijek EUR"),

    # ── SVAKODNEVNI ────────────────────────────────────────────────────────
    ("== Svakodnevni ==", "", "", "", "", "Namirnice, kava, drogerija, kiosk..."),
    ("Svakodnevni", "Vrsta",  "Hrana, Kava, Drogerija, Kiosk, Ostalo", "", "", "Primarni tip troška"),
    ("Svakodnevni", "Dućan",  "Konzum, Spar, Studenac, Lidl, Mlinar, Voćarna, Pekara, Bofrost, Nespresso", "Vrsta", "Hrana", "Dućan za namirnice"),
    ("Svakodnevni", "Dućan",  "DM, Biberon", "Vrsta", "Drogerija", "Drogerija/ljekarnički artikli"),
    ("Svakodnevni", "Dućan",  "Studenac, Tisak, Kraš", "Vrsta", "Kiosk", "Kiosk/trafika"),

    # ── DOM ────────────────────────────────────────────────────────────────
    ("== Dom ==", "", "", "", "", "Komunalije, holding, pričuva, osiguranje nekretnine"),
    ("Dom", "Vrsta",  "HEP, Plin, Voda, Grijanje, Holding, Pričuva, Osiguranje, Ostalo", "", "", "Tip troška stanovanja"),
    ("Dom", "Objekt", "Bulatova, Medulićeva, Kućište, Nena, Mama", "Vrsta", "HEP",       "Za koji objekt/stan"),
    ("Dom", "Objekt", "Bulatova, Kućište",                          "Vrsta", "Plin",      ""),
    ("Dom", "Objekt", "Bulatova",                                   "Vrsta", "Voda",      ""),
    ("Dom", "Objekt", "Bulatova, Nena, Mama",                       "Vrsta", "Grijanje",  ""),
    ("Dom", "Objekt", "Nataša, Saša, Nena, Mama, Medulićeva, Kućište", "Vrsta", "Holding", ""),
    ("Dom", "Objekt", "Nataša, Saša, Nena, Mama, Medulićeva, Kućište", "Vrsta", "Pričuva", ""),
    ("Dom", "Objekt", "(tekst)",                                    "Vrsta", "Osiguranje","Slobodan unos za osiguranje"),

    # ── PRIJEVOZ ───────────────────────────────────────────────────────────
    ("== Prijevoz ==", "", "", "", "", "Gorivo, parking, taxi, servis..."),
    ("Prijevoz", "Vrsta",         "Gorivo, Parking, Taxi, Bolt, Servis, Registracija, HAK, Ostalo", "", "", ""),
    ("Prijevoz", "Vrsta goriva",  "Dizel, Benzin",   "Vrsta", "Gorivo",  "Samo za gorivo"),
    ("Prijevoz", "Stanica",       "Ina, Petrol, (tekst)", "Vrsta", "Gorivo", "Benzinska stanica"),

    # ── ZDRAVLJE ───────────────────────────────────────────────────────────
    ("== Zdravlje ==", "", "", "", "", "Ljekarna, liječnici, osiguranja, mirovinski"),
    ("Zdravlje", "Vrsta",    "Ljekarna, Liječnik, HLK, Optika, Stomatolog, Dopunsko, Životno, Mirovinski, Ostalo", "", "", ""),
    ("Zdravlje", "Osoba",    "Saša, Koka, Nena, Mama", "Vrsta", "Dopunsko",   "Za koga je dopunsko"),
    ("Zdravlje", "Osoba",    "Saša, Koka, Nena, Mama", "Vrsta", "Životno",    ""),
    ("Zdravlje", "Osoba",    "Saša, Koka",             "Vrsta", "Mirovinski", ""),
    ("Zdravlje", "Osiguravatelj", "Allianz, Triglav, Generali, Wustenrot, Jadransko, Passport", "Vrsta", "Dopunsko", ""),
    ("Zdravlje", "Osiguravatelj", "Allianz, Triglav, Generali, Wustenrot, Jadransko",           "Vrsta", "Životno",  ""),

    # ── PRETPLATE ──────────────────────────────────────────────────────────
    ("== Pretplate ==", "", "", "", "", "Streaming, cloud, novine, domene..."),
    ("Pretplate", "Naziv", "Claude, YouTube, HBO, Disney+, Spotify, Audible, Apple, Netflix, PrimeVideo, SkyShowtime, Google, Jutarnji list, Netdomena, Amazon Prime", "", "", "Dodaj što nedostaje!"),

    # ── RATE ───────────────────────────────────────────────────────────────
    ("== Rate ==", "", "", "", "", "Obročno plaćanje — Konzum, Keindl, auto, traperice, PP fondovi..."),
    ("Rate", "Naziv", "(tekst)", "", "", "Što se plaća na rate (npr. Konzum, Keindl, PP SS)"),
    ("Rate", "Rata",  "(tekst)", "", "", "Npr. 3/12 ili 1/60"),

    # ── KUPOVINA ───────────────────────────────────────────────────────────
    ("== Kupovina ==", "", "", "", "", "Online, odjeća, kućanstvo, sport, darovi..."),
    ("Kupovina", "Vrsta", "Online, Odjeća, Kućanstvo, Sport, Dar, Ostalo", "", "", ""),
    ("Kupovina", "Opis",  "(tekst)", "", "", "Što je kupljeno (npr. Temu — jastuci, Igor — poklon za stan)"),

    # ── RESTORAN ───────────────────────────────────────────────────────────
    ("== Restoran i kava ==", "", "", "", "", ""),
    ("Restoran i kava", "Naziv", "(tekst)", "", "", "Naziv restorana ili kafića"),

    # ── OSTALO ─────────────────────────────────────────────────────────────
    ("== Ostalo ==", "", "", "", "", "Sve što ne spada drugdje"),
    ("Ostalo", "Opis", "(tekst)", "", "", "Kratki opis troška"),

    # ── L1 PRIHODI ─────────────────────────────────────────────────────────
    ("== PRIHODI ==", "", "", "", "", "Svi prihodi"),
    ("(L1) Prihodi", "Iznos",  "(broj EUR)", "", "", ""),
    ("(L1) Prihodi", "Račun",  "Mastercard (Koka), Visa (Saša), Kokin tekući (Zaba), Raiffeisen (Saša)", "", "", "Na koji račun dolazi prihod"),
    ("(L1) Prihodi", "Valuta", "EUR, HRK, USD", "", "", ""),

    ("== Plaća i mirovina ==", "", "", "", "", ""),
    ("Plaća i mirovina", "Vrsta", "Plaća, Mirovina, Prijevoz, Prehrana, Regres, Božićnica, Bonus, I stup, II stup, III stup", "", "", ""),
    ("Plaća i mirovina", "Osoba", "Saša, Koka", "", "", ""),

    ("== Najam — Anja ==", "", "", "", "", "Anjina mjesečna rata"),
    ("Najam — Anja", "Rata", "(tekst)", "", "", "Npr. 81/96"),

    ("== Ostali prihodi ==", "", "", "", "", "Zoranov povrat, honorari, povremeni..."),
    ("Ostali prihodi", "Vrsta",  "Honorar, Povrat komunalnih, Ostalo", "", "", ""),
    ("Ostali prihodi", "Izvor",  "(tekst)", "", "", "Tko plaća / odakle dolazi"),

    # ── TRANSFERI ──────────────────────────────────────────────────────────
    ("== TRANSFERI ==", "", "", "", "", "Prebacivanje novca između računa — nije ni prihod ni rashod"),
    ("Transferi", "Iznos",     "(broj EUR)", "", "", ""),
    ("Transferi", "Izvor",     "Kokin tekući (Zaba), Mastercard (Koka), Sašin tekući (PBZ), Visa (Saša), Raiffeisen (Saša), Gotovina", "", "", ""),
    ("Transferi", "Odredište", "Kokin tekući (Zaba), Mastercard (Koka), Sašin tekući (PBZ), Visa (Saša), Raiffeisen (Saša), Gotovina", "", "", ""),
    ("Transferi", "Napomena",  "(tekst)", "", "", ""),
]

# ── Generiraj Excel ──────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Struktura"

# Header
for ci, h in enumerate(HEADERS, 1):
    c = ws.cell(1, ci, h)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.border = BRD; c.alignment = C
ws.row_dimensions[1].height = 20

# Data
for ri, (kat, attr, opcije, ovisi, zavrij, napom) in enumerate(ROWS, 2):
    is_section = kat.startswith("==")
    is_l1      = kat.startswith("(L1)")
    is_dep     = bool(ovisi)

    vals = [kat, attr, opcije, ovisi, zavrij, napom]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(ri, ci, v if v else None)
        c.border = BRD
        c.alignment = L

        if is_section:
            c.fill = SEC_FILL
            c.font = SEC_FONT
            if ci == 1:
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif is_l1:
            c.fill = L1_FILL
            c.font = L1_FONT
        elif is_dep:
            c.fill = DEP_FILL
        else:
            c.fill = REG_FILL

    if is_section:
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=5)
        ws.row_dimensions[ri].height = 18

# Širine kolona
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 38

# Freeze header
ws.freeze_panes = ws.cell(2, 1)

# ── Upute sheet ─────────────────────────────────────────────────────────────
wu = wb.create_sheet("Upute")
upute = [
    ("UPUTE ZA POPUNJAVANJE", ),
    ("", ),
    ("Ovaj file je prijedlog organizacije financija u aplikaciji.", ),
    ("Slobodno ga uredi — možeš:", ),
    ("  • preimenovati kategorije (kolona A)", ),
    ("  • dodati ili ukloniti kategorije (dodaj retke)", ),
    ("  • promijeniti opcije dropdowna (kolona C, comma-odvojene)", ),
    ("  • dodati nove DependsOn veze (kolone D i E)", ),
    ("", ),
    ("BOJE:", ),
    ("  PLAVA = naslov sekcije (kategorija višeg reda)", ),
    ("  ZELENA = zajednički atributi (pojavljuju se na svakom rashodu/prihodu)", ),
    ("  ŽUTA = atribut ovisi o drugom atributu (DependsOn)", ),
    ("  BIJELA = standardni atribut bez ovisnosti", ),
    ("", ),
    ("KOLONE:", ),
    ("  A — Kategorija: naziv kategorije (npr. Dom, Svakodnevni)", ),
    ("  B — Atribut: naziv polja za unos (npr. Vrsta, Dućan, Objekt)", ),
    ("  C — Opcije: dropdown vrijednosti odvojene zarezima", ),
    ("         (tekst) = slobodan unos bez dropdowna", ),
    ("         (broj EUR) = numeričko polje", ),
    ("  D — Ovisi o atributu: popuni ako dropdown ovisi o drugom polju", ),
    ("         Primjer: Dućan ovisi o → Vrsta", ),
    ("  E — Za vrijednost: za koju vrijednost roditelja vrijede ove opcije", ),
    ("         Primjer: Dućan za vrijednost → Hrana (znači: prikaži te dućane samo kad je Vrsta=Hrana)", ),
    ("  F — Napomena: za tebe, ne ulazi u aplikaciju", ),
    ("", ),
    ("KAKO DODATI KATEGORIJU:", ),
    ("  1. Dodaj retke s nazivom kategorije u koloni A", ),
    ("  2. U koloni B upiši atribute koje želiš", ),
    ("  3. U koloni C upiši opcije dropdowna (ili (tekst) za slobodan unos)", ),
    ("  4. Pošalji Saši — on učitava u aplikaciju", ),
    ("", ),
    ("PRIMJER DependsOn:", ),
    ("  Dom | Vrsta   | HEP, Plin, Voda, Holding     |           |           ← primarni dropdown", ),
    ("  Dom | Objekt  | Bulatova, Nena, Mama           | Vrsta     | HEP       ← prikazuje se samo kad je Vrsta=HEP", ),
    ("  Dom | Objekt  | Bulatova, Kućište              | Vrsta     | Plin      ← prikazuje se samo kad je Vrsta=Plin", ),
    ("", ),
    ("Ako nešto fali u dropdownu — dodaš 'Ostalo' i uneseš slobodan tekst.", ),
    ("Aplikacija pamti što si unijela i ponudit će to sljedeći put.", ),
]
for ri, row in enumerate(upute, 1):
    c = wu.cell(ri, 1, row[0])
    if ri == 1:
        c.font = Font(bold=True, size=14)
    elif row[0].startswith("  •") or row[0].startswith("  "):
        c.font = Font(size=11)
    elif row[0].endswith(":") and not row[0].startswith(" "):
        c.font = Font(bold=True, size=11)
wu.column_dimensions['A'].width = 85

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Rows: {len(ROWS)} (incl. section headers)")
