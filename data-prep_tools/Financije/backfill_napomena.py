# -*- coding: utf-8 -*-
"""
backfill_napomena.py  (S107j, 2026-07-22)
=========================================
Popuni PRAZNU `Napomena` iz `Izvod opis` (očišćeno od bankovnog boilerplatea) —
da svaki potvrđeni Review redak ima čitljiv opis pa se može ručno klasificirati
Tip/Podtip. NIKAD ne dira ne-praznu Napomenu (P3 — Kokini originali ostaju).

Čišćenje za account retke: makne "Kreditni transfer ... on-line bankarstvom
(m-zaba)" prefiks + IBAN-e; kartični opis ostaje kakav jest (merchant).

Pokretanje (Review zatvoren):
  Financije\\run.bat backfill_napomena.py --dry   → koliko bi popunilo
  Financije\\run.bat backfill_napomena.py         → popuni + backup
"""

import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")

RE_KT = re.compile(
    r'Kreditni transfer (?:nacionalni|nac\.) u eurima on-line bankarstvom'
    r'(?:\s*\((?:m-zaba|mobilne aplikacije)\))?\s*', re.I)
RE_IBAN = re.compile(r'\bHR\d{15,}\b')
RE_WS = re.compile(r'\s{2,}')


def clean_opis(op: str) -> str:
    s = RE_KT.sub('', op)
    s = RE_IBAN.sub('', s)
    s = RE_WS.sub(' ', s).strip(' -')
    return s or op.strip()


def pick_review(args):
    explicit = [a for a in args if not a.startswith('--')]
    if explicit:
        p = Path(explicit[0])
        if not p.exists():
            sys.exit(f'✗ File ne postoji: {p}')
        return p
    cands = sorted([c for c in DATA_DIR.glob('Financije_review_*.xlsx') if '.pre-' not in c.name],
                   key=lambda p: p.stat().st_mtime, reverse=True)
    if not cands:
        sys.exit('✗ Nema Financije_review_*.xlsx')
    return cands[0]


def main():
    args = sys.argv[1:]
    dry = '--dry' in args
    review = pick_review(args)
    print(f'Review: {review.name}{"  [DRY]" if dry else ""}')
    wb = openpyxl.load_workbook(review)
    ws = wb['Review']
    col = {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}
    cN, cI = col['Napomena'], col['Izvod opis']

    filled = 0
    still_empty = 0
    samples = []
    for r in range(2, ws.max_row + 1):
        nap = ws.cell(r, cN).value
        iop = ws.cell(r, cI).value
        if nap not in (None, '') and str(nap).strip():
            continue
        if iop in (None, '') or not str(iop).strip():
            still_empty += 1
            continue
        cleaned = clean_opis(str(iop))
        if len(samples) < 12:
            samples.append((str(iop)[:40], cleaned[:40]))
        if not dry:
            ws.cell(r, cN, cleaned)
        filled += 1

    print(f'\nPopunit će Napomenu iz Izvod opisa: {filled}')
    print(f'Ostaje prazno (nema ni Izvod opisa — pre-izvod/no-text): {still_empty}')
    print('\nUzorak (Izvod opis → očišćena Napomena):')
    for a, b in samples:
        print(f'  {a:40} → {b}')

    if dry:
        print('\n[DRY] Ništa snimljeno.')
        return
    backup = review.with_name(f'{review.stem}.pre-napomena-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        sys.exit(f'✗ Zatvori Review u Excelu i ponovi. (Backup: {backup.name})')
    print(f'\n✔ {filled} Napomena popunjeno. Backup: {backup.name}')


if __name__ == '__main__':
    main()
