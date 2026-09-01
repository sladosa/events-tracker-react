# -*- coding: utf-8 -*-
"""
uskladi_izvod.py — jedan bankovni izvod protiv baze i protiv Kokinog filea. S124.

ZASTO POSTOJI (i sto zamjenjuje)
    `kosara_naplate.py` (S123) je pitao "zasto se kosara `Datum naplate` ne
    zatvara" BEZ izvoda u ruci, pa je mogao samo razdvojiti slucajeve po
    pravilu. S izvodom pitanje nestaje: izmjereno S124, `MC_2026-06.pdf` daje
    48/48 sparenih redaka i suma 1.244,74 u cent, a od 73 retka kosare ostaju
    2 duplikata i 23 krivo datirana. Dakle papir odgovara na sve.

    Ovaj alat zato ne dijagnosticira nego USKLADJUJE, i dijeli nalaz po tome
    TKO ODLUCUJE:

      potvrdjeno   izvod <-> baza spareno 1:1        nitko (samo kontrola u cent)
      za ispravak  u bazi, izvod dokazuje drugacije   mi, mehanicki
      za uvoz      na izvodu, nema u bazi             mi, transa
      PITANJA      razilazi se, ili banka ne zna      Koka

    Kod Koke ide SAMO zadnja sekcija. Stari file joj je postavljao 73 pitanja
    na koja papir odgovara svih 73 — file koji uvijek pita sve gubi povjerenje
    brze nego sto ga stekne.

SIDRO SPARIVANJA: `Izvod opis`
    Izmjereno S124: za MC retke `Izvod opis` je DOSLOVNO opis s izvoda
    (`PAYPAL *TEMU`, `KONZUM P-3200 RATA 4/12`, `LUFTHAN2202242474447 RATA 1/3`).
    Nije izvedena vrijednost nego prepisana, pa je to najjaci link koji imamo.

    => `Izvod opis` je de facto oznaka POTVRDJENO IZVODOM. Nitko je dosad nije
       citao. Pokrivenost: Visa 96 %, Racun 92 %, Mastercard 91 %. Po mjesecima
       kupovine (MC): 04/2026 36 ima / 3 nema, 05 31/2, 06 47/5, **07 0 od 22**
       — nula jer `MC_2026-07` nije obradjen.

    ⚠ Prazan `Izvod opis` u razdoblju koje izvod POKRIVA je pitanje, ne greska:
      ili je duplikat, ili banka za taj trosak ne zna.

TRI STANJA, NE DVA
    prazan                         Kokina tvrdnja; iznos/datum/oblik privremeni
    popunjen                       banka potvrdila; njen iznos i datum
    prazan a razdoblje pokriveno   PITANJE

⚠ 1:N — ISTI DOGADJAJ, RAZLICIT BROJ REDAKA (razred S114)
    Banka za Kokin jedan redak zna imati N. Izmjereno: `LH 1/3` 63,33 kod nje =
    `LUFTHAN...447 RATA 1/3` 62,01 + `NAKNADA ZA OBROCNU OTPLATU` 1,32 kod
    banke. Dedup po `(datum, iznos)` to NE MOZE vidjeti — kljuc se razlikuje pa
    udju oba. Alat zato pokusava nespareni iznos iz baze SLOZITI iz 2-3
    nesparena retka izvoda u prozoru, i to PRIJAVLJUJE, nikad ne spaja sam.
    ⚠ Postoji i obrnut smjer (ZABA `Anja 73/96`: jedan event nosi uplatu 450,00
      I isplatu 0,70, i to nije greska nego vjeran spoj dvaju redaka izvoda) —
      zato samo prijava.

PRAVILO ZA 1:N (Sasina odluka S124, ispravljena mjerenjem)
    Bankini redci su KOSTUR (iznos, datum, klasifikacija, potvrda), Kokin redak
    DOPUNJAVA (opis, `Rate?`/`Broj rata`/`Rata br`) i zatim nestaje.
    Nikad ne ostaju oba.
    ⚠ Smjer je bio obrnut u prvoj skici. Mjereno: Kokin `LH 1/3` nosi
      `Tip = N/A`, bez `Podtip`, bez `Izvod opis`, datum 30.06.; bankin nosi
      `Putovanja / Karte, osiguranje` + potvrdu + tocan datum 28.06. Zadrzati
      njen redak znacilo bi zadrzati prazniji.
    ⚠ Ciljni oblik VEC POSTOJI u podacima: ostalih 8 rata od 28.06. su jedan
      redak s njenim opisom + bankinim iznosom + klasifikacijom + `Izvod opis`
      + ratom. Pipeline taj spoj radi za 1:1; padne samo na 1:N.

⚠ STATUS SE NE MIJENJA PO PRAVILU NEGO KAO POSLJEDICA POTVRDE
    Odbaceni automat je bio "dospjelo => izvrseno" (dospijece nije dokaz da je
    banka naplatila). Ovdje dokaz nije dospijece nego izvod, pa `Status` prelazi
    u `Izvrsen` SAMO na retku kojem istovremeno upisujemo `Izvod opis` s tog
    izvoda. Redak koji ne mozemo ozigosati ne diramo.

⚠ NE PUSTATI NA RAZDOBLJE CIJI IZVOD NIJE STIGAO
    Ondje je svaki redak legitimno nepotvrdjen, pa bi sekcija PITANJA bila
    cijeli mjesec. Alat radi nad JEDNIM izvodom i prozor uzima iz njega.

Pokretanje:
    python uskladi_izvod.py --izvod ..\\..\\data-prep_data\\Financije\\izvodi\\MC_2026-07.pdf --dry
    python uskladi_izvod.py --izvod <pdf> --dry --env test
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timedelta
from itertools import combinations
from pathlib import Path

from openpyxl.styles import Font, PatternFill
import openpyxl
import pdfplumber

ROOT = Path(__file__).resolve().parents[2]
CAT_PROD = '986a4612-86a2-49fa-b73f-a29e048e5750'
KOKA_DEFAULT = ROOT / 'data-prep_data' / 'Financije' / 'Financije 2026-08-23.xlsx'


# -- DB ----------------------------------------------------------------------
def load_env(which):
    fn = '.env.prod.local' if which == 'prod' else '.env.testing'
    path = ROOT / fn
    if not path.exists():
        sys.exit('Nema ' + fn + ' — bez njega alat ne moze citati bazu.')
    env = {}
    for line in path.read_text(encoding='utf-8').splitlines():
        if '=' in line and not line.strip().startswith('#'):
            k, v = line.split('=', 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    url = env.get('SUPABASE_URL') or env.get('VITE_SUPABASE_URL')
    key = env.get('SUPABASE_SERVICE_ROLE_KEY') or env.get('VITE_SUPABASE_ANON_KEY')
    if not url or not key:
        sys.exit(fn + ' nema SUPABASE_URL / kljuc.')
    return url, key


def rest(url, key, path):
    """PostgREST reze na 1000 redaka BEZ GRESKE, a paginacija bez `order` je
    tiho pogresna — stranice se preklope i istovremeno preskoce (S108)."""
    out, off = [], 0
    while True:
        req = urllib.request.Request(
            url + '/rest/v1/' + path + '&order=id',
            headers={'apikey': key, 'Authorization': 'Bearer ' + key,
                     'Range': str(off) + '-' + str(off + 999)})
        rows = json.load(urllib.request.urlopen(req))
        out += rows
        off += len(rows)
        if len(rows) < 1000:
            return out


def load_db(url, key):
    defs = {d['id']: d['name'] for d in rest(
        url, key, 'attribute_definitions?category_id=eq.' + CAT_PROD + '&select=id,name')}
    events = {e['id']: e for e in rest(
        url, key, 'events?category_id=eq.' + CAT_PROD
        + '&select=id,event_date,session_start,comment,user_id')}
    vals = defaultdict(dict)
    for a in rest(url, key,
                  'event_attributes?attribute_definition_id=in.(' + ','.join(defs) + ')'
                  '&select=event_id,attribute_definition_id,value_text,value_number,'
                  'value_datetime,value_boolean'):
        n = defs.get(a['attribute_definition_id'])
        if not n:
            continue
        v = a['value_text']
        for alt in ('value_number', 'value_datetime', 'value_boolean'):
            if v is None:
                v = a[alt]
        vals[a['event_id']][n] = v
    return [dict(e, attrs=vals.get(eid, {})) for eid, e in events.items()]


def net(attrs):
    return round(float(attrs.get('Isplata') or 0) - float(attrs.get('Uplata') or 0), 2)


def ev_date(r):
    return datetime.strptime(r['event_date'], '%Y-%m-%d').date()


# -- izvod -------------------------------------------------------------------
AMT = r'(-?\d{1,3}(?:\.\d{3})*,\d{2})'
ROW_RE = re.compile(r'^(B\d{10,20})\s+(\d{2}\.\d{2}\.\d{4})\.\s+(.+?)\s+' + AMT + r'\s*$')


def to_f(s):
    return round(float(s.replace('.', '').replace(',', '.')), 2)


def clean_opis(s):
    """DB drzi opis odrezan na prvom zarezu kod tecajnih redaka — izmjereno:
    izvod `AUDIBLE*P99FS38Y3, TECAJ ZABA 22.06.2026. 1 8,99 USD` -> DB
    `AUDIBLE*P99FS38Y3`. Bez ovoga se tecajni redci nikad ne spare."""
    s = re.split(r',\s*TE[C\u010c]AJ\b', s, flags=re.I)[0]
    return re.sub(r'\s+', ' ', s).strip()


def parse_mc(path):
    lines = []
    with pdfplumber.open(path) as pdf:
        for p in pdf.pages:
            lines += (p.extract_text() or '').splitlines()
    dosp = total = None
    rows, card = [], None
    for ln in lines:
        ln = ln.strip()
        m = re.search(r'Datum dospije[c\u0107]a:\s*(\d{2}\.\d{2}\.\d{4})', ln)
        if m:
            dosp = datetime.strptime(m.group(1), '%d.%m.%Y').date()
        m = re.search(r'UKUPNO \(EUR\):\s*' + AMT, ln)
        if m:
            total = to_f(m.group(1))
        m = re.search(r'Kartica broj:\s*\S+\s+(.+?)\s*$', ln)
        if m:
            card = m.group(1).strip()
            continue
        if ln.startswith('UKUPNO'):
            continue
        m = ROW_RE.match(ln)
        if m:
            rows.append({'ref': m.group(1),
                         'datum': datetime.strptime(m.group(2), '%d.%m.%Y').date(),
                         'opis': clean_opis(m.group(3)),
                         'iznos': to_f(m.group(4)),
                         'kartica': card})
    if dosp is None or total is None:
        sys.exit(path.name + ': ne mogu procitati datum dospijeca / UKUPNO.')
    got = round(sum(r['iznos'] for r in rows), 2)
    if abs(got - total) > 0.005:
        # Parser koji procita 47 od 48 redaka daje uvjerljiv i nepotpun nalaz.
        sys.exit(path.name + ': parsirano ' + str(len(rows)) + ' redaka suma '
                 + format(got, '.2f') + ', a na papiru pise ' + format(total, '.2f')
                 + ' (razlika ' + format(got - total, '+.2f') + ') — STOP.')
    return {'dospijece': dosp, 'total': total, 'rows': rows,
            'od': min(r['datum'] for r in rows), 'do': max(r['datum'] for r in rows)}


# -- Kokin file --------------------------------------------------------------
def norm(s):
    return unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode().lower().strip()


def parse_koka_date(v):
    """103 njena retka nose datum kao TEKST (`11.05.23.`, `28.6.23.`) — alat
    koji prima samo `datetime` progutao bi ih bez ijedne poruke (S116)."""
    if v in (None, ''):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = re.match(r'^\s*(\d{1,2})\.\s*(\d{1,2})\.\s*(\d{2,4})\.?\s*$', str(v))
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    y += 2000 if y < 100 else 0
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def index_koka(path):
    """Kolona C je dan kad novac napusti racun; dok naplata nije poznata C je
    prazan a dan troska stoji u G — zato se gledaju OBJE (S113)."""
    if not path.exists():
        print('⚠ Kokin file nije nadjen (' + path.name + ') — usporedba se preskace.')
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    idx = defaultdict(list)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            c = list(row) + [None] * 10
            for raw in (c[2], c[6]):
                d = parse_koka_date(raw)
                if not d:
                    continue
                for amt in (c[3], c[4]):
                    try:
                        a = round(float(amt), 2)
                    except (TypeError, ValueError):
                        continue
                    if a:
                        idx[(d, a)].append((sn, r, str(c[1] or ''), str(c[0] or '')))
    return idx


def koka_par(idx, d, iznos, tol=3):
    """Prazno nije dokaz da redak kod nje ne postoji — samo da nije nadjen."""
    for dd in range(-tol, tol + 1):
        hit = idx.get((d + timedelta(days=dd), abs(iznos)))
        if hit:
            return hit
    return []


# -- usklada -----------------------------------------------------------------
def uskladi(iz, db, izvor, tol):
    """Dvije populacije, ne jedna:

      `mc`     svi retci tog Izvora u bazi — protiv njih se spare retci izvoda,
               da se vidi sto stvarno FALI (sekcija 3). Bez prozora: redak s
               krivim `Datum naplate` je i dalje isti trosak.
      `kosara` retci za koje BAZA TVRDI da pripadaju ovom izvodu
               (`Datum naplate` == dospijece). To je skup koji se mora poklopiti
               s papirom; svaki visak u njemu je ispravak, duplikat ili pitanje.

    ⚠ Prozor oko datuma kupovine NE valja kao populacija: uz toleranciju u nju
      upadnu susjedni izvodi, pa 12 uredno potvrdjenih svibanjskih redaka
      ispadne kao "pitanje". Izmjereno pri prvom pokretanju (28 laznih pitanja).
    """
    mc = [r for r in db if r['attrs'].get('Izvor') == izvor]
    dosp = iz['dospijece'].isoformat()
    kosara = [r for r in mc if str(r['attrs'].get('Datum naplate') or '')[:10] == dosp]

    # ⚠ `Izvod opis` NIJE jedinstven kroz vrijeme. `ZAGREBPARKING.HR APP 3` s
    #   26,60 postoji u vise mjeseci, pa je prvo pokretanje sparilo lipanjski
    #   redak izvoda s retkom iz RUJNA 2025. i predlozilo mu pomak `event_date`
    #   devet mjeseci unaprijed — a pravi lipanjski redak gurnulo u "pitanja".
    #   Sidro govori KOJI TRGOVAC, ne KOJE POJAVLJIVANJE => obavezan i prozor.
    lo, hi = iz['od'] - timedelta(days=tol), iz['do'] + timedelta(days=tol)
    st_left = list(iz['rows'])

    # ⚠ POTVRDJEN REDAK PRIPADA TOCNO JEDNOM IZVODU. Bez ovog uvjeta je
    #   MC_2026-07 "spario" `TV zabava` 6,99 od 26.06. (koju je vec potvrdio
    #   MC_2026-06) sa svojim `PRIME VIDEO` od 26.07. i predlozio da joj se
    #   `event_date` pomakne mjesec dana naprijed — dakle ISPRAVAK KOJI KVARI.
    #   Kandidat je zato redak koji je ili nepotvrdjen, ili potvrdjen a vec
    #   pripisan ovom dospijecu.
    db_left = [r for r in mc if lo <= ev_date(r) <= hi
               and (not r['attrs'].get('Izvod opis')
                    or str(r['attrs'].get('Datum naplate') or '')[:10] == dosp)]
    pairs = []

    # 1. sidro: `Izvod opis` + iznos, unutar prozora, najblizi datum pobjedjuje
    for s in list(st_left):
        best = None
        for r in db_left:
            io = r['attrs'].get('Izvod opis')
            if not io or norm(io) != norm(s['opis']) or net(r['attrs']) != s['iznos']:
                continue
            dd = abs((ev_date(r) - s['datum']).days)
            if best is None or dd < best[0]:
                best = (dd, r)
        if best:
            pairs.append((s, best[1], 'Izvod opis'))
            st_left.remove(s)
            db_left.remove(best[1])

    # 1b. RATE: iznos + broj rate `n/N`. ⚠ Datum tu NE VALJA kao veza — Koka
    #     ratu datira na dospijece (11.07.), banka na dan terecenja (29.07.),
    #     dakle 18 dana razlike; s tolerancijom od 5 dana ispale bi kao "za
    #     uvoz" i uvoz bi ih UDVOSTRUCIO. `RATA n/N` stoji na obje strane
    #     (u tekstu izvoda i u `Rata br`/`Broj rata`) i jednoznacna je.
    for s in list(st_left):
        m = re.search(r'\bRATA\s+(\d+)\s*/\s*(\d+)', s['opis'], re.I)
        if not m:
            continue
        n, N = int(m.group(1)), int(m.group(2))
        for r in db_left:
            a = r['attrs']
            if (a.get('Rate?') is True and net(a) == s['iznos']
                    and int(a.get('Rata br') or 0) == n
                    and int(a.get('Broj rata') or 0) == N):
                pairs.append((s, r, 'rata n/N'))
                st_left.remove(s)
                db_left.remove(r)
                break

    # 2. po iznosu blizu datuma — redak koji izvod jos nije ozigosao
    for s in list(st_left):
        best = None
        for r in db_left:
            if net(r['attrs']) != s['iznos']:
                continue
            dd = abs((ev_date(r) - s['datum']).days)
            if dd <= tol and (best is None or dd < best[0]):
                best = (dd, r)
        if best:
            pairs.append((s, best[1], 'iznos+datum'))
            st_left.remove(s)
            db_left.remove(best[1])

    # 2b. POTVRDJEN REDAK S KRIVIM `Datum naplate`. Filtar kandidata gore stiti
    #     od sparivanja preko izvoda, ali time sakrije i redak koji je potvrdjen
    #     a nosi krivo dospijece — takav ispadne kao "za uvoz", a uvoz bi ga
    #     UDVOSTRUCIO. Izmjereno: `Kokin Temu` 20,72 (14.01.2026.) nosi
    #     `Izvod opis = PAYPAL *TEMU`, ali `Datum naplate = 2026-01-14` umjesto
    #     11.02. — dakle dan kupnje, pravilo za `Racun` na kartici.
    #     ⚠ Prag su 2 dana, ne tolerancija: na 30 dana bi se vratio bas onaj
    #       lazni par (`TV zabava` 26.06. <-> `PRIME VIDEO` 26.07.) zbog kojeg
    #       filtar i postoji. Isti opis + isti iznos + isti dan je ista
    #       transakcija; isti opis mjesec dana kasnije je sljedeca pretplata.
    zabranjeni = [r for r in mc if lo <= ev_date(r) <= hi
                  and r['attrs'].get('Izvod opis')
                  and str(r['attrs'].get('Datum naplate') or '')[:10] != dosp]
    for s in list(st_left):
        for r in zabranjeni:
            a = r['attrs']
            if (norm(a['Izvod opis']) == norm(s['opis']) and net(a) == s['iznos']
                    and abs((ev_date(r) - s['datum']).days) <= 2):
                pairs.append((s, r, 'potvrdjen, krivo dospijece'))
                st_left.remove(s)
                zabranjeni.remove(r)
                break

    spareni_db = {id(r) for _, r, _ in pairs}
    visak = [r for r in kosara if id(r) not in spareni_db]

    # 3. 1:N — visak iz kosare slozen iz 2-3 retka izvoda koji su VEC spareni
    #    drugdje. ⚠ Ovo NIJE korak sparivanja nego detektor duplikata: kod 1:N
    #    su bankini redci u bazi i vec spareni, pa trazenje samo medju
    #    nesparenima ne nadje nista (prvo pokretanje: 0 pogodaka za `LH 1/3`).
    #    ⚠ Izvor NISU samo retci kosare. Za izvod koji jos nije obradjen kosara
    #      je PRAZNA (nijedan redak ne nosi to dospijece), pa bi provjera nad
    #      njom propustila duplikat U NASTAJANJU: `LH 2/3` 63,33 postoji u bazi
    #      samo u Kokinom spojenom obliku, a transa tek donosi bankin razdvojeni
    #      62,01 + 1,32. Uhvatiti ga PRIJE uvoza je jedini jeftin trenutak.
    #    ⚠ ZBROJ SAM PO SEBI NIJE DOKAZ. Bez ogranicenja je provjera "nasla" da
    #      je `LH 2/3` 63,33 = PEVEX 24,29 (02.07.) + TEMU 21,26 (24.07.) +
    #      KONZUM 17,78 (29.07.) — cista slucajnost preko 27 dana, i uz nju
    #      `LH 1/3` (rata 1/3) sparen s bankinim retcima za ratu 2/3.
    #      Zato dva uvjeta koja opisuju sto banka STVARNO radi:
    #        a) razdvojeni redci su ISTOG DANA na izvodu (glavnica + naknada),
    #        b) je li redak rata odlucuje BROJ RATE, ne datum — Koka ratu datira
    #           na dospijece, banka na dan terecenja (18 dana razlike).
    slozeni = []
    izvor_kandidati = visak + [r for r in db_left if r not in visak]
    potroseni = set()
    po_danu = defaultdict(list)
    for s in iz['rows']:
        po_danu[s['datum']].append(s)
    for r in list(izvor_kandidati):
        a = r['attrs']
        n, rd = net(a), ev_date(r)
        rata = (int(a.get('Rata br') or 0), int(a.get('Broj rata') or 0)) if a.get('Rate?') else None
        hit = None
        for dan, grupa in po_danu.items():
            if rata is None and abs((dan - rd).days) > tol:
                continue
            slobodni = [s for s in grupa if s['ref'] not in potroseni]
            for k in (2, 3):
                for combo in combinations(slobodni, k):
                    if abs(sum(c['iznos'] for c in combo) - n) >= 0.005:
                        continue
                    if rata is not None:
                        nn = [re.search(r'\bRATA\s+(\d+)\s*/\s*(\d+)', c['opis'], re.I)
                              for c in combo]
                        if not any(m and (int(m.group(1)), int(m.group(2))) == rata
                                   for m in nn):
                            continue
                    hit = combo
                    break
                if hit:
                    break
            if hit:
                break
        if hit:
            # Retci izvoda se troše — inače oba `LH 1/3` pokažu isti bankin par
            # (447 dvaput), pa izgleda kao da dva viška imaju jedan uzrok.
            for c in hit:
                potroseni.add(c['ref'])
            slozeni.append((r, list(hit)))
            if r in visak:
                visak.remove(r)

    # 4. redak izvoda koji nema para U PROZORU, ali isti trosak postoji u bazi
    #    izvan njega. ⚠ Bez ove provjere bi ispao kao "za uvoz", a uvoz bi ga
    #    UDVOSTRUCIO — dedup po `(datum, iznos)` ga ne bi vidio jer je datum
    #    upravo ono sto je krivo (razred S115: ispravak + uvoz udvostrucuje tiho).
    daleki = {}
    ost = [r for r in mc if id(r) not in {id(x) for _, x, _ in pairs}]
    for s in st_left:
        for r in ost:
            if net(r['attrs']) != s['iznos']:
                continue
            dd = abs((ev_date(r) - s['datum']).days)
            # Prag je 20 dana, ne 45: mjesecne pretplate su 28-31 dan razmaknute
            # pa su na 45 dana davale 16 "kandidata" za jedan `iCloud 2,99` i
            # zatrpale nalaz. Blizi par je stvarna sumnja, mjesecni nije.
            if dd <= 20:
                daleki.setdefault(s['ref'], []).append((r, dd))
    return {'pairs': pairs, 'slozeni': slozeni, 'izvod_bez_para': st_left,
            'visak': visak, 'kosara': kosara, 'mc': mc, 'daleki': daleki}


def ispravci(iz, pairs):
    """Sto na sparenom retku ne odgovara izvodu. `Status` se mijenja SAMO uz
    istovremeni upis `Izvod opis` — nikad kao zakljucak iz dospijeca."""
    out = []
    for s, r, _ in pairs:
        a, promjene = r['attrs'], []
        dn = str(a.get('Datum naplate') or '')[:10]
        if dn != iz['dospijece'].isoformat():
            promjene.append(('Datum naplate', dn or '—', iz['dospijece'].isoformat()))
        if not a.get('Izvod opis'):
            promjene.append(('Izvod opis', '—', s['opis']))
            # Status prelazi SAMO zajedno sa zigom; bez ziga se ne dira.
            if a.get('Status') == 'Planiran':
                promjene.append(('Status', 'Planiran', 'Izvrsen'))
        elif a.get('Status') == 'Planiran':
            promjene.append(('Status', 'Planiran', 'Izvrsen'))
        # ⚠ RATI SE `event_date` NE DIRA, ni "poslije". Pravilo je da sve rate
        #   jedne kupovine dijele `event_date` = DAN KUPNJE (CLAUDE.md), a datum
        #   na izvodu je dan TERECENJA (29.07.) — dakle izvod tu nije autoritet
        #   za `event_date` nego samo za `Datum naplate`. Bez ove iznimke se
        #   9 rata pojavi u popisu "za poslije" kao da je posao koji ceka.
        if ev_date(r) != s['datum'] and a.get('Rate?') is not True:
            promjene.append(('event_date', r['event_date'], s['datum'].isoformat()))
        if promjene:
            out.append((s, r, promjene))
    return out


# -- review file -------------------------------------------------------------
ATTR_ORDER = ['Racun', 'Izvor', 'Smjer', 'Uplata', 'Isplata', 'Tip', 'Podtip',
              'Izvod opis', 'Rate?', 'Broj rata', 'Rata br', 'Datum naplate',
              'Status', 'Stanje', 'Valuta']
DIAG = ['Sekcija', 'Sto mijenjamo', 'Bilo je', 'Dokaz s izvoda', 'Kokin redak']
HDR_FILL = PatternFill('solid', fgColor='FF4472C4')
DIAG_FILL = PatternFill('solid', fgColor='FFFFF2CC')
DEL_FILL = PatternFill('solid', fgColor='FFFFC7CE')
OK_FILL = PatternFill('solid', fgColor='FFC6EFCE')
BOLD_W = Font(bold=True, color='FFFFFFFF')

# ⚠ `event_date` se u ovom fileu NE MIJENJA, iako ga izvod ponegdje demantira.
#   Uvoz ga zna promijeniti (`excelImport.ts:1326`), ali time pomice i
#   `session_start`, a `useActivities` grupira po (user, kategorija,
#   session_start) => dva retka iste minute postaju JEDAN redak liste.
#   Vrijednost tih pomaka je mala (na MC retcima ne diraju saldo), a rizik i
#   uocljivost nisu. Prijedlozi se ispisuju u `Pregled` kao "za poslije".
SKIP_FIELDS = {'event_date'}


def build_rows(rezultati):
    """Sto ide u importabilni list. Vraca (redci, pomaci_datuma, pitanja, uvoz).

    ⚠ Brise se SAMO `LH 1/3` — njegova bankina cetiri retka su vec u bazi.
      `LH 2/3` se NE brise sada: bankini redci za njega dolaze tek transom 4,
      pa bi brisanje ostavilo rupu od 126,66 do tada. Brisanje i uvoz idu
      jednim potezom ili nikako.
    """
    redci, pomaci, pitanja, uvoz, odgodjeni = [], [], [], [], []
    vidjeni = set()
    for iz, u, isp in rezultati:
        for s, r, ch in isp:
            polja = [(f, o, n) for f, o, n in ch if f not in SKIP_FIELDS]
            for f, o, n in ch:
                if f in SKIP_FIELDS:
                    pomaci.append((r, f, o, n, s))
            if polja and r['id'] not in vidjeni:
                vidjeni.add(r['id'])
                redci.append({'r': r, 'ch': polja, 's': s, 'akcija': 'ispravak'})
        for r, combo in u['slozeni']:
            # Kokin spojeni redak: brise se samo ako su bankini vec u bazi.
            bankini_u_bazi = all(
                any(net(x['attrs']) == c['iznos']
                    and x['attrs'].get('Izvod opis')
                    and norm(x['attrs']['Izvod opis']) == norm(c['opis'])
                    for x in u['mc'])
                for c in combo)
            if r['id'] in vidjeni:
                continue
            if not bankini_u_bazi:
                # Duplikat U NASTAJANJU: Kokin spojeni redak je zasad JEDINI
                # zapis tog troska, bankini dolaze tek transom. Brisati ga sada
                # znaci ostaviti rupu; brisanje i uvoz idu jednim potezom.
                odgodjeni.append((r, combo))
                vidjeni.add(r['id'])
                continue
            vidjeni.add(r['id'])
            redci.append({'r': r, 'ch': [], 's': None, 'akcija': 'brisanje',
                          'combo': combo})
            # i obogacivanje bankinih redaka ratom s Kokinog
            a = r['attrs']
            if a.get('Rate?') is not True:
                continue
            for c in combo:
                if not re.search(r'\bRATA\b', c['opis'], re.I):
                    continue
                for x in u['mc']:
                    xa = x['attrs']
                    if (net(xa) == c['iznos'] and xa.get('Izvod opis')
                            and norm(xa['Izvod opis']) == norm(c['opis'])
                            and x['id'] not in vidjeni):
                        vidjeni.add(x['id'])
                        redci.append({'r': x, 's': c, 'akcija': 'dopuna',
                                      # ⚠ OPIS SE NE DIRA. Prepisati bankin
                                      # `LUFTHAN...447 RATA 1/3` Kokinim
                                      # `LH 1/3` dalo bi DVA IDENTICNA retka
                                      # istog dana i istog iznosa — a to u
                                      # listi izgleda kao duplikat, dakle
                                      # tocno ono sto ovim ciscenjem micemo.
                                      # Bankin tekst je jedinstven (447/448) i
                                      # vec sam kaze "RATA 1/3".
                                      'ch': [('Rate?', '—', 'True'),
                                             ('Broj rata', '—', a.get('Broj rata')),
                                             ('Rata br', '—', a.get('Rata br'))],
                                      'iz_retka': r})
                        break
        pitanja += [(iz, r) for r in u['visak']]
        uvoz += [(iz, s) for s in u['izvod_bez_para']]

    # ⚠ Visak jednog izvoda cesto je posao SLJEDECEG. MC_2026-06 prijavi 23
    #   retka kao "banka ih nema" — a 21 od njih MC_2026-07 preuzme kao ispravak
    #   i 2 kao duplikat. Kad bi se visak prenio sirov, Koka bi dobila 30 pitanja
    #   umjesto 7, i to bas ona na koja vec imamo odgovor. Zato se filtrira tek
    #   kad su SVI izvodi obradjeni, i to protiv svega sto je file dirnuo.
    rijeseni = {x['r']['id'] for x in redci}
    rijeseni |= {r['id'] for r, _, _, _, _ in pomaci}
    rijeseni |= {r['id'] for r, _ in odgodjeni}
    vid = set()
    cisto = []
    for iz, r in pitanja:
        if r['id'] in rijeseni or r['id'] in vid:
            continue
        vid.add(r['id'])
        cisto.append((iz, r))
    return redci, pomaci, cisto, uvoz, odgodjeni


def tekst(ws, r, c, v):
    """⚠ openpyxl string koji pocinje s `=` sprema kao FORMULU, a Excel je onda
    ne moze parsirati i nudi "repair" — file se ne otvori. Ulovljeno na
    `Pregled!A31`: objasnjenje se prelomilo tako da je redak poceo s
    `= 63,33), ali razdvojeno...`. Isto vrijedi za `+`, `-` i `@`.
    Svaka pripovjedna celija ide kroz ovo, ne kroz `ws.cell(r, c, v)`."""
    cell = ws.cell(r, c)
    cell.value = v
    if isinstance(v, str) and v[:1] in ('=', '+', '-', '@'):
        cell.data_type = 's'
    return cell


def write_file(out, rezultati, koka, emails):
    redci, pomaci, pitanja, uvoz, odgodjeni = build_rows(rezultati)
    wb = openpyxl.Workbook()

    # ---- Events (jedini list koji se uvozi) --------------------------------
    ws = wb.active
    ws.title = 'Events'
    ws.cell(1, 1, 'ATTRIBUTE LEGEND:').font = Font(bold=True)
    for i, h in enumerate(['Col', 'Area', 'Category_Path', 'Attribute', 'Type', 'Unit'], 1):
        c = ws.cell(2, i, h)
        c.font, c.fill = BOLD_W, HDR_FILL
    tipovi = {'Uplata': 'number', 'Isplata': 'number', 'Rate?': 'boolean',
              'Broj rata': 'number', 'Rata br': 'number',
              'Datum naplate': 'datetime', 'Stanje': 'number'}
    for i, name in enumerate(ATTR_ORDER):
        ws.cell(3 + i, 1, openpyxl.utils.get_column_letter(9 + i))
        ws.cell(3 + i, 2, 'Financije_all')
        ws.cell(3 + i, 3, 'Transakcija')   # ⚠ BEZ imena aree (Activities format)
        ws.cell(3 + i, 4, name)
        ws.cell(3 + i, 5, tipovi.get(name, 'text'))

    r0 = 3 + len(ATTR_ORDER) + 1
    ws.cell(r0, 1, 'EVENT DATA:').font = Font(bold=True)
    hdr = (['event_id', 'Area', 'Category_Path', 'event_date', 'session_start',
            'created_at', 'User', 'comment'] + ATTR_ORDER + ['Delete?'] + DIAG)
    for i, h in enumerate(hdr, 1):
        c = ws.cell(r0 + 1, i, h)
        c.font = BOLD_W
        c.fill = HDR_FILL if i <= 9 + len(ATTR_ORDER) else PatternFill('solid', fgColor='FFBF8F00')

    for j, item in enumerate(redci):
        rr, r, a = r0 + 2 + j, item['r'], item['r']['attrs']
        nove = {f: n for f, _, n in item['ch']}
        ws.cell(rr, 1, r['id'])                      # UUID => UPDATE, ne INSERT
        ws.cell(rr, 2, 'Financije_all')
        ws.cell(rr, 3, 'Transakcija')
        ws.cell(rr, 4, r['event_date'])
        ws.cell(rr, 5, (r['session_start'] or '')[11:16]).number_format = '@'
        ws.cell(rr, 7, emails.get(r['user_id'], ''))  # ⚠ bez emaila redak je „tudji"
        ws.cell(rr, 8, nove.get('comment', r['comment'] or ''))
        for i, name in enumerate(ATTR_ORDER):
            v = nove[name] if name in nove else a.get(name)
            cell = ws.cell(rr, 9 + i)
            if name == 'Rate?':
                cell.value = True if v in (True, 'True', 'true') else (False if v is not None else None)
            elif name == 'Datum naplate':
                d = str(v)[:10] if v else ''
                # ⚠ prava datumska celija, usidrena u PODNE — v. excelDatetime.ts
                cell.value = datetime.strptime(d, '%Y-%m-%d').replace(hour=12) if d else None
                cell.number_format = 'dd.mm.yyyy'
            else:
                cell.value = v
            if name in nove:
                cell.fill = OK_FILL
        base = 9 + len(ATTR_ORDER)
        if item['akcija'] == 'brisanje':
            ws.cell(rr, base, 'DELETE').fill = DEL_FILL
        ws.cell(rr, base + 1, {'ispravak': 'ISPRAVAK', 'brisanje': 'BRISANJE',
                               'dopuna': 'DOPUNA'}[item['akcija']])
        if item['akcija'] == 'brisanje':
            ws.cell(rr, base + 2, 'cijeli redak — duplikat')
            ws.cell(rr, base + 3, format(net(a), '.2f') + ' = '
                    + ' + '.join(format(c['iznos'], '.2f') for c in item['combo']))
            ws.cell(rr, base + 4, ' | '.join(c['opis'] for c in item['combo']))
        else:
            ws.cell(rr, base + 2, ', '.join(f for f, _, _ in item['ch']))
            ws.cell(rr, base + 3, ' | '.join(str(o) for _, o, _ in item['ch']))
            ws.cell(rr, base + 4, (item['s']['opis'] + '  '
                                   + item['s']['datum'].strftime('%d.%m.%Y'))
                    if item['s'] else '')
        k = koka_par(koka, ev_date(r), net(a))
        ws.cell(rr, base + 5, (k[0][0] + ' r.' + str(k[0][1]) + '  "' + k[0][2][:30] + '"')
                if k else '')
        for cc in range(base + 1, base + 1 + len(DIAG)):
            if ws.cell(rr, cc).fill.fgColor.rgb in (None, '00000000'):
                ws.cell(rr, cc).fill = DIAG_FILL

    for i, w in enumerate([38, 13, 13, 12, 11, 10, 30, 30], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for i in range(9, len(hdr) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 17
    ws.freeze_panes = ws.cell(r0 + 2, 1)
    # ⚠ Kolona izvan autofiltera se pri prvom sortu raspari od svog retka.
    ws.auto_filter.ref = ('A' + str(r0 + 1) + ':'
                          + openpyxl.utils.get_column_letter(len(hdr))
                          + str(r0 + 1 + len(redci)))

    # ---- Pitanja (samo za citanje) -----------------------------------------
    wq = wb.create_sheet('Pitanja')
    wq['A1'] = 'Pitanja za Koku — retci koje banka nema ni na jednom izvodu'
    wq['A1'].font = Font(bold=True, size=12)
    for i, t in enumerate([
            '', 'Ovi retci su u bazi, ali ih nijedan bankovni izvod ne potvrdjuje.',
            'Nije nuzno greska — mozda je iznos ili datum malo drukciji pa se ne',
            'poklope. Ali svaki od njih treba jedno "da, to je bilo" ili ispravak.',
            '', 'Prazan komentar znaci da redak nema opis — nije se izgubio, nikad ga',
            'nije ni imao.', ''], start=2):
        tekst(wq, i, 1, t)
    hq = ['datum', 'opis', 'iznos', 'Status', 'racun', 'Tip / Podtip',
          'izvod koji ga je trazio', 'Kokin redak', 'event_id']
    for i, h in enumerate(hq, 1):
        c = wq.cell(10, i, h)
        c.font, c.fill = BOLD_W, HDR_FILL
    for j, (iz, r) in enumerate(sorted(pitanja, key=lambda x: x[1]['event_date'])):
        a = r['attrs']
        k = koka_par(koka, ev_date(r), net(a))
        wq.append([]) if False else None
        vals = [r['event_date'], r['comment'] or '(bez opisa)', net(a),
                a.get('Status'), a.get('Racun'),
                (str(a.get('Tip') or '—') + ' / ' + str(a.get('Podtip') or '—')),
                'dospijece ' + iz['dospijece'].strftime('%d.%m.%Y'),
                (k[0][0] + ' r.' + str(k[0][1])) if k else '—', r['id']]
        for i, v in enumerate(vals, 1):
            wq.cell(11 + j, i, v)
    for i, w in enumerate([12, 30, 10, 11, 22, 30, 24, 18, 38], 1):
        wq.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ---- Za uvoz (najava transe, NE uvozi se ovim fileom) -------------------
    wu = wb.create_sheet('Sto tek dolazi')
    wu['A1'] = 'Retci s izvoda kojih u bazi jos nema — NE uvoze se ovim fileom'
    wu['A1'].font = Font(bold=True, size=12)
    for i, t in enumerate([
            '', 'Ovo je najava, ne zadatak. Ti retci ulaze zasebnim uvozom (transa),',
            'kad se ovaj krug ispravaka slegne. Ovdje su da se vidi da nista nije',
            'zaboravljeno i da se brojka slaze s papirom.', ''], start=2):
        tekst(wu, i, 1, t)
    for i, h in enumerate(['datum', 'opis s izvoda', 'iznos', 'izvod', 'Kokin redak'], 1):
        c = wu.cell(8, i, h)
        c.font, c.fill = BOLD_W, HDR_FILL
    for j, (iz, s) in enumerate(sorted(uvoz, key=lambda x: x[1]['datum'])):
        k = koka_par(koka, s['datum'], s['iznos'])
        for i, v in enumerate([s['datum'], s['opis'], s['iznos'],
                               iz['dospijece'].strftime('%d.%m.%Y'),
                               (k[0][0] + ' r.' + str(k[0][1]) + '  "' + k[0][2][:24] + '"')
                               if k else '—'], 1):
            c = wu.cell(9 + j, i, v)
            if i == 1:
                c.number_format = 'dd.mm.yyyy'
    for i, w in enumerate([12, 40, 10, 12, 34], 1):
        wu.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ---- Pregled (prvi list, ono sto se cita prvo) --------------------------
    wp = wb.create_sheet('Pregled', 0)
    wp['A1'] = 'Uskladjenje s bankovnim izvodima — Mastercard 2026.'
    wp['A1'].font = Font(bold=True, size=14)
    row = 3
    wp.cell(row, 1, 'Svaki izvod je usporedjen redak po redak s bazom.').font = Font(bold=True)
    row += 2
    for i, h in enumerate(['izvod', 'dospijece', 'redaka', 'ukupno s papira',
                           'slaze se?'], 1):
        c = wp.cell(row, i, h)
        c.font, c.fill = BOLD_W, HDR_FILL
    row += 1
    for iz, u, _ in rezultati:
        s_p = sum(x['iznos'] for x, _, _ in u['pairs'])
        s_u = sum(x['iznos'] for x in u['izvod_bez_para'])
        ok = abs(s_p + s_u - iz['total']) < 0.005
        wp.cell(row, 1, iz['ime'])
        wp.cell(row, 2, iz['dospijece'].strftime('%d.%m.%Y'))
        wp.cell(row, 3, len(iz['rows']))
        wp.cell(row, 4, iz['total']).number_format = '#,##0.00'
        c = wp.cell(row, 5, 'DA, u cent' if ok else 'NE — provjeriti')
        c.fill = OK_FILL if ok else DEL_FILL
        row += 1
    row += 1
    for t in ['Sto ovaj file predlaze:',
              '   list `Events`  — ' + str(sum(1 for x in redci if x['akcija'] == 'ispravak'))
              + ' ispravaka, ' + str(sum(1 for x in redci if x['akcija'] == 'dopuna'))
              + ' dopuna, ' + str(sum(1 for x in redci if x['akcija'] == 'brisanje'))
              + ' brisanja  >> TO SE UVOZI',
              '   list `Pitanja` — ' + str(len(pitanja)) + ' redaka za tebe, nista se ne uvozi',
              '   list `Sto tek dolazi` — ' + str(len(uvoz)) + ' redaka, najava sljedeceg uvoza',
              '',
              'Kako uvesti: Activities -> Import -> odaberi ovaj file -> pregledaj -> Apply.',
              'Zeleno obojene celije su ono sto se mijenja. Sve ostalo je vec u bazi',
              'i uvoz ga ne dira.',
              '',
              'Najcesci ispravak je `Datum naplate` — dan kad banka stvarno tereti racun.',
              'Kod Mastercarda je to 11. u mjesecu nakon kupovine, a kod dijela redaka je',
              'bio upisan dan kupovine. Zbog toga se "kosarica" za 11.07. nije zatvarala:',
              'imala je 2.231,02, a banka je skinula 1.244,74.',
              '',
              'BRISANJA (2 retka): tvoj `LH 1/3` 63,33 banka vodi kao DVA retka —',
              'LUFTHANSA rata 62,01 + naknada za obrocnu otplatu 1,32. Oba su vec u bazi,',
              'pa je tvoj treci zapis isti trosak po drugi put. Zbroj je isti',
              '(62,01 plus 1,32 daje 63,33), ali razdvojeno naknada banke ide pod',
              'bankovne troskove, a ne pod putovanja. Broj rate (1/3) se prije',
              'brisanja prepisuje na bankin redak.',
              ] + ([] if not odgodjeni else [
              '',
              'ODGODJENO (' + str(len(odgodjeni)) + ' redaka): isti slucaj kao brisanja gore, ali',
              'bankini redci za njih jos nisu u bazi — dolaze sljedecim uvozom. Brisemo ih',
              'tek tada, jednim potezom, da ne ostane rupa. Popis:']
              + ['   ' + r['event_date'] + '  ' + str(r['comment'])[:24].ljust(26)
                 + format(net(r['attrs']), '8.2f') + '  = '
                 + ' + '.join(format(c['iznos'], '.2f') for c in combo)
                 for r, combo in odgodjeni]) + ([''] if not pomaci else [
              '',
              'NIJE u ovom fileu (za poslije): ' + str(len(pomaci)) + ' redaka ima datum koji se',
              'od izvoda razlikuje za koji dan. Nismo ih dirali jer pomak datuma pomice i',
              'vrijeme zapisa, pa se dva retka istog dana znaju stopiti u jedan. Popis:']
              + ['   ' + r['event_date'] + '  ' + str(r['comment'])[:26].ljust(28)
                 + format(net(r['attrs']), '8.2f') + '   izvod kaze ' + str(n)
                 for r, f, o, n, s in pomaci]):
        tekst(wp, row, 1, t)
        row += 1
    wp.column_dimensions['A'].width = 78
    for col, w in (('B', 13), ('C', 9), ('D', 17), ('E', 16)):
        wp.column_dimensions[col].width = w
    wb.save(out)
    return redci, pomaci, pitanja, uvoz, odgodjeni


# -- ispis -------------------------------------------------------------------
def hr(t=''):
    print('\n' + t)
    print('-' * 100)


def opis_db(r):
    return (r['comment'] or '«bez komentara»')[:34]


def report(iz, u, isp, koka, path):
    dosp = iz['dospijece']
    print('=' * 102)
    print('USKLADA  ' + path.name + '   dospijece ' + dosp.strftime('%d.%m.%Y')
          + '   kupovine ' + iz['od'].strftime('%d.%m.') + '-' + iz['do'].strftime('%d.%m.%Y'))
    print('=' * 102)
    s_p = sum(s['iznos'] for s, _, _ in u['pairs'])
    n_io = sum(1 for _, _, how in u['pairs'] if how == 'Izvod opis')
    print('Izvod  : ' + str(len(iz['rows'])) + ' redaka   ' + format(iz['total'], '.2f')
          + '   (parsirano i provjereno protiv ispisanog UKUPNO)')
    print('Kosara : ' + str(len(u['kosara'])) + ' redaka   '
          + format(sum(net(r['attrs']) for r in u['kosara']), '.2f')
          + '   (baza tvrdi da pripadaju ovom izvodu)')
    print()
    s_u = sum(x['iznos'] for x in u['izvod_bez_para'])
    print('  spareno s izvodom          ' + str(len(u['pairs'])).rjust(3)
          + '   ' + format(s_p, '10.2f')
          + '   (' + str(n_io) + ' preko `Izvod opis`, '
          + str(len(u['pairs']) - n_io) + ' ostalo)')
    print('  jos nije u bazi            ' + str(len(u['izvod_bez_para'])).rjust(3)
          + '   ' + format(s_u, '10.2f'))
    # KONTROLA: spareno + za uvoz mora dati izvod u cent. Manjak nije greska
    # dok je tocno objasnjen sekcijom 3 — greska je kad se ne zbroji.
    print('  ' + '-' * 44)
    if abs(s_p + s_u - iz['total']) < 0.005:
        print('  KONTROLA  ' + format(s_p + s_u, '10.2f')
              + '  == izvod, u cent   (svaki redak izvoda je objasnjen)')
    else:
        print('  KONTROLA  ' + format(s_p + s_u, '10.2f') + '  != izvod '
              + format(iz['total'], '.2f') + '   RAZLIKA '
              + format(s_p + s_u - iz['total'], '+.2f')
              + '  — nista ispod nije pouzdano')

    hr('1 · POTVRDJENO — spareno i tocno datirano, ne dira se')
    ok = [p for p in u['pairs'] if not any(p[1] is r for _, r, _ in isp)]
    print('    ' + str(len(ok)) + ' redaka   '
          + format(sum(s['iznos'] for s, _, _ in ok), '.2f'))

    hr('2 · ZA ISPRAVAK — spareno, ali se polje ne slaze s izvodom   ['
       + str(len(isp)) + ' redaka]   MI, mehanicki')
    if not isp:
        print('    nema')
    for s, r, ch in sorted(isp, key=lambda x: x[1]['event_date']):
        print('  ' + r['event_date'] + '  ' + opis_db(r).ljust(36)
              + format(net(r['attrs']), '9.2f') + '   ' + r['id'][:8]
              + '   izvod: ' + s['opis'][:26])
        for f, old, new in ch:
            print('        ' + f.ljust(15) + str(old)[:22].ljust(24) + '->  ' + str(new)[:36])

    hr('3 · ZA UVOZ — na izvodu, nema ga u bazi   [' + str(len(u['izvod_bez_para']))
       + ' redaka, ' + format(sum(x['iznos'] for x in u['izvod_bez_para']), '.2f')
       + ']   MI, transa')
    if not u['izvod_bez_para']:
        print('    nema')
    for s in sorted(u['izvod_bez_para'], key=lambda x: x['datum']):
        k = koka_par(koka, s['datum'], s['iznos'])
        kk = ('Koka ' + k[0][0] + ' r.' + str(k[0][1]) + ' "' + k[0][2][:20] + '"') if k else 'Koka: nije nadjena'
        print('  ' + s['datum'].strftime('%d.%m.%Y') + '  ' + s['opis'][:34].ljust(36)
              + format(s['iznos'], '9.2f') + '   ' + kk)
        for r, dd in u['daleki'].get(s['ref'], []):
            print('      ⚠ NE UVOZITI SLIJEPO — isti iznos vec u bazi: '
                  + r['event_date'] + ' "' + str(r['comment'])[:24] + '" ('
                  + str(dd) + ' dana razlike, ' + r['id'][:8] + ')')

    hr('4 · DUPLIKAT 1:N — redak baze = zbroj vise redaka izvoda koji su vec u bazi   ['
       + str(len(u['slozeni'])) + ']')
    if not u['slozeni']:
        print('    nema')
    for r, combo in u['slozeni']:
        a = r['attrs']
        print('  VISAK  ' + r['event_date'] + '  ' + opis_db(r).ljust(34)
              + format(net(a), '9.2f') + '   ' + r['id']
              + '   Tip=' + str(a.get('Tip')))
        for c in combo:
            print('   = izvod ' + c['datum'].strftime('%d.%m.') + '  ' + c['opis'][:34].ljust(36)
                  + format(c['iznos'], '9.2f'))
        rt = (str(a.get('Rata br')) + '/' + str(a.get('Broj rata'))) if a.get('Rate?') else '—'
        print('         PRAVILO: bankini redci su kostur; s ovoga preuzeti opis "'
              + str(r['comment']) + '" i ratu ' + rt + ', pa ga obrisati.')

    hr('5 · PITANJA ZA KOKU — u kosari, a banka za njih ne zna   ['
       + str(len(u['visak'])) + ' redaka, '
       + format(sum(net(r['attrs']) for r in u['visak']), '.2f') + ']')
    if not u['visak']:
        print('    nema')
    else:
        print('    Ili im je `Datum naplate` kriv (pripadaju drugom izvodu), ili banka')
        print('    za taj trosak ne zna. Prazan `Izvod opis` = nijedan izvod ih jos nije')
        print('    potvrdio; popunjen = potvrdio ih je NEKI DRUGI izvod, pa je datum kriv.\n')
    for r in sorted(u['visak'], key=lambda x: x['event_date']):
        a = r['attrs']
        k = koka_par(koka, ev_date(r), net(a))
        kk = ('Koka r.' + str(k[0][1])) if k else 'Koka: —'
        print('  ' + r['event_date'] + '  ' + opis_db(r).ljust(34)
              + format(net(a), '9.2f') + '  ' + str(a.get('Status'))[:8].ljust(10)
              + ('potvrdio drugi izvod' if a.get('Izvod opis') else 'NEPOTVRDJEN').ljust(22) + kk)
    print()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--izvod', action='append', required=True,
                    help='moze se ponoviti; s --file svi idu u jedan workbook')
    ap.add_argument('--koka', default=str(KOKA_DEFAULT))
    ap.add_argument('--env', default='prod', choices=['prod', 'test'])
    ap.add_argument('--izvor', default='Mastercard')
    ap.add_argument('--tol', type=int, default=5,
                    help='tolerancija u danima oko prozora izvoda')
    ap.add_argument('--dry', action='store_true', help='samo ispis na ekran')
    ap.add_argument('--file', help='napisi review workbook za Koku')
    args = ap.parse_args()

    paths = [Path(p) for p in args.izvod]
    for p in paths:
        if not p.exists():
            sys.exit('Nema ' + str(p))
        if not p.name.upper().startswith('MC_'):
            sys.exit('Zasad samo MC izvodi (' + p.name + '). Visa/ZABA imaju drugi format.')

    url, key = load_env(args.env)
    db = load_db(url, key)
    koka = index_koka(Path(args.koka))

    rezultati = []
    for p in sorted(paths, key=lambda x: x.name):
        iz = parse_mc(p)
        iz['ime'] = p.name
        u = uskladi(iz, db, args.izvor, args.tol)
        isp = ispravci(iz, u['pairs'])
        rezultati.append((iz, u, isp))
        if args.dry or not args.file:
            report(iz, u, isp, koka, p)

    if not args.file:
        return
    emails = {p['id']: p['email']
              for p in rest(url, key, 'profiles?select=id,email&limit=200')}
    out = Path(args.file)
    redci, pomaci, pitanja, uvoz, odgodjeni = write_file(out, rezultati, koka, emails)
    n = lambda k: sum(1 for x in redci if x['akcija'] == k)  # noqa: E731
    print('\nNapisano: ' + str(out))
    print('  Events          ' + str(len(redci)).rjust(3) + ' redaka  ('
          + str(n('ispravak')) + ' ispravak, ' + str(n('dopuna')) + ' dopuna, '
          + str(n('brisanje')) + ' brisanje)')
    print('  Pitanja         ' + str(len(pitanja)).rjust(3) + ' redaka')
    print('  Sto tek dolazi  ' + str(len(uvoz)).rjust(3) + ' redaka')
    print('  (izostavljeno)  ' + str(len(pomaci)).rjust(3) + ' pomaka `event_date` — v. Pregled')


if __name__ == '__main__':
    main()
