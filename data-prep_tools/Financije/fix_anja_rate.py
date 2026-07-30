# -*- coding: utf-8 -*-
"""
fix_anja_rate.py — 4 rate Anjine posudbe → `Prihodi | Povrat Anja`. S107r, 2026-07-30.

NALAZ (T-S107r-7): od 41 retka s oznakom `X/96` (Anjina posudba, 96 rata po 450 €),
38 je nakon migracije taksonomije u `Prihodi | Povrat Anja`, a 4 su pala u
`Transfer | Anja` jer nisu zadovoljila uvjet mapping reda (`Smjer=Uplata` + 450 €):

  397  `rata 45/96`  Smjer=Isplata, Uplata=450, Isplata=0,30
  3727 `rata 73/96`  Smjer=Isplata, Uplata=450, Isplata=0,70
       → obje kolone iznosa popunjene; `row_amount()` čita `Isplata` PRVO pa vidi 0,30/0,70
  3612 `72/96 (1/2)` Uplata=400
  3613 `72/96 (2/2)` Uplata=50
       → jedna rata plaćena u dva dijela (400+50=450) pa nijedan redak nije 450

Opseg anomalije je zatvoren: u cijelom fileu samo **3** retka imaju popunjene obje kolone
iznosa (397, 3727 i jedna bankovna naknada 0,26/0,17 koja je ispravno klasificirana);
`Smjer` se svugdje inače slaže s popunjenom kolonom ⇒ nema sistemskog rizika za
`Iznos min/max` pravila.

⚠ Zašto jednokratna skripta, a ne pravilo ni `Preimenovanja` red: `Transfer | Anja` je
nakon migracije **valjan** par, a `apply_rules.py` (~linija 516) preskače svaki redak
čiji je par valjan — isto kao u `fix_vocarna_pravilo.py` (S107o). Zato se ovi retci
moraju ispraviti izravno. Fix je stabilan: `Prihodi | Povrat Anja` nije N/A pa ga ni
budući `apply_rules` run ne dira.

⚠ Namjerno se NE dodaje pravilo za budući import: oba obrasca su jednokratna anomalija,
a "Uplata=450 kad Smjer=Isplata" se keyword pravilom ionako ne može izraziti (`row_amount`
čita Isplata prvo).

Ne dira: iznose (0,30/0,70 izgleda kao naknada za transfer), `Smjer`, `Pouzdanost`
(ostaje kakva je bila — VISOKA za Kokine ručne, PRAVILO za pravilom labelirane).

Pokretanje (Excel ZATVOREN):
  python fix_anja_rate.py --dry
  python fix_anja_rate.py              # backup .pre-anjarate-*
"""
from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from apply_rules import find_header_col, pick_file  # noqa: E402

NEW_TIP, NEW_POD = 'Prihodi', 'Povrat Anja'

# (source_key, dio napomene za kontrolu, očekivana Uplata) — traži se po source_key,
# NE po broju retka (retci se pomiču pri sortu/dedupu; nauk iz S107o).
TARGETS = [
    ('fca452ff1fa4', 'Anja 45/96', 450.0),
    ('b0571eae6290', 'Anja 73/96', 450.0),
    ('fa793671341d', '72/96 (1/2)', 400.0),
    ('7cd888fe0847', '72/96 (2/2)', 50.0),
]


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    path = pick_file(args)
    print(f'File: {path.name}{"  [DRY RUN — bez snimanja]" if dry else ""}\n')

    wb = openpyxl.load_workbook(path)
    ws = wb['Review']
    c_key = find_header_col(ws, 'source_key')
    c_tip = find_header_col(ws, 'Tip')
    c_pod = find_header_col(ws, 'Podtip')
    c_nap = find_header_col(ws, 'Napomena')
    c_upl = find_header_col(ws, 'Uplata')
    c_alt = find_header_col(ws, 'Alternativa / nap.')
    c_run = find_header_col(ws, 'Pravilo run')

    by_key = {str(ws.cell(r, c_key).value or '').strip(): r
              for r in range(2, ws.max_row + 1)}
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    changed = 0

    for key, nap_frag, exp_upl in TARGETS:
        r = by_key.get(key)
        if r is None:
            sys.exit(f'✗ source_key {key} nije nađen — je li ovo pravi Review file?')
        nap = str(ws.cell(r, c_nap).value or '')
        if nap_frag not in nap:
            sys.exit(f'✗ red {r} (source_key {key}): Napomena ne sadrži "{nap_frag}" '
                     f'nego "{nap[:50]}" — PREKID, ništa nije upisano.')
        try:
            upl = float(ws.cell(r, c_upl).value)
        except (TypeError, ValueError):
            sys.exit(f'✗ red {r}: Uplata nije broj — PREKID.')
        if abs(upl - exp_upl) > 0.01:
            sys.exit(f'✗ red {r}: Uplata {upl} ≠ očekivano {exp_upl} — PREKID.')

        tip_now = str(ws.cell(r, c_tip).value or '').strip()
        pod_now = str(ws.cell(r, c_pod).value or '').strip()
        if (tip_now, pod_now) == (NEW_TIP, NEW_POD):
            print(f'· red {r} ({nap_frag}) već je {NEW_TIP}|{NEW_POD} — preskačem')
            continue
        print(f'{"bi se ispravio" if dry else "ispravljen"} red {r:>5} ({nap_frag}, '
              f'{upl:g} €): {tip_now}|{pod_now} → {NEW_TIP}|{NEW_POD}')
        if not dry:
            ws.cell(r, c_tip, NEW_TIP)
            ws.cell(r, c_pod, NEW_POD)
            old_alt = str(ws.cell(r, c_alt).value or '').strip()
            mark = f'S107r fix: rata Anja, bio {tip_now}/{pod_now or "—"}'
            ws.cell(r, c_alt, f'{old_alt} | {mark}' if old_alt else mark)
            ws.cell(r, c_run, stamp)
        changed += 1

    print(f'\n{"Bi se ispravilo" if dry else "Ispravljeno"}: {changed} redova')
    if dry or not changed:
        if dry:
            print('DRY RUN — ništa nije snimljeno.')
        return

    backup = path.with_name(f'{path.stem}.pre-anjarate-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori {path.name} u Excelu i ponovi.\n'
                 f'  (Backup je već napravljen: {backup.name})')
    print(f'✔ Snimljeno. Backup: {backup.name}')
    print(f'  Kontrola: filtriraj Pravilo run = {stamp} → mora dati {changed} redova; '
          f'Prihodi|Povrat Anja mora imati 41 + {changed} = {41 + changed}')


if __name__ == '__main__':
    main()
