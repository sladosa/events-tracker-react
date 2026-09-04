# -*- coding: utf-8 -*-
"""
make_financije_import.py — Review → `Activities Events` Excel za ExcelImportModal. S107t.

Ulaz : `Financije_review_*.xlsx` (sheet `Review`) + generirana struktura
       `Financije_all_structure_*.xlsx` (za provjeru imena atributa)
Izlaz: `Financije_all_import_<ts>.xlsx`, sheet `Events` — svi retci CREATE (bez event_id).

Format je preuzet iz stvarnog app exporta (`events_export_preview_20260731_163957.xlsx`)
i provjeren protiv `excelImport.ts`, ne po sjecanju:

  r1              'ATTRIBUTE LEGEND:'
  r2              Col | Area | Category_Path | Attribute | Type | Default | Description
  r3..            jedan redak po atributnoj koloni  (parseLegend: kol A mora biti [A-Z]{1,3})
  (prazan red)
  'EVENT DATA:'   marker
  zaglavlje       event_id|Area|Category_Path|event_date|session_start|created_at|User|leaf comment| <atributi>
  podaci

⚠ CETIRI TIHE RUPE (sve provjerene u kodu; nijedna ne javlja gresku):

 1. `session_start` MORA biti TEKST 'HH:MM' (`excelImport.ts:239` cita ga preko `cellStr`).
    Prava Excel time vrijednost -> `cellStr` vrati puni ISO -> `parseTimeStr` null ->
    fallback `?? {h:9}` -> SVI redovi dobiju 09:00.
 2. Krivo ime atributa se TIHO preskoci (`excelImport.ts:836` — lookup po
    `leafCategoryId||attrName`, bez `else`). `validateLegendHeaders` provjerava samo
    LEGEND vs zaglavlje UNUTAR filea, ne protiv baze. -> zato `--structure` guard ispod.
 3. `Rate?` je `boolean`: `buildAttrData` radi `String(v).toLowerCase()==='true'`, pa bi
    se Review-ov 'DA' spremio kao FALSE. Pisemo pravi Python bool.
 4. Email u koloni G mora biti racun koji IZVODI import, inace se redak klasificira kao
    "tudi" i preskoci (`foreignMode='skip'`). TEST = Sasin, PROD = Kokin (D6).

ODLUKE UGRADENE U GENERATOR:

 · `session_start` = '09:00' + redni broj unutar dana. `useActivities.ts:242` grupira listu
   po `user_id_category_id_session_start`, a leaf je L1 (isti category_id za sve) -> bez
   toga bi se dan s 21 transakcijom slijepio u JEDAN redak. Max je 21/dan, dan ima 1440 min.
   ⚠ Batchevi moraju biti datumski disjunktni (cutoff na granici dana), inace novi batch
   sudari session_start s vec uvezenim danom.
 · Datumski atributi se pisu kao TEKST 'YYYY-MM-DDT12:00' — isti oblik koji app-ova
   `set_attribute` automatika proizvodi (`attributeRules.ts` formatForDatetimeInput +
   setHours(12)). Podne je namjerno: izbjegava pomak dana po vremenskoj zoni. Prava Excel
   date vrijednost bi u `parseDataRows` prosla kroz `.toISOString()` (UTC!) i mogla
   skliznuti dan unazad.
 · Prazna vrijednost se NE pise (nema kolone -> nema retka u `event_attributes`).
   Posebno `Rate?`: pise se samo `True`, nikad `False` — inace 4335 suvisnih zapisa.
 · `Valuta` se ne pise uopce (prazno = EUR; default je maknut iz strukture da se ne bi
   pisao s touched=true na svaku transakciju).
 · Ispadaju retci sa `Smjer = PROVJERI` (2× pocetno stanje 1.1.2023., 1× prazan
   placeholder, 1× nepotpuni duplikat retka 1503 — v. NEXT_SESSION_PROMPT DIO 1).
 · `Rata br` iz `n/m` u Napomeni, uz guard `m == Broj rata` (samovalidirajuce).
 · `Datum kupovine` ostaje prazan — puni se kasnije, zaseban korak nad Reviewom.

Pokretanje:
  python make_financije_import.py --sample 10 --dry
  python make_financije_import.py --sample 10
  python make_financije_import.py --from 2026-01-01 --to 2026-12-31
"""
from __future__ import annotations

import argparse
import datetime as dt
import glob
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from apply_rules import find_header_col, pick_file  # noqa: E402

HERE = Path(__file__).resolve().parent
DATA = HERE.parent.parent / "data-prep_data" / "Financije"
STRUCT_GLOB = str(DATA / "Financije_all_structure_*.xlsx")

AREA = "Financije_all"
LEAF = "Transakcija"
TEST_EMAIL = "sasasladoljev59@gmail.com"

FIXED_HEADERS = ["event_id", "Area", "Category_Path", "event_date",
                 "session_start", "created_at", "User", "leaf comment"]
ATTR_COL_START = 9  # kolona I (FIXED_COL_COUNT 8 + PADDING 0 + 1)

DAY_START_H, DAY_START_M = 9, 0
MAX_PER_DAY = 1440 - (DAY_START_H * 60 + DAY_START_M)

# (Ime atributa u bazi, data_type, kolona u Reviewu ili None)
# Redoslijed = Sort iz strukture, bez `Valuta`.
ATTRS: list[tuple[str, str, str | None]] = [
    ("Racun",          "text",     "Racun"),
    ("Izvor",          "text",     "Izvor"),
    ("Smjer",          "text",     "Smjer"),
    ("Uplata",         "number",   "Uplata"),
    ("Isplata",        "number",   "Isplata"),
    ("Tip",            "text",     "Tip"),
    ("Podtip",         "text",     "Podtip"),
    ("Izvod opis",     "text",     "Izvod opis"),
    ("Rate?",          "boolean",  "Rate?"),
    ("Broj rata",      "number",   "Broj rata"),
    ("Rata br",        "number",   None),          # izveden iz Napomene
    ("Datum naplate",  "datetime", "Datum naplate"),
    ("Status",         "text",     "Status"),
    ("Stanje",         "number",   "Stanje"),
]

NM = re.compile(r"(?<![\d/])(\d{1,3})\s*/\s*(\d{1,3})(?![\d/])")


# ─────────────────────────────────────────────────────────────
# Citanje
# ─────────────────────────────────────────────────────────────

def pick_structure(explicit: str | None) -> Path:
    if explicit:
        return Path(explicit)
    files = [f for f in glob.glob(STRUCT_GLOB) if not os.path.basename(f).startswith("~$")]
    if not files:
        sys.exit(f"✗ Nema structure filea ({STRUCT_GLOB}). Pokreni "
                 f"make_financije_all_structure.py prvo.")
    return Path(max(files, key=os.path.getmtime))


def read_structure_attrs(path: Path) -> dict[str, str]:
    """{AttrName: AttrType} iz generirane strukture — guard za tihu rupu #2."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Structure"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [("" if c is None else str(c).strip()) for c in rows[2]]
    i = {h: n for n, h in enumerate(hdr) if h}
    out: dict[str, str] = {}
    for r in rows[3:]:
        if not r or str(r[i["Type"]] or "").strip() != "Attribute":
            continue
        out[str(r[i["AttrName"]]).strip()] = str(r[i["AttrType"]] or "").strip()
    wb.close()
    return out


def read_review(path: Path) -> tuple[list[dict], dict[str, int]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Review"]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    idx = {h: n for n, h in enumerate(hdr) if h}
    data = []
    for n, r in enumerate(rows, start=2):
        if all(c is None for c in r):
            continue
        data.append({"_row": n, **{h: r[i] for h, i in idx.items()}})
    wb.close()
    return data, idx


# ─────────────────────────────────────────────────────────────
# Transformacija
# ─────────────────────────────────────────────────────────────

def as_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def as_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def as_datetime_text(v):
    """'YYYY-MM-DDT12:00' — isti oblik koji pise set_attribute automatika."""
    if not isinstance(v, dt.datetime):
        return None
    return f"{v:%Y-%m-%d}T12:00"


def rata_br(napomena, broj_rata) -> int | None:
    """Redni broj rate iz `n/m`, samo ako se m poklapa s `Broj rata`."""
    if broj_rata is None:
        return None
    try:
        total = int(broj_rata)
    except (TypeError, ValueError):
        return None
    for m in NM.finditer(str(napomena or "")):
        if int(m.group(2)) == total:
            return int(m.group(1))
    return None


def build_row(src: dict, email: str) -> dict:
    """Jedan Review redak -> dict {ime atributa: vrijednost} + fiksna polja."""
    attrs: dict[str, object] = {}
    for name, dtype, col in ATTRS:
        raw = src.get(col) if col else None
        if name == "Rate?":
            val = True if str(raw or "").strip().upper() == "DA" else None
        elif name == "Rata br":
            val = rata_br(src.get("Napomena"), src.get("Broj rata"))
        elif dtype == "number":
            val = as_number(raw)
        elif dtype == "datetime":
            val = as_datetime_text(raw)
        else:
            val = as_text(raw)
        if val is not None:
            attrs[name] = val

    return {
        "event_id": None,
        "Area": AREA,
        "Category_Path": LEAF,
        "event_date": src["event_date"],
        "session_start": None,           # dodijeljen kasnije, po danu
        "created_at": None,
        "User": email,
        "leaf comment": as_text(src.get("Napomena")),
        "_attrs": attrs,
        "_src_row": src["_row"],
        "_key": src.get("source_key"),
    }


def assign_session_starts(rows: list[dict]) -> None:
    """'09:00' + redni broj unutar dana; ulazni redoslijed se cuva."""
    seen: dict[dt.date, int] = defaultdict(int)
    for r in rows:
        d = r["event_date"].date()
        n = seen[d]
        seen[d] += 1
        if n >= MAX_PER_DAY:
            sys.exit(f"✗ {d}: vise od {MAX_PER_DAY} transakcija — dan se prelio.")
        total = DAY_START_H * 60 + DAY_START_M + n
        r["session_start"] = f"{total // 60:02d}:{total % 60:02d}"


def pick_sample(rows: list[dict], n: int) -> list[dict]:
    """Raznolik uzorak: po jedan redak iz svake karakteristicne skupine.

    Prva 3 mjesta su rezervirana za TRI TRANSAKCIJE ISTOG DANA — bez toga uzorak
    ne testira ono sto je najrizicnije (da postanu 3 reda u listi, a ne jedan;
    `useActivities.ts:242` grupira po user+category+session_start, a leaf je L1).
    """
    def has(r, a):
        return a in r["_attrs"]

    out, used = [], set()

    by_day: dict[dt.date, list[dict]] = defaultdict(list)
    for r in rows:
        by_day[r["event_date"].date()].append(r)
    busy = [d for d, rs in by_day.items() if len(rs) >= 3]
    if busy and n >= 3:
        day = max(busy, key=lambda d: (len(by_day[d]), d))
        for r in by_day[day][:3]:
            out.append(dict(r, _bucket=f"isti dan ({day})"))
            used.add(r["_src_row"])

    buckets: list[tuple[str, callable]] = [
        ("rata s Rata br",     lambda r: has(r, "Rata br") and has(r, "Rate?")),
        ("Mastercard",         lambda r: r["_attrs"].get("Izvor") == "Mastercard"),
        ("Visa",               lambda r: r["_attrs"].get("Izvor") == "Visa"),
        ("Cash",               lambda r: r["_attrs"].get("Izvor") == "Cash"),
        ("Uplata (priljev)",   lambda r: has(r, "Uplata")),
        ("Status Planiran",    lambda r: r["_attrs"].get("Status") == "Planiran"),
        ("sa Stanjem",         lambda r: has(r, "Stanje")),
        ("bez Podtipa (N/A)",  lambda r: not has(r, "Podtip")),
        ("s Izvod opisom",     lambda r: has(r, "Izvod opis")),
        ("bez komentara",      lambda r: not r["leaf comment"]),
        ("dijakritika u tekstu", lambda r: any(
            ch in str(r["leaf comment"] or "") for ch in "čćžšđČĆŽŠĐ")),
    ]

    for label, pred in buckets:
        if len(out) >= n:
            break
        for r in rows:
            if r["_src_row"] in used or not pred(r):
                continue
            r = dict(r, _bucket=label)
            out.append(r)
            used.add(r["_src_row"])
            break
    # dopuni obicnima ako je premalo skupina pogodilo
    for r in rows:
        if len(out) >= n:
            break
        if r["_src_row"] not in used:
            out.append(dict(r, _bucket="—"))
            used.add(r["_src_row"])
    return out[:n]


# ─────────────────────────────────────────────────────────────
# Pisanje
# ─────────────────────────────────────────────────────────────

FILL_HDR = PatternFill("solid", fgColor="FF4472C4")
FILL_LEG = PatternFill("solid", fgColor="FFF2F2F2")


def write_xlsx(rows: list[dict], out_path: Path, attr_names: list[str],
               attr_types: dict[str, str]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"

    # ── LEGEND ────────────────────────────────────────────────
    ws.cell(1, 1, "ATTRIBUTE LEGEND:").font = Font(bold=True, size=12)
    ws.cell(1, 3, "generirano make_financije_import.py — Col je izvor istine za mapiranje")
    for c, h in enumerate(["Col", "Area", "Category_Path", "Attribute",
                           "Type", "Default", "Description"], start=1):
        cell = ws.cell(2, c, h)
        cell.fill = FILL_HDR
        cell.font = Font(bold=True, color="FFFFFFFF")

    for i, name in enumerate(attr_names):
        r = 3 + i
        ws.cell(r, 1, get_column_letter(ATTR_COL_START + i)).font = Font(bold=True)
        ws.cell(r, 2, AREA)
        ws.cell(r, 3, LEAF)
        ws.cell(r, 4, name)
        ws.cell(r, 5, attr_types[name])
        for c in range(1, 8):
            ws.cell(r, c).fill = FILL_LEG

    marker = 3 + len(attr_names) + 1          # jedan prazan red izmedu
    ws.cell(marker, 1, "EVENT DATA:").font = Font(bold=True, size=12)

    # ── Zaglavlje ─────────────────────────────────────────────
    hrow = marker + 1
    headers = FIXED_HEADERS + [f"{n} ({LEAF})" for n in attr_names]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(hrow, c, h)
        cell.fill = FILL_HDR
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(vertical="center")

    # ── Podaci ────────────────────────────────────────────────
    for i, row in enumerate(rows):
        r = hrow + 1 + i
        ws.cell(r, 1, row["event_id"])
        ws.cell(r, 2, row["Area"])
        ws.cell(r, 3, row["Category_Path"])
        # ⚠ PODNE, ne ponoc (S127). `excelImport.normalizeDateCell` cita Date
        #   LOKALNIM getterima (`getFullYear/getMonth/getDate`), a exceljs serial
        #   pretvara u Date CISTIM UTC-om. Celija na ponoc je zato ispravna samo
        #   dok je uvoznik u pozitivnom offsetu: iz zone UTC-X svih 2.738 redaka
        #   tiho sklizne DAN UNAZAD. Podne daje 12 h margine u obje zone — isto
        #   pravilo koje `Datum naplate` vec postuje (`...T12:00`).
        #   Batchevi 2025./2026. su prosli s ponoci samo zato sto je uvoz bio iz CET-a.
        c4 = ws.cell(r, 4, dt.datetime.combine(row["event_date"].date()
                                               if hasattr(row["event_date"], "date")
                                               else row["event_date"],
                                               dt.time(12, 0)))
        c4.number_format = "yyyy-mm-dd"
        ws.cell(r, 5, row["session_start"])   # TEKST, v. rupa #1
        ws.cell(r, 6, row["created_at"])
        ws.cell(r, 7, row["User"])
        ws.cell(r, 8, row["leaf comment"])
        for j, name in enumerate(attr_names):
            v = row["_attrs"].get(name)
            if v is not None:
                ws.cell(r, ATTR_COL_START + j, v)

    widths = {1: 36, 2: 14, 3: 14, 4: 12, 5: 13, 6: 11, 7: 30, 8: 34}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for j, name in enumerate(attr_names):
        ws.column_dimensions[get_column_letter(ATTR_COL_START + j)].width = max(12, len(name) + 6)

    ws.freeze_panes = f"I{hrow + 1}"
    last = get_column_letter(ATTR_COL_START + len(attr_names) - 1)
    ws.auto_filter.ref = f"A{hrow}:{last}{hrow + len(rows)}"

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description="Review -> Activities Events import xlsx")
    ap.add_argument("--review", help="Financije_review_*.xlsx")
    ap.add_argument("--structure", help="Financije_all_structure_*.xlsx (guard imena atributa)")
    ap.add_argument("--out", help="izlazni xlsx")
    ap.add_argument("--email", default=TEST_EMAIL, help=f"kolona G (default {TEST_EMAIL})")
    ap.add_argument("--from", dest="dfrom", help="event_date od (YYYY-MM-DD, uklj.)")
    ap.add_argument("--to", dest="dto", help="event_date do (YYYY-MM-DD, uklj.)")
    ap.add_argument("--sample", type=int, help="izvuci N raznolikih redaka umjesto raspona")
    ap.add_argument("--dry", action="store_true", help="samo report, ne pise file")
    args = ap.parse_args()

    review_path = Path(args.review) if args.review else pick_file([])
    struct_path = pick_structure(args.structure)

    print("=" * 72)
    print("make_financije_import.py")
    print("=" * 72)
    print(f"  REVIEW    : {review_path.name}")
    print(f"  STRUCTURE : {struct_path.name}")
    print(f"  User (G)  : {args.email}")
    print()

    # ── Guard tihe rupe #2: svako ime koje pisemo mora postojati u strukturi ──
    struct_attrs = read_structure_attrs(struct_path)
    attr_names = [n for n, _, _ in ATTRS]
    missing = [n for n in attr_names if n not in struct_attrs]
    if missing:
        sys.exit(f"✗ Atribut(i) kojih NEMA u strukturi: {missing}\n"
                 f"  Import bi ih tiho preskocio. Struktura ima: {sorted(struct_attrs)}")
    for name, dtype, _ in ATTRS:
        if struct_attrs[name] != dtype:
            sys.exit(f"✗ '{name}': struktura kaze '{struct_attrs[name]}', "
                     f"generator pise '{dtype}' — PREKID.")
    print(f"  ✔ {len(attr_names)} atributa provjereno protiv strukture (ime + tip)")

    data, _ = read_review(review_path)
    print(f"  Review    : {len(data)} redaka")

    # ── Filtri ────────────────────────────────────────────────
    excluded = Counter()
    rows = []
    for src in data:
        ed = src.get("event_date")
        if not isinstance(ed, dt.datetime):
            excluded["bez event_date"] += 1
            continue
        if str(src.get("Smjer") or "").strip().upper() == "PROVJERI":
            excluded["Smjer=PROVJERI"] += 1
            continue
        if args.dfrom and ed.date() < dt.date.fromisoformat(args.dfrom):
            excluded["prije --from"] += 1
            continue
        if args.dto and ed.date() > dt.date.fromisoformat(args.dto):
            excluded["poslije --to"] += 1
            continue
        rows.append(build_row(src, args.email))

    for k, v in excluded.items():
        print(f"  izostavljeno: {v:>5}  ({k})")

    if args.sample:
        rows = pick_sample(rows, args.sample)
    assign_session_starts(rows)

    if not rows:
        sys.exit("✗ Nijedan redak nije prosao filtere.")

    # ── Report ────────────────────────────────────────────────
    dates = [r["event_date"].date() for r in rows]
    print()
    print(f"  ZA IMPORT : {len(rows)} redaka   {min(dates)} → {max(dates)}")
    per_day = Counter(dates)
    print(f"  dana      : {len(per_day)}   max transakcija/dan: {max(per_day.values())} "
          f"(session_start do {max(r['session_start'] for r in rows)})")
    fill = Counter()
    for r in rows:
        for k in r["_attrs"]:
            fill[k] += 1
    print()
    print("  Atribut          popunjeno")
    print("  " + "-" * 30)
    for n in attr_names:
        print(f"  {n:<16} {fill.get(n, 0):>6}")

    if args.sample:
        print()
        print("  Uzorak:")
        for r in rows:
            a = r["_attrs"]
            print(f"   r{r['_src_row']:>5} {r['event_date']:%Y-%m-%d} {r['session_start']} "
                  f"| {a.get('Izvor','—'):<10} {a.get('Smjer','—'):<8} "
                  f"| {str(a.get('Tip','—'))[:14]:<14} | {r.get('_bucket','')}")

    if args.dry:
        print("\n  --dry: file NIJE napisan.")
        return

    out = Path(args.out) if args.out else DATA / \
        f"Financije_all_import_{dt.datetime.now():%Y%m%d_%H%M%S}.xlsx"
    write_xlsx(rows, out, attr_names, {n: t for n, t, _ in ATTRS})
    print(f"\n  ZAPISANO: {out}")


if __name__ == "__main__":
    main()
