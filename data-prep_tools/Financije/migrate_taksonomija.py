# -*- coding: utf-8 -*-
"""
migrate_taksonomija.py  (S107r, 2026-07-30)
===========================================
Jednokratna migracija Review workbooka sa STARE na NOVU taksonomiju
(`Taksonomija (2)` sheet, koji je Koka složila po svom ukusu).

Zašto poseban alat: promjena taksonomije ne pogađa samo `Review!Tip/Podtip`, nego
ČETIRI mjesta, a `Preimenovanja` sheet (apply_rules.py) pokriva samo prvo od njih:

  1. `Review!Tip`/`Podtip`      → 2061 red  → radi apply_rules.py --only-renames
                                              (ovaj alat mu samo pripremi mapping)
  2. `Pravila!Tip`/`Podtip`     → 37 od 70 pravila ima par kojeg više nema;
                                  `read_rules()` bi ih TIHO preskočio → izgubio bi se
                                  cijeli S107l/S107h posao (35 mehanički + 2 ručno)
  3. `Review!Tip_AI`/`Podtip_AI`→ ~1033 od 1592 predikcije; `apply_ai.py --harvest`
                                  odbija par koji nije u Taksonomiji → $1,17 + Sašin
                                  pregled bi postali neupotrebljivi
  4. `Neklasificirano!Tip`/`Podtip` → 10 popunjenih redova; `--harvest` bi ih odbio

Sve četiri koriste ISTU mapping tablicu (MAP dolje) → nema divergencije.

Što alat radi (jedan `wb.save()` na kraju = atomično):
  0. backup `<ime>.pre-taks2-<timestamp>.xlsx`
  1. doda `Investicije | Štednja` u `Taksonomija (2)` ako ne postoji (odluka S107r)
  2. provjeri da SVI ciljevi MAP-a postoje u novoj taksonomiji (inače stane)
  3. `Taksonomija` → `Taksonomija_v1` (skriven, audit), `Taksonomija (2)` → `Taksonomija`
  4. `Preimenovanja` → `Preimenovanja_v1` (skriven), novi `Preimenovanja` iz MAP-a
  5. remap `Pravila` (35 mehanički + 2 ručne odluke, v. MANUAL_RULES)
  6. remap `Review!Tip_AI`/`Podtip_AI` — punim MAP-om, S UVJETIMA, jer AI kolone
     žive na Review redovima pa uvjeti (Racun/Smjer/Iznos/Napomena) imaju kontekst
  7. remap `Neklasificirano` — samo bezuvjetnim dijelom MAP-a (klaster nije red,
     nema konteksta); uvjetni parovi se prijave, ne pogađaju

NAKON ovog alata (redom):
  apply_rules.py --only-renames --dry   → mora prijaviti 2061 → pa pravi run
  sync_taxonomy.py                      → dropdowni + DV rasponi + Liste
  apply_rules.py --dry                  → pravila na N/A hrpi → pa pravi run

Pokretanje (file zatvoren u Excelu!):
  ... migrate_taksonomija.py --dry            → samo report, ništa se ne piše
  ... migrate_taksonomija.py                  → pravi run (uz backup)
  ... migrate_taksonomija.py <put.xlsx> [--dry]

Mapping je odobrio Saša 2026-07-30 (Kokine odluke); pokrivenost provjerena:
2061 nevaljanih Review redaka → 2061 pokriveno, 0 nepokrivenih.
"""

import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from apply_rules import (BORDER, HDR_FILL, PREIM_HEADERS, PREIM_HELP, WHITE_BOLD,  # noqa: E402
                         fold, pick_file)
from openpyxl.styles import Alignment  # noqa: E402

NEW_TAX_SHEET = 'Taksonomija (2)'
NEW_PAIRS = [('Investicije', 'Štednja')]     # dodaci u taksonomiju (odluka S107r)

# ── MAPPING ────────────────────────────────────────────────────────────────────
# (stari tip, stari podtip, racun uvjet, smjer uvjet, iznos min, iznos max,
#  napomena uvjet, novi tip, novi podtip, komentar)
# VAŽNO: uvjetni redovi za isti stari par MORAJU biti iznad bezuvjetnog
# (bezuvjetni = "sve ostalo"); apply_rules čita odozgo, prvi match pobjeđuje.
MAP = [
    # --- uvjetni: jedan stari par → više novih ---------------------------------
    ('Povrat', 'Anja', '', 'Uplata', 450, 450, '', 'Prihodi', 'Povrat Anja',
     'rata posudbe 450 € (Koka): samo uplate od 450 su povrat'),
    ('Povrat', 'Anja', '', '', None, None, '', 'Transfer', 'Anja',
     'sve ostalo (uklj. 2 isplate od 450) nije povrat'),
    ('Ostali prihodi', '', '', '', None, None, 'natasa povrat', 'Transfer', 'Natasa',
     'Nataša linija, ne generički prihod'),
    ('Ostali prihodi', '', 'sasin', 'Isplata', None, None, '', 'Investicije', 'Štednja',
     'red "Stednja" 550 € — odljev, ne prihod'),
    ('Ostali prihodi', '', 'kokin', '', None, None, '', 'Prihodi', 'Koka', ''),
    ('Ostali prihodi', '', 'sasin', '', None, None, '', 'Prihodi', 'Saša', ''),
    ('Domaćinstvo', 'Povrat Nataša', '', '', None, None, 'holding', 'Kuća', 'Holding (smeće)',
     'Kokin račun za Kuću koji plaća ona; razdvojivo po napomeni'),
    ('Domaćinstvo', 'Povrat Nataša', '', '', None, None, '', 'Transfer', 'Natasa',
     '3 reda bez "Holding" u napomeni'),
    ('Domaćinstvo', 'Groblja', '', '', None, None, 'nena', 'Transfer', 'Nena', ''),
    ('Domaćinstvo', 'Groblja', '', '', None, None, '', 'Transfer', 'Natasa',
     'grobna naknada se ne očekuje ubuduće (S107r)'),
    # --- bezuvjetni (1:1 preimenovanje) ---------------------------------------
    ('Namirnice', 'Hrana i ostalo', '', '', None, None, '', 'Domaćinstvo', 'Hrana i ostalo',
     'Koka ukinula Tip Namirnice; uklj. sredstva za čišćenje i higijenu'),
    ('Razno', 'Kave/jelo vani', '', '', None, None, '', 'Domaćinstvo', 'Kave/jelo vani', ''),
    ('auto C5', 'parking', '', '', None, None, '', 'Prijevoz', 'Taksi, Zet, Parking',
     'Koka: ne da se pratiti za koji auto je parking'),
    # 0 redaka u Review!Tip, ali 8 predikcija u Tip_AI — bez ovog reda bi ostale
    # nevaljane i apply_ai --harvest bi ih odbio (uhvaćeno --dry runom, S107r)
    ('auto Lacetti', 'parking', '', '', None, None, '', 'Prijevoz', 'Taksi, Zet, Parking',
     'isto kao auto C5|parking; par je Koka ukinula'),
    ('Mirovina', 'Saša', '', '', None, None, '', 'Prihodi', 'Saša', ''),
    ('Projekti', 'Sasa_Informatika', '', '', None, None, '', 'Projekti', 'Sasa', ''),
    ('Razno', 'Odjeća/obuća/potrebstine _Koka', '', '', None, None, '',
     'Razno', 'Odjeća/obuća/ostalo_Koka', ''),
    ('Domaćinstvo', 'Investicije', '', '', None, None, '',
     'Kuća', 'Popravci, održavanje, osiguranje',
     'IKEA/namještaj/kućni hardver; Investicije su od sad samo Dionice/Kripto/Štednja'),
    ('Razno', 'Taksi', '', '', None, None, '', 'Prijevoz', 'Taksi, Zet, Parking', ''),
    ('Mirovina', 'Koka', '', '', None, None, '', 'Prihodi', 'Koka', ''),
    ('Domaćinstvo', 'Holding (smeće)', '', '', None, None, '', 'Kuća', 'Holding (smeće)', ''),
    ('Domaćinstvo', 'Popravci, održavanje, osiguranje', '', '', None, None, '',
     'Kuća', 'Popravci, održavanje, osiguranje', ''),
    ('Osiguranje', 'Osiguranje', '', '', None, None, '', 'Osiguranje', 'Zivotno',
     'Triglav životno; auto i kuća su izdvojeni još u S107h'),
    ('Domaćinstvo', 'Plin', '', '', None, None, '', 'Kuća', 'Plin', ''),
    ('Domaćinstvo', 'Struja', '', '', None, None, '', 'Kuća', 'Struja', ''),
    ('Domaćinstvo', 'Povrat Zoran', '', '', None, None, '', 'Kuća', 'Povrat Zoran', ''),
    ('Informatika', 'Sitni hardware', '', '', None, None, '', 'Informatika', 'Hardver', ''),
    ('Putovanja', 'Karte', '', '', None, None, '', 'Putovanja', 'Karte, osiguranje', ''),
    ('Razno', 'Odjeća/obuća/potrebstine _Sasa', '', '', None, None, '',
     'Razno', 'Odjeća/obuća/ostalo_Sasa', ''),
    ('Razno', 'Kino/Kazalište/Muzeji', '', '', None, None, '', 'Zabava', 'Kino/Kazalište/Muzeji', ''),
    ('Investicije', 'Dionice', '', '', None, None, '', 'Investicije', 'Dionice / Kripto', ''),
    ('Domaćinstvo', 'Voda', '', '', None, None, '', 'Kuća', 'Voda', ''),
    ('Ostavine', 'Advokati', '', '', None, None, '', 'Advokati', 'Ostavine', 'Koka zamijenila razine'),
]

# ── Pravila: 2 slučaja koja se ne mogu mehanički preslikati ────────────────────
# `UPLATA ANJA CRNKOVIĆ` → stari par ima DVA nova cilja; razdvaja se na dva
# pravila, a ono s Iznos uvjetom MORA biti prvo (prvi match pobjeđuje).
# `grobn` → Groblja više ne postoji. Zadržavamo pravilo iznad `NAKNADA` jer
# prag po iznosu ne razdvaja grobnu (26–28 €) od bankovnih naknada (0,13–50 €);
# u najgorem slučaju kaže Natasa umjesto Nena, ali to je vidljiva greška, dok
# bi ga bez pravila `NAKNADA` zakopala među 102 bankovne naknade.
MANUAL_RULES = {
    'UPLATA ANJA CRNKOVIĆ': {
        'split': [
            {'tip': 'Prihodi', 'pod': 'Povrat Anja', 'imin': 450, 'imax': 450},
            {'tip': 'Transfer', 'pod': 'Anja', 'imin': None, 'imax': None},
        ],
    },
    'grobn': {'tip': 'Transfer', 'pod': 'Natasa'},
}


def read_tax(ws) -> dict[str, set[str]]:
    tax: dict[str, set[str]] = {}
    for r in range(2, ws.max_row + 1):
        tip = str(ws.cell(r, 1).value or '').strip()
        pod = str(ws.cell(r, 2).value or '').strip()
        if not tip:
            continue
        tax.setdefault(tip, set())
        if pod and pod != '—':
            tax[tip].add(pod)
    return tax


def valid(tax: dict[str, set[str]], tip: str, pod: str) -> bool:
    return tip in tax and (not pod or pod in tax[tip])


def hdr_map(ws) -> dict[str, int]:
    return {str(ws.cell(1, c).value or '').strip(): c for c in range(1, ws.max_column + 1)}


def blanket_map() -> dict[tuple[str, str], tuple[str, str]]:
    """Samo bezuvjetni redovi MAP-a — za mjesta bez konteksta reda (Neklasificirano)."""
    out = {}
    for st, sp, uv, us, imin, imax, unap, nt, np_, _kom in MAP:
        if not (uv or us or imin is not None or imax is not None or unap):
            out.setdefault((st, sp), (nt, np_))
    return out


def rewrite_preimenovanja(wb) -> int:
    if 'Preimenovanja' in wb.sheetnames:
        old = wb['Preimenovanja']
        old.title = 'Preimenovanja_v1'
        old.sheet_state = 'hidden'
    idx = wb.sheetnames.index('Pravila') + 1
    ws = wb.create_sheet('Preimenovanja', idx)
    col = {h: i for i, (h, _) in enumerate(PREIM_HEADERS, 1)}
    for c, (h, w) in enumerate(PREIM_HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    for r, (st, sp, uv, us, imin, imax, unap, nt, np_, kom) in enumerate(MAP, 2):
        vals = {'Stari Tip': st, 'Stari Podtip': sp or None, 'Racun uvjet': uv or None,
                'Smjer uvjet': us or None, 'Iznos min': imin, 'Iznos max': imax,
                'Napomena uvjet': unap or None, 'Novi Tip': nt, 'Novi Podtip': np_ or None,
                'Redova': None, 'Komentar': kom or None}
        for c in range(1, len(PREIM_HEADERS) + 1):
            ws.cell(r, c, vals.get(PREIM_HEADERS[c - 1][0])).border = BORDER
    help_col = len(PREIM_HEADERS) + 2
    note = ws.cell(2, help_col, PREIM_HELP)
    note.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions[get_column_letter(help_col)].width = 95
    ws.row_dimensions[2].height = 230
    ws.freeze_panes = ws.cell(2, col['Novi Tip']).coordinate
    ws.auto_filter.ref = f'A1:{get_column_letter(len(PREIM_HEADERS))}{len(MAP) + 1}'
    return len(MAP)


def remap_pravila(wb, tax, bmap, dry: bool) -> tuple[int, int, list[str]]:
    ws = wb['Pravila']
    h = hdr_map(ws)
    c_kw, c_tip, c_pod = h['Ključne riječi'], h['Tip'], h['Podtip']
    c_imin, c_imax = h.get('Iznos min'), h.get('Iznos max')
    log: list[str] = []
    auto = manual = 0

    # 1) mehanički remap + jednostavne ručne odluke (bez splita)
    for r in range(2, ws.max_row + 1):
        kw = str(ws.cell(r, c_kw).value or '').strip()
        tip = str(ws.cell(r, c_tip).value or '').strip()
        pod = str(ws.cell(r, c_pod).value or '').strip()
        if not kw or not tip or valid(tax, tip, pod):
            continue
        man = MANUAL_RULES.get(kw)
        if man and 'split' in man:
            continue                 # split pravila obrađuje petlja 2 (umeće redove)
        if man:
            if not dry:
                ws.cell(r, c_tip, man['tip'])
                ws.cell(r, c_pod, man['pod'])
            manual += 1
            log.append(f'  red {r:>3} {kw:<26} {tip}|{pod} → {man["tip"]}|{man["pod"]}  [RUČNA ODLUKA]')
        elif (tip, pod) in bmap:
            nt, np_ = bmap[(tip, pod)]
            if not dry:
                ws.cell(r, c_tip, nt)
                ws.cell(r, c_pod, np_ or None)
            auto += 1
            log.append(f'  red {r:>3} {kw:<26} {tip}|{pod} → {nt}|{np_}')
        elif not man:
            log.append(f'  red {r:>3} {kw:<26} {tip}|{pod} → ⚠ NEMA MAPPINGA (pravilo ostaje nevaljano)')

    # 2) split pravila — obrada odvojeno jer umeće redove (mijenja numeraciju)
    for kw, man in MANUAL_RULES.items():
        if 'split' not in man:
            continue
        target = next((r for r in range(2, ws.max_row + 1)
                       if str(ws.cell(r, c_kw).value or '').strip() == kw), None)
        if target is None:
            log.append(f'  ⚠ pravilo "{kw}" nije nađeno — split preskočen')
            continue
        first, rest = man['split'][0], man['split'][1:]
        if not dry:
            ws.cell(target, c_tip, first['tip'])
            ws.cell(target, c_pod, first['pod'])
            if c_imin:
                ws.cell(target, c_imin, first['imin'])
            if c_imax:
                ws.cell(target, c_imax, first['imax'])
        manual += 1
        log.append(f'  red {target:>3} {kw:<26} → {first["tip"]}|{first["pod"]}'
                   f'  [RUČNA ODLUKA, Iznos {first["imin"]}–{first["imax"]}]')
        for i, extra in enumerate(rest, 1):
            if not dry:
                ws.insert_rows(target + i)
                ws.cell(target + i, c_kw, kw)
                ws.cell(target + i, c_tip, extra['tip'])
                ws.cell(target + i, c_pod, extra['pod'])
                if c_imin:
                    ws.cell(target + i, c_imin, extra['imin'])
                if c_imax:
                    ws.cell(target + i, c_imax, extra['imax'])
            manual += 1
            log.append(f'  red {target + i:>3} {kw:<26} → {extra["tip"]}|{extra["pod"]}'
                       f'  [NOVI RED, sve ostalo]')
    return auto, manual, log


def remap_ai(wb, tax, dry: bool) -> tuple[int, int, Counter]:
    """Tip_AI/Podtip_AI — puni MAP s uvjetima (AI kolone žive na Review redovima)."""
    ws = wb['Review']
    h = hdr_map(ws)
    for need in ('Tip_AI', 'Podtip_AI', 'Racun', 'Smjer', 'Uplata', 'Isplata', 'Napomena'):
        if need not in h:
            sys.exit(f'✗ Review sheet nema kolonu "{need}"')
    izvod = [c for k, c in h.items() if k.startswith('Izvod opis')]

    def txt(r):
        s = fold(ws.cell(r, h['Napomena']).value)
        for c in izvod:
            s += ' | ' + fold(ws.cell(r, c).value)
        return s

    def amt(r):
        for k in ('Isplata', 'Uplata'):
            v = ws.cell(r, h[k]).value
            if v not in (None, ''):
                try:
                    return float(v)
                except (TypeError, ValueError):
                    pass
        return None

    changed = 0
    unmapped: Counter = Counter()
    for r in range(2, ws.max_row + 1):
        tip = str(ws.cell(r, h['Tip_AI']).value or '').strip()
        pod = str(ws.cell(r, h['Podtip_AI']).value or '').strip()
        if not tip or tip == 'NEPOZNATO' or valid(tax, tip, pod):
            continue
        rac, smj, a, t = fold(ws.cell(r, h['Racun']).value), fold(ws.cell(r, h['Smjer']).value), amt(r), None
        hit = None
        for st, sp, uv, us, imin, imax, unap, nt, np_, _k in MAP:
            if (fold(st), fold(sp)) != (fold(tip), fold(pod)):
                continue
            if uv and fold(uv) not in rac:
                continue
            if us and fold(us) != smj:
                continue
            if imin is not None or imax is not None:
                if a is None or (imin is not None and a < imin) or (imax is not None and a > imax):
                    continue
            if unap:
                t = txt(r) if t is None else t
                if fold(unap) not in t:
                    continue
            hit = (nt, np_)
            break
        if hit is None:
            unmapped[(tip, pod)] += 1
            continue
        if not dry:
            ws.cell(r, h['Tip_AI'], hit[0])
            ws.cell(r, h['Podtip_AI'], hit[1] or None)
        changed += 1
    return changed, len(unmapped), unmapped


def remap_neklasificirano(wb, tax, bmap, dry: bool) -> tuple[int, list[str]]:
    if 'Neklasificirano' not in wb.sheetnames:
        return 0, []
    ws = wb['Neklasificirano']
    h = hdr_map(ws)
    if 'Tip' not in h or 'Podtip' not in h:
        return 0, []
    changed, log = 0, []
    for r in range(2, ws.max_row + 1):
        tip = str(ws.cell(r, h['Tip']).value or '').strip()
        pod = str(ws.cell(r, h['Podtip']).value or '').strip()
        if not tip or valid(tax, tip, pod):
            continue
        if (tip, pod) in bmap:
            nt, np_ = bmap[(tip, pod)]
            if not dry:
                ws.cell(r, h['Tip'], nt)
                ws.cell(r, h['Podtip'], np_ or None)
            changed += 1
            log.append(f'  red {r:>3} {tip}|{pod} → {nt}|{np_}')
        else:
            log.append(f'  red {r:>3} {tip}|{pod} → ⚠ samo uvjetni mapping, riješi ručno')
    return changed, log


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    path = pick_file(args)
    print(f'File: {path.name}{"  [DRY RUN — bez snimanja]" if dry else ""}\n')

    wb = openpyxl.load_workbook(path)
    if 'Review' not in wb.sheetnames:
        sys.exit('✗ File nema Review sheet — je li ovo review Excel?')
    if NEW_TAX_SHEET not in wb.sheetnames:
        if 'Taksonomija_v1' in wb.sheetnames:
            sys.exit(f'✗ Nema "{NEW_TAX_SHEET}", ali postoji "Taksonomija_v1" — '
                     'migracija je već napravljena. Prekidam.')
        sys.exit(f'✗ File nema "{NEW_TAX_SHEET}" sheet.')

    # ── 1. dopuna nove taksonomije ────────────────────────────────────────────
    tws = wb[NEW_TAX_SHEET]
    tax = read_tax(tws)
    for tip, pod in NEW_PAIRS:
        if valid(tax, tip, pod) and pod in tax.get(tip, set()):
            print(f'· "{tip} | {pod}" već postoji u {NEW_TAX_SHEET}')
            continue
        r = tws.max_row + 1
        if not dry:
            tws.cell(r, 1, tip)
            tws.cell(r, 2, pod)
        tax.setdefault(tip, set()).add(pod)
        print(f'✔ Dodano u {NEW_TAX_SHEET}: "{tip} | {pod}" (red {r})')
    print(f'  Nova taksonomija: {len(tax)} Tipova, {sum(len(v) or 1 for v in tax.values())} parova\n')

    # ── 2. ciljevi MAP-a moraju postojati ─────────────────────────────────────
    bad = [f'{nt}|{np_}' for *_x, nt, np_, _k in MAP if not valid(tax, nt, np_)]
    if bad:
        sys.exit('✗ Ciljevi kojih nema u novoj taksonomiji: ' + ', '.join(sorted(set(bad))))
    print(f'✔ Svih {len(MAP)} ciljeva mappinga postoji u novoj taksonomiji\n')

    # ── 3. zamjena Taksonomija sheetova ──────────────────────────────────────
    if not dry:
        wb['Taksonomija'].title = 'Taksonomija_v1'
        wb['Taksonomija_v1'].sheet_state = 'hidden'
        tws.title = 'Taksonomija'
        tws.sheet_state = 'visible'
    print('✔ Taksonomija → Taksonomija_v1 (skriven);  '
          f'{NEW_TAX_SHEET} → Taksonomija\n')

    # ── 4. novi Preimenovanja sheet ──────────────────────────────────────────
    if not dry:
        n = rewrite_preimenovanja(wb)
    else:
        n = len(MAP)
    print(f'✔ Preimenovanja → Preimenovanja_v1 (skriven);  novi Preimenovanja: {n} redova')
    cond = sum(1 for _st, _sp, uv, us, imin, imax, unap, *_ in MAP
               if uv or us or imin is not None or imax is not None or unap)
    print(f'  ({cond} uvjetnih + {n - cond} bezuvjetnih)\n')

    # ── 5–7. remapiranja ─────────────────────────────────────────────────────
    bmap = blanket_map()

    print('── Pravila ' + '─' * 60)
    auto, manual, log = remap_pravila(wb, tax, bmap, dry)
    print('\n'.join(log))
    print(f'  → {auto} mehanički + {manual} iz ručnih odluka\n')

    print('── Review Tip_AI/Podtip_AI ' + '─' * 44)
    ai_changed, ai_bad, ai_unmapped = remap_ai(wb, tax, dry)
    print(f'  → {ai_changed} predikcija remapirano')
    if ai_unmapped:
        print(f'  ⚠ {sum(ai_unmapped.values())} redaka u {ai_bad} parova bez mappinga:')
        for (t, p), c in ai_unmapped.most_common(10):
            print(f'      {c:>4}×  {t}|{p or "—"}')
    print()

    print('── Neklasificirano ' + '─' * 52)
    nk_changed, nk_log = remap_neklasificirano(wb, tax, bmap, dry)
    print('\n'.join(nk_log) if nk_log else '  (nema nevaljanih)')
    print(f'  → {nk_changed} redova remapirano\n')

    if dry:
        print('DRY RUN — ništa nije snimljeno.')
        print('Sljedeće: migrate_taksonomija.py (bez --dry), pa '
              'apply_rules.py --only-renames --dry')
        return

    backup = path.with_name(f'{path.stem}.pre-taks2-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori {path.name} u Excelu i ponovi.\n'
                 f'  (Backup je već napravljen: {backup.name})')
    print(f'✔ Snimljeno. Backup: {backup.name}')
    print('\nSljedeći koraci:')
    print('  1. apply_rules.py --only-renames --dry   → mora prijaviti 2061 preimenovanja')
    print('  2. apply_rules.py --only-renames         → pravi run')
    print('  3. sync_taxonomy.py                     → dropdowni + DV rasponi')
    print('  4. apply_rules.py --dry                 → pravila na N/A hrpi')


if __name__ == '__main__':
    main()
