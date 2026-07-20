# -*- coding: utf-8 -*-
"""
merge_missing_account.py  (S107i Faza 4b, 2026-07-20)
=====================================================
Dodaje bankovne ACCOUNT transakcije (ZABA/RF, Racun kanal) kojih NEMA u Review-u
kao nove retke — isti pattern kao merge_pbzvisa.py. Koka je vodila SALDO računa ali
ne baš svaku transakciju pojedinačno → ove account tx (kojih nema u Reviewu) su realni
promet koji itemiziramo.

Izvor: `Izvodi_transakcije.xlsx` Transakcije sheet, Src ZABA*/RF*.
Match protiv Review Racun/Cash redaka (isti racun + smjer + iznos, datum ±7) — što se
ne matcha = NEDOSTAJE → novi redak.

Novi retci: Izvor='Racun', Racun = Kokin ZABA / Sašin RF (po Src), Tip='N/A',
Napomena prazna, Izvod opis = bankovni opis (pravila ga vide), Datum naplate =
event_date (D1, kao backfill za Racun/Cash). Izvor reda='Izvod account:ZABA'/'...:RF'.

⚠ NAPOMENA (S107i nalaz): ZABA tx-ekstrakcija NIJE potpuna (saldo-lanac ne zatvara —
fali ~359-544€/mjesec), pa je ovih 117 DONJA granica stvarno nedostajućih. Per-red
SaldoB / potpuni bank-reconcile čeka popravak parse_zaba_racun (backlog).

Idempotentno (source_key skip). Sort (Opcija B) + DV/autofilter proširenje + stil
s postojećeg account reda.

Pokretanje (Review zatvoren!):
  Financije\\run.bat merge_missing_account.py --dry   → PREVIEW (missing_acct_PREVIEW.xlsx)
  Financije\\run.bat merge_missing_account.py         → append u Review + backup
"""

import hashlib
import shutil
import sys
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")
TX_XLSX  = DATA_DIR / 'izvodi' / 'Izvodi_transakcije.xlsx'
PREVIEW  = DATA_DIR / 'missing_acct_PREVIEW.xlsx'
DELTAS   = tuple(range(-7, 8))

SRC_ACCOUNT = {'ZABA': 'Kokin tekući ZABA', 'RF': 'Sašin tekući RF'}


def src_prefix(src: str) -> str:
    s = src.upper()
    return 'ZABA' if s.startswith('ZABA') else ('RF' if s.startswith('RF') else '?')


def to_date(v):
    try:
        return v.date()
    except Exception:
        return v if hasattr(v, 'year') else None


def pick_review(args) -> Path:
    explicit = [a for a in args if not a.startswith('--')]
    if explicit:
        p = Path(explicit[0])
        if not p.exists():
            sys.exit(f'✗ File ne postoji: {p}')
        return p
    cands = sorted([c for c in DATA_DIR.glob('Financije_review_*.xlsx') if '.pre-' not in c.name],
                   key=lambda p: p.stat().st_mtime, reverse=True)
    if not cands:
        sys.exit('✗ Nema Financije_review_*.xlsx')
    return cands[0]


def hdr_index(ws) -> dict:
    return {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}


def load_account_tx() -> list[dict]:
    wb = openpyxl.load_workbook(TX_XLSX, read_only=True)
    ws = wb['Transakcije']
    h = {str(c.value): (i + 1) for i, c in enumerate(ws[1]) if c.value is not None}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pref = src_prefix(str(row[h['Src'] - 1] or ''))
        if pref not in SRC_ACCOUNT:
            continue
        d = to_date(row[h['Datum'] - 1])
        if d is None:
            continue
        out.append({
            'date': d, 'opis': str(row[h['Opis'] - 1] or ''), 'smjer': str(row[h['Smjer'] - 1] or ''),
            'iznos': round(float(row[h['Iznos'] - 1]), 2), 'src': str(row[h['Src'] - 1] or ''),
            'pref': pref, 'racun': SRC_ACCOUNT[pref],
        })
    wb.close()
    return out


def source_key(tx: dict, taken: set) -> str:
    canon = f"acct|{tx['racun']}|{tx['date']}|{tx['smjer']}|{tx['iznos']}|{tx['opis']}|{tx['src']}"
    key = hashlib.md5(canon.encode('utf-8')).hexdigest()[:12]
    salt = 0
    while key in taken:
        salt += 1
        key = hashlib.md5(f'{canon}#{salt}'.encode('utf-8')).hexdigest()[:12]
    taken.add(key)
    return key


def build_row(tx, col, ncols, skey) -> list:
    row = [None] * ncols
    def put(name, val): row[col[name] - 1] = val
    ed = datetime(tx['date'].year, tx['date'].month, tx['date'].day)
    put('Racun', tx['racun'])
    put('event_date', ed)
    put('Datum naplate', ed)            # D1: Racun/Cash → Datum naplate = event_date
    put('Smjer', tx['smjer'])
    put('Izvor', 'Racun')
    put('Uplata', tx['iznos'] if tx['smjer'] == 'Uplata' else None)
    put('Isplata', tx['iznos'] if tx['smjer'] == 'Isplata' else None)
    put('Stanje', None)
    put('Napomena', None)
    put('Tip', 'N/A'); put('Podtip', None); put('Pouzdanost', 'NEMA')
    put('Tip_O', 'N/A'); put('Podtip_O', None)
    put('Status', 'Izvrsen')
    put('Izvor reda', f'Izvod account:{tx["pref"]}')
    put('source_key', skey)
    put('Izvod opis', tx['opis'][:250])
    put('Izvod file', tx['src'])
    return row


def sort_review(ws, col, ncols):
    n = ws.max_row
    recs = []
    for r in range(2, n + 1):
        d = to_date(ws.cell(r, col['event_date']).value)
        cells = [(ws.cell(r, c).value, ws.cell(r, c)._style) for c in range(1, ncols + 1)]
        recs.append(((d is None, d or datetime.min.date()), cells))
    recs.sort(key=lambda x: x[0])
    for i, (_, cells) in enumerate(recs):
        for c, (val, style) in enumerate(cells, 1):
            cell = ws.cell(2 + i, c)
            cell.value = val
            cell._style = style
    for dv in ws.data_validations.dataValidation:
        sq = str(dv.sqref)
        if sq.startswith('J'):
            dv.sqref = f'J2:J{n}'
        elif sq.startswith('K'):
            dv.sqref = f'K2:K{n}'
    if ws.auto_filter.ref:
        first = ws.auto_filter.ref.split(':')[0]
        ws.auto_filter.ref = f'{first}:{get_column_letter(ncols)}{n}'


def main():
    args = sys.argv[1:]
    dry = '--dry' in args
    review = pick_review(args)
    print(f'Review: {review.name}{"  [DRY — Review se NE dira]" if dry else ""}')

    tx = load_account_tx()
    print(f'Bank account tx (ZABA/RF): {len(tx)}')

    wb = openpyxl.load_workbook(review)
    ws = wb['Review']
    col = hdr_index(ws)
    needed = ('Racun', 'event_date', 'Datum naplate', 'Smjer', 'Izvor', 'Uplata', 'Isplata',
              'Stanje', 'Napomena', 'Tip', 'Podtip', 'Pouzdanost', 'Status', 'Izvor reda',
              'source_key', 'Izvod opis', 'Izvod file', 'Tip_O', 'Podtip_O')
    missing_cols = [c for c in needed if c not in col]
    if missing_cols:
        sys.exit(f'✗ Review nema kolone: {missing_cols}')
    ncols = ws.max_column

    # index postojećih account redaka za dedup (racun, smjer, iznos) -> [(row, date)]
    existing_keys = {str(ws.cell(r, col['source_key']).value or '')
                     for r in range(2, ws.max_row + 1) if ws.cell(r, col['source_key']).value}
    from collections import defaultdict
    idx = defaultdict(list)
    template_row = None
    for r in range(2, ws.max_row + 1):
        izv = str(ws.cell(r, col['Izvor']).value or '')
        rac = str(ws.cell(r, col['Racun']).value or '')
        if izv not in ('Racun', 'Cash') or rac not in SRC_ACCOUNT.values():
            continue
        if template_row is None:
            template_row = r
        smjer = str(ws.cell(r, col['Smjer']).value or '')
        amt = ws.cell(r, col['Uplata' if smjer == 'Uplata' else 'Isplata']).value
        d = to_date(ws.cell(r, col['event_date']).value)
        if amt in (None, '') or d is None:
            continue
        idx[(rac, smjer, round(float(amt), 2))].append((r, d))
    template_styles = ([ws.cell(template_row, c)._style for c in range(1, ncols + 1)]
                       if template_row else None)

    # match → missing
    used = set()
    missing = []
    for t in sorted(tx, key=lambda t: t['date']):
        hit = None
        for (r, d) in idx.get((t['racun'], t['smjer'], t['iznos']), []):
            if r in used:
                continue
            if abs((d - t['date']).days) <= 7:
                hit = r
                break
        if hit is None:
            missing.append(t)
        else:
            used.add(hit)

    taken = set(existing_keys)
    rows_to_add = []
    skipped = 0
    stats = Counter()
    for t in missing:
        canon = f"acct|{t['racun']}|{t['date']}|{t['smjer']}|{t['iznos']}|{t['opis']}|{t['src']}"
        if hashlib.md5(canon.encode('utf-8')).hexdigest()[:12] in existing_keys:
            skipped += 1
            continue
        skey = source_key(t, taken)
        rows_to_add.append(build_row(t, col, ncols, skey))
        stats[t['pref']] += 1

    print(f'\n{"[DRY] " if dry else ""}NEDOSTAJE za dodati: {len(rows_to_add)}'
          + (f'  (preskočeno {skipped} — source_key već u Review)' if skipped else ''))
    for k, v in stats.most_common():
        print(f'  {k}: {v}')
    print(f'Review sada {ws.max_row - 1} → {ws.max_row - 1 + len(rows_to_add)}')
    if not rows_to_add:
        print('Nema ničega za dodati.'); return

    print('\n--- 12 sample (event_date | Racun | Smjer | iznos | Izvod opis[:42]) ---')
    for row in rows_to_add[:12]:
        amt = row[col['Isplata'] - 1] if row[col['Smjer'] - 1] == 'Isplata' else row[col['Uplata'] - 1]
        print(f"  {str(row[col['event_date']-1])[:10]} | {str(row[col['Racun']-1]):18} | "
              f"{row[col['Smjer']-1]:7} | {str(amt):>9} | {str(row[col['Izvod opis']-1] or '')[:42]}")

    start = ws.max_row + 1
    for i, row in enumerate(rows_to_add):
        for c in range(1, ncols + 1):
            cell = ws.cell(start + i, c)
            if row[c - 1] is not None:
                cell.value = row[c - 1]
            if template_styles:
                cell._style = template_styles[c - 1]
    sort_review(ws, col, ncols)

    if dry:
        wb.save(PREVIEW)
        print(f'\n✔ [DRY] PREVIEW: {PREVIEW.name} ({len(rows_to_add)} redaka, usortirano). '
              f'Filtriraj Izvor reda = "Izvod account:*".')
        print('  Review NIJE diran. Kad potvrdiš → bez --dry.')
        return

    backup = review.with_name(f'{review.stem}.pre-acct-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        sys.exit(f'✗ Zatvori file u Excelu i ponovi. (Backup: {backup.name})')
    print(f'\n✔ Snimljeno: {len(rows_to_add)} account redaka dodano. Backup: {backup.name}')
    print('  Sljedeći korak: apply_rules.py')


if __name__ == '__main__':
    main()
