# -*- coding: utf-8 -*-
"""
consolidate_review.py  (S107j, 2026-07-22)
==========================================
Zatvori izvod-matching: sve što je JASNO iz `Izvodi_transakcije.xlsx` upiši u Review
kao nove retke, a dvojbeni ostatak stavi u sheet **`Nematchano_v3`** UNUTAR Review
workbooka (side-by-side izvod↔Review kandidat + Transfer Y/n + saldo-hint), da dalji
rad živi u Review a Izvodi_transakcije.xlsx više ne treba za odluke.

Klasifikacija nematchanih izvod-tx (match = ±2 dana isti Racun/Izvor/Smjer/Iznos):
  DODAJ u Review:
    • ZABA "TROŠKOVI UČINJENI MASTERCARD KARTICOM" (mjesečni lump) → Tip=Transfer/
      izmedju racuna (novac s tekućeg na karticu; itemizirano posebno preko MC izvoda,
      pa Transfer isključuje dvostruko brojanje)
    • MC/Visa kartična kupovina bez para → nova Isplata/Uplata (Izvor=Mastercard/Visa),
      Tip=N/A (klasificira ih apply_rules / Neklasificirano petlja)
    • ZABA/RF account tx bez para → nova Isplata/Uplata (Izvor=Racun), Tip=N/A
  → Nematchano_v3 (ručni pregled, NE dodaje se auto):
    • "možda već u Reviewu" (isti iznos, datum izvan ±7d) — vjerojatan duplikat
    • "Smjer?" (isti iznos ±7d, suprotan smjer)

Verdikt tok (S107k):
  • sitniš < 5 € s kandidatom → AUTO-DUP (izbačen iz v3, samo brojka u reportu)
  • preostali peach redci imaju kolonu **Verdikt** (dropdown DUP/DODAJ/PRESKOČI,
    pre-popunjen prijedlogom — Saša samo overridea gdje se ne slaže)
  • `--harvest`: pročita Verdikt iz postojećeg v3 PRIJE regeneracije:
      DODAJ    → tx se doda u Review (Transfer ako je kolona Transfer='y' ili MC lump)
      DUP      → nearest green kandidat dobiva event_date ← bankovni datum
                 (Tier C date-sync; par time nestaje iz v3 zauvijek)
      PRESKOČI → trajno ignoriran (hidden sheet `V3 preskočeno`)
      (prazno) → ostaje u v3 za sljedeći krug

Saldo-hint: po ZABA mjesecu usporedi Kokin month-end `Stanje` (Racun redovi) s bankovnim
NOVO STANJE — mjesec koji balansira ⇒ dvojbeni ZABA redak je vjerojatno dup; manjak ⇒
kandidat za dodati. Piše se i zaseban `Saldo kontrola` sheet (svi ZABA mjeseci).

Idempotentno (source_key skip). Backup Review prije snimanja.

Pokretanje (Review + Izvodi_transakcije zatvoreni u Excelu!):
  Financije\\run.bat consolidate_review.py --dry             → PREVIEW, Review netaknut
  Financije\\run.bat consolidate_review.py                   → dodaj retke + Nematchano_v3 + Saldo kontrola
  Financije\\run.bat consolidate_review.py --harvest --dry   → pokaži što bi Verdikt odluke napravile
  Financije\\run.bat consolidate_review.py --harvest         → primijeni Verdikt + regeneriraj v3
"""

import hashlib
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, str(Path(__file__).parent))
from enrich_from_izvoda import _parse_zaba_all, _zaba_is_tekuci  # noqa: E402

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")
IZVODI   = DATA_DIR / 'izvodi' / 'Analizirani_izvodi'
TX_XLSX  = DATA_DIR / 'izvodi' / 'Izvodi_transakcije.xlsx'
PREVIEW  = DATA_DIR / 'consolidate_PREVIEW.xlsx'

RACUN_KOKA = 'Kokin tekući ZABA'
RACUN_SASA = 'Sašin tekući RF'
SRC_MAP = {  # Src prefix → (Racun, Izvor)  [post-merge istina]
    'ZABA':    (RACUN_KOKA, 'Racun'),
    'RF':      (RACUN_SASA, 'Racun'),
    'MC':      (RACUN_KOKA, 'Mastercard'),
    'PBZVISA': (RACUN_SASA, 'Visa'),
}
MATCH_DELTAS = (0, 1, -1, 2, -2)
MC_LUMP = 'TROŠKOVI UČINJENI MASTERCARD KARTICOM'
SITNIS = 5.0                      # prag sitniša: < 5 € s kandidatom → auto-DUP
SKIP_SHEET = 'V3 preskočeno'      # hidden sheet s trajno ignoriranim tx ključevima

HDR_FILL   = PatternFill('solid', fgColor='C55A11')
V3_IZV_FILL = PatternFill('solid', fgColor='FCE4D6')   # izvod redak (svijetlo narančasto)
V3_REV_FILL = PatternFill('solid', fgColor='E2EFDA')   # review kandidat (svijetlo zeleno)
WHITE_BOLD = Font(color='FFFFFF', bold=True)
THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def src_prefix(src: str) -> str:
    s = str(src or '').upper()
    for p in ('PBZVISA', 'ZABA', 'MC', 'RF'):
        if s.startswith(p):
            return p
    return '?'


def to_date(v):
    try:
        return v.date()
    except Exception:
        return v if hasattr(v, 'year') else None


def pick_review(args):
    explicit = [a for a in args if not a.startswith('--')]
    if explicit:
        p = Path(explicit[0])
        if not p.exists():
            sys.exit(f'✗ File ne postoji: {p}')
        return p
    cands = sorted([c for c in DATA_DIR.glob('Financije_review_*.xlsx') if '.pre-' not in c.name],
                   key=lambda p: p.stat().st_mtime, reverse=True)
    if not cands:
        sys.exit('✗ Nema Financije_review_*.xlsx')
    return cands[0]


def hdr_index(ws):
    return {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}


def load_tx():
    wb = openpyxl.load_workbook(TX_XLSX, read_only=True)
    ws = wb['Transakcije']
    h = {str(c.value): i for i, c in enumerate(ws[1]) if c.value is not None}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = to_date(row[h['Datum']])
        pref = src_prefix(row[h['Src']])
        if d is None or pref not in SRC_MAP:
            continue
        racun, izvor = SRC_MAP[pref]
        out.append({'date': d, 'opis': str(row[h['Opis']] or ''),
                    'smjer': str(row[h['Smjer']] or ''),
                    'iznos': round(float(row[h['Iznos']]), 2),
                    'src': str(row[h['Src']] or ''), 'pref': pref,
                    'racun': racun, 'izvor': izvor})
    wb.close()
    return out


def zaba_bank_saldo():
    """{'YYYY-MM': (close_date, novo_stanje)} bankovni Tekući račun. close_date =
    zadnja tekući transakcija izvatka (izvadak se zatvara par dana u idući mjesec,
    pa Kokin saldo treba uspoređivati NA taj datum, ne na kalendarski kraj mjeseca)."""
    out = {}
    for f in sorted(IZVODI.glob('ZABA_*.pdf')):
        txs, balances = _parse_zaba_all(f)
        tek = [t for t in txs if _zaba_is_tekuci(t['account'])]
        tb = [b for b in balances if _zaba_is_tekuci(b['account']) and b['novo'] is not None]
        if tek and tb:
            out[f.stem.split('_')[1]] = (max(t['date'] for t in tek), tb[-1]['novo'])
    return out


def tx_canon(tx):
    return f"cons|{tx['racun']}|{tx['date']}|{tx['smjer']}|{tx['iznos']}|{tx['opis']}|{tx['src']}"


def tx_key(tx):
    """Stabilan identitet izvod-transakcije (md5 canona) — koristi se kao
    source_key baza, v3 `key` kolona i V3 preskočeno zapis."""
    return hashlib.md5(tx_canon(tx).encode()).hexdigest()[:12]


def source_key(tx, taken):
    canon = tx_canon(tx)
    key = hashlib.md5(canon.encode()).hexdigest()[:12]
    salt = 0
    while key in taken:
        salt += 1
        key = hashlib.md5(f'{canon}#{salt}'.encode()).hexdigest()[:12]
    taken.add(key)
    return key


def build_row(tx, col, ncols, skey, tip, podtip):
    row = [None] * ncols
    def put(name, val):
        if name in col:
            row[col[name] - 1] = val
    ed = datetime(tx['date'].year, tx['date'].month, tx['date'].day)
    is_lump = tip == 'Transfer'
    put('Racun', tx['racun'])
    put('event_date', ed)
    put('Datum naplate', ed if tx['izvor'] in ('Racun',) else None)  # Racun→=event_date (D1); kartica puni generator
    put('Smjer', tx['smjer'])
    put('Izvor', tx['izvor'])
    put('Uplata', tx['iznos'] if tx['smjer'] == 'Uplata' else None)
    put('Isplata', tx['iznos'] if tx['smjer'] == 'Isplata' else None)
    put('Napomena', None)
    put('Tip', tip); put('Podtip', podtip)
    put('Pouzdanost', 'VISOKA' if is_lump else 'NEMA')
    put('Tip_O', tip); put('Podtip_O', podtip)
    put('Status', 'Izvrsen')
    put('Izvor reda', f'Konsolidacija:{tx["pref"]}')
    put('source_key', skey)
    put('Izvod opis', tx['opis'][:250])
    put('Izvod file', tx['src'])
    return row


def classify(tx, cand_idx):
    """→ ('add', tip, podtip) ILI ('v3', problem)."""
    cands = cand_idx.get((tx['racun'], tx['izvor'], tx['iznos']), [])
    for c in cands:                                  # suprotan smjer, blizu → Smjer?
        if c['smjer'] != tx['smjer'] and abs((c['date'] - tx['date']).days) <= 7:
            return ('v3', 'Smjer? (suprotan smjer u Reviewu)', None)
    near = [c for c in cands if c['smjer'] == tx['smjer'] and abs((c['date'] - tx['date']).days) <= 31]
    if near:
        return ('v3', 'možda već u Reviewu (isti iznos, datum izvan ±7d)', None)
    if tx['pref'] == 'ZABA' and MC_LUMP in tx['opis'].upper():
        return ('add', 'Transfer', 'izmedju racuna')
    return ('add', 'N/A', None)


def append_review_rows(ws, col, ncols, rows_vals):
    """Append redaka u Review sa stilovima template 'Racun' retka."""
    template = next((r for r in range(2, ws.max_row + 1)
                     if str(ws.cell(r, col['Izvor']).value or '') == 'Racun'), 2)
    tstyles = [ws.cell(template, c)._style for c in range(1, ncols + 1)]
    start = ws.max_row + 1
    for i, rowvals in enumerate(rows_vals):
        for c in range(1, ncols + 1):
            cell = ws.cell(start + i, c)
            if rowvals[c - 1] is not None:
                cell.value = rowvals[c - 1]
            cell._style = tstyles[c - 1]


def harvest_v3(wb, ws, col, ncols, all_tx, dry):
    """Pročita Verdikt iz postojećeg Nematchano_v3 i primijeni odluke na Review.
    → (changed: bool). Poziva se PRIJE matchinga/regeneracije v3, pa DODAJ/DUP
    parovi u novom v3 prirodno nestaju (matchaju na Δ=0); PRESKOČI se pamti u
    hidden sheetu. Prazan Verdikt = ostaje za sljedeći krug."""
    if 'Nematchano_v3' not in wb.sheetnames:
        print('--harvest: nema Nematchano_v3 sheeta — preskačem.')
        return False
    v3ws = wb['Nematchano_v3']
    h = {str(c.value): c.column for c in v3ws[1] if c.value}
    if 'Verdikt' not in h or 'key' not in h:
        print('--harvest: v3 sheet je stari format (bez Verdikt/key) — preskačem.')
        return False
    key2tx = {tx_key(t): t for t in all_tx}
    # postojeći source_keyevi (da DODAJ ne duplicira) + index za DUP lookup
    existing_keys, taken = set(), set()
    ridx = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        sk = ws.cell(r, col['source_key']).value
        if sk:
            existing_keys.add(str(sk)); taken.add(str(sk))
        d = to_date(ws.cell(r, col['event_date']).value)
        if d is None:
            continue
        sm = str(ws.cell(r, col['Smjer']).value or '')
        amt = ws.cell(r, col['Uplata' if sm == 'Uplata' else 'Isplata']).value
        if amt not in (None, ''):
            ridx[(str(ws.cell(r, col['Racun']).value or ''),
                  str(ws.cell(r, col['Izvor']).value or ''),
                  sm, d, round(float(amt), 2))].append(r)
    # prikupi peach grupe (Izvod red + njegovi Review kandidati ispod)
    groups, r, n = [], 2, v3ws.max_row
    while r <= n:
        if str(v3ws.cell(r, h['Source']).value or '') != 'Izvod':
            r += 1
            continue
        verd = str(v3ws.cell(r, h['Verdikt']).value or '').strip().upper().replace('PRESKOCI', 'PRESKOČI')
        key = str(v3ws.cell(r, h['key']).value or '')
        transfer = str(v3ws.cell(r, h['Transfer']).value or '').strip().lower() == 'y'
        cands, rr = [], r + 1
        while rr <= n and str(v3ws.cell(rr, h['Source']).value or '').startswith('Review'):
            # 'Review (matchan)' pripada DRUGOJ tx — info-only, ne smije se sinkati
            dd = to_date(v3ws.cell(rr, h['Datum']).value)
            iz = v3ws.cell(rr, h['Iznos']).value
            if str(v3ws.cell(rr, h['Source']).value or '') == 'Review' \
                    and dd is not None and iz not in (None, ''):
                cands.append((str(v3ws.cell(rr, h['Racun']).value or ''),
                              str(v3ws.cell(rr, h['Izvor']).value or ''),
                              str(v3ws.cell(rr, h['Smjer']).value or ''),
                              dd, round(float(iz), 2)))
            rr += 1
        groups.append((verd, key, transfer, cands))
        r = rr
    # primijeni
    stats = Counter()
    add_rows, problems = [], []
    for verd, key, transfer, cands in groups:
        if verd == '':
            stats['bez verdikta'] += 1
            continue
        t = key2tx.get(key)
        if t is None:
            problems.append(f'key {key}: tx više ne postoji u Izvodi_transakcije')
            stats['greška'] += 1
            continue
        if verd == 'PRESKOČI':
            stats['PRESKOČI'] += 1
            if not dry:
                append_skip(wb, t)
        elif verd == 'DODAJ':
            if tx_key(t) in existing_keys:
                stats['DODAJ (već u Reviewu)'] += 1
                continue
            is_lump = t['pref'] == 'ZABA' and MC_LUMP in t['opis'].upper()
            tip, podtip = ('Transfer', 'izmedju racuna') if (transfer or is_lump) else ('N/A', None)
            add_rows.append(build_row(t, col, ncols, source_key(t, taken), tip, podtip))
            stats['DODAJ'] += 1
        elif verd == 'DUP':
            row = next((rw for c in cands for rw in ridx.get(c, [])), None)
            if row is None:
                problems.append(f'{t["date"]} {t["iznos"]:.2f} {t["opis"][:40]}: '
                                'nema SLOBODNOG green kandidata za DUP sync — ako je '
                                'kandidat "(matchan)", vjerojatno treba DODAJ')
                stats['greška'] += 1
                continue
            old = to_date(ws.cell(row, col['event_date']).value)
            stats['DUP → datum-sync'] += 1
            if not dry:
                if old != t['date']:
                    nd = datetime(t['date'].year, t['date'].month, t['date'].day)
                    if to_date(ws.cell(row, col['Datum naplate']).value) == old:
                        ws.cell(row, col['Datum naplate'], nd)
                    ws.cell(row, col['event_date'], nd)
                # potvrđeni par → popuni prazan Izvod opis/file (kao date_accuracy)
                if ws.cell(row, col['Izvod opis']).value in (None, ''):
                    ws.cell(row, col['Izvod opis'], t['opis'][:250])
                    ws.cell(row, col['Izvod file'], t['src'])
        else:
            problems.append(f'key {key}: nepoznat Verdikt "{verd}"')
            stats['greška'] += 1
    if add_rows and not dry:
        append_review_rows(ws, col, ncols, add_rows)
    print(f'--harvest ({len(groups)} peach grupa): '
          + ', '.join(f'{k}: {v}' for k, v in stats.most_common()) or 'ništa')
    for p in problems[:10]:
        print(f'   ⚠ {p}')
    if dry:
        print('   [DRY] odluke NISU primijenjene (samo brojke).')
        return False
    return bool(add_rows or stats['DUP → datum-sync'] or stats['PRESKOČI'])


def sort_review(ws, col, ncols):
    n = ws.max_row
    recs = []
    for r in range(2, n + 1):
        d = to_date(ws.cell(r, col['event_date']).value)
        cells = [(ws.cell(r, c).value, ws.cell(r, c)._style) for c in range(1, ncols + 1)]
        recs.append(((d is None, d or datetime.min.date()), cells))
    recs.sort(key=lambda x: x[0])
    for i, (_, cells) in enumerate(recs):
        for c, (val, style) in enumerate(cells, 1):
            cell = ws.cell(2 + i, c)
            cell.value = val
            cell._style = style
    for dv in ws.data_validations.dataValidation:
        sq = str(dv.sqref)
        if sq.startswith('J'):
            dv.sqref = f'J2:J{n}'
        elif sq.startswith('K'):
            dv.sqref = f'K2:K{n}'
    if ws.auto_filter.ref:
        first = ws.auto_filter.ref.split(':')[0]
        ws.auto_filter.ref = f'{first}:{get_column_letter(ncols)}{n}'


def v3_verdict(tx, cands, problem, koka_at_close, bank_saldo):
    """→ (keep: bool, analiza: str, prijedlog: str). Zadrži samo PROBLEMATIČNE.
    Sitniš < 5 € s kandidatom → auto-DUP (izbačen). ZABA: mjesec balansira ⇒ dup;
    inače provjeri (Δ). Kartica/RF: kandidat isti iznos ≤7 dana ⇒ dup (date-shift);
    inače provjeri. Prijedlog pre-popuni Verdikt dropdown (Saša overridea)."""
    is_smjer = problem.startswith('Smjer?')
    # DUP ima smisla samo za SLOBODAN kandidat unutar ±31d (isti prozor kao classify);
    # daleki slobodni kandidat (pretplata prije godinu dana) NIJE dup ovog tx-a
    free = [c for c in cands if c['smjer'] == tx['smjer'] and not c.get('used')
            and abs((c['date'] - tx['date']).days) <= 31]
    if not is_smjer and tx['iznos'] < SITNIS and free:
        return (False, f'AUTO-DUP — sitniš <{SITNIS:.0f}€ sa slobodnim kandidatom', '')
    ym = f'{tx["date"]:%Y-%m}'
    if tx['pref'] == 'ZABA' and ym in bank_saldo and koka_at_close.get(ym) is not None:
        # ZABA account: saldo je KONTEKST, ne per-red oracle — timing (Koka upiše par
        # dana nakon zatvaranja izvatka) i agregacija (2×33.98 → 67.96) prave prividne
        # manjkove. Iznos-kandidat je za sitne iznose (parking/naknade) slučajan. Zato
        # samo balansiran mjesec = siguran DUP; ostalo PROVJERI (Saša odlučuje uz
        # Izvor reda + side-by-side + Δ mjeseca).
        k, b = koka_at_close[ym], bank_saldo[ym][1]
        diff = round(k - b, 2)
        if abs(diff) < 0.01:
            return (False, f'DUP — {ym} mjesec balansira ({k:.2f})', '')
        # Koka > banka ⇒ u Reviewu fali trošak → prijedlog DODAJ za Isplatu;
        # bez slobodnog kandidata DUP nema na što sinkati → DODAJ
        sugg = 'DODAJ' if (not free or (diff > 0 and tx['smjer'] == 'Isplata')) else 'DUP'
        return (True, f'PROVJERI — {ym} Koka {k:.2f} vs banka {b:.2f} (Δ{diff:+.2f})',
                '' if is_smjer else sugg)
    if is_smjer:
        return (True, problem, '')
    nd = min((abs((c['date'] - tx['date']).days) for c in free), default=None)
    if nd is not None and nd <= 7:
        return (False, f'DUP — isti iznos u Reviewu {nd} dana blizu (slobodan)', '')
    if nd is not None:
        return (True, f'PROVJERI — slobodan kandidat {nd} dana daleko', 'DUP')
    if any(c['smjer'] == tx['smjer'] and abs((c['date'] - tx['date']).days) <= 31
           for c in cands):
        return (True, 'PROVJERI — blizak kandidat već matchan drugim izvod-tx → vjerojatno FALI',
                'DODAJ')
    return (True, 'PROVJERI — kandidati samo daleko (>31d) → vjerojatno FALI (npr. mjesec '
            'pretplate koji nije upisan)', 'DODAJ')


def load_skip_keys(wb):
    if SKIP_SHEET not in wb.sheetnames:
        return set()
    return {str(r[0]) for r in wb[SKIP_SHEET].iter_rows(min_row=2, values_only=True) if r[0]}


def append_skip(wb, tx):
    ws = (wb[SKIP_SHEET] if SKIP_SHEET in wb.sheetnames else wb.create_sheet(SKIP_SHEET))
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        for c, hh in enumerate(('key', 'Datum', 'Iznos', 'Smjer', 'Opis', 'Kad'), 1):
            ws.cell(1, c, hh).font = Font(bold=True)
        ws.sheet_state = 'hidden'
    r = ws.max_row + 1
    ws.cell(r, 1, tx_key(tx)); ws.cell(r, 2, str(tx['date'])); ws.cell(r, 3, tx['iznos'])
    ws.cell(r, 4, tx['smjer']); ws.cell(r, 5, tx['opis'][:80])
    ws.cell(r, 6, f'{datetime.now():%Y-%m-%d %H:%M}')


def write_v3(wb, v3_items, koka_at_close, bank_saldo):
    """Piše SAMO problematične (dup/sitniš se izbaci), recent-first (2026 gore).
    Peach `Izvod` red = odluka (Verdikt dropdown); green `Review` = read-only kopija."""
    skip_keys = load_skip_keys(wb)
    if 'Nematchano_v3' in wb.sheetnames:
        del wb['Nematchano_v3']
    ws = wb.create_sheet('Nematchano_v3')
    heads = ('Source', 'Datum', 'Iznos', 'Smjer', 'Racun', 'Izvor',
             'Opis / Napomena', 'Tip (Review)', 'Δ dana', 'Transfer',
             'Analiza / saldo', 'Verdikt', 'Src', 'key')
    for c, hh in enumerate(heads, 1):
        cell = ws.cell(1, c, hh)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
    for c, w in zip('ABCDEFGHIJKLMN', (9, 11, 10, 8, 18, 11, 46, 16, 7, 9, 48, 11, 20, 13)):
        ws.column_dimensions[c].width = w
    dv = DataValidation(type='list', formula1='"DUP,DODAJ,PRESKOČI"', allow_blank=True,
                        promptTitle='Verdikt', prompt='DUP = isti kao green red (event_date se '
                        'sinka na bankovni datum). DODAJ = novi red u Review. PRESKOČI = trajno ignoriraj.')
    ws.add_data_validation(dv)
    # verdikt + zadrži samo problematične, sortiraj recent-first
    kept, dropped, sitnis_n, skipped_n = [], 0, 0, 0
    for tx, cands, problem in v3_items:
        if tx_key(tx) in skip_keys:
            skipped_n += 1
            continue
        keep, analiza, sugg = v3_verdict(tx, cands, problem, koka_at_close, bank_saldo)
        if keep:
            kept.append((tx, cands, analiza, sugg))
        else:
            dropped += 1
            if 'sitniš' in analiza:
                sitnis_n += 1
    kept.sort(key=lambda x: x[0]['date'], reverse=True)   # 2026/2025 gore
    r = 2
    for tx, cands, analiza, sugg in kept:
        ws.cell(r, 1, 'Izvod'); ws.cell(r, 2, tx['date']).number_format = 'DD.MM.YYYY'
        ws.cell(r, 3, tx['iznos']).number_format = '#,##0.00'
        ws.cell(r, 4, tx['smjer']); ws.cell(r, 5, tx['racun']); ws.cell(r, 6, tx['izvor'])
        ws.cell(r, 7, tx['opis'][:120]); ws.cell(r, 10, 'n'); ws.cell(r, 11, analiza)
        ws.cell(r, 12, sugg); ws.cell(r, 13, tx['src']); ws.cell(r, 14, tx_key(tx))
        dv.add(ws.cell(r, 12))
        for c in range(1, 15):
            ws.cell(r, c).fill = V3_IZV_FILL
        r += 1
        # slobodni kandidati prvi, pa po blizini; zauzeti = 'Review (matchan)' info-only
        for c in sorted(cands, key=lambda c: (c.get('used', False),
                                              abs((c['date'] - tx['date']).days)))[:2]:
            ws.cell(r, 1, 'Review (matchan)' if c.get('used') else 'Review')
            ws.cell(r, 2, c['date']).number_format = 'DD.MM.YYYY'
            ws.cell(r, 3, c['iznos']).number_format = '#,##0.00'
            ws.cell(r, 4, c['smjer']); ws.cell(r, 5, c['racun']); ws.cell(r, 6, c['izvor'])
            ws.cell(r, 7, (c['napomena'] or '')[:120]); ws.cell(r, 8, c['tip'])
            ws.cell(r, 9, (c['date'] - tx['date']).days)
            for cc in range(1, 15):
                ws.cell(r, cc).fill = V3_REV_FILL
            r += 1
    ws.freeze_panes = 'F2'                      # split: pinaj A–E + header
    ws.auto_filter.ref = f'A1:N{max(2, r - 1)}'
    ws.sheet_view.tabColor = 'C55A11'
    print(f'  Nematchano_v3: {len(kept)} za odluku ({dropped} auto-dup izbačeno, '
          f'od toga {sitnis_n} sitniš <{SITNIS:.0f}€; {skipped_n} ranije PRESKOČENO)')


def write_saldo_kontrola(wb, koka_monthend, bank_saldo):
    if 'Saldo kontrola' in wb.sheetnames:
        del wb['Saldo kontrola']
    ws = wb.create_sheet('Saldo kontrola')
    for c, hh in enumerate(('Izvadak', 'Datum zatvaranja', 'Koka Stanje @ zatvaranju',
                            'Banka NOVO STANJE', 'Δ (Koka−banka)', 'Status'), 1):
        cell = ws.cell(1, c, hh)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
    for c, w in zip('ABCDEF', (10, 16, 24, 20, 16, 26)):
        ws.column_dimensions[c].width = w
    r = 2
    for ym in sorted(bank_saldo):
        close, b = bank_saldo[ym]
        k = koka_monthend.get(ym)
        ws.cell(r, 1, ym); ws.cell(r, 2, close).number_format = 'DD.MM.YYYY'
        ws.cell(r, 4, b).number_format = '#,##0.00'
        if k is not None:
            diff = round(k - b, 2)
            ws.cell(r, 3, k).number_format = '#,##0.00'
            ws.cell(r, 5, diff).number_format = '#,##0.00'
            ws.cell(r, 6, 'OK — balansira' if abs(diff) < 0.01 else 'RAZLIKA — provjeriti')
        else:
            ws.cell(r, 6, 'nema Kokinog Stanja za datum')
        r += 1
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{max(2, r - 1)}'


def main():
    args = sys.argv[1:]
    dry = '--dry' in args
    review = pick_review(args)
    print(f'Review: {review.name}{"  [DRY — Review NETAKNUT]" if dry else ""}')

    tx = load_tx()
    wb = openpyxl.load_workbook(review)
    ws = wb['Review']
    col = hdr_index(ws)
    ncols = ws.max_column

    # harvest Verdikta iz postojećeg v3 PRIJE matchinga (odluke mijenjaju Review,
    # pa ih regeneracija v3 prirodno više ne vidi)
    harvest_changed = False
    if '--harvest' in args:
        harvest_changed = harvest_v3(wb, ws, col, ncols, tx, dry)

    # indeksi: match (±2) + kandidati (isti iznos) + Kokin month-end Stanje
    match_idx = defaultdict(list)     # (racun,izvor,smjer,date,iznos) -> [row]
    cand_idx = defaultdict(list)      # (racun,izvor,iznos) -> [cand]
    existing_keys = set()
    koka_series = []                  # [(date, stanje)] ZABA Racun redovi (za saldo @ close)
    for r in range(2, ws.max_row + 1):
        sk = ws.cell(r, col['source_key']).value
        if sk:
            existing_keys.add(str(sk))
        d = to_date(ws.cell(r, col['event_date']).value)
        if d is None:
            continue
        rac = str(ws.cell(r, col['Racun']).value or '')
        izv = str(ws.cell(r, col['Izvor']).value or '')
        sm = str(ws.cell(r, col['Smjer']).value or '')
        amt = ws.cell(r, col['Uplata' if sm == 'Uplata' else 'Isplata']).value
        if amt not in (None, ''):
            a = round(float(amt), 2)
            match_idx[(rac, izv, sm, d, a)].append(r)
            cand_idx[(rac, izv, a)].append(
                {'date': d, 'smjer': sm, 'iznos': a, 'racun': rac, 'izvor': izv, 'row': r,
                 'napomena': ws.cell(r, col['Napomena']).value, 'tip': ws.cell(r, col['Tip']).value})
        st = ws.cell(r, col['Stanje']).value
        if 'ZABA' in rac.upper() and izv == 'Racun' and st not in (None, ''):
            koka_series.append((d, round(float(st), 2)))
    koka_series.sort()

    def koka_stanje_at(dt):           # zadnje Kokino Stanje s datumom ≤ dt
        v = None
        for dd, s in koka_series:
            if dd <= dt:
                v = s
            else:
                break
        return v

    bank_saldo = zaba_bank_saldo()    # ym -> (close_date, novo)
    koka_at_close = {ym: koka_stanje_at(close) for ym, (close, _) in bank_saldo.items()}

    used = set()
    to_add, v3 = [], []
    for t in sorted(tx, key=lambda t: t['date']):
        row = None
        for dd in MATCH_DELTAS:
            key = (t['racun'], t['izvor'], t['smjer'], t['date'] + timedelta(days=dd), t['iznos'])
            row = next((r for r in match_idx.get(key, []) if r not in used), None)
            if row:
                break
        if row:
            used.add(row)
            continue                                 # već u Review
        kind, a, b = classify(t, cand_idx)
        if kind == 'add':
            to_add.append((t, a, b))
        else:
            v3.append((t, cand_idx.get((t['racun'], t['izvor'], t['iznos']), []), a))

    # označi kandidate koje je već potrošio neki drugi tx match (Δ0..±2) —
    # takav red pripada DRUGOJ transakciji pa NIJE dup-kandidat za sync
    for lst in cand_idx.values():
        for c in lst:
            c['used'] = c['row'] in used
    # sitniš bez ijednog SLOBODNOG kandidata → auto-DODAJ (kandidat zauzet =
    # tx stvarno fali u Reviewu; klasifikaciju odradi apply_rules)
    still_v3, sitnis_added = [], 0
    for t, cands, prob in v3:
        free = [c for c in cands if c['smjer'] == t['smjer'] and not c['used']
                and abs((c['date'] - t['date']).days) <= 31]
        if not prob.startswith('Smjer?') and t['iznos'] < SITNIS and not free:
            to_add.append((t, 'N/A', None))
            sitnis_added += 1
        else:
            still_v3.append((t, cands, prob))
    v3 = still_v3

    # build add-rows (source_key dedup)
    taken = set(existing_keys)
    rows_to_add = []
    stats = Counter()
    for t, tip, podtip in to_add:
        if tx_key(t) in existing_keys:
            continue
        skey = source_key(t, taken)
        rows_to_add.append(build_row(t, col, ncols, skey, tip, podtip))
        stats[('Transfer' if tip == 'Transfer' else t['izvor'])] += 1

    print(f'\nIzvod tx: {len(tx)}  |  DODAJ: {len(rows_to_add)}'
          + (f' (od toga {sitnis_added} sitniš auto-DODAJ, kandidat zauzet)' if sitnis_added else '')
          + f'  |  Nematchano_v3: {len(v3)}')
    for k, n in stats.most_common():
        print(f'   + {k}: {n}')
    # saldo kontrola sažetak
    off = [ym for ym in bank_saldo
           if koka_at_close.get(ym) is not None and abs(koka_at_close[ym] - bank_saldo[ym][1]) >= 0.01]
    print(f'Saldo kontrola: {len(bank_saldo)} ZABA izvadaka, {len(off)} s razlikom Koka↔banka'
          + (f' ({", ".join(off[:8])}{"…" if len(off) > 8 else ""})' if off else ''))

    # append + sort + v3 + saldo sheet
    append_review_rows(ws, col, ncols, rows_to_add)
    if rows_to_add or harvest_changed:
        sort_review(ws, col, ncols)
    ws.freeze_panes = 'F2'                        # split: pinaj A–E + header (Review)
    write_v3(wb, v3, koka_at_close, bank_saldo)
    write_saldo_kontrola(wb, koka_at_close, bank_saldo)

    if dry:
        wb.save(PREVIEW)
        print(f'\n✔ [DRY] PREVIEW: {PREVIEW.name} ({len(rows_to_add)} dodano, v3={len(v3)}). Review NETAKNUT.')
        return
    backup = review.with_name(f'{review.stem}.pre-consolidate-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        sys.exit(f'✗ Zatvori Review u Excelu i ponovi. (Backup: {backup.name})')
    print(f'\n✔ Snimljeno: +{len(rows_to_add)} redaka, Nematchano_v3 ({len(v3)}), Saldo kontrola.')
    print(f'  Backup: {backup.name}')
    print('  Izvodi_transakcije.xlsx više ne treba za odluke — sve u Review. Sljedeći: apply_rules.py')


if __name__ == '__main__':
    main()
