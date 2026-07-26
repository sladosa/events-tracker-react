"""
apply_label_fixes.py — jednokratne ispravke labela dogovorene sa Sašom 2026-07-26.

Kontekst: prvi AI eval (v. NEXT_SESSION_PROMPT O9) otkrio je da dio "grešaka modela"
nisu greške modela nego nedosljednosti u labelama i parovi koji ne postoje u
Taksonomiji. Saša je donio odluke; ovdje se primjenjuju.

Pravila se primjenjuju REDOM — prvi pogodak vrijedi.

  1. Transfer|prazno + Ašo/Aso        → Zdravlje | Sport_Sasa        (personal trener, 20 €/h)
  2. Transfer|prazno + Dionice        → Investicije | Dionice        (NOVI par u Taksonomiji)
  3. Transfer|prazno + bankomat/goto  → Transfer | cash - bankomat
  4. Transfer|prazno  ostatak         → Transfer | izmedju racuna    (uklj. Visa, Saša uplata)
  5. Putovanja|prazno                 → Putovanja | Restoran         (hrana na putu, bez dijeljenja)
  6. Domaćinstvo|prazno (pričuva)     → Transfer | izmedju racuna    (Koka dobila novac natrag)
  7. Zdravlje|prazno (KEINDL)         → Zdravlje | Sport_Sasa
  8. Informatika|prazno (APPLE.COM)   → Informatika | Cloud backup   (presedan S107h)
  9. auto C5|prazno (HAK)             → registracija; SS=Saša→Lacetti, ostalo→C5
 10. Razno|prazno (Promjena guma)     → auto C5 | popravci
 11. BIBERON (svi)                    → Projekti | Sasa_Informatika  (bio 33/22 nedosljedan)
 12. KONZUM + RADNIČKA + isplata<30 € → Projekti | Sasa_Informatika  (ručak uz posao)

NE dira `Tip_O`/`Podtip_O` (to je originalni snapshot). Upisuje timestamp u `Pravilo run`
i dopisuje marker u `Alternativa / nap.` po P3 (puna vrijednost se ne gazi).

  python apply_label_fixes.py --dry     # samo pokaži
  python apply_label_fixes.py           # upiši (backup automatski)
"""
from __future__ import annotations

import argparse
import collections
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
import ai_classify as A                                          # noqa: E402

MARKER = 'fix-2026-07-26'
NEW_PAIR = ('Investicije', 'Dionice')


def rule_for(txt: str, tip: str, pod: str, isplata: float) -> tuple[str, str, str] | None:
    """Vraća (novi_tip, novi_podtip, ime_pravila) ili None."""
    u = txt.upper()
    if tip == 'Transfer' and not pod:
        if 'AŠO' in u or 'ASO ' in u or u.strip() == 'ASO':
            return 'Zdravlje', 'Sport_Sasa', '1 Ašo (trener)'
        if 'DIONICE' in u:
            return 'Investicije', 'Dionice', '2 Dionice'
        if any(k in u for k in ('BANKOMAT', 'ISPLATAGOTO', 'PODIZANJE')):
            return 'Transfer', 'cash - bankomat', '3 bankomat'
        return 'Transfer', 'izmedju racuna', '4 Transfer ostatak'
    if tip == 'Putovanja' and not pod:
        return 'Putovanja', 'Restoran', '5 Putovanja'
    if tip == 'Domaćinstvo' and not pod:
        return 'Transfer', 'izmedju racuna', '6 pričuva'
    if tip == 'Zdravlje' and not pod:
        return 'Zdravlje', 'Sport_Sasa', '7 KEINDL'
    if tip == 'Informatika' and not pod:
        return 'Informatika', 'Cloud backup', '8 Apple/iCloud'
    if tip == 'auto C5' and not pod:
        if 'HAK SS' in u:                       # SS = Saša Sladoljev → njegov auto
            return 'auto Lacetti', 'registracija', '9 HAK (SS=Saša)'
        if 'HAK DPS' in u:                      # DPS = Dubravka Pavić-Sladoljev = Koka
            return 'auto C5', 'registracija', '9 HAK (DPS=Koka)'
        # 2023./2024. jedna uplata od 66 € pokriva OBA auta (po 33). Ne razdvajam
        # u dva retka radi dva zapisa — činjenica se čuva u napomeni.
        return 'auto C5', 'registracija', '9 HAK 66EUR = oba auta po 33, knjizeno na C5'
    if tip == 'Razno' and not pod:
        return 'auto C5', 'popravci', '10 promjena guma'
    if 'BIBERON' in u:
        return 'Projekti', 'Sasa_Informatika', '11 BIBERON'
    # 'RATA nn/ mm' = velika kupovina razbijena na rate (10 × 18 € = 180 €), ne ručak.
    # Sašin prag <30 € postoji baš da "ne uđu velike kupovine" — rate su iznimka koju
    # sam prag ne hvata jer je svaka pojedina rata mala.
    if 'KONZUM' in u and 'RADNI' in u and 0 < isplata < 30 and 'RATA' not in u:
        return 'Projekti', 'Sasa_Informatika', '12 Konzum Radnička <30 (bez rata)'
    return None


def add_taxonomy_pair(wb, dry: bool) -> bool:
    """Doda Investicije | Dionice ako ga nema. Taksonomija: Tip=A, Podtip=B (fiksni indeks!)."""
    ws = wb['Taksonomija']
    for r in ws.iter_rows(min_row=2, values_only=True):
        if A.clean(r[0]) == NEW_PAIR[0] and A.clean(r[1] if len(r) > 1 else '') == NEW_PAIR[1]:
            return False
    row = ws.max_row + 1
    while row > 2 and not A.clean(ws.cell(row - 1, 1).value):
        row -= 1
    if not dry:
        ws.cell(row, 1, NEW_PAIR[0])
        ws.cell(row, 2, NEW_PAIR[1])
    print(f'  + Taksonomija: novi par "{NEW_PAIR[0]} | {NEW_PAIR[1]}" (red {row})')
    return True


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry', action='store_true')
    args = ap.parse_args()

    review = A.pick_review()
    wb = openpyxl.load_workbook(review)
    ws = wb['Review']
    H = A.header_map(ws)
    for c in ('Tip', 'Podtip', 'Napomena', 'Izvod opis', 'Isplata',
              'Alternativa / nap.', 'Pravilo run'):
        if c not in H:
            sys.exit(f'✗ Nema kolone "{c}" u Review sheetu.')
    col = {k: H[k] + 1 for k in H}                     # openpyxl je 1-based

    stamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    hits: collections.Counter = collections.Counter()
    changed = 0
    samples: dict[str, list[str]] = collections.defaultdict(list)

    for r in range(2, ws.max_row + 1):
        tip = A.clean(ws.cell(r, col['Tip']).value)
        if not tip or tip == 'N/A':
            continue
        pod = A.clean(ws.cell(r, col['Podtip']).value)
        txt = (A.clean(ws.cell(r, col['Napomena']).value) + ' / '
               + A.clean(ws.cell(r, col['Izvod opis']).value)).strip(' /')
        try:
            isp = float(ws.cell(r, col['Isplata']).value or 0)
        except (TypeError, ValueError):
            isp = 0.0

        res = rule_for(txt, tip, pod, isp)
        if not res:
            continue
        nt, npod, name = res
        if (nt, npod) == (tip, pod):
            continue                                    # već točno — ne diramo
        hits[name] += 1
        changed += 1
        if len(samples[name]) < 3:
            samples[name].append(f'red {r}: {tip} | {pod}  →  {nt} | {npod}   [{txt[:40]}]')
        if args.dry:
            continue
        ws.cell(r, col['Tip'], nt)
        ws.cell(r, col['Podtip'], npod)
        ws.cell(r, col['Pravilo run'], stamp)
        prev = A.clean(ws.cell(r, col['Alternativa / nap.']).value)
        note = f'{MARKER}: {name}'
        ws.cell(r, col['Alternativa / nap.'],
                f'{prev} | {note}' if prev else note)   # P3: ne gazi punu vrijednost

    print(f'Review: {review.name}\n')
    for name in sorted(hits):
        print(f'  {hits[name]:>4}×  {name}')
        for s in samples[name]:
            print(f'          {s}')
    print(f'\n  UKUPNO promijenjeno: {changed} redaka')
    add_taxonomy_pair(wb, args.dry)

    if args.dry:
        print('\n✔ [DRY] Ništa nije upisano.')
        return

    backup = review.with_name(f'{review.stem}.pre-labelfix-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        sys.exit(f'✗ Zatvori Review u Excelu i ponovi. (Backup: {backup.name})')
    print(f'\n✔ Upisano. Backup: {backup.name}')
    print('  SLJEDEĆE: pokreni `sync_taxonomy.py` da dropdowni vide novi par.')


if __name__ == '__main__':
    main()
