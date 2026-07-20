# -*- coding: utf-8 -*-
"""
reconcile_izvoda.py  (S107i Faza 4, 2026-07-20)
================================================
Kontrola pokrivenosti migracije: koje izvod-transakcije JOŠ nisu zastupljene u Review-u,
i ZAŠTO (dijagnoza po retku). Piše u `Izvodi_transakcije.xlsx`:
  • `Nematchano_v1` — freeze trenutnog `Nematchano` (pre-Faza3/4 baseline, jednom).
  • `Nematchano_v2` — izvod-tx bez para u Reviewu + `Problem` kolona (dijagnoza).
  • `Coverage`      — sažetak matched/NEDOSTAJE po tipu izvora + breakdown Problema.
REVIEW SE NE DIRA (samo se čita).

ISPRAVNO mapiranje Src → (Racun, Izvor) — post-merge istina:
  ZABA→Kokin ZABA/Racun, RF→Sašin RF/Racun, MC→Kokin ZABA/Mastercard, PBZVISA→Sašin RF/Visa

Dijagnoza `Problem` (prioritetom):
  • "Smjer? (u Reviewu suprotan smjer)"  — isti racun/iznos, datum ±7, ali suprotan Smjer
      → parser vjerojatno krivo odredio Priljev/Odljev (ZABA nalaz S107i).
  • "Smjer? (opis=priljev, vodi se Isplata)" — inflow keyword u opisu, a Smjer=Isplata.
  • "možda već u Reviewu (isti iznos, datum ΔNd)" — postoji red istog iznosa, datum >7d.
  • "kartična kupovina (nije itemizirana)" — MC/Visa kupovina, Koka je nije pojedinačno vodila.
  • "nedostaje (nema kandidata)" — stvarno fali.

⚠ NAPOMENA (S107i): ZABA saldo-lanac NE zatvara + parser ima Smjer greške →
`Nematchano_v2` je donja granica; "Smjer?" redovi = kandidati za fix parse_zaba_racun (backlog).

Pokretanje (Izvodi_transakcije.xlsx zatvoren!):
  Financije\\run.bat reconcile_izvoda.py --dry   → samo report
  Financije\\run.bat reconcile_izvoda.py         → + upiše sheetove (backup prvo)
"""

import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")
TX_XLSX  = DATA_DIR / 'izvodi' / 'Izvodi_transakcije.xlsx'

HDR_FILL   = PatternFill('solid', fgColor='4472C4')
WHITE_BOLD = Font(color='FFFFFF', bold=True)
SMJER_FILL = PatternFill('solid', fgColor='FFC7CE')   # crveno za Smjer probleme
THIN       = Side(style='thin')
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SRC_MAP = {
    'ZABA':    ('Kokin tekući ZABA', 'Racun'),
    'RF':      ('Sašin tekući RF',   'Racun'),
    'MC':      ('Kokin tekući ZABA', 'Mastercard'),
    'PBZVISA': ('Sašin tekući RF',   'Visa'),
}
DELTAS = (0, 1, -1, 2, -2)
IN_KW  = ('mirovinsk', 'priljev', 'povrat sredstava', 'uplata mirovinsk',
          'uplata sladoljev', 'uplata zoran', 'uplata anja', 'uplata mama',
          'priljev iz inozemstva', 'dividend', 'kamata', 'placa', 'plaća', 'plaće')


def src_prefix(src: str) -> str:
    s = src.upper()
    for p in ('PBZVISA', 'ZABA', 'MC', 'RF'):
        if s.startswith(p):
            return p
    return '?'


def to_date(v):
    try:
        return v.date()
    except Exception:
        return v if hasattr(v, 'year') else None


def load_tx() -> list[dict]:
    wb = openpyxl.load_workbook(TX_XLSX, read_only=True)
    ws = wb['Transakcije']
    h = {str(c.value): i for i, c in enumerate(ws[1]) if c.value is not None}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = to_date(row[h['Datum']])
        pref = src_prefix(str(row[h['Src']] or ''))
        if d is None or pref not in SRC_MAP:
            continue
        racun, izvor = SRC_MAP[pref]
        out.append({
            'date': d, 'opis': str(row[h['Opis']] or ''), 'smjer': str(row[h['Smjer']] or ''),
            'iznos': round(float(row[h['Iznos']]), 2), 'kartica': str(row[h['Kartica']] or ''),
            'src': str(row[h['Src']] or ''), 'pref': pref, 'racun': racun, 'izvor': izvor,
        })
    wb.close()
    return out


def load_review():
    review = sorted([c for c in DATA_DIR.glob('Financije_review_*.xlsx') if '.pre-' not in c.name],
                    key=lambda p: p.stat().st_mtime, reverse=True)
    if not review:
        sys.exit('✗ Nema Financije_review_*.xlsx')
    path = review[0]
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Review']
    h = {str(c.value).strip(): i for i, c in enumerate(ws[1]) if c.value is not None}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = to_date(row[h['event_date']])
        if d is None:
            continue
        smjer = str(row[h['Smjer']] or '')
        amt = row[h['Uplata'] if smjer == 'Uplata' else h['Isplata']]
        if amt in (None, ''):
            continue
        rows.append({'date': d, 'smjer': smjer, 'iznos': round(float(amt), 2),
                     'racun': str(row[h['Racun']] or ''), 'izvor': str(row[h['Izvor']] or '')})
    wb.close()
    return path.name, rows


def main() -> None:
    dry = '--dry' in sys.argv[1:]
    tx = load_tx()
    review_name, rrows = load_review()
    print(f'Izvod tx: {len(tx)}  |  Review: {review_name} ({len(rrows)} redaka){"  [DRY]" if dry else ""}')

    # match index (za coverage) + amount index (za dijagnozu)
    idx = defaultdict(list)
    amt_idx = defaultdict(list)   # (racun, izvor, amount) -> [(smjer, date)]
    for i, r in enumerate(rrows):
        idx[(r['racun'], r['izvor'], r['smjer'], r['date'], r['iznos'])].append(i)
        amt_idx[(r['racun'], r['izvor'], r['iznos'])].append((r['smjer'], r['date']))
    used = set()
    for t in tx:
        hit = None
        for dd in DELTAS:
            key = (t['racun'], t['izvor'], t['smjer'], t['date'] + timedelta(days=dd), t['iznos'])
            hit = next((i for i in idx.get(key, []) if i not in used), None)
            if hit is not None:
                break
        t['matched'] = hit is not None
        if hit is not None:
            used.add(hit)

    def diagnose(t) -> str:
        cands = amt_idx.get((t['racun'], t['izvor'], t['iznos']), [])
        for sm, d in cands:
            if sm != t['smjer'] and abs((d - t['date']).days) <= 7:
                return 'Smjer? (u Reviewu suprotan smjer, isti iznos/datum)'
        low = t['opis'].lower()
        if t['smjer'] == 'Isplata' and any(k in low for k in IN_KW):
            return 'Smjer? (opis=priljev, vodi se Isplata)'
        near = [d for sm, d in cands if sm == t['smjer'] and abs((d - t['date']).days) <= 31]
        if near:
            return 'možda već u Reviewu (isti iznos, datum izvan ±7d)'
        if t['pref'] in ('MC', 'PBZVISA'):
            return 'kartična kupovina (nije itemizirana u Reviewu)'
        return 'nedostaje (nema kandidata u Reviewu)'

    missing = [t for t in tx if not t['matched']]
    for t in missing:
        t['problem'] = diagnose(t)

    # ── Report ───────────────────────────────────────────────────────────────
    cov = defaultdict(lambda: [0, 0])
    for t in tx:
        cov[t['pref']][0 if t['matched'] else 1] += 1
    print('\nCOVERAGE:')
    for p in ('ZABA', 'RF', 'MC', 'PBZVISA'):
        m, u = cov.get(p, [0, 0])
        print(f'  {p:8} matched {m:>5}  NEDOSTAJE {u:>5}')
    print(f'\nNEDOSTAJE ukupno: {len(missing)}  — Problem breakdown:')
    for prob, n in Counter(t['problem'] for t in missing).most_common():
        print(f'  {n:>4}  {prob}')

    if dry:
        print('\n[DRY] Ništa zapisano. Bez --dry: Nematchano_v1 + Nematchano_v2(Problem) + Coverage.')
        return

    backup = TX_XLSX.with_name(f'{TX_XLSX.stem}.pre-reconcile-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(TX_XLSX, backup)
    wb = openpyxl.load_workbook(TX_XLSX)

    if 'Nematchano' in wb.sheetnames and 'Nematchano_v1' not in wb.sheetnames:
        dst = wb.copy_worksheet(wb['Nematchano'])
        dst.title = 'Nematchano_v1'
        print('✔ Zamrznut Nematchano_v1 (baseline)')

    for name in ('Nematchano_v2', 'Coverage'):
        if name in wb.sheetnames:
            del wb[name]

    ws = wb.create_sheet('Nematchano_v2')
    heads = ('Datum', 'Smjer', 'Iznos', 'Opis', 'Kartica', 'Src', 'Racun', 'Izvor', 'Problem')
    for c, hh in enumerate(heads, 1):
        cell = ws.cell(1, c, hh)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
    order = {'Smjer?': 0}
    for r, t in enumerate(sorted(missing, key=lambda t: (0 if t['problem'].startswith('Smjer') else 1,
                                                          t['pref'], t['date'])), 2):
        ws.cell(r, 1, t['date']).number_format = 'DD.MM.YYYY'
        ws.cell(r, 2, t['smjer']); ws.cell(r, 3, t['iznos']).number_format = '#,##0.00'
        ws.cell(r, 4, t['opis']); ws.cell(r, 5, t['kartica']); ws.cell(r, 6, t['src'])
        ws.cell(r, 7, t['racun']); ws.cell(r, 8, t['izvor']); ws.cell(r, 9, t['problem'])
        if t['problem'].startswith('Smjer'):
            for c in range(1, 10):
                ws.cell(r, c).fill = SMJER_FILL
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:I{max(2, len(missing) + 1)}'
    for col, w in zip('ABCDEFGHI', (12, 8, 11, 52, 24, 24, 20, 11, 44)):
        ws.column_dimensions[col].width = w

    cs = wb.create_sheet('Coverage')
    cs.append(['Tip', 'matched', 'NEDOSTAJE'])
    for p in ('ZABA', 'RF', 'MC', 'PBZVISA'):
        m, u = cov.get(p, [0, 0])
        cs.append([p, m, u])
    cs.append([])
    cs.append(['Problem', 'Broj'])
    for prob, n in Counter(t['problem'] for t in missing).most_common():
        cs.append([prob, n])
    for c in ('A', 'B', 'C'):
        cs.column_dimensions[c].width = 46 if c == 'A' else 12
    for cell in cs[1] + cs[5]:
        cell.fill, cell.font = HDR_FILL, WHITE_BOLD

    try:
        wb.save(TX_XLSX)
    except PermissionError:
        sys.exit(f'✗ Zatvori {TX_XLSX.name} u Excelu i ponovi. (Backup: {backup.name})')
    print(f'✔ Zapisano: Nematchano_v2 ({len(missing)} + Problem) + Coverage. Backup: {backup.name}')


if __name__ == '__main__':
    main()
