"""
fix_pitanja_koka.py — primjena Kokinih odgovora iz sheeta "Pitanja za Koku". S107x.

Izvor odgovora: radna kopija `Financije_review-prolaz-s-Kokom.xlsx` (Odluka + Njena
napomena po pitanju, popunjeno uživo s Kokom). Cilj: PRAVI Review
(`Financije_review_20260710_1448.xlsx`).

Primijenjeno (sve po `source_key`, ne po broju retka):
  • #1 (89631181b5c9, red 4996 Parking)   — event_date+Datum naplate 07.08.2026 → 07.07.2026
  • #4 (e4fedafd629e, red 2787 Mirovina)  — event_date+Datum naplate 07.01.2025 → 07.02.2025
  • #4 (967ae386d73b, red 2788 Triglav)   — event_date+Datum naplate 07.01.2025 → 07.02.2025
  • #2 (4cd738a307f2, red 4997 MC 21,88)  — BRIŠE (duplikat reda 4247, PAYPAL *TEMU)
  • #5 (a7828c74171f, red 3609 Anja 450)  — BRIŠE (rata 72/96 već pokrivena 3612+3613)
  • #7 (da81b7fe4096, red 2004 Mirovina)  — BRIŠE (duplikat reda 2001)
  • #3, #6, #8-14 — bez izmjene u Reviewu; Odluka/Njena napomena samo se prepisuju u
    "Pitanja za Koku" sheet pravog Reviewa radi traga

Svaki par/redak se prije diranja verificira po `source_key`, iznosu i trenutnom datumu;
bilo koje odstupanje = izlaz bez pisanja.

Pokretanje:
  python fix_pitanja_koka.py --dry     # popis, ništa se ne piše
  python fix_pitanja_koka.py           # pravi run (oba Excela ZATVORENA; backup .pre-pitanja-*)
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
DATA = HERE.parent.parent / 'data-prep_data' / 'Financije'
REAL_REVIEW = DATA / 'Financije_review_20260710_1448.xlsx'
ANSWERS_COPY = DATA / 'Financije_review-prolaz-s-Kokom.xlsx'
QUESTIONS_SHEET = 'Pitanja za Koku'

# (source_key, novi_datum, opis, očekivani_iznos, kolona_iznosa, stari_datum_očekivan)
DATE_FIXES = [
    ('89631181b5c9', date(2026, 7, 7), 'red 4996 Parking', 1.60, 'Isplata', date(2026, 8, 7)),
    ('e4fedafd629e', date(2025, 2, 7), 'red 2787 Mirovina', 1125.07, 'Uplata', date(2025, 1, 7)),
    ('967ae386d73b', date(2025, 2, 7), 'red 2788 Triglav', 1260.58, 'Uplata', date(2025, 1, 7)),
]

# (source_key, opis, očekivani_iznos, kolona_iznosa)
DELETES = [
    ('4cd738a307f2', 'red 4997 MC 21,88 — duplikat reda 4247 (PAYPAL *TEMU)', 21.88, 'Isplata'),
    ('a7828c74171f', 'red 3609 Anja 450 — rata 72/96 pokrivena 3612+3613', 450.00, 'Uplata'),
    ('da81b7fe4096', 'red 2004 Mirovina — duplikat reda 2001', 1158.99, 'Uplata'),
]

MARK = 'S107x fix: Koka potvrdila'


def clean(v) -> str:
    return str(v).strip() if v is not None else ''


def as_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def header_map(ws) -> dict[str, int]:
    return {str(c.value).strip(): i for i, c in enumerate(next(ws.iter_rows(max_row=1)), 1) if c.value}


def index_by_key(ws, col: int) -> dict[str, int]:
    by_key: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        k = clean(ws.cell(r, col).value)
        if k:
            by_key.setdefault(k, r)
    return by_key


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--dry', action='store_true', help='pokaži plan, ne piši ništa')
    args = ap.parse_args()

    if not REAL_REVIEW.exists():
        sys.exit(f'✗ Nema {REAL_REVIEW}')
    if not ANSWERS_COPY.exists():
        sys.exit(f'✗ Nema {ANSWERS_COPY}')

    # ── učitaj Kokine odgovore iz radne kopije ──
    awb = openpyxl.load_workbook(ANSWERS_COPY, data_only=True, read_only=True)
    if QUESTIONS_SHEET not in awb.sheetnames:
        sys.exit(f'✗ Radna kopija nema sheet "{QUESTIONS_SHEET}"')
    aws = awb[QUESTIONS_SHEET]
    aH = header_map(aws)
    for c in ('Br', 'Odluka', 'Njena napomena'):
        if c not in aH:
            sys.exit(f'✗ "{QUESTIONS_SHEET}" u radnoj kopiji nema kolonu "{c}"')
    answers: dict[int, tuple[str, str]] = {}
    for r in range(2, aws.max_row + 1):
        br = aws.cell(r, aH['Br']).value
        if br is None:
            continue
        odluka = clean(aws.cell(r, aH['Odluka']).value)
        napomena = clean(aws.cell(r, aH['Njena napomena']).value)
        answers[int(br)] = (odluka, napomena)
    awb.close()

    print(f'Odgovora učitano iz radne kopije: {len(answers)}/14')

    # ── otvori pravi Review ──
    wb = openpyxl.load_workbook(REAL_REVIEW)
    ws = wb['Review']
    H = header_map(ws)
    for c in ('source_key', 'event_date', 'Datum naplate', 'Uplata', 'Isplata', 'Alternativa / nap.'):
        if c not in H:
            sys.exit(f'✗ Review nema kolonu "{c}"')
    by_key = index_by_key(ws, H['source_key'])

    if QUESTIONS_SHEET not in wb.sheetnames:
        sys.exit(f'✗ Pravi Review nema sheet "{QUESTIONS_SHEET}" (očekivan iz make_pitanja_koka.py)')
    pws = wb[QUESTIONS_SHEET]
    pH = header_map(pws)
    for c in ('Br', 'Odluka', 'Njena napomena'):
        if c not in pH:
            sys.exit(f'✗ "{QUESTIONS_SHEET}" u pravom Reviewu nema kolonu "{c}"')

    # ── verifikacija PRIJE ijedne izmjene ──
    problems: list[str] = []
    date_plan = []
    for key, new_dt, label, exp_amt, amt_col, exp_old_dt in DATE_FIXES:
        r = by_key.get(key)
        if not r:
            problems.append(f'{label}: source_key {key} nije nađen u Reviewu')
            continue
        amt = ws.cell(r, H[amt_col]).value
        if amt is None or round(float(amt), 2) != exp_amt:
            problems.append(f'{label} (red {r}): {amt_col} = {amt}, očekivano {exp_amt}')
            continue
        cur_ed = as_date(ws.cell(r, H['event_date']).value)
        if cur_ed != exp_old_dt:
            problems.append(f'{label} (red {r}): event_date = {cur_ed}, očekivano {exp_old_dt} '
                            f'(je li već ispravljeno?)')
            continue
        date_plan.append((r, key, new_dt, label))

    delete_plan = []
    for key, label, exp_amt, amt_col in DELETES:
        r = by_key.get(key)
        if not r:
            problems.append(f'{label}: source_key {key} nije nađen u Reviewu')
            continue
        amt = ws.cell(r, H[amt_col]).value
        if amt is None or round(float(amt), 2) != exp_amt:
            problems.append(f'{label} (red {r}): {amt_col} = {amt}, očekivano {exp_amt}')
            continue
        delete_plan.append((r, key, label))

    if problems:
        print('\n✗ Verifikacija nije prošla — NIŠTA nije dirnuto:')
        for p in problems:
            print(f'  {p}')
        sys.exit(1)

    print(f'\nVerificirano {len(date_plan)}/3 popravka datuma, {len(delete_plan)}/3 brisanja:')
    print('-' * 96)
    for r, key, new_dt, label in date_plan:
        print(f'  DATUM   red {r:>5}  → {new_dt.isoformat()}   {label}')
    for r, key, label in delete_plan:
        print(f'  BRIŠEM  red {r:>5}   {label}')
    print('-' * 96)
    print(f'  Pitanja za Koku — odgovora za prepis: {len(answers)}')

    if args.dry:
        print('\n… --dry: ništa nije zapisano.')
        return

    # ── upis: datumi, pa prepis Pitanja, pa brisanje odozdo prema gore ──
    for r, key, new_dt, label in date_plan:
        ws.cell(r, H['event_date']).value = new_dt
        ws.cell(r, H['Datum naplate']).value = new_dt
        old_note = clean(ws.cell(r, H['Alternativa / nap.']).value)
        note = f'{old_note} | {MARK}' if old_note else MARK
        ws.cell(r, H['Alternativa / nap.']).value = note

    for r in range(2, pws.max_row + 1):
        br = pws.cell(r, pH['Br']).value
        if br is None or int(br) not in answers:
            continue
        odluka, napomena = answers[int(br)]
        pws.cell(r, pH['Odluka']).value = odluka
        pws.cell(r, pH['Njena napomena']).value = napomena

    for r in sorted((p[0] for p in delete_plan), reverse=True):
        ws.delete_rows(r, 1)

    if ws.auto_filter.ref:
        first = ws.auto_filter.ref.split(':')[0]
        ws.auto_filter.ref = f'{first}:{get_column_letter(ws.max_column)}{ws.max_row}'

    backup = REAL_REVIEW.with_name(f'{REAL_REVIEW.stem}.pre-pitanja-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(REAL_REVIEW, backup)
    try:
        wb.save(REAL_REVIEW)
    except PermissionError:
        backup.unlink(missing_ok=True)
        sys.exit('✗ Zatvori Review u Excelu i ponovi. (Ništa nije zapisano.)')

    print(f'\n✔ {len(date_plan)} datuma ispravljeno · {len(delete_plan)} redaka obrisano · '
          f'{len(answers)} odgovora prepisano · Review sad {ws.max_row - 1} redaka')
    print(f'  Backup: {backup.name}')


if __name__ == '__main__':
    main()
