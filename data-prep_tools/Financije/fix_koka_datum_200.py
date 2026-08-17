# -*- coding: utf-8 -*-
"""
fix_koka_datum_200.py  (S110, 2026-08-17)  — ONE-OFF
================================================================================
Kokina tipfelerica u GODINI: podizanje 200,00 s bankomata datirano `2026-05-29`
umjesto `2025-05-29`.

DOKAZ (ne pretpostavka):
  • izvod `ZABA_2025-05.pdf` ima DVA podizanja po 200,00 — 19.05. i 29.05.2025.
    Review ima samo prvo (xl3325, `koka EU:1772`, Tip `Transfer`).
  • izvod `ZABA_2026-05.pdf` nema nijednu transakciju od 200,00.
  • sporni redak je `koka EU:1780` — OSAM redaka iza EU:1772 u Kokinom fileu, a
    datiran godinu dana kasnije. Njeno `Stanje` na njemu je 925,33, dakle u
    NJENOM lancu redak sjedi u svibnju 2025. (u svibnju 2026. joj je saldo ~2.900).
    Kriva je samo ćelija s datumom — ne i njen saldo.

POSLJEDICA: app sortira po datumu, Koka ne — pa je app taj redak odveo godinu
dana naprijed. Zbog toga se model razilazio +200,00 u svibnju 2025. i −200,00
u svibnju 2026. Nakon popravka ostatak kroz 19 mjeseci pada s −200,14 na −0,14.

--------------------------------------------------------------------------------
ŠTO MIJENJA — dvije ćelije, ništa više:
    event_date     2026-05-29 → 2025-05-29
    Datum naplate  2026-05-29 → 2025-05-29     (D1b: Izvor=Racun ⇒ isti dan)

ŠTO NAMJERNO NE DIRA:
  • `Stanje` (925,33) — Kokin lanac je NEOVISNI SVJEDOK protiv kojeg mjerimo
    app. Ispravljati ga značilo bi mjeriti se protiv vlastite korekcije.
  • `source_key` ('268b5fbe9013') — ⚠ zvuči kao da ga treba preračunati (ovisi
    o `seq_per_day`), ali NE treba: taj ključ je već vezan uz redak koji je
    UVEZEN u bazu. Preračunavanje bi prekinulo tu vezu i idempotencija
    `merge_pbzvisa.py` bi redak vidjela kao nov. Zastario ključ je ovdje
    ispravniji od svježeg.
  • `Tip`/`Napomena` — ostaju `N/A`/prazno. Sad kad je datum točan,
    `enrich_from_izvoda.py` može redak spariti s izvodom i `apply_rules.py`
    (pravilo #23) dati mu `Transfer`/`cash - bankomat`, kao njegovom blizancu.
    To je posao tih alata, ne ovog — ovdje bih vrijednosti izmislio.

Pokretanje:
    Financije\\run.bat fix_koka_datum_200.py --dry     (samo pokaže)
    Financije\\run.bat fix_koka_datum_200.py
"""
from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DATA = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")
REVIEW = DATA / 'Financije_review_20260710_1448.xlsx'

# Potpis retka — namjerno uzak. Ako se ne poklopi TOČNO jedan redak, alat staje:
# krivi redak s krivim datumom je gora šteta od nepokrenutog popravka.
SIG = {'Racun': 'Kokin tekući ZABA', 'Izvor reda': 'koka EU:1780',
       'source_key': '268b5fbe9013', 'Isplata': 200}
OLD = datetime(2026, 5, 29)
NEW = datetime(2025, 5, 29)


def main() -> None:
    dry = '--dry' in sys.argv[1:]
    if not REVIEW.exists():
        sys.exit(f'✗ Nema {REVIEW}')

    wb = openpyxl.load_workbook(REVIEW)
    ws = wb['Review']
    hdr = {c.value: c.column for c in ws[1] if c.value}
    for need in ('event_date', 'Datum naplate', *SIG):
        if need not in hdr:
            sys.exit(f'✗ Nedostaje kolona "{need}" u Review sheetu.')

    hits = []
    for r in range(2, ws.max_row + 1):
        if all(ws.cell(r, hdr[k]).value == v for k, v in SIG.items()) \
                and ws.cell(r, hdr['event_date']).value == OLD:
            hits.append(r)

    if len(hits) != 1:
        sys.exit(f'✗ Očekivan točno 1 redak s potpisom, nađeno {len(hits)}: {hits}\n'
                 f'  Review se u međuvremenu promijenio — provjeri prije nego pustiš popravak.')

    row = hits[0]
    print(f'Redak xl{row}:')
    print(f"  event_date     {ws.cell(row, hdr['event_date']).value} → {NEW}")
    print(f"  Datum naplate  {ws.cell(row, hdr['Datum naplate']).value} → {NEW}")
    print(f"  Stanje         {ws.cell(row, hdr['Stanje']).value}  (ne dira se — Kokin svjedok)")

    if dry:
        print('\n--dry: ništa nije zapisano.')
        return

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup = REVIEW.with_name(f'{REVIEW.stem}.pre-datum200-{stamp}.xlsx')
    shutil.copy2(REVIEW, backup)
    print(f'\nBackup: {backup.name}')

    ws.cell(row, hdr['event_date']).value = NEW
    ws.cell(row, hdr['Datum naplate']).value = NEW
    wb.save(REVIEW)
    print(f'✓ Zapisano u {REVIEW.name}')
    print('\n⚠ Baza NIJE dirana — event 29.05.2026. i dalje postoji.')
    print('  Ispravlja se kroz Edit Activity (delta-shift i provjera kolizije su tamo).')


if __name__ == '__main__':
    main()
