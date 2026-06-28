"""
Classify Financije events with Tip='N/A' by keyword matching on Napomena/comment.
Outputs an xlsx review file for Koka to verify before bulk update.

Usage:
  cd data-prep_tools
  python Financije/classify_na_events.py                      # PROD
  python Financije/classify_na_events.py --env .env.local     # TEST
"""

import os, sys, re
from pathlib import Path
from collections import Counter
from datetime import datetime

# ── Setup ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
REPO_ROOT  = SCRIPT_DIR.parent.parent
sys.path.insert(0, str(SCRIPT_DIR.parent / "Tools"))

def load_env(env_file: str) -> dict:
    env = {}
    with open(env_file, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env

# ── Keyword classification rules ─────────────────────────────────────────────
# Each rule: (tip, podtip_or_None, [keywords])
# First match wins — order matters (more specific first)

RULES = [
    # ── Domaćinstvo ──
    ("Domaćinstvo", "Struja",               ["struja", "hep", "elektr"]),
    ("Domaćinstvo", "Voda",                  ["voda", "vodovod"]),
    ("Domaćinstvo", "Holding (smeće)",       ["holding", "čistoća", "cistoća", "smeće", "smece"]),
    ("Domaćinstvo", "Plin",                  ["plin", "gpz", "gradska plinara"]),
    ("Domaćinstvo", "Bankovni troškovi",     ["provizija", "naknada za", "bankovni", "kamate", "kamata", "pristupnica"]),
    ("Domaćinstvo", "Popravci i održavanje", ["vodoinstalater", "keramičar", "elektricar", "električar", "majstor", "popravak kuć"]),
    ("Domaćinstvo", "Investicije",           ["investicij"]),
    ("Domaćinstvo", "Povrat Nataša",         ["nataša", "natasa", "natasha"]),
    ("Domaćinstvo", "Povrat Zoran",          ["zoran"]),

    # ── Informatika ──
    ("Informatika", "T-com",                 ["t-com", "tcom", "t com", "iskon", "optima"]),
    ("Informatika", "T-mobile",              ["t-mobile", "tmobile", "t mobile"]),
    ("Informatika", "Disney",                ["disney"]),
    ("Informatika", "Sky",                   ["sky"]),
    ("Informatika", "Prime",                 ["prime", "amazon"]),
    ("Informatika", "HBOmax",                ["hbo"]),
    ("Informatika", "Youtube",               ["youtube", "google"]),
    ("Informatika", "AudibleKoka",           ["audible koka", "audiblekoka"]),
    ("Informatika", "AudibleSasa",           ["audible saš", "audible sas", "audiblesaš", "audiblesas"]),
    ("Informatika", "Cloud backup",          ["cloud", "backup", "backblaze", "icloud"]),
    ("Informatika", "Microsoft",             ["microsoft", "office 365", "onedrive"]),
    ("Informatika", "HP",                    ["hewlett", " hp "]),
    ("Informatika", "Saša projekti",         ["claude", "anthropic", "perplexity", "autocad", "autodesk", "openai", "chatgpt"]),
    ("Informatika", None,                    ["netflix"]),  # Netflix is under T-com? or separate?

    # ── Ostavine ──
    ("Ostavine", "Advokati",                 ["advokat", "odvjetnik", "javni bilježnik", "javni biljeznik", "notar"]),

    # ── auto C5 ──
    ("auto C5", "gorivo",                    ["c5 gorivo", "citroen gorivo"]),
    ("auto C5", "registracija",              ["c5 registr", "citroen registr"]),
    ("auto C5", "popravci",                  ["c5 poprav", "citroen poprav", "c5 servis", "citroen servis"]),

    # ── auto Lacetti ──
    ("auto Lacetti", "gorivo",               ["lacetti gorivo", "chevrolet gorivo"]),
    ("auto Lacetti", "registracija",         ["lacetti registr"]),
    ("auto Lacetti", "popravci",             ["lacetti poprav", "lacetti servis"]),

    # ── Generic auto (can't determine which car) ──
    ("auto C5", "gorivo",                    ["gorivo", "benzin", "dizel", "ina ", "petrol", "tifon", "mol ", "lukoil"]),
    ("auto C5", "parking",                   ["parking"]),
    ("auto C5", "registracija",              ["registracija", "tehnički pregled", "tehnicki pregled"]),
    ("auto C5", None,                        ["auto", "gume", "vulkanizer"]),

    # ── Putovanja ──
    ("Putovanja", "karte",                   ["karta", "karte", "avion", "bus ", "vlak", "hž ", "flixbus"]),
    ("Putovanja", "smještaj",                ["hotel", "hostel", "apartman", "booking", "airbnb", "smještaj", "smjestaj"]),
    ("Putovanja", "restoran",                ["restoran", "restaurant"]),
    ("Putovanja", None,                      ["cestarina", "hac ", "autoput", "enc ", "putarina"]),

    # ── Zdravlje ──
    ("Zdravlje", "PP",                       ["posmrtn", "pogreb"]),
    ("Zdravlje", "Lječnička komora",         ["komora", "lječnič", "liječnič", "ljecnic"]),
    ("Zdravlje", "Medical",                  ["ljekarna", "lekarna", "apoteka", "doktor", "liječnik", "lijecnik",
                                              "bolnica", "klinika", "ordinacija", "medical", "zdravstven",
                                              "laboratorij", "rtg", "ct ", "ultrazvuk", "mamograf"]),
    ("Zdravlje", "PassSport",               ["passport", "pass sport", "teretana", "gym", "fitness"]),
    ("Zdravlje", "Sportski rekviziti",       ["sportsk", "tenisic", "patike", "dres"]),

    # ── Ostalo — specific ──
    ("Ostalo", "Odjeća/obuća",               ["odjeća", "odjeca", "obuća", "obuca", "cipele", "čizme", "cizme",
                                              "jakna", "hlače", "hlace", "košulja", "kosulja", "h&m", "zara",
                                              "c&a", "reserved", "new yorker"]),
    ("Ostalo", "Pokloni",                    ["poklon", "dar ", "rođendan", "rodjendan", "božić", "bozic"]),
    ("Ostalo", "Kave/jelo vani",             ["kava", "kafe", "café", "caffe", "kavana", "pizz",
                                              "mcdonald", "mcdonalds", "burger", "pekara", "slastičar"]),
    ("Ostalo", "Temu",                       ["temu"]),
    ("Ostalo", "Taksi",                      ["taksi", "taxi", "uber", "bolt"]),
    ("Ostalo", "Kino/Kazalište/Muzeji",      ["kino", "kazalište", "kazaliste", "muzej", "koncert", "predstava"]),

    # ── Uplata specific ──
    ("Mirovina", "Saša",                     ["mirovina saš", "mirovina sas", "i stup saš", "i stup sas"]),
    ("Mirovina", "Koka",                     ["mirovina kok", "mirovina drag", "i stup kok", "i stup drag"]),
    ("Mirovina", None,                       ["mirovina", "i stup", "ii stup"]),
    ("Najam", "Anja",                        ["anja"]),
    ("Transfer", None,                       ["transfer", "prijenos", "prebacivanje"]),
    ("Ostali prihodi", None,                 ["povrat poreza", "porezna uprava"]),
]


def classify(text: str) -> tuple[str | None, str | None]:
    """Return (tip, podtip) based on keyword match, or (None, None)."""
    if not text:
        return (None, None)
    lower = text.lower()
    for tip, podtip, keywords in RULES:
        for kw in keywords:
            if kw in lower:
                return (tip, podtip)
    return (None, None)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--env", default=".env.prod.local")
    parser.add_argument("--area", default="Financije")
    parser.add_argument("--limit", type=int, default=0, help="0 = all")
    args = parser.parse_args()

    env_path = REPO_ROOT / args.env
    if not env_path.exists():
        print(f"ERROR: {env_path} not found")
        sys.exit(1)

    env = load_env(str(env_path))
    url = env.get("SUPABASE_URL") or env.get("VITE_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: need SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY")
        sys.exit(1)

    from supabase import create_client
    sb = create_client(url, key)

    # Find area
    areas_res = sb.table("areas").select("id, name").ilike("name", f"{args.area}%").execute()
    if not areas_res.data:
        print(f"ERROR: no area matching '{args.area}'")
        sys.exit(1)

    for area in areas_res.data:
        print(f"\n{'='*70}")
        print(f"Area: {area['name']} ({area['id'][:8]}...)")
        print(f"{'='*70}")
        process_area(sb, area, args.limit)


def process_area(sb, area, limit):
    area_id = area["id"]

    # Get categories for this area
    cats = sb.table("categories").select("id, name").eq("area_id", area_id).execute().data
    cat_ids = [c["id"] for c in cats]
    if not cat_ids:
        print("  No categories found")
        return

    # Get attr defs for slug lookup
    all_attr_defs = []
    for cid in cat_ids:
        res = sb.table("attribute_definitions").select("id, name, slug, category_id").eq("category_id", cid).execute()
        all_attr_defs.extend(res.data)

    tip_def_ids = [a["id"] for a in all_attr_defs if a["name"] == "Tip"]
    nap_def_ids = [a["id"] for a in all_attr_defs if a["name"] == "Napomena"]
    smjer_def_ids = [a["id"] for a in all_attr_defs if a["name"] == "Smjer"]

    if not tip_def_ids:
        print("  No 'Tip' attribute found")
        return

    # Get all events
    query = sb.table("events").select("id, event_date, comment, category_id").in_("category_id", cat_ids).order("event_date")
    if limit > 0:
        query = query.limit(limit)

    # Paginate events (Supabase 1000-row cap)
    all_events = []
    page_size = 1000
    offset = 0
    while True:
        res = query.range(offset, offset + page_size - 1).execute()
        all_events.extend(res.data)
        if len(res.data) < page_size:
            break
        offset += page_size

    print(f"  Total events: {len(all_events)}")

    # Get event_attributes for Tip, Napomena, Smjer (chunked)
    event_ids = [e["id"] for e in all_events]
    relevant_def_ids = tip_def_ids + nap_def_ids + smjer_def_ids

    all_attrs = []
    chunk_size = 200
    for i in range(0, len(event_ids), chunk_size):
        chunk = event_ids[i:i+chunk_size]
        res = sb.table("event_attributes").select(
            "event_id, attribute_definition_id, value_text"
        ).in_("event_id", chunk).in_("attribute_definition_id", relevant_def_ids).execute()
        all_attrs.extend(res.data)

    # Build lookup maps
    tip_map = {}    # event_id → tip value
    nap_map = {}    # event_id → napomena value
    smjer_map = {}  # event_id → smjer value
    tip_set = set(tip_def_ids)
    nap_set = set(nap_def_ids)
    smjer_set = set(smjer_def_ids)

    for a in all_attrs:
        eid = a["event_id"]
        did = a["attribute_definition_id"]
        val = a["value_text"] or ""
        if did in tip_set:
            tip_map[eid] = val
        elif did in nap_set:
            nap_map[eid] = val
        elif did in smjer_set:
            smjer_map[eid] = val

    # Filter N/A events
    na_events = []
    for e in all_events:
        tip_val = tip_map.get(e["id"], "")
        if tip_val == "N/A" or tip_val == "" or tip_val is None:
            na_events.append({
                "event_id": e["id"],
                "date": e["event_date"],
                "comment": e["comment"] or "",
                "napomena": nap_map.get(e["id"], ""),
                "smjer": smjer_map.get(e["id"], ""),
                "old_tip": tip_val,
            })

    print(f"  N/A or empty Tip events: {len(na_events)}")

    # Classify
    classified = 0
    unclassified = 0
    tip_counts = Counter()
    rows = []

    for ev in na_events:
        # Try classification on napomena first, then comment
        search_text = f"{ev['napomena']} {ev['comment']}"
        new_tip, new_podtip = classify(search_text)

        if new_tip:
            classified += 1
            tip_counts[new_tip] += 1
        else:
            unclassified += 1
            tip_counts["??? NEPOZNATO"] += 1

        rows.append({
            **ev,
            "new_tip": new_tip or "???",
            "new_podtip": new_podtip or "",
        })

    print(f"\n  Classified: {classified} / {len(na_events)}")
    print(f"  Unclassified: {unclassified}")
    print(f"\n  Distribution:")
    for tip, cnt in tip_counts.most_common():
        print(f"    {tip:30s} {cnt:5d}")

    # Write xlsx
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("\n  pip install openpyxl to get xlsx output")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Klasifikacija"

    headers = ["date", "smjer", "napomena", "comment", "old_tip", "new_tip", "new_podtip", "KOREKCIJA_TIP", "KOREKCIJA_PODTIP"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")

    red_fill = PatternFill("solid", fgColor="FFC7CE")
    green_fill = PatternFill("solid", fgColor="C6EFCE")

    for r, row in enumerate(rows, 2):
        ws.cell(r, 1, row["date"])
        ws.cell(r, 2, row["smjer"])
        ws.cell(r, 3, row["napomena"])
        ws.cell(r, 4, row["comment"])
        ws.cell(r, 5, row["old_tip"])
        tip_cell = ws.cell(r, 6, row["new_tip"])
        ws.cell(r, 7, row["new_podtip"])
        # Empty columns for Koka to fill corrections
        ws.cell(r, 8, "")  # KOREKCIJA_TIP
        ws.cell(r, 9, "")  # KOREKCIJA_PODTIP

        if row["new_tip"] == "???":
            tip_cell.fill = red_fill
        else:
            tip_cell.fill = green_fill

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Summary sheet
    ws2 = wb.create_sheet("Sažetak")
    ws2.cell(1, 1, "Tip").font = Font(bold=True)
    ws2.cell(1, 2, "Broj").font = Font(bold=True)
    for r, (tip, cnt) in enumerate(tip_counts.most_common(), 2):
        ws2.cell(r, 1, tip)
        ws2.cell(r, 2, cnt)

    out_dir = SCRIPT_DIR.parent.parent / "data-prep_data" / "Financije"
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = out_dir / f"classify_NA_{area['name']}_{ts}.xlsx"
    wb.save(str(out_path))
    print(f"\n  Output: {out_path}")


if __name__ == "__main__":
    main()
