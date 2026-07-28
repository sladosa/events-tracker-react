"""
fix_vocarna_pravilo.py — pravilo `voce i povrce` iznad #43 AGRAM + ispravak reda. S107o.

NALAZ (S107n §2l): `Agram - voce i povrce Zagreb`, 10,33 € — vočarna koja se slučajno
zove Agram, pa ju je pravilo #43 (`AGRAM` → `auto Lacetti | registracija`) pokupilo.

Priority-order pattern (S107l, `grobn` iznad `NAKNADA`): specifičnije pravilo se UMEĆE
IZNAD preširokog, jer u `apply_rules.py` prvi match pobjeđuje, a NOT/exclusion sintaksa
ne postoji.

⚠ Pravilo samo NE POPRAVLJA postojeći redak. `apply_rules.py` (linija ~516) preskače
svaki redak čiji je par VALJAN u Taksonomiji — a `auto Lacetti | registracija` jest
valjan, samo je kriv. Zato ovaj alat radi oboje:
  1. umeće pravilo u `Pravila` (štiti budući import)
  2. ispravlja taj jedan redak (traži ga po `source_key`, ne po broju — retci se pomiču)

Pokretanje:
  python fix_vocarna_pravilo.py --dry
  python fix_vocarna_pravilo.py            # Excel ZATVOREN; backup .pre-vocarna-*
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
DATA = HERE.parent.parent / 'data-prep_data' / 'Financije'

KEYWORD = 'voce i povrce'
NEW_TIP, NEW_POD = 'Namirnice', 'Hrana i ostalo'
ANCHOR = 'AGRAM'                     # pravilo iznad kojeg se umeće
ROW_KEY = 'fad46efbb6cd'             # Voćarna, 2026-03-13, 10,33 €
NOTE = 'ispravak S107o: vočarna, ne autocentar'


def pick_review() -> Path:
    cands = [p for p in DATA.glob('Financije_review_*.xlsx')
             if '.pre-' not in p.name and not p.name.startswith('~$')]
    if not cands:
        sys.exit(f'✗ Nema Review filea u {DATA}')
    return max(cands, key=lambda p: p.stat().st_mtime)


def clean(v) -> str:
    return str(v).strip() if v is not None else ''


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--dry', action='store_true')
    args = ap.parse_args()

    review = pick_review()
    print(f'Review: {review.name}')
    wb = openpyxl.load_workbook(review)
    pr, ws = wb['Pravila'], wb['Review']

    # ── 1. pravilo ──
    hdr = [clean(c.value) for c in next(pr.iter_rows(max_row=1))]
    if hdr[:3] != ['Ključne riječi', 'Tip', 'Podtip']:
        sys.exit(f'✗ Neočekivan header Pravila sheeta: {hdr[:3]}')
    anchor_row = next((r for r in range(2, pr.max_row + 1)
                       if clean(pr.cell(r, 1).value).upper() == ANCHOR), None)
    if anchor_row is None:
        sys.exit(f'✗ Pravilo "{ANCHOR}" nije nađeno — provjeri Pravila sheet.')
    already = next((r for r in range(2, pr.max_row + 1)
                    if clean(pr.cell(r, 1).value).lower() == KEYWORD), None)
    if already:
        print(f'✔ Pravilo "{KEYWORD}" već postoji (red {already}) — preskačem umetanje.')
    else:
        print(f'  Pravilo "{KEYWORD}" → {NEW_TIP} | {NEW_POD}')
        print(f'    umeće se na red {anchor_row}, IZNAD "{ANCHOR}" '
              f'(#{anchor_row - 1} → #{anchor_row}), prvi match pobjeđuje')

    # ── 2. redak ──
    H = {clean(c.value): i for i, c in enumerate(next(ws.iter_rows(max_row=1)), 1) if c.value}
    row = next((r for r in range(2, ws.max_row + 1)
                if clean(ws.cell(r, H['source_key']).value) == ROW_KEY), None)
    if row is None:
        sys.exit(f'✗ Redak sa source_key {ROW_KEY} nije nađen — je li već popravljen ili obrisan?')
    if KEYWORD not in clean(ws.cell(row, H['Izvod opis']).value).lower():
        sys.exit(f'✗ Red {row}: "Izvod opis" ne sadrži "{KEYWORD}" — pogrešan redak, izlazim.')
    old = (clean(ws.cell(row, H['Tip']).value), clean(ws.cell(row, H['Podtip']).value))
    print(f'  Red {row} ({clean(ws.cell(row, H["Napomena"]).value)}, '
          f'{ws.cell(row, H["Isplata"]).value} €): '
          f'{old[0]} | {old[1]}  →  {NEW_TIP} | {NEW_POD}')
    if old == (NEW_TIP, NEW_POD):
        print('    (već ispravljen)')

    if args.dry:
        print('\n… --dry: ništa nije zapisano.')
        return

    changed = False
    if not already:
        pr.insert_rows(anchor_row, 1)
        for c, v in enumerate((KEYWORD, NEW_TIP, NEW_POD, None, NOTE), 1):
            pr.cell(anchor_row, c, v)
        changed = True
    if old != (NEW_TIP, NEW_POD):
        stamp = datetime.now().strftime('%Y-%m-%d %H:%M')
        ws.cell(row, H['Tip'], NEW_TIP)
        ws.cell(row, H['Podtip'], NEW_POD)
        ws.cell(row, H['Pouzdanost'], 'PRAVILO')
        ws.cell(row, H['Alternativa / nap.'], f'pravilo "{KEYWORD}" — {NOTE}')
        ws.cell(row, H['Pravilo run'], stamp)      # ostaje u tvom audit filteru
        changed = True
    if not changed:
        print('\n✔ Nema promjena — Review netaknut.')
        return

    backup = review.with_name(f'{review.stem}.pre-vocarna-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        backup.unlink(missing_ok=True)
        sys.exit('✗ Zatvori Review u Excelu i ponovi. (Ništa nije zapisano.)')
    print(f'\n✔ Snimljeno. Backup: {backup.name}')


if __name__ == '__main__':
    main()
