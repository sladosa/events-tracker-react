# -*- coding: utf-8 -*-
"""
fix_tcom_tmobile_swap.py  (S107g, 2026-07-16, jednokratna korekcija)
=====================================================================
Preimenovanja je blanket-preimenovao SVE stare 'Informatika/T-com' (41) i
'Informatika/T-mobile' (40) retke u Komunikacije_T-com / Komunikacije_T-mobile
prema STAROM (Kokinom) Napomena-labelu. Taj label je za pokoji redak bio krivo
upisan — stvarni sadržaj Izvod opisa ('usluge fiksne mreže' vs 'usluge u mobilnoj
mreži') to otkriva. Analiza: T-com bucket 28 fiksna/1 mobilna/12 bez teksta;
T-mobile bucket 28 mobilna/1 fiksna/11 bez teksta — samo 1+1 stvarni mismatch.

Ograničeno SAMO na Tip_O=Informatika, Podtip_O in (T-com, T-mobile) — ne dira
ništa drugo. Ako i 'fiksn' i 'mobiln' matchaju (nejasno) ili nijedno — NE DIRA.

Pokretanje (file zatvoren u Excelu!):
  Financije\\run.bat fix_tcom_tmobile_swap.py            → najnoviji review file
  Financije\\run.bat fix_tcom_tmobile_swap.py --dry      → samo pokaži
"""

import shutil
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")

TARGET = {
    'fiksna':  ('Informatika', 'Komunikacije_T-com (internet, MaxTv)'),
    'mobilna': ('Informatika', 'Komunikacije_T-mobile'),
}


def fold(s) -> str:
    s = unicodedata.normalize('NFD', str(s or ''))
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower()


def find_header_col(ws, header: str) -> int:
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() == header:
            return c
    sys.exit(f'✗ Kolona "{header}" nije nađena u Review sheetu.')


def pick_file(args: list[str]) -> Path:
    explicit = [a for a in args if not a.startswith('--')]
    if explicit:
        return Path(explicit[0])
    candidates = sorted(DATA_DIR.glob('Financije_review_*.xlsx'),
                         key=lambda p: p.stat().st_mtime, reverse=True)
    candidates = [c for c in candidates if '.pre-' not in c.name]
    if not candidates:
        sys.exit(f'✗ Nema Financije_review_*.xlsx u {DATA_DIR}')
    return candidates[0]


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    path = pick_file(args)
    print(f'File: {path.name}{"  [DRY RUN]" if dry else ""}')

    wb = openpyxl.load_workbook(path)
    ws = wb['Review']
    col_nap   = find_header_col(ws, 'Napomena')
    col_izvod = find_header_col(ws, 'Izvod opis')
    col_tip   = find_header_col(ws, 'Tip')
    col_pod   = find_header_col(ws, 'Podtip')
    col_alt   = find_header_col(ws, 'Alternativa / nap.')
    col_tip_o = find_header_col(ws, 'Tip_O')
    col_pod_o = find_header_col(ws, 'Podtip_O')

    n_fixed = n_skipped = 0
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, col_tip_o).value or '').strip() != 'Informatika':
            continue
        pod_o = str(ws.cell(r, col_pod_o).value or '').strip()
        if pod_o not in ('T-com', 'T-mobile'):
            continue
        text = fold(ws.cell(r, col_nap).value) + ' | ' + fold(ws.cell(r, col_izvod).value)
        fiksna = 'fiksn' in text and 'mrez' in text
        mobilna = 'mobiln' in text and 'mrez' in text
        if fiksna == mobilna:
            continue   # oba ili nijedno — nejasno, ne diramo
        want = 'fiksna' if fiksna else 'mobilna'
        want_tip, want_pod = TARGET[want]
        cur_pod = str(ws.cell(r, col_pod).value or '').strip()
        if cur_pod == want_pod:
            continue   # već ispravno
        n_fixed += 1
        print(f'  red {r}: bio {pod_o} → stvarno "{want}" mreža → {want_tip}/{want_pod}')
        if not dry:
            ws.cell(r, col_tip, want_tip)
            ws.cell(r, col_pod, want_pod)
            old_alt = str(ws.cell(r, col_alt).value or '').strip()
            mark = f'RUČNO S107g: {pod_o}→{want_pod} (Izvod opis otkrio krivi stari label)'
            ws.cell(r, col_alt, f'{old_alt} | {mark}' if old_alt else mark)

    print(f'\n{"Bi se ispravilo" if dry else "Ispravljeno"}: {n_fixed} redova')

    if dry or not n_fixed:
        return

    backup = path.with_name(f'{path.stem}.pre-tcomswap-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori file u Excelu i ponovi. (Backup: {backup.name})')
    print(f'✔ Snimljeno. Backup: {backup.name}')


if __name__ == '__main__':
    main()
