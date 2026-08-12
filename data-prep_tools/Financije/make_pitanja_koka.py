# -*- coding: utf-8 -*-
"""
make_pitanja_koka.py — sheet `Pitanja za Koku`. S107x, 2026-08-12.

Otvorena pitanja iz migracije skupljena na jedno mjesto, u obliku na koji Koka može
odgovoriti. Nije njen alat — ovo je stranica koju Saša otvori dok sjedi s njom
(2026-08-12: ona još ne otvara Sašine fileove; TEST baza, PROD tek nakon odluke).

Obrazac je preuzet od `Nematchano_v3` (S107k), koji je s `Verdikt` dropdownom otišao
41 → 0: kontekst u retku, jedna kolona za odluku, `--harvest` je primijeni.

⚠ Ključna kolona je `U čemu je nejasnoća` — pitanje mora reći ŠTO nije bilo moguće
zaključiti i ZAŠTO. „Provjeri redak 2787" nije pitanje na koje se može odgovoriti.

Dvije vrste pitanja, namjerno tim redoslijedom:
  · `redak`  — konkretna transakcija koju prepozna; ide PRVO (lakše kreće)
  · `mjesec` — `Saldo kontrola` razlika; nema retka za pokazati, pita se sjećanje

Idempotentno: sheet se briše i ponovo kreira. `Odluka`/`Njena napomena` se pritom
GUBE — ako je već popunjavala, prvo napravi `--harvest` (nije još implementiran;
za sada kopiraj odgovore prije regeneracije).

Pokretanje:
    python make_pitanja_koka.py --dry
    python make_pitanja_koka.py
Backup: `.pre-pitanja-*` prije snimanja.
"""
from __future__ import annotations

import shutil
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent))
from apply_rules import find_header_col, pick_file  # noqa: E402

SHEET = 'Pitanja za Koku'
HDR_FILL = PatternFill('solid', fgColor='C55A11')
WHITE_BOLD = Font(color='FFFFFF', bold=True)
FILL_REDAK = PatternFill('solid', fgColor='FFF2CC')   # konkretna transakcija
FILL_MJESEC = PatternFill('solid', fgColor='DDEBF7')  # razlika na razini mjeseca
ODLUKE = ('točno je', 'datum je kriv', 'iznos je kriv', 'duplikat — obriši',
          'nedostaje zapis', 'ne sjećam se')

COLS = [('Br', 4), ('Vrsta', 8), ('Datum', 11), ('Iznos', 11), ('Račun / Izvor', 20),
        ('Njen opis', 26), ('Što kaže banka', 30), ('U čemu je nejasnoća', 62),
        ('Odluka', 18), ('Njena napomena', 30), ('Ref (red / source_key)', 26)]

# ── Pitanja ────────────────────────────────────────────────────────────────
# (vrsta, datum, iznos, račun/izvor, njen opis, što kaže banka, nejasnoća, ref)
PITANJA: list[tuple] = [
    ('redak', date(2026, 8, 7), -1.60, 'Kokin ZABA / Racun', 'Parking',
     'nema na izvodu',
     'Redak stoji na 07.08., ali ga tvoj `Stanje` lanac smješta između 04. i 08.07. — '
     'i tamo se slaže u cent s obje strane. Je li datum 07.07. umjesto 07.08.?',
     'red 4996 · 89631181b5c9'),
    ('redak', date(2026, 12, 1), -21.88, 'Kokin ZABA / Mastercard', '(prazno)',
     'nema na izvodu',
     'Datum je 01.12.2026., ali `Datum naplate` je 11.02.2026. — deset mjeseci PRIJE '
     'kupnje. Jedina Mastercard transakcija od 21,88 € u tom razdoblju je PAYPAL *TEMU '
     'od 31.12.2025., a nju već imamo. Je li ovo isti trošak upisan dvaput?',
     'red 4997 · 4cd738a307f2 (usporedi red 4247)'),
    ('redak', date(2025, 11, 26), -700.00, 'Kokin ZABA / Racun',
     'Podizanje gotovog novca - debitnom karticom', 'nema na izvodu',
     'Podizanje od 700 € ne postoji na ZABA izvodu. Banka u 11.–12.2025. bilježi '
     'bankomate 100 + 150 + 100 + 200 €. Je li ovih 700 zbroj više podizanja, ili je '
     'iznos drukčiji?',
     'red 4101 · 7887ece5a2de'),
    ('redak', date(2025, 1, 7), 1125.07, 'Kokin ZABA / Racun', 'Mirovina',
     'banka je vidi tek u veljači',
     'Mirovina (1.125,07) i Triglav (1.260,58) stoje na 07.01.2025., ali ih banka knjiži '
     'tek u veljači — točno tih 2.385,65 € nedostaje u veljači, koliko ih u siječnju ima '
     'viška. Jesu li stvarno stigle 07.01., ili si ih upisala unaprijed?',
     'redovi 2787 + 2788'),
    ('redak', date(2025, 7, 25), 450.00, 'Kokin ZABA / Racun', 'Anja 72/96',
     'M-ZABA UPLATA ANJA CRNKOVIĆ 400 + 50',
     'Ratu 72/96 imamo dvaput: tvojih 450 € u jednom retku, i bankovnih 400 + 50 € u dva '
     'retka istog dana. Je li Anja tu ratu platila u dva dijela (pa tvoj redak treba '
     'obrisati), ili je bilo dvije uplate?',
     'red 3609 vs redovi 3612 + 3613'),
    ('redak', date(2024, 10, 4), -236.04, 'Kokin ZABA / Racun', 'Allianz Lacetti',
     'nema na izvodu u tom razdoblju',
     'Allianz za Lacetti od 236,04 € banka ne bilježi u listopadu 2024. Je li plaćen '
     'drugim putem (kartica, gotovina), ili je datum drukčiji?',
     'red 2368 · 0f93ec36ee38'),
    ('redak', date(2024, 7, 8), 1047.51, 'Kokin ZABA / Racun', 'Mirovina',
     'izvod potvrđuje samo jednu uplatu',
     'Na 08.07.2024. stoje dvije Mirovine — 1.047,51 i 1.158,99 €. Izvod ima samo jednu '
     '(UPLATA MIROVINSKOG PRIMANJA). Jesu li to dva stupa koja stvarno stižu isti dan, '
     'ili je jedna upisana dvaput?',
     'redovi 2001 + 2004'),
    # ── razlike na razini mjeseca (`Saldo kontrola`) ───────────────────────
    ('mjesec', date(2026, 1, 30), 359.43, 'Kokin ZABA', 'izvadak 2026-01',
     'tvoje stanje 645,96 · banka 286,53',
     'Na dan zatvaranja izvatka tvoje stanje je 359,43 € VIŠE od bankovnog. To je '
     'najveća neobjašnjena razlika. Sjećaš li se čega u siječnju 2026. — možda trošak '
     'koji nije upisan?',
     'Saldo kontrola, redak 2026-01'),
    ('mjesec', date(2024, 10, 1), 149.00, 'Kokin ZABA', 'izvadak 2024-09',
     'tvoje stanje 2.014,48 · banka 1.865,48',
     'Razlika od 149,00 €. Okrugao iznos — možda jedan trošak koji nije zabilježen. '
     'Sjećaš li se čega u rujnu 2024.?',
     'Saldo kontrola, redak 2024-09'),
    ('mjesec', date(2024, 1, 1), 49.00, 'Kokin ZABA', 'izvadak 2023-12',
     'tvoje stanje 893,83 · banka 844,83',
     'Razlika je točno 49,00 € = jedna mjesečna Multisport uplata. Multisport nedostaje '
     'u 4 mjeseca (07/2023, 11/2023, 06/2024, 03/2025) — je li u tim mjesecima ipak '
     'plaćen, samo nije upisan?',
     'Saldo kontrola, redak 2023-12'),
    ('mjesec', date(2024, 3, 1), 49.00, 'Kokin ZABA', 'izvadak 2024-02',
     'tvoje stanje 5.214,61 · banka 5.165,61',
     'Opet točno 49,00 € — ista priča kao gore (Multisport). Jedan odgovor vjerojatno '
     'pokriva oba mjeseca.',
     'Saldo kontrola, redak 2024-02'),
    ('mjesec', date(2026, 6, 1), 8.40, 'Kokin ZABA', 'izvadak 2026-05',
     'tvoje stanje 2.965,24 · banka 2.956,84',
     'Sitna razlika 8,40 €. Ako se ne sjećaš, u redu je — zabilježit ćemo da ostaje '
     'nerazjašnjena.',
     'Saldo kontrola, redak 2026-05'),
    ('mjesec', date(2025, 12, 1), 1.60, 'Kokin ZABA', 'izvadak 2025-11',
     'tvoje stanje 1.412,92 · banka 1.411,32',
     'Sitna razlika 1,60 € (iznos parkinga). Ako se ne sjećaš, ostaje nerazjašnjena.',
     'Saldo kontrola, redak 2025-11'),
    ('mjesec', date(2025, 7, 1), 0.70, 'Kokin ZABA', 'izvadak 2025-06',
     'tvoje stanje 987,21 · banka 986,51',
     'Sitna razlika 0,70 €. Ako se ne sjećaš, ostaje nerazjašnjena.',
     'Saldo kontrola, redak 2025-06'),
]


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    path = pick_file(args)
    print(f'File: {path.name}{"  [DRY RUN — bez snimanja]" if dry else ""}\n')

    wb = openpyxl.load_workbook(path)
    if 'Review' not in wb.sheetnames:
        sys.exit('✗ Nema `Review` sheeta — je li ovo pravi file?')

    # kontrola: postoje li referencirani retci (da pitanja ne pokazuju u prazno)
    ws_r = wb['Review']
    c_key = find_header_col(ws_r, 'source_key')
    keys = {str(ws_r.cell(r, c_key).value or '').strip()
            for r in range(2, ws_r.max_row + 1)}
    for _, _, _, _, _, _, _, ref in PITANJA:
        for tok in ref.replace('·', ' ').split():
            if len(tok) == 12 and all(ch in '0123456789abcdef' for ch in tok):
                if tok not in keys:
                    sys.exit(f'✗ source_key {tok} nije u Reviewu — PREKID.')

    if SHEET in wb.sheetnames:
        print(f'· postojeći `{SHEET}` se briše i ponovo kreira '
              f'(⚠ odgovori u njemu se gube)')
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, 1)

    for c, (name, w) in enumerate(COLS, 1):
        cell = ws.cell(1, c, name)
        cell.fill, cell.font = HDR_FILL, WHITE_BOLD
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[chr(64 + c)].width = w
    ws.row_dimensions[1].height = 30

    dv = DataValidation(type='list', formula1=f'"{",".join(ODLUKE)}"', allow_blank=True,
                        promptTitle='Odluka',            # ≤32 znaka
                        prompt='Odaberi što je s ovim zapisom. Ako ništa ne odgovara, '
                               'ostavi prazno i upiši u kolonu desno.')  # ≤255
    ws.add_data_validation(dv)

    for i, (vrsta, d, iznos, rac, nom, banka, nejasno, ref) in enumerate(PITANJA, 1):
        r = i + 1
        fill = FILL_REDAK if vrsta == 'redak' else FILL_MJESEC
        vals = [i, vrsta, d, iznos, rac, nom, banka, nejasno, None, None, ref]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.fill = fill
            cell.alignment = Alignment(vertical='top', wrap_text=c in (6, 7, 8, 10))
            if c == 3:
                cell.number_format = 'DD.MM.YYYY'
            elif c == 4:
                cell.number_format = '#,##0.00'
        ws.cell(r, 8).font = Font(bold=True)
        dv.add(ws.cell(r, 9))
        ws.row_dimensions[r].height = 46

    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:K{len(PITANJA) + 1}'
    ws.sheet_view.tabColor = 'C55A11'

    n_red = sum(1 for p in PITANJA if p[0] == 'redak')
    print(f'`{SHEET}`: {len(PITANJA)} pitanja '
          f'({n_red} na razini retka, {len(PITANJA) - n_red} na razini mjeseca)')
    for i, p in enumerate(PITANJA, 1):
        print(f'   {i:>2}. [{p[0]:<6}] {p[1]}  {p[2]:>+10.2f}  {p[7]}')

    if dry:
        print('\nDRY RUN — ništa nije snimljeno.')
        return

    backup = path.with_name(f'{path.stem}.pre-pitanja-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori {path.name} u Excelu i ponovi.\n'
                 f'  (backup je već napravljen: {backup.name})')
    print(f'\n✔ Snimljeno. Backup: {backup.name}')


if __name__ == '__main__':
    main()
