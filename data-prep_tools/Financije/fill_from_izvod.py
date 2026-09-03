# -*- coding: utf-8 -*-
"""
fill_from_izvod.py  (S113, 2026-08-21)
======================================
Puni retke u app-ov Excel (delta sheet ili obični Activities export) **iz
bankovnog izvoda**, umjesto da ih čovjek prepisuje kolonu po kolonu iz Kokinog
filea. Kolone se pogađaju **po imenu zaglavlja**, ne po položaju — raspored
stupaca (profil, skrivene kolone) zato ne igra nikakvu ulogu.

ZAŠTO IZ IZVODA, A NE IZ KOKINOG FILEA
  Izvod je autoritet za **iznos i datum** (v. CLAUDE.md „Politika izvora"), a i
  jedini je izvor koji se da provjeriti: RF izvod uz svaki redak nosi ISPISANO
  tekuće stanje, pa alat na kraju javi na kojoj brojci kontrolni stupac mora
  završiti. Kokin file ostaje autoritet za opis i klasifikaciju.

ŠTO ALAT NE RADI
  • ne uvozi ništa — piše samo Excel, uvoz ide kroz app kao i dosad
  • ne dira postojeće retke (samo prazne retke predloška i, ako ih ponestane,
    dopisuje nove ispod)
  • ne izmišlja `session_start` ondje gdje ga app već zapisao — prazni retci
    predloška ga imaju, a dopisani ga dobivaju iz istog pojasa (14:00+n)

⚠ AUTOFILTER: dopisani redak IZVAN `auto_filter.ref` se pri prvom sortu raspari
  od svog zaglavlja (CLAUDE.md). Alat zato širi raspon filtra na nove retke.

⚠ BOOLEAN: `Rate?` mora biti PRAVI bool. Sve osim doslovnog `true` uvozi se kao
  FALSE — `'DA'` tiho postane netočno (`excelImport.ts:1244`).

Pokretanje (target = file skinut iz appa, ostaje netaknut):
    python fill_from_izvod.py <target.xlsx> --rf   <RF_2026-07.pdf>   [--od 2026-08-05]
    python fill_from_izvod.py <target.xlsx> --visa <PBZVIZA_2026-07.pdf> --naplata 2026-08-07
Rezultat: `<target>_filled.xlsx` pokraj originala + izvještaj na ekran.
"""

import argparse
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')   # sys.exit poruke idu ovuda
sys.path.insert(0, str(Path(__file__).parent))

AREA_COL, PATH_COL, DATE_COL, SESS_COL, USER_COL, COMMENT_COL = 2, 3, 4, 5, 7, 8
TIME_START_H = 14          # isti pojas koji piše deltaSheet.ts (DELTA_TIME_START_H)

# Klasifikacija Racun-redaka po tekstu izvoda. Namjerno kratko i vidljivo:
# krivo-ali-valjano klasificiran redak `apply_rules.py` više NE MOŽE popraviti
# (preskače retke s valjanim parom), pa se svaki pogodak ispisuje u izvještaju.
# Zadnji stupac je `leaf comment` — ime pod kojim redak VEĆ postoji u bazi.
# Sirovi tekst izvoda (`RAIFFEISENDOBROVOLJNIMIROVINSKIFONDPETRINJSKA...`) je
# točan ali nečitljiv, a lista se čita okom; original ostaje u `Izvod opis`.
RF_RULES = [
    (re.compile(r'PBZCARD|PBZ CARD', re.I),        'Transfer',    'izmedju racuna',    'Visa'),
    (re.compile(r'DOBROVOLJNIMIROVINSKI', re.I),   'Prihodi',     'Saša',              'Mirovina III stup'),
    (re.compile(r'ZAVODZAMIROVINSKO', re.I),       'Prihodi',     'Saša',              'Mirovina I stup'),
    (re.compile(r'MIROVINSKO ?OSIGURAVAJUCE', re.I),'Prihodi',    'Saša',              'Mirovina II stup'),
    (re.compile(r'NAKNADA', re.I),                 'Domaćinstvo', 'Bankovni troškovi', 'Naknada'),
]
# ZABA tekući račun. Isti princip kao RF_RULES: kratko, vidljivo, i sve što ne
# pogodi ide u N/A (krivo-ali-valjano klasificiran redak `apply_rules.py` više
# ne može popraviti, pa je N/A pošteniji od nagađanja).
# ⚠ `Podizanje gotovog novca` je `Transfer | cash - bankomat` — taj redak JE u
#   saldu (novac je otišao s računa), a gotovinski trošak koji slijedi NIJE
#   (v. CLAUDE.md: Cash je izbačen iz filtra salda).
ZABA_RULES = [
    (re.compile(r'MASTERCARD KARTICOM', re.I),        'Transfer',    'izmedju racuna',    'Mastercard'),
    # ⚠ SIDRO NA POCETAK REDAKA JE NUZNO, i to je izmjereno a ne pretpostavljeno.
    #   Bilo je `Naknada za vođenje`; kolovoz 2026. donio je i `Naknada za
    #   kreditni transfer` (3 retka po 0,35), pa je pravilo prosireno na
    #   `Naknada za `. Time je odmah pokupilo i `... (m-zaba) Naknada za
    #   uređenje voda - SPLIT ... NUV - 1. rata za 2026.` (7,43) — a to NIJE
    #   bankovna naknada nego vodnogospodarska davanja, dakle trosak kucanstva.
    #   Bankine vlastite naknade svoj redak POCINJU tim tekstom; tudje ga nose
    #   iza prefiksa naloga. Razlika je u polozaju, pa je i pravilo takvo.
    #   (CLAUDE.md „Pretraga po kljucnoj rijeci prekomjerno hvata".)
    (re.compile(r'^Naknada za ', re.I),               'Domaćinstvo', 'Bankovni troškovi', 'Naknada'),
    (re.compile(r'Podizanje gotovog novca', re.I),    'Transfer',    'cash - bankomat',   'Bankomat'),
    (re.compile(r'MIROVINSKOG PRIMANJA', re.I),       'Prihodi',     'Koka',              'Mirovina'),
    # Imenovani uplatitelji: tekst izvoda ih nosi doslovno, a povijest je
    # jednoglasna (Zoran 20/20, Anja 17/17). Jači ključ od iznosa — iznos im se
    # mijenja svaki mjesec, ime ne.
    (re.compile(r'UPLATA ZORAN SLADOLJEV', re.I),     'Kuća',        'Povrat Zoran',      'Zoran povrat'),
    (re.compile(r'UPLATA ANJA CRNKOVI', re.I),        'Prihodi',     'Povrat Anja',       'Anja'),
]

NA = 'N/A'                 # legitimna vrijednost, ne blokira uvoz (S107q)
# Granice oko znamenki: bez njih `rezije voda za 07/2026` daje „ratu 07/202".
_RATA_IZ_IZVODA = re.compile(r'(?<!\d)(\d{1,3})\s*/\s*(\d{1,3})(?!\d)')
DATE_TOL = 3               # dana tolerancije pri prepoznavanju istog retka


def _as_date(v) -> date:
    """Parseri vraćaju datum kao `date`, `datetime` ili ISO string — ovisno o
    izvodu. Jedan oblik dalje, inače dedup uspoređuje različite tipove."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return date.fromisoformat(str(v)[:10])


def _load_tolerant(path: Path):
    """openpyxl padne na nevaljanom `errorStyle="error"` koji je app-ov export
    pisao do S113 (OOXML pozna samo `stop|warning|information`; Excel ga
    progura, openpyxl ne). Fileovi skinuti prije popravka i dalje postoje, pa
    ih alat popravlja u memoriji umjesto da traži novi export."""
    try:
        return openpyxl.load_workbook(path)
    except Exception:
        import io as _io
        import zipfile
        src = zipfile.ZipFile(path)
        buf = _io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as out:
            for item in src.infolist():
                data = src.read(item.filename)
                if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                    data = data.decode('utf-8').replace(
                        'errorStyle="error"', 'errorStyle="stop"').encode('utf-8')
                out.writestr(item, data)
        buf.seek(0)
        print('  (popravljen nevaljan errorStyle iz starijeg exporta)')
        return openpyxl.load_workbook(buf)


class Target:
    """App-ov Events list: gdje je zaglavlje, koje kolone postoje, gdje su prazni retci."""

    def __init__(self, path: Path):
        self.path = path
        self.wb = _load_tolerant(path)
        self.ws = self.wb['Events'] if 'Events' in self.wb.sheetnames else self.wb.worksheets[0]

        self.header_row = 0
        for r in range(1, self.ws.max_row + 1):
            if str(self.ws.cell(r, 1).value or '').strip() == 'event_id':
                self.header_row = r
                break
        if not self.header_row:
            sys.exit('✗ Ne nalazim zaglavlje EVENT DATA (kolona A "event_id").')

        # `Uplata (Transakcija)` → ključ `uplata`; kolona se traži po imenu atributa
        self.col = {}
        self.last_col = 0
        for c in range(1, self.ws.max_column + 1):
            v = str(self.ws.cell(self.header_row, c).value or '').strip()
            if not v:
                continue
            self.last_col = c
            self.col[v.split(' (')[0].strip().lower()] = c

        # redak s podacima = ima Area; „pravi" = ima i datum. Ostatak su prazni
        # retci predloška (nose prepisan Area/email/vrijeme, čekaju unos).
        self.real_rows, self.blank_rows = [], []
        for r in range(self.header_row + 1, self.ws.max_row + 1):
            has_area = bool(str(self.ws.cell(r, AREA_COL).value or '').strip())
            has_time = bool(str(self.ws.cell(r, SESS_COL).value or '').strip())
            # ⚠ Prazan redak predloška poznaje se i po samom vremenu: export
            #   usklađenog računa (prazan prozor) zna ostaviti `Area` prazan, a
            #   redak je i dalje predložak. Da ga alat ne vidi, tiho bi dopisivao
            #   ispod umjesto da ga popuni.
            if not has_area and not has_time:
                continue
            (self.real_rows if self.ws.cell(r, DATE_COL).value else self.blank_rows).append(r)

    def find(self, name: str):
        return self.col.get(name.lower())

    def put(self, row: int, name: str, value):
        """Upiši samo ako list tu kolonu uopće ima — inače tiho preskoči i javi."""
        c = self.find(name)
        if c is None:
            return False
        self.ws.cell(row, c).value = value
        return True

    def accounts(self) -> set:
        """Vrijednosti `Racun` koje list već nosi — i na postojećim retcima i na
        prepisanima u praznima."""
        c = self.find('Racun')
        if c is None:
            return set()
        vals = set()
        for r in self.real_rows + self.blank_rows:
            v = str(self.ws.cell(r, c).value or '').strip()
            if v:
                vals.add(v)
        return vals

    def existing_keys(self) -> set:
        """(datum, iznos) postojećih redaka — ključ za dedup protiv izvoda."""
        keys = set()
        cu, ci = self.find('Uplata'), self.find('Isplata')
        for r in self.real_rows:
            d = self.ws.cell(r, DATE_COL).value
            d = d.date() if isinstance(d, datetime) else d
            for c in (cu, ci):
                if c is None:
                    continue
                v = self.ws.cell(r, c).value
                if isinstance(v, (int, float)) and v:
                    keys.add((str(d), round(float(v), 2)))
        return keys


# OCR jedne RF stranice traje ~25 s, pa se rezultat kešira po md5 filea: izvod se
# ne mijenja, a alat se pokreće više puta (dry → provjera → pravi upis).
CACHE_DIR = (Path(__file__).parents[2] / 'data-prep_data' / 'Financije'
             / '_arhiva' / 'ocr_cache')


def _rf_parse_cached(pdf: Path) -> list[dict]:
    import hashlib
    import json
    md5 = hashlib.md5(pdf.read_bytes()).hexdigest()
    cache = CACHE_DIR / f'{pdf.stem}_{md5}.json'
    if cache.exists():
        print(f'  (OCR iz keša: {cache.name})')
        return json.loads(cache.read_text(encoding='utf-8'))
    from rf_ocr import parse_rf_ocr
    txs = parse_rf_ocr(pdf)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache.write_text(json.dumps(txs, ensure_ascii=False, default=str), encoding='utf-8')
    return txs


def rf_rows(pdf: Path, od: date | None, do: date | None) -> list[dict]:
    """Racun retci s RF izvoda — nose i ISPISANO stanje (`stanje_izvod`)."""
    out = []
    for t in _rf_parse_cached(pdf):
        d = _as_date(t['date'])
        if (od and d < od) or (do and d > do):
            continue
        tip, podtip, kom = NA, NA, short_opis(t['opis'])
        for rx, ti, po, km in RF_RULES:
            if rx.search(t['opis']):
                tip, podtip, kom = ti, po, km
                break
        out.append({
            'date': d, 'smjer': t['smjer'], 'iznos': round(float(t['iznos']), 2),
            # `opis` je ono sto ide u `Izvod opis`; `opis_puni` ostaje sirov jer
            # se po njemu jos sparuje i klasificira unutar ovog prolaza.
            'opis': skrati_opis(t['opis']), 'opis_puni': t['opis'],
            'izvor': 'Racun', 'tip': tip, 'podtip': podtip,
            'komentar': kom,
            'naplata': d,                       # D1b: Izvor=Racun ⇒ naplata istog dana
            'stanje': t.get('stanje_izvod'),
        })
    return out


class KokaOpisi:
    """Kokini opisi iz njene Excelice, spareni s retcima izvoda po (iznos, datum).

    ⚠ PODJELA AUTORITETA (CLAUDE.md „Politika izvora"): izvod je autoritet za
    IZNOS i DATUM, njen redak za OPIS. Izvod piše
    `SUPER KONZUM P-3200 - RADNICKA CESTA 1 - ZAGREB`, ona piše `Konzum` — i to
    je tekst koji ona prepoznaje u popisu. Strojno kracenje opisa je pogadanje;
    njen tekst je podatak.

    Sparuje se s tolerancijom na datum, jer se dan knjizenja i dan kupovine
    razlikuju po nekoliko dana. Svaki njen redak se trosi najvise jednom, inace
    bi dvije iste kupovine istog dana obje pokupile isti opis.
    """

    def __init__(self, path: Path, prije: int = 3, poslije: int = 45):
        # Prozor je NESIMETRICAN: ona kupovinu upisuje na dan kupnje ILI na dan
        # kad je karticni racun naplacen (vidjeno oboje u istom fileu), a nikad
        # prije kupovine. Simetricna tolerancija od par dana zato ne uhvati nista
        # za karticne retke.
        self.prije, self.poslije = prije, poslije
        self.by_amount: dict[float, list] = {}
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < 6:
                    continue
                _, opis, d, upl, isp, _ = row[:6]
                # ⚠ Njene DVIJE kolone datuma. C (`Datum`) je dan kad novac
                #   napusti račun; kad naplata još nije poznata — kartična
                #   kupovina koja čeka mjesečni račun — C je PRAZAN, a dan
                #   troška stoji u koloni G. Bez te rezerve alat ne vidi
                #   nijedan redak koji još čeka naplatu, a to su upravo oni
                #   najsvježiji.
                if not isinstance(d, datetime) and len(row) > 6 and isinstance(row[6], datetime):
                    d = row[6]
                if not isinstance(d, datetime) or not opis:
                    continue
                iznos = upl if isinstance(upl, (int, float)) and upl else isp
                if not isinstance(iznos, (int, float)) or not iznos:
                    continue
                self.by_amount.setdefault(round(float(iznos), 2), []).append(
                    {'date': d.date(), 'opis': str(opis).strip(), 'used': False})
        self.hits = 0
        self.misses = 0

    def find(self, d: date, iznos: float,
             prije: int | None = None, poslije: int | None = None) -> str | None:
        """⚠ Prozor se smije stegnuti po izvoru. Zadani −3/+45 dana je za
        KARTIČNE retke, gdje dan kupovine i dan naplate stvarno stoje mjesec i
        pol razmaknuti. Na tekućem računu takva tolerancija nije velikodušna
        nego opasna: `Cash 100,00` se ponavlja svakih par tjedana, pa bi prvi
        bankomat pokupio opis nekog kasnijeg — i to tiho, jer se iznos i dalje
        slaže. Ondje je njen datum bankin datum, pa prozor ide na 0/+1."""
        prije = self.prije if prije is None else prije
        poslije = self.poslije if poslije is None else poslije
        cands = [c for c in self.by_amount.get(round(iznos, 2), []) if not c['used']]
        if not cands:
            self.misses += 1
            return None
        best = min(cands, key=lambda c: abs((c['date'] - d).days))
        delta = (best['date'] - d).days
        if delta < -prije or delta > poslije:
            self.misses += 1
            return None
        best['used'] = True
        self.hits += 1
        return best['opis']


def short_opis(raw: str) -> str:
    """Kratko ime trgovca iz opisa izvoda — `leaf comment` čita čovjek u popisu.

    Izvod piše `SUPER KONZUM P-3200 - RADNIČKA CESTA 1 - ZAGREB`; Koka piše
    `Konzum`. Popis od 45 redaka s punim adresama je nečitljiv, a informacija
    nije izgubljena: sirovi tekst ide u `Izvod opis` i ondje ostaje za
    pretraživanje i pravila.

    ⚠ Skraćuje se samo prikaz. Ništa se ne odbacuje i ništa se ne pogađa —
    ako od opisa ne ostane ništa upotrebljivo, vraća se izvorni tekst.
    """
    t = re.sub(r'\[kartica:[^\]]*\]', '', raw)           # trag kartice ide u Izvod opis
    t = re.sub(r'^RATA\s*\d+\s*/\s*\d+\s*-\s*', '', t.strip())
    t = re.split(r'\s+-\s*', t)[0]                        # sve iza prve crtice je adresa
    t = re.sub(r'(?<!\w)[A-ZČĆŠĐŽ]{1,3}[- ]?\d{2,5}(?!\w)', ' ', t)   # šifre poslovnica (P-3200, K018)
    t = re.sub(r'\s+(ZAGREB|ZG)(?!\w)', ' ', t)               # grad se ponavlja na svakom retku
    t = re.sub(r'[-\s]{2,}', ' ', t)                       # ostatak nakon izbacene sifre
    t = re.sub(r'[\s,.-]+$', '', t).strip()
    # Velika slova se NE diraju: „DM-DROGERIE MARKT“ je čitljivo, a „Dm-drogerie
    # Markt“ je pogrešno na način koji izgleda namjerno. Kraćenje je jedini posao
    # ove funkcije; ako od opisa ne ostane ništa, vraća se izvorni tekst.
    return t[:40] if t else raw[:40]


# Uvod koji ZABA lijepi na SVAKI nalog. 66 znakova bez ijedne informacije —
# a `Izvod opis` je stupac koji covjek cita u tablici i u kojem trazi trgovca.
# ⚠ Sidro na POCETAK je bitno: redak bankine vlastite naknade glasi
#   `Naknada za kreditni transfer nacionalni u eurima on-line bankarstvom
#   (m-zaba) M160...` — ondje ta rijec nije uvod nego SADRZAJ, pa se ne dira.
# ⚠ `(m-zaba)` OSTAJE: to je kanal naloga, jedina informacija u uvodu.
# ⚠ Skracivanje je sigurno za sparivanje: `presedani.kljuc_izvoda` isti uvod
#   ionako skida prije nego napravi kljuc, pa se stari (dugi) i novi (kratki)
#   zapisi i dalje poklapaju. Bez toga bi svaki uvoz od danas prestao nalaziti
#   presedane u vlastitoj povijesti.
_UVOD = re.compile(r'^Kreditni transfer nacionalni u eurima on-?line bankarstvom\s*', re.I)


def skrati_opis(s: str) -> str:
    """`Kreditni transfer ... bankarstvom (m-zaba) Bmove d.o.o. ...`
    -> `(m-zaba) Bmove d.o.o. ...`"""
    return _UVOD.sub('', str(s or '')).strip()


def zaba_rows(pdf: Path, od: date | None, do: date | None,
              koka: 'KokaOpisi | None' = None) -> list[dict]:
    """Retci tekućeg računa sa ZABA izvatka. Parser sam validira smjer i
    potpunost protiv ISPISANIH salda (`_validate_zaba`) — ako format izvatka
    ikad promijeni, javi se na stderr umjesto da tiho da manje redaka."""
    from enrich_from_izvoda import parse_zaba_racun
    out = []
    for t in parse_zaba_racun(pdf):
        d = _as_date(t['date'])
        if (od and d < od) or (do and d > do):
            continue
        tip, podtip, kom = NA, NA, short_opis(t['opis'])
        for rx, ti, po, km in ZABA_RULES:
            if rx.search(t['opis']):
                tip, podtip, kom = ti, po, km
                # Rata je jedino sto se u fiksnu oznaku smije dopisati, i to
                # SAMO ako je izvod stvarno nosi: `ANJA CRNKOVIC 85/96` -> `Anja
                # 85/96`. Prepisan broj rate iz presedana bio bi pogodak
                # (proslomjesecni), pa dolazi iskljucivo iz teksta izvoda.
                m = _RATA_IZ_IZVODA.search(t['opis'])
                if m and not _RATA_IZ_IZVODA.search(kom):
                    kom = '%s %s/%s' % (kom, m.group(1), m.group(2))
                break
        # ⚠ Tekući račun je gori od kartice: banka svaki nalog opisuje istim
        #   tekstom (`Kreditni transfer nacionalni u eurima on-line bankarstvo`),
        #   pa strojno kraćenje ne razlikuje T-com od T-mobilea. Njen redak zna
        #   koji je koji, i to je jedini izvor koji to zna — Tip/Podtip ostaju
        #   s pravila, mijenja se samo tekst koji čovjek čita u popisu.
        kom_koka = koka.find(d, round(float(t['iznos']), 2),
                             prije=0, poslije=1) if koka else None
        out.append({
            'date': d, 'smjer': t['smjer'], 'iznos': round(float(t['iznos']), 2),
            # `opis` je ono sto ide u `Izvod opis`; `opis_puni` ostaje sirov jer
            # se po njemu jos sparuje i klasificira unutar ovog prolaza.
            'opis': skrati_opis(t['opis']), 'opis_puni': t['opis'],
            'izvor': 'Racun', 'tip': tip, 'podtip': podtip,
            'komentar': kom_koka or kom,
            'opis_izvor': 'koka' if kom_koka else 'izvod',
            'naplata': d, 'stanje': None,
        })
    return out


def zaba_printed_balance(pdf: Path) -> tuple[float | None, date | None]:
    """Ispisano `NOVO STANJE` i datum zadnje transakcije — kontrolni cilj koji
    NIJE naš izračun. ⚠ Izvod se ne zatvara krajem mjeseca (S110), pa datum
    dolazi iz zadnjeg retka, ne iz imena datoteke."""
    from enrich_from_izvoda import _parse_zaba_all, _zaba_is_tekuci
    txs, balances = _parse_zaba_all(pdf)
    novo = next((b['novo'] for b in balances if _zaba_is_tekuci(b['account'])), None)
    tekuci = [t for t in txs if _zaba_is_tekuci(t['account'])]
    last = max((_as_date(t['date']) for t in tekuci), default=None)
    return novo, last


def visa_rows(pdf: Path, naplata: date, od: date | None, do: date | None,
              koka: 'KokaOpisi | None' = None) -> list[dict]:
    """Kartične stavke s PBZ Visa računa. Saldo NE diraju — plaća ih jedna
    skupna naplata, koja stiže s RF izvoda kao običan Racun redak."""
    from enrich_from_izvoda import parse_pbz_visa
    out = []
    for t in parse_pbz_visa(pdf):
        d = _as_date(t['date'])
        if (od and d < od) or (do and d > do):
            continue
        if t['smjer'] != 'Isplata':
            continue                            # 'PRIMLJENA UPLATA' je protustavka naplate
        m = re.search(r'RATA\s*(\d+)\s*/\s*(\d+)', t['opis'])
        kom_koka = koka.find(d, round(float(t['iznos']), 2)) if koka else None
        kom_visa = kom_koka or short_opis(t['opis'])
        out.append({
            'date': d, 'smjer': 'Isplata', 'iznos': round(float(t['iznos']), 2),
            'opis': t['opis'], 'izvor': 'Visa', 'tip': NA, 'podtip': NA,
            'komentar': kom_visa,
            'opis_izvor': 'koka' if kom_koka else 'izvod',
            'naplata': naplata,                 # dan kad je banka skinula skupnu naplatu
            'rata': (int(m.group(1)), int(m.group(2))) if m else None,
            'stanje': None,
        })
    return out


# -----------------------------------------------------------------------------
# IZVOR: Kokina Excelica (S116)
# -----------------------------------------------------------------------------
# WARN  ZASTO OVAJ IZVOR UOPCE POSTOJI, KAD JE IZVOD AUTORITET
#   Za kolovoz 2026. izvoda NEMA i nece ga biti do rujna. Politika izvora
#   („izvodi rjesavaju staro, Koka novo") to predvida, a D-2 je odluka:
#   njeni retci ulaze sada, izvod ih poslije provjerava. Ovdje je zato njen
#   file autoritet i za IZNOS i za DATUM - jedini put kad to vrijedi.
#
# WARN  PROVJERA MORA BITI MEHANICKA (D-2). Njeni se iznosi razlikuju od bankinih
#   na ~4 % redaka, a karticne stavke ne diraju saldo => takva greska NIKAD ne
#   ispliva sama. Alat zato ispise lanac koji njeni retci daju, da se poklapanje
#   s kontrolnim brojem vidi PRIJE uvoza, a ne mjesecima poslije.
#
# WARN  NJEN MODEL NIJE NAS MODEL. Ona svaku karticnu stavku tereti tekucem racunu
#   pojedinacno; banka skine JEDNU skupnu naplatu. Zbroj se poklapa u cent
#   (45 MC stavki 11.08. = 1.332,52 = iznos s `MC_2026-07.pdf`), ali model se ne
#   poklapa: kod nas su karticne stavke potovi (`Izvor = Mastercard`), a saldo
#   mice samo skupna naplata (`Izvor = Racun`). Zato `Izvor` odreduje KOLONA A
#   njenog sheeta, a skupna naplata se uzima s karticnog izvoda i NIKAD ne
#   sintetizira (CLAUDE.md).

KOKA_IZVOR = {
    'kokin tekuci': 'Racun',
    'sasin tekuci': 'Racun',
    'mastercard':   'Mastercard',
    'visa':         'Visa',
}


def _koka_norm(s: str) -> str:
    """`Kokin tekuci` <- `Kokin tekuci` s dijakriticima.

    WARN Usporedba imena racuna MORA ici preko ovoga. Njena kolona A pise
    `Kokin tekuci`/`Sasin tekuci` s kvacicama, a argument s komandne linije ih
    kroz `run.bat` zna izgubiti; obicna `==` usporedba bi tada nasla NULA redaka
    i alat bi javio „0 novih" - sto se cita kao „nema sto uvesti", a ne kao
    „nije ni usporedeno" (isti razred kao S114 brojac)."""
    d = unicodedata.normalize('NFKD', s.strip().lower())
    d = ''.join(ch for ch in d if not unicodedata.combining(ch))
    return d.replace('đ', 'd')


# `Konzum 5/12`, `Anja 73/96`, `LH 2/3` - rata N od M, na kraju opisa.
RATA_RE = re.compile(r'\s(\d{1,3})/(\d{1,3})\s*$')


def _koka_klasa(opis: str) -> tuple[str, str]:
    """Tip/Podtip iz IZBROJANE povijesti Reviewa, nikad iz teksta izvoda (S114).

    `PO_OPISU` je jedini izvor mapiranja i dijeli se s `klasificiraj_transu.py`
    - dva popisa koji se raziduju bila bi gora od nijednog."""
    from klasificiraj_transu import PO_OPISU
    low = opis.strip().lower()
    for prefiks, tip, podtip, _dokaz in PO_OPISU:
        if low == prefiks or low.startswith(prefiks):
            return tip, podtip
    return NA, NA


def koka_rows(path: Path, sheet: str, tip_racuna: str,
              od: date | None, do: date | None,
              klasificiraj: bool, osim: set[int] | None = None) -> list[dict]:
    """Retci jednog racuna iz Kokine Excelice, u obliku koji `write_rows` pise."""
    izvor = KOKA_IZVOR.get(_koka_norm(tip_racuna))
    if izvor is None:
        sys.exit('X Nepoznat tip racuna %r. Poznati: %s'
                 % (tip_racuna, ', '.join(sorted(KOKA_IZVOR))))

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        sys.exit('X %s nema list %r. Ima: %s' % (path.name, sheet, wb.sheetnames))
    ws = wb[sheet]

    out: list[dict] = []
    buducnost: list[tuple[int, date, float, str]] = []
    preskoceni: list[int] = []
    bez_datuma = 0
    granica = date.today().year + 1
    want = _koka_norm(tip_racuna)

    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 6 or not row[0]:
            continue
        if _koka_norm(str(row[0])) != want:
            continue
        # WARN Rucno iskljucen redak. Postoji jer se tipfeler u datumu ne smije
        #   „popraviti i uvesti": `Parking 1,60` datiran 07.08. vec je u bazi kao
        #   07.07. (i njen vlastiti stupac `Stanje` ga racuna medu srpanjskima).
        #   Uvoz bi ga udvostrucio, a razlika od 1,60 se ne primijeti. Odluka je
        #   covjekova, ali BROJ RETKA ostaje zapisan u naredbi - to je trag.
        if osim and r in osim:
            preskoceni.append(r)
            continue

        opis = str(row[1] or '').strip()
        c_datum = row[2] if isinstance(row[2], datetime) else None      # dan naplate
        g_datum = row[6] if len(row) > 6 and isinstance(row[6], datetime) else None
        upl, isp = row[3], row[4]

        iznos = upl if isinstance(upl, (int, float)) and upl else isp
        if not isinstance(iznos, (int, float)) or not iznos:
            continue
        smjer = 'Uplata' if (isinstance(upl, (int, float)) and upl) else 'Isplata'

        # D1b: event_date = DAN KUPOVINE uvijek. Na tekucem racunu to je isti dan
        # kad novac ode (kolona C); na kartici je kolona G, a C je dan naplate.
        kupovina = (g_datum or c_datum)
        if kupovina is None:
            bez_datuma += 1
            continue
        kupovina = kupovina.date()
        naplata = c_datum.date() if c_datum else None

        # WARN Tipfeler u godini nije poziv da ga popravis i uvezes (S115). Dva
        #   njena retka datirana 2036-04-08 vec postoje u bazi kao 2026-04-08;
        #   ispravak godine + uvoz udvostrucio bi ih, i to tiho - padaju prije
        #   sidra pa ne micu nijednu kontrolnu brojku. Alat ih IZDVAJA, ne popravlja.
        if kupovina.year > granica:
            buducnost.append((r, kupovina, round(float(iznos), 2), opis))
            continue

        if od and kupovina < od:
            continue
        if do and kupovina > do:
            continue

        tip, podtip = _koka_klasa(opis) if klasificiraj else (NA, NA)

        rec = {
            'date':     kupovina,
            'komentar': opis,
            'opis':     '',              # `Izvod opis` - nema ga, redak nije s izvoda
            'izvor':    izvor,
            'smjer':    smjer,
            'iznos':    round(float(iznos), 2),
            'tip':      tip,
            'podtip':   podtip,
            'naplata':  naplata,
            # Naplata poznata => novac je otisao. Prazna kolona C je upravo ono
            # sto u nasem modelu znaci `Planiran` (S113).
            'status':   'Izvrsen' if naplata else 'Planiran',
            'red':      r,
        }
        m = RATA_RE.search(opis)
        if m and int(m.group(1)) <= int(m.group(2)):
            rec['rata'] = (int(m.group(1)), int(m.group(2)))
        out.append(rec)

    if buducnost:
        print('\nWARN %d redaka s godinom > %d - NISU uzeti:' % (len(buducnost), granica))
        for r, d, iz, op in buducnost:
            print('   red %d: %s  %9.2f  %s' % (r, d, iz, op))
        print('   Prije ispravka godine provjeri postoji li redak vec pod ispravnim')
        print('   datumom - ispravak + uvoz bi ga udvostrucio tiho (S115).')
    if bez_datuma:
        print('  (%d redaka bez ijednog datuma preskoceno - ni C ni G)' % bez_datuma)
    if preskoceni:
        print('  (--osim: rucno iskljuceni redci %s)'
              % ', '.join(str(x) for x in preskoceni))

    return out


def koka_lanac(path: Path, sheet: str, od: date, pocetno: float,
               do: date | None = None, osim: set[int] | None = None) -> None:
    """Ispisi lanac koji NJENI retci daju od zadanog stanja - mehanicka provjera D-2.

    WARN Zbraja se SVE sto na tom listu stoji poslije `od`, ukljucujuci karticne
    stavke koje kod nas ne micu saldo. Njen lanac je drugi model, i bas zato je
    svjedok: ako se zavrsna brojka poklopi s kontrolnim brojem, oba modela kazu
    isto iako broje razlicito."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    saldo = pocetno
    n = 0
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 6 or not row[0]:
            continue
        if osim and r in osim:
            continue
        # WARN Lanac SALDA gleda samo kolonu C (dan kad novac napusti racun).
        #   Kolona G je dan troska i za jos nenaplacene karticne stavke je jedini
        #   datum koji redak ima - ali te stavke racun jos nisu teretile. Uzeti i
        #   njih znaci brojati buduce naplate kao da su se dogodile, i lanac
        #   promasi kontrolni broj za tocno njihov zbroj.
        c = row[2] if isinstance(row[2], datetime) else None
        eff = c
        if eff is None or eff.date() <= od or eff.date().year > date.today().year + 1:
            continue
        # WARN Bez gornje granice lanac pokupi i njene BUDUCE retke: karticne
        #   stavke koje jos cekaju naplatu nose datum kraja mjeseca u koloni G,
        #   pa lanac zavrsi na brojci koja se ne poklapa ni s cim - a izgleda
        #   kao da se model raziso.
        if do and eff.date() > do:
            continue
        upl, isp = row[3], row[4]
        if isinstance(upl, (int, float)) and upl:
            saldo += float(upl)
        elif isinstance(isp, (int, float)) and isp:
            saldo -= float(isp)
        else:
            continue
        n += 1
    print('> Kokin lanac: %.2f @ %s + %d redaka%s (list %s) = %.2f'
          % (pocetno, od, n, (' do %s' % do) if do else '', sheet, saldo))
    print('  Usporedi s kontrolnim brojem transe PRIJE uvoza - poslije se razlika')
    print('  na karticnim retcima ne vidi ni na plocici ni u kontrolnom stupcu.')


def strip_comments(wb) -> list[str]:
    """Makni Excel bilješke iz radne kopije i vrati njihov tekst.

    ⚠ openpyxl sprema bilješku kao `xl/comments/comment1.xml` s APSOLUTNOM
    putanjom u relacijama; app-ov citac (exceljs) ocekuje relativnu, ne nade
    dio i pukne s `Cannot read properties of undefined (reading 'comments')`.
    File se time ne da uvesti — a bilješka je informacija za covjeka, ne podatak.

    Zato se izbacuje iz radne kopije, a tekst se ISPISE: original izvoza je i
    dalje nosi, i podrijetlo otvarajuceg stanja ne smije nestati bez traga.
    """
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    out.append(cell.comment.text)
                    cell.comment = None
    return out


def objasni_spojene(tg: 'Target', kandidati: list[dict], svi_izvod: list[dict]) -> list[dict]:
    """Kokin JEDAN redak = N redaka izvoda (razred S114/S124).

    ⚠ Dedup po (datum, iznos) ovo NE MOZE vidjeti: kljuc se razlikuje, pa oba
      retka izvoda prodju kao „novo" i iznos ude DRUGI PUT. Izmjereno na
      `ZABA_2026-08.pdf`: njen `Parking 3,20 @ 02.08.` su bankina dva naloga po
      1,60, a `Parking 2,40 @ 04.08.` su 0,80 + 1,60 — cetiri retka, 5,60 €,
      koja bi usla uz vec postojeca dva.

    Alat NE SPAJA i ne brise nista — samo prepozna da je iznos vec pokriven i
    izostavi te retke iz upisa, pa to ISPISE. Odluka o tome zadrzava li se njen
    spojeni redak ili se zamjenjuje bankinima je covjekova (v. CLAUDE.md
    „Bankini redci su KOSTUR").
    """
    from itertools import combinations
    kljucevi_izvoda = {(r['date'].isoformat(), r['iznos']) for r in svi_izvod}
    cu, ci = tg.find('Uplata'), tg.find('Isplata')

    neobjasnjeni = []
    for red in tg.real_rows:
        d = tg.ws.cell(red, DATE_COL).value
        d = d.date() if isinstance(d, datetime) else d
        for c, smjer in ((cu, 'Uplata'), (ci, 'Isplata')):
            if c is None:
                continue
            v = tg.ws.cell(red, c).value
            if not isinstance(v, (int, float)) or not v:
                continue
            kljuc = (str(d)[:10], round(float(v), 2))
            if kljuc not in kljucevi_izvoda:
                neobjasnjeni.append({'red': red, 'datum': kljuc[0], 'iznos': kljuc[1],
                                     'smjer': smjer,
                                     'opis': tg.ws.cell(red, COMMENT_COL).value})

    preostali, spojevi = list(kandidati), []
    for nb in neobjasnjeni:
        isti_dan = [r for r in preostali
                    if r['date'].isoformat() == nb['datum'] and r['smjer'] == nb['smjer']]
        spoj = None
        for k in (2, 3):
            for combo in combinations(isti_dan, k):
                if abs(sum(x['iznos'] for x in combo) - nb['iznos']) < 0.005:
                    spoj = combo
                    break
            if spoj:
                break
        if not spoj:
            continue
        for x in spoj:
            preostali.remove(x)
        spojevi.append((nb['red'], list(spoj)))
        print('   ⊃ %s %9.2f  "%s" (r%d) je vec zbroj %s — tih %d redaka izvoda '
              'NE upisujem'
              % (nb['datum'], nb['iznos'], str(nb['opis'])[:26], nb['red'],
                 ' + '.join('%.2f' % x['iznos'] for x in spoj), len(spoj)))
        print('       (jedan njen redak = N bankinih; zamjena njenog bankinima je '
              'zaseban potez, ne uvoz)')
    return preostali, spojevi


def zigosi_postojece(tg: 'Target', svi_izvod: list[dict], spojevi: list,
                     mail_vlasnika: str = '', izvor: str = 'Racun') -> tuple[int, list]:
    """Upisi `Izvod opis` na retke koje baza VEC IMA, a izvod ih potvrdjuje.

    ⚠ `Izvod opis` je de facto oznaka POTVRDJENO IZVODOM (S124) — jedini nacin
      da se poslije razlikuju tri stanja: prazan = Kokina nepotvrdjena tvrdnja ·
      popunjen = banka potvrdila · prazan a razdoblje pokriveno izvodom =
      PITANJE. Bez zigosanja trece stanje ne postoji, pa svaki iduci izvod
      ponovno pita ono na sto je ovaj vec odgovorio.

    ⚠ Zigose se SAMO tocan par (datum + iznos + smjer). Tolerancija na datum bi
      ovdje bila opasna na nacin koji se ne vidi: `Cash 100,00` se ponavlja
      svakih par tjedana (S114), pa bi prvi bankomat pokupio potvrdu nekog
      kasnijeg — iznos se i dalje slaze, i nista ne bi javilo gresku.

    ⚠ Prazna celija se popunjava, popunjena se NE DIRA. Postojeca potvrda je
      dokaz nekog drugog izvoda; prepisati je znacilo bi premjestiti redak u
      pogresan izvod (S124: „potvrdjen redak pripada tocno jednom izvodu").

    ⚠ Spojeni redak (jedan njen = N bankinih) dobiva OBA teksta spojena s ` + `.
      Uzeti „opis od prvog" bilo bi tise i krace, ali bi tvrdilo da redak
      odgovara jednom nalogu — a on odgovara dvama. Polu-istina u stupcu koji
      sluzi kao dokaz gora je od duljeg teksta.
    """
    c_izvod = tg.find('Izvod opis')
    if c_izvod is None:
        print('  (list nema kolonu `Izvod opis` — zigosanje preskoceno)')
        return 0, []
    cu, ci, c_izvor = tg.find('Uplata'), tg.find('Isplata'), tg.find('Izvor')

    slobodni = [dict(r, uzet=False) for r in svi_izvod]
    spoj_po_retku = {}
    for red, combo in spojevi:
        spoj_po_retku[red] = combo
        for x in combo:
            for k in slobodni:
                if (not k['uzet'] and k['date'] == x['date']
                        and abs(k['iznos'] - x['iznos']) < 0.005):
                    k['uzet'] = True
                    break

    n, tudji, bez_para = 0, [], []
    for red in tg.real_rows:
        if str(tg.ws.cell(red, c_izvod).value or '').strip():
            continue
        # ⚠ Delta sheet od S123 nosi ISPOD praznih redaka i sekciju kosare —
        #   karticne retke koji saldo ne micu. Njih ZABA izvadak ne moze
        #   potvrditi ni u principu, pa bi ih prijavio kao „nema para" i
        #   proizveo pitanje ondje gdje pitanja nema.
        if c_izvor and str(tg.ws.cell(red, c_izvor).value or '').strip() != izvor:
            continue
        d = tg.ws.cell(red, DATE_COL).value
        d = d.date() if isinstance(d, datetime) else d
        u = tg.ws.cell(red, cu).value if cu else None
        i = tg.ws.cell(red, ci).value if ci else None
        iznos = float(u or i or 0)
        smjer = 'Uplata' if (u or 0) else 'Isplata'
        if not iznos:
            continue

        if red in spoj_po_retku:
            tekst = ' + '.join(x['opis'] for x in spoj_po_retku[red])
        else:
            par = next((k for k in slobodni if not k['uzet']
                        and k['date'] == d and k['smjer'] == smjer
                        and abs(k['iznos'] - iznos) < 0.005), None)
            if not par:
                bez_para.append((red, d, iznos))
                continue
            par['uzet'] = True
            tekst = par['opis']
        tg.ws.cell(red, c_izvod).value = tekst
        n += 1
        mail = str(tg.ws.cell(red, USER_COL).value or '').strip()
        if mail_vlasnika and mail and mail != mail_vlasnika:
            tudji.append((red, mail))

    print('Zigosanje: %d postojecih redaka dobilo `Izvod opis` '
          '(potvrda ovim izvodom)' % n)
    for red, d, iznos in bez_para:
        print('   ? r%-3d %s %9.2f — izvod nema tocan par, celija ostaje prazna'
              % (red, d, iznos))
    if bez_para:
        print('       (prazna celija u razdoblju koje izvod POKRIVA je pitanje, '
              'ne greska: ili duplikat, ili banka za taj trosak ne zna)')
    return n, tudji


def prosiri_kosaru(tg: 'Target', novi_kraj: int) -> bool:
    """Kontrola kosare (`Σ kosara`) ima FIKSAN raspon iz trenutka izvoza. Dopise
    li alat retke ispod sekcije, oni u zbroj ne udu — a kontrola tada pokazuje
    uvjerljiv broj koji ne pokriva sve retke, sto je gore od nikakvog.

    ⚠ Od S126 kontrola stoji IZNAD sekcije, pa se sekcija smije siriti prema
      dolje bez pomicanja icega. Ovdje se zato mijenja samo kraj raspona.
    """
    lbl = tg.last_col - 1                       # kolona oznaka (`Delete?`)
    for r in range(tg.header_row, tg.ws.max_row + 1):
        if not str(tg.ws.cell(r, lbl).value or '').strip().startswith('Σ'):
            continue
        c = tg.ws.cell(r, tg.last_col)
        f = str(c.value or '')
        if not f.startswith('='):
            continue
        novi = re.sub(r'(\$[A-Z]+\$\d+:\$[A-Z]+\$)\d+',
                      lambda m: m.group(1) + str(novi_kraj), f)
        if novi != f:
            c.value = novi
            print('  Σ kosare prosirena do retka %d' % novi_kraj)
            return True
    print('  ⚠ List nema kontrolu kosare (`Σ kosara`) — dopisani karticni retci '
          'nisu ni u jednom zbroju. Izvezi delta sheet iz Aree koja ima '
          '`split.due_slug`, inace kosaru nema s cim usporediti.')
    return False


def write_rows(tg: Target, rows: list[dict], racun: str, dry: bool,
               ref: 'Target | None' = None) -> tuple[int, int]:
    """Popuni prazne retke predloška; kad ih ponestane, dopiši nove ispod."""
    template = tg.blank_rows[0] if tg.blank_rows else (tg.real_rows[-1] if tg.real_rows else None)
    area = tg.ws.cell(template, AREA_COL).value if template else None
    path = tg.ws.cell(template, PATH_COL).value if template else None
    mail = tg.ws.cell(template, USER_COL).value if template else None

    # `Area` i `Category_Path` su jedina dva polja bez kojih uvoz redak ne vidi
    # kao redak. Ako ih predložak nema (v. prazan prozor), posuđuju se iz
    # referentnog exporta — ondje su pravi, jer dolaze iz istih zapisa.
    if (not area or not path) and ref and ref.real_rows:
        src = ref.real_rows[0]
        area = area or ref.ws.cell(src, AREA_COL).value
        path = path or ref.ws.cell(src, PATH_COL).value
        mail = mail or ref.ws.cell(src, USER_COL).value
        print(f'  (Area/Category_Path preuzeti iz reference: {area} > {path})')
    if not area or not path:
        print('✗ Predložak nema Area/Category_Path, a nema ni reference iz koje bi ih uzeo.')
        sys.exit('  Izvezi delta sheet ponovno ili dodaj --protiv <app export>.')

    # Format datuma se preuzima s postojećeg retka — app ga piše, alat ga ne izmišlja.
    date_fmt = 'yyyy-mm-dd'
    if tg.real_rows:
        date_fmt = tg.ws.cell(tg.real_rows[0], DATE_COL).number_format or date_fmt

    used_blanks, appended = 0, 0
    # WARN Od S123 delta sheet ima ISPOD praznih redaka jos i sekciju kosare s
    #   vlastitim kontrolnim retcima (Suma planirano / naplaceno s izvoda /
    #   razlika). Ti retci nemaju `Area`, pa ih `Target` ne vidi ni kao prave ni
    #   kao prazne — brojanje „max(real+blank)+1" bi zato dopisivalo TOCNO preko
    #   njih. Dopisuje se ispod SVEGA na listu.
    next_free = max(list(tg.real_rows) + list(tg.blank_rows)
                    + [tg.ws.max_row, tg.header_row]) + 1

    # ⚠ Vrijeme dopisanog retka mora biti slobodno na CIJELOM listu, ne samo
    #   „iza broja praznih". Prazni retci koje je popunila prethodna tranša su
    #   potrošeni, ali njihova vremena i dalje stoje na listu — brojanje od
    #   `len(blank_rows)` bi zato palo točno na njih. A kolizija nije sitnica:
    #   ona je zaštita od dvostrukog uvoza istog filea, pa lažna kolizija
    #   izgleda kao pad featurea („All skipped").
    used_times = set()
    for r in tg.real_rows + tg.blank_rows:
        v = tg.ws.cell(r, SESS_COL).value
        if v:
            used_times.add(str(v)[:5])

    def free_time() -> str:
        for m in range(TIME_START_H * 60, 24 * 60):
            t = f'{m // 60:02d}:{m % 60:02d}'
            if t not in used_times:
                used_times.add(t)
                return t
        sys.exit('✗ Nema slobodnog vremena u danu — podijeli uvoz na dva filea.')

    # ⚠ Kontrolni stupac (`Stanje (kontrola)`) postoji SAMO na pripremljenim
    #   praznim retcima — dopisani su izvan njegovog raspona. Retci koji miču
    #   saldo (`Izvor = Racun`) zato idu u prazne PRVI; kartične stavke smiju
    #   ispasti ispod jer ih kontrolni stupac ionako ne broji. Obrnuti redoslijed
    #   dao bi kontrolni broj koji izgleda uvjerljivo, a ne uključuje sve.
    rows = sorted(rows, key=lambda r: (r['izvor'] != 'Racun', r['date'], r['iznos']))

    # ⚠ KARTICNI REDAK NE SMIJE U PRAZNE RETKE GLAVNOG BLOKA (S126).
    #   Prije je smio („kontrolni stupac ih ionako ne broji"), i to je bilo
    #   tocno dok sekcija nije imala VLASTITU kontrolu. Sada ima: `Σ kosara`
    #   pokriva samo retke sekcije, pa bi kartcni redak u praznom retku ispao
    #   iz OBA zbroja — iz kontrolnog stupca jer nije `Racun`, i iz kosare jer
    #   nije u njezinu rasponu. Izmjereno: uz --zaba (30) i --mc (45) na 40
    #   praznih redaka, 10 MC stavki palo bi upravo tako, i kontrola kosare bi
    #   pokazala uvjerljiv broj manji za njihov zbroj.
    saldo_rows = [r for r in rows if r['izvor'] == 'Racun']
    if len(saldo_rows) < len(rows) and len(saldo_rows) < len(tg.blank_rows):
        print('  (kartcni retci se dopisuju ISPOD sekcije, ne u prazne retke — '
              'inace ne bi usli u `Σ kosara`)')

    for i, r in enumerate(rows):
        if r['izvor'] == 'Racun' and i < len(tg.blank_rows):
            row = tg.blank_rows[i]
            used_blanks += 1
        else:
            row = next_free
            next_free += 1
            appended += 1
            tg.ws.cell(row, SESS_COL).value = free_time()
            tg.ws.cell(row, SESS_COL).number_format = '@'

        # ⚠ Area / Category_Path / User se popunjavaju UVIJEK kad fale — ne samo
        #   na dopisanim retcima. Prazan prozor daje prazne retke predloška BEZ
        #   Area, a redak bez Area parser preskoči **bez ijedne poruke**: uvoz
        #   javi „0 New, 0 Modify" i file izgleda prazan iako je pun.
        for col, val in ((AREA_COL, area), (PATH_COL, path), (USER_COL, mail)):
            if not str(tg.ws.cell(row, col).value or '').strip():
                tg.ws.cell(row, col).value = val

        # ⚠ event_date MORA biti prava datumska ćelija, ne ISO tekst. Kontrolni
        #   stupac uspoređuje `"<="&$D<red>` protiv datumskih ćelija; tekstualni
        #   redak u tom rasponu ne zadovolji nijedan kriterij, pa se **ne broji**
        #   — stupac ostane na prethodnoj brojci i izgleda kao da formula fali.
        cd = tg.ws.cell(row, DATE_COL)
        cd.value = datetime(r['date'].year, r['date'].month, r['date'].day)
        cd.number_format = date_fmt
        tg.ws.cell(row, COMMENT_COL).value = r['komentar']
        tg.put(row, 'Racun',  racun)
        tg.put(row, 'Izvor',  r['izvor'])       # ⚠ prazni retci nose 'Racun' — mora se prepisati
        tg.put(row, 'Smjer',  r['smjer'])
        tg.put(row, 'Uplata'  if r['smjer'] == 'Uplata' else 'Isplata', r['iznos'])
        tg.put(row, 'Tip',    r['tip'])
        tg.put(row, 'Podtip', r['podtip'])
        # Retci s izvoda su se vec dogodili. Kokini nose vlastiti status: prazna
        # kolona C (naplata jos nepoznata) je upravo ono sto u nasem modelu znaci
        # `Planiran` (S113) - a planirani retci MORAJU ostati vidljivi, inace ih
        # korisnik dopise iz bankovne aplikacije i dobijemo ih dvaput.
        tg.put(row, 'Status', r.get('status', 'Izvrsen'))
        tg.put(row, 'Izvod opis', r['opis'])
        cn = tg.find('Datum naplate')
        if cn:
            tg.ws.cell(row, cn).value = r['naplata']
            tg.ws.cell(row, cn).number_format = 'd.m.yyyy'
        if r.get('rata'):
            tg.put(row, 'Rate?', True)          # PRAVI bool — 'DA' bi se uvezlo kao FALSE
            tg.put(row, 'Broj rata', r['rata'][1])
            tg.put(row, 'Rata br',   r['rata'][0])

    if appended:
        prosiri_kosaru(tg, next_free - 1)

    spilled_racun = [r for i, r in enumerate(rows)
                     if i >= len(tg.blank_rows) and r['izvor'] == 'Racun']
    if spilled_racun:
        print()
        print(f'⚠ {len(spilled_racun)} Racun redaka nije stalo u prazne retke — '
              f'kontrolni stupac ih NE broji, pa mu brojka na dnu nije potpuna. '
              f'Podijeli uvoz ili izvezi delta sheet s više praznih redaka.')

    # Autofilter mora obuhvatiti dopisane retke, inače ih prvi sort raspari.
    if appended and tg.ws.auto_filter.ref:
        last = next_free - 1
        tg.ws.auto_filter.ref = (f'A{tg.header_row}:'
                                 f'{get_column_letter(tg.last_col)}{last}')
    return used_blanks, appended


def main() -> None:
    ap = argparse.ArgumentParser(description='Popuni app-ov Excel retcima s bankovnog izvoda.')
    ap.add_argument('target', type=Path, help='xlsx skinut iz appa (delta sheet ili export)')
    ap.add_argument('--rf',      type=Path, help='RF izvod (PDF, OCR)')
    ap.add_argument('--zaba',    type=Path, help='ZABA izvadak tekućeg računa (PDF)')
    ap.add_argument('--visa',    type=Path, help='PBZ Visa račun (PDF)')
    ap.add_argument('--mc',      type=Path,
                    help='Mastercard izvod (PDF) — retci idu u sekciju KOSARA, '
                         'ispod nje, i ulaze u `Σ kosara`. Trazi --presedan '
                         '(dedup ide protiv baze, ne protiv lista).')
    ap.add_argument('--mc-izvor', dest='mc_izvor', default='Mastercard',
                    help='vrijednost atributa Izvor za --mc retke')
    ap.add_argument('--mc-tol', dest='mc_tol', type=int, default=5,
                    help='tolerancija u danima oko prozora MC izvoda')
    ap.add_argument('--naplata', help='datum skupne naplate Vise, YYYY-MM-DD (s RF izvoda)')
    ap.add_argument('--od',      help='uzmi retke od datuma (YYYY-MM-DD)')
    ap.add_argument('--do',      help='uzmi retke do datuma (YYYY-MM-DD)')
    ap.add_argument('--racun',   help='vrijednost atributa Racun (zadano po izvoru)')
    ap.add_argument('--koka',    type=Path,
                    help='Kokina Excelica — iz nje se preuzimaju OPISI (iznos i datum '
                         'ostaju s izvoda)')
    ap.add_argument('--iz-koke', type=Path, dest='iz_koke',
                    help='Kokina Excelica kao IZVOR redaka (iznos i datum su njeni). '
                         'Za razdoblje koje izvod jos ne pokriva - v. D-2.')
    ap.add_argument('--sheet',   default='koka EU',
                    help='list u Kokinoj Excelici (zadano: "koka EU")')
    ap.add_argument('--tip-racuna', dest='tip_racuna', default='Kokin tekuci',
                    help='vrijednost njene kolone A: Kokin tekuci | Sasin tekuci | '
                         'Mastercard | Visa')
    ap.add_argument('--klasificiraj', action='store_true',
                    help='popuni Tip/Podtip iz izbrojane povijesti (PO_OPISU); '
                         'bez toga svi retci dobiju N/A - legitimno, ne blokira uvoz')
    ap.add_argument('--osim', default='',
                    help='brojevi redaka Kokinog lista koje NE uzimati, npr. "2564". '
                         'Za redak za koji je provjereno da je vec u bazi pod drugim '
                         'datumom - ispravak ide u NJEN file, ne u uvoz (S115).')
    ap.add_argument('--lanac', metavar='DATUM=IZNOS',
                    help='ispisi lanac njenog lista od tog stanja, npr. 2026-07-30=13815.33')
    ap.add_argument('--protiv',  type=Path,
                    help='dodatni app-ov export protiv kojeg se radi dedup (retci koje '
                         'target sheet ne pokazuje — npr. kartične stavke)')
    ap.add_argument('--zigosi', action='store_true',
                    help='upisi `Izvod opis` i na retke koje baza vec ima, a ovaj '
                         'izvod ih potvrdjuje (mijenja postojece retke -> uvoz ih '
                         'javi kao Modify)')
    ap.add_argument('--presedan', choices=['prod', 'test'],
                    help='predloži Tip/Podtip brojanjem povijesti TOG računa u bazi '
                         '(v. presedani.py); bez njega neprepoznati redak ostaje N/A')
    ap.add_argument('--dry',     action='store_true', help='samo ispiši što bi upisao')
    a = ap.parse_args()

    if not a.rf and not a.visa and not a.zaba and not a.iz_koke and not a.mc:
        sys.exit('X Zadaj barem jedan izvor: --rf, --zaba, --visa, --mc ili --iz-koke.')
    if a.zaba and (a.rf or a.visa):
        sys.exit('X ZABA i RF su razliciti racuni - jedan file, jedan racun.')
    if a.iz_koke and (a.rf or a.zaba or a.visa):
        # Autoritet za iznos i datum ne moze biti i izvod i njen file u istom
        # prolazu: gdje se razilaze (~4 % redaka) nema pravila koje bi presudilo.
        sys.exit('X --iz-koke se ne kombinira s izvodom. Jedan prolaz, jedan autoritet.')
    # WARN Normalizacija (`_koka_norm`) sluzi SAMO za usporedbu njene kolone A.
    #   Vrijednost atributa `Racun` koja ide u bazu nosi dijakritike i mora se
    #   poklopiti u znak - inace redak zavrsi pod novim, cetvrtim racunom, a
    #   plocica to prikaze kao uredan racun sa svojim saldom.
    racun = a.racun or (
        ('Kokin tekući ZABA' if _koka_norm(a.tip_racuna).startswith('kokin')
         else 'Sašin tekući RF') if a.iz_koke
        # `--mc` ide na isti racun kao `--zaba`: Mastercard tereti Kokin tekuci
        # jednom skupnom naplatom, pa karticni redak nosi TAJ racun i
        # `Izvor = Mastercard`. (Visa je Sasina i tereti RF — otud podjela.)
        else ('Kokin tekući ZABA' if (a.zaba or a.mc) else 'Sašin tekući RF'))
    if a.visa and not a.naplata:
        sys.exit('✗ --visa traži --naplata (dan kad je banka skinula skupnu naplatu; piše na RF izvodu).')

    od = date.fromisoformat(a.od) if a.od else None
    do = date.fromisoformat(a.do) if a.do else None

    koka = KokaOpisi(a.koka) if a.koka else None
    tg = Target(a.target)
    print(f'Target: {a.target.name} — {len(tg.real_rows)} postojećih redaka, '
          f'{len(tg.blank_rows)} praznih redaka predloška')

    # ⚠ Dedup vidi samo ono što je NA listu. Delta sheet nosi isključivo retke
    #   koji miču saldo, pa kartične stavke ondje ne postoje — a baza ih ima.
    #   Bez druge reference kartični izvod ulazi drugi put, i to tiho: saldo se
    #   ne mijenja, pa se greška ne pokaže ni na pločici ni u kontrolnom stupcu.
    extra_keys: set = set()
    ref: Target | None = None
    if a.protiv:
        ref = Target(a.protiv)
        extra_keys = ref.existing_keys()
        print(f'Dedup i protiv: {a.protiv.name} — {len(ref.real_rows)} redaka')

    # ⚠ Delta sheet je uvijek sheet JEDNOG računa, a dedup se radi protiv redaka
    #   koji su na njemu. Izvod drugog računa u tom fileu prolazi bez ijedne
    #   poruke: ništa se ne poklopi, sve izgleda kao „novo", i dobiješ tuđe
    #   retke pod svojim računom — plus kontrolni stupac koji ih ne broji.
    have = tg.accounts()
    if len(have) == 1 and racun not in have:
        print(f'✗ File je za račun {next(iter(have))!r}, a izvor traži {racun!r}.')
        sys.exit('  Izvezi delta sheet za taj račun, ili zadaj --racun ako znaš što radiš.')
    if len(have) > 1:
        print(f'  ⚠ File nosi više računa ({", ".join(sorted(have))}) — dedup je '
              f'zato slabiji; provjeri popis prije upisa.')

    rows: list[dict] = []
    if a.rf:
        rf = rf_rows(a.rf, od, do)
        # Dedup protiv redaka koji su već na listu: prozor delta sheeta pokazuje
        # što baza ima, pa se dvostruki uvoz vidi OVDJE, prije nego što nastane.
        keys = tg.existing_keys() | extra_keys
        # ⚠ Dedup po TOČNOM (datum, iznos) propušta isti redak zaveden pod
        #   susjednim datumom — a to nije rijetkost nego pravilo: Kokin file i
        #   izvod se za knjiženje znaju razići za dan (`Mirovina III stup`:
        #   izvod 09.07., baza 10.07.). Bez tolerancije bi takav redak ušao
        #   DRUGI PUT, a saldo bi se razišao za točno taj iznos.
        dup, near, new = [], [], []
        for r in rf:
            k = (r['date'].isoformat(), r['iznos'])
            if k in keys:
                dup.append(r)
                continue
            hit = next((d for (d, a) in keys
                        if a == r['iznos']
                        and abs((date.fromisoformat(d) - r['date']).days) <= DATE_TOL), None)
            (near.append((r, hit)) if hit else new.append(r))
        rf = new
        print(f'RF izvod: {len(rf)} novih, {len(dup)} već na listu (preskočeno)')
        for r in dup:
            print(f'   = {r["date"]} {r["iznos"]:>9.2f}  {r["opis"][:45]}')
        for r, hit in near:
            print(f'   ≈ {r["date"]} {r["iznos"]:>9.2f}  {r["opis"][:40]}')
            print(f'       ⚠ isti iznos već postoji na {hit} — PRESKOČENO kao isti redak '
                  f'pod drugim datumom. Autoritet za datum je izvod, pa je popravak '
                  f'datuma u bazi (ne novi redak).')
        rows += rf
    if a.zaba:
        svi_zb = zaba_rows(a.zaba, od, do, koka)
        keys = tg.existing_keys() | extra_keys
        dup = [r for r in svi_zb if (r['date'].isoformat(), r['iznos']) in keys]
        zb  = [r for r in svi_zb if (r['date'].isoformat(), r['iznos']) not in keys]
        prije_1n = len(zb)
        zb, spojevi = objasni_spojene(tg, zb, svi_zb)
        if a.zigosi:
            _mail = str(tg.ws.cell(tg.blank_rows[0], USER_COL).value or '').strip()                 if tg.blank_rows else ''
            _n, _tudji = zigosi_postojece(tg, svi_zb, spojevi, _mail, 'Racun')
            for red, mail in _tudji:
                print('   WARN r%d je redak korisnika %s — uvoz ga smije ispraviti '
                      'samo kroz „Ispravi kao vlasnik Aree" (S125). Bez tog izbora '
                      'se preskace, i to bez poruke o pravima.' % (red, mail))
        spojeno = prije_1n - len(zb)
        print(f'ZABA izvod: {len(zb)} novih, {len(dup)} već na listu (preskočeno)'
              + (f', {spojeno} pokriveno spojenim retkom' if spojeno else ''))
        novo, last = zaba_printed_balance(a.zaba)
        if novo is not None:
            print(f'► Ispisano NOVO STANJE: {novo:.2f}'
                  + (f' na {last} (zadnja transakcija na izvodu)' if last else ''))
            print('  ⚠ Kontrolni stupac se s tim poklapa SAMO ako prozor sheeta seže '
                  'do tog datuma i ako prije njega ne fali nijedan redak.')
        rows += zb
    if a.visa:
        vs = visa_rows(a.visa, date.fromisoformat(a.naplata), od, do, koka)
        s = round(sum(r['iznos'] for r in vs), 2)
        print(f'Visa račun: {len(vs)} kupovina, Σ {s:.2f} — mora biti jednako '
              f'skupnoj naplati na RF izvodu, inače košara nije potpuna')
        keys = tg.existing_keys() | extra_keys
        dup = [r for r in vs if (r['date'].isoformat(), r['iznos']) in keys]
        vs  = [r for r in vs if (r['date'].isoformat(), r['iznos']) not in keys]
        if dup:
            print(f'  {len(dup)} stavki već postoji (preskočeno):')
            for r in dup:
                print(f'   = {r["date"]} {r["iznos"]:>9.2f}  {r["opis"][:45]}')
        if not a.protiv:
            print('  ⚠ BEZ --protiv: delta sheet ne sadrži kartične retke, pa dedup '
                  'nema što usporediti. Izvezi iz appa retke s Izvor=Visa za razdoblje '
                  'izvoda i predaj ih kao --protiv, inače stavke koje baza već ima '
                  'ulaze DRUGI PUT — a saldo to ne osjeti.')
        rows += vs

    if a.mc:
        # ⚠ Dedup ide protiv BAZE, ne protiv lista: delta sheet nosi samo one
        #   karticne retke koji su u kosari (dospijece jos nije proslo), a baza
        #   ih ima jos. Bez toga bi svaka vec uvezena stavka usla drugi put —
        #   i to tiho, jer karticni redak saldo ne dira pa se greska ne pokaze
        #   ni na plocici ni u kontrolnom stupcu.
        if not a.presedan:
            sys.exit('X --mc trazi --presedan prod|test (dedup se radi protiv baze).')
        from uskladi_izvod import (load_db as _mld, load_env as _mle, parse_mc,
                                   uskladi as _uskladi)
        _u, _k = _mle(a.presedan)
        izm = parse_mc(a.mc)
        us = _uskladi(izm, _mld(_u, _k), a.mc_izvor, a.mc_tol)
        # `daleki` = redak izvoda bez para U PROZORU, ali isti trosak postoji u
        # bazi izvan njega. Uvoz bi ga UDVOSTRUCIO, a dedup po (datum, iznos) ga
        # ne vidi jer je datum upravo ono sto je krivo (razred S115).
        sumnjivi = [x for x in us['izvod_bez_para'] if x['ref'] in us['daleki']]
        novi = [x for x in us['izvod_bez_para'] if x['ref'] not in us['daleki']]
        print('MC izvod %s: dospijece %s, %d redaka / %.2f'
              % (a.mc.name, izm['dospijece'], len(izm['rows']),
                 sum(x['iznos'] for x in izm['rows'])))
        print('  %d za uvoz, %d vec u bazi, %d SUMNJIVIH (ne uvozim)'
              % (len(novi), len(izm['rows']) - len(us['izvod_bez_para']), len(sumnjivi)))
        for x in sumnjivi:
            bliski = ', '.join('%s (%d d)' % (r['event_date'], dd)
                               for r, dd in us['daleki'][x['ref']][:3])
            print('   ? %s %9.2f  %-34s — isti iznos vec u bazi: %s'
                  % (x['datum'], x['iznos'], x['opis'][:34], bliski))
        if sumnjivi:
            print('       (ne uvozim ih: vjerojatno isti trosak pod krivim datumom. '
                  'Prvo ispravak datuma u bazi, pa uvoz — obrnuto udvostrucuje.)')
        for x in novi:
            m = re.search(r'RATA\s+(\d+)\s*/\s*(\d+)', x['opis'], re.I)
            kom = (koka.find(x['datum'], x['iznos']) if koka else None) or short_opis(x['opis'])
            rows.append({
                'date': x['datum'], 'smjer': 'Isplata', 'iznos': x['iznos'],
                'opis': x['opis'], 'opis_puni': x['opis'],
                'izvor': a.mc_izvor, 'tip': NA, 'podtip': NA,
                'komentar': kom, 'opis_izvor': 'koka' if koka and kom != short_opis(x['opis']) else 'izvod',
                # ⚠ `Planiran` jer dospijece JOS NIJE PROSLO. Status prelazi u
                #   Izvrsen tek kad se kosara slozi sa skupnom naplatom na
                #   izvatku tekuceg — nikad kao zakljucak iz dospijeca.
                'naplata': izm['dospijece'], 'status': 'Planiran', 'stanje': None,
                'rata': (int(m.group(1)), int(m.group(2))) if m else None,
            })

    if a.iz_koke:
        osim = {int(x) for x in a.osim.replace(';', ',').split(',') if x.strip()}
        kr = koka_rows(a.iz_koke, a.sheet, a.tip_racuna, od, do, a.klasificiraj, osim)
        keys = tg.existing_keys() | extra_keys
        # WARN Dedup po (datum, iznos) NE hvata skoro-duplikate: kad dva izvora
        #   opisuju isti dogadaj razlicitim iznosom (Koka 1.265,59, banka
        #   1.285,59 - zamijenjena znamenka), kljuc se razlikuje i oba retka udu
        #   (S111, 9 takvih na jednom racunu). Zato i tolerancija na datum, i
        #   zato se `~` retci ISPISUJU umjesto da se tiho progutaju.
        dup, near, new_rows = [], [], []
        for r in kr:
            k = (r['date'].isoformat(), r['iznos'])
            if k in keys:
                dup.append(r)
                continue
            hit = next((d for (d, am) in keys
                        if am == r['iznos']
                        and abs((date.fromisoformat(d) - r['date']).days) <= DATE_TOL), None)
            (near.append((r, hit)) if hit else new_rows.append(r))
        print('\nKokin file (%s / %s): %d novih, %d vec na listu, %d blizu postojecem'
              % (a.iz_koke.name, a.tip_racuna, len(new_rows), len(dup), len(near)))
        for r in dup:
            print('   = %s %9.2f  %s' % (r['date'], r['iznos'], r['komentar'][:45]))
        for r, hit in near:
            print('   ~ %s %9.2f  %s' % (r['date'], r['iznos'], r['komentar'][:40]))
            print('       WARN isti iznos vec postoji na %s - PRESKOCENO. Provjeri je li'
                  % hit)
            print('       to isti redak pod drugim datumom prije nego ga vratis rucno.')
        for r in new_rows:
            rata = ('  rata %d/%d' % r['rata']) if r.get('rata') else ''
            print('   + red %-5d %s %-11s %9.2f  %-26s %s/%s%s'
                  % (r['red'], r['date'], r['izvor'], r['iznos'],
                     r['komentar'][:26], r['tip'], r['podtip'], rata))
        rows += new_rows

        if a.lanac:
            d_txt, _, i_txt = a.lanac.partition('=')
            koka_lanac(a.iz_koke, a.sheet, date.fromisoformat(d_txt), float(i_txt),
                       do, osim)

    if not rows:
        print('Nema ničega za upisati.')
        return

    if a.presedan:
        from presedani import Presedani
        from uskladi_izvod import load_db as _load_db, load_env as _load_env
        _url, _key = _load_env(a.presedan)
        # WARN Povijest je ono STARIJE od prozora koji uskladujemo. Bez te
        #   granice bi se prozor klasificirao sam sobom — to nije presedan nego
        #   jeka vlastitog uvoza.
        _prije = min(r['date'] for r in rows).isoformat()
        _svi = [d for d in _load_db(_url, _key)
                if str(d['attrs'].get('Racun')) == racun]
        _po_izvoru: dict = {}
        _pogodaka, _dvojbeni = 0, []
        for r in rows:
            if r['tip'] != NA:
                continue
            izv = r['izvor']
            if izv not in _po_izvoru:
                _po_izvoru[izv] = Presedani(
                    [d for d in _svi if str(d['attrs'].get('Izvor')) == izv],
                    prije=_prije)
            hit = _po_izvoru[izv].nadji(r['iznos'], r.get('opis_puni') or r['opis'],
                                        r['smjer'])
            if not hit:
                continue
            r['tip'], r['podtip'] = hit['tip'], hit['podtip']
            r['presedan'] = hit['dokaz']
            # WARN Kokin opis se NE prepisuje presedanom — ona je autoritet za
            #   znacenje (CLAUDE.md „Politika izvora"). Prepisuje se samo
            #   strojni tekst izvoda, koji ionako ne kaze nista.
            if hit['comment'] and r.get('opis_izvor') != 'koka':
                r['komentar'] = hit['comment']
            if hit['alternative']:
                _dvojbeni.append((r, hit['alternative']))
            _pogodaka += 1
        print()
        print('Presedani (%s, povijest racuna %r prije %s): %d od %d redaka '
              'klasificirano brojanjem'
              % (a.presedan, racun, _prije, _pogodaka, len(rows)))
        for r in rows:
            if r.get('presedan'):
                print('   > %s %9.2f  %-26s %-24s (%s)'
                      % (r['date'], r['iznos'], (r['tip'] + '/' + r['podtip'])[:26],
                         r['komentar'][:24], r['presedan']))
        for r, alt in _dvojbeni:
            print('   WARN %s %9.2f  par je jednoglasan, ALI komentar nije: %s'
                  % (r['date'], r['iznos'], ', '.join(alt)))
            print('        komentar zato NIJE upisan — odaberi ga rukom u fileu.')

    rows.sort(key=lambda r: (r['date'], r['iznos']))
    print(f'\nZa upis: {len(rows)} redaka')
    for r in rows:
        flag = '' if r['tip'] != NA else '  ⚠ Tip=N/A'
        if r.get('opis_izvor') == 'izvod':
            flag += '  ⚠ nema Kokinog opisa'
        print(f'   + {r["date"]} {r["smjer"]:<7} {r["iznos"]:>9.2f}  '
              f'{r["tip"]}/{r["podtip"]:<16} {r["komentar"][:34]:<34}{flag}')

    if koka:
        print()
        print(f'Kokini opisi: {koka.hits} spareno, {koka.misses} bez para '
              f'(za te retke ostaje skraceni tekst izvoda).')
    na = sum(1 for r in rows if r['tip'] == NA)
    if na:
        print(f'\n⚠ {na} redaka ide s Tip=N/A — legitimno za uvoz, ali ih netko '
              f'mora klasificirati (u appu ili pravilima).')

    last_stanje = next((r['stanje'] for r in reversed(rows) if r.get('stanje') is not None), None)
    if last_stanje is not None:
        print(f'\n► Kontrolni stupac mora završiti na {last_stanje:.2f} — '
              f'to je ISPISANO stanje s izvoda, ne naš izračun.')

    if a.dry:
        print('\n--dry: ništa nije zapisano.')
        return

    used, app = write_rows(tg, rows, racun, a.dry, ref)
    out = a.target.with_name(f'{a.target.stem}_filled.xlsx')
    for note in strip_comments(tg.wb):
        first = note.strip().splitlines()[0] if note.strip() else ''
        print(f'  (bilješka maknuta radi uvoza — original je čuva: {first})')
    tg.wb.save(out)
    print(f'\n✓ {used} praznih redaka popunjeno, {app} dopisano ispod')
    print(f'✓ {out}')
    print('  Original je netaknut. Otvori novi file, provjeri kontrolni stupac, pa uvezi.')


if __name__ == '__main__':
    main()
