# -*- coding: utf-8 -*-
"""
fix_datum_naplate_statement.py — `Datum naplate` iz bankovnog statementa. S107x, 2026-08-12.

NALAZ (Faza 1a, `verify_saldo_model.py --nalazi`): 27 MC redaka ima `Datum naplate`
PRIJE `event_date` — nemoguće. Provjera protiv statementa (`Izvod file`) pokazala je da
je to samo vidljivi vrh: **57 redaka ne odgovara svom statementu**, od čega je 30 „tiho
krivo" (naplata je poslije kupnje pa izgleda uredno).

UZROK: `kartice_datum_naplate.py` namjerno NE dira retke koji već imaju popunjen
`Datum naplate` (linija ~130, `stats['već popunjeno (ne diram)']`). Kod ovih 57 je
vrijednost bila upisana ranijim alatom, pa je stara vrijednost preživjela; poslije je
`enrich_from_izvoda.py` pridružio statement, a `date_accuracy.py` (S107k) pomaknuo
`event_date` na bankovni datum. Naplatu nitko nije preračunao.

PRAVILO (isto kao `naplata_for_mc` u `kartice_datum_naplate.py`):
    Datum naplate = 11. u mjesecu NAKON statementa iz `Izvod file`
Statement je nezavisna istina — 1.016 redaka se već slaže s njim, 57 ne.

OPSEG: samo `Izvor = Mastercard` s `Izvod file` koji počinje `MC_`.
668 MC redaka bez statementa se NE dira (njima vrijedi cutoff pravilo).
`Izvor = Visa` se NE dira (drukčiji mehanizam — v. `kartice_datum_naplate.py`).

⚠ 7 od 57 ide u SUPROTNOM smjeru (pohranjena naplata je mjesec KASNIJE od očekivane).
Prikazuju se u zasebnom bloku; bez `--include-obrnute` se NE diraju.

Pokretanje:
    python fix_datum_naplate_statement.py --dry                    # samo prikaz
    python fix_datum_naplate_statement.py                          # 50 redaka (bez obrnutih)
    python fix_datum_naplate_statement.py --include-obrnute        # svih 57
Backup: `.pre-naplata-*` prije snimanja.
"""
from __future__ import annotations

import shutil
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from apply_rules import find_header_col, pick_file  # noqa: E402

MC_DAY = 11          # Kokino pravilo: naplata = 11. u mjesecu nakon statementa
MIN_MATCH = 900      # sanity: toliko redaka se MORA već slagati, inače je pravilo krivo


def ym_add(ym: str, k: int) -> str:
    y, m = int(ym[:4]), int(ym[5:7])
    m += k
    while m > 12:
        y, m = y + 1, m - 12
    while m < 1:
        y, m = y - 1, m + 12
    return f'{y}-{m:02d}'


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    return v if isinstance(v, date) else None


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    include_obrnute = '--include-obrnute' in args
    path = pick_file([a for a in args if a not in ('--dry', '--include-obrnute')])
    print(f'File: {path.name}{"  [DRY RUN — bez snimanja]" if dry else ""}\n')

    wb = openpyxl.load_workbook(path)
    ws = wb['Review']
    c_izv = find_header_col(ws, 'Izvor')
    c_ed = find_header_col(ws, 'event_date')
    c_dn = find_header_col(ws, 'Datum naplate')
    c_file = find_header_col(ws, 'Izvod file')
    c_key = find_header_col(ws, 'source_key')
    c_nap = find_header_col(ws, 'Napomena')
    c_op = find_header_col(ws, 'Izvod opis')

    naprijed, obrnute, prazne = [], [], []
    match = bez_filea = 0

    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, c_izv).value or '').strip() != 'Mastercard':
            continue
        f = str(ws.cell(r, c_file).value or '').strip()
        if not f.startswith('MC_'):
            bez_filea += 1
            continue
        ed = to_date(ws.cell(r, c_ed).value)
        dn = to_date(ws.cell(r, c_dn).value)
        if ed is None:
            continue
        ym = f.split('.pdf')[0].replace('MC_', '')
        try:
            nxt = ym_add(ym, 1)
            ocek = date(int(nxt[:4]), int(nxt[5:7]), MC_DAY)
        except ValueError:
            sys.exit(f'✗ red {r}: neispravan `Izvod file` "{f}" — PREKID.')
        if dn is None:
            prazne.append((r, ed, dn, ocek, f))
        elif dn == ocek:
            match += 1
        elif dn < ocek:
            naprijed.append((r, ed, dn, ocek, f))
        else:
            obrnute.append((r, ed, dn, ocek, f))

    print(f'MC redaka sa statementom : {match + len(naprijed) + len(obrnute) + len(prazne)}'
          f'   (bez statementa, ne diram: {bez_filea})')
    print(f'  već se slaže           : {match}')
    print(f'  naplata PRERANA        : {len(naprijed)}   → popravlja se')
    print(f'  naplata PREKASNA       : {len(obrnute)}   → {"popravlja se" if include_obrnute else "NE dira se (bez --include-obrnute)"}')
    if prazne:
        print(f'  prazna naplata         : {len(prazne)}   → popunjava se')

    if match < MIN_MATCH:
        sys.exit(f'✗ Samo {match} redaka se slaže sa statementom (očekivano ≥{MIN_MATCH}) — '
                 f'pravilo je vjerojatno krivo za ovaj file. PREKID, ništa nije upisano.')

    def opis(r):
        return (str(ws.cell(r, c_nap).value or '')
                or str(ws.cell(r, c_op).value or ''))[:30]

    nemoguci = sum(1 for r, ed, dn, _, _ in naprijed if dn and dn < ed)
    print(f'\n— PRERANA naplata ({len(naprijed)}), od toga {nemoguci} "nemogućih" '
          f'(naplata prije kupnje) —')
    for r, ed, dn, ocek, f in naprijed:
        flag = ' ⚠ nemoguće' if dn and dn < ed else ''
        print(f'   red {r:>5}  kupnja {ed}  naplata {dn} → {ocek}  {f:<20}'
              f' {opis(r)}{flag}')

    if obrnute:
        print(f'\n— PREKASNA naplata ({len(obrnute)}) — pohranjena je KASNIJE od očekivane;'
              f'\n  moguće legitimno (kupnja tik uz granicu statementa) — pregledaj prije diranja —')
        for r, ed, dn, ocek, f in obrnute:
            print(f'   red {r:>5}  kupnja {ed}  naplata {dn} → {ocek}  {f:<20} {opis(r)}')

    meta = list(naprijed) + list(prazne) + (list(obrnute) if include_obrnute else [])
    print(f'\n{"Bi se promijenilo" if dry else "Mijenja se"}: {len(meta)} redaka')
    if dry or not meta:
        if dry:
            print('DRY RUN — ništa nije snimljeno.')
        return

    # kontrolni otisak: samo `Datum naplate` smije se promijeniti
    kljucevi = {r: str(ws.cell(r, c_key).value or '') for r, *_ in meta}
    for r, ed, dn, ocek, f in meta:
        ws.cell(r, c_dn).value = ocek
        ws.cell(r, c_dn).number_format = 'DD.MM.YYYY'

    backup = path.with_name(f'{path.stem}.pre-naplata-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori {path.name} u Excelu i ponovi.\n'
                 f'  (backup je već napravljen: {backup.name})')
    print(f'✔ Snimljeno. Backup: {backup.name}')
    print(f'  promijenjenih source_keyeva: {len(set(kljucevi.values()))}')
    print(f'  raspodjela po godini: '
          f'{dict(sorted(Counter(ed.year for _, ed, *_ in meta).items()))}')


if __name__ == '__main__':
    main()
