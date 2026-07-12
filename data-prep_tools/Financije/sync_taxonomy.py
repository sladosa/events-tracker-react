# -*- coding: utf-8 -*-
"""
sync_taxonomy.py  (S107c, 2026-07-12)
=====================================
Sinkronizira dropdowne u Financije REVIEW Excelu s (editiranim) `Taksonomija` sheetom.

Problem koji rješava: dropdowni Tip/Podtip na Review sheetu su generirani pri exportu
(normalize_financije.py) iz TAXONOMY dicta. Kad Saša/Koka za vrijeme ručnog reviewa
dodaju novi Tip ili Podtip u `Taksonomija` sheet, dropdowni to NE vide.
Ova skripta čita `Taksonomija` sheet i regenerira:
  - skriveni `Liste` sheet + named range-ove (TipList, Tip_<sanitized>)
  - Data Validation na Tip/Podtip kolonama Review sheeta (dependent INDIRECT dropdown)
  - Conditional Formatting (crveni mismatch Podtip↔Tip, žuti prazan/N-A Tip)

RUČNI RAD U REVIEW SHEETU SE NE DIRA — mijenjaju se samo validacije/formatting/skriveni sheet.
Prije snimanja radi backup kopiju (`<ime>.pre-sync-<timestamp>.xlsx`).

Pokretanje (file smije biti zatvoren u Excelu!):
  PYTHONUTF8=1 C:/0_Sasa/events-tracker/venv/Scripts/python.exe sync_taxonomy.py
      → uzima NAJNOVIJI Financije_review_*.xlsx iz data-prep_data/Financije
  ... sync_taxonomy.py <puna putanja do .xlsx>   → eksplicitni file
"""

import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")

RED_FILL = PatternFill('solid', fgColor='FFC7CE')
YEL_FILL = PatternFill('solid', fgColor='FFEB9C')

# Dijakritike → ASCII (mora ostati usklađeno s normalize_financije.sanitize_name)
DIACRITICS = {'ć': 'c', 'č': 'c', 'š': 's', 'ž': 'z', 'đ': 'd',
              'Ć': 'C', 'Č': 'C', 'Š': 'S', 'Ž': 'Z', 'Đ': 'D'}


def sanitize_name(s: str) -> str:
    for a, b in DIACRITICS.items():
        s = s.replace(a, b)
    return re.sub(r'[^A-Za-z0-9_]', '_', s)


def pick_file() -> Path:
    if len(sys.argv) > 1:
        p = Path(sys.argv[1])
        if not p.exists():
            sys.exit(f'✗ File ne postoji: {p}')
        return p
    candidates = sorted(DATA_DIR.glob('Financije_review_*.xlsx'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    candidates = [c for c in candidates if '.pre-sync-' not in c.name]
    if not candidates:
        sys.exit(f'✗ Nema Financije_review_*.xlsx u {DATA_DIR}')
    return candidates[0]


def read_taxonomy(wb) -> dict[str, list[str]]:
    """Čita Taksonomija sheet → {tip: [podtipovi]} uz očuvanje redoslijeda prvog pojavljivanja."""
    if 'Taksonomija' not in wb.sheetnames:
        sys.exit('✗ Nema sheeta "Taksonomija" u fileu.')
    tws = wb['Taksonomija']
    taxonomy: dict[str, list[str]] = {}
    problems: list[str] = []
    seen_pairs: set[tuple[str, str]] = set()

    for r in range(2, tws.max_row + 1):
        tip = tws.cell(r, 1).value
        pod = tws.cell(r, 2).value
        tip = str(tip).strip() if tip not in (None, '') else ''
        pod = str(pod).strip() if pod not in (None, '') else ''
        if pod == '—':
            pod = ''
        if not tip and not pod:
            continue
        if not tip:
            problems.append(f'  red {r}: Podtip "{pod}" bez Tipa — PRESKOČEN')
            continue
        if (tip, pod) in seen_pairs:
            problems.append(f'  red {r}: duplikat ({tip}, {pod or "—"}) — preskočen')
            continue
        seen_pairs.add((tip, pod))
        taxonomy.setdefault(tip, [])
        if pod:
            taxonomy[tip].append(pod)

    if problems:
        print('⚠ Upozorenja pri čitanju Taksonomije:')
        print('\n'.join(problems))
    if not taxonomy:
        sys.exit('✗ Taksonomija sheet je prazan.')

    # Kolizija sanitiziranih imena (dva Tipa → isti named range) = fatalno
    by_sanitized: dict[str, list[str]] = {}
    for tip in taxonomy:
        by_sanitized.setdefault(sanitize_name(tip), []).append(tip)
    for san, tips in by_sanitized.items():
        if len(tips) > 1:
            sys.exit(f'✗ Tipovi {tips} se sanitiziraju u isto ime "{san}" — preimenuj jedan od njih.')
    return taxonomy


def build_substitute_chain(cell_ref: str, tips: list[str]) -> str:
    """SUBSTITUTE lanac za Excel stranu sanitizacije — pokriva SAMO znakove
    koji se stvarno pojavljuju u imenima Tipova (DV formula limit 255 znakova)."""
    specials: dict[str, str] = {}   # char → zamjena, insertion order
    for tip in tips:
        for ch in tip:
            if re.match(r'[A-Za-z0-9_]', ch) or ch in specials:
                continue
            specials[ch] = DIACRITICS.get(ch, '_')
    sub = cell_ref
    for a, b in specials.items():
        sub = f'SUBSTITUTE({sub},"{a}","{b}")'
    return sub


def find_header_col(ws, header: str) -> int:
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() == header:
            return c
    sys.exit(f'✗ Kolona "{header}" nije nađena u headeru Review sheeta.')


def main() -> None:
    path = pick_file()
    print(f'File: {path.name}')

    wb = openpyxl.load_workbook(path)
    taxonomy = read_taxonomy(wb)
    tips = list(taxonomy.keys())
    print(f'Taksonomija: {len(tips)} Tipova, '
          f'{sum(len(v) for v in taxonomy.values())} Podtipova')

    if 'Review' not in wb.sheetnames:
        sys.exit('✗ Nema sheeta "Review".')
    ws = wb['Review']
    col_tip = find_header_col(ws, 'Tip')
    col_pod = find_header_col(ws, 'Podtip')
    tip_ltr, pod_ltr = get_column_letter(col_tip), get_column_letter(col_pod)
    last = ws.max_row

    # ── 1. Regeneriraj Liste sheet + named range-ovi ─────────────────────────
    # Prvo makni stare defined names (TipList + svi Tip_*)
    for name in [n for n in list(wb.defined_names) if n == 'TipList' or n.startswith('Tip_')]:
        del wb.defined_names[name]
    if 'Liste' in wb.sheetnames:
        del wb['Liste']
    lst = wb.create_sheet('Liste')
    lst.cell(1, 1, 'Tip')
    for i, tip in enumerate(tips, 2):
        lst.cell(i, 1, tip)
    wb.defined_names.add(DefinedName('TipList', attr_text=f'Liste!$A$2:$A${1 + len(tips)}'))

    col = 2
    for tip, podtips in taxonomy.items():
        values = podtips if podtips else ['—']
        lst.cell(1, col, tip)
        for i, p in enumerate(values, 2):
            lst.cell(i, col, p)
        ltr = get_column_letter(col)
        wb.defined_names.add(DefinedName(
            f'Tip_{sanitize_name(tip)}', attr_text=f'Liste!${ltr}$2:${ltr}${1 + len(values)}'))
        col += 1
    lst.sheet_state = 'veryHidden'

    # ── 2. Zamijeni Data Validation na Review Tip/Podtip kolonama ────────────
    sub = build_substitute_chain(f'{tip_ltr}2', tips)
    dv_pod_formula = f'INDIRECT("Tip_"&{sub})'
    if len(dv_pod_formula) > 255:
        sys.exit(f'✗ Podtip DV formula ima {len(dv_pod_formula)} znakova (limit 255) — '
                 f'previše specijalnih znakova u imenima Tipova. Pojednostavi imena.')

    ws.data_validations.dataValidation = []   # makni SVE postojeće DV (Review ima samo ova 2)
    dv_tip = DataValidation(type='list', formula1='=TipList', allowBlank=True, showErrorMessage=False)
    ws.add_data_validation(dv_tip)
    dv_tip.add(f'{tip_ltr}2:{tip_ltr}{last}')
    dv_pod = DataValidation(type='list', formula1=dv_pod_formula, allowBlank=True, showErrorMessage=False)
    ws.add_data_validation(dv_pod)
    dv_pod.add(f'{pod_ltr}2:{pod_ltr}{last}')

    # ── 3. Regeneriraj Conditional Formatting (mismatch crveno, prazan/N-A žuto) ──
    ws.conditional_formatting = ConditionalFormattingList()
    ws.conditional_formatting.add(
        f'{pod_ltr}2:{pod_ltr}{last}',
        FormulaRule(
            formula=[f'AND({pod_ltr}2<>"",{pod_ltr}2<>"—",ISERROR(MATCH({pod_ltr}2,INDIRECT("Tip_"&{sub}),0)))'],
            fill=RED_FILL,
        ),
    )
    ws.conditional_formatting.add(
        f'{tip_ltr}2:{tip_ltr}{last}',
        FormulaRule(formula=[f'OR({tip_ltr}2="",{tip_ltr}2="N/A")'], fill=YEL_FILL),
    )

    # ── 4. Backup + save in place ─────────────────────────────────────────────
    backup = path.with_name(f'{path.stem}.pre-sync-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti {path.name} — zatvori file u Excelu i pokreni ponovno.\n'
                 f'  (Backup je već napravljen: {backup.name})')

    print(f'✔ Dropdowni sinkronizirani. Backup: {backup.name}')
    print('  Tipovi u dropdownu:')
    for tip, podtips in taxonomy.items():
        print(f'    {tip}: {" | ".join(podtips) if podtips else "—"}')


if __name__ == '__main__':
    main()
