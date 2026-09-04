# -*- coding: utf-8 -*-
"""
pregled_stanja.py — jedan file koji odgovara na pitanje "je li stanje tocno". S128.

--------------------------------------------------------------------------------
ZASTO GENERIRAN FILE, A NE OZNAKE U KOKINOJ EXCELICI

Sasina ideja je bila oznaciti sidra i sporne retke u `Financije 2026-08-16.xlsx`,
s autofilterom. Izmjereno je zasto to ne radi: **u njenom fileu nema sto oznaciti.**
Njen lanac zatvara, jer njen model tereti racun svakom karticnom stavkom. Nas
tereti jednom SKUPNOM naplatom s izvoda -- redak kojeg kod nje NEMA. Polovica
spornog materijala (skupne naplate, razdvojeni parking) u njenom fileu ne postoji,
pa bi oznake sjele na ispravne retke, a ondje gdje redak fali ne bi bilo nicega.

Zato: pregled se GENERIRA iz tri izvora (baza, izvodi, njen file) i pokazuje
razliku medju njima. Njena datoteka ostaje netaknuta.

--------------------------------------------------------------------------------
LISTOVI

  Pregled   svih ~30 izvoda: prozor, app, banka, Δ, n, sidro, status.
            Δ je promet MJESECA i NE prolazi kroz sidro (v. `promet_check.py`),
            pa vrijedi i za zasidrene mjesece gdje saldo po konstrukciji da nulu.
  Sporno    samo mjeseci s Δ ≠ 0, redak po redak: sto ima banka a nema baza i
            obrnuto. Jedan redak = jedna odluka. Autofilter.
  2023      zasto 2023. ne zatvara i zasto to nije greska -- s brojkama, da se
            pitanje ne otvara ponovo.

--------------------------------------------------------------------------------
⚠ SPARIVANJE JE ALAT ZA TRAZENJE, NE PRESUDA
Parovi se traze po (iznos, datum ±3 dana), pohlepno. To je dovoljno da se sporni
retci izdvoje, ali NIJE dokaz da je bas taj par ispravan -- `Cash 100,00` se
ponavlja svakih par tjedana (S114). Presuda je `uskladi_izvod.py` nad tim izvodom.

⚠ Kolona izvan autofiltera se pri prvom sortu raspari od retka ⇒ `auto_filter.ref`
pokriva sve kolone.

Pokretanje:
    python pregled_stanja.py                 # PROD, svi izvodi
    python pregled_stanja.py --od=2025-01
"""
from __future__ import annotations

import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).parent))
from make_saldo_anchors import EPS, GROUP_SLUG, GROUP_VALUE, printed_series  # noqa: E402
from uskladi_izvod import parse_koka_date, tekst  # noqa: E402
from verify_rpc_vs_model import (AREA_ID, ENV_FILE, FILTERS_IZVRSENO, Supa,  # noqa: E402
                                 load_env, pull_db)
from enrich_from_izvoda import _parse_zaba_all, _zaba_is_tekuci  # noqa: E402

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parents[2]
IZVODI = ROOT / 'data-prep_data' / 'Financije' / 'izvodi' / 'Analizirani_izvodi'
KOKA = ROOT / 'data-prep_data' / 'Financije' / 'Financije 2026-08-16.xlsx'
IZLAZ = ROOT / 'data-prep_data' / 'Financije'

HEAD = PatternFill('solid', fgColor='1F4E79')
OK = PatternFill('solid', fgColor='E2EFDA')
LOSE = PatternFill('solid', fgColor='FCE4D6')
SIDRO = PatternFill('solid', fgColor='FFF2CC')
BOLD_W = Font(bold=True, color='FFFFFF')


def koka_redci():
    """Njeni retci, samo za KONTEKST u listu `Sporno` — nikad kao presuda.
    Kolona C je dan kad novac napusti racun, G dan troska dok naplata nije
    poznata; gledaju se OBJE (S113)."""
    if not KOKA.exists():
        print('⚠ Kokin file nije nadjen — stupac "kod Koke" ostaje prazan.')
        return []
    wb = openpyxl.load_workbook(KOKA, read_only=True, data_only=True)
    out = []
    for sn in wb.sheetnames:
        for i, row in enumerate(wb[sn].iter_rows(min_row=2, values_only=True), start=2):
            c = list(row) + [None] * 12
            f = lambda x: float(x) if isinstance(x, (int, float)) else 0.0  # noqa: E731
            s = round(f(c[3]) - f(c[4]), 2)
            if not s:
                continue
            for raw in (c[2], c[6]):
                d = parse_koka_date(raw)
                if d:
                    out.append({'d': d, 's': s, 'opis': str(c[1] or '')[:44],
                                'gdje': sn + '!' + str(i), 'racun': str(c[0] or '')})
                    break
    return out


def banka_tx(stem):
    txs, _ = _parse_zaba_all(IZVODI / (stem + '.pdf'))
    return [{'d': t['date'],
             's': round(t['iznos'] if t['smjer'] == 'Uplata' else -t['iznos'], 2),
             'opis': t['opis']}
            for t in txs if _zaba_is_tekuci(t['account'])]


def spari(a, b, tol_dana=3):
    """Pohlepno po (iznos, datum ±tol). Vraca (nespareni_a, nespareni_b)."""
    slobodni = list(range(len(b)))
    ostali_a = []
    for x in a:
        nasao = None
        for dd in range(0, tol_dana + 1):
            for sgn in ((0,) if dd == 0 else (-1, 1)):
                for j in slobodni:
                    if (abs(b[j]['s'] - x['s']) < EPS
                            and b[j]['d'] == x['d'] + timedelta(days=sgn * dd)):
                        nasao = j
                        break
                if nasao is not None:
                    break
            if nasao is not None:
                break
        if nasao is None:
            ostali_a.append(x)
        else:
            slobodni.remove(nasao)
    return ostali_a, [b[j] for j in slobodni]


def main():
    lo = next((a[5:] for a in sys.argv[1:] if a.startswith('--od=')), None)
    hi = next((a[5:] for a in sys.argv[1:] if a.startswith('--do=')), None)

    sp = Supa(load_env(ENV_FILE))
    print('citam bazu…')
    db_all, _, _ = pull_db(sp)
    db = [r for r in db_all if r['racun'] == GROUP_VALUE and r['izvor'] == 'Racun'
          and r['status'] != 'Planiran' and r['signed']]
    anchors = {a['confirmed_on']: float(a['amount']) for a in sp.select_all(
        'balance_anchors?area_id=eq.' + AREA_ID + '&group_slug=eq.' + GROUP_SLUG
        + '&select=amount,confirmed_on,group_value&order=id')
        if a['group_value'] == GROUP_VALUE}
    print('citam izvode…')
    series = printed_series()
    kr = koka_redci()

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------- Pregled --
    ws = wb.active
    ws.title = 'Pregled'
    tekst(ws, 1, 1, 'PREGLED TOCNOSTI STANJA — ' + GROUP_VALUE
          + '   (generirano ' + datetime.now().strftime('%d.%m.%Y. %H:%M') + ')')
    ws.cell(1, 1).font = Font(bold=True, size=12)
    tekst(ws, 2, 1, 'Δ je promet MJESECA (app − banka) i ne prolazi kroz sidro — '
                    'vrijedi i za zasidrene mjesece. Mjesec s Δ ≠ 0 otvara se '
                    'listom "Sporno", presuda je uskladi_izvod.py.')
    ws.cell(2, 1).alignment = Alignment(wrap_text=True)
    ws.merge_cells('A2:I2')

    kol = ['izvod', 'prozor od', 'prozor do', 'app', 'banka', 'Δ (app−banka)',
           'redaka', 'sidro na taj dan', 'status']
    r0 = 4
    for j, k in enumerate(kol, start=1):
        c = tekst(ws, r0, j, k)
        c.fill, c.font = HEAD, BOLD_W

    sporni = []
    r = r0
    for prev, cur in zip(series, series[1:]):
        if (lo and cur['ym'] < lo) or (hi and cur['ym'] > hi):
            continue
        od, do = prev['close'], cur['close']
        red = [x for x in db if od < x['date'] <= do]
        app = round(sum(x['signed'] for x in red), 2)
        bnk = round(cur['novo'] - cur['pocetno'], 2)
        d = round(app - bnk, 2)
        sid = anchors.get(do.isoformat())
        r += 1
        ws.cell(r, 1, cur['ym'])
        ws.cell(r, 2, od)
        ws.cell(r, 3, do)
        ws.cell(r, 4, app)
        ws.cell(r, 5, bnk)
        ws.cell(r, 6, d)
        ws.cell(r, 7, len(red))
        ws.cell(r, 8, sid if sid is not None else None)
        if abs(d) < EPS:
            st, fill = ('zatvara u cent' + (' · zapecaceno sidrom' if sid else '')), OK
        else:
            st, fill = 'RAZILAZI SE — v. list Sporno', LOSE
            sporni.append((cur['ym'], od, do, d))
        tekst(ws, r, 9, st)
        for j in range(1, 10):
            ws.cell(r, j).fill = fill if abs(d) >= EPS or not sid else (SIDRO if sid else OK)
        for j in (2, 3):
            ws.cell(r, j).number_format = 'DD.MM.YYYY'
        for j in (4, 5, 6, 8):
            ws.cell(r, j).number_format = '#,##0.00'

    ws.auto_filter.ref = 'A' + str(r0) + ':I' + str(r)
    ws.freeze_panes = 'A' + str(r0 + 1)
    for j, w in enumerate([10, 12, 12, 13, 13, 15, 8, 17, 34], start=1):
        ws.column_dimensions[chr(64 + j)].width = w

    # -------------------------------------------------------------- Sporno --
    ws2 = wb.create_sheet('Sporno')
    tekst(ws2, 1, 1, 'SPORNI MJESECI — redak po redak. "samo BANKA" = redak koji '
                     'baza nema (fali). "samo BAZA" = redak koji banka nema (visak '
                     'ili krivi datum).')
    ws2.cell(1, 1).font = Font(bold=True)
    ws2.merge_cells('A1:H1')
    kol2 = ['mjesec', 'strana', 'datum', 'iznos', 'opis', 'kod Koke?',
            'gdje kod Koke', 'trag (id / —)']
    for j, k in enumerate(kol2, start=1):
        c = tekst(ws2, 3, j, k)
        c.fill, c.font = HEAD, BOLD_W

    r = 3
    for ym, od, do, d in sporni:
        stem = 'ZABA_' + ym
        if not (IZVODI / (stem + '.pdf')).exists():
            continue
        b = [x for x in banka_tx(stem) if od < x['d'] <= do]
        a = [{'d': x['date'], 's': x['signed'], 'opis': (x['comment'] or '')[:60],
              'id': x['id']} for x in db if od < x['date'] <= do]
        samo_baza, samo_banka = spari(a, b)
        r += 1
        c = tekst(ws2, r, 1, ym + '   Δ = ' + format(d, '.2f')
                  + '   (nespareno: baza ' + str(len(samo_baza))
                  + ', banka ' + str(len(samo_banka)) + ')')
        c.font = Font(bold=True)
        for x in samo_banka:
            r += 1
            k = [y for y in kr if y['d'] == x['d'] and abs(y['s'] - x['s']) < EPS]
            ws2.cell(r, 1, ym)
            tekst(ws2, r, 2, 'samo BANKA (fali u bazi)')
            ws2.cell(r, 3, x['d']).number_format = 'DD.MM.YYYY'
            ws2.cell(r, 4, x['s']).number_format = '#,##0.00'
            tekst(ws2, r, 5, x['opis'][:80])
            tekst(ws2, r, 6, 'DA' if k else 'ne')
            tekst(ws2, r, 7, k[0]['gdje'] if k else '')
            tekst(ws2, r, 8, '—')
            for j in range(1, 9):
                ws2.cell(r, j).fill = LOSE
        for x in samo_baza:
            r += 1
            k = [y for y in kr if y['d'] == x['d'] and abs(y['s'] - x['s']) < EPS]
            ws2.cell(r, 1, ym)
            tekst(ws2, r, 2, 'samo BAZA (visak ili krivi datum)')
            ws2.cell(r, 3, x['d']).number_format = 'DD.MM.YYYY'
            ws2.cell(r, 4, x['s']).number_format = '#,##0.00'
            tekst(ws2, r, 5, x['opis'])
            tekst(ws2, r, 6, 'DA' if k else 'ne')
            tekst(ws2, r, 7, k[0]['gdje'] if k else '')
            tekst(ws2, r, 8, x['id'])
            for j in range(1, 9):
                ws2.cell(r, j).fill = SIDRO

    ws2.auto_filter.ref = 'A3:H' + str(max(r, 4))
    ws2.freeze_panes = 'A4'
    for j, w in enumerate([10, 30, 12, 12, 62, 10, 16, 38], start=1):
        ws2.column_dimensions[chr(64 + j)].width = w

    # ---------------------------------------------------------------- 2023 --
    ws3 = wb.create_sheet('2023')
    txt = [
        ('ZASTO 2023. NE ZATVARA — i zasto to nije greska', True),
        ('', False),
        ('Kokin model i nas model broje isti novac na razlicitom mjestu:', False),
        ('   · kod nje svaka karticna kupovina ODMAH tereti racun', False),
        ('   · kod nas karticna kupovina ne dira racun (pot), a racun tereti '
         'JEDAN redak: skupna naplata banke', False),
        ('', False),
        ('Oba modela daju isti saldo — ako obje strane imaju svoje retke. Njoj '
         'lanac zatvara jer ima svoje. Nama treba jedan redak mjesecno '
         '("TROSKOVI UCINJENI MASTERCARD KARTICOM"), koji dolazi SAMO s izvoda.', False),
        ('', False),
        ('Izmjereno na PROD-u:', True),
        ('   · prvu skupnu MC naplatu ona ima tek 11.12.2023. (926,52)', False),
        ('   · za sijecanj–studeni 2023. tog retka nema NIGDJE: ni kod nje, ni na '
         'izvodu (prvi ZABA izvod koji imamo je 2023-12)', False),
        ('   · njenih karticnih troskova 2023. imamo svih 528 (17.264,41), ali su '
         'u nasem modelu potovi i ne miču saldo', False),
        ('   · razlika koju to proizvodi: 15.752,07', False),
        ('', False),
        ('POSLJEDICA: 2023. nije popravljiva radom — nedostaje IZVOR, ne trud.', True),
        ('Dvije opcije: (a) skinuti ZABA izvode za 2023. iz e-bankarstva, ili '
         '(b) ostaviti kako je. Sidro 844,83 @ 01.01.2024. 2023. ionako zapecati, '
         'pa greska ne moze iscuriti u 2024. i dalje — sto je i dokazano: 2024. '
         'iznad tog sidra zatvara 9/12 mjeseci u cent.', False),
        ('', False),
        ('2023. je uvezena zbog analize i AI sloja, ne zbog salda.', False),
    ]
    for i, (t, b) in enumerate(txt, start=1):
        c = tekst(ws3, i, 1, t)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if b:
            c.font = Font(bold=True)
    ws3.column_dimensions['A'].width = 118

    out = IZLAZ / ('pregled_stanja_' + datetime.now().strftime('%Y%m%d_%H%M') + '.xlsx')
    wb.save(out)
    print('\nzapisano: ' + out.name)
    print('  Pregled : ' + str(len(series) - 1) + ' izvoda, ' + str(len(sporni))
          + ' spornih')
    print('  Sporno  : ' + str(max(r - 3, 0)) + ' redaka')


if __name__ == '__main__':
    main()
