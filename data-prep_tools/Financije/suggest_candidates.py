# -*- coding: utf-8 -*-
"""
suggest_candidates.py  (S107j, 2026-07-22)
==========================================
Rule-authoring pomoćnik: skenira N/A retke Review-a (Tip prazan/N/A) koji IMAJU
tekst (Izvod opis / Napomena), grupira ih po normaliziranom merchant/keyword ključu,
i nudi **top N najčešćih** (default 20 — da ne preplavi) u sheetu `Neklasificirano`
s Tip/Podtip **dropdownima** (isti named-range mehanizam kao Review). Ti popuniš
Tip/Podtip za ponuđene → `--harvest` ih prebaci u `Pravila` sheet → `apply_rules.py`
klasificira → sljedeći `suggest` krug je kraći. Petlja dok tail ne postane premalen.

Fokus po godini (`--year 2026`) — cilj: zatvoriti tekuću godinu pa je poslati u PROD.

Pokretanje (Review zatvoren u Excelu!):
  Financije\\run.bat suggest_candidates.py --year 2026            → napiši Neklasificirano (2026)
  Financije\\run.bat suggest_candidates.py --year 2026 --top 25  → koliko klastera nudi
  Financije\\run.bat suggest_candidates.py --preview --year 2026 → samo ispiši, bez pisanja
  Financije\\run.bat suggest_candidates.py --harvest             → popunjeni redovi → Pravila
"""

import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")
SHEET = 'Neklasificirano'
HDR_FILL = PatternFill('solid', fgColor='C55A11')      # narančasti header (razlika od Review plavog)
WHITE_BOLD = Font(color='FFFFFF', bold=True)
THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# tokeni bez informacije (bankovni boilerplate + reference) — ne ulaze u ključ
BOILER = {
    'KREDITNI', 'TRANSFER', 'NACIONALNI', 'NAC', 'EURIMA', 'ON-LINE', 'ONLINE',
    'BANKARSTVOM', 'MOBILNE', 'APLIKACIJE', 'M-ZABA', 'E-ZABA', 'INOZEMSTVO',
    'INOZEMSTVA', 'EUR', 'HRK', 'COST', 'ELEC', 'GASB', 'OTHR', 'MODEL', 'POZIV',
    'BROJ', 'IZNOS', 'VALUTA', 'PLAĆANJA', 'PLACANJA', 'REFERENCA', 'PRIJENOS',
    'PUTEM', 'USLUGE', 'ZA', 'NA', 'OD', 'DO', 'ILI', 'I', 'U', 'PO', 'SE',
    'HR', 'D.O.O', 'D.O.O.', 'D.D', 'D.D.', 'J.D.O.O', 'ZAGREB', 'RATA', 'RATE',
    'THE', 'COM', 'ON',
}
# hyphen u razredu → 'on-line'/'m-zaba' ostaju cijeli i padnu u BOILER
RE_WORD = re.compile(r"[A-Za-zČĆŠĐŽčćšđž][A-Za-zČĆŠĐŽčćšđž\.\*&/'\-]*")
RE_IBAN = re.compile(r'^HR\d{5,}$', re.I)
RE_RATA = re.compile(r'\bRAT[AE]\b|\b\d{1,2}\s*/\s*\d{1,2}\b', re.I)   # 'RATA 02/12', '1/4'


def to_year(v):
    try:
        return v.year
    except Exception:
        return None


def norm_key(text: str) -> str:
    """Merchant/keyword ključ iz opisa: makni rata-marker/IBAN/ref/brojke i
    boilerplate, uzmi merchant. Ključ = prvi token ako je ≥5 slova (spaja
    'AFRODITA'/'AFRODITA BEAUTY', 'KEINDL'/'KEINDL SPORT'), inače prva 2 tokena.
    Primjeri: STUDENAC1134→STUDENAC, KEINDL SPORT RATA 2/12→KEINDL,
    …transfer… HRVATSKI TELEKOM→HRVATSKI TELEKOM."""
    if not text:
        return ''
    up = RE_RATA.sub(' ', text.upper())
    toks = []
    for w in RE_WORD.findall(up):
        w = w.strip('.').strip('*').strip('/').strip('-')
        w = re.sub(r'\d+$', '', w)            # STUDENAC1134 → STUDENAC
        if len(w) < 3 or RE_IBAN.match(w) or w in BOILER:
            continue
        toks.append(w)
        if len(toks) >= 2:
            break
    if not toks:
        return ''
    if len(toks[0]) >= 5:                      # distinktivan merchant → 1 token
        return toks[0]
    return ' '.join(toks)


def pick_review(args):
    explicit = [a for a in args if not a.startswith('--')]
    if explicit:
        p = Path(explicit[0])
        if not p.exists():
            sys.exit(f'✗ File ne postoji: {p}')
        return p
    cands = sorted([c for c in DATA_DIR.glob('Financije_review_*.xlsx') if '.pre-' not in c.name],
                   key=lambda p: p.stat().st_mtime, reverse=True)
    if not cands:
        sys.exit('✗ Nema Financije_review_*.xlsx')
    return cands[0]


def hdr_index(ws):
    return {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}


def existing_rule_keywords(wb) -> set:
    """Ključevi već pokriveni u Pravila/Preimenovanja — da ih ne nudimo opet."""
    kw = set()
    if 'Pravila' in wb.sheetnames:
        ws = wb['Pravila']
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v:
                kw.add(norm_key(str(v)))
    return {k for k in kw if k}


def collect_candidates(ws, col, year, source):
    """→ dict key → {count, examples:[(opis,iznos,smjer,izvor)]}."""
    groups = defaultdict(lambda: {'count': 0, 'ex': []})
    for r in range(2, ws.max_row + 1):
        tip = str(ws.cell(r, col['Tip']).value or '')
        if tip not in ('', 'N/A'):
            continue
        if year is not None and to_year(ws.cell(r, col['event_date']).value) != year:
            continue
        izv = str(ws.cell(r, col['Izvor']).value or '')
        if source and source.lower() not in izv.lower():
            continue
        opis = str(ws.cell(r, col['Izvod opis']).value or '').strip()
        nap = str(ws.cell(r, col['Napomena']).value or '').strip()
        text = opis or nap
        key = norm_key(text)
        if not key:
            continue
        sm = str(ws.cell(r, col['Smjer']).value or '')
        amt = ws.cell(r, col['Uplata' if sm == 'Uplata' else 'Isplata']).value
        g = groups[key]
        g['count'] += 1
        if len(g['ex']) < 1:
            g['ex'].append((text[:60], amt, sm, izv))
    return groups


def write_sheet(wb, review, ranked, ncols_review):
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    heads = ('Ključne riječi', 'Broj', 'Primjer opisa', 'Primjer iznos', 'Izvor',
             'Tip', 'Podtip', 'Iznos min', 'Iznos max', 'Napomena', 'Komentar')
    for c, h in enumerate(heads, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
    for col, w in zip('ABCDEFGHIJK', (26, 6, 46, 12, 12, 16, 22, 10, 10, 24, 20)):
        ws.column_dimensions[col].width = w
    for i, (key, g) in enumerate(ranked, 2):
        ex = g['ex'][0] if g['ex'] else ('', None, '', '')
        ws.cell(i, 1, key)
        ws.cell(i, 2, g['count'])
        ws.cell(i, 3, ex[0])
        ws.cell(i, 4, ex[1])
        if ex[1] is not None:
            ws.cell(i, 4).number_format = '#,##0.00'
        ws.cell(i, 5, ex[3])
    last = len(ranked) + 1
    # Tip dropdown (F) + zavisni Podtip (G) — isti named-range mehanizam kao Review
    dv_tip = DataValidation(type='list', formula1='TipList', allow_blank=True)
    dv_tip.sqref = f'F2:F{max(2, last)}'
    ws.add_data_validation(dv_tip)
    dv_sub = DataValidation(
        type='list', allow_blank=True,
        formula1='INDIRECT("Tip_"&SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(F2," ","_"),"ć","c"),"/","_"))')
    dv_sub.sqref = f'G2:G{max(2, last)}'
    ws.add_data_validation(dv_sub)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:K{max(2, last)}'
    ws.sheet_view.tabColor = 'C55A11'


def harvest(wb) -> int:
    """Popunjeni Neklasificirano redovi (Tip zadan) → dopiši u Pravila (dedup po
    Ključne riječi). Vraća broj dodanih pravila."""
    if SHEET not in wb.sheetnames:
        sys.exit(f'✗ Nema {SHEET} sheeta — prvo pokreni bez --harvest.')
    ns = wb[SHEET]
    nh = hdr_index(ns)
    pr = wb['Pravila']
    ph = hdr_index(pr)
    existing = {str(pr.cell(r, ph['Ključne riječi']).value or '').strip().lower()
                for r in range(2, pr.max_row + 1) if pr.cell(r, ph['Ključne riječi']).value}
    added = 0
    dst = pr.max_row + 1
    for r in range(2, ns.max_row + 1):
        kw = str(ns.cell(r, nh['Ključne riječi']).value or '').strip()
        tip = str(ns.cell(r, nh['Tip']).value or '').strip()
        if not kw or not tip or tip == 'N/A':
            continue
        if kw.lower() in existing:
            continue
        pr.cell(dst, ph['Ključne riječi'], kw)
        pr.cell(dst, ph['Tip'], tip)
        pr.cell(dst, ph['Podtip'], ns.cell(r, nh['Podtip']).value)
        if 'Napomena' in ph:
            pr.cell(dst, ph['Napomena'], ns.cell(r, nh['Napomena']).value)
        if 'Komentar' in ph:
            pr.cell(dst, ph['Komentar'], ns.cell(r, nh['Komentar']).value)
        if 'Iznos min' in ph:
            pr.cell(dst, ph['Iznos min'], ns.cell(r, nh['Iznos min']).value)
        if 'Iznos max' in ph:
            pr.cell(dst, ph['Iznos max'], ns.cell(r, nh['Iznos max']).value)
        existing.add(kw.lower())
        added += 1
        dst += 1
    return added


def main():
    args = sys.argv[1:]
    preview = '--preview' in args
    do_harvest = '--harvest' in args
    year = None
    if '--year' in args:
        year = int(args[args.index('--year') + 1])
    source = None
    if '--source' in args:
        source = args[args.index('--source') + 1]
    top = 20
    if '--top' in args:
        top = int(args[args.index('--top') + 1])

    review = pick_review([a for a in args if a not in (
        '--preview', '--harvest', '--year', '--source', '--top', str(year), str(top), source or '')])
    print(f'Review: {review.name}'
          f'{f"  [god. {year}]" if year else ""}'
          f'{f"  [izvor {source}]" if source else ""}'
          f'{"  [PREVIEW]" if preview else ""}')

    wb = openpyxl.load_workbook(review)
    ws = wb['Review']
    col = hdr_index(ws)

    if do_harvest:
        n = harvest(wb)
        if preview:
            print(f'[PREVIEW] {n} popunjenih kandidata → Pravila (nije snimljeno).')
            return
        backup = review.with_name(f'{review.stem}.pre-harvest-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
        shutil.copy2(review, backup)
        wb.save(review)
        print(f'✔ {n} pravila dodano u Pravila sheet. Backup: {backup.name}')
        print('  Sljedeći korak: apply_rules.py --dry  (pa bez --dry)')
        return

    groups = collect_candidates(ws, col, year, source)
    covered = existing_rule_keywords(wb)
    ranked = sorted(((k, g) for k, g in groups.items() if k not in covered),
                    key=lambda kv: -kv[1]['count'])[:top]
    total_na = sum(g['count'] for g in groups.values())
    shown = sum(g['count'] for _, g in ranked)
    print(f'\nN/A s tekstom u opsegu: {total_na}  |  top {len(ranked)} klastera pokriva {shown} redaka')
    print(f'{"KLJUČ":26} {"BROJ":>5}  PRIMJER')
    for key, g in ranked:
        ex = g['ex'][0] if g['ex'] else ('', None, '', '')
        print(f'  {key:26} {g["count"]:>5}  {str(ex[3]):10} {str(ex[0])[:42]}')

    if preview:
        print('\n[PREVIEW] Ništa zapisano.')
        return
    write_sheet(wb, review, ranked, ws.max_column)
    backup = review.with_name(f'{review.stem}.pre-neklas-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        sys.exit(f'✗ Zatvori Review u Excelu i ponovi. (Backup: {backup.name})')
    print(f'\n✔ Sheet "{SHEET}" ({len(ranked)} redaka) zapisan. Backup: {backup.name}')
    print('  Popuni Tip/Podtip (dropdown) → suggest_candidates.py --harvest → apply_rules.py')


if __name__ == '__main__':
    main()
