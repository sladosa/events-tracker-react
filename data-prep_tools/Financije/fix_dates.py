# -*- coding: utf-8 -*-
import sys, os
import openpyxl
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
INPUT_PATH  = os.path.join(SCRIPT_DIR, "Financije 2026_2.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "Financije 2026_3.xlsx")

# Rows with confirmed context-based year correction (row_index: (year, note))
MANUAL_FIXES = {
    93: (2023, "TIPFELER '11.07.223' → 2023-07-11 (kontekst: R92=2023-07-08, R95=2023-07-21)"),
    94: (2023, "TIPFELER '11.07.224' → 2023-07-11 (kontekst: R93=2023-07-11, R95=2023-07-21)"),
}

def parse_string_date(s, row_idx):
    """Returns (date_str YYYY-MM-DD, note)"""
    s = s.strip()
    s_clean = s.rstrip('.')
    parts = s_clean.split('.')
    if len(parts) < 3:
        return None, f"Nepoznat format: {s!r}"
    day_s, mon_s, yr_s = parts[0], parts[1], parts[2]
    try:
        day = int(day_s)
        mon = int(mon_s)
        yr_raw = int(yr_s)
    except ValueError:
        return None, f"Nije broj: {s!r}"

    # Manual context override
    if row_idx in MANUAL_FIXES:
        year, note = MANUAL_FIXES[row_idx]
    elif yr_raw < 100:
        year = 2000 + yr_raw
        note = ""
    else:
        # Typo year — fall back to last 2 digits
        year = 2000 + (yr_raw % 100)
        note = f"TIPFELER u godini: '{s}' → pretpostavljam {year}"

    try:
        dt = datetime(year, mon, day)
        return dt.strftime('%Y-%m-%d'), note
    except ValueError as e:
        return None, f"Nevazeci datum {day}.{mon}.{year}: {e}"


def convert_value(val, row_idx):
    """Returns (iso_str, note)"""
    if val is None:
        return '', ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d'), ''
    if isinstance(val, str):
        return parse_string_date(val, row_idx)
    try:
        from datetime import date
        if isinstance(val, date):
            return val.strftime('%Y-%m-%d'), ''
    except Exception:
        pass
    return '', f"Nepoznat tip: {type(val).__name__} = {val!r}"


wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
ws = wb["sasa EU"]

print(f"sasa EU: {ws.max_row} rows, {ws.max_column} cols")

col_date_iso = ws.max_column + 1  # col G
col_note     = ws.max_column + 2  # col H

ws.cell(1, col_date_iso).value = 'Datum_ISO'
ws.cell(1, col_note).value = 'Napomena'

string_count = 0
noted_rows = []

for row_idx in range(2, ws.max_row + 1):
    raw_val = ws.cell(row_idx, 3).value
    iso_str, note = convert_value(raw_val, row_idx)
    ws.cell(row_idx, col_date_iso).value = iso_str
    if note:
        ws.cell(row_idx, col_note).value = note
    if isinstance(raw_val, str) and raw_val.strip():
        string_count += 1
    if note:
        noted_rows.append((row_idx, raw_val, iso_str, note))

print(f"String datumi konvertirani: {string_count}")
if noted_rows:
    print(f"\nRedovi s napomenom ({len(noted_rows)}):")
    for r, v, iso, n in noted_rows:
        print(f"  R{r}: {v!r} → {iso}  [{n}]")
else:
    print("Nema problematicnih redova.")

wb.save(OUTPUT_PATH)
print(f"\nSpremnljeno: {OUTPUT_PATH}")
