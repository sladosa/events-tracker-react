# -*- coding: utf-8 -*-
"""
date_accuracy.py  (S107k, 2026-07-23)
=====================================
Tier A date-sync: Review `event_date` ← točan bankovni datum, za retke koji su
već matchani s izvod-transakcijom na ±1/±2 dana (isti Racun/Izvor/Smjer/Iznos).
Izvod je precizniji od Kokine/Sašine procjene datuma kupovine/knjiženja.

Što radi:
  • matcha SVE izvod tx (Izvodi_transakcije.xlsx) na Review istim algoritmom kao
    consolidate_review (±2 dana, greedy po datumu) — parovi s Δ=0 se ne diraju
  • Δ=±1/±2: event_date ← bankovni datum; ako je `Datum naplate` bio == starom
    event_date (Racun/Cash backfill D1), pomiče se zajedno s njim
  • usput popuni prazan `Izvod opis`/`Izvod file` na matchanim retcima
  • na kraju re-sort Reviewa po event_date (redoslijed ostaje kronološki)

NE dira: retke bez izvod-matcha, parove s Δ=0, MC `Datum naplate` (11. u mjesecu
— nikad nije == event_date pa ga follow-up logika ne hvata).

Tier C (match preko ±2 dana) ide kroz Nematchano_v3 Verdikt=DUP u
consolidate_review.py --harvest, ne ovdje.

Pokretanje (Review ZATVOREN u Excelu!):
  Tools\\venv\\Scripts\\python.exe Financije\\date_accuracy.py --dry   → samo brojke, bez pisanja
  Tools\\venv\\Scripts\\python.exe Financije\\date_accuracy.py         → backup + upis
"""

import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, str(Path(__file__).parent))
from consolidate_review import (  # noqa: E402
    DATA_DIR, hdr_index, load_tx, pick_review, sort_review, to_date,
)

MATCH_DELTAS = (0, 1, -1, 2, -2)


def main():
    args = sys.argv[1:]
    dry = '--dry' in args
    review = pick_review(args)
    print(f'Review: {review.name}{"  [DRY — Review NETAKNUT]" if dry else ""}')

    tx = load_tx()
    print(f'Izvod transakcija: {len(tx)}')
    wb = openpyxl.load_workbook(review)
    ws = wb['Review']
    col = hdr_index(ws)
    ncols = ws.max_column

    # indeks Review redaka: (racun, izvor, smjer, date, iznos) -> [row]
    index = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        d = to_date(ws.cell(r, col['event_date']).value)
        if d is None:
            continue
        rac = str(ws.cell(r, col['Racun']).value or '')
        izv = str(ws.cell(r, col['Izvor']).value or '')
        sm = str(ws.cell(r, col['Smjer']).value or '')
        amt = ws.cell(r, col['Uplata' if sm == 'Uplata' else 'Isplata']).value
        if amt not in (None, ''):
            index[(rac, izv, sm, d, round(float(amt), 2))].append(r)

    # greedy match (identično consolidate_review) + skupljanje shiftova
    used = set()
    delta_hist = Counter()
    shifts = []          # (row, old_date, new_date, tx)
    opis_fill = 0
    for t in sorted(tx, key=lambda t: t['date']):
        row, hit = None, 0
        for dd in MATCH_DELTAS:
            key = (t['racun'], t['izvor'], t['smjer'], t['date'] + timedelta(days=dd), t['iznos'])
            row = next((r for r in index.get(key, []) if r not in used), None)
            if row:
                hit = dd
                break
        if row is None:
            continue
        used.add(row)
        delta_hist[hit] += 1
        if hit != 0:
            old = to_date(ws.cell(row, col['event_date']).value)
            shifts.append((row, old, t['date'], t))
        if ws.cell(row, col['Izvod opis']).value in (None, ''):
            ws.cell(row, col['Izvod opis'], t['opis'][:250])
            ws.cell(row, col['Izvod file'], t['src'])
            opis_fill += 1

    matched = sum(delta_hist.values())
    print(f'\nMatchano: {matched}/{len(tx)}  |  Δ=0 (već točni): {delta_hist[0]}')
    print(f'Za pomak: {len(shifts)}  '
          f'(Δ+1: {delta_hist[1]}, Δ-1: {delta_hist[-1]}, Δ+2: {delta_hist[2]}, Δ-2: {delta_hist[-2]})')
    print(f'Prazan Izvod opis popunjen usput: {opis_fill}')

    # raspodjela po izvoru + godini, za osjećaj
    by_src = Counter(t['pref'] for _, _, _, t in shifts)
    by_year = Counter(t['date'].year for _, _, _, t in shifts)
    if shifts:
        print(f'Pomaci po izvoru: {dict(by_src.most_common())}')
        print(f'Pomaci po godini: {dict(sorted(by_year.items()))}')
        print('\n--- 12 primjera (red | staro → novo | izvor | opis[:45]) ---')
        for row, old, new, t in shifts[:12]:
            print(f'  r{row:<5} {old} → {new}  {t["pref"]:<7} {t["opis"][:45]}')

    # primjena: event_date + Datum naplate follow-up
    naplata_follow = 0
    for row, old, new, t in shifts:
        ws.cell(row, col['event_date'], datetime(new.year, new.month, new.day))
        dn = to_date(ws.cell(row, col['Datum naplate']).value)
        if dn == old:
            ws.cell(row, col['Datum naplate'], datetime(new.year, new.month, new.day))
            naplata_follow += 1
    print(f'\nDatum naplate pomaknut zajedno s event_date (bio == starom): {naplata_follow}')

    if dry:
        print(f'\n✔ [DRY] Ništa nije pisano. Za pravi run makni --dry.')
        return
    if not shifts and not opis_fill:
        print('\n✔ Nema promjena — ništa za snimiti.')
        return
    if shifts:
        sort_review(ws, col, ncols)
        ws.freeze_panes = 'F2'
    backup = review.with_name(f'{review.stem}.pre-dateacc-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        sys.exit(f'✗ Zatvori Review u Excelu i ponovi. (Backup: {backup.name})')
    print(f'\n✔ Snimljeno: {len(shifts)} event_date pomaknuto, {naplata_follow} Datum naplate, '
          f'{opis_fill} Izvod opis. Review re-sortiran.')
    print(f'  Backup: {backup.name}')


if __name__ == '__main__':
    main()
