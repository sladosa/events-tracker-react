"""
apply_ai.py — prijenos AI prijedloga (`Tip_AI`/`Podtip_AI`) u prave `Tip`/`Podtip`.

Mehanizam odluke (odluka S107o): kolona **`AI odluka`** odmah desno od `Podtip_AI`,
dropdown `OK` / `NE` / `?`.

  OK  → prijedlog je točan. `--harvest` ga prepiše u Tip/Podtip i **očisti ćeliju**.
  NE  → prijedlog je kriv, a točan još ne znaš. Ostaje u koloni = otvoreno pitanje.
  ?   → treba razgovor. Ostaje u koloni.
  (prazno) → nepregledano. `--harvest` ga NE dira.

Zato nakon harvesta filter "`AI odluka` nije prazno" pokazuje točno ono što je ostalo
za riješiti — kolona se sama prazni kako posao napreduje (uzor: `Nematchano_v3`, 41→0).

Ako znaš točan Tip/Podtip, **ne treba ti `NE`** — samo upiši ispravno u `Tip`/`Podtip`
(dropdown na J/K već radi). Harvest preskače svaki redak koji više nije `N/A`.

Pokretanje (Python iz Tools/venv, PYTHONUTF8=1):
  python apply_ai.py --report                 # samo brojke; radi i s otvorenim Excelom
  python apply_ai.py --init                   # kreira kolonu "AI odluka" (Excel ZATVOREN)
  python apply_ai.py --harvest --dry          # što bi se preneslo, bez pisanja
  python apply_ai.py --harvest                # pravi upis (Excel ZATVOREN; backup .pre-aiapply-*)

Tvrda pravila:
  • NIKAD ne piše preko postojećeg `Tip` — samo `N/A`/prazno.
  • Par `Tip | Podtip` mora postojati u Taksonomiji, inače se redak preskače uz prijavu.
  • Provenijencija ide u `Labela iz` (`AI:visoka 2026-07-28 10:15`), **ne** u `Pravilo run` —
    ta kolona kod `ai_classify.py --eval` znači "labelu je stavilo keyword pravilo" pa bi
    AI labele ušle u vlastiti eval set (mjerenje modela protiv sebe samog).
"""
from __future__ import annotations

import argparse
import collections
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
DATA = HERE.parent.parent / 'data-prep_data' / 'Financije'

DECISION_COL = 'AI odluka'
AFTER_COL = 'Podtip_AI'          # kolona iza koje se umeće (lijevo od DV/CF na J/K!)
DEC_WIDTH = 11
VALUES = ('OK', 'NE', '?')
UNKNOWN = 'NEPOZNATO'

# Narančasto = "tvoj potez" (plavo = ljudske kolone, ljubičasto = AI kolone).
HDR_FILL = PatternFill('solid', fgColor='ED7D31')
WHITE_BOLD = Font(color='FFFFFF', bold=True)
_THIN = Side(style='thin')
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
OK_FILL = PatternFill('solid', fgColor='C6EFCE')      # zeleno
NO_FILL = PatternFill('solid', fgColor='FFC7CE')      # crveno


# ── Pomoćno ──────────────────────────────────────────────────────────────────

def pick_review() -> Path:
    cands = [p for p in DATA.glob('Financije_review_*.xlsx')
             if '.pre-' not in p.name and not p.name.startswith('~$')]
    if not cands:
        sys.exit(f'✗ Nema Review filea u {DATA}')
    return max(cands, key=lambda p: p.stat().st_mtime)


def header_map(ws) -> dict[str, int]:
    """Ime kolone → 1-based indeks."""
    return {str(c.value).strip(): i
            for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)
            if c.value}


def clean(v) -> str:
    return str(v).strip() if v is not None else ''


def load_taxonomy(wb) -> set[tuple[str, str]]:
    ws = wb['Taksonomija']
    pairs = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        tip = clean(r[0])
        pod = clean(r[1]) if len(r) > 1 else ''
        if tip and tip != 'N/A':
            pairs.add((tip, pod))
    return pairs


def is_na(tip: str) -> bool:
    return (not tip) or tip == 'N/A'


# ── Kolona odluke ────────────────────────────────────────────────────────────

def ensure_decision_column(ws, dry: bool = False) -> tuple[int | None, bool]:
    """Umeće `AI odluka` odmah desno od `Podtip_AI`. Vraća (1-based indeks, je li kreirana).

    Umeće se DESNO od J/K, pa data validation (`TipList`, `INDIRECT`) i conditional
    formatting, koji žive na J2:J*/K2:K*, ostaju gdje jesu — openpyxl ih ne pomiče.
    `insert_cols` pomiče ćelije, ali NE i `column_dimensions` (širine/outline), pa se
    one prenose ručno (ista zamka kao `ai_classify.ensure_ai_columns`).
    """
    H = header_map(ws)
    if DECISION_COL in H:
        print(f'✔ Kolona "{DECISION_COL}" već postoji ({get_column_letter(H[DECISION_COL])}) '
              f'— ništa za raditi.')
        return H[DECISION_COL], False
    if AFTER_COL not in H:
        sys.exit(f'✗ Review nema kolonu "{AFTER_COL}" — pokreni prvo ai_classify.py --run.')
    at = H[AFTER_COL] + 1
    if dry:
        print(f'… (dry) kolona "{DECISION_COL}" bi se kreirala na '
              f'{get_column_letter(at)} (desno od {AFTER_COL})')
        return None, False

    snap = {}
    for i in range(1, ws.max_column + 1):
        L = get_column_letter(i)
        if L in ws.column_dimensions:
            d = ws.column_dimensions[L]
            snap[i] = (d.width, d.outlineLevel, d.hidden)

    ws.insert_cols(at, 1)

    for L in list(ws.column_dimensions):
        del ws.column_dimensions[L]
    for i, (w, lvl, hid) in snap.items():
        d = ws.column_dimensions[get_column_letter(i if i < at else i + 1)]
        d.width, d.outlineLevel, d.hidden = w, lvl, hid

    cell = ws.cell(1, at, DECISION_COL)
    cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
    ws.column_dimensions[get_column_letter(at)].width = DEC_WIDTH

    L = get_column_letter(at)
    rng = f'{L}2:{L}{ws.max_row}'
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type='list', formula1=f'"{",".join(VALUES)}"', allow_blank=True,
        errorTitle='Nepoznata odluka',            # ≤32 znaka
        error='Dopušteno: OK, NE ili ?')          # ≤255 znakova
    ws.add_data_validation(dv)
    dv.add(rng)
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"OK"'], fill=OK_FILL))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"NE"'], fill=NO_FILL))

    # Sve što nosi status mora biti UNUTAR autofiltera, inače se pri sortu raspari od reda.
    if ws.auto_filter.ref:
        first = ws.auto_filter.ref.split(':')[0]
        ws.auto_filter.ref = f'{first}:{get_column_letter(ws.max_column)}{ws.max_row}'

    _check_dv_alignment(ws)
    print(f'✔ Kreirana kolona "{DECISION_COL}" na {L} (dropdown {"/".join(VALUES)}, '
          f'OK zeleno / NE crveno, unutar autofiltera)')
    return at, True


def _check_dv_alignment(ws) -> None:
    """Dropdowni Tip/Podtip moraju i dalje sjediti na kolonama Tip/Podtip."""
    import re
    H = header_map(ws)
    want = {get_column_letter(H['Tip']): 'Tip', get_column_letter(H['Podtip']): 'Podtip'}
    for dv in ws.data_validations.dataValidation:
        if DECISION_COL in str(dv.formula1) or set(VALUES) <= set(str(dv.formula1)):
            continue
        for rng in dv.sqref.ranges:
            col = re.match(r'([A-Z]+)', str(rng)).group(1)
            if col not in want and str(dv.formula1).startswith(('=TipList', 'INDIRECT')):
                print(f'⚠ Data validation ({str(dv.formula1)[:40]}) sjedi na koloni {col}, '
                      f'a to više nije Tip/Podtip — provjeri raspored kolona!')


# ── Report ───────────────────────────────────────────────────────────────────

def report(wb) -> None:
    ws = wb['Review']
    H = header_map(ws)
    if 'Tip_AI' not in H:
        sys.exit('✗ Nema AI kolona — pokreni prvo ai_classify.py --run.')
    tax = load_taxonomy(wb)
    has_dec = DECISION_COL in H

    band = collections.Counter()
    dec = collections.Counter()
    pairs = collections.defaultdict(collections.Counter)
    for r in ws.iter_rows(min_row=2, values_only=True):
        conf = clean(r[H['Pouzdanost_AI'] - 1])
        if not clean(r[H['AI run'] - 1]):
            continue
        band[conf] += 1
        if has_dec:
            dec[clean(r[H[DECISION_COL] - 1]).upper() or '(prazno)'] += 1
        tip, pod = clean(r[H['Tip_AI'] - 1]), clean(r[H['Podtip_AI'] - 1])
        if tip and tip != UNKNOWN:
            pairs[conf][(tip, pod)] += 1

    print(f'\nAI prijedlozi: {sum(band.values())} redaka')
    for c in ('visoka', 'srednja', 'niska'):
        n = band.get(c, 0)
        print(f'  {c:<9}{n:>6}   ({len(pairs[c])} različitih Tip|Podtip parova)')
    if has_dec:
        print(f'\nKolona "{DECISION_COL}": ' +
              ' · '.join(f'{k} {v}' for k, v in sorted(dec.items())))
    else:
        print(f'\nKolona "{DECISION_COL}" još ne postoji — pokreni --init.')

    print('\nTraka "visoka" po paru (to je jedinica pregleda, ne redak):')
    print('-' * 62)
    for (tip, pod), n in pairs['visoka'].most_common():
        flag = '' if (tip, pod) in tax else '   ⚠ para NEMA u Taksonomiji'
        print(f'{n:>5}×  {tip} | {pod}{flag}')


# ── Harvest ──────────────────────────────────────────────────────────────────

def harvest(wb, dry: bool) -> tuple[int, list[str]]:
    ws = wb['Review']
    H = header_map(ws)
    for c in ('Tip', 'Podtip', 'Tip_AI', 'Podtip_AI', 'Pouzdanost_AI', 'Labela iz', DECISION_COL):
        if c not in H:
            sys.exit(f'✗ Review nema kolonu "{c}".')
    tax = load_taxonomy(wb)
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M')

    applied, skipped, by_pair = 0, [], collections.Counter()
    for row in range(2, ws.max_row + 1):
        if clean(ws.cell(row, H[DECISION_COL]).value).upper() != 'OK':
            continue
        tip_now = clean(ws.cell(row, H['Tip']).value)
        tip = clean(ws.cell(row, H['Tip_AI']).value)
        pod = clean(ws.cell(row, H['Podtip_AI']).value)
        conf = clean(ws.cell(row, H['Pouzdanost_AI']).value) or '?'

        if not is_na(tip_now):
            skipped.append(f'red {row}: Tip je već "{tip_now}" — OK ignoriran (ručna labela pobjeđuje)')
            continue
        if not tip or tip == UNKNOWN:
            skipped.append(f'red {row}: Tip_AI je "{tip or "prazno"}" — nema što prenijeti')
            continue
        if (tip, pod) not in tax:
            skipped.append(f'red {row}: par "{tip} | {pod}" NIJE u Taksonomiji')
            continue

        by_pair[(tip, pod)] += 1
        applied += 1
        if not dry:
            ws.cell(row, H['Tip']).value = tip
            ws.cell(row, H['Podtip']).value = pod or None
            ws.cell(row, H['Labela iz']).value = f'AI:{conf} {stamp}'
            ws.cell(row, H[DECISION_COL]).value = None      # self-clearing

    print(f'\n{"(dry) " if dry else ""}Preneseno u Tip/Podtip: {applied}')
    for (tip, pod), n in by_pair.most_common():
        print(f'  {n:>5}×  {tip} | {pod}')
    if skipped:
        print(f'\n⚠ Preskočeno {len(skipped)}:')
        for s in skipped[:30]:
            print(f'  {s}')
        if len(skipped) > 30:
            print(f'  … i još {len(skipped) - 30}')
    return applied, skipped


# ── main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--init', action='store_true', help=f'kreiraj kolonu "{DECISION_COL}"')
    ap.add_argument('--harvest', action='store_true', help='primijeni OK retke u Tip/Podtip')
    ap.add_argument('--report', action='store_true', help='samo brojke, ništa se ne piše')
    ap.add_argument('--dry', action='store_true', help='pokaži što bi se dogodilo, bez pisanja')
    ap.add_argument('--fix-freeze', action='store_true',
                    help='vrati freeze_panes na F2 (sad je F4855 — zamrznuto 4854 redaka)')
    args = ap.parse_args()
    if not (args.init or args.harvest or args.report or args.fix_freeze):
        args.report = True

    review = pick_review()
    print(f'Review: {review.name}')
    wb = openpyxl.load_workbook(review)
    ws = wb['Review']

    if args.report:
        report(wb)
        return

    changed = False
    if args.init:
        _, created = ensure_decision_column(ws, dry=args.dry)
        changed = changed or created
    if args.fix_freeze:
        print(f'freeze_panes: {ws.freeze_panes} → F2')
        if not args.dry:
            ws.freeze_panes = 'F2'
            changed = True
    if args.harvest:
        applied, _ = harvest(wb, dry=args.dry)
        changed = changed or bool(applied)

    if args.dry:
        print('\n… --dry: ništa nije zapisano.')
        return
    if not changed:
        print('\n✔ Nema promjena — Review netaknut.')
        return

    backup = review.with_name(f'{review.stem}.pre-aiapply-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        backup.unlink(missing_ok=True)
        sys.exit('✗ Zatvori Review u Excelu i ponovi. (Ništa nije zapisano.)')
    print(f'\n✔ Snimljeno. Backup: {backup.name}')


if __name__ == '__main__':
    main()
