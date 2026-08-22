"""Popuni Tip/Podtip na retcima koje je uvoz upravo stvorio (tranša 3).

Jednokratna skripta, u duhu `fix_vocarna_pravilo.py` / `fix_anja_rate.py`:
`apply_rules.py` radi nad Review workbookom, a ovo je app-ov **izvještaj o
uvozu** — radni file s `row_hash`om, koji se uvozi natrag.

⚠ PODJELA AUTORITETA. Mapiranje NIJE izmišljeno: svaki par je izvučen iz
`Financije_review_*.xlsx` prebrojavanjem kako je isti Kokin tekst klasificiran
u povijesti (T-com 40/41, Parking 118/118, Zoran povrat 41/41...). Gdje povijest
nije jednoglasna ili retka nema, odluku je dao čovjek — ne skripta.

⚠ Parovi se PROVJERAVAJU protiv `DropdownData` lista app-ovog exporta, ne protiv
popisa u ovom fileu. Podtip koji ne postoji u `validation_rules` uvezao bi se kao
običan tekst i nigdje ne bi javio grešku — dropdown bi ga poslije odbio, a
podatak bi već bio u bazi.
"""
from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')   # sys.exit poruke idu ovuda

NA = 'N/A'

# ── Mapiranje po Kokinom tekstu (`leaf comment`) ─────────────────────────────
# Ključ se uspoređuje malim slovima, po točnom pogotku ili prefiksu.
PO_OPISU: list[tuple[str, str, str, str]] = [
    # (prefiks opisa,      Tip,           Podtip,                                 dokaz)
    ('parking',            'Prijevoz',    'Taksi, Zet, Parking',                  '118/118'),
    ('t-com',              'Informatika', 'Komunikacije_T-com (internet, MaxTv)', '40/41'),
    ('t-mobile',           'Informatika', 'Komunikacije_T-mobile',                '41/42'),
    ('saša holding',       'Kuća',        'Holding (smeće)',                      '39/39'),
    ('nataša holding',     'Kuća',        'Holding (smeće)',                      '41/41'),
    ('nataša povrat',      'Transfer',    'Natasa',                               '5/5'),
    ('nataša popvrat',     'Transfer',    'Natasa',                               '5/5 (tipfeler)'),
    ('zoran povrat',       'Kuća',        'Povrat Zoran',                         '41/41'),
    ('povrat poreza',      'Porezi',      'porez/prirez/dohodak',                 '4/5'),
    ('pp saša',            'Zdravlje',    'PP (Posmrtna pripomoc)',               '5/5'),
    ('pp koka',            'Zdravlje',    'PP (Posmrtna pripomoc)',               '5/5'),
    ('hlk',                'Zdravlje',    'Lječnička komora_Koka',                '16/18'),
    ('anja 84/96',         'Prihodi',     'Povrat Anja',                          '41/41, niz 83/96 -> 84/96'),
    ('anja povrat',        'Transfer',    'Anja',                                 '19/19 za ne-ratne'),
    ('nena',               'Prihodi',     'Koka',                                 'Sašina odluka'),
    ('mall.hr povrat',     'Prihodi',     'Koka',                                 'Sašina odluka'),
]

# ── Retci koje banka vodi drukčije nego ona ──────────────────────────────────
# ⚠ Ona vodi JEDAN redak `Parking 1,40`, banka DVA naloga po `0,70`. Sparivanje
#   po (iznos, datum) ih zato ne nađe i ostaju sa strojnim tekstom izvoda
#   (`Kreditni transfer nacionalni u eurima...`), koji u povijesti vodi na
#   `Domaćinstvo / Bankovni troškovi` — dakle u KRIVI razred, i to uvjerljivo.
#   Popravlja se i opis, ne samo klasifikacija: file mora ostati sam sebi
#   dosljedan da bi se sutra dao pročitati.
PAR_PARKING = ('Prijevoz', 'Taksi, Zet, Parking')
PAROVI_PARKING: list[tuple[date, float]] = [
    (date(2026, 7, 13), 0.70),
    (date(2026, 7, 27), 0.70),
    (date(2026, 7, 30), 0.70),
]

# ── Retci bez ijednog opisa ──────────────────────────────────────────────────
# Skupna naplata kartice nema Kokin redak (nije njena stavka nego bankina) i u
# bazi stoji bez komentara, pa je mapiranje po opisu ne može uhvatiti.
# ⚠ Ona je „zrcalno pravilo" u čistom obliku: ULAZI u saldo (novac je stvarno
#   otišao s računa), IZLAZI iz razreza po Tipu (razrez nose pojedinačne
#   kartične stavke). Zato `Transfer` — 31/31 u povijesti, neprekinut mjesečni
#   niz 11.01.–11.06.2026., a ovaj redak je 11.07.
PO_IZNOSU: list[tuple[date, float, str, str, str]] = [
    (date(2026, 7, 11), 1244.74, 'Transfer', 'izmedju racuna',
     '31/31, mjesečni niz do 11.06.2026.'),
]


def ucitaj_taksonomiju(ref: Path) -> dict[str, set[str]]:
    """Dozvoljeni parovi iz `DropdownData` app-ovog exporta.

    Zaglavlje kolone je `tip=<Tip>`, ispod njega idu podtipovi tog Tipa — to je
    ista mapa koju Excel koristi za `depends_on` dropdown.
    """
    wb = openpyxl.load_workbook(ref, data_only=True, read_only=True)
    if 'DropdownData' not in wb.sheetnames:
        sys.exit(f'✗ {ref.name} nema DropdownData list — treba app-ov export, ne izvještaj.')
    ws = wb['DropdownData']
    rows = list(ws.iter_rows(values_only=True))
    tax: dict[str, set[str]] = {}
    for c, hdr in enumerate(rows[0]):
        if not isinstance(hdr, str) or not hdr.startswith('tip='):
            continue
        tip = hdr[4:].strip()
        tax[tip] = {str(r[c]).strip() for r in rows[1:]
                    if c < len(r) and r[c] not in (None, 'None', '')}
    return tax


def nadi_zaglavlje(ws) -> tuple[int, dict[str, int]]:
    for r in range(1, 60):
        if str(ws.cell(r, 1).value or '').strip() == 'event_id':
            return r, {str(ws.cell(r, c).value or '').strip(): c
                       for c in range(1, ws.max_column + 1)}
    sys.exit('✗ Nije nađen redak zaglavlja (`event_id` u koloni A).')


def makni_biljeske(wb) -> list[str]:
    """⚠ openpyxl sprema bilješku s APSOLUTNOM putanjom u relacijama; exceljs je
    ne nađe i cijeli file postane neuvoziv (S113). Tekst se ispisuje da se ne
    izgubi bez traga."""
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    out.append(f'{ws.title}!{cell.coordinate}: {cell.comment.text.strip()}')
                    cell.comment = None
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('target', type=Path, help='import_report_*.xlsx iz aplikacije')
    ap.add_argument('--taksonomija', type=Path, required=True,
                    help='app-ov export s DropdownData listom (npr. transa3.xlsx)')
    ap.add_argument('--dry', action='store_true', help='samo ispiši što bi upisao')
    a = ap.parse_args()

    tax = ucitaj_taksonomiju(a.taksonomija)
    print(f'Taksonomija: {len(tax)} Tipova, {sum(len(v) for v in tax.values())} podtipova '
          f'({a.taksonomija.name})')

    # Provjera mapiranja PRIJE otvaranja targeta: krivi par se ne smije ni
    # pokušati upisati.
    # ⚠ Provjeravaju se SVI parovi koje skripta zna upisati, ne samo tablica po
    #   opisu: par upisan mimo `validation_rules` ne javlja grešku pri uvozu,
    #   nego tek kad ga netko poslije otvori u dropdownu — a tada je već u bazi.
    svi = ([(t, p) for _, t, p, _ in PO_OPISU]
           + [(t, p) for _, _, t, p, _ in PO_IZNOSU]
           + [PAR_PARKING])
    lose = [(t, p) for t, p in svi if p not in tax.get(t, set())]
    if lose:
        for t, p in lose:
            print(f'✗ Par ne postoji u taksonomiji: {t} / {p}')
        sys.exit(1)
    print(f'✓ Svih {len(svi)} parova postoji u taksonomiji.')

    wb = openpyxl.load_workbook(a.target)
    if 'Events' not in wb.sheetnames:
        sys.exit('✗ File nema `Events` list.')
    ws = wb['Events']
    hdr_row, col = nadi_zaglavlje(ws)

    def c(name: str) -> int:
        if name not in col:
            sys.exit(f'✗ Nema kolone {name!r}. Ima: {sorted(k for k in col if k)}')
        return col[name]

    c_date, c_kom = c('event_date'), c('leaf comment')
    c_tip, c_pod = c('Tip (Transakcija)'), c('Podtip (Transakcija)')
    c_isp = c('Isplata (Transakcija)')

    promjene, preskoceno, bez_pravila = [], [], []
    for r in range(hdr_row + 1, ws.max_row + 1):
        d = ws.cell(r, c_date).value
        if not isinstance(d, datetime):
            continue
        dd, kom = d.date(), str(ws.cell(r, c_kom).value or '').strip()
        isp = ws.cell(r, c_isp).value
        tip_sad = str(ws.cell(r, c_tip).value or '').strip()

        # ⚠ Redak koji je već klasificiran se NE dira. Kriterij je isti kao u
        #   `apply_rules.py`: prazan ili N/A. Prepisivanje valjane vrijednosti
        #   bi značilo da skripta zna bolje od čovjeka koji ju je upisao.
        if tip_sad and tip_sad != NA:
            preskoceno.append((r, kom, tip_sad))
            continue

        novi_kom = None
        po_iznosu = next(((t, p, dk) for pd, pi, t, p, dk in PO_IZNOSU
                          if dd == pd and isinstance(isp, (int, float))
                          and abs(float(isp) - pi) < 0.005), None)
        if po_iznosu:
            tip, pod, dokaz = po_iznosu
        elif any(dd == pd and isinstance(isp, (int, float)) and abs(float(isp) - pi) < 0.005
                 for pd, pi in PAROVI_PARKING):
            novi_kom = 'Parking'
            tip, pod = PAR_PARKING
            dokaz = 'pola njenog retka 1,40'
        else:
            pogodak = next(((t, p, dk) for pre, t, p, dk in PO_OPISU
                            if kom.lower().startswith(pre)), None)
            if not pogodak:
                bez_pravila.append((r, dd, kom))
                continue
            tip, pod, dokaz = pogodak

        promjene.append((r, dd, novi_kom or kom, tip, pod, dokaz, novi_kom))

    print(f'\nZa upis: {len(promjene)} redaka  ·  '
          f'već klasificirano: {len(preskoceno)}  ·  bez pravila: {len(bez_pravila)}')
    for r, dd, kom, tip, pod, dokaz, novi in promjene:
        opis = f'{kom} (opis prepisan)' if novi else kom
        print(f'   r{r:<4} {dd}  {opis:<30} → {tip} / {pod:<38} [{dokaz}]')
    for r, dd, kom in bez_pravila:
        print(f'   ⚠ r{r:<4} {dd}  {kom[:34]:<34} — nema pravila, ostaje N/A')

    if a.dry:
        print('\n--dry: ništa nije zapisano.')
        return

    for r, _, _, tip, pod, _, novi in promjene:
        ws.cell(r, c_tip).value = tip
        ws.cell(r, c_pod).value = pod
        if novi:
            ws.cell(r, c_kom).value = novi

    for b in makni_biljeske(wb):
        print(f'  (bilješka maknuta radi uvoza: {b})')

    out = a.target.with_name(a.target.stem + '_klasificirano.xlsx')
    wb.save(out)
    print(f'\n✓ {len(promjene)} redaka klasificirano')
    print(f'✓ {out}')
    print('  Original je netaknut. Otvori novi file, pregledaj, pa uvezi.')


if __name__ == '__main__':
    main()
