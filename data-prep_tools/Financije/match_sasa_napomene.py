"""
match_sasa_napomene.py
======================
Matchira 'Sto' opise iz source sheeta 'Za Sasu' s export datotekom
po kljucu (datum, iznos_isplata), upisuje rezultat u col R (18)
export datoteke za rucni pregled.

Source:  Financije 2026-06.xlsx   sheet 'Za Sasu'
           A=Datum  B=Nacin placanja  C=Sto  D=EU(iznos)
Export:  events_export_20260607_2.xlsx  sheet 'Events'
           D=event_date  J=Uplata  K=Isplata  N=Napomena (existing)

Output:  events_export_matched.xlsx   (original nije dotaknut)
           col R = matched 'Sto' za pregled
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# ── Paths ──────────────────────────────────────────────────────────────────────
DATA_DIR    = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")
SOURCE_FILE = DATA_DIR / "Financije 2026-06.xlsx"
EXPORT_FILE = DATA_DIR / "events_export_20260607_2.xlsx"
OUTPUT_FILE = DATA_DIR / "events_export_matched.xlsx"

# Export column indices (1-based)
EXP_HDR_ROW  = 16
EXP_DATE_COL =  4   # D: event_date
EXP_NAP_COL  = 14   # N: Napomena (existing, read-only here)
EXP_UPL_COL  = 10   # J: Uplata
EXP_ISP_COL  = 11   # K: Isplata
EXP_OUT_COL  = 18   # R: output — Sto (Za Sasu) za pregled

# Source sheet (index 4 = 'Za Sasu')
SRC_SHEET_IDX = 4
SRC_HDR_ROW   = 1
SRC_DATE_COL  = 1   # A: Datum
SRC_STO_COL   = 3   # C: Sto
SRC_EUR_COL   = 4   # D: EU (iznos isplate)

# Colours
FILL_ORANGE = PatternFill("solid", fgColor="FFC000")   # duplikat match
FILL_GREEN  = PatternFill("solid", fgColor="E2EFDA")   # jedinstven match


# ── Helpers ────────────────────────────────────────────────────────────────────
def to_date_str(val):
    """Normalizira razne formate datuma na 'YYYY-MM-DD'."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s[:10] if len(s) >= 10 else None


def to_amount(val):
    """Zaokruzuje iznos na 2 decimale, vraca None za prazno/0."""
    if val is None or val == "":
        return None
    try:
        f = float(val)
        return round(f, 2) if f != 0 else None
    except (ValueError, TypeError):
        return None


# ── 1. Ucitaj source: gradi (datum, iznos) -> [Sto1, Sto2, ...] ───────────────
print("Citam source sheet 'Za Sasu'...")
src_wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
src_ws = src_wb.worksheets[SRC_SHEET_IDX]

# key: (date_str, amount) -> lista 'Sto' opisa
src_map = defaultdict(list)
src_loaded = 0

for r_idx, row in enumerate(
    src_ws.iter_rows(min_row=SRC_HDR_ROW + 1, values_only=True), start=SRC_HDR_ROW + 1
):
    datum = to_date_str(row[SRC_DATE_COL - 1])
    sto   = row[SRC_STO_COL - 1]
    eur   = to_amount(row[SRC_EUR_COL - 1])

    if not datum or not eur:
        continue

    sto_str = str(sto).strip() if sto else "(prazno)"
    src_map[(datum, eur)].append(sto_str)
    src_loaded += 1

print(f"  Ucitano source redova: {src_loaded}")
print(f"  Unikatnih kljuceva (datum+iznos): {len(src_map)}")

# Provjera: duplikati u sourceu (isti datum+iznos vise puta)
dup_src = {k: v for k, v in src_map.items() if len(v) > 1}
if dup_src:
    print(f"  Source duplikati (isti datum+iznos): {len(dup_src)}")
    for k, vs in list(dup_src.items())[:5]:
        print(f"    {k[0]} {k[1]:>8.2f} -> {vs}")


# ── 2. Procesiraj export datoteku ─────────────────────────────────────────────
print("\nProcesiram export datoteku...")
exp_wb = openpyxl.load_workbook(EXPORT_FILE, data_only=True)
exp_ws = exp_wb.worksheets[0]

# Postavi header za col R
hdr_cell = exp_ws.cell(EXP_HDR_ROW, EXP_OUT_COL)
hdr_cell.value = "Sto (Za Sasu)"
hdr_cell.font  = Font(bold=True)

matched   = 0
duplicate = 0
unmatched = 0

for r in range(EXP_HDR_ROW + 1, exp_ws.max_row + 1):
    date_val = exp_ws.cell(r, EXP_DATE_COL).value
    isp_val  = exp_ws.cell(r, EXP_ISP_COL).value
    upl_val  = exp_ws.cell(r, EXP_UPL_COL).value

    if date_val is None:
        continue   # prazni redovi na kraju

    date_str = to_date_str(date_val)
    isp_amt  = to_amount(isp_val)
    upl_amt  = to_amount(upl_val)

    out_cell = exp_ws.cell(r, EXP_OUT_COL)

    # Pokusaj match: Isplata prvi, pa Uplata ako nema
    candidates = []
    if isp_amt:
        candidates = src_map.get((date_str, isp_amt), [])
    if not candidates and upl_amt:
        candidates = src_map.get((date_str, upl_amt), [])

    if not candidates:
        unmatched += 1
        continue

    if len(candidates) == 1:
        out_cell.value = candidates[0]
        out_cell.fill  = FILL_GREEN
        matched += 1
    else:
        # Vise mogucih matcheva — oznaci za rucnu provjeru
        out_cell.value = "⚠ DUPLIKAT: " + " | ".join(candidates)
        out_cell.fill  = FILL_ORANGE
        duplicate += 1


# ── 3. Spremi output ───────────────────────────────────────────────────────────
exp_wb.save(OUTPUT_FILE)
print(f"\n=== REZULTAT ===")
print(f"  Matchiranih:       {matched}")
print(f"  Duplikat matcheva: {duplicate}  (narancasto - provjeri rucno)")
print(f"  Bez matcha:        {unmatched}")
print(f"\nSpremieno: {OUTPUT_FILE.name}")
print("Col R = 'Sto (Za Sasu)' za pregled — kopirati rucno u col N (Napomena) gdje zelite")
