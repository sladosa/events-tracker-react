# -*- coding: utf-8 -*-
"""
fix_lazne_rate.py — HLK/APN mjesec/godina krivo prepoznat kao rata. S107t, 2026-08-01.

NALAZ: `normalize_financije.py` parsira `X/N` iz Napomene kao "rata X od N"
(v. FINANCIJE_MIGRACIJA.md §6.3). Kokine mjesečne obveze pisane kao **mjesec/godina**
padaju pod isti obrazac:

  'HLK 3/26'        2026-03-06  → Rate?=DA, Broj rata=26   (ožujak 2026, ne rata 3 od 26)
  'APN porez 1/25'  2025-02-19  → Rate?=DA, Broj rata=25
  'HLK članarina 01-02/24' 2024-02-08 → Rate?=DA, Broj rata=24

To nisu obročne kupovine nego godišnja članarina Liječničke komore i APN porez, plaćani
mjesečno. Ostave li se, generator budućih `Planiran` rata (`rataAutomation.ts`) jednog dana
pokuša izvesti 25-ratni plan iz članarine.

Ista klasa koju je S107s već izmjerio u obrnutom smjeru ("goli `n/m` daje 31 lažni pozitiv,
datumi `03/23`") — ovih 32 je već ušlo u podatke.

DETEKCIJA (samovalidirajuća, ne popis brojeva redaka):
  Rate?=DA  ∧  2000 + `Broj rata` == godina event_date-a  ∧  Napomena ima `n/N` s n ≤ 12
Razdvajanje je čisto: nijedan pravi plan (N = 2,3,4,5,6,10,12,48,60,96) ne pogađa uvjet,
a svih 32 pogođenih nose 'HLK' ili 'APN' u Napomeni (dodatni guard ispod).

⚠ Zašto skripta, a ne pravilo: `apply_rules.py` dira `Tip`/`Podtip`, a ovdje se ispravljaju
`Rate?`/`Broj rata` — pravila ta polja uopće ne pišu.

Ne dira: `Tip`/`Podtip` (klasifikacija je točna — Zdravlje/Lječnička komora_Koka,
Porezi/porez-prirez-dohodak), iznose, datume, `Napomena` (tekst `3/26` ostaje kao trag).

Pokretanje (Excel ZATVOREN):
  python fix_lazne_rate.py --dry
  python fix_lazne_rate.py              # backup .pre-lazne-rate-*
"""
from __future__ import annotations

import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from apply_rules import find_header_col, pick_file  # noqa: E402

EXPECTED = 32          # izmjereno 2026-08-01; odstupanje = PREKID
GUARD_TOKENS = ('HLK', 'APN')   # svaki pogođeni redak mora nositi jedan od njih

NM = re.compile(r'(?<![\d/])(\d{1,3})\s*/\s*(\d{1,3})(?![\d/])')


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    path = pick_file(args)
    print(f'File: {path.name}{"  [DRY RUN — bez snimanja]" if dry else ""}\n')

    wb = openpyxl.load_workbook(path)
    ws = wb['Review']
    c_rate = find_header_col(ws, 'Rate?')
    c_broj = find_header_col(ws, 'Broj rata')
    c_date = find_header_col(ws, 'event_date')
    c_nap = find_header_col(ws, 'Napomena')
    c_key = find_header_col(ws, 'source_key')
    c_alt = find_header_col(ws, 'Alternativa / nap.')
    c_run = find_header_col(ws, 'Pravilo run')

    hits, suspect = [], []
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, c_rate).value or '').strip().upper() != 'DA':
            continue
        n_raw = ws.cell(r, c_broj).value
        ed = ws.cell(r, c_date).value
        if n_raw is None or not isinstance(ed, datetime):
            continue
        try:
            n_tot = int(n_raw)
        except (TypeError, ValueError):
            continue
        if 2000 + n_tot != ed.year:
            continue                      # pravi plan — broj rata nije godina
        nap = str(ws.cell(r, c_nap).value or '')
        months = [m for m in NM.finditer(nap)
                  if int(m.group(2)) == n_tot and int(m.group(1)) <= 12]
        (hits if months else suspect).append((r, nap, n_tot, ed))

    # ── Guardovi: sve mora izgledati točno kako je izmjereno ──────────────
    if suspect:
        for r, nap, n_tot, ed in suspect:
            print(f'  ?? red {r}: Broj rata={n_tot} == godina {ed.year}, '
                  f'ali Napomena "{nap[:40]}" nema n/N s n ≤ 12')
        sys.exit(f'✗ {len(suspect)} redaka pogađa godinu ali ne obrazac mjeseca — '
                 f'PREKID, ništa nije upisano. Provjeri ručno.')

    bad = [(r, nap) for r, nap, _, _ in hits
           if not any(t in nap.upper() for t in GUARD_TOKENS)]
    if bad:
        for r, nap in bad:
            print(f'  ?? red {r}: "{nap[:50]}" ne sadrži {" ni ".join(GUARD_TOKENS)}')
        sys.exit('✗ Pogođen redak izvan poznatog obrasca (HLK/APN) — PREKID.')

    if len(hits) != EXPECTED:
        sys.exit(f'✗ Nađeno {len(hits)} redaka, očekivano {EXPECTED} — '
                 f'PREKID. Je li ovo pravi Review file / je li netko već pokrenuo fix?')

    stamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    for r, nap, n_tot, ed in hits:
        key = str(ws.cell(r, c_key).value or '')
        print(f'{"bi se ocistio" if dry else "ocisceno"} red {r:>5} '
              f'({ed:%Y-%m-%d}, Broj rata={n_tot:>2}, "{nap[:26]}", key={key})')
        if dry:
            continue
        # ⚠ ws.cell(r, c, None) NE brise sadrzaj (zamka iz S107e) — mora .value
        ws.cell(r, c_rate).value = None
        ws.cell(r, c_broj).value = None
        old_alt = str(ws.cell(r, c_alt).value or '').strip()
        mark = f'S107t fix: nije rata (mjesec/godina {n_tot}), bio Rate?=DA/Broj rata={n_tot}'
        ws.cell(r, c_alt).value = f'{old_alt} | {mark}' if old_alt else mark
        ws.cell(r, c_run).value = stamp

    print(f'\n{"Bi se ocistilo" if dry else "Ocisceno"}: {len(hits)} redova')
    if dry:
        print('DRY RUN — nista nije snimljeno.')
        return

    backup = path.with_name(f'{path.stem}.pre-lazne-rate-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori {path.name} u Excelu i ponovi.\n'
                 f'  (Backup je vec napravljen: {backup.name})')
    print(f'✔ Snimljeno. Backup: {backup.name}')
    print(f'  Kontrola: Rate?=DA mora pasti 661 → {661 - len(hits)}; '
          f'filtriraj Pravilo run = {stamp} → {len(hits)} redova')


if __name__ == '__main__':
    main()
