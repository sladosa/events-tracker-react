# -*- coding: utf-8 -*-
"""
fix_keks_trener.py — 20 KEKS Pay uplata treneru → `Zdravlje | Sport_Sasa`. S107x, 2026-08-12.

NALAZ (Saša, pri pregledu `saldo_model_nalazi.xlsx`): među `TRANSFER-BEZ-PARA` retcima
puno je napomena `KEKS` s iznosom −20,00 € — to je osobni trener („Ašo"/„Aša").

Potvrda u podacima:
  · `Ašo/Aša` — 27 redaka, svi −20,00, svi VEĆ `Zdravlje | Sport_Sasa` (izravna uplata)
  · `KEKS` −20,00 — 22 retka, od toga 20 u `Transfer | izmedju racuna`
  · tih 20 ide 21.07.2023. → 14.03.2024., kadenca je TJEDNA (21.07, 28.07, 03.08, 18.08…)
  ⇒ Saša je s izravne uplate prešao na KEKS Pay i od tada trener nije bio prepoznat.

Preostala 2 KEKS −20,00 NISU trener i namjerno se ne diraju:
  · 2025-10-29 `KUPOVINAKEKSPAY Zagreb`  · 2026-04-06 `Picek` (Mastercard)
  Saša je potvrdio da je 2025./2026. prestao s trenerom.

Zašto ih je manje od 52 godišnje: put, bolest, praznici + dio plaćen gotovinom.
Nije rupa u podacima — v. razgovor S107x.

UČINAK: saldo se NE mijenja (ni iznos ni datum se ne diraju) — mijenja se samo razrez
po Tipu. Po §2.14 spec-a `Transfer` retci se IZBACUJU iz `breakdown` pločice, pa je
trener dosad bio nevidljiv u statistici potrošnje.

⚠ Zašto jednokratna skripta, a ne `Pravila` red: `Transfer | izmedju racuna` je VALJAN
par, a `apply_rules.py` preskače svaki redak s valjanim parom — ista zamka kao
`fix_vocarna_pravilo.py` (S107o) i `fix_anja_rate.py` (S107r).

⚠ Namjerno se NE dodaje keyword pravilo za budući import: `KEKS` je aplikacija za
plaćanje, ne trgovac — isti razlog zbog kojeg su `PAYPAL`/`KEKS PAY`/`GLS` izbačeni iz
pravila u S107l. Uvjet „KEKS + točno 20 €" vrijedi samo za ovo razdoblje.

Pokretanje:
    python fix_keks_trener.py --dry
    python fix_keks_trener.py
Backup: `.pre-keks-*` prije snimanja.
"""
from __future__ import annotations

import shutil
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from apply_rules import find_header_col, pick_file  # noqa: E402

NEW_TIP, NEW_POD = 'Zdravlje', 'Sport_Sasa'
OLD_TIP, OLD_POD = 'Transfer', 'izmedju racuna'
IZNOS = -20.0

# (source_key, očekivani event_date) — traži se po source_key, NE po broju retka
# (retci se pomiču pri sortu/dedupu; nauk iz S107o).
TARGETS = [
    ('a3df1d7c1f94', '2023-07-21'), ('89cc59ba858b', '2023-07-28'),
    ('444edb1e07b1', '2023-08-03'), ('4e1263a55f61', '2023-08-18'),
    ('634c435a926c', '2023-08-23'), ('15e1354af77d', '2023-09-01'),
    ('0a0c9b56013f', '2023-09-07'), ('3488b0d7ea37', '2023-09-15'),
    ('49dfc25f2fb5', '2023-09-23'), ('78bc6a5a2afa', '2023-10-06'),
    ('4066791af93d', '2023-11-09'), ('27dec7f77357', '2023-11-29'),
    ('a8ad59b3ceb3', '2023-12-07'), ('5fd5d045bd70', '2023-12-23'),
    ('ec128f57cdba', '2024-02-08'), ('c1b367e3dbc1', '2024-02-15'),
    ('509c5ffbbf21', '2024-02-21'), ('ea2c9097c968', '2024-02-29'),
    ('35ddb9b8cc7c', '2024-03-07'), ('5ff5d0c476e2', '2024-03-14'),
]


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    return v if isinstance(v, date) else None


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    path = pick_file(args)
    print(f'File: {path.name}{"  [DRY RUN — bez snimanja]" if dry else ""}\n')

    wb = openpyxl.load_workbook(path)
    ws = wb['Review']
    c_key = find_header_col(ws, 'source_key')
    c_tip = find_header_col(ws, 'Tip')
    c_pod = find_header_col(ws, 'Podtip')
    c_nap = find_header_col(ws, 'Napomena')
    c_isp = find_header_col(ws, 'Isplata')
    c_ed = find_header_col(ws, 'event_date')
    c_alt = find_header_col(ws, 'Alternativa / nap.')
    c_run = find_header_col(ws, 'Pravilo run')

    by_key = {str(ws.cell(r, c_key).value or '').strip(): r
              for r in range(2, ws.max_row + 1)}
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    changed = 0

    for key, exp_date in TARGETS:
        r = by_key.get(key)
        if r is None:
            sys.exit(f'✗ source_key {key} nije nađen — je li ovo pravi Review file?')
        nap = str(ws.cell(r, c_nap).value or '').strip()
        if nap.upper() != 'KEKS':
            sys.exit(f'✗ red {r} ({key}): Napomena je "{nap[:40]}", očekivano "KEKS" — '
                     f'PREKID, ništa nije upisano.')
        ed = to_date(ws.cell(r, c_ed).value)
        if ed is None or str(ed) != exp_date:
            sys.exit(f'✗ red {r} ({key}): event_date {ed} ≠ očekivano {exp_date} — PREKID.')
        try:
            isp = float(ws.cell(r, c_isp).value)
        except (TypeError, ValueError):
            sys.exit(f'✗ red {r}: Isplata nije broj — PREKID.')
        if abs(-isp - IZNOS) > 0.01:
            sys.exit(f'✗ red {r}: Isplata {isp} ≠ 20,00 — PREKID.')

        tip_now = str(ws.cell(r, c_tip).value or '').strip()
        pod_now = str(ws.cell(r, c_pod).value or '').strip()
        if (tip_now, pod_now) == (NEW_TIP, NEW_POD):
            print(f'· red {r} ({exp_date}) već je {NEW_TIP}|{NEW_POD} — preskačem')
            continue
        if (tip_now, pod_now) != (OLD_TIP, OLD_POD):
            sys.exit(f'✗ red {r} ({exp_date}): očekivan {OLD_TIP}|{OLD_POD}, '
                     f'nađen {tip_now}|{pod_now} — PREKID.')

        print(f'{"bi se ispravio" if dry else "ispravljen"} red {r:>5} ({exp_date}, '
              f'{IZNOS:g} €): {tip_now}|{pod_now} → {NEW_TIP}|{NEW_POD}')
        if not dry:
            ws.cell(r, c_tip, NEW_TIP)
            ws.cell(r, c_pod, NEW_POD)
            old_alt = str(ws.cell(r, c_alt).value or '').strip()
            mark = f'S107x fix: trener preko KEKS Pay, bio {tip_now}/{pod_now}'
            ws.cell(r, c_alt, f'{old_alt} | {mark}' if old_alt else mark)
            ws.cell(r, c_run, stamp)
        changed += 1

    print(f'\n{"Bi se ispravilo" if dry else "Ispravljeno"}: {changed} redaka '
          f'({changed * -IZNOS:.2f} € prelazi iz Transfera u trošak Zdravlje|Sport_Sasa)')
    if dry or not changed:
        if dry:
            print('DRY RUN — ništa nije snimljeno.')
        return

    backup = path.with_name(f'{path.stem}.pre-keks-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori {path.name} u Excelu i ponovi.\n'
                 f'  (backup je već napravljen: {backup.name})')
    print(f'✔ Snimljeno. Backup: {backup.name}')


if __name__ == '__main__':
    main()
