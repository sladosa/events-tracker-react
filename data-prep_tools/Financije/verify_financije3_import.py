# -*- coding: utf-8 -*-
"""
verify_financije3_import.py
============================
Provjerava da li su SVI redovi iz izvornih sheetova ('koka EU', 'sasa EU')
ispravno zavrsili u bazi kao Financije_3 eventi.

Reproducira istu transformaciju kao make_financije3_import.py (date korekcija,
leaf_comment, Tip klasifikacija, Smjer, vrijeme sesije) i match-a svaki izvorni
red s odgovarajucim DB eventom iz export filea preko kljuca (event_date, session_start)
- taj par je jedinstven po dizajnu import skripte.

Output: Financije3_verify.xlsx - jedan red po izvornom retku:
  source_sheet, source_row, izvorne vrijednosti + event_id + DB vrijednosti + status
"""

import sys
from pathlib import Path
from datetime import date, datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

sys.stdout.reconfigure(encoding='utf-8')

# Reuse transformation logic from the import script
sys.path.insert(0, str(Path(__file__).parent))
from make_financije3_import import (
    SHEETS, MIN_DATE, MAX_DATE,
    parse_datum, correct_date, classify_tip, get_smjer, assign_times,
)

# ── Paths ──────────────────────────────────────────────────────────────────────
SOURCE  = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije\Financije 2026-06.xlsx")
EXPORT  = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije\events_export_20260607_193138.xlsx")
OUTPUT  = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije\Financije3_verify.xlsx")

# ── Styles ─────────────────────────────────────────────────────────────────────
BLUE_FILL  = PatternFill("solid", fgColor="4472C4")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
WHITE_FONT = Font(color="FFFFFF", bold=True)


def num(v):
    if v is None:
        return None
    try:
        return round(float(v), 2)
    except (ValueError, TypeError):
        return None


# ── Re-read source sheets, tracking source row + reproducing transform ─────────

def read_source_with_transform(ws, sheet_name, racun_label):
    """Mirrors make_financije3_import.read_sheet but also keeps source_row + raw cols."""
    rows = []
    last_valid_date = MIN_DATE

    for r in range(2, ws.max_row + 1):
        opis    = ws.cell(r, 2).value
        datum   = ws.cell(r, 3).value
        uplata  = ws.cell(r, 4).value
        isplata = ws.cell(r, 5).value
        stanje  = ws.cell(r, 6).value

        d = parse_datum(datum)

        if d is None and uplata is None and isplata is None and opis is None:
            continue  # fully empty row - not in import either

        datum_greska = None
        status_note = ''
        if d is None:
            original = str(datum).strip() if datum is not None else '(prazno)'
            datum_greska = original
            d = last_valid_date
        elif d < MIN_DATE or d > MAX_DATE:
            corrected, note = correct_date(d, last_valid_date)
            datum_greska = note
            d = corrected
        else:
            last_valid_date = d

        if datum_greska and MIN_DATE <= d <= MAX_DATE:
            last_valid_date = d

        if uplata is None and isplata is None:
            rows.append({
                'source_sheet': sheet_name, 'source_row': r,
                'opis': opis, 'datum_raw': datum, 'uplata_raw': uplata,
                'isplata_raw': isplata, 'stanje_raw': stanje,
                'skipped': True, 'skip_reason': 'balance row (no Uplata/Isplata)',
            })
            continue

        opis_str = str(opis).strip() if opis else ''
        if datum_greska:
            napomena = f'[DATUM_GREŠKA: {datum_greska}] {opis_str}'.strip()
        else:
            napomena = opis_str
        prefix = 'ZABA' if 'ZABA' in racun_label else 'RF'
        comment = f'{prefix}: {opis_str}' if opis_str else prefix

        rows.append({
            'source_sheet': sheet_name, 'source_row': r,
            'opis': opis, 'datum_raw': datum, 'uplata_raw': uplata,
            'isplata_raw': isplata, 'stanje_raw': stanje,
            'skipped': False,
            'date': d.isoformat(), 'date_obj': d,
            'racun': racun_label,
            'uplata': num(uplata), 'isplata': num(isplata), 'stanje': num(stanje),
            'valuta': 'EUR', 'napomena': napomena, 'comment': comment,
            'smjer': get_smjer(uplata, isplata), 'tip': classify_tip(opis_str),
        })
    return rows


# ── Load DB-exported events ────────────────────────────────────────────────────

def load_export_events(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Events']

    headers = None
    col = None
    events = []
    for vals in ws.iter_rows(values_only=True):
        if headers is None:
            if vals and vals[0] == 'event_id':
                headers = list(vals)
                col = {h: i for i, h in enumerate(headers) if h}
            continue

        if vals[col['event_id']] is None:
            continue
        ev_date = vals[col['event_date']]
        if isinstance(ev_date, datetime):
            ev_date = ev_date.date()
        session_start = vals[col['session_start']]
        events.append({
            'event_id':       vals[col['event_id']],
            'event_date':     ev_date.isoformat() if ev_date else None,
            'session_start':  str(session_start).strip() if session_start else None,
            'leaf_comment':   vals[col.get('leaf comment')],
            'racun':          vals[col.get('Racun (Transakcija)')],
            'uplata':         num(vals[col.get('Uplata (Transakcija)')]),
            'isplata':        num(vals[col.get('Isplata (Transakcija)')]),
            'stanje':         num(vals[col.get('Stanje (Transakcija)')]),
            'valuta':         vals[col.get('Valuta (Transakcija)')],
            'napomena':       vals[col.get('Napomena (Transakcija)')],
            'smjer':          vals[col.get('Smjer (Transakcija)')],
            'tip':            vals[col.get('Tip (Transakcija)')],
        })
    return events


# ── Output workbook ─────────────────────────────────────────────────────────────

OUT_HEADERS = [
    'source_sheet', 'source_row', 'src_Opis', 'src_Datum', 'src_Uplata', 'src_Isplata', 'src_Stanje',
    'status', 'event_id', 'event_date', 'session_start',
    'db_leaf_comment', 'db_Racun', 'db_Uplata', 'db_Isplata', 'db_Stanje',
    'db_Valuta', 'db_Napomena', 'db_Smjer', 'db_Tip',
    'mismatch_fields',
]


def write_output(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Verify'

    for c, h in enumerate(OUT_HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16

    for r_idx, row in enumerate(rows, 2):
        for c_idx, key in enumerate(OUT_HEADERS, 1):
            ws.cell(r_idx, c_idx, row.get(key))
        status = row.get('status')
        fill = {
            'MATCH': GREEN_FILL,
            'MISMATCH': AMBER_FILL,
            'NOT FOUND': RED_FILL,
            'SKIPPED (balance row)': None,
        }.get(status)
        if fill:
            ws.cell(r_idx, 8).fill = fill

    ws.freeze_panes = 'A2'
    wb.save(OUTPUT)


# ── Main ────────────────────────────────────────────────────────────────────────

def main():
    if not SOURCE.exists():
        print(f'ERROR: source not found: {SOURCE}')
        sys.exit(1)
    if not EXPORT.exists():
        print(f'ERROR: export not found: {EXPORT}')
        sys.exit(1)

    wb_src = openpyxl.load_workbook(SOURCE, data_only=True)

    all_rows = []
    for sheet_name, racun_label in SHEETS.items():
        if sheet_name not in wb_src.sheetnames:
            print(f'WARNING: sheet "{sheet_name}" not found, skipping')
            continue
        ws = wb_src[sheet_name]
        all_rows.extend(read_source_with_transform(ws, sheet_name, racun_label))

    # Reproduce the exact same global sort + time assignment as the import script
    transformable = [r for r in all_rows if not r['skipped']]
    skipped_rows  = [r for r in all_rows if r['skipped']]

    transformable.sort(key=lambda e: e['date'])
    assign_times(transformable)

    # Build lookup of DB events by (event_date, session_start) - unique key
    db_events = load_export_events(EXPORT)
    db_by_key = {}
    dupes = []
    for ev in db_events:
        key = (ev['event_date'], ev['session_start'])
        if key in db_by_key:
            dupes.append(key)
        db_by_key[key] = ev

    matched_ids = set()
    out_rows = []

    FIELD_CHECKS = [
        ('racun', 'racun'), ('uplata', 'uplata'), ('isplata', 'isplata'),
        ('stanje', 'stanje'), ('valuta', 'valuta'), ('napomena', 'napomena'),
        ('comment', 'leaf_comment'), ('smjer', 'smjer'), ('tip', 'tip'),
    ]

    for row in transformable:
        key = (row['date'], row['time'])
        db = db_by_key.get(key)
        out = {
            'source_sheet': row['source_sheet'],
            'source_row':   row['source_row'],
            'src_Opis':     row['opis'],
            'src_Datum':    row['datum_raw'],
            'src_Uplata':   row['uplata_raw'],
            'src_Isplata':  row['isplata_raw'],
            'src_Stanje':   row['stanje_raw'],
        }
        if db is None:
            out['status'] = 'NOT FOUND'
        else:
            matched_ids.add(db['event_id'])
            mismatches = []
            for src_key, db_key in FIELD_CHECKS:
                src_val = row[src_key]
                db_val = db[db_key]
                # Empty string vs NULL: DB doesn't persist empty attribute values
                if (src_val or None) != (db_val or None):
                    mismatches.append(f'{db_key}: src={src_val!r} db={db_val!r}')
            out.update({
                'event_id':        db['event_id'],
                'event_date':      db['event_date'],
                'session_start':   db['session_start'],
                'db_leaf_comment': db['leaf_comment'],
                'db_Racun':        db['racun'],
                'db_Uplata':       db['uplata'],
                'db_Isplata':      db['isplata'],
                'db_Stanje':       db['stanje'],
                'db_Valuta':       db['valuta'],
                'db_Napomena':     db['napomena'],
                'db_Smjer':        db['smjer'],
                'db_Tip':          db['tip'],
            })
            if mismatches:
                out['status'] = 'MISMATCH'
                out['mismatch_fields'] = '; '.join(mismatches)
            else:
                out['status'] = 'MATCH'
        out_rows.append(out)

    for row in skipped_rows:
        out_rows.append({
            'source_sheet': row['source_sheet'],
            'source_row':   row['source_row'],
            'src_Opis':     row['opis'],
            'src_Datum':    row['datum_raw'],
            'src_Uplata':   row['uplata_raw'],
            'src_Isplata':  row['isplata_raw'],
            'src_Stanje':   row['stanje_raw'],
            'status':       'SKIPPED (balance row)',
        })

    write_output(out_rows)

    # Summary
    n_match     = sum(1 for r in out_rows if r['status'] == 'MATCH')
    n_mismatch  = sum(1 for r in out_rows if r['status'] == 'MISMATCH')
    n_notfound  = sum(1 for r in out_rows if r['status'] == 'NOT FOUND')
    n_skipped   = sum(1 for r in out_rows if r['status'] == 'SKIPPED (balance row)')
    db_unmatched = [ev for ev in db_events if ev['event_id'] not in matched_ids]

    print(f'\n{"="*60}')
    print(f'OUTPUT: {OUTPUT}')
    print(f'Source rows processed: {len(all_rows)}')
    print(f'  MATCH:    {n_match}')
    print(f'  MISMATCH: {n_mismatch}')
    print(f'  NOT FOUND: {n_notfound}')
    print(f'  SKIPPED (balance row): {n_skipped}')
    print(f'\nDB events total: {len(db_events)}')
    print(f'DB events matched to a source row: {len(matched_ids)}')
    print(f'DB events with NO matching source row (orphans in DB): {len(db_unmatched)}')
    if dupes:
        print(f'\nWARNING: duplicate (event_date, session_start) keys in DB export: {len(dupes)}')
        for k in dupes[:10]:
            print(f'  {k}')
    if db_unmatched:
        print(f'\nFirst orphan DB events:')
        for ev in db_unmatched[:10]:
            print(f"  {ev['event_id']}  {ev['event_date']} {ev['session_start']}  {ev['leaf_comment']}")
    print(f'{"="*60}')


if __name__ == '__main__':
    main()
