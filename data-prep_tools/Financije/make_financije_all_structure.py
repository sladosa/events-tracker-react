#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
make_financije_all_structure.py  —  S107s

Generira Structure Excel za novu areu `Financije_all` (Structure tab → Import).

Ulaz:
  1) BASE   — Activities/Structure export postojece PROD aree `Financije`
              (sheet `Structure`); iz njega se preuzimaju atributi koji su
              ispravni 1:1 (Racun, Izvor, Status, Smjer, Uplata/Isplata,
              Rate?/Broj rata, Stanje, Valuta, Napomena).
  2) REVIEW — `Financije_review_*.xlsx`, sheet `Taksonomija` (Tip | Podtip)
              iz kojeg se REGENERIRAJU `Tip` i `Podtip` retci.

Izlaz: `Financije_all_structure_<ts>.xlsx` — sheetovi `Structure` + `Automations`.

Sto se mijenja u odnosu na BASE (v. MODS ispod):
  - Area `Financije` -> `Financije_all`
  - `Tip`    : 13 starih opcija -> svi Tipovi iz Taksonomije
  - `Podtip` : options_map regeneriran iz Taksonomije (1 redak po Tipu + `*`)
  - `Napomena` -> preimenovan u `Izvod opis` (slug `izvod_opis`)
  - NOVO: `Datum naplate`, `Datum kupovine` (datetime)
  - `Uplata`/`Isplata` dobivaju Unit=EUR (prikaz u formi, 0 zapisa u bazi)
  - `Valuta` gubi default `EUR` (default bi se s touched=true pisao na SVAKI
    novi event -> 1 suvisan redak u event_attributes po transakciji)
  - novi `Sort` (lanac ovisnosti odozgo, auto-polja na dnu)
  - `Automations`: set_attribute pravilo za `Datum naplate` po Izvoru

Pokretanje (v. ENRICH_PLAN §4 — NE preko run.bat, cmd gusi zarez u argumentima):
    data-prep_tools/Tools/venv/Scripts/python.exe \
        data-prep_tools/Financije/make_financije_all_structure.py --dry
"""

from __future__ import annotations

import argparse
import datetime as dt
import glob
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# Putanje
# ─────────────────────────────────────────────────────────────

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
DATA = os.path.join(ROOT, "data-prep_data", "Financije")

BASE_GLOB = os.path.join(DATA, "events_export_preview_*.xlsx")
REVIEW_GLOB = os.path.join(DATA, "Financije_review_*.xlsx")

NEW_AREA = "Financije_all"
OLD_AREA = "Financije"
LEAF = "Transakcija"

# ─────────────────────────────────────────────────────────────
# MODS — sve odluke S107s na jednom mjestu
# ─────────────────────────────────────────────────────────────

# Redoslijed u formi: prvo sto korisnik BIRA, na dnu sto se SAMO POPUNI.
SORT_ORDER = [
    "Racun",          # 1  ┐ lanac ovisnosti
    "Izvor",          # 2  ┤ (Racun -> Izvor -> Status)
    "Smjer",          # 3  ┘ gata Uplata/Isplata i Rate?
    "Uplata",         # 4
    "Isplata",        # 5
    "Tip",            # 6  ┐
    "Podtip",         # 7  ┘ Tip -> Podtip
    "Izvod opis",     # 8  (bivsa `Napomena`)
    "Rate?",          # 9  ┐
    "Broj rata",      # 10 ┤ ukupno N
    "Rata br",        # 11 ┘ redni broj OVE uplate (1..N)
    "Datum naplate",  # 12 auto (set_attribute + rata automatika)
    "Status",         # 13 auto (default_map po Izvoru)
    "Stanje",         # 14 skriven u formi
    "Valuta",         # 15 rijetko
]

# `Datum kupovine` NAMJERNO IZOSTAVLJEN (odluka S107t, poništava D1a):
# postojao je samo kao zakrpa za iznimku iz D1 ("rata nosi event_date = datum
# naplate"). Iznimka je ukinuta — sve rate jedne kupovine dijele event_date =
# dan kupnje, a razlikuje ih `Datum naplate` — pa bi `Datum kupovine` bio
# doslovno jednak event_date-u na svakom retku. Ako zatreba kao povijesni
# anker, dodaje se natrag jednim nedestruktivnim Structure importom.

# Preimenovanja: stari AttrName -> (novi AttrName, novi Slug)
RENAME = {
    "Napomena": ("Izvod opis", "izvod_opis"),
}

# Unit koji se postavlja (prikaz uz number input; ne trosi zapise u bazi)
SET_UNIT = {
    "Uplata": "EUR",
    "Isplata": "EUR",
    "Stanje": "EUR",
}

# Atributi kojima se BRISE default_value
CLEAR_DEFAULT = {"Valuta"}

# Novi atributi. Bez `Variants` = jedan redak; s `Variants` = po redak za svaku
# (DependsOn, WhenValue) kombinaciju (isti obrazac kojim BASE opisuje `Broj rata`).
NEW_ATTRS = [
    {
        "AttrName": "Rata br",
        "Slug": "rata_br",
        "AttrType": "number",
        "Val.Type": "suggest",
        # Vidljiv pod istim uvjetom kao `Broj rata` (rate=TRUE). Prazne
        # TextOptions -> options_map {'*': [], 'TRUE': []}, tj. cisto gatanje
        # vidljivosti bez dropdowna (v. structureImport.ts:327).
        "Variants": [
            {"DependsOn": "rate", "WhenValue": "*"},
            {"DependsOn": "rate", "WhenValue": "TRUE"},
        ],
        "Description": "Redni broj OVE uplate unutar plana (1..Broj rata). "
                       "Smije biti prazan (nepoznato ili zbirna uplata za vise rata).",
    },
    {
        "AttrName": "Datum naplate",
        "Slug": "datum_naplate",
        "AttrType": "datetime",
        "Description": "Dan kad banka/kartica stvarno skine iznos "
                       "(Racun/Cash = isti dan, kartice = sljedeci mjesec; "
                       "na ratama = dan naplate te rate).",
    },
]

# CommentTemplate — `{napomena}` iz BASE-a IZBACEN (odluka Sasa, S107s):
# `Izvod opis` je pri unosu uvijek prazan (izvod stize tek kasnije), pa bi
# auto-comment zavrsavao praznim repom ("ZABA/Razno/Pokloni/").
# NAPOMENA: template se koristi SAMO kad je Event Note prazan
# (AddActivityPage.tsx:698), pa Kokina biljeska uvijek pobjeduje.
COMMENT_TEMPLATE = "{racun}/{tip}/{podtip}"

# Automations — set_attribute za `Datum naplate` po Izvoru.
# next:11 = Kokino pravilo za Mastercard; next:3 = PBZ/Visa s RF racuna
# (isti brojevi kao postojeci `automations.rata.date_map` {RF:3, ZABA:11}).
AUTOMATION_ROWS = [
    {
        "Area": NEW_AREA,
        "RuleName": "Datum naplate po Izvoru",
        "Action": "set_attribute",
        "TargetAttr": "datum_naplate",
        "MapAttr": "izvorplacanja",
        "DateMap": "Racun=same | Cash=same | Mastercard=next:11 | Visa=next:3",
    },
    # Post-Finish rata modal. Do S107t ova konfiguracija NIJE prolazila Structure
    # roundtripom (isla je SQL-om) pa se tiho gubila pri prijenosu aree — v.
    # Sasin princip "sve ide importom". Dani naplate su isti brojevi kao u
    # `datum_naplate` pravilu gore (ZABA/Mastercard 11., RF/Visa 3.).
    {
        "Area": NEW_AREA,
        "RuleName": "Generiraj rate",
        "Action": "rata",
        "TargetAttr": "datum_naplate",   # kamo ide datum naplate svake rate
        "MapAttr": "izvorplacanja",
        "DateMap": "Mastercard=11 | Visa=3",
        "TriggerAttr": "rate",
        "CountAttr": "brojrata",
        "AmountAttr": "isplata",
        "OverrideAttrs": "status=Planiran",
        "CommentAttr": "",
        "IndexAttr": "rata_br",          # redni broj rate 1..N
    },
]

# Kolone Structure sheeta — redoslijed kao u app exportu
COLUMNS = [
    "Type", "IsLeaf", "Area", "SharedWith", "CategoryPath", "Sort",
    "AttrName", "Slug", "AttrType", "IsRequired", "Val.Type", "Default",
    "Val.Max (no)", "Unit", "TextOptions/Val.Min", "DependsOn", "WhenValue",
    "Description", "CommentTemplate",
]

HEADER_ROW = 3  # findHeader() u structureImport.ts skenira redove 1–12

# ─────────────────────────────────────────────────────────────
# Helperi
# ─────────────────────────────────────────────────────────────


def pick(pattern: str, exclude_backups: bool = True) -> str:
    """Najnoviji file po mtime; preskace `.pre-*` backupe."""
    files = glob.glob(pattern)
    if exclude_backups:
        files = [f for f in files if ".pre-" not in os.path.basename(f)]
        files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        sys.exit(f"NEMA fajla za uzorak: {pattern}")
    return max(files, key=os.path.getmtime)


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def find_header_row(ws, needle: str = "type") -> int:
    """Trazi red zaglavlja skeniranjem prvih 12 redova (isto kao app)."""
    for r in range(1, 13):
        vals = [cell_str(c.value).lower() for c in ws[r]]
        if needle in vals and ("categorypath" in vals or "chain" in vals):
            return r
    sys.exit("Ne mogu naci zaglavlje u BASE Structure sheetu.")


def read_base(path: str) -> list[dict]:
    """Cita BASE Structure sheet -> lista redaka (dict po nazivu kolone)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Structure" not in wb.sheetnames:
        sys.exit(f"BASE nema sheet 'Structure': {path}")
    ws = wb["Structure"]
    hr = find_header_row(ws)
    hdr = [cell_str(c.value) for c in ws[hr]]
    idx = {h: i for i, h in enumerate(hdr) if h}

    rows = []
    for r in ws.iter_rows(min_row=hr + 1, values_only=True):
        typ = cell_str(r[idx["Type"]]) if "Type" in idx else ""
        if typ.lower() not in ("area", "category", "attribute"):
            continue
        rows.append({h: cell_str(r[i]) for h, i in idx.items()})
    wb.close()
    return rows


def read_taxonomy(path: str) -> "dict[str, list[str]]":
    """Cita `Taksonomija` sheet -> {Tip: [Podtip, ...]} u redoslijedu pojave."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Taksonomija" not in wb.sheetnames:
        sys.exit(f"REVIEW nema sheet 'Taksonomija': {path}")
    ws = wb["Taksonomija"]
    taks: "dict[str, list[str]]" = {}
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        tip = cell_str(r[0]) if len(r) > 0 else ""
        pod = cell_str(r[1]) if len(r) > 1 else ""
        if not tip:
            continue
        taks.setdefault(tip, [])
        if pod:
            taks[tip].append(pod)
    wb.close()

    bad = [v for v in list(taks) + [p for ps in taks.values() for p in ps] if "|" in v]
    if bad:
        sys.exit(f"Taksonomija sadrzi '|' (separator TextOptions): {bad}")
    return taks


def blank_row() -> dict:
    return {c: "" for c in COLUMNS}


# ─────────────────────────────────────────────────────────────
# Gradnja
# ─────────────────────────────────────────────────────────────


def build_rows(base: list[dict], taks: "dict[str, list[str]]") -> list[dict]:
    new_path = f"{NEW_AREA} > {LEAF}"

    # ── Area + Category retci ────────────────────────────────
    out: list[dict] = []

    r = blank_row()
    r.update({"Type": "Area", "CategoryPath": NEW_AREA, "Sort": "1",
              "CommentTemplate": COMMENT_TEMPLATE})
    out.append(r)

    r = blank_row()
    r.update({"Type": "Category", "IsLeaf": "TRUE", "CategoryPath": new_path,
              "Sort": "1", "CommentTemplate": COMMENT_TEMPLATE})
    out.append(r)

    # ── Atributi iz BASE-a, grupirani po AttrName ────────────
    groups: "dict[str, list[dict]]" = {}
    for row in base:
        if row.get("Type", "").lower() != "attribute":
            continue
        name = row.get("AttrName", "")
        if not name:
            continue
        groups.setdefault(name, []).append(row)

    # Tip/Podtip se REGENERIRAJU iz Taksonomije — BASE verzija se odbacuje
    groups.pop("Tip", None)
    groups.pop("Podtip", None)

    # Preimenovanja
    for old, (new_name, new_slug) in RENAME.items():
        if old in groups:
            rows = groups.pop(old)
            for rw in rows:
                rw["AttrName"] = new_name
                rw["Slug"] = new_slug
            groups[new_name] = rows

    # Novi atributi
    for spec in NEW_ATTRS:
        spec = dict(spec)
        variants = spec.pop("Variants", [{}])
        rows = []
        for var in variants:
            rw = blank_row()
            rw.update({"Type": "Attribute", "IsRequired": "FALSE", "Val.Type": "none"})
            rw.update(spec)
            rw.update(var)
            rows.append(rw)
        groups[spec["AttrName"]] = rows

    # Tip — jedan redak, sve opcije
    tip_row = blank_row()
    tip_row.update({
        "Type": "Attribute", "AttrName": "Tip", "Slug": "tip",
        "AttrType": "text", "IsRequired": "FALSE", "Val.Type": "suggest",
        "TextOptions/Val.Min": "|".join(taks.keys()),
    })
    groups["Tip"] = [tip_row]

    # Podtip — `*` fallback + jedan redak po Tipu (i za Tipove bez Podtipa,
    # da dropdown bude prazan umjesto da padne na `*`)
    pod_rows = []
    star = blank_row()
    star.update({
        "Type": "Attribute", "AttrName": "Podtip", "Slug": "podtip",
        "AttrType": "text", "IsRequired": "FALSE", "Val.Type": "suggest",
        "DependsOn": "tip", "WhenValue": "*",
    })
    pod_rows.append(star)
    for tip, pods in taks.items():
        rw = blank_row()
        rw.update({
            "Type": "Attribute", "AttrName": "Podtip", "Slug": "podtip",
            "AttrType": "text", "IsRequired": "FALSE", "Val.Type": "suggest",
            "TextOptions/Val.Min": "|".join(pods),
            "DependsOn": "tip", "WhenValue": tip,
        })
        pod_rows.append(rw)
    groups["Podtip"] = pod_rows

    # ── Emit u SORT_ORDER redoslijedu ────────────────────────
    unknown = [n for n in groups if n not in SORT_ORDER]
    missing = [n for n in SORT_ORDER if n not in groups]
    if unknown:
        sys.exit(f"Atribut(i) iz BASE-a nisu u SORT_ORDER: {unknown}")
    if missing:
        sys.exit(f"SORT_ORDER trazi atribut(e) kojih nema: {missing}")

    for sort_no, name in enumerate(SORT_ORDER, start=1):
        for rw in groups[name]:
            o = blank_row()
            o.update(rw)
            o["Type"] = "Attribute"
            o["CategoryPath"] = new_path
            o["Sort"] = str(sort_no)
            o["Area"] = ""
            o["SharedWith"] = ""
            o["IsLeaf"] = ""
            o["CommentTemplate"] = ""
            if name in SET_UNIT:
                o["Unit"] = SET_UNIT[name]
            if name in CLEAR_DEFAULT:
                o["Default"] = ""
            out.append(o)

    return out


# ─────────────────────────────────────────────────────────────
# Pisanje
# ─────────────────────────────────────────────────────────────

FILL_HDR = PatternFill("solid", fgColor="FF4472C4")
FILL_AREA = PatternFill("solid", fgColor="FFFFF2CC")
FILL_ATTR = PatternFill("solid", fgColor="FFE6F2FF")
FILL_DEP = PatternFill("solid", fgColor="FFE2EFDA")


def write_xlsx(rows: list[dict], out_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Structure"

    ws.cell(1, 1, f"Financije_all — struktura za Structure Import "
                  f"(generirano {dt.datetime.now():%Y-%m-%d %H:%M})").font = Font(bold=True, size=12)
    ws.cell(2, 1, "Zuto = Area/Category · plavo = atributi · zeleno = DependsOn/WhenValue. "
                  "Import je NEDESTRUKTIVAN (nista se ne brise).").font = Font(italic=True, size=9)

    for c, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(HEADER_ROW, c, name)
        cell.fill = FILL_HDR
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(vertical="center")

    dep_col = COLUMNS.index("DependsOn") + 1
    when_col = COLUMNS.index("WhenValue") + 1

    for i, row in enumerate(rows):
        r = HEADER_ROW + 1 + i
        is_area = row["Type"] in ("Area", "Category")
        for c, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(r, c, row.get(name) or None)
            if is_area:
                cell.fill = FILL_AREA
            elif c in (dep_col, when_col):
                cell.fill = FILL_DEP
            elif c >= COLUMNS.index("AttrName") + 1:
                cell.fill = FILL_ATTR

    widths = {"Type": 11, "IsLeaf": 8, "Area": 10, "SharedWith": 14,
              "CategoryPath": 26, "Sort": 6, "AttrName": 16, "Slug": 16,
              "AttrType": 10, "IsRequired": 11, "Val.Type": 10, "Default": 10,
              "Val.Max (no)": 12, "Unit": 7, "TextOptions/Val.Min": 60,
              "DependsOn": 13, "WhenValue": 16, "Description": 44,
              "CommentTemplate": 34}
    for c, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(name, 12)

    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(COLUMNS))}{HEADER_ROW + len(rows)}"

    # ── Automations sheet (zaglavlje MORA biti u redu 1) ─────
    wa = wb.create_sheet("Automations")
    acols = ["Area", "RuleName", "Action", "TargetAttr", "MapAttr", "DateMap",
             "TriggerAttr", "CountAttr", "AmountAttr", "OverrideAttrs",
             "CommentAttr", "IndexAttr"]
    for c, name in enumerate(acols, start=1):
        cell = wa.cell(1, c, name)
        cell.fill = FILL_HDR
        cell.font = Font(bold=True, color="FFFFFFFF")
    for i, rule in enumerate(AUTOMATION_ROWS, start=2):
        for c, name in enumerate(acols, start=1):
            wa.cell(i, c, rule.get(name, ""))
    note = len(AUTOMATION_ROWS) + 3
    for i, line in enumerate((
        "Import ZAMJENJUJE navedene automatike svake Aree koja se pojavi u sheetu.",
        "set_attribute DateMap: 'vrijednost=same' ili 'vrijednost=next:N' (N = 1..31).",
        "rata DateMap: 'vrijednost=DAN' gdje je DAN broj 1..31 (dan naplate u mjesecu).",
        "rata: sve rate dijele event_date (dan kupnje); razlikuje ih TargetAttr (datum naplate) "
        "i pomak session_start-a za 1 min.",
    )):
        wa.cell(note + i, 1, line)
    for c, w in zip("ABCDEFGHIJKL",
                    (16, 20, 12, 16, 16, 34, 13, 12, 12, 20, 13, 12)):
        wa.column_dimensions[c].width = w

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────
# Report
# ─────────────────────────────────────────────────────────────


def report(rows: list[dict], taks: "dict[str, list[str]]") -> None:
    attrs = [r for r in rows if r["Type"] == "Attribute"]
    names = []
    for r in attrs:
        if r["AttrName"] not in names:
            names.append(r["AttrName"])

    print(f"  Area          : {NEW_AREA}  (leaf L1 '{LEAF}')")
    print(f"  Taksonomija   : {len(taks)} Tipova / "
          f"{sum(len(v) for v in taks.values())} Podtipova")
    print(f"  Redaka ukupno : {len(rows)}  (Area 1 + Category 1 + Attribute {len(attrs)})")
    print(f"  Atributa      : {len(names)}")
    print()
    print("  Sort  Atribut          Tip        Redaka  Napomena")
    print("  " + "-" * 68)
    for name in names:
        grp = [r for r in attrs if r["AttrName"] == name]
        note = ""
        if grp[0].get("DependsOn"):
            note = f"depends_on={grp[0]['DependsOn']}"
        if grp[0].get("Unit"):
            note = (note + " " if note else "") + f"unit={grp[0]['Unit']}"
        if name in [s["AttrName"] for s in NEW_ATTRS]:
            note = (note + " " if note else "") + "** NOVO **"
        if name in [new for new, _ in RENAME.values()]:
            note = (note + " " if note else "") + "** PREIMENOVANO **"
        print(f"  {grp[0]['Sort']:>4}  {name:<16} {grp[0]['AttrType']:<10} "
              f"{len(grp):>6}  {note}")
    print()
    for old, (new, slug) in RENAME.items():
        print(f"  preimenovano  : {old} -> {new} (slug '{slug}')")
    print(f"  automations   : {len(AUTOMATION_ROWS)} pravilo/a")
    for a in AUTOMATION_ROWS:
        if a["Action"] == "rata":
            print(f"                  rata: trigger={a['TriggerAttr']} count={a['CountAttr']} "
                  f"amount={a['AmountAttr']} index={a['IndexAttr']}")
            print(f"                        naplata -> {a['TargetAttr']} "
                  f"({a['MapAttr']}: {a['DateMap']}), override {a['OverrideAttrs']}")
        else:
            print(f"                  set_attribute: {a['TargetAttr']} <- "
                  f"{a['MapAttr']}: {a['DateMap']}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Generira Structure Excel za Financije_all")
    ap.add_argument("--base", help="Structure/Activities export postojece aree (xlsx)")
    ap.add_argument("--review", help="Financije_review_*.xlsx (sheet Taksonomija)")
    ap.add_argument("--out", help="izlazni xlsx")
    ap.add_argument("--dry", action="store_true", help="samo report, ne pise file")
    args = ap.parse_args()

    base_path = args.base or pick(BASE_GLOB)
    review_path = args.review or pick(REVIEW_GLOB)

    print("=" * 72)
    print("make_financije_all_structure.py")
    print("=" * 72)
    print(f"  BASE   : {os.path.basename(base_path)}")
    print(f"  REVIEW : {os.path.basename(review_path)}")
    print()

    base = read_base(base_path)
    taks = read_taxonomy(review_path)
    rows = build_rows(base, taks)

    report(rows, taks)

    if args.dry:
        print()
        print("  --dry: file NIJE napisan.")
        return

    out = args.out or os.path.join(
        DATA, f"Financije_all_structure_{dt.datetime.now():%Y%m%d_%H%M%S}.xlsx")
    write_xlsx(rows, out)
    print()
    print(f"  ZAPISANO: {out}")


if __name__ == "__main__":
    main()
