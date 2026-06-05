# -*- coding: utf-8 -*-
import sys, re, os
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT  = r"C:\0_Sasa\events-tracker-react\Claude-temp_R\Financije 2026-06.xlsx"
OUTPUT = os.path.join(SCRIPT_DIR, "Financije_2_ZaSasu_import.xlsx")
AREA   = "Financije_2"

# ── Fills / fonts ──────────────────────────────────────────────────────────────
PINK_FILL   = PatternFill("solid", fgColor="FFE6F0")
BLUE_FILL   = PatternFill("solid", fgColor="E6F2FF")
ORANGE_FILL = PatternFill("solid", fgColor="FFC000")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
LEG_FILL    = PatternFill("solid", fgColor="7030A0")
SEP_FILL    = PatternFill("solid", fgColor="FFD0E0")
WHITE_FONT  = Font(color="FFFFFF", bold=True)
TITLE_FONT  = Font(bold=True, size=12)
BOLD_FONT   = Font(bold=True)
THIN        = Side(style='thin')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
L_ALIGN     = Alignment(horizontal='left',   vertical='center')
C_ALIGN     = Alignment(horizontal='center', vertical='center')
R_ALIGN     = Alignment(horizontal='right',  vertical='center')

def parse_datum(datum):
    """Return datetime or None. Handles Excel Date objects and Croatian string formats."""
    if isinstance(datum, datetime):
        return datetime(datum.year, datum.month, datum.day)
    if datum is None:
        return None
    s = str(datum).strip().rstrip('.')
    # Try "D.M.YY" or "D.M.YYYY" — also handles typos like "1.3.326" → 2026
    import re as _re
    m = _re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d+)$', s)
    if m:
        d, mo, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 100:
            yr += 2000
        elif yr > 2100:
            yr = (yr % 100) + 2000   # '326' → 26 + 2000 = 2026
        try:
            return datetime(yr, mo, d)
        except ValueError:
            pass
    return None

def col_letter(n):
    s = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

# ── Attr column definitions ────────────────────────────────────────────────────
# (category_path_no_area, attr_name, data_type)
ATTR_COLS = [
    # Rashodi L1 — Iznos/Račun/Valuta inherited by all Rashodi leaves
    ("Rashodi",                  "Iznos",        "number"),
    ("Rashodi",                  "Račun",        "suggest"),
    ("Rashodi",                  "Valuta",       "suggest"),
    # Dom
    ("Rashodi > Dom",            "Vrsta",        "suggest"),
    ("Rashodi > Dom",            "Objekt",       "suggest"),
    # Svakodnevni
    ("Rashodi > Svakodnevni",    "Vrsta",        "suggest"),
    ("Rashodi > Svakodnevni",    "Dućan",        "suggest"),
    # Restoran
    ("Rashodi > Restoran",       "Naziv",        "text"),
    # Prijevoz
    ("Rashodi > Prijevoz",       "Vrsta",        "suggest"),
    ("Rashodi > Prijevoz",       "Vozilo",       "suggest"),
    # Zdravlje
    ("Rashodi > Zdravlje",       "Vrsta",        "suggest"),
    ("Rashodi > Zdravlje",       "Osoba",        "suggest"),
    # Trening (za manualni unos; Passport → Zdravlje)
    ("Rashodi > Trening",        "Vrsta",        "suggest"),
    ("Rashodi > Trening",        "Osoba",        "suggest"),
    # Pretplate (osobne digitalne)
    ("Rashodi > Pretplate",      "Naziv",        "suggest"),
    # Razvoj (Saša: AI alati + AI ručkovi)
    ("Rashodi > Razvoj",         "Vrsta",        "suggest"),
    ("Rashodi > Razvoj",         "Naziv",        "text"),
    # Kupovina
    ("Rashodi > Kupovina",       "Vrsta",        "suggest"),
    ("Rashodi > Kupovina",       "Opis",         "text"),
    # Telekomunikacije
    ("Rashodi > Telekomunikacije","Operater",    "suggest"),
    # Rate
    ("Rashodi > Rate",           "Naziv",        "text"),
    ("Rashodi > Rate",           "Rata",         "text"),
    # Porezi
    ("Rashodi > Porezi",         "Vrsta",        "suggest"),
    # Putovanje
    ("Rashodi > Putovanje",      "Vrsta",        "suggest"),
    ("Rashodi > Putovanje",      "Destinacija",  "text"),
    # Ostalo
    ("Rashodi > Ostalo",         "Opis",         "text"),
    # Transferi
    ("Transferi",                "Iznos",        "number"),
    ("Transferi",                "Izvor",        "suggest"),
    ("Transferi",                "Napomena",     "text"),
    ("Transferi",                "Odredište",    "suggest"),
    ("Transferi",                "Valuta",       "suggest"),
]

FIXED_COUNT    = 8
ATTR_COL_START = FIXED_COUNT + 1  # col I

ATTR_KEY_TO_COL = {
    (cat, attr): ATTR_COL_START + i
    for i, (cat, attr, _) in enumerate(ATTR_COLS)
}

def ancestors(path):
    parts = [p.strip() for p in path.split('>')]
    result = []
    for i in range(1, len(parts) + 1):
        result.append(' > '.join(parts[:i]))
    return set(result)

def is_relevant(attr_cat_path, event_cat_path):
    return attr_cat_path in ancestors(event_cat_path)

# ── Payment method → account label ────────────────────────────────────────────
NACIN_TO_RACUN = {
    'Master': 'Mastercard (Koka)',
    'Visa':   'Visa (Saša)',
    'Zaba':   'Kokin tekući (Zaba)',
    'RF':     'Raiffeisen (Saša)',
}

def detect_osoba_from_nacin(nacin):
    """Koka pays with Master/Zaba; Saša pays with RF/Visa."""
    if nacin in ('Master', 'Zaba'): return 'Koka'
    if nacin in ('RF', 'Visa'):     return 'Saša'
    return None


# ── CATEGORIZE ────────────────────────────────────────────────────────────────
def categorize(sto, nacin, iznos):
    """Returns (category_path, attrs_dict, comment)."""
    sto   = str(sto).strip()   if sto   else ''
    nacin = str(nacin).strip() if nacin else ''
    iznos_f = float(iznos) if iznos is not None else 0.0
    racun = NACIN_TO_RACUN.get(nacin, nacin)

    # ── TRANSFERI ────────────────────────────────────────────────────────────
    card_dest = {
        'Visa': 'Visa (Saša)', 'Master': 'Mastercard (Koka)',
        'Mastercard': 'Mastercard (Koka)',
    }
    if sto in card_dest:
        return ('Transferi', {
            ('Transferi','Iznos'): iznos, ('Transferi','Valuta'): 'EUR',
            ('Transferi','Izvor'): racun, ('Transferi','Odredište'): card_dest[sto],
        }, sto)
    if sto == 'Koka':
        return ('Transferi', {
            ('Transferi','Iznos'): iznos, ('Transferi','Valuta'): 'EUR',
            ('Transferi','Izvor'): racun, ('Transferi','Odredište'): 'Kokin tekući (Zaba)',
            ('Transferi','Napomena'): 'Saša → Koka',
        }, sto)
    if sto in ('Gotovina', 'Cash', 'Sašin novčanik'):
        return ('Transferi', {
            ('Transferi','Iznos'): iznos, ('Transferi','Valuta'): 'EUR',
            ('Transferi','Izvor'): racun, ('Transferi','Odredište'): 'Gotovina',
            ('Transferi','Napomena'): sto,
        }, sto)

    # ── Rashodi base attrs (L1, inherited by all Rashodi leaves) ──────────────
    base = {
        ('Rashodi','Iznos'): iznos,
        ('Rashodi','Račun'): racun,
        ('Rashodi','Valuta'): 'EUR',
    }

    # ── PUTOVANJE ────────────────────────────────────────────────────────────
    if sto in ('Amsterdam', 'Amsteradam'):
        p = 'Rashodi > Putovanje'
        return (p, {**base, (p,'Destinacija'):'Amsterdam', (p,'Vrsta'):'Ostalo'}, sto)
    if sto == 'Aviokarte':
        p = 'Rashodi > Putovanje'
        return (p, {**base, (p,'Vrsta'):'Aviokarte'}, sto)
    if sto == 'Hotel':
        p = 'Rashodi > Putovanje'
        return (p, {**base, (p,'Vrsta'):'Hotel'}, sto)

    # ── RAZVOJ (Saša: AI tools + AI ručkovi) ─────────────────────────────────
    razvoj_pretplate = {
        'Claude': 'Claude', 'Perplexity': 'Perplexity',
    }
    if sto in razvoj_pretplate:
        p = 'Rashodi > Razvoj'
        return (p, {**base, (p,'Vrsta'):'Pretplata', (p,'Naziv'):razvoj_pretplate[sto]}, sto)

    if sto == 'Biberon':
        # 5–20 EUR → AI ručak u firmi Atomic Intelligence; van tog raspona → ljekarna
        if 5.0 <= iznos_f <= 20.0:
            p = 'Rashodi > Razvoj'
            return (p, {**base, (p,'Vrsta'):'Ručak', (p,'Naziv'):'AI ručak'}, sto)
        else:
            p = 'Rashodi > Zdravlje'
            return (p, {**base, (p,'Vrsta'):'Ljekarna'}, sto)

    # ── SVAKODNEVNI ───────────────────────────────────────────────────────────
    hrana_ducani = {
        'Konzum':'Konzum', 'Konzum dostava':'Konzum', 'Spar':'Spar',
        'Studenac':'Studenac', 'Lidl':'Lidl', 'Mlinar':'Mlinar',
        'Voćarna':'Voćarna', 'Bofrost':'Bofrost', 'Igomat':'Igomat',
        'Pekara':'Pekara', 'Kraš':'Kraš', 'Nespresso':'Nespresso',
        'McDonalds':'McDonalds', 'Kruh':None, 'Hrana':None,
    }
    if sto in hrana_ducani:
        p = 'Rashodi > Svakodnevni'
        a = {**base, (p,'Vrsta'):'Hrana'}
        if hrana_ducani[sto]: a[(p,'Dućan')] = hrana_ducani[sto]
        return (p, a, sto)

    if sto == 'DM':
        p = 'Rashodi > Svakodnevni'
        return (p, {**base, (p,'Vrsta'):'Drogerija', (p,'Dućan'):'DM'}, sto)

    # ── RESTORAN ──────────────────────────────────────────────────────────────
    restoran_map = {
        'Pizzeria':'Pizzeria', 'Afrodita':'Afrodita', 'Dubravica':'Dubravica',
        'Vidikovac':'Vidikovac', 'Fisherija':'Fisherija',
        'Restoran Time':'Restoran Time', 'Veronika':'Veronika',
        'Maslina':'Maslina', 'Picek':'Picek', 'Nautic pizza':'Nautic pizza',
        'Mullef':'Mullef', 'Chipoteka':'Chipoteka', 'Batak':'Batak',
        'Kava': None,
    }
    if sto in restoran_map:
        p = 'Rashodi > Restoran'
        a = {**base}
        if restoran_map[sto]: a[(p,'Naziv')] = restoran_map[sto]
        return (p, a, sto)

    # ── DOM (komunalije, pričuva, holding, osiguranje nekretnine) ─────────────
    dom_map = {
        'Bulatova HEP':    ('HEP',      'Bulatova'),
        'Bulatova plin':   ('Plin',     'Bulatova'),
        'Plin razlika':    ('Plin',     'Bulatova'),
        'Electrocoin':     ('Electrocoin', 'Kućište'),
        'Saša Holding':    ('Holding',  'Saša'),
        'Nataša Holding':  ('Holding',  'Nataša'),
        'Medulićeva':      ('Pričuva',  'Medulićeva'),
    }
    if sto in dom_map:
        p = 'Rashodi > Dom'
        v, o = dom_map[sto]
        return (p, {**base, (p,'Vrsta'):v, (p,'Objekt'):o}, sto)

    # ── PRIJEVOZ ──────────────────────────────────────────────────────────────
    # Gorivo: Koka (Zaba/Master) → Citroen diesel; Saša (RF/Visa) → Lacetti benzin
    if sto in ('Ina', 'Petrol', 'Gorivo', 'Gorivo dizel'):
        p = 'Rashodi > Prijevoz'
        if nacin in ('Zaba', 'Master'):
            vrsta, vozilo = 'Gorivo diesel', 'Citroen'
        else:
            vrsta, vozilo = 'Gorivo benzin', 'Lacetti'
        return (p, {**base, (p,'Vrsta'):vrsta, (p,'Vozilo'):vozilo}, sto)

    prijevoz_vrsta = {
        'Parking':'Parking', 'Mjesečni parking':'Parking', 'Lacetti parking':'Parking',
        'Taxi':'Taxi', 'Bolt':'Bolt',
        'Carglass':'Servis', 'Tehnički C5':'Servis', 'Šatrak':'Servis',
        'HAK':'HAK',
    }
    if sto in prijevoz_vrsta:
        p = 'Rashodi > Prijevoz'
        return (p, {**base, (p,'Vrsta'):prijevoz_vrsta[sto]}, sto)

    # ── ZDRAVLJE ──────────────────────────────────────────────────────────────
    zdravlje_vrsta = {
        'Ljekarna':'Ljekarna', 'Ljekarma':'Ljekarna',   # Ljekarma = typo
        'D-vitamin':'Ljekarna', 'Yasenka':'Ljekarna', 'Magnezij':'Ljekarna',
        'Kreatin':'Ljekarna', 'Lijekovi za mamu':'Ljekarna',
        'HLK':'HLK',
        'Naočale':'Optika',
        'Synlab':'Liječnik',
    }
    if sto in zdravlje_vrsta:
        p = 'Rashodi > Zdravlje'
        return (p, {**base, (p,'Vrsta'):zdravlje_vrsta[sto]}, sto)

    # PassSport/Passport variants → Zdravlje, Vrsta: Passport
    # SS prefix = Saša, DPS prefix = Koka; otherwise detect from nacin
    passport_re = re.compile(r'[Pp]ass[Ss]?port', re.I)
    if passport_re.search(sto):
        p = 'Rashodi > Zdravlje'
        a = {**base, (p,'Vrsta'):'Passport'}
        if re.search(r'\bSS\b', sto):        a[(p,'Osoba')] = 'Saša'
        elif re.search(r'\bDPS\b', sto):     a[(p,'Osoba')] = 'Koka'
        else:
            osoba = detect_osoba_from_nacin(nacin)
            if osoba: a[(p,'Osoba')] = osoba
        return (p, a, sto)

    # Životno osiguranje (Allianz, Generali)
    if re.search(r'[Aa]llianz', sto) or re.search(r'[Gg]enerali', sto):
        p = 'Rashodi > Zdravlje'
        return (p, {**base, (p,'Vrsta'):'Životno'}, sto)

    # Mirovinski fond (PP ...)
    if sto.startswith('PP '):
        p = 'Rashodi > Zdravlje'
        a = {**base, (p,'Vrsta'):'Mirovinski'}
        if re.search(r'\bSS\b', sto) or 'Saša' in sto: a[(p,'Osoba')] = 'Saša'
        elif re.search(r'\bDPS\b', sto) or 'Koka' in sto: a[(p,'Osoba')] = 'Koka'
        return (p, a, sto)

    # ── PRETPLATE (osobne digitalne) ──────────────────────────────────────────
    # Note: Claude → Razvoj (above); ove su osobne pretplate
    pretplate_map = {
        'Youtube':'YouTube', 'HBO':'HBO', 'Disney':'Disney+',
        'Spotify':'Spotify',
        'Apple':'Apple', 'Apple Cloud':'Apple', 'iCloude':'Apple', 'Cloud':'Apple',
        'Sky':'SkyShowtime', 'SkyShowtime':'SkyShowtime', 'Skyshow':'SkyShowtime',
        'Google':'Google',
        'Audible SS':'Audible', 'Audible DPS':'Audible', 'DPS audible':'Audible',
        'Audible':'Audible', 'SS audible':'Audible',
        'Jutarnji list':'Jutarnji list',
        'Netdomena Igor':'Netdomena', 'Netdomena':'Netdomena',
        'Amazon Prime':'PrimeVideo', 'PrimeVideo':'PrimeVideo', 'Prime Video':'PrimeVideo',
    }
    if sto in pretplate_map:
        p = 'Rashodi > Pretplate'
        return (p, {**base, (p,'Naziv'):pretplate_map[sto]}, sto)

    # ── TELEKOMUNIKACIJE ──────────────────────────────────────────────────────
    if sto in ('T-com', 'T-mobile', 'A1'):
        p = 'Rashodi > Telekomunikacije'
        return (p, {**base, (p,'Operater'):sto}, sto)

    # ── RATE (pattern "Naziv X/Y") ────────────────────────────────────────────
    rate_m = re.match(r'^(.+?)\s+(\d+/\d+)$', sto)
    if rate_m:
        naziv = rate_m.group(1).strip()
        rata  = rate_m.group(2)
        p = 'Rashodi > Rate'
        return (p, {**base, (p,'Naziv'):naziv, (p,'Rata'):rata}, sto)

    # ── KUPOVINA ──────────────────────────────────────────────────────────────
    kupovina_vrsta = {
        'Temu':'Online', 'Decathlon':'Sport', 'Nordletics':'Sport',
        'Galeb gaće':'Odjeća', 'H&M':'Odjeća',
        'Ikea':'Kućanstvo', 'Čaše':'Kućanstvo', 'Purex':'Kućanstvo',
        'Gitara':'Ostalo', 'Miš za komp':'Ured', 'Korica':'Ostalo',
        'Myprotein':'Sport', 'Body shop':'Ostalo',
        'Video game museum':'Ostalo', 'Cinestar':'Ostalo',
        'GLS':'Ostalo', 'Paket':'Ostalo', 'Tisak':'Ostalo',
        'Sljeme':'Sport', 'Kreatin':'Ostalo',  # backup if not caught above
        'Igor':'Dar',
        'HP':'Ured',     # HP toner/printer
    }
    if sto in kupovina_vrsta:
        p = 'Rashodi > Kupovina'
        a = {**base, (p,'Vrsta'):kupovina_vrsta[sto], (p,'Opis'):sto}
        return (p, a, sto)

    # ── BANKOVNA NAKNADA → Ostalo ─────────────────────────────────────────────
    if sto in ('Naknada', 'RF naknada', 'Zaba naknada', 'HP naknada'):
        p = 'Rashodi > Ostalo'
        return (p, {**base, (p,'Opis'):'Bankovna naknada'}, sto)

    # ── OSTALO (catch-all) ────────────────────────────────────────────────────
    p = 'Rashodi > Ostalo'
    return (p, {**base, (p,'Opis'):sto}, sto)


# ── Read source ────────────────────────────────────────────────────────────────
src_wb = openpyxl.load_workbook(INPUT, data_only=True)
src_ws = src_wb['Za Sašu']

rows = []
for r in range(2, src_ws.max_row + 1):
    datum  = src_ws.cell(r, 1).value
    nacin  = src_ws.cell(r, 2).value
    sto    = src_ws.cell(r, 3).value
    iznos  = src_ws.cell(r, 4).value
    if not datum or not sto: continue
    rows.append((datum, nacin, sto, iznos))

print(f"Source rows: {len(rows)}")

# ── Build output workbook ──────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Events'

row = 1

# LEGEND title
c = ws.cell(row, 1, 'ATTRIBUTE LEGEND:')
c.font = TITLE_FONT
ws.cell(row, 3).value = 'see Structure sheet for more details'
ws.cell(row, 3).font  = Font(italic=True, color='666666')
row += 1

# LEGEND header
for ci, h in enumerate(['Col','Area','Category_Path','Attribute','Type','Unit'], 1):
    c = ws.cell(row, ci, h)
    c.fill = LEG_FILL; c.font = WHITE_FONT; c.border = BORDER; c.alignment = C_ALIGN
row += 1

# LEGEND rows
legend_start = row
for i, (cat_path, attr_name, dtype) in enumerate(ATTR_COLS):
    letter = col_letter(ATTR_COL_START + i)
    is_sep = (i == 0) or (cat_path != ATTR_COLS[i-1][0])
    unit   = 'EUR' if (attr_name == 'Iznos' and dtype == 'number') else ''
    data   = [letter, AREA, cat_path, attr_name, dtype, unit]
    for ci, v in enumerate(data, 1):
        c = ws.cell(row, ci, v if v else None)
        c.fill   = SEP_FILL if is_sep else PINK_FILL
        c.font   = BOLD_FONT if is_sep else Font()
        c.border = BORDER
        c.alignment = L_ALIGN
    row += 1
legend_end = row - 1

for r_idx in range(legend_start, legend_end + 1):
    i = r_idx - legend_start
    cat_path = ATTR_COLS[i][0]
    is_sep = (i == 0) or (cat_path != ATTR_COLS[i-1][0])
    if not is_sep:
        ws.row_dimensions[r_idx].outlineLevel = 1
        ws.row_dimensions[r_idx].hidden = True

row += 1  # blank separator

# EVENT DATA title
event_title_row = row
ws.cell(row, 1, 'EVENT DATA:').font = TITLE_FONT
ws.cell(row, 3, 'Sum (if relevant) ->').alignment = Alignment(horizontal='right')
row += 1

# EVENT DATA header
event_header_row = row
fixed_hdrs = ['event_id','Area','Category_Path','event_date','session_start','created_at','User','leaf comment']
attr_hdrs  = [f"{attr} ({cat.split(' > ')[-1]})" for cat, attr, _ in ATTR_COLS]
for ci, h in enumerate(fixed_hdrs + attr_hdrs, 1):
    c = ws.cell(row, ci, h)
    c.fill = HEADER_FILL; c.font = WHITE_FONT; c.border = BORDER; c.alignment = C_ALIGN
row += 1

event_data_start = row

# DATA ROWS
unmapped = {}
for datum, nacin, sto, iznos in rows:
    cat_path, attrs, comment = categorize(sto, nacin, iznos)

    ev_date = parse_datum(datum)
    if ev_date is None:
        print(f"  SKIP (bad date): row datum={datum!r}  sto={sto!r}")
        continue

    fixed_vals = [
        None, AREA, cat_path, ev_date, '09:00', '09:00:01', '', comment,
    ]
    for ci, v in enumerate(fixed_vals, 1):
        c = ws.cell(row, ci, v)
        c.border = BORDER; c.alignment = L_ALIGN
        if ci in (1, 2, 3, 7):
            c.fill = PINK_FILL
        elif ci == 4:
            c.fill = BLUE_FILL
            c.number_format = 'YYYY-MM-DD'
        else:
            c.fill = BLUE_FILL

    for i, (ac_path, attr_name, dtype) in enumerate(ATTR_COLS):
        col_num  = ATTR_COL_START + i
        val      = attrs.get((ac_path, attr_name))
        relevant = is_relevant(ac_path, cat_path)
        c = ws.cell(row, col_num, val)
        c.border    = BORDER
        c.fill      = BLUE_FILL if relevant else ORANGE_FILL
        c.alignment = R_ALIGN if dtype == 'number' else L_ALIGN
        if dtype == 'number' and val is not None:
            c.number_format = '0.##'

    if cat_path == 'Rashodi > Ostalo':
        unmapped[sto] = unmapped.get(sto, 0) + 1

    ws.row_dimensions[row].height = 18
    row += 1

event_data_end = row - 1

# SUBTOTAL formulas
for i, (ac_path, attr_name, dtype) in enumerate(ATTR_COLS):
    if dtype != 'number': continue
    col_num = ATTR_COL_START + i
    letter  = col_letter(col_num)
    c = ws.cell(event_title_row, col_num)
    c.value     = f'=SUBTOTAL(9,{letter}{event_data_start}:{letter}{event_data_end})'
    c.alignment = R_ALIGN

ws.auto_filter.ref = f"A{event_header_row}:{col_letter(ATTR_COL_START+len(ATTR_COLS)-1)}{event_data_end}"
ws.freeze_panes   = ws.cell(event_data_start, ATTR_COL_START)

widths = {'A':10,'B':14,'C':26,'D':12,'E':9,'F':10,'G':22,'H':28}
for k, w in widths.items():
    ws.column_dimensions[k].width = w
for i in range(len(ATTR_COLS)):
    ws.column_dimensions[col_letter(ATTR_COL_START + i)].width = 13
ws.column_dimensions['G'].outlineLevel = 1


# ── STRUCTURE SHEET ────────────────────────────────────────────────────────────
ss = wb.create_sheet('Structure')

STR_HDRS = ['Type','CategoryPath','Sort','AttrName','Slug','AttrType',
            'IsRequired','Val.Type','Default','ValMax','Unit','TextOptions',
            'DependsOn','WhenValue','Description']

RACUN_OPTS = ('Kokin tekući (Zaba)|Mastercard (Koka)|Sašin tekući (PBZ)'
              '|Visa (Saša)|Raiffeisen (Saša)|Gotovina|KEKS (Saša)')

# (type, fullPath, attrName, slug, attrType, valType, default, unit, textOptions)
STRUCT = [
    # ── Area ──────────────────────────────────────────────────────────────────
    ('Area',      f'{AREA}',                                  '','','','','','',''),

    # ── Prihodi ───────────────────────────────────────────────────────────────
    ('Category',  f'{AREA} > Prihodi',                        '','','','','','',''),
    ('Attribute', f'{AREA} > Prihodi',  'Iznos',  'iznos',  'number',  'none',    '',    'EUR', ''),
    ('Attribute', f'{AREA} > Prihodi',  'Račun',  'racun',  'text', 'suggest', '',    '',    RACUN_OPTS),
    ('Attribute', f'{AREA} > Prihodi',  'Valuta', 'valuta', 'text', 'suggest', 'EUR', '',    'EUR|HRK|USD'),

    ('Category',  f'{AREA} > Prihodi > Plaća i mirovina',     '','','','','','',''),
    ('Attribute', f'{AREA} > Prihodi > Plaća i mirovina', 'Vrsta', 'vrsta', 'text','suggest','','',
     'Plaća|Mirovina|Prijevoz|Prehrana|Regres|Bonus|Božićnica|I stup|II stup|III stup'),
    ('Attribute', f'{AREA} > Prihodi > Plaća i mirovina', 'Osoba', 'osoba', 'text','suggest','','','Saša|Koka'),

    ('Category',  f'{AREA} > Prihodi > Najam — Anja',         '','','','','','',''),
    ('Attribute', f'{AREA} > Prihodi > Najam — Anja', 'Rata', 'rata', 'text','none','','',''),

    ('Category',  f'{AREA} > Prihodi > Ostali prihodi',       '','','','','','',''),
    ('Attribute', f'{AREA} > Prihodi > Ostali prihodi', 'Vrsta', 'vrsta', 'text','suggest','','','Honorar|Povrat|Ostalo'),
    ('Attribute', f'{AREA} > Prihodi > Ostali prihodi', 'Izvor', 'izvor', 'text','none','','',''),

    # ── Rashodi ───────────────────────────────────────────────────────────────
    ('Category',  f'{AREA} > Rashodi',                        '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi',  'Iznos',  'iznos',  'number', 'none',    '',    'EUR', ''),
    ('Attribute', f'{AREA} > Rashodi',  'Račun',  'racun',  'text',   'suggest', '',    '',    RACUN_OPTS),
    ('Attribute', f'{AREA} > Rashodi',  'Valuta', 'valuta', 'text',   'suggest', 'EUR', '',    'EUR|HRK|USD'),

    ('Category',  f'{AREA} > Rashodi > Dom',                  '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Dom', 'Vrsta',  'vrsta',  'text','suggest','','',
     'HEP|Plin|Voda|Grijanje|Struja|Holding|Pričuva|Osiguranje|Electrocoin'),
    ('Attribute', f'{AREA} > Rashodi > Dom', 'Objekt', 'objekt', 'text','suggest','','',
     'Bulatova|Medulićeva|Kućište|Nena|Mama'),

    ('Category',  f'{AREA} > Rashodi > Svakodnevni',          '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Svakodnevni', 'Vrsta', 'vrsta', 'text','suggest','','',
     'Hrana|Drogerija|Kiosk|Ostalo'),
    ('Attribute', f'{AREA} > Rashodi > Svakodnevni', 'Dućan', 'ducan', 'text','suggest','','',
     'Konzum|Spar|Studenac|Lidl|Mlinar|Temu|DM|McDonalds|Bofrost|Igomat|Pekara|Kraš|Nespresso|Voćarna'),

    ('Category',  f'{AREA} > Rashodi > Restoran',             '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Restoran', 'Naziv', 'naziv', 'text','none','','',''),

    ('Category',  f'{AREA} > Rashodi > Prijevoz',             '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Prijevoz', 'Vrsta',  'vrsta',  'text','suggest','','',
     'Gorivo diesel|Gorivo benzin|Parking|Taxi|HAK|Osiguranje|Servis|Registracija|Bolt'),
    ('Attribute', f'{AREA} > Rashodi > Prijevoz', 'Vozilo', 'vozilo', 'text','suggest','','',
     'Citroen|Lacetti|Ostalo'),

    ('Category',  f'{AREA} > Rashodi > Zdravlje',             '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Zdravlje', 'Vrsta', 'vrsta', 'text','suggest','','',
     'Ljekarna|Liječnik|HLK|Optika|Passport|Dopunsko|Životno|Mirovinski'),
    ('Attribute', f'{AREA} > Rashodi > Zdravlje', 'Osoba', 'osoba', 'text','suggest','','',
     'Saša|Koka|Nena|Mama'),

    ('Category',  f'{AREA} > Rashodi > Trening',              '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Trening', 'Vrsta', 'vrsta', 'text','suggest','','',
     'Fitnes|Bazen|Tenis|Ostalo'),
    ('Attribute', f'{AREA} > Rashodi > Trening', 'Osoba', 'osoba', 'text','suggest','','',
     'Saša|Koka'),

    ('Category',  f'{AREA} > Rashodi > Pretplate',            '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Pretplate', 'Naziv', 'naziv', 'text','suggest','','',
     'YouTube|HBO|Disney+|Spotify|Audible|Apple|Netflix|SkyShowtime|Google|Jutarnji list|Netdomena|PrimeVideo'),

    ('Category',  f'{AREA} > Rashodi > Razvoj',               '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Razvoj', 'Vrsta', 'vrsta', 'text','suggest','','',
     'Pretplata|Ručak|Alat'),
    ('Attribute', f'{AREA} > Rashodi > Razvoj', 'Naziv', 'naziv', 'text','none','','',''),

    ('Category',  f'{AREA} > Rashodi > Kupovina',             '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Kupovina', 'Vrsta', 'vrsta', 'text','suggest','','',
     'Online|Odjeća|Sport|Dar|Ured|Kućanstvo|Ostalo'),
    ('Attribute', f'{AREA} > Rashodi > Kupovina', 'Opis',  'opis',  'text','none','','',''),

    ('Category',  f'{AREA} > Rashodi > Telekomunikacije',     '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Telekomunikacije', 'Operater','operater','text','suggest','','',
     'T-com|T-mobile|A1'),

    ('Category',  f'{AREA} > Rashodi > Rate',                 '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Rate', 'Naziv', 'naziv', 'text','none','','',''),
    ('Attribute', f'{AREA} > Rashodi > Rate', 'Rata',  'rata',  'text','none','','',''),

    ('Category',  f'{AREA} > Rashodi > Porezi',               '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Porezi', 'Vrsta', 'vrsta', 'text','suggest','','',
     'Porez i prirez|APN|Porez na dohodak|Porez na nekretninu|Komunalni doprinos'),

    ('Category',  f'{AREA} > Rashodi > Putovanje',            '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Putovanje', 'Vrsta',       'vrsta',       'text','suggest','','',
     'Aviokarte|Hotel|Prijevoz|Hrana|Ostalo'),
    ('Attribute', f'{AREA} > Rashodi > Putovanje', 'Destinacija', 'destinacija', 'text','none','','',''),

    ('Category',  f'{AREA} > Rashodi > Ostalo',               '','','','','','',''),
    ('Attribute', f'{AREA} > Rashodi > Ostalo', 'Opis', 'opis', 'text','none','','',''),

    # ── Transferi ─────────────────────────────────────────────────────────────
    ('Category',  f'{AREA} > Transferi',                      '','','','','','',''),
    ('Attribute', f'{AREA} > Transferi', 'Iznos',     'iznos',     'number', 'none',    '',    'EUR', ''),
    ('Attribute', f'{AREA} > Transferi', 'Izvor',     'izvor',     'text',   'suggest', '',    '',    RACUN_OPTS),
    ('Attribute', f'{AREA} > Transferi', 'Napomena',  'napomena',  'text',   'none',    '',    '',    ''),
    ('Attribute', f'{AREA} > Transferi', 'Odredište', 'odrediste', 'text',   'suggest', '',    '',    RACUN_OPTS),
    ('Attribute', f'{AREA} > Transferi', 'Valuta',    'valuta',    'text',   'suggest', 'EUR', '',    'EUR|HRK|USD'),
]

for ci, h in enumerate(STR_HDRS, 1):
    c = ss.cell(1, ci, h)
    c.fill = HEADER_FILL; c.font = WHITE_FONT; c.border = BORDER; c.alignment = C_ALIGN

for ri, (typ, path, aname, slug, atype, valtype, defval, unit, textopts) in enumerate(STRUCT, 2):
    row_data = [typ, path, '', aname, slug, atype, '', valtype, defval, '', unit, textopts, '', '', '']
    for ci, v in enumerate(row_data, 1):
        c = ss.cell(ri, ci, v if v else None)
        c.border = BORDER; c.alignment = L_ALIGN

ss.column_dimensions['A'].width = 12
ss.column_dimensions['B'].width = 50
ss.column_dimensions['D'].width = 16
ss.column_dimensions['E'].width = 16
ss.column_dimensions['F'].width = 10
ss.column_dimensions['H'].width = 10
ss.column_dimensions['L'].width = 60

wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")
print(f"Events rows: {event_data_end - event_data_start + 1}")
print(f"Attr columns: {len(ATTR_COLS)} (I–{col_letter(ATTR_COL_START+len(ATTR_COLS)-1)})")
print(f"Structure rows: {len(STRUCT)} ({sum(1 for r in STRUCT if r[0]=='Area')} areas, "
      f"{sum(1 for r in STRUCT if r[0]=='Category')} categories, "
      f"{sum(1 for r in STRUCT if r[0]=='Attribute')} attributes)")
print(f"\nOstalo (catch-all) — provjeri mapiranje:")
for k, v in sorted(unmapped.items(), key=lambda x: -x[1]):
    print(f"  {v:3}x  {k!r}")
