# -*- coding: utf-8 -*-
"""
make_financije3_import.py
=========================
Generira xlsx za import u Events Tracker — Area 'Financije_3'.

Izvor: Claude-temp_R/Financije 2026-06.xlsx (sheetovi: koka EU, sasa EU)

Struktura:
  L1: Transakcija  — Racun, Uplata, Isplata, Stanje, Valuta
  L2: Kategorija   — Napomena, Smjer, Tip  (leaf)

Svaki red = jedna bankovna transakcija = jedna sesija.

Datum filter: MIN_DATE (2023-01-01) — MAX_DATE (2026-12-31)
"""

import sys, re, os
from pathlib import Path
from datetime import date, datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

sys.stdout.reconfigure(encoding='utf-8')

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
INPUT      = Path(r"C:\0_Sasa\events-tracker-react\Claude-temp_R\Financije 2026-06.xlsx")
OUTPUT_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")
OUTPUT     = OUTPUT_DIR / "Financije3_import.xlsx"

# ── Config ─────────────────────────────────────────────────────────────────────
AREA       = "Financije_3"
L1         = "Transakcija"
L2         = "Kategorija"
USER_EMAIL = "sasasladoljev59@gmail.com"
MIN_DATE   = date(2023, 1, 1)
MAX_DATE   = date(2026, 12, 31)

# ── Source sheets ──────────────────────────────────────────────────────────────
SHEETS = {
    'koka EU': 'Kokin tekući ZABA',
    'sasa EU': 'Sašin tekući RF',
}

# ── Tip keyword mapping (order matters — first match wins) ────────────────────
TIP_MAP = [
    ('Mirovina',   ['mirovina']),
    ('PP',         [' pp ', 'pp ', 'posmrtna']),
    ('Rate',       [r'\d+/\d+']),            # regex: "41/96", "6/12"
    ('Prevoz',     ['parking', 'taxi', 'gorivo', 'benzin', 'carglass',
                    'uber', 'dizel', 'hak', 'autostrada', 'prijevoz']),
    ('Dom/hrana',  ['konzum', 'kruh', 'plodine', 'lidl', 'kaufland', 'spar',
                    'interspar', 'temu', 'hrana', 'pizzeria', 'restoran',
                    'dostava', 'tommy', 'billa', 'dm']),
    ('Zdravlje',   ['lijek', 'ljekarnica', 'doktor', 'ordinacija',
                    'zdravlje', 'dopunsko', 'mama', 'mamu']),
    ('Zabava',     ['audible', 'sky', 'netflix', 'spotify', 'multisport',
                    'passport', 'claude', 'amazon', 'hp', 'steam',
                    'disney', 'kreatin', 'audible']),
    ('Putovanje',  ['amsterdam', 'putovanje', 'hotel', 'airbnb',
                    'booking', 'let ', 'avionske']),
    ('Povrat',     ['povrat', 'refund']),
    ('Naknada',    ['naknada', 'e-zaba']),
    ('Transfer',   ['transfer', 'visa', 'mastercard']),
]

TIP_OPTIONS = (
    'Dom/hrana|Prevoz|Zdravlje|Zabava|Putovanje|Mirovina|'
    'Povrat|Naknada|Rate|PP|Transfer|Ostalo|N/A'
)

# ── Styles ─────────────────────────────────────────────────────────────────────
PURPLE_FILL = PatternFill("solid", fgColor="7030A0")
BLUE_FILL   = PatternFill("solid", fgColor="4472C4")
AMBER_FILL  = PatternFill("solid", fgColor="FFC000")
GRAY_FILL   = PatternFill("solid", fgColor="D9D9D9")
WHITE_FONT  = Font(color="FFFFFF", bold=True)
BOLD_FONT   = Font(bold=True)
THIN        = Side(style='thin')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Attribute definitions ──────────────────────────────────────────────────────
# (cat_path_no_area, attr_name, data_type, unit, val_type, options, default, description)
ATTRS = [
    # L1: Transakcija
    (L1, 'Racun',   'text',   '',    'suggest',
     'Kokin tekući ZABA|Sašin tekući RF', '', ''),
    (L1, 'Uplata',  'number', 'EUR', '',        '', '', ''),
    (L1, 'Isplata', 'number', 'EUR', '',        '', '', ''),
    (L1, 'Stanje',  'number', 'EUR', '',        '', '', 'Stanje računa nakon transakcije'),
    (L1, 'Valuta',  'text',   '',    'suggest', 'EUR|HRK|USD', 'EUR', ''),
    # L2: Kategorija (leaf)
    (f'{L1} > {L2}', 'Napomena', 'text', '', '',        '', '', 'Originalni opis iz bankovnog izvoda'),
    (f'{L1} > {L2}', 'Smjer',    'text', '', 'suggest', 'Uplata|Isplata|PROVJERI', '', ''),
    (f'{L1} > {L2}', 'Tip',      'text', '', 'suggest', TIP_OPTIONS, 'N/A', ''),
]

# Fixed columns before attributes: event_id, Area, Cat_Path, date, time, created, user, comment
FIXED = 8
ATTR_COL_START = FIXED + 1  # col I = index 9

# ── Helpers ────────────────────────────────────────────────────────────────────

def col_letter(n: int) -> str:
    s = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def parse_datum(val) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if val is None:
        return None
    s = str(val).strip().rstrip('.')
    m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d+)$', s)
    if m:
        d, mo, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 100:
            yr += 2000
        elif yr > 2100:
            yr = (yr % 100) + 2000
        try:
            return date(yr, mo, d)
        except ValueError:
            pass
    return None


def classify_tip(opis: str | None) -> str:
    if not opis:
        return 'N/A'
    low = ' ' + opis.lower() + ' '
    for tip, keywords in TIP_MAP:
        for kw in keywords:
            if kw.startswith(r'\d'):
                if re.search(kw, low):
                    return tip
            elif kw in low:
                return tip
    return 'N/A'


def get_smjer(uplata, isplata) -> str:
    u = uplata and float(uplata) > 0
    i = isplata and float(isplata) > 0
    if u and i:
        return 'PROVJERI'
    if u:
        return 'Uplata'
    if i:
        return 'Isplata'
    return 'PROVJERI'


# ── Structure sheet ────────────────────────────────────────────────────────────

def write_structure_sheet(ws):
    headers = [
        'Type', 'IsLeaf', 'Area', 'SharedWith', 'CategoryPath', 'Sort',
        'AttrName', 'Slug', 'AttrType', 'IsRequired', 'Val.Type', 'Default',
        'Val.Max (no)', 'Unit', 'TextOptions/Val.Min', 'DependsOn', 'WhenValue',
        'Description',
    ]
    # Header row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        ws.column_dimensions[col_letter(c)].width = 20

    def row(type_, cat_path='', sort='', attr='', slug='', atype='',
            required='', vtype='', default='', vmax='', unit='',
            options='', dep='', when='', desc=''):
        return [type_, '', '', '', cat_path, sort, attr, slug, atype,
                required, vtype, default, vmax, unit, options, dep, when, desc]

    rows = [
        row('Area', f'{AREA}'),
        row('Category', f'{AREA} > {L1}'),
    ]
    cat_sort: dict[str, int] = {}
    l2_inserted = False
    for cat, name, atype, unit, vtype, options, default, desc in ATTRS:
        full_path = f'{AREA} > {cat}'
        cat_sort[cat] = cat_sort.get(cat, 0) + 1
        sort_val = cat_sort[cat]
        if cat == f'{L1} > {L2}' and not l2_inserted:
            rows.append(row('Category', f'{AREA} > {L1} > {L2}'))
            l2_inserted = True
        rows.append(row('Attribute', full_path, sort_val, name, '', atype,
                        '', vtype, default, '', unit, options, '', '', desc))

    for r_idx, r_data in enumerate(rows, 2):
        for c_idx, val in enumerate(r_data, 1):
            ws.cell(r_idx, c_idx, val)


# ── Events sheet ───────────────────────────────────────────────────────────────

def write_events_sheet(ws, events):
    # ── LEGEND ────────────────────────────────────────────────────────────────
    ws.cell(1, 1, 'ATTRIBUTE LEGEND:')
    ws.cell(1, 1).font = BOLD_FONT

    leg_headers = ['Col', 'Area', 'Category_Path', 'Attribute', 'Type', 'Unit']
    for c, h in enumerate(leg_headers, 1):
        cell = ws.cell(2, c, h)
        cell.fill = PURPLE_FILL
        cell.font = WHITE_FONT

    for i, (cat, name, atype, unit, *_) in enumerate(ATTRS):
        r = 3 + i
        col_ltr = col_letter(ATTR_COL_START + i)
        display_type = 'suggest' if ATTRS[i][4] == 'suggest' else atype
        for c, val in enumerate([col_ltr, AREA, cat, name, display_type, unit], 1):
            cell = ws.cell(r, c, val)
            cell.fill = GRAY_FILL if i % 2 == 0 else PatternFill("solid", fgColor="EDE7F6")

    legend_end = 2 + len(ATTRS)

    # ── EVENT DATA ─────────────────────────────────────────────────────────────
    data_start = legend_end + 2
    ws.cell(data_start, 1, 'EVENT DATA:')
    ws.cell(data_start, 1).font = BOLD_FONT

    ev_headers = ['event_id', 'Area', 'Category_Path', 'event_date',
                  'session_start', 'created_at', 'User', 'leaf_comment']
    ev_headers += [name for _, name, *_ in ATTRS]

    hdr_row = data_start + 1
    for c, h in enumerate(ev_headers, 1):
        cell = ws.cell(hdr_row, c, h)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT

    leaf_path = f'{L1} > {L2}'
    for r_idx, ev in enumerate(events, hdr_row + 1):
        ws.cell(r_idx, 1, '')                          # event_id (empty = create)
        ws.cell(r_idx, 2, AREA)
        ws.cell(r_idx, 3, leaf_path)
        ws.cell(r_idx, 4, ev['date'])
        ws.cell(r_idx, 5, ev['time'])
        ws.cell(r_idx, 6, '')                          # created_at
        ws.cell(r_idx, 7, USER_EMAIL)
        ws.cell(r_idx, 8, '')                          # leaf comment
        # Attributes: Racun, Uplata, Isplata, Stanje, Valuta, Napomena, Smjer, Tip
        ws.cell(r_idx, 9,  ev['racun'])
        ws.cell(r_idx, 10, ev['uplata'])
        ws.cell(r_idx, 11, ev['isplata'])
        ws.cell(r_idx, 12, ev['stanje'])
        ws.cell(r_idx, 13, ev['valuta'])
        ws.cell(r_idx, 14, ev['napomena'])
        ws.cell(r_idx, 15, ev['smjer'])
        ws.cell(r_idx, 16, ev['tip'])


# ── Read source sheet ──────────────────────────────────────────────────────────

def read_sheet(ws, racun_label: str, skipped: list) -> list:
    events = []
    for r in range(2, ws.max_row + 1):
        opis    = ws.cell(r, 2).value
        datum   = ws.cell(r, 3).value
        uplata  = ws.cell(r, 4).value
        isplata = ws.cell(r, 5).value
        stanje  = ws.cell(r, 6).value

        d = parse_datum(datum)

        # Skip empty rows
        if d is None and uplata is None and isplata is None and opis is None:
            continue

        # Skip bad dates
        if d is None:
            skipped.append(f'  row {r}: bad date {repr(datum)!s}')
            continue

        # Skip out-of-range dates
        if d < MIN_DATE or d > MAX_DATE:
            skipped.append(f'  row {r}: out-of-range date {d}  opis={opis}')
            continue

        # Skip opening-balance rows (no amounts)
        if uplata is None and isplata is None:
            skipped.append(f'  row {r}: no Uplata+Isplata (balance row)  date={d}')
            continue

        # Normalize floats
        def num(v):
            if v is None:
                return None
            try:
                f = float(v)
                return round(f, 2)
            except (ValueError, TypeError):
                return None

        events.append({
            'date':     d.isoformat(),
            'date_obj': d,
            'racun':    racun_label,
            'uplata':   num(uplata),
            'isplata':  num(isplata),
            'stanje':   num(stanje),
            'valuta':   'EUR',
            'napomena': str(opis).strip() if opis else '',
            'smjer':    get_smjer(uplata, isplata),
            'tip':      classify_tip(str(opis) if opis else None),
        })
    return events


# ── Session time assignment ────────────────────────────────────────────────────

def assign_times(events: list):
    date_counter: dict[str, int] = {}
    for ev in events:
        d = ev['date']
        count = date_counter.get(d, 0)
        date_counter[d] = count + 1
        total_min = 9 * 60 + count   # start 09:00, +1 min each
        ev['time'] = f'{total_min // 60:02d}:{total_min % 60:02d}'


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    if not INPUT.exists():
        print(f'ERROR: source file not found: {INPUT}')
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb_src = openpyxl.load_workbook(INPUT, data_only=True)

    all_events = []
    all_skipped = []

    for sheet_name, racun_label in SHEETS.items():
        if sheet_name not in wb_src.sheetnames:
            print(f'WARNING: sheet "{sheet_name}" not found, skipping')
            continue
        ws = wb_src[sheet_name]
        skipped = []
        evts = read_sheet(ws, racun_label, skipped)
        all_events.extend(evts)
        all_skipped.extend([f'[{sheet_name}] {s}' for s in skipped])
        tip_na = sum(1 for e in evts if e['tip'] == 'N/A')
        tip_ok = len(evts) - tip_na
        print(f'{sheet_name}: {len(evts)} events  '
              f'(Tip classified: {tip_ok}, N/A: {tip_na}  |  skipped: {len(skipped)})')

    # Sort by date then source order, then assign unique session times
    all_events.sort(key=lambda e: e['date'])
    assign_times(all_events)

    # Tip distribution
    from collections import Counter
    tip_dist = Counter(e['tip'] for e in all_events)
    smjer_dist = Counter(e['smjer'] for e in all_events)

    # Build output workbook
    wb_out = openpyxl.Workbook()

    ws_events = wb_out.active
    ws_events.title = 'Activities Events'
    write_events_sheet(ws_events, all_events)

    ws_struct = wb_out.create_sheet('Structure')
    write_structure_sheet(ws_struct)

    wb_out.save(OUTPUT)

    # Summary
    print(f'\n{"="*60}')
    print(f'OUTPUT: {OUTPUT}')
    print(f'Total events: {len(all_events)}')
    print(f'\nTip distribution:')
    for tip, cnt in tip_dist.most_common():
        bar = '█' * (cnt // 10)
        print(f'  {tip:<15} {cnt:>4}  {bar}')
    print(f'\nSmjer: {dict(smjer_dist)}')
    if all_skipped:
        print(f'\nSkipped ({len(all_skipped)} rows):')
        for s in all_skipped[:20]:
            print(s)
        if len(all_skipped) > 20:
            print(f'  ... and {len(all_skipped) - 20} more')
    print(f'{"="*60}')


if __name__ == '__main__':
    main()
