"""
fix_duplikati_rata.py — jednokratni popravak 8 dupliranih rata (636,36 €). S107o.

NALAZ (S107n, ENRICH_PLAN §2l): kad Koka ratu vodi mjesečno, a izvod sve rate knjiži
na datum kupovine, rate 2..N postoje dvaput. Dedup (`merge_pbzvisa`) i v3 Verdikt pass
(±2 dana) to strukturno ne mogu uhvatiti — rata je mjesec dana odmaknuta. Detekcija ide
po `Datum naplate` + iznos (moguće tek otkad je 100 % popunjen, S107k).

DUP semantika (odluka Saša, ista kao S107k):
  • Kokin (ručni) redak OSTAJE i dobiva `Izvod opis` + `Izvod file` izvodnog (P3: samo ako su prazni)
  • izvodni redak se BRIŠE iz Reviewa, a njegov `source_key` ide u `V3 preskočeno`

Zašto skip zapis nije kozmetika: `merge_pbzvisa.py` je idempotentan tako što preskače
`source_key`eve KOJI POSTOJE u Reviewu. Obrisan redak = izgubljena idempotencija →
sljedeći run bi duplikat vratio. Zato je `merge_pbzvisa` dopunjen da čita `V3 preskočeno`
(isti registar koji `consolidate_review.py` već koristi za trajno odbačene transakcije).

Parovi se traže po `source_key` (stabilan identitet), NE po broju retka — retci se pomiču
pri svakom re-sortu. Svaki par se prije diranja provjerava po iznosu, `Datum naplate` i
očekivanom tekstu; bilo koje odstupanje = izlaz bez pisanja.

Pokretanje:
  python fix_duplikati_rata.py --dry     # popis, ništa se ne piše
  python fix_duplikati_rata.py           # pravi run (Excel ZATVOREN; backup .pre-duprata-*)
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
DATA = HERE.parent.parent / 'data-prep_data' / 'Financije'
SKIP_SHEET = 'V3 preskočeno'

# (source_key izvodnog, source_key Kokinog, iznos, dio Kokine Napomene) — iz §2l tablice.
# Redovi u komentaru su stanje 2026-07-27; provjera ide po ključu, ne po broju.
PAIRS = [
    ('0d9055789042', 'e85e5161754d',  24.50, 'Plodine 6/6'),        # 3665 ↔ 4271
    ('ef469f5b270a', 'b8bbee7dcd82', 157.03, 'AC Šatrak 5/6'),      # 3854 ↔ 4327
    ('ae7fd54dc33a', 'ea39aea22450', 157.03, 'Auto Šatrak 6/6'),    # 3855 ↔ 4450
    ('40d51e05d796', '682e142aea33',  62.50, 'Traperice 2/4'),      # 4418 ↔ 4528
    ('694adf915059', 'c6512f581fa7',  62.50, 'Traperice 3/4'),      # 4419 ↔ 4634
    ('512c6d7b4c34', 'd4c531f49752',  62.50, 'Traperice 3/4'),      # 4420 ↔ 4752
    ('ca7a5069d796', '6b615ec91887',  55.15, 'Reg C5 2/3'),         # 4505 ↔ 4609
    ('6c31d115afa8', '9daa305e1c2f',  55.15, 'Reg C5 2/3'),         # 4506 ↔ 4720
]
# Provjereni LAŽNI pozitivni — ovdje samo da ostane zapisano zašto ih nema gore:
# redovi 929/933 (RATA ZAKS 7,96 €) poklapaju se s "e-Zaba" bankovnim troškom istog
# iznosa i datuma naplate, ali su različiti merchanti. NE DIRATI.


def pick_review() -> Path:
    cands = [p for p in DATA.glob('Financije_review_*.xlsx')
             if '.pre-' not in p.name and not p.name.startswith('~$')]
    if not cands:
        sys.exit(f'✗ Nema Review filea u {DATA}')
    return max(cands, key=lambda p: p.stat().st_mtime)


def clean(v) -> str:
    return str(v).strip() if v is not None else ''


def append_skip(wb, key: str, datum, iznos, smjer: str, opis: str) -> None:
    """Registar trajno odbačenih izvod-transakcija (format iz consolidate_review.py)."""
    if SKIP_SHEET in wb.sheetnames:
        ws = wb[SKIP_SHEET]
    else:
        ws = wb.create_sheet(SKIP_SHEET)
        for c, h in enumerate(('key', 'Datum', 'Iznos', 'Smjer', 'Opis', 'Kad'), 1):
            ws.cell(1, c, h).font = Font(bold=True)
        ws.sheet_state = 'hidden'
    r = ws.max_row + 1
    ws.cell(r, 1, key)
    ws.cell(r, 2, str(datum)[:10])
    ws.cell(r, 3, iznos)
    ws.cell(r, 4, smjer)
    ws.cell(r, 5, opis[:80])
    ws.cell(r, 6, f'{datetime.now():%Y-%m-%d %H:%M}')


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--dry', action='store_true', help='pokaži plan, ne piši ništa')
    args = ap.parse_args()

    review = pick_review()
    print(f'Review: {review.name}')
    wb = openpyxl.load_workbook(review)
    ws = wb['Review']
    H = {str(c.value).strip(): i
         for i, c in enumerate(next(ws.iter_rows(max_row=1)), 1) if c.value}
    for c in ('source_key', 'Isplata', 'Datum naplate', 'Napomena', 'Izvod opis',
              'Izvod file', 'Smjer', 'event_date'):
        if c not in H:
            sys.exit(f'✗ Review nema kolonu "{c}".')

    by_key: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        k = clean(ws.cell(r, H['source_key']).value)
        if k:
            by_key.setdefault(k, r)

    # ── verifikacija svih 8 parova PRIJE ijedne izmjene ──
    plan, problems = [], []
    for izv_k, koka_k, iznos, napomena in PAIRS:
        ri, rk = by_key.get(izv_k), by_key.get(koka_k)
        if not ri or not rk:
            problems.append(f'{izv_k}/{koka_k}: source_key nije nađen '
                            f'(izvodni={ri}, ručni={rk}) — je li već popravljeno?')
            continue
        gi = lambda n: ws.cell(ri, H[n]).value
        gk = lambda n: ws.cell(rk, H[n]).value
        if round(float(gi('Isplata') or 0), 2) != iznos or round(float(gk('Isplata') or 0), 2) != iznos:
            problems.append(f'red {ri}/{rk}: iznos ne odgovara ({gi("Isplata")}/{gk("Isplata")} ≠ {iznos})')
            continue
        if str(gi('Datum naplate'))[:10] != str(gk('Datum naplate'))[:10]:
            problems.append(f'red {ri}/{rk}: "Datum naplate" se razlikuje '
                            f'({gi("Datum naplate")} vs {gk("Datum naplate")})')
            continue
        if napomena.lower() not in clean(gk('Napomena')).lower():
            problems.append(f'red {rk}: Napomena "{clean(gk("Napomena"))}" ne sadrži "{napomena}"')
            continue
        plan.append((ri, rk, iznos, clean(gi('Izvod opis')), clean(gi('Izvod file')),
                     clean(gk('Napomena')), izv_k, gi('event_date'), clean(gi('Smjer'))))

    if problems:
        print('\n✗ Verifikacija nije prošla — NIŠTA nije dirnuto:')
        for p in problems:
            print(f'  {p}')
        sys.exit(1)

    print(f'\nVerificirano {len(plan)}/8 parova (po source_key, iznosu, Datumu naplate i Napomeni):')
    print('-' * 96)
    total = 0.0
    for ri, rk, iznos, opis, _f, nap, _k, _d, _s in plan:
        total += iznos
        print(f'  BRIŠEM  red {ri:>5}  {iznos:>8.2f} €  {opis[:44]}')
        print(f'  OSTAJE  red {rk:>5}  {"":>8}    {nap[:44]}  ← dobiva "Izvod opis"')
    print('-' * 96)
    print(f'  Ukupno uklonjeno dvostruko: {total:.2f} €')

    if args.dry:
        print('\n… --dry: ništa nije zapisano.')
        return

    # ── upis: prvo prepis opisa, pa brisanje odozdo prema gore ──
    for ri, rk, _i, opis, ffile, _n, key, dat, smjer in plan:
        if not clean(ws.cell(rk, H['Izvod opis']).value):          # P3 — puno se ne gazi
            ws.cell(rk, H['Izvod opis']).value = opis
        if not clean(ws.cell(rk, H['Izvod file']).value):
            ws.cell(rk, H['Izvod file']).value = ffile
        append_skip(wb, key, dat, _i, smjer or 'Isplata', opis)

    for ri in sorted((p[0] for p in plan), reverse=True):
        ws.delete_rows(ri, 1)

    if ws.auto_filter.ref:                    # raspon mora pratiti novi broj redaka
        first = ws.auto_filter.ref.split(':')[0]
        ws.auto_filter.ref = f'{first}:{get_column_letter(ws.max_column)}{ws.max_row}'

    backup = review.with_name(f'{review.stem}.pre-duprata-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        backup.unlink(missing_ok=True)
        sys.exit('✗ Zatvori Review u Excelu i ponovi. (Ništa nije zapisano.)')
    print(f'\n✔ Obrisano {len(plan)} redaka · {len(plan)} ključeva u "{SKIP_SHEET}" · '
          f'Review sad {ws.max_row - 1} redaka')
    print(f'  Backup: {backup.name}')


if __name__ == '__main__':
    main()
