# -*- coding: utf-8 -*-
"""
merge_pbzvisa.py  (S107i, 2026-07-20)
=====================================
Mergea PBZ Visa Gold transakcije (iz `Nematchano` sheeta Izvodi_transakcije.xlsx)
kao NOVE retke u Financije REVIEW Excel. Rješava "1/1539 match" problem: enrich je
PBZ Visa matchao na Kokin ZABA/Mastercard, a Sašini Visa retci su Racun=Sašin RF /
Izvor=Visa — pa je praktički sve završilo u Nematchano.

ODLUKE (Saša + Koka, S107i — v. ENRICH_PLAN §2d/§2f):
  • Dedup TAG-AGNOSTIČKI: postojećih 220 Sašinih Visa redaka matcha se protiv SVIH
    PBZ tx (bez obzira na Kartica tag) — jer je Saša u Excel bilježio kupovine s
    OBJE kartice (121 njegovih redaka nosi Kokinu karticu). Matchane PBZ tx se
    PRESKAČU (već u Review); ostale postaju novi retci.
  • BEZ tvrdog person-splita (Odluka 2a): svi novi retci Racun='Sašin tekući RF',
    Izvor='Visa'. Karta se čuva samo kao audit trag u 'Izvor reda' (PBZ Visa:SAŠA/
    Koka/lump), NE određuje Racun ni osobu. Osoba se pojavljuje kroz Tip/Podtip
    klasifikaciju (pravila) gdje ima signala.
  • Lump 'PRIMLJENA UPLATA' (Uplata, prazna kartica) → Tip='Transfer' /
    Podtip='izmedju racuna', Pouzdanost='VISOKA' (mjesečna naplata Visa računa sa
    Sašinog RF — bez duplog brojanja troška).
  • Kupovine → Tip='N/A' (čekaju apply_rules.py). Merchant tekst ide u 'Izvod opis'
    (rules ga pretražuju), Napomena OSTAJE prazna (da rule može upisati čistu labelu).
  • RATA retci → Rate?='DA', Broj rata=N (parsirano iz "RATA X/N"); svaki kao
    zaseban red (mjesečni teret), datum = datum IZVORNE kupovine (kako parser daje).
  • Datum naplate = prazno (puni ga import-generator / PBZ dospijeće — kao i za
    postojećih 220 Sašinih Visa redaka).

POREDAK (Opcija B, S107i): Review je sortiran po event_date pa se nakon dodavanja
CIJELI Review presortira po event_date (stabilno — postojeći retci istog datuma
zadrže redoslijed). Novi retci naslijede stil (formati datum/iznos, obrubi) s
postojećeg Sašinog Visa reda; DV dropdowni (Tip/Podtip) i autofilter se prošire na
sve retke — inače novi retci nemaju dropdown ni sort-safe filter.

Idempotentno: svaki novi red dobije stabilan source_key (md5 identiteta); pravi run
preskače retke čiji source_key već postoji u Review → ponovni run ne duplicira.

Pokretanje (Review zatvoren u Excelu!):
  Financije\\run.bat merge_pbzvisa.py --dry   → NE dira Review; piše
        `pbzvisa_PREVIEW.xlsx` (kopija Review + dodani retci) za pregled + report
  Financije\\run.bat merge_pbzvisa.py         → pravi run: append u Review + backup
  ... merge_pbzvisa.py <review.xlsx> [--dry]
"""

import hashlib
import re
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")
TX_XLSX  = DATA_DIR / 'izvodi' / 'Izvodi_transakcije.xlsx'
PREVIEW  = DATA_DIR / 'pbzvisa_PREVIEW.xlsx'   # pick_file drugih alata ovo IGNORIRA

RACUN_SASA = 'Sašin tekući RF'
IZVOR      = 'Visa'
DELTAS     = (0, 1, -1, 2, -2, 3, -3)          # datum tolerancija za dedup (plateau na ±2)
RE_RATA    = re.compile(r'\bRATA\s+(\d+)\s*/\s*(\d+)\b', re.IGNORECASE)


def is_sasa(kartica: str, opis: str) -> bool:
    return 'SAŠA' in str(kartica or '').upper() or 'SAŠA' in str(opis or '').upper()


def pick_file(args: list[str]) -> Path:
    explicit = [a for a in args if not a.startswith('--')]
    if explicit:
        p = Path(explicit[0])
        if not p.exists():
            sys.exit(f'✗ File ne postoji: {p}')
        return p
    cands = sorted(DATA_DIR.glob('Financije_review_*.xlsx'),
                   key=lambda p: p.stat().st_mtime, reverse=True)
    cands = [c for c in cands if '.pre-' not in c.name]
    if not cands:
        sys.exit(f'✗ Nema Financije_review_*.xlsx u {DATA_DIR}')
    return cands[0]


def hdr_index(ws) -> dict[str, int]:
    """Header -> 1-based column index (Review konvencija: red 1 = headeri)."""
    return {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}


def to_date(v):
    try:
        return v.date()
    except Exception:
        return v if hasattr(v, 'year') else None


def load_pbz_tx() -> list[dict]:
    """PBZVISA transakcije iz Nematchano sheeta."""
    wb = openpyxl.load_workbook(TX_XLSX, read_only=True)
    if 'Nematchano' not in wb.sheetnames:
        sys.exit('✗ Nema "Nematchano" sheeta u Izvodi_transakcije.xlsx (pokreni enrich_from_izvoda.py).')
    ws = wb['Nematchano']
    h = {str(c.value): (i + 1) for i, c in enumerate(ws[1]) if c.value is not None}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        src = str(row[h['Src'] - 1] or '')
        if not src.upper().startswith('PBZVISA'):
            continue
        d = to_date(row[h['Datum'] - 1])
        if d is None:
            continue
        opis = str(row[h['Opis'] - 1] or '')
        kartica = str(row[h['Kartica'] - 1] or '')
        smjer = str(row[h['Smjer'] - 1] or '')
        out.append({
            'date': d, 'opis': opis, 'kartica': kartica, 'smjer': smjer,
            'iznos': round(float(row[h['Iznos'] - 1]), 2), 'src': src,
        })
    wb.close()
    return out


def build_existing_index(ws, col) -> tuple[dict, int, int | None]:
    """(date, amount) -> [row] index postojećih Sašinih Visa redaka + broj + template red
    (prvi Visa red — s njega novi retci naslijede stil/formate)."""
    idx: dict[tuple, list[int]] = defaultdict(list)
    n = 0
    template = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, col['Racun']).value or '') != RACUN_SASA:
            continue
        if str(ws.cell(r, col['Izvor']).value or '') != IZVOR:
            continue
        if template is None:
            template = r
        d = to_date(ws.cell(r, col['event_date']).value)
        smjer = str(ws.cell(r, col['Smjer']).value or '')
        amt = ws.cell(r, col['Isplata' if smjer == 'Isplata' else 'Uplata']).value
        n += 1
        if d is None or amt in (None, ''):
            continue
        idx[(d, round(float(amt), 2))].append(r)
    return idx, n, template


def source_key(tx: dict, taken: set[str]) -> str:
    canon = f"pbzvisa|{tx['date']}|{tx['smjer']}|{tx['iznos']}|{tx['opis']}|{tx['src']}"
    base = hashlib.md5(canon.encode('utf-8')).hexdigest()
    key = base[:12]
    salt = 0
    while key in taken:                     # kolizija (identičan tx dvaput) → produži
        salt += 1
        key = hashlib.md5(f'{canon}#{salt}'.encode('utf-8')).hexdigest()[:12]
    taken.add(key)
    return key


def build_row(tx: dict, col: dict, ncols: int, skey: str) -> list:
    """Puni red (lista dužine ncols, 1-based mapiranje kroz col)."""
    row = [None] * ncols
    def put(name, val):
        row[col[name] - 1] = val

    is_lump = (not tx['kartica'].strip()) and tx['smjer'] == 'Uplata'
    who = 'lump' if is_lump else ('SAŠA' if is_sasa(tx['kartica'], tx['opis']) else 'Koka')

    put('Racun', RACUN_SASA)
    put('event_date', datetime(tx['date'].year, tx['date'].month, tx['date'].day))
    put('Datum naplate', None)
    put('Smjer', tx['smjer'])
    put('Izvor', IZVOR)
    put('Uplata', tx['iznos'] if tx['smjer'] == 'Uplata' else None)
    put('Isplata', tx['iznos'] if tx['smjer'] == 'Isplata' else None)
    put('Stanje', None)
    put('Napomena', None)                   # prazno — Izvod opis nosi merchant; rule puni čistu labelu
    if is_lump:
        put('Tip', 'Transfer'); put('Podtip', 'izmedju racuna'); put('Pouzdanost', 'VISOKA')
        put('Tip_O', 'Transfer'); put('Podtip_O', 'izmedju racuna')
    else:
        put('Tip', 'N/A'); put('Podtip', None); put('Pouzdanost', 'NEMA')
        put('Tip_O', 'N/A'); put('Podtip_O', None)
    put('Alternativa / nap.', None)
    m = RE_RATA.search(tx['opis'])
    put('Rate?', 'DA' if m else None)
    put('Broj rata', int(m.group(2)) if m else None)
    put('Status', 'Izvrsen')
    put('Izvor reda', f'PBZ Visa:{who}')
    put('Labela iz', None)
    put('Problem', None)
    put('source_key', skey)
    put('Izvod opis', tx['opis'][:250])
    put('Izvod file', tx['src'])
    put('Pravilo run', None)
    return row


def dedup(tx_list: list[dict], idx: dict) -> tuple[list[dict], int]:
    """Vrati (nove_tx, broj_matchanih). Greedy 1-na-1: postojeći red 'potroši' jednu tx."""
    used_existing: set[int] = set()
    matched_tx: set[int] = set()
    for i, tx in enumerate(tx_list):
        for dd in DELTAS:
            cand = idx.get((tx['date'] + timedelta(days=dd), tx['iznos']), [])
            row = next((r for r in cand if r not in used_existing), None)
            if row is not None:
                used_existing.add(row)
                matched_tx.add(i)
                break
    new_tx = [tx for i, tx in enumerate(tx_list) if i not in matched_tx]
    return new_tx, len(matched_tx)


def sort_review(ws, col, ncols: int) -> None:
    """Presortiraj cijeli Review data-region po event_date (stabilno), očuvaj stil
    svake ćelije, pa proširi DV (Tip/Podtip) + autofilter na sve retke."""
    n = ws.max_row
    records = []
    for r in range(2, n + 1):
        d = to_date(ws.cell(r, col['event_date']).value)
        cells = [(ws.cell(r, c).value, ws.cell(r, c)._style) for c in range(1, ncols + 1)]
        records.append(((d is None, d or datetime.min.date()), cells))
    records.sort(key=lambda rec: rec[0])          # stabilan sort (ties = originalni redoslijed)
    for i, (_, cells) in enumerate(records):
        r = 2 + i
        for c, (val, style) in enumerate(cells, 1):
            cell = ws.cell(r, c)
            cell.value = val
            cell._style = style
    # DV dropdowni: proširi Tip (J) i Podtip (K) sqref na sve retke
    for dv in ws.data_validations.dataValidation:
        sq = str(dv.sqref)
        if sq.startswith('J'):
            dv.sqref = f'J2:J{n}'
        elif sq.startswith('K'):
            dv.sqref = f'K2:K{n}'
    # autofilter na sve kolone + retke
    if ws.auto_filter.ref:
        first = ws.auto_filter.ref.split(':')[0]
        ws.auto_filter.ref = f'{first}:{get_column_letter(ncols)}{n}'


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    review = pick_file(args)
    print(f'Review: {review.name}{"  [DRY — Review se NE dira]" if dry else ""}')

    tx_list = load_pbz_tx()
    print(f'PBZ Visa tx (Nematchano): {len(tx_list)}')

    wb = openpyxl.load_workbook(review)
    ws = wb['Review']
    needed = ('Racun', 'event_date', 'Datum naplate', 'Smjer', 'Izvor', 'Uplata',
              'Isplata', 'Stanje', 'Napomena', 'Tip', 'Podtip', 'Pouzdanost',
              'Alternativa / nap.', 'Rate?', 'Broj rata', 'Status', 'Izvor reda',
              'Labela iz', 'Problem', 'source_key', 'Izvod opis', 'Izvod file',
              'Tip_O', 'Podtip_O', 'Pravilo run')
    col = hdr_index(ws)
    missing = [c for c in needed if c not in col]
    if missing:
        sys.exit(f'✗ Review nema kolone: {missing}')
    ncols = ws.max_column

    # postojeći source_keyevi (idempotencija) + dedup index
    existing_keys = {str(ws.cell(r, col['source_key']).value or '')
                     for r in range(2, ws.max_row + 1)
                     if ws.cell(r, col['source_key']).value}
    idx, n_existing_visa, template_row = build_existing_index(ws, col)
    print(f'Postojeći Sašini Visa redovi: {n_existing_visa}')
    template_styles = ([ws.cell(template_row, c)._style for c in range(1, ncols + 1)]
                       if template_row else None)

    new_tx, matched = dedup(tx_list, idx)
    print(f'Dedup: {matched} PBZ tx matcha postojeće (preskoči) | {len(new_tx)} → novi retci')

    # izgradi retke (idempotentno preskoči već-prisutne source_keyeve)
    taken = set(existing_keys)
    rows_to_add = []
    skipped_dup = 0
    stats = Counter()
    for tx in new_tx:
        # probni source_key za provjeru postoji-li već (bez trošenja salt-a na taken)
        canon = f"pbzvisa|{tx['date']}|{tx['smjer']}|{tx['iznos']}|{tx['opis']}|{tx['src']}"
        probe = hashlib.md5(canon.encode('utf-8')).hexdigest()[:12]
        if probe in existing_keys:
            skipped_dup += 1
            continue
        skey = source_key(tx, taken)
        rows_to_add.append(build_row(tx, col, ncols, skey))
        is_lump = (not tx['kartica'].strip()) and tx['smjer'] == 'Uplata'
        who = 'lump (Transfer)' if is_lump else ('SAŠA→N/A' if is_sasa(tx['kartica'], tx['opis']) else 'Koka→N/A')
        stats[who] += 1
        if RE_RATA.search(tx['opis']):
            stats['(od toga RATA)'] += 1

    print(f'\n{"[DRY] " if dry else ""}Za dodati: {len(rows_to_add)} redaka'
          + (f'  (preskočeno {skipped_dup} — source_key već u Review)' if skipped_dup else ''))
    for k, v in stats.most_common():
        print(f'  {k:18}: {v}')
    print(f'Review sada {ws.max_row - 1} → {ws.max_row - 1 + len(rows_to_add)} redaka')

    if not rows_to_add:
        print('Nema ničega za dodati.'); return

    # sample retci
    print('\n--- 8 sample novih redaka (event_date | Smjer | iznos | Tip/Podtip | Izvor reda | Izvod opis[:38]) ---')
    for row in rows_to_add[:8]:
        ed = row[col['event_date'] - 1]
        amt = row[col['Isplata'] - 1] if row[col['Smjer'] - 1] == 'Isplata' else row[col['Uplata'] - 1]
        print(f"  {str(ed)[:10]} | {row[col['Smjer']-1]:7} | {str(amt):>8} | "
              f"{str(row[col['Tip']-1]):8}/{str(row[col['Podtip']-1] or '—'):14} | "
              f"{str(row[col['Izvor reda']-1]):14} | {str(row[col['Izvod opis']-1] or '')[:38]}")

    # zapiši retke (append) — novi retci naslijede stil (formati/obrubi) s Visa template reda
    start = ws.max_row + 1
    for i, row in enumerate(rows_to_add):
        for c in range(1, ncols + 1):
            cell = ws.cell(start + i, c)
            if row[c - 1] is not None:
                cell.value = row[c - 1]
            if template_styles:
                cell._style = template_styles[c - 1]

    # Opcija B: presortiraj cijeli Review po event_date + proširi DV/autofilter
    sort_review(ws, col, ncols)

    if dry:
        wb.save(PREVIEW)
        print(f'\n✔ [DRY] PREVIEW zapisan: {PREVIEW.name} ({len(rows_to_add)} novih redaka, '
              f'usortiranih po event_date među postojeće).')
        print('  Otvori ga u Excelu, filtriraj Izvor reda = "PBZ Visa:*" da vidiš samo nove.')
        print('  Review NIJE diran. Kad potvrdiš → pokreni bez --dry.')
        return

    backup = review.with_name(f'{review.stem}.pre-pbzvisa-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori file u Excelu i ponovi. (Backup: {backup.name})')
    print(f'\n✔ Snimljeno: {len(rows_to_add)} redaka dodano. Backup: {backup.name}')
    print('  Sljedeći korak: apply_rules.py (pravila će klasificirati nove N/A retke).')


if __name__ == '__main__':
    main()
