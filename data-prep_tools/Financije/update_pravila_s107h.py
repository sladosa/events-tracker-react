# -*- coding: utf-8 -*-
"""
update_pravila_s107h.py  (S107h, 2026-07-17)
=============================================
One-off: primjenjuje dogovorene izmjene Pravila sheeta iz S107h sesije
(chat pregled reda po redu sa Sašom + Kokom):

  - AMAZON red obrisan (samo 2 retka u cijelom fileu, iznosi ne odgovaraju
    Prime pretplati — ostaje ručno za review, ne pravilo)
  - APPLE.COM zamijenjen s 2 Iznos-range reda → Informatika/Cloud backup
    (iznosi 2.99 i 7.99/9.99 su mjesečna iCloud pretplata, ne "Zabava";
    potvrđeno postojećim ručno klasificiranim redom 2291 iste cijene)
  - AUDIBLE razdvojen na 2 reda po Iznos min/max (Koka: Audible_Sasa je
    skuplja pretplata) — ispod/iznad praga 10 (jasan razmak 8.99→13.21)
  - Dodane 'Iznos min'/'Iznos max' header kolone (F/G) + pomaknut help note u H2

Pokretanje (file zatvoren u Excelu!): Financije\\run.bat update_pravila_s107h.py
Idempotentno: siguran za ponovno pokretanje (regenerira cijeli Pravila body
iz FINAL_RULES liste ispod).
"""

import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")

HDR_FILL   = PatternFill('solid', fgColor='4472C4')
WHITE_BOLD = Font(color='FFFFFF', bold=True)
THIN       = Side(style='thin')
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [('Ključne riječi', 32), ('Tip', 16), ('Podtip', 24), ('Napomena', 18),
           ('Komentar', 40), ('Iznos min', 10), ('Iznos max', 10)]

HELP_TEXT = (
    'PRAVILA KLASIFIKACIJE — jedan red = jedno pravilo; odozgo prema dolje, PRVI match pobjeđuje.\n'
    'Ključne riječi: "konzum" = tekst sadrži konzum; "telekom & racun" = sadrži OBJE riječi.\n'
    'Zvjezdica * NIJE wildcard — traži se doslovno (radi samo ako tekst stvarno ima "*", npr.\n'
    '"GOOGLE *YouTube"). Za "sadrži riječ bilo gdje" piši samo riječ, bez zvjezdica.\n'
    'Case i dijakritike se ignoriraju (Č=č=c). OR se piše kao dva odvojena reda pravila.\n'
    'Pravilo se primjenjuje SAMO na redove gdje je Tip prazan ili N/A — ručni rad se ne dira.\n'
    'Napomena kolona (opcionalno): čista GOTOVA labela, upisuje se SAMO u redove s praznom\n'
    'Napomenom — na kraju hrani comment polje uvezenog eventa u appu, NE piši ovdje "odredi\n'
    'koje?" ili slične podsjetnike sebi.\n'
    'Komentar kolona (opcionalno, S107h): NIJE label, ne ide u comment — dopisuje se uz\n'
    '"pravilo #N: <ključne riječi>" u Alternativa / nap. koloni Reviewa, za tvoje filtriranje/\n'
    'suženje kasnije (npr. "provjeri policu", "TODO razdvoji po X"). Sigurno mjesto za bilješke.\n'
    'Iznos min / Iznos max (opcionalno, S107h): dodatni uvjet uz ključne riječi — red mora\n'
    'imati Isplata/Uplata unutar raspona (jedno od dva polje smije ostati prazno = bez granice).\n'
    'Za razdvajanje istog merchanta po cijeni (npr. dvije razine pretplate = dvije osobe).\n'
    'Tekst za pretragu = Napomena + "Izvod opis" kolone (nakon enrich_from_izvoda.py).\n'
    'Tip i Podtip moraju postojati u Taksonomija sheetu (inače se pravilo preskače uz upozorenje).\n'
    'Prije primjene: skripta jednom snima Tip_O/Podtip_O original kolone + resetira na N/A\n'
    'redove čiji Tip/Podtip par više ne postoji u Taksonomiji (oznaka "TAKS:" u Alternativa).\n'
    'Nakon izmjena pokreni: Financije\\run.bat apply_rules.py  (--dry za probu bez snimanja;\n'
    '--all = i report konflikata pravila s već klasificiranim redovima, bez pisanja)'
)

# (ključne riječi, Tip, Podtip, Napomena, Komentar, Iznos min, Iznos max)
FINAL_RULES = [
    ('temu',                      'Razno',        'Temu',                              'Kokin Temu', None, None, None),
    ('bolt.eu',                   'Razno',        'Taksi',                             'Prevoz',     None, None, None),
    ('konzum',                    'Namirnice',    'Hrana i ostalo',                     'Konzum',     None, None, None),
    ('bauhaus',                   'Domaćinstvo',  'Popravci, održavanje, osiguranje',   'Bauhaus',    None, None, None),
    ('prime video',               'Zabava',       'Prime',                             'TV zabava',  None, None, None),
    ('skyshowtime',               'Zabava',       'Sky',                               'TV zabava',  None, None, None),
    ('google*youtube',            'Zabava',       'Youtube',                           'TV zabava',  None, None, None),
    ('apple.com',                 'Informatika',  'Cloud backup',                      'iCloud',     None, 2.9,  3.1),
    ('apple.com',                 'Informatika',  'Cloud backup',                      'iCloud',     None, 7.5,  10.5),
    ('AUDIBLE',                   'Zabava',       'Audible_Koka',                      None,         None, None, 10),
    ('AUDIBLE',                   'Zabava',       'Audible_Sasa',                      None,         None, 10,   None),
    ('spotify',                   'Zabava',       'Spotify',                           None,         None, None, None),
    ('KINDLE ',                   'Zabava',       'Kindle_Koka',                       None,         None, None, None),
    ('UPLATA ANJA CRNKOVIĆ',      'Povrat',       'Anja',                              None,         None, None, None),
    ('KUPOVINACLAUDE.AISUBSANFR', 'Projekti',     'Sasa_Informatika',                  'AI Claude',  None, None, None),
    ('KUPOVINAANTHROPIC SANFR',   'Projekti',     'Sasa_Informatika',                  'AI Claude',  None, None, None),
    ('allianz & lacetti',         'auto Lacetti', 'registracija',                      None,         None, None, None),
    ('allianz',                   'auto C5',      'registracija',                      None,         None, None, None),
    ('generali',                  'Domaćinstvo',  'Popravci, održavanje, osiguranje',  None,         None, None, None),
    ('triglav',                   'Osiguranje',   'Osiguranje',                        None,         None, None, None),
    ('OTP Leasing',               'auto C5',      'leasing',                           None,         None, None, None),
    ('parking',                   'auto C5',      'parking',                           None,         None, None, None),
    ('Podizanje gotovog novca',   'Transfer',     'cash - bankomat',                   None,         None, None, None),
    ('ISPLATAGOTO',               'Transfer',     'cash - bankomat',                   None,         None, None, None),
    ('HP INC.',                   'Informatika',  'HP',                                None,         None, None, None),
]


def pick_file() -> Path:
    candidates = sorted(DATA_DIR.glob('Financije_review_*.xlsx'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    candidates = [c for c in candidates if '.pre-' not in c.name]
    if not candidates:
        sys.exit(f'✗ Nema Financije_review_*.xlsx u {DATA_DIR}')
    return candidates[0]


def main() -> None:
    path = pick_file()
    print(f'File: {path.name}')
    wb = openpyxl.load_workbook(path)
    if 'Pravila' not in wb.sheetnames:
        sys.exit('✗ Nema Pravila sheeta.')
    ws = wb['Pravila']

    # obriši sav postojeći sadržaj (header + body + stara nota) — regeneriramo iz FINAL_RULES
    ws.delete_rows(1, ws.max_row)

    for c, (h, w) in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
        ws.column_dimensions[chr(64 + c)].width = w

    for r, rule in enumerate(FINAL_RULES, 2):
        for c, v in enumerate(rule, 1):
            ws.cell(r, c, v).border = BORDER

    note = ws.cell(2, 8, HELP_TEXT)
    note.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['H'].width = 95
    ws.row_dimensions[2].height = 210
    ws.freeze_panes = 'A2'

    backup = path.with_name(f'{path.stem}.pre-pravilaS107h-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori {path.name} u Excelu i ponovi. (Backup: {backup.name})')
    print(f'✔ Pravila sheet ažuriran: {len(FINAL_RULES)} pravila. Backup: {backup.name}')
    print('  Sljedeći korak: Financije\\run.bat apply_rules.py --dry')


if __name__ == '__main__':
    main()
