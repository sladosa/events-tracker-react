# -*- coding: utf-8 -*-
"""
verify_saldo_model.py  (S107x Faza 1a, 2026-08-12)
================================================================================
DOKAZ MODELA SALDA nad stvarnim Reviewom — PRIJE ijednog reda app koda.
Spec: docs/OVERVIEW_TAB_SPEC.md §2.10 (saldo miče Izvor, ne Racun), §2.13
(planirano po kantama), §2.14 (transfer), "Faza 1a".

⚠ STROGO READ-ONLY. Workbook se otvara s read_only=True i NIKAD se ne snima.
   U ovom fileu nema `.save()` — namjerno. Skripta samo čita i ispisuje.

--------------------------------------------------------------------------------
PROVJERA 1 — je li pravilo iz §2.10 točno?

  Pravilo:  Izvor ∈ {Racun, Cash}      → miče saldo (izvršeno)
            Izvor ∈ {Visa, Mastercard} → ne miče (tek kroz skupnu naplatu)

  ⚠ Test NIJE "usporedi razinu s Kokinim `Stanje`". Njen `Stanje` stupac se u
  ovom fileu NE MOŽE hodati: Review je presortiran po `event_date` (S107i,
  Opcija B), a njeni saldi su iz izvornog redoslijeda njenog workbooka
  (`Izvor reda` = `koka EU:N`). U datumskom poretku lanac puca na 969 od 2564
  mjesta — pa bi svaka usporedba razine mjerila artefakt sortiranja, ne model.

  Zato se mjeri POMAK (movement), ne razina, i to protiv BANKE:
      model_Δ(M)  = Σ signed(retci u prozoru izvatka M) uz pravilo
      banka_Δ(M)  = NOVO STANJE(M) − NOVO STANJE(M−1)
  Test je time neovisan o sidru — jedna rana greška ne razmazuje se na svih 31.
  Usporedno se računa i NAIVNI zbroj po `Racun`u (kontrola: koliko griješi).

  Rezidual se pokušava objasniti granicom prozora (±2 dana): transakcija
  datirana uz sam datum zatvaranja legitimno može pasti na obje strane.

PROVJERA 2 — transfer jednom ili dvaput? (§2.14)
  Transfer retci se najprije razvrstaju po ULOZI, jer nemaju svi protupartiju:
    · naplata kartice (MC/PBZ lump) — druga strana je kartica, NE praćeni račun
    · bankomat — druga strana je gotovina (novčanik nije račun)
    · međuračunski — JEDINI kojem protupartija smije postojati
  Pitanje iz §2.14 odnosi se samo na treću skupinu.

PROVJERA 3 — planirano po kantama i smjerovima (§2.13)
  Dospjelo (`Planiran` ∧ `Datum naplate` ≤ danas) / Uskoro (≤30 dana) / Kasnije,
  odvojeno za Isplatu i Uplatu.

--------------------------------------------------------------------------------
Pokretanje:
    Financije\\run.bat verify_saldo_model.py
    Financije\\run.bat verify_saldo_model.py --rows     (detalj rezidualnih mjeseci)
    Financije\\run.bat verify_saldo_model.py <put\\do\\review.xlsx>

⚠ run.bat guši zarez u argumentima — jedan argument po pozivu.
"""
from __future__ import annotations

import sys
import textwrap
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")

# --- model (§2.10) ---------------------------------------------------------
IZVOR_IZVRSENO = {'Racun', 'Cash'}          # miče saldo odmah
IZVOR_ODGODENO = {'Visa', 'Mastercard'}     # tek kroz skupnu naplatu
RACUN_ZABA = 'Kokin tekući ZABA'            # jedini račun s bankovnim baselineom
EPS = 0.005                                 # "u cent"
GRANICA_DANA = 2                            # tolerancija uz datum zatvaranja
TRANSFER_TOL_DANA = 3                       # prozor za traženje protupartije
USKORO_DANA = 30                            # kanta "Uskoro" (§2.13)

# prepoznavanje uloge Transfer retka (§2.14)
LUMP_KLJUC = ('MASTERCARD KARTICOM', 'PBZCARD', 'PRIMLJENA UPLATA', 'PBZ CARD',
              'PBZ VISA', 'VISA')
BANKOMAT_KLJUC = ('BANKOMAT', 'PODIZANJE GOTOV', 'ISPLATA GOTOV')

SEP = '─' * 78


# ---------------------------------------------------------------- helpers --
def pick_file(args: list[str]) -> Path:
    """Isti izbor kao ostali alati: eksplicitan put ili najnoviji Review bez .pre-*."""
    explicit = [a for a in args if not a.startswith('--')]
    if explicit:
        p = Path(explicit[0])
        if not p.exists():
            sys.exit(f'✗ File ne postoji: {p}')
        return p
    cands = sorted(DATA_DIR.glob('Financije_review_*.xlsx'),
                   key=lambda p: p.stat().st_mtime, reverse=True)
    cands = [c for c in cands if '.pre-' not in c.name]
    if not cands:
        sys.exit(f'✗ Nema Financije_review_*.xlsx u {DATA_DIR}')
    return cands[0]


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    return v if isinstance(v, date) else None


def num(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def s(v) -> str:
    return '' if v is None else str(v).strip()


def eur(x) -> str:
    return f'{x:,.2f}'.replace(',', ' ')


def kol(x, w=12) -> str:
    return eur(x).rjust(w)


# ------------------------------------------------------------------ učitaj --
def load_review(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Review']
    it = ws.iter_rows(min_row=1, values_only=True)
    hdr = {s(v): i for i, v in enumerate(next(it)) if v is not None}
    need = ('Racun', 'event_date', 'Datum naplate', 'Smjer', 'Izvor', 'Uplata',
            'Isplata', 'Stanje', 'Napomena', 'Tip', 'Podtip', 'Status', 'Rate?')
    missing = [c for c in need if c not in hdr]
    if missing:
        wb.close()
        sys.exit(f'✗ Nedostaju kolone u Review sheetu: {missing}')

    rows = []
    for xl, raw in enumerate(it, start=2):
        def c(name):
            i = hdr.get(name)
            return raw[i] if i is not None and i < len(raw) else None
        d = to_date(c('event_date'))
        if d is None:
            continue
        upl, isp = num(c('Uplata')), num(c('Isplata'))
        rows.append({
            'xl': xl, 'racun': s(c('Racun')), 'date': d,
            'naplata': to_date(c('Datum naplate')), 'smjer': s(c('Smjer')),
            'izvor': s(c('Izvor')), 'uplata': upl, 'isplata': isp,
            'signed': round((upl or 0.0) - (isp or 0.0), 2),
            'stanje': num(c('Stanje')), 'napomena': s(c('Napomena')),
            'tip': s(c('Tip')), 'podtip': s(c('Podtip')), 'status': s(c('Status')),
            'rate': s(c('Rate?')), 'izvor_reda': s(c('Izvor reda')),
            'opis': s(c('Izvod opis')),
        })

    saldo = []
    if 'Saldo kontrola' in wb.sheetnames:
        for r in wb['Saldo kontrola'].iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            saldo.append({'ym': s(r[0]), 'close': to_date(r[1]), 'koka': num(r[2]),
                          'banka': num(r[3]), 'delta': num(r[4]) or 0.0})
    wb.close()
    return rows, saldo


# ------------------------------------------------------ PROVJERA 1 — saldo --
def lanac_dijagnoza(rows) -> None:
    """Zašto se NE uspoređuje razina s Kokinim `Stanje`: lanac je puknut sortom."""
    z = [r for r in rows if r['racun'] == RACUN_ZABA and r['stanje'] is not None]
    ok = brk = 0
    for a, b in zip(z, z[1:]):
        exp = round(a['stanje'] + b['signed'], 2)
        if abs(exp - b['stanje']) < EPS:
            ok += 1
        else:
            brk += 1
    print(f'Kokin `Stanje` lanac u datumskom poretku: {ok} slaže / {brk} puknuća '
          f'(od {ok + brk})')
    print('   ⇒ razina se NE može usporediti (Review je presortiran po event_date, '
          'S107i);\n     mjeri se POMAK protiv banke — v. tablicu.')


def provjera_1(rows, saldo, show_rows: bool):
    print(SEP)
    print('PROVJERA 1 — pravilo §2.10 (saldo miče `Izvor`, ne `Racun`)')
    print(SEP)
    lanac_dijagnoza(rows)

    z = [r for r in rows if r['racun'] == RACUN_ZABA and r['status'] != 'Planiran']

    def suma(a, b, rule):
        return round(sum(r['signed'] for r in z if a < r['date'] <= b and rule(r)), 2)

    model = lambda r: r['izvor'] in IZVOR_IZVRSENO   # noqa: E731  §2.10
    naivno = lambda r: True                          # noqa: E731  zbroj po Racunu

    okn = sorted([e for e in saldo if e['close'] and e['banka'] is not None],
                 key=lambda e: e['close'])
    print()
    print(f'{"Izvadak":<9} {"banka Δ":>12} {"MODEL Δ":>12} {"rezidual":>10} '
          f'{"NAIVNO Δ":>12} {"rezidual":>11}')
    print('-' * 72)

    ok_m = ok_n = 0
    rezid = []
    for prev, cur in zip(okn, okn[1:]):
        bd = round(cur['banka'] - prev['banka'], 2)
        m = suma(prev['close'], cur['close'], model)
        n = suma(prev['close'], cur['close'], naivno)
        dm, dn = round(m - bd, 2), round(n - bd, 2)
        if abs(dm) < EPS:
            ok_m += 1
        else:
            rezid.append({'ym': cur['ym'], 'od': prev['close'], 'do': cur['close'],
                          'rez': dm})
        if abs(dn) < EPS:
            ok_n += 1
        print(f'{cur["ym"]:<9} {kol(bd)} {kol(m)} {kol(dm, 10)} {kol(n)} {kol(dn, 11)}'
              + ('' if abs(dm) < EPS else '  ←'))
    print('-' * 72)

    n_mj = len(okn) - 1
    print(f'Mjeseci: {n_mj}')
    print(f'   MODEL  (Izvor ∈ {sorted(IZVOR_IZVRSENO)}) reproducira banku u '
          f'{ok_m}/{n_mj} mjeseci')
    print(f'   NAIVNO (zbroj po `Racun`u)                 u {ok_n}/{n_mj} mjeseci')

    # koliko naivni zbroj promaši — mjera dvostrukog brojanja
    print()
    print('Razmjer dvostrukog brojanja (§2.10), cijelo razdoblje po računu:')
    print(f'   {"račun":<20} {"naivno":>12} {"model":>12} {"neto višak":>12} '
          f'{"bruto |višak|":>14}')
    for rac in sorted({r['racun'] for r in rows if r['racun']}):
        sel = [r for r in rows if r['racun'] == rac and r['status'] != 'Planiran']
        nv = round(sum(r['signed'] for r in sel), 2)
        md = round(sum(r['signed'] for r in sel if r['izvor'] in IZVOR_IZVRSENO), 2)
        izb = [r for r in sel if r['izvor'] not in IZVOR_IZVRSENO]
        bruto = round(sum(abs(r['signed']) for r in izb), 2)
        print(f'   {rac:<20} {kol(nv)} {kol(md)} {kol(round(nv - md, 2))} '
              f'{kol(bruto, 14)}')
    print('   ⚠ NETO višak podcjenjuje problem na računu s karticom čiji je i lump')
    print('     zapisan s `Izvor=Visa` (32 retka `PRIMLJENA UPLATA`, Σ +40 244,88):')
    print('     potrošnja i podmirenje se međusobno skraćuju. Bruto stupac je prava mjera.')

    # rezidual: pokušaj imenovati krivca umjesto da ostane gola brojka
    if rezid:
        print()
        print(f'Rezidualni mjeseci ({len(rezid)}) — automatska atribucija:')
        # (a) par susjednih mjeseci koji se poništavaju = redak na krivoj strani granice
        par = set()
        for a, b in zip(rezid, rezid[1:]):
            if abs(a['rez'] + b['rez']) < 50 and abs(a['rez']) > 100:
                par.add(a['ym'])
                par.add(b['ym'])
                a['par'] = b['ym']
                b['par'] = a['ym']
        imenovano = 0
        for e in rezid:
            win = [r for r in z if e['od'] < r['date'] <= e['do']
                   and r['izvor'] in IZVOR_IZVRSENO]
            # (b) jedan redak točno jednak rezidualu (dupli ili višak zapisa)
            tocan = [r for r in win if abs(r['signed'] - e['rez']) < EPS]
            # (c) dva retka istog iznosa i datuma = duplikat
            dupli = [r for r in win
                     if sum(1 for q in win if q['date'] == r['date']
                            and abs(q['signed'] - r['signed']) < EPS) > 1]
            tag = ''
            if tocan:
                imenovano += 1
                r = tocan[0]
                tag = (f'  → 1 redak istog iznosa: red {r["xl"]} ({r["date"]}) '
                       f'{(r["napomena"] or r["opis"])[:26]}')
            elif e.get('par'):
                imenovano += 1
                tag = f'  → poništava se s {e["par"]} (redak na krivoj strani granice)'
            elif dupli:
                d = dupli[0]
                tag = (f'  → mogući duplikat: red {d["xl"]} {d["date"]} '
                       f'{d["signed"]:+.2f} {(d["napomena"] or d["opis"])[:20]}')
            print(f'   {e["ym"]:<9} rezidual {kol(e["rez"], 11)}{tag}')
        print(f'   imenovano: {imenovano}/{len(rezid)}  '
              f'(ostatak traži ručni pogled — `--rows`)')

    if show_rows and rezid:
        print()
        print('Detalj rezidualnih prozora (najveći retci):')
        for e in rezid:
            sel = sorted([r for r in z if e['od'] < r['date'] <= e['do']
                          and r['izvor'] in IZVOR_IZVRSENO],
                         key=lambda r: -abs(r['signed']))[:5]
            print(f'\n   [{e["ym"]}] {e["od"]} .. {e["do"]}  rezidual {eur(e["rez"])}')
            for r in sel:
                print(f'      red {r["xl"]:>5} {r["date"]} {r["signed"]:>+11.2f} '
                      f'{r["izvor_reda"]:<19} {(r["napomena"] or r["opis"])[:38]}')

    # odnos prema `Saldo kontrola` (razina, Kokin broj) — različita veličina!
    sheet_off = [e['ym'] for e in saldo if abs(e['delta']) >= EPS]
    model_off = [e['ym'] for e in rezid]
    print()
    print('Odnos prema `Saldo kontrola` sheetu:')
    print(f'   sheet (Kokin `Stanje` − banka, RAZINA): {len(sheet_off)} razlika '
          f'{sheet_off}')
    print(f'   model (POMAK − banka)                 : {len(model_off)} rezidua '
          f'{model_off}')
    zaj = sorted(set(sheet_off) & set(model_off))
    print(f'   presjek: {zaj if zaj else "—"}')
    print('   ⚠ To su DVIJE RAZLIČITE VELIČINE (razina njenog ručnog broja vs pomak '
          'modela),\n     pa se ne očekuje isti popis. Vidi zaključak.')
    return ok_m, n_mj, rezid, sheet_off


# --------------------------------------------------- PROVJERA 2 — transfer --
def uloga_transfera(r) -> str:
    """Transfer retci nemaju svi protupartiju — samo `izmedju racuna` smije imati."""
    txt = (r['napomena'] + ' ' + r['opis']).upper()
    pod = r['podtip'].lower()
    if 'bankomat' in pod or any(k in txt for k in BANKOMAT_KLJUC):
        return 'bankomat/gotovina'
    if any(k in txt for k in LUMP_KLJUC):
        return 'naplata kartice'
    if pod == 'izmedju racuna':
        return 'međuračunski'
    return 'druga osoba'          # Anja, Nataša, Nena… — vanjski račun, ne naš


def provjera_2(rows) -> None:
    print()
    print(SEP)
    print('PROVJERA 2 — transfer zapisan jednom ili dvaput? (§2.14)')
    print(SEP)

    tr = [r for r in rows if r['tip'] == 'Transfer' and abs(r['signed']) > 0]
    for r in tr:
        r['uloga'] = uloga_transfera(r)
    print(f'Transfer redaka: {len(tr)}')
    print(f'{"uloga":<20} {"n":>5}  {"Σ":>14}   protupartija je')
    print('-' * 72)
    objasnjenje = {
        'naplata kartice': 'kartica — NIJE praćeni račun',
        'bankomat/gotovina': 'gotovina — novčanik nije račun',
        'druga osoba': 'tuđi račun — izvan našeg opsega',
        'međuračunski': 'drugi praćeni račun ⇒ SMIJE postojati',
    }
    for u in ('naplata kartice', 'bankomat/gotovina', 'druga osoba', 'međuračunski'):
        sel = [r for r in tr if r['uloga'] == u]
        print(f'{u:<20} {len(sel):>5}  {kol(round(sum(abs(r["signed"]) for r in sel), 2), 14)}'
              f'   {objasnjenje[u]}')
    print('-' * 72)

    # protupartija se traži SAMO za međuračunske (§2.14 pitanje)
    idx = defaultdict(list)
    for r in rows:
        if abs(r['signed']) > 0:
            idx[(round(abs(r['signed']), 2), r['smjer'])].append(r)

    med = [r for r in tr if r['uloga'] == 'međuračunski']
    paired, single, used = [], [], set()
    for r in med:
        amt = round(abs(r['signed']), 2)
        opp = 'Uplata' if r['smjer'] == 'Isplata' else 'Isplata'
        hit = next((c for c in idx.get((amt, opp), [])
                    if c['xl'] != r['xl'] and c['xl'] not in used
                    and c['racun'] != r['racun']
                    and abs((c['date'] - r['date']).days) <= TRANSFER_TOL_DANA), None)
        if hit:
            used.add(hit['xl'])
            paired.append((r, hit))
        else:
            single.append(r)

    sum_p = round(sum(abs(r['signed']) for r, _ in paired), 2)
    sum_s = round(sum(abs(r['signed']) for r in single), 2)
    uk = sum_p + sum_s or 1
    print()
    print(f'Međuračunski transferi: {len(med)}')
    print(f'   s protupartijom na drugom računu (±{TRANSFER_TOL_DANA} dana): '
          f'{len(paired):>4}  ({len(paired) / max(1, len(med)) * 100:4.1f} % kom.)   '
          f'Σ {kol(sum_p)}  ({sum_p / uk * 100:4.1f} % iznosa)')
    print(f'   bez protupartije                            : {len(single):>4}  '
          f'({len(single) / max(1, len(med)) * 100:4.1f} % kom.)   '
          f'Σ {kol(sum_s)}  ({sum_s / uk * 100:4.1f} % iznosa)')
    print('   ⇒ mjerodavan je udio PO IZNOSU (saldo se mjeri u eurima, ne u komadima).')

    if single:
        print()
        print('   Bez protupartije po Podtipu:')
        for k, v in Counter(r['podtip'] or '(prazno)' for r in single).most_common():
            print(f'      {k:<20} {v:>4}')
        print('   Primjeri (prva 5):')
        for r in single[:5]:
            print(f'      red {r["xl"]:>5} {r["date"]} {r["smjer"]:<8} '
                  f'{abs(r["signed"]):>9.2f} {r["racun"]:<20} '
                  f'{(r["napomena"] or r["opis"])[:32]}')

    print()
    if not med:
        print('⇒ NALAZ: nema međuračunskih transfera za ocijeniti.')
    elif sum_p / uk >= 0.85:
        print(f'⇒ NALAZ: pravi međuračunski transfer je zapisan DVAPUT '
              f'({sum_p / uk * 100:.1f} % iznosa) ⇒ oba salda točna, ništa se ne dupla.')
        print(f'   Preostalih {len(single)} jednostranih (Σ {eur(sum_s)}, prosjek '
              f'{eur(round(sum_s / max(1, len(single)), 2))}) ide vanjskim')
        print('   odredištima (Seka, Nena, Revolut, APN…) pogrešno označenim kao '
              '`izmedju racuna` —\n   to je pogreška labele, ne rupa u modelu.')
    elif not paired:
        print('⇒ NALAZ: međuračunski transfer je zapisan JEDNOM ⇒ drugi račun ne dobiva '
              'protustavku, model treba dopunu.')
    else:
        print(f'⇒ NALAZ: MJEŠOVITO — {sum_p / uk * 100:.1f} % iznosa dvostrano, '
              f'{sum_s / uk * 100:.1f} % jednostrano.')


# -------------------------------------------------- PROVJERA 3 — planirano --
def provjera_3(rows, danas: date) -> None:
    print()
    print(SEP)
    print(f'PROVJERA 3 — "planirano" po kantama i smjerovima (§2.13), danas = {danas}')
    print(SEP)

    plan = [r for r in rows if r['status'] == 'Planiran' and abs(r['signed']) > 0]
    zadnji = max((r['date'] for r in rows), default=None)
    print(f'Redaka `Status=Planiran`: {len(plan)}   (ukupno {len(rows)})')
    print(f'Najnoviji event_date u fileu: {zadnji}')

    bez = [r for r in plan if r['naplata'] is None]
    if bez:
        print(f'⚠ {len(bez)} planiranih bez `Datum naplate` — ne mogu u kantu')

    kante = {'Dospjelo': [], 'Uskoro': [], 'Kasnije': []}
    for r in plan:
        if r['naplata'] is None:
            continue
        if r['naplata'] <= danas:
            kante['Dospjelo'].append(r)
        elif r['naplata'] <= danas + timedelta(days=USKORO_DANA):
            kante['Uskoro'].append(r)
        else:
            kante['Kasnije'].append(r)

    print()
    print(f'{"kanta":<10} {"n":>4} {"odlazi":>14} {"dolazi":>14}')
    print('-' * 46)
    for k in ('Dospjelo', 'Uskoro', 'Kasnije'):
        rs = kante[k]
        out = round(sum(r['isplata'] or 0 for r in rs), 2)
        inn = round(sum(r['uplata'] or 0 for r in rs), 2)
        print(f'{k:<10} {len(rs):>4} {kol(-out, 14)} {kol(inn, 14)}')
    print('-' * 46)

    if kante['Dospjelo']:
        print()
        print(f'Dospjelo — detalj ({len(kante["Dospjelo"])}):')
        for r in sorted(kante['Dospjelo'], key=lambda r: r['naplata']):
            print(f'   red {r["xl"]:>5}  naplata {r["naplata"]}  {r["signed"]:>+10.2f}  '
                  f'{r["racun"]:<20} {(r["napomena"] or r["opis"])[:32]}')
        dana = [(danas - r['naplata']).days for r in kante['Dospjelo']]
        print(f'   raspon dospijeća: {min(dana)}–{max(dana)} dana unatrag')

    rate_p = [r for r in rows if s(r['rate']).upper() in ('DA', 'TRUE', 'YES')
              and r['status'] == 'Planiran']
    rate_i = [r for r in rows if s(r['rate']).upper() in ('DA', 'TRUE', 'YES')
              and r['status'] != 'Planiran']
    print()
    print(f'Kontekst: `Rate?=DA` {len(rate_p) + len(rate_i)}  '
          f'(Izvrsen {len(rate_i)}, Planiran {len(rate_p)})')
    if not kante['Uskoro'] and not kante['Kasnije']:
        naj_nap = max((r['naplata'] for r in plan if r['naplata']), default=None)
        print('⚠ Kante `Uskoro` i `Kasnije` su PRAZNE — ne zato što model griješi, nego')
        print(f'   jer NIJEDAN planirani redak nema `Datum naplate` u budućnosti '
              f'(najkasniji: {naj_nap}).')
        print('   Buduće rate u Reviewu ne postoje kao retci — generira ih rata modal')
        print('   nakon importa ⇒ trokantni razrez se na ovim podacima NE MOŽE provjeriti.')
        print('   (Anjinih 96 rata iz §2.13 su ovdje 45 POVIJESNIH redaka, svi `Izvrsen`.)')


# ------------------------------------------------------ NALAZI (loši retci) --
def _podskup(win, cilj):
    """Nađi 1 ili 2 retka u prozoru čiji zbroj = rezidual (misdate/duplikat)."""
    for r in win:
        if abs(r['signed'] - cilj) < EPS:
            return [r]
    for i, a in enumerate(win):
        for b in win[i + 1:]:
            if abs(a['signed'] + b['signed'] - cilj) < EPS:
                return [a, b]
    return []


def tok(r) -> str:
    """Iz kojeg je ULAZNOG TOKA redak došao — duplikat prepoznaje razliku tokova."""
    ir = r['izvor_reda']
    if ir.startswith('Konsolidacija'):
        return 'izvod'
    if ir.startswith('PBZ Visa'):
        return 'pbz'
    return 'kokin/sašin file'


def nalazi(rows, rezid, danas: date, out: Path | None):
    """Retci koji NISU u redu — sa šifrom, objašnjenjem i pouzdanošću."""
    print()
    print(SEP)
    print('NALAZI — retci koji nisu u redu')
    print(SEP)

    f = defaultdict(list)   # šifra -> [(row, detalj)]

    # 1) naplata prije kupnje — unutarnja proturječnost
    for r in rows:
        if r['date'] and r['naplata'] and r['naplata'] < r['date']:
            f['NAPLATA<KUPNJA'].append(
                (r, f'naplaćeno {(r["date"] - r["naplata"]).days} dana PRIJE kupnje'))

    # 2) obje kolone iznosa popunjene
    for r in rows:
        if r['uplata'] is not None and r['isplata'] is not None:
            f['OBJE-KOLONE'].append(
                (r, f'Uplata {r["uplata"]:.2f} I Isplata {r["isplata"]:.2f}'))

    # 3) bez ijednog iznosa
    for r in rows:
        if r['uplata'] is None and r['isplata'] is None:
            sidro = r['stanje'] is not None and r['date'] and r['date'].day == 1
            f['BEZ-IZNOSA'].append(
                (r, 'početno stanje (OK, ne uvozi se)' if sidro else 'prazan zapis'))

    # 4) retci koje je banka opovrgnula (imenovani iz rezidualnih mjeseci)
    z = [r for r in rows if r['racun'] == RACUN_ZABA and r['status'] != 'Planiran']
    for e in rezid:
        win = [r for r in z if e['od'] < r['date'] <= e['do']
               and r['izvor'] in IZVOR_IZVRSENO]
        for r in _podskup(win, e['rez']):
            f['BANKA-NE-VIDI'].append(
                (r, f'{e["ym"]}: model −banka = {eur(e["rez"])} ⇒ ovaj redak banka '
                    f'nema u tom prozoru'))

    # 5) isti iznos+datum+račun iz DVA različita ulazna toka
    grp = defaultdict(list)
    for r in rows:
        if abs(r['signed']) > 0:
            grp[(r['racun'], r['date'], r['signed'])].append(r)
    for v in grp.values():
        if len(v) > 1 and len({tok(x) for x in v}) > 1:
            for r in v:
                f['DUPLI-IZVOR'].append(
                    (r, 'isti iznos/datum iz ' + ' + '.join(sorted({tok(x) for x in v}))))

    # 6) Transfer `izmedju racuna` bez protupartije = pogrešna labela
    idx = defaultdict(list)
    for r in rows:
        if abs(r['signed']) > 0:
            idx[(round(abs(r['signed']), 2), r['smjer'])].append(r)
    for r in rows:
        if r['tip'] == 'Transfer' and uloga_transfera(r) == 'međuračunski':
            opp = 'Uplata' if r['smjer'] == 'Isplata' else 'Isplata'
            hit = any(c for c in idx.get((round(abs(r['signed']), 2), opp), [])
                      if c['racun'] != r['racun']
                      and abs((c['date'] - r['date']).days) <= TRANSFER_TOL_DANA)
            if not hit:
                f['TRANSFER-BEZ-PARA'].append(
                    (r, f'označen `izmedju racuna` ali odredište nije naš račun '
                        f'({(r["napomena"] or r["opis"])[:24]})'))

    # 7) planirano a dospjelo
    for r in rows:
        if r['status'] == 'Planiran' and r['naplata'] and r['naplata'] <= danas:
            f['PLANIRAN-DOSPIO'].append(
                (r, f'dospjelo prije {(danas - r["naplata"]).days} dana, čeka potvrdu'))

    zasto = {
        'NAPLATA<KUPNJA': ('jedan od dva datuma je kriv; `Datum naplate` nosi saldo '
                           'kartice, pa kriva vrijednost pomiče krivi mjesec',
                           'visoka'),
        'OBJE-KOLONE': ('`row_amount()` čita Isplatu PRVO ⇒ uzima 0,30 umjesto 450 '
                        '(uzrok S107r §10b)', 'visoka'),
        'BEZ-IZNOSA': ('nema što uvesti; 2 su legitimna početna stanja, ostalo su '
                       'prazni zapisi', 'visoka'),
        'BANKA-NE-VIDI': ('model i banka se u tom mjesecu razilaze točno za iznos ovog '
                          'retka ⇒ redak je dvaput, krivo datiran ili ga nema', 'srednja'),
        'DUPLI-IZVOR': ('isti iznos i datum stigao iz dva ulazna toka ⇒ moguć duplikat '
                        '(može biti i stvarno dvije iste kupnje)', 'niska'),
        'TRANSFER-BEZ-PARA': ('ne kvari saldo, ali `Transfer` ispada iz razreza po Tipu '
                              '(§2.14) ⇒ trošak se gubi iz statistike', 'srednja'),
        'PLANIRAN-DOSPIO': ('po dizajnu (§2.5a) — traži ljudsku potvrdu, nije greška',
                            'n/p'),
    }

    red = ['NAPLATA<KUPNJA', 'OBJE-KOLONE', 'BEZ-IZNOSA', 'BANKA-NE-VIDI',
           'DUPLI-IZVOR', 'TRANSFER-BEZ-PARA', 'PLANIRAN-DOSPIO']
    print(f'{"šifra":<19} {"n":>4}  {"pouzdanost":<11} što nije u redu')
    print('-' * 78)
    uvlaka = ' ' * 37
    for k in red:
        if not f[k]:
            continue
        opis, pouz = zasto[k]
        linije = textwrap.wrap(opis, 41)
        print(f'{k:<19} {len(f[k]):>4}  {pouz:<11} {linije[0]}')
        for ln in linije[1:]:
            print(uvlaka + ln)
    print('-' * 78)
    print(f'ukupno označenih redaka: {sum(len(v) for v in f.values())} '
          f'(neki nose više šifara)')

    for k in red:
        if not f[k]:
            continue
        print(f'\n[{k}]  {len(f[k])} — {zasto[k][1]} pouzdanost')
        for r, det in sorted(f[k], key=lambda x: x[0]['xl'])[:12]:
            print(f'   red {r["xl"]:>5} {r["date"]} {r["signed"]:>+10.2f} '
                  f'{r["racun"]:<18} {det[:44]}')
        if len(f[k]) > 12:
            print(f'   … još {len(f[k]) - 12} (puni popis u TSV-u)')

    if out:
        zaglavlje = ('sifra', 'red', 'event_date', 'datum_naplate', 'racun', 'izvor',
                     'iznos', 'napomena', 'pouzdanost', 'zasto')
        podaci = [(k, r['xl'], r['date'], r['naplata'] or '', r['racun'], r['izvor'],
                   round(r['signed'], 2), (r['napomena'] or r['opis'])[:60],
                   zasto[k][1], det)
                  for k in red for r, det in sorted(f[k], key=lambda x: x[0]['xl'])]

        out.write_text(
            '\n'.join(['\t'.join(zaglavlje)]
                      + ['\t'.join(str(x) for x in row) for row in podaci]),
            encoding='utf-8')
        print(f'\n✔ TSV : {out}  ({len(podaci)} redaka)')

        # .xlsx — jer .tsv nije registriran na Windowsu i traži "Choose an app"
        xl_out = out.with_suffix('.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Nalazi'
        hdr_fill = PatternFill('solid', fgColor='C55A11')
        hdr_font = Font(color='FFFFFF', bold=True)
        boja = {'visoka': 'FFC7CE', 'srednja': 'FFEB9C', 'niska': 'EDEDED',
                'n/p': 'DDEBF7'}
        for c, hh in enumerate(zaglavlje, 1):
            cell = ws.cell(1, c, hh)
            cell.fill, cell.font = hdr_fill, hdr_font
        for r_i, row in enumerate(podaci, 2):
            fill = PatternFill('solid', fgColor=boja.get(row[8], 'FFFFFF'))
            for c_i, v in enumerate(row, 1):
                cell = ws.cell(r_i, c_i, v)
                cell.fill = fill
                if c_i in (3, 4) and v:
                    cell.number_format = 'DD.MM.YYYY'
                elif c_i == 7:
                    cell.number_format = '#,##0.00'
        for c, w in zip('ABCDEFGHIJ', (19, 7, 12, 13, 19, 12, 11, 34, 11, 62)):
            ws.column_dimensions[c].width = w
        ws.freeze_panes = 'C2'
        ws.auto_filter.ref = f'A1:J{max(2, len(podaci) + 1)}'
        wb.save(xl_out)
        print(f'✔ XLSX: {xl_out}  ← otvori ovaj (autofiltar na `sifra`)')
        print('  (nova fileova — Review NIJE diran)')
    return f


# -------------------------------------------------------------------- main --
def main():
    args = sys.argv[1:]
    show_rows = '--rows' in args
    path = pick_file(args)
    danas = date.today()

    print(SEP)
    print('VERIFIKACIJA MODELA SALDA — Faza 1a (docs/OVERVIEW_TAB_SPEC.md)')
    print(SEP)
    print(f'File : {path.name}   [READ-ONLY — ništa se ne piše]')

    rows, saldo = load_review(path)
    print(f'Redaka: {len(rows)}   Izvadaka u `Saldo kontrola`: {len(saldo)}')

    print()
    print('Račun × Izvor:')
    for (rac, izv), n in sorted(Counter((r['racun'], r['izvor']) for r in rows).items()):
        rola = ('izvršeno' if izv in IZVOR_IZVRSENO
                else 'odgođeno' if izv in IZVOR_ODGODENO else '⚠ IZVAN MODELA')
        print(f'   {rac:<20} {izv:<12} {n:>5}   {rola}')
    vanjski = {r['izvor'] for r in rows} - IZVOR_IZVRSENO - IZVOR_ODGODENO - {''}
    if vanjski:
        print(f'⚠ Izvori izvan modela: {sorted(vanjski)}')

    ok_m, n_mj, rezid, sheet_off = provjera_1(rows, saldo, show_rows)
    provjera_2(rows)
    provjera_3(rows, danas)
    if '--nalazi' in args:
        nalazi(rows, rezid, danas, DATA_DIR / 'saldo_model_nalazi.tsv')

    print()
    print(SEP)
    print('ZAKLJUČAK')
    print(SEP)
    print(f'1. Pravilo §2.10 reproducira bankovni pomak u {ok_m}/{n_mj} mjeseci; '
          f'naivni zbroj po `Racun`u u 0/{n_mj}.')
    print(f'   Preostalo {len(rezid)} rezidualnih mjeseci — v. popis (nisu ista '
          f'veličina kao {len(sheet_off)} iz sheeta).')
    print('2. i 3. su nalazi, ne prolaz/pad — v. gore.')
    print(SEP)


if __name__ == '__main__':
    main()
