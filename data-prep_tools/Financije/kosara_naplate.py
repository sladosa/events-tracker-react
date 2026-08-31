# -*- coding: utf-8 -*-
"""
kosara_naplate.py — koji kartični redak nosi krivi `Datum naplate`. S123, 2026-08-31.

NALAZ KOJI GA JE TRAŽIO
    Košara `Datum naplate = 11.07.2026` ∧ `Izvor = Mastercard` nosi 73 retka i
    **2.231,02 €**, a banka je tog dana skinula **1.244,74 €**. Izmjereno ovim
    alatom (2026-08-31), razlika se razlaže na ČETIRI različita pitanja:

      40 redaka   946,48   OK — slažu se s pravilom
      21 redak    832,86   RATA — pravilo ne vrijedi, treba plan otplate
      11 redaka   431,10   KRIVI MJESEC ⇒ pripadaju izvodu 11.08.
       1 redak     20,58   KRIVI MJESEC ⇒ pripada izvodu 11.06.

    ⚠ Ni nakon micanja krivo datiranih ne zatvara se: 946,48 + 832,86 = 1.779,34
      naspram 1.244,74. Dakle ni sve rate ne pripadaju ovom izvodu. Ostatak može
      razriješiti samo `MC_2026-06.pdf` — pravilo je iscrpljeno.

    Zato alat ne „popravlja košaru" nego **razdvaja slučajeve** i za svaki kaže
    odakle dokaz dolazi.

    ⚠ RATA NIJE KUPOVINA. Sve rate jedne kupovine dijele `event_date` = dan
      kupnje, a razlikuje ih `Datum naplate` iz plana otplate. Pravilo „11. u
      sljedećem mjesecu" bi ih redom proglasilo krivima i poslalo čovjeka da
      „popravi" 21 vjerojatno ispravan redak — zato se za njih ne izvodi.

ZAŠTO IZLAZ IDE U APP FORMATU
    Ispravak mora natrag u bazu, a jedini put koji app ima je Excel import.
    Datoteka zato ima LEGEND + EVENT DATA sekciju s `event_id`-em, pa je uvoz
    UPDATE postojećih redaka, ne novi zapisi.

    ⚠ U atributnoj koloni `Datum naplate` stoji **POSTOJEĆA** vrijednost, ne
      prijedlog. Slučajan uvoz ovog filea zato ne mijenja ništa — a to je
      namjerno: prijedlog koji se sam uveze je tvrdnja koju nitko nije potvrdio.
      Prijedlog stoji u dijagnostičkim kolonama desno; `--predlozi` ga upiše u
      atributnu kolonu za retke s jednoznačnom dijagnozom.

PRAVILO NAPLATE (isto kao `fix_datum_naplate_statement.py` i `set_attribute`)
    Mastercard → 11. u mjesecu NAKON mjeseca kupovine
    Visa       →  3. u mjesecu NAKON mjeseca kupovine
    ⚠ Pravilo vrijedi za tipičan ciklus. Izvod je jači dokaz od pravila: kad se
      razilaze, PDF je u pravu (S110: izvod se ne zatvara na kalendarski kraj).

POKAZIVAČ NA KOKIN FILE
    Za svaki redak se traži par u njezinoj Excelici po `(datum, iznos)` i ispisuje
    `Sheet` + `redak`. ⚠ Njezina kolona `Datum` (C) je dan kad novac napusti račun;
    dok naplata nije poznata, C je prazan a dan troška stoji u koloni G — zato se
    gledaju OBJE. Bez para ostaje prazno: „nije nađeno" nije isto što i „ne postoji".

Pokretanje:
    python kosara_naplate.py --naplata 2026-07-11
    python kosara_naplate.py --naplata 2026-07-11 --banka 1244.74
    python kosara_naplate.py --mjesec 2026-08 --izvor Visa
    python kosara_naplate.py --naplata 2026-07-11 --predlozi     # upiše prijedlog
    python kosara_naplate.py --naplata 2026-07-11 --env test
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
import urllib.request
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
AREA_PROD = 'de8662e6-54f7-4ded-ab42-a786e7456067'
CAT_PROD = '986a4612-86a2-49fa-b73f-a29e048e5750'
KOKA_DEFAULT = ROOT / 'data-prep_data' / 'Financije' / 'Financije 2026-08-23.xlsx'

# Redoslijed atributnih kolona u izlazu. Fiksan je namjerno: import ih čita po
# LEGEND-u, ali čovjek ih čita očima i mora ih naći na istom mjestu kao u appu.
ATTR_ORDER = [
    'Racun', 'Izvor', 'Smjer', 'Uplata', 'Isplata', 'Tip', 'Podtip',
    'Izvod opis', 'Rate?', 'Broj rata', 'Rata br', 'Datum naplate',
    'Status', 'Stanje', 'Valuta',
]
DIAG_COLS = ['Dijagnoza', 'Predlozena naplata', 'Koka sheet', 'Koka redak',
             'Koka opis', 'Koka datum']

HDR_FILL = PatternFill('solid', fgColor='FF4472C4')
DIAG_FILL = PatternFill('solid', fgColor='FFFFF2CC')
BAD_FILL = PatternFill('solid', fgColor='FFFFC7CE')
OK_FILL = PatternFill('solid', fgColor='FFC6EFCE')


# ── DB ──────────────────────────────────────────────────────────────────────
def load_env(which: str) -> tuple[str, str]:
    fn = '.env.prod.local' if which == 'prod' else '.env.testing'
    path = ROOT / fn
    if not path.exists():
        sys.exit(f'Nema {fn} — bez njega alat ne može čitati bazu.')
    env = {}
    for line in path.read_text(encoding='utf-8').splitlines():
        if '=' in line and not line.strip().startswith('#'):
            k, v = line.split('=', 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    url = env.get('SUPABASE_URL') or env.get('VITE_SUPABASE_URL')
    key = env.get('SUPABASE_SERVICE_ROLE_KEY') or env.get('VITE_SUPABASE_ANON_KEY')
    if not url or not key:
        sys.exit(f'{fn} nema SUPABASE_URL / ključ.')
    return url, key


def rest(url: str, key: str, path: str) -> list[dict]:
    """⚠ PostgREST reže na 1000 redaka BEZ GREŠKE — svaki upit se stranicira,
    i to s `order`, inače se stranice preklope i istovremeno preskoče (S108)."""
    out, off = [], 0
    while True:
        req = urllib.request.Request(
            f'{url}/rest/v1/{path}&order=id',
            headers={'apikey': key, 'Authorization': f'Bearer {key}',
                     'Range': f'{off}-{off + 999}'},
        )
        rows = json.load(urllib.request.urlopen(req))
        out += rows
        off += len(rows)
        if len(rows) < 1000:
            return out


def load_rows(url: str, key: str) -> tuple[list[dict], dict]:
    defs = {d['name']: d for d in rest(url, key, f'attribute_definitions?category_id=eq.{CAT_PROD}&select=id,name,slug,data_type')}
    by_id = {d['id']: d['name'] for d in defs.values()}
    events = {e['id']: e for e in rest(url, key, f'events?category_id=eq.{CAT_PROD}&select=id,event_date,session_start,created_at,comment,user_id')}
    vals = defaultdict(dict)
    for a in rest(url, key, f'event_attributes?attribute_definition_id=in.({",".join(d["id"] for d in defs.values())})&select=id,event_id,attribute_definition_id,value_text,value_number,value_datetime,value_boolean'):
        n = by_id.get(a['attribute_definition_id'])
        if not n:
            continue
        v = a['value_text']
        if v is None:
            v = a['value_number']
        if v is None:
            v = a['value_datetime']
        if v is None:
            v = a['value_boolean']
        vals[a['event_id']][n] = v
    rows = []
    for eid, ev in events.items():
        rows.append({**ev, 'attrs': vals.get(eid, {})})
    return rows, defs


# ── Kokin file ──────────────────────────────────────────────────────────────
def norm(s: str) -> str:
    return unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode().lower().strip()


def parse_koka_date(v) -> str | None:
    """⚠ 103 njena retka nose datum kao TEKST (`11.05.23.`, `28.6.23.`,
    `29.2.2024.`) — alat koji prima samo `datetime` progutao bi ih bez poruke."""
    if v is None or v == '':
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    m = re.match(r'^\s*(\d{1,2})\.\s*(\d{1,2})\.\s*(\d{2,4})\.?\s*$', str(v))
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    try:
        return date(y, mo, d).strftime('%Y-%m-%d')
    except ValueError:
        return None


def index_koka(path: Path) -> dict:
    """Ključ `(datum, iznos)` → popis (sheet, redak, opis, koja datumska kolona).
    Gledaju se OBJE datumske kolone (C i G) i oba iznosa (uplata i isplata)."""
    if not path.exists():
        print(f'⚠ Kokin file nije nađen ({path.name}) — stupci `Koka *` ostaju prazni.')
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    idx = defaultdict(list)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cells = list(row) + [None] * 10
            opis = cells[1]
            for col_letter, raw in (('C', cells[2]), ('G', cells[6])):
                dt = parse_koka_date(raw)
                if not dt:
                    continue
                for amt in (cells[3], cells[4]):
                    try:
                        a = round(float(amt), 2)
                    except (TypeError, ValueError):
                        continue
                    if a == 0:
                        continue
                    idx[(dt, a)].append((sn, r, str(opis or ''), col_letter))
    wb.close()
    return idx


# ── Dijagnoza ───────────────────────────────────────────────────────────────
def expected_naplata(izvor: str, event_date: str) -> str | None:
    """MC → 11. sljedećeg mjeseca, Visa → 3. sljedećeg. Račun/Cash → isti dan."""
    y, m, _ = (int(x) for x in event_date.split('-'))
    if izvor in ('Racun', 'Cash'):
        return event_date
    day = {'Mastercard': 11, 'Visa': 3}.get(izvor)
    if day is None:
        return None
    m += 1
    if m == 13:
        m, y = 1, y + 1
    return date(y, m, day).strftime('%Y-%m-%d')


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument('--naplata', help='točan datum naplate košare, YYYY-MM-DD')
    g.add_argument('--mjesec', help='svi retci s naplatom u mjesecu, YYYY-MM')
    ap.add_argument('--izvor', default='Mastercard', help='Mastercard | Visa | sve')
    ap.add_argument('--banka', type=float, help='iznos skupne naplate s izvoda (kontrola)')
    ap.add_argument('--koka', default=str(KOKA_DEFAULT), help='putanja do Kokine Excelice')
    ap.add_argument('--env', default='prod', choices=['prod', 'test'])
    ap.add_argument('--out', help='izlazna xlsx (zadano: uz Kokin file)')
    ap.add_argument('--predlozi', action='store_true',
                    help='upiši prijedlog U atributnu kolonu (samo jednoznačne slučajeve)')
    args = ap.parse_args()

    url, key = load_env(args.env)
    print(f'Čitam bazu ({args.env})…')
    rows, defs = load_rows(url, key)

    def naplata_of(r):
        v = r['attrs'].get('Datum naplate')
        return str(v)[:10] if v else ''

    sel = []
    for r in rows:
        n = naplata_of(r)
        if args.naplata and n != args.naplata:
            continue
        if args.mjesec and not n.startswith(args.mjesec):
            continue
        if args.izvor != 'sve' and r['attrs'].get('Izvor') != args.izvor:
            continue
        sel.append(r)
    if not sel:
        sys.exit('Nijedan redak ne odgovara filtru — provjeri --naplata / --izvor.')
    sel.sort(key=lambda r: (r['event_date'], r['session_start'] or ''))

    koka = index_koka(Path(args.koka))

    # ── dijagnoza + zbrojevi ────────────────────────────────────────────────
    tot = 0.0
    buckets = defaultdict(lambda: [0, 0.0])
    for r in sel:
        a = r['attrs']
        iz = a.get('Izvor') or ''
        stored = naplata_of(r)
        exp = expected_naplata(iz, r['event_date'])
        iznos = float(a.get('Isplata') or 0) - float(a.get('Uplata') or 0)
        tot += iznos
        # ⚠ RATA NIJE KUPOVINA. Njezin `Datum naplate` dolazi iz plana otplate
        #   (rata N dospijeva N-ti mjesec), a ne iz mjeseca kupovine — sve rate
        #   jedne kupovine dijele `event_date` = dan kupnje. Pravilo bi ih zato
        #   redom proglasilo krivima i poslalo čovjeka da „popravi" 11 ispravnih
        #   redaka. Dijagnoza se za njih ne izvodi nego se izričito odbija.
        je_rata = bool(a.get('Rate?')) or a.get('Rata br') not in (None, '')
        if not stored:
            dg = 'BEZ DATUMA'
        elif exp is None:
            dg = 'NEPOZNAT IZVOR'
        elif je_rata:
            dg = 'RATA — pravilo ne vrijedi, treba plan otplate'
        elif stored == exp:
            dg = 'OK (slaže se s pravilom)'
        else:
            dg = f'KRIVI MJESEC (pravilo kaže {exp})'
        if a.get('Status') == 'Planiran':
            dg += ' · PLANIRAN'
        r['_dg'], r['_exp'], r['_iznos'] = dg, exp or '', iznos
        b = buckets[dg.split(' ·')[0]]
        b[0] += 1
        b[1] += iznos

        par = koka.get((r['event_date'], round(abs(iznos), 2)), [])
        r['_koka'] = par[0] if len(par) == 1 else (('VIŠE PARA', len(par), '', '') if par else None)

    print(f'\nKošara: {len(sel)} redaka, Σ = {tot:,.2f}')
    for k, (n, s) in sorted(buckets.items(), key=lambda x: -x[1][1]):
        print(f'   {n:4} × {k:34} Σ={s:10,.2f}')
    if args.banka is not None:
        print(f'\n   banka s izvoda      = {args.banka:10,.2f}')
        print(f'   RAZLIKA             = {tot - args.banka:10,.2f}')
        if abs(tot - args.banka) > 0.005:
            print('   ⚠ Košara se ne zatvara — ne potvrđuj retke dok se ne razjasni.')
    nadjeno = sum(1 for r in sel if r['_koka'] and r['_koka'][0] != 'VIŠE PARA')
    print(f'\n   par u Kokinom fileu: {nadjeno}/{len(sel)}'
          f'   (bez para nije dokaz da ga nema — v. zaglavlje)')

    # ── izlaz ───────────────────────────────────────────────────────────────
    out = Path(args.out) if args.out else Path(args.koka).parent / (
        f'kosara_{(args.naplata or args.mjesec).replace("-", "")}_{args.izvor.lower()}.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Events'

    ws.cell(1, 1, 'ATTRIBUTE LEGEND:')
    for i, h in enumerate(['Col', 'Area', 'Category_Path', 'Attribute', 'Type', 'Unit'], start=1):
        c = ws.cell(2, i, h)
        c.font = Font(bold=True, color='FFFFFFFF')
        c.fill = HDR_FILL
    for i, name in enumerate(ATTR_ORDER):
        col = 9 + i                                  # I, J, K…
        d = defs.get(name)
        ws.cell(3 + i, 1, openpyxl.utils.get_column_letter(col))
        ws.cell(3 + i, 2, 'Financije_all')
        ws.cell(3 + i, 3, 'Transakcija')             # ⚠ BEZ imena aree (Activities format)
        ws.cell(3 + i, 4, name)
        ws.cell(3 + i, 5, d['data_type'] if d else 'text')

    r0 = 3 + len(ATTR_ORDER) + 1
    ws.cell(r0, 1, 'EVENT DATA:')
    hdr = ['event_id', 'Area', 'Category_Path', 'event_date', 'session_start',
           'created_at', 'User', 'comment'] + ATTR_ORDER + DIAG_COLS
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(r0 + 1, i, h)
        c.font = Font(bold=True, color='FFFFFFFF')
        c.fill = HDR_FILL if i <= 8 + len(ATTR_ORDER) else PatternFill('solid', fgColor='FFBF8F00')

    emails = {}
    for p in rest(url, key, 'profiles?select=id,email&limit=100'):
        emails[p['id']] = p['email']

    for j, r in enumerate(sel):
        rr = r0 + 2 + j
        a = r['attrs']
        ws.cell(rr, 1, r['id'])                       # UUID ⇒ UPDATE, ne INSERT
        ws.cell(rr, 2, 'Financije_all')
        ws.cell(rr, 3, 'Transakcija')
        ws.cell(rr, 4, r['event_date'])
        ws.cell(rr, 5, (r['session_start'] or '')[11:16])   # ⚠ TEKST "HH:MM"
        ws.cell(rr, 5).number_format = '@'
        ws.cell(rr, 7, emails.get(r['user_id'], ''))       # ⚠ bez emaila redak je „tuđi"
        ws.cell(rr, 8, r['comment'] or '')
        for i, name in enumerate(ATTR_ORDER):
            v = a.get(name)
            if name == 'Datum naplate':
                cur = str(v)[:10] if v else ''
                # Prijedlog ide u atributnu kolonu SAMO uz --predlozi i samo za
                # jednoznačan slučaj; inače ondje stoji postojeća vrijednost i
                # uvoz ovog filea ništa ne mijenja.
                if args.predlozi and r['_dg'].startswith('KRIVI MJESEC'):  # rate su isključene dijagnozom
                    v = r['_exp']
                else:
                    v = cur
            ws.cell(rr, 9 + i, v if v is not None else '')
        base = 9 + len(ATTR_ORDER)
        ws.cell(rr, base, r['_dg']).fill = OK_FILL if r['_dg'].startswith('OK') else BAD_FILL
        ws.cell(rr, base + 1, r['_exp'])
        k = r['_koka']
        if k and k[0] != 'VIŠE PARA':
            ws.cell(rr, base + 2, k[0])
            ws.cell(rr, base + 3, k[1])
            ws.cell(rr, base + 4, k[2][:60])
            ws.cell(rr, base + 5, f'kol. {k[3]}')
        elif k:
            ws.cell(rr, base + 2, f'{k[1]} mogućih parova — provjeri ručno')
        for cc in range(base, base + len(DIAG_COLS)):
            if not ws.cell(rr, cc).fill.fgColor.rgb or ws.cell(rr, cc).fill.fgColor.rgb == '00000000':
                ws.cell(rr, cc).fill = DIAG_FILL

    for i, w in enumerate([38, 14, 14, 12, 10, 10, 30, 34], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for i in range(9, 9 + len(ATTR_ORDER) + len(DIAG_COLS)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18
    ws.freeze_panes = ws.cell(r0 + 2, 1)
    ws.auto_filter.ref = (f'A{r0 + 1}:'
                          f'{openpyxl.utils.get_column_letter(len(hdr))}{r0 + 1 + len(sel)}')

    info = wb.create_sheet('Nalaz')
    info['A1'] = 'Kako čitati ovaj file'
    info['A1'].font = Font(bold=True, size=12)
    for i, line in enumerate([
        '',
        f'Košara: {args.naplata or args.mjesec} · Izvor = {args.izvor} · {len(sel)} redaka · Σ {tot:,.2f}',
        (f'Banka s izvoda: {args.banka:,.2f} · RAZLIKA {tot - args.banka:,.2f}'
         if args.banka is not None else 'Iznos s izvoda nije zadan (--banka).'),
        '',
        'Kolone lijevo od `Dijagnoza` su format app importa — file se smije uvesti natrag.',
        'U koloni `Datum naplate` stoji POSTOJEĆA vrijednost, pa uvoz bez izmjena ne mijenja ništa.',
        'Ispravak: prepiši `Predlozena naplata` u `Datum naplate`, pa uvezi (Activities → Import).',
        '',
        'Dijagnoza je izvedena iz PRAVILA (MC 11., Visa 3. sljedećeg mjeseca).',
        'Izvod je jači dokaz od pravila — kad se razilaze, vjeruj PDF-u.',
        '',
        '`Koka sheet/redak` je par nađen po (datum, iznos) u njezinoj Excelici.',
        'Prazno = nije nađen par, NIJE dokaz da redak kod nje ne postoji.',
    ], start=2):
        info.cell(i, 1, line)
    info.column_dimensions['A'].width = 100

    wb.save(out)
    print(f'\nZapisano: {out}')
    if args.predlozi:
        n = sum(1 for r in sel if r['_dg'].startswith('KRIVI MJESEC'))
        print(f'⚠ `--predlozi`: prijedlog upisan u `Datum naplate` za {n} redaka.')
        print('  Provjeri ih prije uvoza — pravilo nije izvod.')


if __name__ == '__main__':
    main()
