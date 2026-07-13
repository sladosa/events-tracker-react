# -*- coding: utf-8 -*-
"""
enrich_from_izvoda.py  (S107c prototip, 2026-07-12)
===================================================
Izvod enrichment (FINANCIJE_MIGRACIJA.md §12.5): čita bankovne e-izvode (PDF) iz
`data-prep_data/Financije/izvodi/`, matcha transakcije na redove REVIEW Excela
(datum ±2 dana + iznos + smjer, isti mehanizam kao Za Sašu labele) i upisuje
opis s izvoda u NOVE kolone Review sheeta:

    'Izvod opis'  — opis prometa s izvoda (primatelj/merchant)
    'Izvod file'  — iz kojeg filea/retka dolazi (kontrola)

RUČNI RAD SE NE DIRA — pune se samo 'Izvod *' kolone. Nakon ovoga pokreni
apply_rules.py: pravila pretražuju i 'Izvod *' kolone → auto Tip/Podtip.

Izvor transakcija (S107d): ako postoji `izvodi/Izvodi_transakcije.xlsx`
(output inventory_izvoda.py — pokreni NJEGA prvo!) čita se on; inače legacy
mode parsira PDF-ove iz izvodi/ po filename prefixu. Nematchane transakcije
(potencijalno RETCI KOJI FALE u Kokinom Excelu) se upisuju u 'Nematchano'
sheet Izvodi_transakcije.xlsx.

Podržani formati (registry SOURCE_TYPES — proširivati po potrebi):
  ZABA     — izvadak Kokinog tekućeg računa ✅ (Izvor=Racun redovi)
  MC       — ZABA Mastercard izvod KARTICE ("Obavijest o učinjenim troškovima")
             ✅ (Izvor=Mastercard; uklj. Sašinu dodatnu karticu — [kartica: SAŠA])
  PBZVISA  — PBZ Card Visa Gold mjesečni račun ✅ (matcha na Izvor=Mastercard
             jer Koka nema 'Visa' izvor — praktički sve završi u Nematchano)
  RF       — Sašin Raiffeisen: NEMA tekst-sloja → OCR (rf_ocr.py, strip-based
             RapidOCR + stanje-chain validacija; sumnjivi redovi = '[OCR?]')

Pokretanje (review file zatvoren u Excelu!):
  Financije\\run.bat enrich_from_izvoda.py            → najnoviji review
  ... enrich_from_izvoda.py --dry                     → bez snimanja, samo report
  ... enrich_from_izvoda.py <review.xlsx> [--dry]
"""

import re
import shutil
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import pdfplumber
from openpyxl.styles import Border, Font, PatternFill, Side

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR   = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")
IZVODI_DIR = DATA_DIR / 'izvodi'
TX_XLSX    = IZVODI_DIR / 'Izvodi_transakcije.xlsx'   # output inventory_izvoda.py

RACUN_KOKA = 'Kokin tekući ZABA'
RACUN_SASA = 'Sašin tekući RF'

HDR_FILL   = PatternFill('solid', fgColor='4472C4')
WHITE_BOLD = Font(color='FFFFFF', bold=True)
THIN       = Side(style='thin')
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RE_DATE   = re.compile(r'^(\d{2})\.(\d{2})\.(\d{4})\.$')
RE_REF    = re.compile(r'^[A-Z]\d{12,18}$')
RE_AMOUNT = re.compile(r'^-?\d{1,3}(?:\.\d{3})*,\d{2}$')


def parse_amount(s: str) -> float:
    return round(float(s.replace('.', '').replace(',', '.')), 2)


# ── Parser: ZABA izvadak tekućeg računa ────────────────────────────────────────

def parse_zaba_racun(path: Path) -> list[dict]:
    """→ [{date, opis, iznos, smjer('Uplata'/'Isplata'), src}]
    Transakcijska linija: 'dd.mm.yyyy. REFERENCA Opis ... IZNOS' + eventualne
    continuation linije opisa. Priljev/Odljev se razlikuje po x-poziciji iznosa
    (granica = sredina između header riječi 'Priljev' i 'Odljev')."""
    txs: list[dict] = []
    with pdfplumber.open(path) as pdf:
        for pno, page in enumerate(pdf.pages, 1):
            words = page.extract_words()
            # gate: stranica mora imati tablicu prometa
            priljev_x = odljev_x = None
            for w in words:
                if w['text'] == 'Priljev':
                    priljev_x = (w['x0'] + w['x1']) / 2
                elif w['text'] == 'Odljev':
                    odljev_x = (w['x0'] + w['x1']) / 2
            if priljev_x is None or odljev_x is None:
                continue
            boundary = (priljev_x + odljev_x) / 2

            # grupiraj riječi u linije po y (top zaokružen)
            lines: dict[float, list] = defaultdict(list)
            for w in words:
                lines[round(w['top'], 0)].append(w)
            current: dict | None = None
            cont_left = 0
            for top in sorted(lines):
                ws_ = sorted(lines[top], key=lambda w: w['x0'])
                tokens = [w['text'] for w in ws_]
                m = RE_DATE.match(tokens[0]) if tokens else None
                if m:
                    d = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
                    # zadnji token koji izgleda kao iznos
                    amt_w = next((w for w in reversed(ws_) if RE_AMOUNT.match(w['text'])), None)
                    if amt_w is None:
                        current = None
                        continue
                    smjer = 'Uplata' if (amt_w['x0'] + amt_w['x1']) / 2 < boundary else 'Isplata'
                    mid = [t for t, w in zip(tokens[1:], ws_[1:]) if w is not amt_w]
                    if mid and RE_REF.match(mid[0]):
                        mid = mid[1:]          # makni referencu (O16023...)
                    current = {
                        'date': d, 'opis': ' '.join(mid),
                        'iznos': parse_amount(amt_w['text']), 'smjer': smjer,
                        'src': f'{path.name}:p{pno}',
                    }
                    txs.append(current)
                    cont_left = 2              # opis se prelama u max ~2 dodatna reda
                elif current and cont_left > 0 and tokens:
                    first = tokens[0]
                    if first.isupper() and ('STANJE' in first or first in ('IZVADAK', 'S/N:')):
                        current = None
                        continue
                    current['opis'] += ' ' + ' '.join(tokens)
                    cont_left -= 1
                else:
                    current = None
    return txs


# RF (Sašin Raiffeisen) nema tekst-sloj → OCR parser u rf_ocr.py (S107d)
from rf_ocr import parse_rf_ocr  # noqa: E402


# ── Parser: ZABA Mastercard IZVOD KARTICE ("Obavijest o učinjenim troškovima") ─

RE_MC_REF  = re.compile(r'^B\d{12,18}$')
RE_KARTICA = re.compile(r'^Kartica broj:\s*\S+\s+(.+)$')


def parse_zaba_kartica(path: Path) -> list[dict]:
    """ZABA Mastercard izvod KARTICE (S107d). Transakcijska linija:
    'Bref DD.MM.YYYY. OPIS [orig-valuta detalji] IZNOS_EUR'. Dvije sekcije
    kartica (Kokina + Sašina dodatna) — nositelj iz 'Kartica broj:' linije.
    Datum na izvodu = datum KUPOVINE = event_date u Review (D1)."""
    txs: list[dict] = []
    holder = ''
    with pdfplumber.open(path) as pdf:
        for pno, page in enumerate(pdf.pages, 1):
            for line in (page.extract_text() or '').split('\n'):
                m = RE_KARTICA.match(line.strip())
                if m:
                    holder = m.group(1).strip()
                    continue
                toks = line.split()
                if len(toks) < 4 or not RE_MC_REF.match(toks[0]):
                    continue
                md = RE_DATE.match(toks[1])
                if not md or not RE_AMOUNT.match(toks[-1]):
                    continue
                d = date(int(md.group(3)), int(md.group(2)), int(md.group(1)))
                iznos = parse_amount(toks[-1])
                opis = ' '.join(toks[2:-1])
                # devizna linija: '..., TEČAJ ZABA dd.mm. 1 19,03 USD' rep se reže
                opis = re.sub(r',?\s*TEČAJ ZABA.*$', '', opis).strip()
                if 'SAŠA' in holder.upper():
                    opis += ' [kartica: SAŠA]'
                txs.append({
                    'date': d, 'opis': opis, 'iznos': abs(iznos),
                    'smjer': 'Uplata' if iznos < 0 else 'Isplata',
                    'kartica': holder, 'src': f'{path.name}:p{pno}',
                })
    return txs


# ── Parser: PBZ Visa Gold (PBZ Card mjesečni račun, SPECIFIKACIJA PROMETA) ─────

RE_PBZ_DATE   = re.compile(r'^(\d{2})\.(\d{2})\.(\d{2})\.$')
RE_PBZ_REF    = re.compile(r'^\d{9,11}$')
RE_PBZ_AMOUNT = re.compile(r'^[+-]?\d{1,3}(?:\.\d{3})*,\d{2}$')
RE_PBZ_UKUPNO = re.compile(r'^Ukupno troškova VISA\s+(.+?)\s+[\d.,]+\s+EUR$')


def parse_pbz_visa(path: Path) -> list[dict]:
    """PBZ Card Visa Gold mjesečni račun (S107d). Spec linija:
    'DD.MM.YY. REF OPIS IZNOS' (+prefiks = uplata/odobrenje). Nositelj kartice
    se doznaje TEK iz 'Ukupno troškova VISA <ime>' linije NAKON bloka →
    transakcije se skupljaju u bucket pa označe retroaktivno. Datum = datum
    kupovine (RATA retci nose datum IZVORNE kupovine — mogu biti mjeseci stari)."""
    txs: list[dict] = []
    bucket: list[dict] = []

    def flush(holder: str) -> None:
        for t in bucket:
            t['kartica'] = holder
            if 'SAŠA' in holder.upper():
                t['opis'] += ' [kartica: SAŠA]'
        txs.extend(bucket)
        bucket.clear()

    with pdfplumber.open(path) as pdf:
        for pno, page in enumerate(pdf.pages, 1):
            for line in (page.extract_text() or '').split('\n'):
                m = RE_PBZ_UKUPNO.match(line.strip())
                if m:
                    flush(m.group(1).strip())
                    continue
                toks = line.split()
                if len(toks) < 3 or not RE_PBZ_DATE.match(toks[0]):
                    continue
                if not RE_PBZ_AMOUNT.match(toks[-1]):
                    continue
                md = RE_PBZ_DATE.match(toks[0])
                d = date(2000 + int(md.group(3)), int(md.group(2)), int(md.group(1)))
                raw = toks[-1]
                mid = toks[1:-1]
                if mid and RE_PBZ_REF.match(mid[0]):
                    mid = mid[1:]
                bucket.append({
                    'date': d, 'opis': ' '.join(mid),
                    'iznos': abs(parse_amount(raw.lstrip('+'))),
                    'smjer': 'Uplata' if raw.startswith('+') else 'Isplata',
                    'kartica': '', 'src': f'{path.name}:p{pno}',
                })
    flush('')  # rep nakon zadnjeg 'Ukupno' (npr. 'PRIMLJENA UPLATA - HVALA')
    return txs


# Registry tipova izvora — koristi ga i inventory_izvoda.py (klasifikacija +
# rename prefix). Koka SVE kartične retke u Excelu vodi pod Izvor='Mastercard'
# (nema 'Visa' za Koku) → i MC i PBZVISA matchaju na 'Mastercard'.
SOURCE_TYPES: dict[str, tuple] = {
    # prefix: (parser, racun za match, Izvor za match)
    'ZABA':    (parse_zaba_racun,   RACUN_KOKA, 'Racun'),
    'MC':      (parse_zaba_kartica, RACUN_KOKA, 'Mastercard'),
    'PBZVISA': (parse_pbz_visa,     RACUN_KOKA, 'Mastercard'),
    'RF':      (parse_rf_ocr,       RACUN_SASA, 'Racun'),
}
# prefix match: duži prefiksi prvi (PBZVISA prije P*)
PARSERS: list[tuple[str, callable, str, str]] = [
    (pref, fn, rac, izv) for pref, (fn, rac, izv)
    in sorted(SOURCE_TYPES.items(), key=lambda kv: -len(kv[0]))
]


# ── Review match & write ───────────────────────────────────────────────────────

def pick_file(args: list[str]) -> Path:
    explicit = [a for a in args if not a.startswith('--')]
    if explicit:
        p = Path(explicit[0])
        if not p.exists():
            sys.exit(f'✗ File ne postoji: {p}')
        return p
    candidates = sorted(DATA_DIR.glob('Financije_review_*.xlsx'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    candidates = [c for c in candidates if '.pre-' not in c.name]   # svi backup fileovi
    if not candidates:
        sys.exit(f'✗ Nema Financije_review_*.xlsx u {DATA_DIR}')
    return candidates[0]


def find_smjer_col(ws) -> int:
    """'Smjer' header s auto-repairom: 2026-07-13 nađen pregažen slučajnim
    pasteom ('run.bat sync_taxonomy.py'). Ako headera nema, kolona se prepozna
    po podacima (samo Uplata/Isplata/PROVJERI) i header se vrati uz warning."""
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() == 'Smjer':
            return c
    for c in range(1, ws.max_column + 1):
        vals = {str(ws.cell(r, c).value or '') for r in range(2, min(ws.max_row, 200) + 1)}
        vals.discard('')
        if vals and vals <= {'Uplata', 'Isplata', 'PROVJERI'}:
            old = ws.cell(1, c).value
            ws.cell(1, c, 'Smjer')
            print(f'  ⚠ Header "Smjer" bio pregažen ({old!r}) — popravljen (kolona {c}).')
            return c
    sys.exit('✗ Kolona "Smjer" nije nađena u Review sheetu (ni po podacima).')


def find_header_col(ws, header: str, create: bool = False) -> int:
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() == header:
            return c
    if not create:
        sys.exit(f'✗ Kolona "{header}" nije nađena u Review sheetu.')
    c = ws.max_column + 1
    cell = ws.cell(1, c, header)
    cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 34 if 'opis' in header else 22
    return c


def to_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def load_tx_from_excel() -> list[tuple[dict, str, str]]:
    """Transakcije iz Izvodi_transakcije.xlsx (output inventory_izvoda.py)
    umjesto ponovnog parsiranja PDF-ova."""
    wb = openpyxl.load_workbook(TX_XLSX, read_only=True)
    ws = wb['Transakcije']
    hdr = {str(c.value): i for i, c in enumerate(ws[1]) if c.value}
    all_tx: list[tuple[dict, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = to_date(row[hdr['Datum']])
        if d is None:
            continue
        tx = {
            'date': d, 'opis': str(row[hdr['Opis']] or ''),
            'iznos': round(float(row[hdr['Iznos']]), 2),
            'smjer': str(row[hdr['Smjer']] or ''),
            'kartica': str(row[hdr['Kartica']] or ''),
            'src': str(row[hdr['Src']] or ''),
        }
        all_tx.append((tx, str(row[hdr['Racun']] or ''), str(row[hdr['Izvor']] or '')))
    wb.close()
    return all_tx


def write_unmatched_sheet(unmatched: list[dict]) -> None:
    """Nematchano sheet u Izvodi_transakcije.xlsx — transakcije s izvoda kojih
    NEMA u Review = potencijalno retci koji fale u Kokinom Excelu."""
    wb = openpyxl.load_workbook(TX_XLSX)
    if 'Nematchano' in wb.sheetnames:
        del wb['Nematchano']
    ws = wb.create_sheet('Nematchano')
    headers = ('Datum', 'Smjer', 'Iznos', 'Opis', 'Kartica', 'Src')
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
    for w, col in zip((12, 9, 11, 60, 26, 30), 'ABCDEF'):
        ws.column_dimensions[col].width = w
    for r, t in enumerate(sorted(unmatched, key=lambda t: t['date']), 2):
        ws.cell(r, 1, t['date']).number_format = 'DD.MM.YYYY'
        ws.cell(r, 2, t['smjer'])
        ws.cell(r, 3, t['iznos']).number_format = '#,##0.00'
        ws.cell(r, 4, t['opis'])
        ws.cell(r, 5, t.get('kartica', ''))
        ws.cell(r, 6, t['src'])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{max(2, len(unmatched) + 1)}'
    try:
        wb.save(TX_XLSX)
        print(f'✔ Nematchano sheet ({len(unmatched)}) upisan u {TX_XLSX.name}')
    except PermissionError:
        print(f'  ⚠ {TX_XLSX.name} otvoren u Excelu — Nematchano sheet NIJE upisan.')


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    path = pick_file(args)
    print(f'Review: {path.name}{"  [DRY RUN]" if dry else ""}')

    # 1. Transakcije: Izvodi_transakcije.xlsx (inventory_izvoda.py) ako postoji,
    #    inače legacy mode — parsiraj PDF-ove iz izvodi/ po filename prefixu
    all_tx: list[tuple[dict, str, str]] = []   # (tx, racun, izvor)
    if TX_XLSX.exists():
        all_tx = load_tx_from_excel()
        print(f'Izvor transakcija: {TX_XLSX.name} ({len(all_tx)} transakcija)')
    else:
        pdfs = sorted(IZVODI_DIR.glob('*.pdf'))
        if not pdfs:
            sys.exit(f'✗ Nema PDF-ova u {IZVODI_DIR} (ni {TX_XLSX.name})')
        for pdf in pdfs:
            parser = next(((fn, rac, izv) for pref, fn, rac, izv in PARSERS
                           if pdf.name.upper().startswith(pref)), None)
            if parser is None:
                print(f'  ⚠ {pdf.name}: nepoznat prefix — preskočen (pokreni inventory_izvoda.py)')
                continue
            fn, racun, izvor = parser
            txs = fn(pdf)
            if txs:
                print(f'  ✔ {pdf.name}: {len(txs)} transakcija')
            all_tx.extend((t, racun, izvor) for t in txs)
    if not all_tx:
        sys.exit('✗ Nijedna transakcija parsirana — nema se što matchati.')

    # 2. Učitaj Review i indeksiraj kandidate: (racun, izvor, smjer, datum, iznos) → [row]
    wb = openpyxl.load_workbook(path)
    ws = wb['Review']
    col = {h: find_header_col(ws, h) for h in
           ('Racun', 'event_date', 'Izvor', 'Uplata', 'Isplata', 'Napomena')}
    col['Smjer'] = find_smjer_col(ws)
    col_iopis = find_header_col(ws, 'Izvod opis', create=True)
    col_ifile = find_header_col(ws, 'Izvod file', create=True)

    index: dict[tuple, list[int]] = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        d = to_date(ws.cell(r, col['event_date']).value)
        if d is None:
            continue
        racun = str(ws.cell(r, col['Racun']).value or '')
        izvor = str(ws.cell(r, col['Izvor']).value or '')
        smjer = str(ws.cell(r, col['Smjer']).value or '')
        iznos = ws.cell(r, col['Uplata' if smjer == 'Uplata' else 'Isplata']).value
        if iznos is None:
            continue
        index[(racun, izvor, smjer, d, round(float(iznos), 2))].append(r)

    used: set[int] = set()
    matched, unmatched = 0, []
    for tx, racun, izvor in all_tx:
        row = None
        for delta in (0, 1, -1, 2, -2):
            key = (racun, izvor, tx['smjer'], tx['date'] + timedelta(days=delta), tx['iznos'])
            row = next((r for r in index.get(key, []) if r not in used), None)
            if row:
                break
        if row is None:
            unmatched.append(tx)
            continue
        used.add(row)
        matched += 1
        if not dry:
            ws.cell(row, col_iopis, tx['opis'][:250])
            ws.cell(row, col_ifile, tx['src'])

    # 3. Report + save
    print(f'\nMatchano: {matched}/{len(all_tx)} transakcija → kolone "Izvod opis"/"Izvod file"')
    if unmatched:
        print(f'Nematchano ({len(unmatched)}) — lump/kartični retci koji nisu u Review,'
              f' ILI retci koji FALE u Kokinom Excelu:')
        for t in unmatched[:15]:
            print(f'  {t["date"]} {t["smjer"]:<7} {t["iznos"]:>10.2f}  {t["opis"][:60]}')
        if len(unmatched) > 15:
            print(f'  ... i još {len(unmatched) - 15}')

    if not dry and unmatched and TX_XLSX.exists():
        write_unmatched_sheet(unmatched)
    if dry or not matched:
        return
    backup = path.with_name(f'{path.stem}.pre-izvod-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori file u Excelu i ponovi. (Backup: {backup.name})')
    print(f'✔ Snimljeno. Backup: {backup.name}')
    print('  Sljedeći korak: apply_rules.py (pravila vide i "Izvod opis" kolonu).')


if __name__ == '__main__':
    main()
