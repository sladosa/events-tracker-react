# -*- coding: utf-8 -*-
"""
fix_sportski_rekviziti_split.py  (S107g, 2026-07-16, jednokratna korekcija)
============================================================================
Preimenovanja sheet je blanket-preimenovao SVE stare 'Zdravlje/Sportski rekviziti'
retke (29) u 'Razno/Odjeća/obuća/potrebstine _Sasa' (1:1 par-mapping, ne gleda sadržaj).
Ta stara kategorija je bila mješavina: Multisport pretplata (24), Kreatin/MyProtein (3),
Decathlon (3). Saša je odlučio (2026-07-16):
  - "multisport" u Napomeni (24 retka)          → Zdravlje / Sport_Sasa
  - Napomena == "Kreatin" (3 retka, MyProtein)   → Namirnice / Hrana i ostalo
  - Decathlon (3 retka)                          → NE DIRAJ (ostaje Razno/Odjeća...)

Prepoznaje redove preko Podtip_O == 'Sportski rekviziti' (snapshot kolona iz
apply_rules.py) — nezavisno o trenutnom Tip/Podtip, radi i ako se ponovno pokrene.

Pokretanje (file zatvoren u Excelu!):
  Financije\\run.bat fix_sportski_rekviziti_split.py            → najnoviji review file
  Financije\\run.bat fix_sportski_rekviziti_split.py --dry      → samo pokaži
"""

import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")


def fold(s) -> str:
    return str(s or '').strip().lower()


def find_header_col(ws, header: str) -> int:
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() == header:
            return c
    sys.exit(f'✗ Kolona "{header}" nije nađena u Review sheetu.')


def pick_file(args: list[str]) -> Path:
    explicit = [a for a in args if not a.startswith('--')]
    if explicit:
        return Path(explicit[0])
    candidates = sorted(DATA_DIR.glob('Financije_review_*.xlsx'),
                         key=lambda p: p.stat().st_mtime, reverse=True)
    candidates = [c for c in candidates if '.pre-' not in c.name]
    if not candidates:
        sys.exit(f'✗ Nema Financije_review_*.xlsx u {DATA_DIR}')
    return candidates[0]


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    path = pick_file(args)
    print(f'File: {path.name}{"  [DRY RUN]" if dry else ""}')

    wb = openpyxl.load_workbook(path)
    ws = wb['Review']
    col_nap  = find_header_col(ws, 'Napomena')
    col_tip  = find_header_col(ws, 'Tip')
    col_pod  = find_header_col(ws, 'Podtip')
    col_alt  = find_header_col(ws, 'Alternativa / nap.')
    col_pod_o = find_header_col(ws, 'Podtip_O')

    n_multi = n_kreatin = n_decathlon = 0
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, col_pod_o).value or '').strip() != 'Sportski rekviziti':
            continue
        nap = fold(ws.cell(r, col_nap).value)
        if 'multisport' in nap:
            n_multi += 1
            if not dry:
                ws.cell(r, col_tip, 'Zdravlje')
                ws.cell(r, col_pod, 'Sport_Sasa')
                old_alt = str(ws.cell(r, col_alt).value or '').strip()
                mark = 'RUČNO S107g: multisport split (bio Razno/Odjeća)'
                ws.cell(r, col_alt, f'{old_alt} | {mark}' if old_alt else mark)
        elif nap == 'kreatin':
            n_kreatin += 1
            if not dry:
                ws.cell(r, col_tip, 'Namirnice')
                ws.cell(r, col_pod, 'Hrana i ostalo')
                old_alt = str(ws.cell(r, col_alt).value or '').strip()
                mark = 'RUČNO S107g: Kreatin split (bio Razno/Odjeća)'
                ws.cell(r, col_alt, f'{old_alt} | {mark}' if old_alt else mark)
        else:
            n_decathlon += 1   # Decathlon i ostalo — namjerno NE diramo

    print(f'{"Bi se ispravilo" if dry else "Ispravljeno"}: {n_multi} multisport → Zdravlje/Sport_Sasa, '
          f'{n_kreatin} Kreatin → Namirnice/Hrana i ostalo')
    print(f'Netaknuto (Decathlon i ostalo): {n_decathlon}')

    if dry or not (n_multi or n_kreatin):
        return

    backup = path.with_name(f'{path.stem}.pre-sportfix-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori file u Excelu i ponovi. (Backup: {backup.name})')
    print(f'✔ Snimljeno. Backup: {backup.name}')


if __name__ == '__main__':
    main()
