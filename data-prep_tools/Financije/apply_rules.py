# -*- coding: utf-8 -*-
"""
apply_rules.py  (S107c, 2026-07-12; dorade S107e, 2026-07-14)
=============================================================
Primjenjuje EDITABILNA keyword pravila (`Pravila` sheet) na neklasificirane redove
Financije REVIEW Excela — cilj: minimalan ručni rad u Tip/Podtip klasifikaciji.

Kako radi (svaki pravi run, ovim redom):
  1. Prvi put: kreira `Pravila` sheet (header + primjeri + upute) i stane —
     Saša/Koka upišu pravila pa pokrenu ponovno.
  2. Prvi put kad postoje nevaljani Tip/Podtip parovi: kreira `Preimenovanja`
     sheet (S107f) pred-popunjen svim starim parovima + auto-prijedlozima novih
     imena iz Taksonomije, pa stane — Saša pregleda/popuni i pokrene ponovno.
  3. SNAPSHOT (jednom, S107e): ako ne postoje, kreira `Tip_O`/`Podtip_O` kolone
     na kraju Review sheeta = kopija Tip/Podtip PRIJE ikakvog pisanja pravila
     (trajni trag originala; nikad se više ne ažuriraju).
  4. PREIMENOVANJA (S107f): red s nevaljanim parom koji ima mapping u
     `Preimenovanja` sheetu → dobije Novi Tip/Podtip UMJESTO reseta —
     Pouzdanost OSTAJE (VISOKA se čuva!), 'PREIM: bio <tip>/<pod>' se dodaje
     u Alternativa. 'Racun uvjet' kolona = per-osoba split (kokin/sasin).
  5. VALIDACIJA TAKSONOMIJE (S107e): preostali red čiji Tip/Podtip par NE
     postoji u Taksonomiji → reset na N/A + Pouzdanost='NEMA' +
     'TAKS: bio <tip>/<pod>' u Alternativa (original ostaje u _O kolonama).
     VISOKA klasifikacije s valjanim parovima se NE diraju. Prazan Podtip valjan.
  6. PRAVILA: čita pravila (odozgo prema dolje, PRVI match pobjeđuje)
     i primjenjuje ih SAMO na redove gdje je Tip prazan ili 'N/A'
     (uklj. svježe resetirane iz koraka 5) → ručni rad se NIKAD ne gazi.

Sintaksa ključnih riječi (kolona A `Pravila` sheeta):
  konzum                  → tekst sadrži "konzum"
  telekom & racun         → tekst sadrži OBJE riječi (bilo gdje, bilo koji red)
  (case-insensitive, dijakritike izjednačene: č/ć→c itd.; OR = više redova pravila)

Tekst po kojem se traži = Napomena kolona + 'Izvod opis*' kolone
(enrichment kolone iz enrich_from_izvoda.py; namjerno NE 'Izvod reda'/'Izvod file').

Napomena output (S107e): pravilo može imati i `Napomena` kolonu — čista ljudska
labela (npr. "Konzum"). Upisuje se SAMO ako je Napomena reda prazna (P3 princip).

Označavanje: pogođeni red dobije Pouzdanost='PRAVILO' i 'pravilo #N: <ključne riječi>'
u koloni Alternativa / nap. — filtriraj Pouzdanost=PRAVILO za brzu kontrolu.

Kolona `Pravilo run` (S107g): timestamp (YYYY-MM-DD HH:MM) upisan na SVAKI red koji
je OVAJ run promijenio (rename, TAKS reset ili pravilo) — filtriraj po zadnjem
timestampu da vidiš točno što je zadnji run dirao, neovisno od starijih runova.
Kreira se jednom (kao Tip_O/Podtip_O), vrijednost se prepisuje na svakom runu.

Validacija pravila: Tip/Podtip pravila moraju postojati u `Taksonomija` sheetu,
inače se pravilo preskače uz upozorenje (da se u Review ne upiše nevaljan par).

Pokretanje (file zatvoren u Excelu!):
  Financije\\run.bat apply_rules.py            → najnoviji Financije_review_*.xlsx
  Financije\\run.bat apply_rules.py --dry      → samo pokaži što bi se promijenilo
  Financije\\run.bat apply_rules.py --all      → + REPORT konflikata na već
                                                 klasificiranim redovima (ne piše!)
  ... apply_rules.py <putanja.xlsx> [--dry] [--all]  → eksplicitni file
"""

import re
import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")

HDR_FILL   = PatternFill('solid', fgColor='4472C4')
WHITE_BOLD = Font(color='FFFFFF', bold=True)
THIN       = Side(style='thin')
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DIACRITICS = {'ć': 'c', 'č': 'c', 'š': 's', 'ž': 'z', 'đ': 'd'}

SEED_RULES = [  # primjeri — slobodno obriši/zamijeni (kolone: ključne riječi, Tip, Podtip, Napomena, Komentar)
    ('hrvatski telekom', 'Informatika', 'T-com',           'T-com',   'primjer: opis s izvoda'),
    ('gpz',              'Domaćinstvo', 'Plin',            'Plin',    'primjer'),
    ('holding',          'Domaćinstvo', 'Holding (smeće)', 'Holding', 'primjer'),
    ('mirovinsk',        'Mirovina',    'Koka',            '',        'PRIMJER — OPREZ: mirovina može biti i Sašina (odvoji po Racun)!'),
]

HELP_TEXT = (
    'PRAVILA KLASIFIKACIJE — jedan red = jedno pravilo; odozgo prema dolje, PRVI match pobjeđuje.\n'
    'Ključne riječi: "konzum" = tekst sadrži konzum; "telekom & racun" = sadrži OBJE riječi.\n'
    'Case i dijakritike se ignoriraju (Č=č=c). OR se piše kao dva odvojena reda pravila.\n'
    'Pravilo se primjenjuje SAMO na redove gdje je Tip prazan ili N/A — ručni rad se ne dira.\n'
    'Napomena kolona (opcionalno): čista labela, upisuje se SAMO u redove s praznom Napomenom.\n'
    'Tekst za pretragu = Napomena + "Izvod opis" kolone (nakon enrich_from_izvoda.py).\n'
    'Tip i Podtip moraju postojati u Taksonomija sheetu (inače se pravilo preskače uz upozorenje).\n'
    'Prije primjene: skripta jednom snima Tip_O/Podtip_O original kolone + resetira na N/A\n'
    'redove čiji Tip/Podtip par više ne postoji u Taksonomiji (oznaka "TAKS:" u Alternativa).\n'
    'Nakon izmjena pokreni: Financije\\run.bat apply_rules.py  (--dry za probu bez snimanja;\n'
    '--all = i report konflikata pravila s već klasificiranim redovima, bez pisanja)'
)


def fold(s: str) -> str:
    s = str(s or '').lower()
    for a, b in DIACRITICS.items():
        s = s.replace(a, b)
    return s


def pick_file(args: list[str]) -> Path:
    explicit = [a for a in args if not a.startswith('--')]
    if explicit:
        p = Path(explicit[0])
        if not p.exists():
            sys.exit(f'✗ File ne postoji: {p}')
        return p
    candidates = sorted(DATA_DIR.glob('Financije_review_*.xlsx'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    candidates = [c for c in candidates if '.pre-' not in c.name]   # svi backup fileovi
    if not candidates:
        sys.exit(f'✗ Nema Financije_review_*.xlsx u {DATA_DIR}')
    return candidates[0]


def create_pravila_sheet(wb, path: Path) -> None:
    ws = wb.create_sheet('Pravila', 2)   # odmah iza Taksonomije
    headers = [('Ključne riječi', 32), ('Tip', 16), ('Podtip', 24), ('Napomena', 18), ('Komentar', 40)]
    for c, (h, w) in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
        ws.column_dimensions[chr(64 + c)].width = w
    for r, rule in enumerate(SEED_RULES, 2):
        for c, v in enumerate(rule, 1):
            ws.cell(r, c, v).border = BORDER
    note = ws.cell(2, 7, HELP_TEXT)
    note.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['G'].width = 95
    ws.row_dimensions[2].height = 150
    ws.freeze_panes = 'A2'
    wb.save(path)
    print(f'✔ Kreiran "Pravila" sheet u {path.name} (s {len(SEED_RULES)} primjera).')
    print('  Upiši/uredi pravila pa pokreni skriptu ponovno da se primijene.')


def read_taxonomy(wb) -> dict[str, set[str]]:
    tws = wb['Taksonomija']
    tax: dict[str, set[str]] = {}
    for r in range(2, tws.max_row + 1):
        tip = str(tws.cell(r, 1).value or '').strip()
        pod = str(tws.cell(r, 2).value or '').strip()
        if pod == '—':
            pod = ''
        if not tip:
            continue
        tax.setdefault(tip, set())
        if pod:
            tax[tip].add(pod)
    return tax


def read_rules(wb, tax: dict[str, set[str]]) -> list[dict]:
    ws = wb['Pravila']
    hdr = {str(ws.cell(1, c).value or '').strip().lower(): c
           for c in range(1, ws.max_column + 1)}
    c_kw  = hdr.get('ključne riječi', 1)
    c_tip = hdr.get('tip', 2)
    c_pod = hdr.get('podtip', 3)
    c_nap = hdr.get('napomena')          # None u starom (4-kolonskom) Pravila sheetu
    rules, skipped = [], 0
    for r in range(2, ws.max_row + 1):
        kw  = str(ws.cell(r, c_kw).value or '').strip()
        tip = str(ws.cell(r, c_tip).value or '').strip()
        pod = str(ws.cell(r, c_pod).value or '').strip()
        nap = str(ws.cell(r, c_nap).value or '').strip() if c_nap else ''
        if not kw and not tip:
            continue
        if not kw or not tip:
            print(f'⚠ Pravila red {r}: treba i ključne riječi i Tip — preskočen'); skipped += 1
            continue
        if tip not in tax:
            print(f'⚠ Pravila red {r}: Tip "{tip}" ne postoji u Taksonomiji — preskočen'); skipped += 1
            continue
        if pod and pod not in tax[tip]:
            print(f'⚠ Pravila red {r}: Podtip "{pod}" ne postoji pod Tipom "{tip}" — preskočen'); skipped += 1
            continue
        terms = [fold(t.strip()) for t in kw.split('&') if t.strip()]
        rules.append({'row': r, 'kw': kw, 'terms': terms, 'tip': tip, 'pod': pod, 'nap': nap})
    if skipped:
        print(f'  ({skipped} pravila preskočeno — ispravi pa ponovi)')
    return rules


def find_header_col(ws, header: str) -> int:
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() == header:
            return c
    sys.exit(f'✗ Kolona "{header}" nije nađena u Review sheetu.')


# ── Preimenovanja (S107f): stari Tip/Podtip par → novi, umjesto reseta ─────────

PREIM_HELP = (
    'PREIMENOVANJA — spašava redove čiji stari Tip/Podtip par više ne postoji u Taksonomiji.\n'
    'Umjesto reseta na N/A red dobije Novi Tip/Podtip; Pouzdanost se NE dira (VISOKA ostaje).\n'
    'Racun uvjet (opcionalno): red se primijeni samo ako Racun sadrži taj tekst — per-osoba\n'
    'split, npr. "kokin" → Kokin tekući ZABA, "sasin" → Sašin tekući RF.\n'
    'Redovi se čitaju odozgo; PRVI koji odgovara (stari par + uvjet) pobjeđuje.\n'
    'Prazan Novi Tip = par se NE preimenuje → ide na reset na N/A (kao dosad).\n'
    'Novi par mora postojati u Taksonomiji (inače se red preskače uz upozorenje).\n'
    'Auto-prijedlozi su popunjeni gdje je kandidat bio očit — SVAKI RED PROVJERI prije runa!\n'
    'Nakon popune: Financije\\run.bat apply_rules.py --dry (proba) pa bez --dry.'
)


def collect_invalid_pairs(ws, col_tip: int, col_pod: int, tax: dict[str, set[str]]) -> Counter:
    """Counter[(tip, pod)] za sve Review redove čiji par ne postoji u Taksonomiji."""
    pairs: Counter = Counter()
    for r in range(2, ws.max_row + 1):
        tip = str(ws.cell(r, col_tip).value or '').strip()
        if tip in ('', 'N/A'):
            continue
        pod = str(ws.cell(r, col_pod).value or '').strip()
        if pod == '—':
            pod = ''
        if tip in tax and (not pod or pod in tax[tip]):
            continue
        pairs[(tip, pod)] += 1
    return pairs


def find_candidates(old_tip: str, old_pod: str, tax: dict[str, set[str]]) -> list[tuple[str, str]]:
    """Kandidati u Taksonomiji čije ime SADRŽI stari podtip (ili tip, ako podtipa nema)."""
    key = fold(old_pod or old_tip)
    out: list[tuple[str, str]] = []
    for t in tax:
        if old_pod:
            out += [(t, p) for p in sorted(tax[t]) if key in fold(p)]
        elif key in fold(t):
            out.append((t, ''))
    return out


def create_preimenovanja_sheet(wb, path: Path, pairs: Counter, tax: dict[str, set[str]]) -> None:
    """Pred-popunjen sheet: svaki nevaljani par + auto-prijedlog gdje je očit.
    2 kandidata koji se razlikuju po koka/sasa → dva reda s Racun uvjetom."""
    ws = wb.create_sheet('Preimenovanja', wb.sheetnames.index('Pravila') + 1)
    headers = [('Stari Tip', 18), ('Stari Podtip', 26), ('Racun uvjet', 12),
               ('Novi Tip', 18), ('Novi Podtip', 34), ('Redova', 8), ('Komentar', 60)]
    for c, (h, w) in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
        ws.column_dimensions[chr(64 + c)].width = w

    r = 2
    for (tip, pod), n in pairs.most_common():
        cands = find_candidates(tip, pod, tax)
        rows: list[tuple[str, str, str, str]] = []   # (uvjet, novi tip, novi pod, komentar)
        if len(cands) == 1:
            rows = [('', *cands[0], 'auto-prijedlog (jedini kandidat) — PROVJERI')]
        elif len(cands) == 2:
            by_who = {}
            for t, p in cands:
                s = fold(f'{t} {p}')
                who = 'kokin' if 'koka' in s else ('sasin' if 'sasa' in s else None)
                by_who[who] = (t, p)
            if None not in by_who and len(by_who) == 2:
                rows = [(who, *by_who[who], 'auto-prijedlog per-osoba — PROVJERI uvjet!')
                        for who in sorted(by_who)]
        if not rows:
            if cands:
                hint = 'kandidati: ' + ', '.join(f'{t}/{p or "—"}' for t, p in cands[:4])
            else:
                hint = 'nema kandidata u Taksonomiji — prazno = reset na N/A'
            rows = [('', '', '', hint)]
        for i, (uvjet, nt, np_, kom) in enumerate(rows):
            for c, v in enumerate((tip, pod, uvjet, nt, np_, n if i == 0 else None, kom), 1):
                ws.cell(r, c, v).border = BORDER
            r += 1

    note = ws.cell(2, 9, PREIM_HELP)
    note.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['I'].width = 95
    ws.row_dimensions[2].height = 145
    ws.freeze_panes = 'A2'
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori {path.name} u Excelu i ponovi.')
    print(f'✔ Kreiran "Preimenovanja" sheet u {path.name}: {len(pairs)} starih parova '
          f'({sum(pairs.values())} redova), auto-prijedlozi popunjeni gdje su bili očiti.')
    print('  Pregledaj/popuni Novi Tip/Podtip pa pokreni skriptu ponovno (--dry za probu).')


def read_renames(wb, tax: dict[str, set[str]]) -> list[dict]:
    ws = wb['Preimenovanja']
    hdr = {str(ws.cell(1, c).value or '').strip().lower(): c
           for c in range(1, ws.max_column + 1)}
    cols = (hdr.get('stari tip', 1), hdr.get('stari podtip', 2), hdr.get('racun uvjet', 3),
            hdr.get('novi tip', 4), hdr.get('novi podtip', 5))
    out, skipped = [], 0
    for r in range(2, ws.max_row + 1):
        st, sp, uv, nt, np_ = (str(ws.cell(r, c).value or '').strip() for c in cols)
        sp, np_ = ('' if v == '—' else v for v in (sp, np_))
        if not nt:
            continue                     # nepopunjen red — taj par ide na reset (namjerno)
        if not st:
            print(f'⚠ Preimenovanja red {r}: Novi Tip bez Starog Tipa — preskočen'); skipped += 1
            continue
        if nt not in tax or (np_ and np_ not in tax[nt]):
            print(f'⚠ Preimenovanja red {r}: novi par "{nt}/{np_ or "—"}" ne postoji u Taksonomiji — preskočen')
            skipped += 1
            continue
        out.append({'row': r, 'old': (fold(st), fold(sp)), 'uvjet': fold(uv),
                    'nt': nt, 'np': np_, 'label': f'{st}/{sp or "—"}'})
    if skipped:
        print(f'  ({skipped} preimenovanja preskočeno — ispravi pa ponovi)')
    return out


def expand_autofilter(ws) -> None:
    """Autofilter mora obuhvatiti sve kolone — sort inače ne nosi nove kolone s redom."""
    if ws.auto_filter.ref:
        first = ws.auto_filter.ref.split(':')[0]
        ws.auto_filter.ref = f'{first}:{get_column_letter(ws.max_column)}{ws.max_row}'


def ensure_snapshot(ws, col_tip: int, col_pod: int, dry: bool) -> bool:
    """Tip_O/Podtip_O kolone = kopija Tip/Podtip PRIJE prvog pisanja pravila.
    Kreira se jednom (postojanje kolona = marker); nikad se ne ažurira."""
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() == 'Tip_O':
            return False
    if dry:
        print('… (dry) Tip_O/Podtip_O snapshot kolone bi se kreirale pri pravom runu')
        return False
    c_tip_o, c_pod_o = ws.max_column + 1, ws.max_column + 2
    for c, h in ((c_tip_o, 'Tip_O'), (c_pod_o, 'Podtip_O')):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
    for r in range(2, ws.max_row + 1):
        v_tip, v_pod = ws.cell(r, col_tip).value, ws.cell(r, col_pod).value
        if v_tip is not None:
            ws.cell(r, c_tip_o, v_tip)
        if v_pod is not None:
            ws.cell(r, c_pod_o, v_pod)
    expand_autofilter(ws)
    print(f'✔ Snapshot: Tip_O/Podtip_O kolone kreirane (kopija Tip/Podtip prije pravila)')
    return True


def ensure_run_column(ws, dry: bool) -> tuple[int | None, bool]:
    """'Pravilo run' kolona — timestamp na svaki red koji OVAJ run promijeni.
    Kreira se jednom (kao Tip_O), vrijednosti se prepisuju na svakom runu.
    Vraća (col_index, created_now) — created_now forsira save i kad nema drugih promjena."""
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() == 'Pravilo run':
            return c, False
    if dry:
        print('… (dry) "Pravilo run" kolona bi se kreirala pri pravom runu')
        return None, False
    c_run = ws.max_column + 1
    cell = ws.cell(1, c_run, 'Pravilo run')
    cell.fill, cell.font, cell.border = HDR_FILL, WHITE_BOLD, BORDER
    ws.column_dimensions[get_column_letter(c_run)].width = 16
    expand_autofilter(ws)
    print('✔ Kreirana kolona "Pravilo run" (timestamp po redu za zadnji run)')
    return c_run, True


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    report_all = '--all' in args
    path = pick_file(args)
    print(f'File: {path.name}{"  [DRY RUN — bez snimanja]" if dry else ""}'
          f'{"  [--all: report konflikata]" if report_all else ""}')

    wb = openpyxl.load_workbook(path)
    if 'Review' not in wb.sheetnames or 'Taksonomija' not in wb.sheetnames:
        sys.exit('✗ File nema Review/Taksonomija sheet — je li ovo review Excel?')

    if 'Pravila' not in wb.sheetnames:
        create_pravila_sheet(wb, path)
        return

    tax = read_taxonomy(wb)

    ws = wb['Review']
    col_nap  = find_header_col(ws, 'Napomena')
    col_tip  = find_header_col(ws, 'Tip')
    col_pod  = find_header_col(ws, 'Podtip')
    col_conf = find_header_col(ws, 'Pouzdanost')
    col_alt  = find_header_col(ws, 'Alternativa / nap.')
    col_rac  = find_header_col(ws, 'Racun')

    # ── 0. PREIMENOVANJA sheet: prvi put pred-popuni stare parove i stani ─────
    invalid_pairs = collect_invalid_pairs(ws, col_tip, col_pod, tax)
    if invalid_pairs and 'Preimenovanja' not in wb.sheetnames:
        create_preimenovanja_sheet(wb, path, invalid_pairs, tax)
        return
    renames = read_renames(wb, tax) if 'Preimenovanja' in wb.sheetnames else []
    if renames:
        print(f'Preimenovanja: {len(renames)} valjanih mappinga')

    rules = read_rules(wb, tax)
    if not rules and not renames:
        sys.exit('✗ Nema valjanih pravila u "Pravila" sheetu (ni preimenovanja).')
    print(f'Pravila: {len(rules)} valjanih')
    # SAMO 'Izvod opis*' kolone — NE 'Izvod reda' (koka EU:...) ni 'Izvod file'
    # (ZABA_*.pdf) jer bi ključne riječi poput "zaba"/"koka" lažno matchale sve.
    izvod_cols = [c for c in range(1, ws.max_column + 1)
                  if str(ws.cell(1, c).value or '').startswith('Izvod opis')]

    # ── 1. SNAPSHOT (jednom): Tip_O/Podtip_O = original prije pravila ─────────
    snap_created = ensure_snapshot(ws, col_tip, col_pod, dry)
    col_run, run_col_created = ensure_run_column(ws, dry)
    run_stamp = datetime.now().strftime('%Y-%m-%d %H:%M')

    hits_per_rule: dict[int, int] = {}

    def match_text(r: int) -> str:
        text = fold(ws.cell(r, col_nap).value)
        for c in izvod_cols:
            text += ' | ' + fold(ws.cell(r, c).value)
        return text

    def find_rule(r: int) -> tuple[int, dict] | None:
        text = match_text(r)
        if not text.strip(' |'):
            return None
        for i, rule in enumerate(rules):
            if all(t in text for t in rule['terms']):
                return i, rule
        return None

    # ── 2. PREIMENOVANJA + VALIDACIJA: nevaljan par → PRAVILO (ako pogađa) →
    #        novi par (Preimenovanja) → reset na N/A. Pravilo nadvladava blanket
    #        rename kad par-mapping pogađa preširoko (S107g).
    reset_rows: set[int] = set()
    tax_samples: list[str] = []
    renamed = 0
    ren_hits: dict[int, int] = {}
    ren_samples: list[str] = []
    overridden = 0
    over_samples: list[str] = []
    for r in range(2, ws.max_row + 1):
        tip_now = str(ws.cell(r, col_tip).value or '').strip()
        if tip_now in ('', 'N/A'):
            continue
        pod_now = str(ws.cell(r, col_pod).value or '').strip()
        if pod_now == '—':
            pod_now = ''
        if tip_now in tax and (not pod_now or pod_now in tax[tip_now]):
            continue                          # valjan par (prazan Podtip = valjan)
        found = find_rule(r)
        if found:
            i, rule = found
            overridden += 1
            hits_per_rule[i] = hits_per_rule.get(i, 0) + 1
            if not dry:
                ws.cell(r, col_tip, rule['tip'])
                if rule['pod']:
                    ws.cell(r, col_pod, rule['pod'])
                else:
                    ws.cell(r, col_pod).value = None
                ws.cell(r, col_conf, 'PRAVILO')
                old_alt = str(ws.cell(r, col_alt).value or '').strip()
                mark = f'PRAVILO #{i + 1} nadvladao Preimenovanja: bio {tip_now}/{pod_now or "—"}'
                ws.cell(r, col_alt, f'{old_alt} | {mark}' if old_alt else mark)
                if rule['nap'] and not str(ws.cell(r, col_nap).value or '').strip():
                    ws.cell(r, col_nap, rule['nap'])
                if col_run:
                    ws.cell(r, col_run, run_stamp)
            if len(over_samples) < 8:
                over_samples.append(f'  red {r}: {tip_now}/{pod_now or "—"} → pravilo #{i + 1} '
                                     f'"{rule["kw"]}" → {rule["tip"]}/{rule["pod"] or "—"}')
            continue
        racun_f = fold(ws.cell(r, col_rac).value)
        m = next((m for m in renames
                  if m['old'] == (fold(tip_now), fold(pod_now))
                  and (not m['uvjet'] or m['uvjet'] in racun_f)), None)
        if m:
            renamed += 1
            ren_hits[m['row']] = ren_hits.get(m['row'], 0) + 1
            if not dry:
                ws.cell(r, col_tip, m['nt'])
                if m['np']:
                    ws.cell(r, col_pod, m['np'])
                else:
                    ws.cell(r, col_pod).value = None
                old_alt = str(ws.cell(r, col_alt).value or '').strip()
                mark = f'PREIM: bio {tip_now}/{pod_now or "—"}'
                ws.cell(r, col_alt, f'{old_alt} | {mark}' if old_alt else mark)
                if col_run:
                    ws.cell(r, col_run, run_stamp)
            if len(ren_samples) < 8:
                ren_samples.append(f'  red {r}: {tip_now}/{pod_now or "—"} → {m["nt"]}/{m["np"] or "—"}')
            continue
        reset_rows.add(r)
        if not dry:
            ws.cell(r, col_tip, 'N/A')
            ws.cell(r, col_pod).value = None   # cell(r,c,None) NE briše — mora preko .value
            ws.cell(r, col_conf, 'NEMA')
            ws.cell(r, col_alt, f'TAKS: bio {tip_now}/{pod_now or "—"}')
            if col_run:
                ws.cell(r, col_run, run_stamp)
        if len(tax_samples) < 8:
            tax_samples.append(f'  red {r}: {tip_now}/{pod_now or "—"} → N/A (nije u Taksonomiji)')
    if renamed:
        print(f'{"Bi se preimenovalo" if dry else "Preimenovano"} (Preimenovanja sheet): '
              f'{renamed} redova — Pouzdanost netaknuta, oznaka "PREIM:" u Alternativa')
        for m in renames:
            n = ren_hits.get(m['row'], 0)
            if n:
                print(f'  {m["label"]} → {m["nt"]}/{m["np"] or "—"}: {n}×')
        print('\n'.join(ren_samples))
        if renamed > len(ren_samples):
            print(f'  ... i još {renamed - len(ren_samples)}')
    if reset_rows:
        print(f'{"Bi se resetiralo" if dry else "Resetirano"} na N/A (Taksonomija validacija): '
              f'{len(reset_rows)} redova — original u Tip_O/Podtip_O, oznaka "TAKS:" u Alternativa')
        print('\n'.join(tax_samples))
        if len(reset_rows) > len(tax_samples):
            print(f'  ... i još {len(reset_rows) - len(tax_samples)}')
    if overridden:
        print(f'{"Bi pravilo nadvladalo" if dry else "Pravilo nadvladalo"} Preimenovanja (par bio '
              f'preširok): {overridden} redova')
        print('\n'.join(over_samples))
        if overridden > len(over_samples):
            print(f'  ... i još {overridden - len(over_samples)}')

    # ── 3. PRAVILA na Tip prazan/N/A (uklj. svježe resetirane) ────────────────
    changed = nap_filled = 0
    samples: list[str] = []
    conflicts: list[str] = []
    n_conflicts = 0
    for r in range(2, ws.max_row + 1):
        tip_now = str(ws.cell(r, col_tip).value or '').strip()
        unclassified = tip_now in ('', 'N/A') or r in reset_rows
        if not unclassified and not report_all:
            continue                          # ručno/pipeline klasificiran — NE diraj
        text = fold(ws.cell(r, col_nap).value)
        for c in izvod_cols:
            text += ' | ' + fold(ws.cell(r, c).value)
        if not text.strip(' |'):
            continue
        for i, rule in enumerate(rules):
            if all(t in text for t in rule['terms']):
                if not unclassified:
                    # --all: klasificiran red — samo REPORT ako se pravilo ne slaže
                    pod_now = str(ws.cell(r, col_pod).value or '').strip()
                    if tip_now != rule['tip'] or (rule['pod'] and pod_now != rule['pod']):
                        n_conflicts += 1
                        if len(conflicts) < 20:
                            conflicts.append(f'  red {r}: sada {tip_now}/{pod_now or "—"} '
                                             f'vs pravilo #{i + 1} "{rule["kw"]}" → {rule["tip"]}/{rule["pod"] or "—"}')
                    break
                if not dry:
                    ws.cell(r, col_tip, rule['tip'])
                    if rule['pod']:
                        ws.cell(r, col_pod, rule['pod'])
                    ws.cell(r, col_conf, 'PRAVILO')
                    ws.cell(r, col_alt, f'pravilo #{i + 1}: {rule["kw"]}')
                    if rule['nap'] and not str(ws.cell(r, col_nap).value or '').strip():
                        ws.cell(r, col_nap, rule['nap'])   # P3: prazno se puni, puno NE
                        nap_filled += 1
                    if col_run:
                        ws.cell(r, col_run, run_stamp)
                hits_per_rule[i] = hits_per_rule.get(i, 0) + 1
                changed += 1
                if len(samples) < 8:
                    samples.append(f'  red {r}: "{str(ws.cell(r, col_nap).value or "")[:40]}" → {rule["tip"]}/{rule["pod"] or "—"}')
                break

    print(f'\n{"Bi se promijenilo" if dry else "Promijenjeno"}: {changed} redova'
          + (f' (+ {nap_filled} Napomena popunjeno)' if nap_filled else ''))
    for i, rule in enumerate(rules):
        n = hits_per_rule.get(i, 0)
        if n:
            print(f'  #{i + 1} "{rule["kw"]}" → {rule["tip"]}/{rule["pod"] or "—"}: {n}×')
    if samples:
        print('Primjeri:')
        print('\n'.join(samples))
    if report_all:
        print(f'\n--all KONFLIKTI (klasificiran red vs pravilo — SAMO report, ništa se ne piše): {n_conflicts}')
        print('\n'.join(conflicts))
        if n_conflicts > len(conflicts):
            print(f'  ... i još {n_conflicts - len(conflicts)}')

    if dry or not (changed or reset_rows or renamed or overridden or snap_created or run_col_created):
        return

    backup = path.with_name(f'{path.stem}.pre-rules-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    try:
        wb.save(path)
    except PermissionError:
        sys.exit(f'✗ Ne mogu snimiti — zatvori file u Excelu i ponovi. (Backup: {backup.name})')
    print(f'✔ Snimljeno. Backup: {backup.name}')
    print('  Kontrola: filtriraj Pouzdanost = PRAVILO (pravila) i "TAKS:" u Alternativa (reseti);'
          f' ili "Pravilo run" = {run_stamp} za sve što je OVAJ run dirao.')


if __name__ == '__main__':
    main()
