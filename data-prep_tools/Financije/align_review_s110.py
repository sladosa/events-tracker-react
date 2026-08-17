# -*- coding: utf-8 -*-
"""
align_review_s110.py  (S110, 2026-08-17)  — ONE-OFF
================================================================================
Uskladi Review s onim što je RUČNO uneseno u aplikaciju u S110.

ZAŠTO UOPĆE: Review je izvor uvoza. Ako se ono što je ispravljeno u bazi ne
prenese ovamo, sljedeći uvoz (2024, 2023, Kokina delta) vratio bi staro stanje
— tiho, jer update-guard (D7) gleda `row_hash` i mirne duše prepiše ručnu
ispravku. Divergencija izvora i baze se ne javlja kao greška; ona se otkrije
mjesecima kasnije kao „pa ovo sam već popravio".

--------------------------------------------------------------------------------
DVIJE PROMJENE, OBJE PRESLIKA ONOGA ŠTO JE VEĆ U BAZI

1. xl4792 — podizanje 200,00 s bankomata (`koka EU:1780`), datum popravljen u
   `fix_koka_datum_200.py`. Klasifikacija je bila `N/A` jer redak nema opis.
   Sad kad je datum točan, izvod `ZABA_2025-05.pdf` pokazuje da je to
   „Podizanje gotovog novca - debitnom karticom na bankomatu" — isto kao
   njegov blizanac xl3325 (19.05.2025.), koji pravilo #23 klasificira kao
   `Transfer` / `cash - bankomat`.
   ⚠ `Transfer` nije kozmetika: ulazi u saldo (novac je stvarno otišao s
     računa), ali IZLAZI iz razreza po Tipu. Kao `N/A` bi tih 200 zauvijek
     visjelo kao neklasificiran TROŠAK, a podizanje gotovine nije potrošnja.

2. xl4994 — parking 1,60 (`koka EU:2564`), `Status = Planiran`. To je ostatak
   iz vremena kad je redak bio datiran `2026-08-07`, dakle u budućnosti; sad je
   `2026-07-07` i u bazi je unesen kao `Izvrsen`.
   ⚠ `Planiran` je ovdje aktivno štetan: filtar salda je `Status ≠ Planiran`,
     pa bi redak ispao iz izračuna i razlika od 1,60 bi se vratila.

ŠTO NE DIRA: `Stanje`, `source_key`, datume (v. `fix_koka_datum_200.py` za
obrazloženje — isti razlozi vrijede i ovdje).

Pokretanje:
    Financije\\run.bat align_review_s110.py --dry
    Financije\\run.bat align_review_s110.py
"""
from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DATA = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")
REVIEW = DATA / 'Financije_review_20260710_1448.xlsx'

# Svaki zahvat: uski potpis (da ne pogodi krivi redak) + ciljane promjene.
# `Pouzdanost` = 'VISOKA' jer je klasifikacija potvrđena protiv izvoda, ne
# pogođena pravilom — vrijednost postoji u rječniku (1012 redaka je koriste).
FIXES = [
    {
        'ime': '200,00 bankomat (EU:1780)',
        'potpis': {'Izvor reda': 'koka EU:1780', 'source_key': '268b5fbe9013',
                   'Isplata': 200, 'Tip': 'N/A'},
        'promjene': {
            'Tip': 'Transfer',
            'Podtip': 'cash - bankomat',
            'Napomena': 'podizanje s bankomata',
            'Pouzdanost': 'VISOKA',
            'Alternativa / nap.': 'S110: datum 2026→2025 (Kokina tipfelerica), '
                                  'klasificirano po izvodu ZABA_2025-05 kao blizanac EU:1772',
        },
    },
    {
        'ime': 'parking 1,60 (EU:2564)',
        'potpis': {'Izvor reda': 'koka EU:2564', 'source_key': '89631181b5c9',
                   'Isplata': 1.6, 'Status': 'Planiran'},
        'promjene': {
            'Status': 'Izvrsen',
            'Alternativa / nap.': 'D8 default — Lacetti? | PREIM: bio auto C5/parking | '
                                  'S107x fix: Koka potvrdila | S110: Planiran→Izvrsen '
                                  '(bio ostatak od datuma 2026-08-07), unesen u bazu kroz app',
        },
    },
]


def main() -> None:
    dry = '--dry' in sys.argv[1:]
    if not REVIEW.exists():
        sys.exit(f'✗ Nema {REVIEW}')

    wb = openpyxl.load_workbook(REVIEW)
    ws = wb['Review']
    hdr = {c.value: c.column for c in ws[1] if c.value}

    plan = []
    for fix in FIXES:
        for k in (*fix['potpis'], *fix['promjene']):
            if k not in hdr:
                sys.exit(f'✗ Nedostaje kolona "{k}".')
        hits = [r for r in range(2, ws.max_row + 1)
                if all(ws.cell(r, hdr[k]).value == v for k, v in fix['potpis'].items())]
        if len(hits) != 1:
            sys.exit(f"✗ [{fix['ime']}] očekivan 1 redak, nađeno {len(hits)}: {hits}\n"
                     f'  Review se promijenio ili je zahvat već primijenjen — provjeri ručno.')
        plan.append((hits[0], fix))

    for row, fix in plan:
        print(f"xl{row} — {fix['ime']}")
        for k, v in fix['promjene'].items():
            old = ws.cell(row, hdr[k]).value
            print(f'    {k:<20} {old!r} → {v!r}' if old != v else f'    {k:<20} (već {v!r})')

    if dry:
        print('\n--dry: ništa nije zapisano.')
        return

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup = REVIEW.with_name(f'{REVIEW.stem}.pre-align110-{stamp}.xlsx')
    shutil.copy2(REVIEW, backup)
    print(f'\nBackup: {backup.name}')

    for row, fix in plan:
        for k, v in fix['promjene'].items():
            ws.cell(row, hdr[k]).value = v
    wb.save(REVIEW)
    print(f'✓ Zapisano u {REVIEW.name} — {len(plan)} retka usklađena s bazom.')


if __name__ == '__main__':
    main()
