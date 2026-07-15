# -*- coding: utf-8 -*-
"""Backfill 'Datum naplate' = event_date za Izvor Racun/Cash (D1 pravilo).

Kod direktnih transakcija s računa i gotovine naplata pada na dan kupovine,
pa je 'Datum naplate' doslovna kopija event_date. Kartice se NE diraju:
Mastercard već ima vrijednosti iz Kokinog Excela, Visa se puni tek pri
generiranju import Excela (next:N ili stvarni datumi lump isplata).

Ne mijenja NIŠTA osim praznih 'Datum naplate' ćelija na Racun/Cash redovima.

Pokretanje (review file zatvoren u Excelu!):
  Financije\\run.bat backfill_datum_naplate.py            → najnoviji review
  ... backfill_datum_naplate.py --dry                     → bez snimanja, samo report
  ... backfill_datum_naplate.py <review.xlsx> [--dry]
"""

import shutil
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = Path(r"C:\0_Sasa\events-tracker-react\data-prep_data\Financije")

IZVORI_BACKFILL = {'Racun', 'Cash'}


def pick_file(args: list[str]) -> Path:
    explicit = [a for a in args if not a.startswith('--')]
    if explicit:
        p = Path(explicit[0])
        if not p.exists():
            sys.exit(f'✗ File ne postoji: {p}')
        return p
    candidates = sorted(DATA_DIR.glob('Financije_review_*.xlsx'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    candidates = [c for c in candidates if '.pre-' not in c.name]
    if not candidates:
        sys.exit(f'✗ Nema Financije_review_*.xlsx u {DATA_DIR}')
    return candidates[0]


def find_header_col(ws, header: str) -> int:
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() == header:
            return c
    sys.exit(f'✗ Kolona "{header}" nije nađena u Review sheetu.')


def main() -> None:
    args = sys.argv[1:]
    dry = '--dry' in args
    path = pick_file(args)
    print(f'Review: {path.name}{"  [DRY RUN]" if dry else ""}')

    wb = openpyxl.load_workbook(path)
    ws = wb['Review']
    col_izvor = find_header_col(ws, 'Izvor')
    col_edate = find_header_col(ws, 'event_date')
    col_dnapl = find_header_col(ws, 'Datum naplate')

    filled = Counter()          # po Izvoru — koliko je popunjeno
    skipped_has_value = 0       # Racun/Cash koji već imaju Datum naplate
    empty_other = Counter()     # prazni na ostalim Izvorima (samo report, ne dira se)
    no_event_date = 0

    for r in range(2, ws.max_row + 1):
        izvor = str(ws.cell(r, col_izvor).value or '').strip()
        if not izvor:
            continue
        dnapl = ws.cell(r, col_dnapl).value
        has_dnapl = dnapl is not None and str(dnapl).strip() != ''
        if izvor not in IZVORI_BACKFILL:
            if not has_dnapl:
                empty_other[izvor] += 1
            continue
        if has_dnapl:
            skipped_has_value += 1
            continue
        src = ws.cell(r, col_edate)
        if not isinstance(src.value, (datetime, date)):
            no_event_date += 1
            continue
        if not dry:
            dst = ws.cell(r, col_dnapl)
            dst.value = src.value
            dst.number_format = src.number_format
        filled[izvor] += 1

    total = sum(filled.values())
    print(f'\nPopunjeno Datum naplate = event_date: {total}')
    for izvor, n in sorted(filled.items()):
        print(f'  {izvor:<6} {n}')
    if skipped_has_value:
        print(f'Preskočeno (Racun/Cash već ima vrijednost): {skipped_has_value}')
    if no_event_date:
        print(f'⚠ Racun/Cash bez valjanog event_date — NIJE popunjeno: {no_event_date}')
    if empty_other:
        ostali = ', '.join(f'{k} {n}' for k, n in sorted(empty_other.items()))
        print(f'Prazan Datum naplate na ostalim Izvorima (namjerno ne diram): {ostali}')

    if dry or not total:
        return
    backup = path.with_name(f'{path.stem}.pre-naplata-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(path, backup)
    print(f'Backup: {backup.name}')
    try:
        wb.save(path)
        print(f'✔ Snimljeno: {path.name}')
    except PermissionError:
        sys.exit(f'✗ {path.name} je otvoren u Excelu — zatvori ga pa ponovi. '
                 f'(Backup je svejedno kreiran.)')


if __name__ == '__main__':
    main()
