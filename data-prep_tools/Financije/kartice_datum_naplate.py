# -*- coding: utf-8 -*-
"""
kartice_datum_naplate.py  (S107k, 2026-07-23; bivši visa_datum_naplate.py)
==========================================================================
Puni prazan `Datum naplate` na KARTIČNIM (Visa + Mastercard) retcima Reviewa.

VISA — verificirani obrazac (30/30 statementa u cent, 2026-07-23):
  suma kupovina statementa M == PRIMLJENA UPLATA u statementu M+1
  == PBZCARD isplata sa Sašinog RF računa ISTI DAN (≈4.–8. u mjesecu M+1)
Izvori datuma, po prioritetu:
  1. lump redci (Izvor reda 'PBZ Visa:lump')      → Datum naplate = event_date (transfer sam sebi)
  2. `Izvod file` = PBZVISA_YYYY-MM               → EGZAKTNO: datum uplate statementa ym
     (lump u statementu ym+1, fallback RF PBZCARD isplata u mjesecu ym+1)
  3. bez Izvod filea (stari Sašini ručni redci)   → CUTOFF PRAVILO: dan ≤ 3 → statement
     prethodnog mjeseca, inače tekućeg; datum = poznata uplata ili sintetski 5. u ym+1

MASTERCARD — Kokino pravilo (potvrđeno u podacima: 1650/1653 njenih redaka = 11. u mjesecu):
  naplata statementa M = 11. u mjesecu M+1. Prazni MC redci su oni dodani konsolidacijom
  (build_row kartici ne puni naplatu) — svi imaju `Izvod file` = MC_YYYY-MM → EGZAKTNO;
  bez filea → cutoff pravilo (dan ≤ 3 → prethodni statement) → 11. u ym+1.

P3: puni SAMO prazne ćelije — postojeće vrijednosti se nikad ne gaze.

Pokretanje (Review ZATVOREN u Excelu!):
  Tools\\venv\\Scripts\\python.exe Financije\\kartice_datum_naplate.py --dry   → samo brojke
  Tools\\venv\\Scripts\\python.exe Financije\\kartice_datum_naplate.py         → backup + upis
"""

import shutil
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, str(Path(__file__).parent))
from consolidate_review import DATA_DIR, TX_XLSX, hdr_index, pick_review, to_date  # noqa: E402

CUTOFF_DAY = 3          # kupovina dana ≤ 3 → statement prethodnog mjeseca (Koka: "do 5.")
SYNTH_DAY = 5           # Visa sintetski datum naplate: 5. u mjesecu nakon statementa
MC_DAY = 11             # Mastercard: naplata = 11. u mjesecu nakon statementa (Kokino pravilo)


def ym_add(ym: str, k: int) -> str:
    y, m = int(ym[:4]), int(ym[5:7])
    m += k
    y += (m - 1) // 12
    m = (m - 1) % 12 + 1
    return f'{y:04d}-{m:02d}'


def build_payment_map() -> dict[str, tuple[date, str]]:
    """statement ym → (datum naplate, izvor podatka). Lump 'PRIMLJENA UPLATA' u
    statementu M plaća statement M-1; RF 'PBZCARD' isplata u mjesecu M plaća M-1."""
    wb = openpyxl.load_workbook(TX_XLSX, read_only=True)
    ws = wb['Transakcije']
    h = {str(c.value): i for i, c in enumerate(ws[1]) if c.value}
    pay: dict[str, tuple[date, str]] = {}
    rf_pay: dict[str, tuple[date, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = to_date(row[h['Datum']])
        if d is None:
            continue
        src = str(row[h['Src']] or '')
        opis = str(row[h['Opis']] or '').upper()
        if src.startswith('PBZVISA') and 'PRIMLJENA UPLATA' in opis:
            stm_ym = src.split('.pdf')[0].replace('PBZVISA_', '')
            pay[ym_add(stm_ym, -1)] = (d, 'lump')
        elif src.startswith('RF') and 'PBZCARD' in opis.replace(' ', '') \
                and str(row[h['Smjer']] or '') == 'Isplata':
            rf_pay[ym_add(f'{d:%Y-%m}', -1)] = (d, 'RF isplata')
    wb.close()
    for ym, v in rf_pay.items():        # RF popunjava rupe (npr. zadnji statement)
        pay.setdefault(ym, v)
    return pay


def naplata_for(edate: date, izvod_file: str, pay: dict) -> tuple[date, str]:
    """→ (datum, kategorija reporta)."""
    if izvod_file.startswith('PBZVISA_'):
        ym = izvod_file.split('.pdf')[0].replace('PBZVISA_', '')
        if ym in pay:
            return (pay[ym][0], f'egzaktno ({pay[ym][1]})')
        nxt = ym_add(ym, 1)
        return (date(int(nxt[:4]), int(nxt[5:7]), SYNTH_DAY), 'statement bez poznate uplate → sintetski')
    ym = f'{edate:%Y-%m}' if edate.day > CUTOFF_DAY else ym_add(f'{edate:%Y-%m}', -1)
    if ym in pay:
        return (pay[ym][0], 'cutoff pravilo → poznata uplata')
    nxt = ym_add(ym, 1)
    return (date(int(nxt[:4]), int(nxt[5:7]), SYNTH_DAY), 'cutoff pravilo → sintetski')


def naplata_for_mc(edate: date, izvod_file: str) -> tuple[date, str]:
    """MC: naplata = 11. u mjesecu nakon statementa. → (datum, kategorija reporta)."""
    if izvod_file.startswith('MC_'):
        ym = izvod_file.split('.pdf')[0].replace('MC_', '')
        kind = 'MC egzaktno (statement iz Izvod file)'
    else:
        ym = f'{edate:%Y-%m}' if edate.day > CUTOFF_DAY else ym_add(f'{edate:%Y-%m}', -1)
        kind = 'MC cutoff pravilo'
    nxt = ym_add(ym, 1)
    return (date(int(nxt[:4]), int(nxt[5:7]), MC_DAY), kind)


def main():
    args = sys.argv[1:]
    dry = '--dry' in args
    review = pick_review(args)
    print(f'Review: {review.name}{"  [DRY — Review NETAKNUT]" if dry else ""}')

    pay = build_payment_map()
    print(f'Poznate uplate statementa: {len(pay)} '
          f'({min(pay)}..{max(pay)}, od toga RF-fallback: '
          f'{sum(1 for v in pay.values() if v[1] == "RF isplata")})')

    wb = openpyxl.load_workbook(review)
    ws = wb['Review']
    col = hdr_index(ws)

    stats = Counter()
    by_year = Counter()
    filled = 0
    for r in range(2, ws.max_row + 1):
        izvor = str(ws.cell(r, col['Izvor']).value or '')
        if izvor not in ('Visa', 'Mastercard'):
            continue
        dn_cell = ws.cell(r, col['Datum naplate'])
        if dn_cell.value not in (None, ''):
            stats[f'{izvor}: već popunjeno (ne diram)'] += 1
            continue
        ed_cell = ws.cell(r, col['event_date'])
        ed = to_date(ed_cell.value)
        if ed is None:
            stats[f'{izvor}: bez event_date (⚠)'] += 1
            continue
        if izvor == 'Mastercard':
            nd, kind = naplata_for_mc(ed, str(ws.cell(r, col['Izvod file']).value or ''))
        elif str(ws.cell(r, col['Izvor reda']).value or '') == 'PBZ Visa:lump':
            nd, kind = ed, 'lump = event_date'
        else:
            nd, kind = naplata_for(ed, str(ws.cell(r, col['Izvod file']).value or ''), pay)
        stats[kind] += 1
        by_year[ed.year] += 1
        filled += 1
        if not dry:
            dn_cell.value = datetime(nd.year, nd.month, nd.day)
            dn_cell.number_format = ed_cell.number_format
    print(f'\nKartičnih redaka za popuniti: {filled}')
    for k, n in stats.most_common():
        print(f'  {k}: {n}')
    print(f'Po godini (event_date): {dict(sorted(by_year.items()))}')

    if dry:
        print('\n✔ [DRY] Ništa nije pisano. Za pravi run makni --dry.')
        return
    if not filled:
        print('\n✔ Nema promjena — ništa za snimiti.')
        return
    backup = review.with_name(f'{review.stem}.pre-kartnaplata-{datetime.now():%Y%m%d_%H%M%S}.xlsx')
    shutil.copy2(review, backup)
    try:
        wb.save(review)
    except PermissionError:
        sys.exit(f'✗ Zatvori Review u Excelu i ponovi. (Backup: {backup.name})')
    print(f'\n✔ Snimljeno: {filled} Datum naplate popunjeno. Backup: {backup.name}')


if __name__ == '__main__':
    main()
