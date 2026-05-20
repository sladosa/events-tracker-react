# -*- coding: utf-8 -*-
"""
garmin_daily_metrics_to_xlsx.py
================================
Generates an app-importable xlsx for Area: Health_Sasa
  L1: Daily_metrics  ->  L2 leaf: Garmin_data

Data sources:
  UDS          DI-Connect-Aggregator/UDSFile_*.json
               -> hr_rest, hr_min, body_battery, steps, mod_min, vig_min,
                  calories_active, avg_stress
  VO2max       DI-Connect-Metrics/ActivityVo2Max_*.json + MetricsMaxMetData_*.json
               -> vo2max
  Training     DI-Connect-Metrics/TrainingHistory_*.json
               -> training_status

Sleep/HRV: not available in current GDPR export (DI-Connect-Wellness missing).
           Add back when files become available.

Usage:
  python garmin_daily_metrics_to_xlsx.py
  python garmin_daily_metrics_to_xlsx.py --out Health_daily_v2.xlsx
  python garmin_daily_metrics_to_xlsx.py --from 2020-01-01 --to 2025-12-31

Output: data-prep_data/Health/Health_daily_import.xlsx
"""

import sys, os, json, glob, argparse
from pathlib import Path
from datetime import date, datetime
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
GARMIN_DIR   = SCRIPT_DIR.parent.parent / "data-prep_data" / "DataFromGarmin"
DI_CONNECT   = GARMIN_DIR / "DI_CONNECT"
UDS_DIR      = DI_CONNECT / "DI-Connect-Aggregator"
METRICS_DIR  = DI_CONNECT / "DI-Connect-Metrics"
OUTPUT_DIR   = SCRIPT_DIR.parent.parent / "data-prep_data" / "Health"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── App config ─────────────────────────────────────────────────────────────────
AREA       = "Health_Sasa"
CAT_PATH   = "Daily_metrics > Garmin_data"
USER_EMAIL = "sasasladoljev59@gmail.com"

# ── Attribute definitions ──────────────────────────────────────────────────────
# (display_name, dict_key, attr_type, unit, description)
ATTRS = [
    # UDS — heart rate
    ("HR Rest",          "hr_rest",           "number",  "bpm",   "Garmin resting heart rate"),
    ("HR Min",           "hr_min",            "number",  "bpm",   "Lowest HR recorded during the day"),
    # UDS — body battery
    ("Body Battery High","body_battery_high",  "number",  "",      "Peak body battery level (0-100)"),
    ("Body Battery Low", "body_battery_low",   "number",  "",      "Lowest body battery level (0-100)"),
    # Metrics — VO2max
    ("VO2max",           "vo2max",            "number",  "ml/kg/min", "VO2max estimate from activity"),
    # UDS — daily activity
    ("Steps",            "steps",             "number",  "",      "Total daily steps"),
    ("Active Min Mod",   "mod_min",           "number",  "min",   "Moderate intensity minutes (WHO goal: 150/week)"),
    ("Active Min Vig",   "vig_min",           "number",  "min",   "Vigorous intensity minutes (WHO goal: 75/week)"),
    ("Calories Active",  "calories_active",   "number",  "kcal",  "Active calories burned"),
    # UDS — stress (2019+)
    ("Avg Stress",       "avg_stress",        "number",  "",      "Average daily stress level (0-100, Garmin algorithm)"),
    # TrainingHistory (2022+)
    ("Training Status",  "training_status",   "suggest", "",      "Garmin training status assessment"),
]

TRAINING_STATUS_OPTIONS = "MAINTAINING|PRODUCTIVE|UNPRODUCTIVE|RECOVERY|PEAKING|DETRAINING|STRAINED|OVERREACHING"

# ── Formatting ─────────────────────────────────────────────────────────────────
def _fill(h): return PatternFill("solid", fgColor=h)
def _font(**kw): return Font(**kw)
def _side(): return Side(style="thin")
def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())

BORDER    = _border()
ALIGN_L   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
ALIGN_C   = Alignment(horizontal="center", vertical="center")
FILL_BLUE = _fill("4472C4")
FILL_PURP = _fill("7030A0")
FILL_ATTR = _fill("DDEBF7")
FILL_STRU = _fill("E2EFDA")
FONT_WHT  = _font(bold=True, color="FFFFFF")
FONT_BOLD = _font(bold=True, size=12)

def col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

FIXED      = 8       # cols A-H fixed
ATTR_START = FIXED + 1   # col I onward

# ── Data loading ───────────────────────────────────────────────────────────────

def load_uds(date_from: date, date_to: date) -> dict:
    """Load UDS files -> {date_str: {hr_rest, hr_min, body_battery_*, steps, ...}}"""
    files = sorted(UDS_DIR.glob("UDSFile_*.json"))
    print(f"UDS files: {len(files)}")
    daily = {}
    for fp in files:
        for rec in json.loads(fp.read_text(encoding="utf-8")):
            cal = rec.get("calendarDate")
            if not cal:
                continue
            try:
                d = date.fromisoformat(str(cal)[:10])
            except ValueError:
                continue
            if d < date_from or d > date_to:
                continue
            dkey = d.isoformat()
            row = {}

            rhr = rec.get("restingHeartRate") or rec.get("currentDayRestingHeartRate")
            if rhr: row["hr_rest"] = int(rhr)

            minh = rec.get("minHeartRate")
            if minh: row["hr_min"] = int(minh)

            bb = rec.get("bodyBattery")
            if isinstance(bb, dict):
                for stat in bb.get("bodyBatteryStatList", []):
                    t = stat.get("bodyBatteryStatType", "")
                    v = stat.get("statsValue")
                    if t == "HIGHEST" and v is not None:
                        row["body_battery_high"] = int(v)
                    elif t == "LOWEST" and v is not None:
                        row["body_battery_low"] = int(v)

            steps = rec.get("totalSteps")
            if steps: row["steps"] = int(steps)

            mod = rec.get("moderateIntensityMinutes")
            if mod is not None: row["mod_min"] = int(mod)

            vig = rec.get("vigorousIntensityMinutes")
            if vig is not None: row["vig_min"] = int(vig)

            cal_act = rec.get("activeKilocalories")
            if cal_act: row["calories_active"] = round(float(cal_act))

            stress_agg = rec.get("allDayStress", {}).get("aggregatorList", [])
            for agg in stress_agg:
                if agg.get("type") == "TOTAL":
                    lvl = agg.get("averageStressLevel")
                    if lvl is not None:
                        row["avg_stress"] = int(lvl)
                    break

            # Merge: newer file wins for same date
            existing = daily.get(dkey, {})
            existing.update({k: v for k, v in row.items() if v is not None})
            daily[dkey] = existing

    print(f"UDS: {len(daily)} unique dates")
    return daily


def load_vo2max(date_from: date, date_to: date) -> dict:
    """Load VO2max per calendar date from ActivityVo2Max and MetricsMaxMetData."""
    vo2 = {}
    for pattern in ("ActivityVo2Max_*.json", "MetricsMaxMetData_*.json"):
        for fp in sorted(METRICS_DIR.glob(pattern)):
            records = json.loads(fp.read_text(encoding="utf-8"))
            if not isinstance(records, list):
                records = [records]
            for rec in records:
                cal = rec.get("calendarDate")
                val = rec.get("vo2MaxValue")
                if not cal or val is None:
                    continue
                try:
                    d = date.fromisoformat(str(cal)[:10])
                except ValueError:
                    continue
                if date_from <= d <= date_to:
                    vo2.setdefault(d.isoformat(), float(val))
    print(f"VO2max: {len(vo2)} dates")
    return vo2


def load_training_status(date_from: date, date_to: date) -> dict:
    """Load TrainingHistory -> {date_str: training_status}"""
    files = sorted(METRICS_DIR.glob("TrainingHistory_*.json"))
    print(f"TrainingHistory files: {len(files)}")
    daily = {}
    for fp in files:
        for rec in json.loads(fp.read_text(encoding="utf-8")):
            cal = rec.get("calendarDate")
            status = rec.get("trainingStatus", "")
            if not cal or not status or status == "NO_STATUS":
                continue
            try:
                d = date.fromisoformat(str(cal)[:10])
            except ValueError:
                continue
            if date_from <= d <= date_to:
                daily[d.isoformat()] = status
    print(f"TrainingHistory: {len(daily)} dates with status")
    return daily


def _sec_to_min(v):
    if v is None: return None
    try: return round(float(v) / 60, 1)
    except (TypeError, ValueError): return None


# ── Merge per date ─────────────────────────────────────────────────────────────

def merge_daily(uds: dict, vo2: dict, training: dict) -> list:
    attr_keys = {a[1] for a in ATTRS}
    events = []
    all_dates = sorted(set(uds) | set(vo2) | set(training))
    for dkey in all_dates:
        row = {"date": dkey}
        row.update(uds.get(dkey, {}))
        if dkey in vo2:
            row["vo2max"] = vo2[dkey]
        if dkey in training:
            row["training_status"] = training[dkey]
        if not any(row.get(k) is not None for k in attr_keys):
            continue
        events.append(row)
    return events


# ── Build Events sheet ─────────────────────────────────────────────────────────

def build_events_sheet(ws, events: list):
    ws.title = "Events"
    n_attr       = len(ATTRS)
    attr_letters = [col_letter(ATTR_START + i) for i in range(n_attr)]
    attr_names   = [a[0] for a in ATTRS]
    attr_keys    = [a[1] for a in ATTRS]
    row = 1

    # ── ATTRIBUTE LEGEND ──────────────────────────────────────────────────────
    ws.cell(row, 1, "ATTRIBUTE LEGEND:").font = FONT_BOLD
    row += 1
    for ci, h in enumerate(["Col", "Area", "Category_Path", "Attribute", "Type", "Unit"], 1):
        c = ws.cell(row, ci, h)
        c.fill = FILL_PURP; c.font = FONT_WHT; c.alignment = ALIGN_C; c.border = BORDER
    row += 1
    for i, (name, key, dtype, unit, _desc) in enumerate(ATTRS):
        data = [attr_letters[i], AREA, CAT_PATH, name, dtype, unit or None]
        for ci, v in enumerate(data, 1):
            c = ws.cell(row, ci, v)
            c.fill = FILL_ATTR; c.border = BORDER; c.alignment = ALIGN_L
        row += 1
    row += 1  # blank separator

    # ── EVENT DATA ────────────────────────────────────────────────────────────
    ws.cell(row, 1, "EVENT DATA:").font = FONT_BOLD
    row += 1
    header_row = row
    fixed_hdrs = ["event_id", "Area", "Category_Path", "event_date", "session_start",
                  "created_at", "User", "leaf comment"]
    for ci, h in enumerate(fixed_hdrs + attr_names, 1):
        c = ws.cell(row, ci, h)
        c.fill = FILL_BLUE; c.font = FONT_WHT; c.alignment = ALIGN_C; c.border = BORDER
    ws.row_dimensions[row].height = 28
    row += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ev in events:
        fixed_vals = [None, AREA, CAT_PATH, ev["date"], "07:00",
                      None, USER_EMAIL, None]
        attr_vals = [ev.get(k) for k in attr_keys]
        for ci, v in enumerate(fixed_vals + attr_vals, 1):
            ws.cell(row, ci, v)
        row += 1

    # ── AutoFilter on header row ──────────────────────────────────────────────
    last_col = col_letter(FIXED + n_attr)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row}"

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [12, 14, 34, 12, 14, 12, 28, 16] + [16] * n_attr
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[col_letter(i)].width = w
    # Freeze panes at col I (first attr col) and below header
    ws.freeze_panes = f"I{header_row + 1}"


# ── Build Structure sheet ──────────────────────────────────────────────────────

STRUCTURE_HEADERS = [
    "Type", "IsLeaf", "Area", "SharedWith", "CategoryPath", "Sort",
    "AttrName", "Slug", "AttrType", "IsRequired", "Val.Type", "Default",
    "Val.Max (no)", "Unit", "TextOptions/Val.Min", "DependsOn", "WhenValue", "Description"
]

def build_structure_sheet(ws):
    ws.title = "Structure"
    for ci, h in enumerate(STRUCTURE_HEADERS, 1):
        c = ws.cell(1, ci, h)
        c.fill = FILL_STRU; c.font = _font(bold=True); c.border = BORDER; c.alignment = ALIGN_C
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["E"].width = 46
    ws.column_dimensions["R"].width = 46
    row = 2

    def srow(type_, cat_path, is_leaf="", attr_name="", slug="", attr_type="",
             unit="", val_type="", text_options="", description="", sort=0):
        vals = [type_, is_leaf, AREA, "", f"{AREA} > {cat_path}", sort,
                attr_name, slug, attr_type, "FALSE", val_type, "", "", unit,
                text_options, "", "", description]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row, ci, v or None)
            c.border = BORDER; c.alignment = ALIGN_L
        return row + 1

    # L1 category
    row = srow("Category", "Daily_metrics", sort=1)
    # L2 leaf
    row = srow("Category", "Daily_metrics > Garmin_data", is_leaf="TRUE", sort=1)
    # Attributes
    for i, (name, slug, dtype, unit, desc) in enumerate(ATTRS, 1):
        vtype = "suggest" if dtype == "suggest" else ""
        opts  = TRAINING_STATUS_OPTIONS if slug == "training_status" else ""
        row = srow("Attribute", "Daily_metrics > Garmin_data",
                   attr_name=name, slug=slug, attr_type=dtype,
                   unit=unit, val_type=vtype, text_options=opts,
                   description=desc, sort=i)


# ── Build Info sheet ───────────────────────────────────────────────────────────

def build_info_sheet(ws, n_events: int, n_vo2: int, n_train: int):
    ws.title = "Info"
    rows = [
        ("Generated",       datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Script",          "garmin_daily_metrics_to_xlsx.py"),
        ("Area",            AREA),
        ("Category",        CAT_PATH),
        ("Events total",    n_events),
        ("VO2max dates",    n_vo2),
        ("Training dates",  n_train),
        ("",                ""),
        ("SLEEP/HRV",       "Not available - DI-Connect-Wellness missing from GDPR export"),
        ("",                ""),
        ("UDS source",      str(UDS_DIR)),
        ("Metrics source",  str(METRICS_DIR)),
    ]
    for r, (k, v) in enumerate(rows, 1):
        ws.cell(r, 1, k).font = _font(bold=True)
        ws.cell(r, 2, str(v) if v is not None else "")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 65


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--from", dest="date_from", default="2014-01-01")
    ap.add_argument("--to",   dest="date_to",   default="2025-12-31")
    ap.add_argument("--out",  default="Health_daily_import.xlsx")
    args = ap.parse_args()

    date_from = date.fromisoformat(args.date_from)
    date_to   = date.fromisoformat(args.date_to)
    print(f"Date range: {date_from} -> {date_to}")

    uds      = load_uds(date_from, date_to)
    vo2      = load_vo2max(date_from, date_to)
    training = load_training_status(date_from, date_to)
    events   = merge_daily(uds, vo2, training)
    print(f"Total events: {len(events)}")

    if not events:
        print("ERROR: no events — check paths"); sys.exit(1)

    out_path = OUTPUT_DIR / args.out
    wb = openpyxl.Workbook()
    build_events_sheet(wb.active, events)
    build_structure_sheet(wb.create_sheet())
    build_info_sheet(wb.create_sheet(), len(events), len(vo2), len(training))
    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    n_train = len(training)
    print(f"  {len(events)} events | {len(vo2)} VO2max | {n_train} training status")
    print(f"  Attributes ({len(ATTRS)}): {', '.join(a[0] for a in ATTRS)}")


if __name__ == "__main__":
    main()
