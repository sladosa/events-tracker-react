# -*- coding: utf-8 -*-
"""
health_lab_review.py
=====================
Reads Health area from Supabase (PROD by default).
Generates a review + import Excel for cleaning up Medical Visits.

Sheets in output:
  "Review - Lab Results"     -- all 45 Lab Results events, full comment + parsing suggestions
  "Review - Medical Visits"  -- all 13 existing Medical Visits + parsing suggestions for doktor/vrsta/iznos
  "Events"                   -- proposed NEW Medical Visit events (app-importable format)
  "Structure"                -- Medical Visit structure (needed for app import)

Usage:
  python health_lab_review.py
  python health_lab_review.py --env .env.prod.local --out Health_medvisit_review.xlsx
"""

import sys, os, re, json, argparse
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
REPO_ROOT  = SCRIPT_DIR.parent.parent
DATA_DIR   = REPO_ROOT / "data-prep_data" / "Health"

# ── Styles ────────────────────────────────────────────────────────────────────
def fill(hex6): return PatternFill("solid", fgColor=hex6)
def font(bold=False, color="000000", size=11): return Font(bold=bold, color=color, size=size)
def border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

FILL_PURPLE  = fill("7030A0")
FILL_BLUE    = fill("4472C4")
FILL_GREEN   = fill("E2EFDA")   # confident parse
FILL_YELLOW  = fill("FFF2CC")   # uncertain parse
FILL_ORANGE  = fill("FCE4D6")   # no medical info
FILL_GREY    = fill("F2F2F2")   # existing visit reference
FILL_WHITE   = fill("FFFFFF")
FILL_HEADER  = fill("D9E1F2")

FONT_WHITE   = font(bold=True, color="FFFFFF")
FONT_BOLD    = font(bold=True)
FONT_NORMAL  = font()

ALIGN_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_R  = Alignment(horizontal="right",  vertical="center")
BORDER   = border()

# ── Known doctors / places ────────────────────────────────────────────────────
# Patterns use \w* suffix to handle Croatian declension (Filipovica, Mihaljevica...)
DOCTOR_MAP = [
    (r"\bBates\b",              "Bates"),
    (r"\bBAT[EI]S\b",           "Bates"),
    (r"\bBreyer\w*\b",          "Breyer"),
    (r"\bBedalov\w*\b",         "Bedalov"),
    (r"sv\.Duh",                "sv.Duh - Davorka"),
    (r"\bDavorka\b",            "sv.Duh - Davorka"),
    (r"\bNemetov\w*\b",         "Nemetova"),
    (r"\bKB\s*DUBRAVA\b",       "KB Dubrava"),
    (r"\bDubrava\b",            "KB Dubrava"),
    (r"\bBajera?\b",            "Bajer"),
    (r"\bFilipovic\w*\b",       "Filipovic"),
    (r"\bGalovic\w*\b",         "Galovic"),
    (r"\bMihaljevic\w*\b",      "Mihaljevic"),
    (r"\bIvkovic\w*\b",         "Alan Ivkovic"),
    (r"\bZrinka\b",             "Zrinka"),
    (r"\bDragan\w*\b",          "dr. Dragan Mihaljevic"),
    (r"\bBek\b",                "dr. Bek"),
    (r"\bBekic\w*\b",           "dr. Bekic"),
    (r"\bZoran\w*\b",           "Zoran Filipovic"),
]

# ── Known visit types ─────────────────────────────────────────────────────────
VISIT_MAP = [
    (r"sistematski\s*pregled",  "Sistematski pregled"),
    (r"sistematski",            "Sistematski pregled"),
    (r"\bUZV\b",                "UZV"),
    (r"\bRTG\b",                "RTG"),
    (r"\bEKG\b",                "EKG"),
    (r"\bMR\b",                 "MR"),
    (r"ergometrij",             "Ergometrija"),
    (r"kolonoskopij",           "Kolonoskopija"),
    (r"urolo",                  "Urološka kontrola"),
    (r"biopsij",                "Biopsija"),
    (r"implanta",               "Implantacija"),
    (r"udarni val",             "Udarne valove"),
    (r"udlaga",                 "Zubna udlaga"),
    (r"kruna",                  "Zubna kruna"),
    (r"operacij",               "Operacija"),
    (r"vadjenje slijepog",      "Operacija"),
    (r"Kardiolo",               "Kardiološki pregled"),
    (r"\bkontrol[ae]\b",        "Kontrola"),
    (r"\bpregled\b",            "Pregled"),
]

EUR_RE = re.compile(r"(\d+[\.,]?\d*)\s*eur", re.IGNORECASE)

# ── Editirana napomena: strip doctor + EUR, keep visit type for context ────────

def make_editirana_napomena(full_text: str) -> str:
    """Remove doctor names and EUR amounts from text; keep visit type keywords."""
    if not full_text:
        return ""
    text = full_text

    # Remove EUR amounts (e.g. "180eur", "550 eur", "396EUR")
    text = re.sub(r"\b\d+[\.,]?\d*\s*eur\b", "", text, flags=re.IGNORECASE)

    # Remove all doctor name patterns (handles Croatian declension)
    for pattern, _label in DOCTOR_MAP:
        text = re.sub(pattern, "", text, flags=re.IGNORECASE)

    # Remove connective orphans left after doctor removal
    text = re.sub(r"\bkod\s+dr\.?\s*",    "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bdr\.?\s+",          "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bkod\s+",            "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bu\s+poliklinici\b", "", text, flags=re.IGNORECASE)

    # Clean up punctuation artifacts
    text = re.sub(r"[ \t]+",       " ",  text)         # collapse spaces
    text = re.sub(r"\s*,\s*,+",   ",",  text)          # double commas
    text = re.sub(r"\s*-\s*,",    ",",  text)          # "- ,"
    text = re.sub(r",\s*-\s*$",   "",   text)          # trailing ", -"
    text = re.sub(r"\s*-\s*$",    "",   text)          # trailing " -"
    text = re.sub(r"^\s*[-,]\s*", "",   text)          # leading "- " or ", "
    text = re.sub(r"\(\s*\)",     "",   text)          # empty parens "()"
    text = text.strip().strip(",").strip("-").strip()
    return text

# ── Parsing ───────────────────────────────────────────────────────────────────

def parse_comment(comment: str) -> dict:
    """
    Returns parsed fields from a comment string.
    comment format: "[AutoFlags |] manual note"
    """
    if not comment:
        return dict(auto_flags="", manual_note="", doktor="", vrsta="", iznos=None, napomena="", confidence="none")

    # Split auto-flags from manual note (auto-flags end at " | " separator)
    if " | " in comment:
        auto_flags, manual_note = comment.split(" | ", 1)
    else:
        # If comment only has H/L markers and nothing useful
        auto_flags_re = re.compile(r'^([\w\s·]+ [HL](?: · [\w\s·]+ [HL])*)$')
        if auto_flags_re.match(comment.strip()):
            auto_flags = comment.strip()
            manual_note = ""
        else:
            auto_flags = ""
            manual_note = comment.strip()

    note = manual_note.strip()

    # Extract EUR amount
    iznos = None
    m = EUR_RE.search(note)
    if m:
        try:
            iznos = float(m.group(1).replace(",", "."))
        except ValueError:
            pass

    # Extract doktor
    doktor = ""
    for pattern, label in DOCTOR_MAP:
        if re.search(pattern, note, re.IGNORECASE):
            doktor = label if not doktor else doktor + ", " + label
            break  # first match wins for doktor

    # Extract vrsta
    vrsta = ""
    for pattern, label in VISIT_MAP:
        if re.search(pattern, note, re.IGNORECASE):
            vrsta = label
            break

    # "vadjenje krv" alone (blood draw) without any real visit type → lower confidence
    is_blood_draw_only = bool(re.search(r"vadjenje krvi?", note, re.IGNORECASE)) and not vrsta

    # Confidence level
    if vrsta and doktor:
        confidence = "high"
    elif vrsta or (doktor and not is_blood_draw_only):
        confidence = "medium"
    elif doktor and is_blood_draw_only:
        confidence = "low"     # just a blood draw location, not really a visit
    else:
        confidence = "none"

    # Napomena = full manual note (let user decide what to keep)
    napomena = note if note else ""

    edit_nap = make_editirana_napomena(note)

    return dict(
        auto_flags=auto_flags,
        manual_note=note,
        doktor=doktor,
        vrsta=vrsta,
        iznos=iznos,
        napomena=napomena,
        edit_nap=edit_nap,
        confidence=confidence,
    )

# ── Excel helpers ─────────────────────────────────────────────────────────────

def cell(ws, row, col, value="", fill=None, font=None, align=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill   = fill
    if font:   c.font   = font
    if align:  c.alignment = align
    if border: c.border = border
    if number_format: c.number_format = number_format
    return c

def hdr(ws, row, col, value, bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = bg or FILL_HEADER
    c.font   = FONT_BOLD
    c.alignment = ALIGN_C
    c.border = BORDER
    return c

# ── Sheet: Review - Lab Results ───────────────────────────────────────────────

def build_review_lab(ws, events, existing_visit_dates):
    ws.title = "Review - Lab Results"

    CONF_FILL = {
        "high":   FILL_GREEN,
        "medium": FILL_YELLOW,
        "low":    FILL_ORANGE,
        "none":   FILL_WHITE,
    }
    CONF_LABEL = {
        "high":   "✓ new visit",
        "medium": "? check",
        "low":    "– blood draw",
        "none":   "– no visit",
    }

    headers = ["Date", "Auto-flags (H/L)", "Manual note (full)", "⚙ Confidence",
               "→ Doktor", "→ Vrsta", "→ Iznos (EUR)", "→ Napomena",
               "Existing visit\non same date?"]
    col_widths = [14, 38, 50, 14, 22, 24, 12, 42, 12]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        hdr(ws, 1, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for ri, ev in enumerate(events, 2):
        date_str = ev["session_start"][:10]
        p = parse_comment(ev.get("comment") or "")
        has_visit = date_str in existing_visit_dates
        row_fill = CONF_FILL[p["confidence"]]

        vals = [
            date_str,
            p["auto_flags"],
            p["manual_note"],
            CONF_LABEL[p["confidence"]],
            p["doktor"],
            p["vrsta"],
            p["iznos"],
            p["napomena"],
            "YES – exists" if has_visit else "",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill   = row_fill
            c.alignment = ALIGN_L
            c.border = BORDER

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ── Sheet: Review - Medical Visits ────────────────────────────────────────────

def build_review_visits(ws, events):
    ws.title = "Review - Medical Visits"

    headers = ["Date", "Napomena (full)", "→ Doktor", "→ Vrsta", "→ Iznos (EUR)",
               "Editirana Napomena (→ import)"]
    col_widths = [14, 60, 22, 24, 12, 52]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        hdr(ws, 1, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for ri, ev in enumerate(events, 2):
        date_str = ev["session_start"][:10]
        napomena = ev.get("napomena") or ""
        p = parse_comment(napomena)

        vals = [date_str, napomena, p["doktor"], p["vrsta"], p["iznos"], p["edit_nap"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill      = FILL_GREY
            c.alignment = ALIGN_L
            c.border    = BORDER

# ── Sheet: Events (import-ready Medical Visits) ───────────────────────────────
# Format mirrors make_health_events.py — the app parser requires:
#   Row 1:  "ATTRIBUTE LEGEND:" in col A
#   Row 2:  legend header (Col, Area, Category_Path, Attribute, Type, Unit)
#   Row 3+: one row per attr: col_letter | area | cat_path | attr_name | type | unit
#   (blank row)
#   Next:   "EVENT DATA:" in col A
#   Next:   header row (event_id, Area, Category_Path, event_date, ...)
#   Next+:  data rows — col A empty = CREATE new event

AREA_DEFAULT = "Health_Sasa"
CAT_PATH     = "Medical > Medical Visit"   # bez area name — col C in data rows
FIXED_COUNT  = 8                           # cols A–H

# Attribute columns: (display_name, data_type, unit)
VISIT_ATTRS = [
    ("Doktor",   "suggest", ""),
    ("Vrsta",    "suggest", ""),
    ("Iznos",    "number",  "EUR"),
    ("Napomena", "text",    ""),
]

def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def build_events_sheet(ws, proposed_rows, existing_rows, user_email, area=AREA_DEFAULT):
    ws.title = "Events"

    n_attr     = len(VISIT_ATTRS)
    attr_start = FIXED_COUNT + 1          # col I = index 9
    total_cols = FIXED_COUNT + n_attr     # 12
    attr_letters = [_col_letter(attr_start + i) for i in range(n_attr)]  # I J K L

    row = 1

    # ── ATTRIBUTE LEGEND title ────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="ATTRIBUTE LEGEND:").font = Font(bold=True, size=12)
    row += 1

    # Legend header row
    for ci, h in enumerate(["Col", "Area", "Category_Path", "Attribute", "Type", "Unit"], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = FILL_PURPLE; c.font = FONT_WHITE
        c.alignment = ALIGN_C; c.border = BORDER
    row += 1

    # One legend row per attribute — col A = Excel column letter (I, J, K, L)
    VISIT_FILL = fill("DDEBF7")
    for i, (attr_name, dtype, unit) in enumerate(VISIT_ATTRS):
        data = [attr_letters[i], area, CAT_PATH, attr_name, dtype, unit]
        for ci, v in enumerate(data, 1):
            c = ws.cell(row=row, column=ci, value=v or None)
            c.fill = VISIT_FILL; c.border = BORDER; c.alignment = ALIGN_L
        row += 1

    row += 1  # blank separator

    # ── EVENT DATA title ──────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="EVENT DATA:").font = Font(bold=True, size=12)
    row += 1

    # ── EVENT DATA header ─────────────────────────────────────────────────────
    event_header_row = row
    cat_short = CAT_PATH.split(" > ")[-1]   # "Medical Visit"
    fixed_hdrs = ["event_id", "Area", "Category_Path", "event_date", "session_start",
                  "created_at", "User", "leaf comment"]
    attr_hdrs  = [f"{name} ({cat_short})" for name, _, _ in VISIT_ATTRS]

    for ci, h in enumerate(fixed_hdrs + attr_hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = FILL_BLUE; c.font = FONT_WHITE
        c.alignment = ALIGN_C; c.border = BORDER
    ws.row_dimensions[row].height = 28
    row += 1

    event_data_start = row

    # ── Data row writer ───────────────────────────────────────────────────────
    FILL_AMBER = fill("FFF2CC")   # existing events
    FILL_NEW_H = fill("E2EFDA")   # new, high confidence
    FILL_NEW_M = fill("FFFACD")   # new, medium confidence

    def write_data_row(pr, row_num, row_fill):
        date_str = pr["date"][:10] if isinstance(pr["date"], str) else pr["date"].strftime("%Y-%m-%d")
        time_str = pr.get("time", "08:00")
        fixed_vals = [
            None,        # event_id: empty = CREATE new
            area,        # Area
            CAT_PATH,    # Category_Path (no area prefix)
            date_str,    # event_date
            time_str,    # session_start (time only)
            None,        # created_at
            user_email,  # User
            None,        # leaf comment
        ]
        attr_vals = [
            pr.get("doktor",   "") or None,
            pr.get("vrsta",    "") or None,
            pr.get("iznos",    None),
            pr.get("edit_nap", "") or None,
        ]
        for ci, v in enumerate(fixed_vals + attr_vals, 1):
            c = ws.cell(row=row_num, column=ci, value=v)
            c.fill = row_fill; c.alignment = ALIGN_L; c.border = BORDER

    # Existing Medical Visit events (amber) — sorted by date ascending
    for pr in existing_rows:
        write_data_row(pr, row, FILL_AMBER)
        row += 1

    # New events from Lab Results (green/yellow)
    for pr in proposed_rows:
        row_fill = FILL_NEW_H if pr.get("confidence") == "high" else FILL_NEW_M
        write_data_row(pr, row, row_fill)
        row += 1

    event_data_end = row - 1

    # ── Autofilter + freeze ───────────────────────────────────────────────────
    last_col = _col_letter(total_cols)
    ws.auto_filter.ref = f"A{event_header_row}:{last_col}{event_data_end}"
    ws.freeze_panes    = f"I{event_data_start}"

    # ── Column widths ─────────────────────────────────────────────────────────
    for ltr, w in zip("ABCDEFGH", [10, 12, 30, 12, 10, 10, 30, 14]):
        ws.column_dimensions[ltr].width = w
    for ltr, w in zip(attr_letters, [22, 24, 8, 52]):
        ws.column_dimensions[ltr].width = w

# ── Sheet: Structure ──────────────────────────────────────────────────────────

def build_structure_sheet(ws, area=AREA_DEFAULT):
    ws.title = "Structure"

    headers = ["Type", "CategoryPath", "Sort", "AttrName", "Slug", "AttrType",
               "IsRequired", "Val.Type", "Default", "ValMax", "Unit",
               "TextOptions", "DependsOn", "WhenValue", "Description"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = FILL_BLUE
        c.font = FONT_WHITE
        c.alignment = ALIGN_C
        c.border = BORDER

    rows = [
        # Type | CategoryPath (WITH area) | Sort | AttrName | Slug | AttrType | IsRequired | ...
        ["area",      f"{area}",                                    "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["category",  f"{area} > Medical",                          "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["category",  f"{area} > Medical > Medical Visit",          "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["attribute", f"{area} > Medical > Medical Visit", 10, "Doktor",   "doktor",   "text",   "false", "suggest", "", "", "",
         "Bates,Bedalov,Breyer,sv.Duh - Davorka,Filipovic,Galovic,Mihaljevic,Alan Ivkovic,Zrinka,Bajer,KB Dubrava,Nemetova,Ostalo",
         "", "", "Doctor or specialist seen"],
        ["attribute", f"{area} > Medical > Medical Visit", 20, "Vrsta",    "vrsta",    "text",   "false", "suggest", "", "", "",
         "Sistematski pregled,Kontrola,UZV,RTG,EKG,MR,Ergometrija,Kolonoskopija,Urološka kontrola,Kardiološki pregled,Operacija,Biopsija,Implantacija,Udarne valove,Pregled,Ostalo",
         "", "", "Type of medical visit or procedure"],
        ["attribute", f"{area} > Medical > Medical Visit", 30, "Iznos",    "iznos",    "number", "false", "",        "", "", "EUR",
         "", "", "", "Cost of the visit or procedure"],
        ["attribute", f"{area} > Medical > Medical Visit", 40, "Napomena", "napomena", "text",   "false", "",        "", "", "",
         "", "", "", "Notes: diagnosis, therapy, doctor recommendations"],
    ]

    for ri, row_vals in enumerate(rows, 2):
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = ALIGN_L
            c.border    = BORDER

    col_widths = [10, 40, 5, 16, 16, 10, 10, 10, 10, 10, 6, 60, 14, 14, 36]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

# ── Main ──────────────────────────────────────────────────────────────────────

def load_env(env_file):
    env = {}
    with open(env_file, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--env",      default=".env.prod.local")
    parser.add_argument("--email",    default="sasasladoljev59@gmail.com")
    parser.add_argument("--out",      default="")
    parser.add_argument("--area-out", default=AREA_DEFAULT,
                        help="Area name used in the import Excel (default: Health_Sasa)")
    args = parser.parse_args()

    env_path = REPO_ROOT / args.env
    if not env_path.exists():
        print(f"ERROR: env file not found: {env_path}")
        sys.exit(1)

    env = load_env(str(env_path))
    url = env.get("SUPABASE_URL") or env.get("VITE_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required")
        sys.exit(1)

    try:
        from supabase import create_client
    except ImportError:
        print("ERROR: pip install supabase")
        sys.exit(1)

    sb = create_client(url, key)

    # ── Resolve user_id ───────────────────────────────────────────────────────
    profile = sb.table("profiles").select("id, email").eq("email", args.email).execute()
    if not profile.data:
        print(f"ERROR: no profile for {args.email}")
        sys.exit(1)
    user_id    = profile.data[0]["id"]
    user_email = profile.data[0]["email"]
    print(f"User: {user_email} ({user_id})")

    # ── Find Health area ──────────────────────────────────────────────────────
    areas = sb.table("areas").select("id, name").eq("user_id", user_id).execute().data
    health_area = next((a for a in areas if "health" in a["name"].lower()), None)
    if not health_area:
        print(f"ERROR: Health area not found. Areas: {[a['name'] for a in areas]}")
        sys.exit(1)
    print(f"Area: {health_area['name']} ({health_area['id']})")

    # ── Load all categories ───────────────────────────────────────────────────
    cats = sb.table("categories").select("id, name, slug, parent_category_id, level") \
              .eq("area_id", health_area["id"]).execute().data
    cat_by_id   = {c["id"]: c for c in cats}
    cat_by_name = {c["name"].lower(): c for c in cats}

    lab_cat    = cat_by_name.get("lab results")
    visit_cat  = cat_by_name.get("medical visit")
    if not lab_cat:
        print(f"ERROR: 'Lab Results' category not found. Cats: {[c['name'] for c in cats]}")
        sys.exit(1)
    if not visit_cat:
        print(f"ERROR: 'Medical Visit' category not found.")
        sys.exit(1)

    print(f"Lab Results cat:  {lab_cat['id']}")
    print(f"Medical Visit cat: {visit_cat['id']}")

    # ── Load Lab Results events (full comment, no truncation) ─────────────────
    print("Loading Lab Results events...")
    lab_events = sb.table("events") \
        .select("id, session_start, comment") \
        .eq("user_id", user_id) \
        .eq("category_id", lab_cat["id"]) \
        .order("session_start", desc=True) \
        .execute().data
    print(f"  {len(lab_events)} events found")

    # ── Load existing Medical Visit events ────────────────────────────────────
    print("Loading Medical Visit events...")
    visit_event_ids_q = sb.table("events") \
        .select("id, session_start") \
        .eq("user_id", user_id) \
        .eq("category_id", visit_cat["id"]) \
        .execute().data

    visit_event_ids = [e["id"] for e in visit_event_ids_q]
    existing_visit_dates = {e["session_start"][:10] for e in visit_event_ids_q}
    print(f"  {len(visit_event_ids_q)} existing Medical Visit events")

    # Load napomena attr def for Medical Visit
    visit_attrs = sb.table("attribute_definitions") \
        .select("id, name, slug") \
        .eq("category_id", visit_cat["id"]) \
        .execute().data
    visit_attr_by_slug = {a["slug"]: a for a in visit_attrs}
    napomena_attr = visit_attr_by_slug.get("napomena")

    # Load napomena values for existing visits
    visit_events_full = []
    if visit_event_ids and napomena_attr:
        chunk = 50
        for i in range(0, len(visit_event_ids), chunk):
            ids = visit_event_ids[i:i+chunk]
            ea = sb.table("event_attributes") \
                .select("event_id, value_text") \
                .in_("event_id", ids) \
                .eq("attribute_definition_id", napomena_attr["id"]) \
                .execute().data
            ea_by_event = {e["event_id"]: e["value_text"] for e in ea}
            for ev in visit_event_ids_q[i:i+chunk]:
                visit_events_full.append({
                    "session_start": ev["session_start"],
                    "napomena": ea_by_event.get(ev["id"], ""),
                })

    # ── Build proposed new Medical Visit rows ─────────────────────────────────
    proposed = []
    for ev in lab_events:
        date_str = ev["session_start"][:10]
        p = parse_comment(ev.get("comment") or "")
        if p["confidence"] in ("high", "medium") and date_str not in existing_visit_dates:
            proposed.append({
                "date":     date_str,
                "doktor":   p["doktor"],
                "vrsta":    p["vrsta"],
                "iznos":    p["iznos"],
                "napomena": p["napomena"],
                "confidence": p["confidence"],
            })

    print(f"\nParsing summary:")
    from collections import Counter
    conf_counts = Counter(parse_comment(ev.get("comment",""))["confidence"] for ev in lab_events)
    for k, v in sorted(conf_counts.items()):
        print(f"  {k}: {v}")
    print(f"\nProposed NEW Medical Visit events: {len(proposed)}")
    for p in proposed:
        print(f"  {p['date']}  [{p['confidence']}]  Doktor: {p['doktor']!r}  Vrsta: {p['vrsta']!r}  Iznos: {p['iznos']}")

    # ── Build existing_rows for Events sheet ──────────────────────────────────
    # Parse each existing Medical Visit's napomena → doktor/vrsta/iznos/edit_nap
    existing_rows_for_import = []
    for ev in visit_events_full:
        napomena = ev.get("napomena") or ""
        p = parse_comment(napomena)
        existing_rows_for_import.append({
            "date":     ev["session_start"],
            "doktor":   p["doktor"],
            "vrsta":    p["vrsta"],
            "iznos":    p["iznos"],
            "edit_nap": p["edit_nap"],
            "confidence": "existing",
        })
    # Sort by date ascending so they appear chronologically
    existing_rows_for_import.sort(key=lambda x: x["date"])
    proposed_sorted = sorted(proposed, key=lambda x: x["date"])

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    ws_events   = wb.create_sheet("Events")
    ws_struct   = wb.create_sheet("Structure")
    ws_lab_rev  = wb.create_sheet("Review - Lab Results")
    ws_vis_rev  = wb.create_sheet("Review - Medical Visits")

    area_out = args.area_out
    print(f"\nOutput area name: '{area_out}'")

    build_events_sheet(ws_events, proposed_sorted, existing_rows_for_import, user_email, area=area_out)
    build_structure_sheet(ws_struct, area=area_out)
    build_review_lab(ws_lab_rev, lab_events, existing_visit_dates)
    build_review_visits(ws_vis_rev, visit_events_full)

    # ── Save ──────────────────────────────────────────────────────────────────
    out_path = args.out or str(DATA_DIR / "Health_medvisit_review.xlsx")
    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    print(f"\nNext steps:")
    print(f"  1. Open {Path(out_path).name} in Excel")
    print(f"  2. Review 'Review - Lab Results' (green=confident, yellow=check, orange=blood draw only)")
    print(f"  3. Review 'Review - Medical Visits' (existing events — update manually in app)")
    print(f"  4. Edit 'Events' sheet — remove rows you don't want, fix doktor/vrsta/napomena")
    print(f"  5. Import via app: Activities → Import → select this file")
    print(f"     (Import reads 'Events' + 'Structure' sheets; Review tabs are ignored)")


if __name__ == "__main__":
    main()
