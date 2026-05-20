"""
compare_xlsx.py — Uspoređuje dva Events xlsx exporta.
Pronalazi EVENT DATA header, čita podatke i ispisuje razlike.

Korištenje:
    python compare_xlsx.py file1.xlsx file2.xlsx
"""

import sys
import pandas as pd
import openpyxl


def find_data_header_row(ws):
    """Pronađi redni broj (1-based) retka koji ima 'event_id' u koloni A."""
    for row in ws.iter_rows():
        if row[0].value == 'event_id':
            return row[0].row
    raise ValueError("Nije pronađen header red s 'event_id'")


def load_event_data(path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['Events']

    header_row = find_data_header_row(ws)

    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(min_row=header_row, values_only=True)):
        if headers is None:
            headers = [str(c) if c is not None else f'col_{i}' for i, c in enumerate(row)]
            continue
        if row[0] is None:   # prazni redovi ispod podataka
            break
        rows.append(row)

    wb.close()
    df = pd.DataFrame(rows, columns=headers)
    df['event_date'] = pd.to_datetime(df['event_date'], errors='coerce').dt.date
    return df


def compare(path1: str, path2: str):
    print(f"\nUčitavam:\n  A: {path1}\n  B: {path2}\n")

    df1 = load_event_data(path1)
    df2 = load_event_data(path2)

    label1 = path1.split('\\')[-1].split('/')[-1]
    label2 = path2.split('\\')[-1].split('/')[-1]

    # ── Osnovne statistike ──────────────────────────────────────────
    print(f"{'':30s}  {'A':>12s}  {'B':>12s}")
    print("-" * 58)
    print(f"{'Broj redova':<30s}  {len(df1):>12,}  {len(df2):>12,}")

    if 'event_date' in df1.columns and len(df1):
        print(f"{'Najstariji datum':<30s}  {str(df1['event_date'].min()):>12s}  {str(df2['event_date'].min()):>12s}")
        print(f"{'Najnoviji datum':<30s}  {str(df1['event_date'].max()):>12s}  {str(df2['event_date'].max()):>12s}")

    # Numeričke kolone (ne event_id, Area, Category_Path, User, comment)
    skip = {'event_id', 'Area', 'Category_Path', 'event_date',
            'session_start', 'created_at', 'User', 'leaf comment', 'user_email'}
    num_cols = [c for c in df1.columns if c not in skip]

    print()
    print(f"{'Kolona':<30s}  {'Sum A':>12s}  {'Sum B':>12s}  {'Razlika':>10s}")
    print("-" * 68)

    diffs = []
    for col in num_cols:
        try:
            s1 = pd.to_numeric(df1[col], errors='coerce').sum()
            s2 = pd.to_numeric(df2[col], errors='coerce').sum()
            diff = s2 - s1
            flag = " ← RAZLIKA!" if abs(diff) > 0.01 else ""
            short = col[:28] if '(' in col else col
            print(f"  {short:<28s}  {s1:>12,.1f}  {s2:>12,.1f}  {diff:>+10.1f}{flag}")
            if flag:
                diffs.append(col)
        except Exception:
            pass

    # ── Usporedba po event_id ────────────────────────────────────────
    print()
    if 'event_id' in df1.columns and 'event_id' in df2.columns:
        ids1 = set(df1['event_id'].dropna())
        ids2 = set(df2['event_id'].dropna())
        only_in_1 = ids1 - ids2
        only_in_2 = ids2 - ids1
        if only_in_1:
            print(f"⚠  Samo u A ({len(only_in_1)} eventa): {list(only_in_1)[:5]}{'...' if len(only_in_1)>5 else ''}")
        if only_in_2:
            print(f"⚠  Samo u B ({len(only_in_2)} eventa): {list(only_in_2)[:5]}{'...' if len(only_in_2)>5 else ''}")
        if not only_in_1 and not only_in_2:
            print("✓  Isti set event_id-ova u oba fila")
    else:
        print("(event_id kolona nije pronađena za usporedbu)")

    print()
    if diffs:
        print(f"⚠  UKUPNO {len(diffs)} kolona s razlikom u sumi")
    else:
        print("✓  Sume svih kolona identične — import izgleda OK")


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Korištenje: python compare_xlsx.py file1.xlsx file2.xlsx")
        sys.exit(1)
    compare(sys.argv[1], sys.argv[2])
