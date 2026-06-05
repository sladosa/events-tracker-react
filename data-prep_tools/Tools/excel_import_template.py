# -*- coding: utf-8 -*-
"""
excel_import_template.py
=========================
Reference template za generiranje xlsx datoteke koja se moze importirati
u Events Tracker app (Activities → Import Excel).

KOPIRAJ OVO KAO OSNOVU za novi importer. Ne mijenjaj strukturu LEGEND / EVENT DATA.

Testirano: Health_medvisit_v3.xlsx (S77)

─────────────────────────────────────────────────────────────────────────────
OBAVEZNA STRUKTURA EVENTS SHEETA
─────────────────────────────────────────────────────────────────────────────

Row 1:  col A = "ATTRIBUTE LEGEND:"   (bilo koji tekst koji sadrzi tu frazu)
Row 2:  header: "Col" | "Area" | "Category_Path" | "Attribute" | "Type" | "Unit"
Row 3+: jedan red po atributu:
          col A = Excel column letter (I, J, K, L...)  ← MORA biti 1-3 uppercase slova
          col B = Area name  (npr. "Health_Sasa")
          col C = Category_Path BEZ area name  (npr. "Medical > Medical Visit")
          col D = Attribute name  (mora matchati header u EVENT DATA sekciji)
          col E = data type  (text / number / suggest / boolean / datetime)
          col F = unit  (opciono, npr. "EUR", "kg")

(blank row)

Next:   col A = "EVENT DATA:"   (bilo koji tekst koji sadrzi tu frazu)
Next:   header row:
          A=event_id | B=Area | C=Category_Path | D=event_date | E=session_start
          F=created_at | G=User | H=leaf comment | I+=attr columns
Next+:  data rows:
          col A = prazan  → CREATE novi event
          col A = UUID    → UPDATE postojeci event (rijetko koristi)
          col B = Area name
          col C = Category_Path (BEZ area name)
          col D = datum  YYYY-MM-DD string
          col E = vrijeme  HH:MM string  (npr. "08:00")
          col F = prazan
          col G = user email
          col H = leaf comment (opciono)
          col I+ = vrijednosti atributa (isti redosljed kao u LEGEND)

STRUCTURE SHEET (opciono, ali potreban za kreiranje kategorija):
  Ako kategorija ne postoji u bazi, import ce prikazati "confirm-structure" modal.
  Sheet "Structure" mora biti prisutan da bi taj modal ponudio "Create categories".
  Format: Type | CategoryPath (SA area name) | Sort | AttrName | Slug | AttrType |
          IsRequired | Val.Type | Default | ValMax | Unit | TextOptions |
          DependsOn | WhenValue | Description

  ⚠️  KRITICNO: AttrType vs Val.Type za suggest atribute
  ──────────────────────────────────────────────────────────
  U bazi, 'suggest' NIJE validan data_type. Suggest atributi se pohranjuju kao:
    data_type = 'text'  +  validation_rules = { "type": "suggest", "suggest": [...] }

  U Structure sheetu mora biti:
    AttrType (col F) = 'text'     ← uvijek 'text' za suggest (ovo je DB data_type)
    Val.Type (col H) = 'suggest'  ← ovo triggerira kreiranje validation_rules
    TextOptions (col L) = 'Opt1|Opt2|Opt3'

  ❌ POGREŠKA: AttrType = 'suggest' → structureImport.ts pokušava INSERT data_type='suggest'
     → DB odbija (nije valjani tip) → atribut se tiho preskače → u bazi nema suggest opcija
  ✅ ISPRAVNO:  AttrType = 'text'  + Val.Type = 'suggest' + TextOptions = 'Opt1|Opt2'

─────────────────────────────────────────────────────────────────────────────
CESTE GRESKE
─────────────────────────────────────────────────────────────────────────────

❌  Col A legend reda nije uppercase slovo → parser preskoči red, atribut se ne mapira
❌  "ATTRIBUTE LEGEND" naslov u merged celiji → parser ga ne vidi (merge skriva vrijednost)
❌  "— EVENT DATA —" umjesto "EVENT DATA:" → .includes('EVENT DATA') radi, ali
    section label u krivom retku može zbuniti parser
❌  Category_Path u LEGEND sadrzi area name → mapiranje ne radi
❌  Attr name u LEGEND ne matchira header u EVENT DATA → ValidationError
❌  Col A u data redovima nije prazan ni UUID → parser ga pokusa kao event_id
❌  Col B (Area) prazan u data redu → parser preskoci red (feature, ne bug)

─────────────────────────────────────────────────────────────────────────────
MINIMALNI WORKING PRIMJER (copy-paste za novi importer)
─────────────────────────────────────────────────────────────────────────────
"""

import sys, os
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# KONFIGURIRAJ OVO ZA SVAKI NOVI IMPORTER
# ─────────────────────────────────────────────────────────────────────────────

AREA      = "MojaArea"              # Area name u bazi
USER_EMAIL = "sasasladoljev59@gmail.com"
OUTPUT    = "import_output.xlsx"

# Atributi — redosljed odredjuje koji Excel stupac dobivaju (pocevsi od I)
# (ime, data_type, unit)
ATTR_DEFS = [
    ("Naziv",    "text",    ""),
    ("Iznos",    "number",  "EUR"),
    ("Napomena", "text",    ""),
]

# Svaki event: {"cat_path": "L1 > Leaf", "date": "YYYY-MM-DD",
#               "time": "08:00", "comment": "", attrs...}
EVENTS = [
    {"cat_path": "L1 > Leaf", "date": "2025-01-15", "time": "08:00",
     "Naziv": "Test event", "Iznos": 42.0, "Napomena": "Napomena tekst"},
    {"cat_path": "L1 > Leaf", "date": "2025-02-20", "time": "09:30",
     "Naziv": "Drugi event", "Iznos": None, "Napomena": ""},
]

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS (ne mijenjaj)
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex6):  return PatternFill("solid", fgColor=hex6)
def _font(**kw):  return Font(**kw)
def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

BORDER     = _border()
ALIGN_L    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_C    = Alignment(horizontal="center", vertical="center")
FILL_BLUE  = _fill("4472C4")
FILL_PURP  = _fill("7030A0")
FILL_ATTR  = _fill("DDEBF7")
FONT_WHITE = _font(bold=True, color="FFFFFF")
FONT_BOLD  = _font(bold=True, size=12)

def col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

FIXED_COUNT = 8   # cols A-H
ATTR_START  = FIXED_COUNT + 1   # col I

# ─────────────────────────────────────────────────────────────────────────────
# BUILD EVENTS SHEET
# ─────────────────────────────────────────────────────────────────────────────

def build_events_sheet(ws):
    ws.title = "Events"
    n_attr       = len(ATTR_DEFS)
    attr_letters = [col_letter(ATTR_START + i) for i in range(n_attr)]
    total_cols   = FIXED_COUNT + n_attr
    row = 1

    # ── ATTRIBUTE LEGEND title ────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="ATTRIBUTE LEGEND:").font = FONT_BOLD
    row += 1

    # ── ATTRIBUTE LEGEND header ───────────────────────────────────────────────
    for ci, h in enumerate(["Col","Area","Category_Path","Attribute","Type","Unit"], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = FILL_PURP; c.font = FONT_WHITE; c.alignment = ALIGN_C; c.border = BORDER
    row += 1

    # ── ATTRIBUTE LEGEND rows — col A = Excel column letter ───────────────────
    # Jedan red po atributu, cat_path BEZ area name
    for i, (attr_name, dtype, unit) in enumerate(ATTR_DEFS):
        # Atributi koji idu na isti leaf imaju isti cat_path
        # Atributi višeg nivoa (parent) imaju cat_path do tog nivoa
        cat_path = "L1 > Leaf"   # ← PRILAGODI PO KATEGORIJI
        data = [attr_letters[i], AREA, cat_path, attr_name, dtype, unit]
        for ci, v in enumerate(data, 1):
            c = ws.cell(row=row, column=ci, value=v or None)
            c.fill = FILL_ATTR; c.border = BORDER; c.alignment = ALIGN_L
        row += 1

    row += 1  # blank separator

    # ── EVENT DATA title ──────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="EVENT DATA:").font = FONT_BOLD
    row += 1

    # ── EVENT DATA header ─────────────────────────────────────────────────────
    event_header_row = row
    fixed_hdrs = ["event_id","Area","Category_Path","event_date","session_start",
                  "created_at","User","leaf comment"]
    attr_hdrs  = [f"{name}" for name, _, _ in ATTR_DEFS]   # ili "Name (Category)"

    for ci, h in enumerate(fixed_hdrs + attr_hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = FILL_BLUE; c.font = FONT_WHITE
        c.alignment = ALIGN_C; c.border = BORDER
    ws.row_dimensions[row].height = 28
    row += 1

    event_data_start = row

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ev in EVENTS:
        fixed_vals = [
            None,              # event_id: PRAZAN = CREATE novi event
            AREA,              # Area
            ev["cat_path"],    # Category_Path (BEZ area name!)
            ev["date"],        # event_date YYYY-MM-DD
            ev.get("time", "08:00"),  # session_start HH:MM
            None,              # created_at (prazan)
            USER_EMAIL,        # User email
            ev.get("comment") or None,  # leaf comment
        ]
        attr_vals = [ev.get(name) for name, _, _ in ATTR_DEFS]

        for ci, v in enumerate(fixed_vals + attr_vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.alignment = ALIGN_L; c.border = BORDER
        row += 1

    event_data_end = row - 1

    # ── Autofilter + freeze ───────────────────────────────────────────────────
    last = col_letter(total_cols)
    ws.auto_filter.ref = f"A{event_header_row}:{last}{event_data_end}"
    ws.freeze_panes    = f"I{event_data_start}"

    # ── Column widths ─────────────────────────────────────────────────────────
    for ltr, w in zip("ABCDEFGH", [10, 12, 30, 12, 10, 10, 30, 40]):
        ws.column_dimensions[ltr].width = w
    for ltr in attr_letters:
        ws.column_dimensions[ltr].width = 18

# ─────────────────────────────────────────────────────────────────────────────
# BUILD STRUCTURE SHEET (non-destructive — kreira samo ako ne postoji)
# ─────────────────────────────────────────────────────────────────────────────

def build_structure_sheet(ws):
    ws.title = "Structure"
    headers = ["Type","CategoryPath","Sort","AttrName","Slug","AttrType",
               "IsRequired","Val.Type","Default","ValMax","Unit",
               "TextOptions","DependsOn","WhenValue","Description"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = FILL_BLUE; c.font = FONT_WHITE
        c.alignment = ALIGN_C; c.border = BORDER

    # CategoryPath u Structure sheetu = SA area name
    rows = [
        ["area",      f"{AREA}",                "",  "", "", "",       "", "", "", "", "", "", "", "", ""],
        ["category",  f"{AREA} > L1",           "",  "", "", "",       "", "", "", "", "", "", "", "", ""],
        ["category",  f"{AREA} > L1 > Leaf",    "",  "", "", "",       "", "", "", "", "", "", "", "", ""],
        # Atributi: sort | AttrName | Slug | AttrType | IsRequired | Val.Type | ... | TextOptions | ... | Description
        ["attribute", f"{AREA} > L1 > Leaf",    10,  "Naziv",    "naziv",    "text",   "false", "",       "", "", "", "", "", "", ""],
        ["attribute", f"{AREA} > L1 > Leaf",    20,  "Iznos",    "iznos",    "number", "false", "",       "", "", "EUR", "", "", "", ""],
        ["attribute", f"{AREA} > L1 > Leaf",    30,  "Napomena", "napomena", "text",   "false", "",       "", "", "", "", "", "", ""],
    ]
    for ri, rv in enumerate(rows, 2):
        for ci, v in enumerate(rv, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = ALIGN_L; c.border = BORDER

    col_widths = [10, 40, 5, 16, 16, 10, 10, 10, 10, 10, 6, 60, 14, 14, 36]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[col_letter(ci)].width = w

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_events_sheet(wb.create_sheet("Events"))
    build_structure_sheet(wb.create_sheet("Structure"))
    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")
    print("Import: Activities → Import Excel → odaberi xlsx")
    print("App cita: Events sheet + Structure sheet (ostale ignorira)")
